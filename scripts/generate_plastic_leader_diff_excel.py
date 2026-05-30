# -*- coding: utf-8 -*-
"""生成间接人工成本塑料破碎分摊两列差异对比 Excel。"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).resolve().parent.parent / '间接人工成本-塑料破碎分摊列差异对比.xlsx'

COL1 = '列一：生产班组长(塑料破碎)分摊'
COL2 = '列二：生产班组长(塑料破碎)分摊（不考虑期初库存和库存结余）'

HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(name='微软雅黑', bold=True, size=12)
DATA_FONT = Font(name='微软雅黑', size=11)
WRAP = Alignment(wrap_text=True, vertical='center')
CENTER = Alignment(wrap_text=True, horizontal='center', vertical='center')
THIN = Side(style='thin', color='BFBFBF')


def border_all():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = border_all()


def style_data_area(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = DATA_FONT
            cell.alignment = WRAP
            cell.border = border_all()


def set_col_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_table(ws, title, headers, rows, col_widths):
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    for c, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c, value=h)
    style_header_row(ws, 2, len(headers))

    for r_idx, row in enumerate(rows, start=3):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    end_row = 2 + len(rows)
    style_data_area(ws, 3, end_row, len(headers))
    set_col_widths(ws, col_widths)
    ws.freeze_panes = 'A3'


def build_workbook():
    wb = Workbook()

    # Sheet1 总览差异
    ws1 = wb.active
    ws1.title = '总览差异表'
    write_table(
        ws1,
        '4.1 总览差异表（按对比维度）',
        ['对比维度', COL1, COL2],
        [
            ['列含义', '按「考虑期初库存和库存结余」口径的班组长提成成本占比分摊固定薪酬', '按「不考虑期初库存和库存结余」口径的班组长提成成本占比分摊固定薪酬'],
            ['薪酬池', 'pool_leader = 人员基础配置 × 月均固定成本 × 预测期数', 'pool_leader（与列一完全相同）'],
            ['分摊公式', '本行班组长提成成本 / Σ班组长提成成本 × pool_leader', '本行班组长提成成本(不考虑期初) / Σ班组长提成成本(不考虑期初) × pool_leader'],
            ['分摊权重分子', '班组长提成成本', '班组长提成成本(不考虑期初库存和库存结余)'],
            ['分摊权重分母', 'leader_total_commission（一次拆解产物+一破，提成成本>0 求和）', 'leader_total_no_opening（一次拆解产物+一破，不考虑期初提成成本>0 求和）'],
            ['参与分摊的类别', '一次拆解产物、一破', '一次拆解产物、一破'],
            ['不参与分摊的类别', '旧机、打包铁、屏（值为 0）', '旧机、打包铁、屏（旧机回退为列一值，通常为 0）'],
            ['是否计入「分摊固定成本合计」', '是', '否（单独展示，仅供对比分析）'],
            ['内部字段名', '生产班组长(塑料破碎)分摊固定成本', '生产班组长(塑料破碎)分摊固定成本(不考虑期初库存和库存结余)'],
            ['Excel 列写入条件', '类别=一次拆解产物 或 一破 时写入', '所有行均读取；未单独计算时回退为列一值'],
        ],
        [22, 48, 48],
    )

    # Sheet2 按类别差异
    ws2 = wb.create_sheet('按类别差异表')
    write_table(
        ws2,
        '4.2 按类别差异表（取数与结果是否相同）',
        ['类别', '列一 数量/重量来源', '列二 数量/重量来源', '数量是否相同', '两列分摊金额是否必然相同'],
        [
            ['旧机', '—（不参与塑料破碎池）', '—（不参与塑料破碎池）', '—', '是（均为 0 或回退为 0）'],
            ['一次拆解产物', '计算结果(KG)\ndisassembly_data', '计算结果(KG)（同左）', '相同', '是（分子分母同口径，占比一致 → 分摊金额相同）'],
            ['一破', '深加工结果(KG)\n深加工产值（考虑库存）', '数量/重量(不考虑期初库存和库存结余)\n深加工产值（不考虑期初）', '可能不同', '否（一破数量口径不同，占比不同 → 分摊金额可能不同）'],
            ['打包铁', '—（不参与塑料破碎池）', '—（不参与塑料破碎池）', '—', '是（均为 0）'],
            ['屏', '—（不参与塑料破碎池）', '—（不参与塑料破碎池）', '—', '是（均为 0）'],
        ],
        [14, 34, 34, 14, 28],
    )

    # Sheet3 一破专项
    ws3 = wb.create_sheet('一破专项差异')
    write_table(
        ws3,
        '4.3 一破类别专项差异表（两列差异的唯一来源）',
        ['对比项', COL1, COL2],
        [
            ['数量/重量字段', '深加工结果(KG)', '数量/重量(不考虑期初库存和库存结余)'],
            ['数量数据来源', 'calculate_deep_processing_product_output_value_data()\n（考虑期初库存和库存结余）', 'calculate_deep_processing_product_output_value_without_stock_data()\n（不考虑期初库存和库存结余）'],
            ['班组长提成成本', '深加工结果(KG) × 班组长提成单价', '不考虑期初 KG × 班组长提成单价'],
            ['单价', '相同：(深加工产物编码, 一破) → 班组长提成单价', '相同：(深加工产物编码, 一破) → 班组长提成单价'],
            ['何时两列数值相同', '当一破「考虑库存」与「不考虑期初」的 KG 完全相同时', '—'],
            ['何时两列数值不同', '当一破两种口径的 KG 存在差异时', '此时列二反映剔除期初库存/库存结余影响后的分摊'],
        ],
        [22, 42, 42],
    )

    # Sheet4 公式对照
    ws4 = wb.create_sheet('公式对照')
    write_table(
        ws4,
        '4.4 公式并列对照',
        ['项目', '内容'],
        [
            ['列一公式', '生产班组长(塑料破碎)分摊 = (班组长提成成本 / leader_total_commission) × pool_leader'],
            ['列二公式', '生产班组长(塑料破碎)分摊（不考虑期初库存和库存结余） = (班组长提成成本(不考虑期初) / leader_total_no_opening) × pool_leader'],
            ['班组长提成成本', '考虑库存口径的数量/重量 × 班组长提成单价'],
            ['班组长提成成本(不考虑期初)', '不考虑期初口径的数量/重量 × 班组长提成单价'],
            ['pool_leader', '人员基础配置 × 月均固定成本 × 预测期数（两列共用）'],
        ],
        [28, 90],
    )

    # Sheet5 差异结论
    ws5 = wb.create_sheet('差异结论')
    write_table(
        ws5,
        '4.5 差异结论（快速查阅）',
        ['分类', '说明'],
        [
            ['相同点-1', '共用同一薪酬池 pool_leader（岗位、人员配置、月均固定成本、预测期数均相同）'],
            ['相同点-2', '共用同一套班组长提成单价（计件人工标准 labor_cost_data）'],
            ['相同点-3', '参与分摊的物料类别相同（一次拆解产物 + 一破）'],
            ['相同点-4', '分摊方法相同（按班组长提成成本占比 × 薪酬池）'],
            ['不同点-1', '列二的分摊权重使用「不考虑期初库存和库存结余」口径的班组长提成成本'],
            ['不同点-2', '仅「一破」类别的数量/重量取数口径不同，这是两列分摊金额产生差异的唯一业务来源'],
            ['不同点-3', '列二不计入「分摊固定成本合计」；列一计入'],
            ['业务含义-列一', '反映含期初库存/库存结余影响下的塑料破碎班组长固定成本分摊'],
            ['业务含义-列二', '反映剔除期初库存/库存结余影响后的塑料破碎班组长固定成本分摊'],
            ['分析建议', '一次拆解产物两列结果一致；若需分析库存口径影响，重点对比「一破」行的两列差异'],
        ],
        [16, 90],
    )

    return wb


if __name__ == '__main__':
    wb = build_workbook()
    wb.save(OUTPUT)
    print(f'已生成: {OUTPUT}')

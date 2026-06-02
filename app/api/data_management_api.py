from flask import Blueprint, jsonify, request, send_file, session, current_app
import pandas as pd
import traceback
import io
import os
import importlib
import threading
import uuid
import copy
from datetime import datetime, timedelta
from data.base_data.mapping_data import get_mapping_dataframe, MAPPING_TABLE_DATA as MAPPING_DATA
from data.base_data.deduction_data import DEDUCTION_CODES
import data.base_data.deep_processing_data as dpd
from data.base_data.product_data import (
    PRODUCT_DISASSEMBLY_DATA,
    save_product_disassembly_to_disk,
    apply_product_disassembly_snapshot,
    reload_product_disassembly_json_if_stale,
)
from data.base_data.saleable_data import get_saleable_data_dataframe
from app.utils.data_utils import safe_json_convert
from app.utils.auth_utils import (
    login_required,
    require_can_edit,
    require_base_data_maintainer,
    is_base_data_maintainer,
    get_current_user,
)
from app.models.compatibility import AppDataManagerAdapter
from app.core.data_processor import supplement_old_machine_from_mapping
from app.models.database import db, User, Group, UserSession, SessionDataset
from app.models.session_data_manager import SessionDataManager, SessionDataManagerFactory
from app.services.sync_history_service import save_sync_entry
from app.services.sync_history_service import save_sync_entry

# 同步时复制的数据键（排除运行时状态）
SYNC_DATA_KEYS = [
    # 用户界面设置（预测期数等）
    'user_ui_settings',
    # 基础数据（只读，由计算引擎生成）
    'source_data', 'mapping_data', 'extracted_data', 'disassembly_data',
    'calculated_data', 'deep_processing_data', 'saleable_data',
    # 被减扣数据（系统只读 + 手工编辑相关）
    'deducted_data', 'deducted_data_manual', 'original_deducted_data', 'deducted_data_modified',
    'deducted_modifications', 'modification_timestamp', 'deep_processing_data_source',
    # 提取结果数据（手工编辑相关）
    'extracted_data_manual', 'original_extracted_data', 'extracted_data_modified',
    'extracted_modification_timestamp', 'extracted_modifications',
    # 可销售量数据（手工编辑相关）
    'saleable_data_manual', 'saleable_data_modified',
    'saleable_auto_init_timestamp', 'saleable_auto_sync_timestamp',
    'saleable_modification_timestamp', 'saleable_silent_init_timestamp',
    # 计算结果数据
    'subsidy_income_data', 'cost_forecast_data',
    'disassembly_product_output_value_data',
    # 数据修改记录
    'data_modifications',
    # 基础数据快照（映射表、产品拆解系数等）
    'base_product_disassembly', 'base_deep_processing_coefficients',
]

# 同步时需要按前缀通配复制的缓存键模板（period 为动态值）
SYNC_CACHE_KEY_PATTERNS = [
    'screen_cost_allocation_result_v2_',
    'screen_cost_allocation_result_',       # v1 缓存键
    'production_cost_allocation_result_v2_',
    'production_cost_allocation_result_',   # v1 缓存键
    'disassembly_product_cost_result_v2_',
    'disassembly_product_cost_result_',     # v1 缓存键
    'deep_processing_product_cost_result_v1_',
]

def get_session_data_manager():
    """获取会话数据管理器的便利函数"""
    from flask import session
    session_id = session.get('session_id')
    return AppDataManagerAdapter.get_instance(session_id)


def _align_deducted_with_disassembly(
    app_data,
    df,
    recalculate_kg: bool,
    recalculate_kg_when_tai_changed: bool = False,
):
    """将 DataFrame 与 disassembly_data 对齐；df 会被复制，不修改会话中的原始表。"""
    from app.utils.deducted_disassembly_align import align_deducted_inventory_tai_from_disassembly

    if df is None or df.empty:
        return df
    out = df.copy()
    disassembly = app_data.get_data('disassembly_data')
    if disassembly is not None and not disassembly.empty:
        out = align_deducted_inventory_tai_from_disassembly(
            out,
            disassembly,
            recalculate_kg=recalculate_kg,
            recalculate_kg_when_tai_changed=recalculate_kg_when_tai_changed,
        )
    return out


def _build_deducted_readonly_dataframe(app_data):
    """
    被减扣数据（只读）展示用 DataFrame。
    与可销售量数据模式一致：始终返回最新系统计算数据。
    已手工修改时仅同步 TAI（KG 保持公式重算），未修改时完全重算。
    """
    deducted_data = app_data.get_data('deducted_data')
    if deducted_data is None or (hasattr(deducted_data, 'empty') and deducted_data.empty):
        # 兼容旧数据：若无 deducted_data，尝试从 deducted_data_manual 读取
        manual = app_data.get_data('deducted_data_manual')
        if manual is not None and not (hasattr(manual, 'empty') and manual.empty):
            return _align_deducted_with_disassembly(app_data, manual, recalculate_kg=True)
        return None

    modified = bool(app_data.get_data('deducted_data_modified'))
    if modified:
        # 已修改：只同步 TAI，KG 由公式重算（保证只读视图反映最新源数据）
        return _align_deducted_with_disassembly(
            app_data,
            deducted_data,
            recalculate_kg=False,
            recalculate_kg_when_tai_changed=True,
        )
    # 未修改：完全重算
    return _align_deducted_with_disassembly(app_data, deducted_data, recalculate_kg=True)


def _build_deducted_manual_dataframe(app_data):
    """被减扣数据（手工）展示用；未修改时返回 None。"""
    if not bool(app_data.get_data('deducted_data_modified')):
        return None
    manual = app_data.get_data('deducted_data_manual')
    if manual is None or manual.empty:
        return None
    return _align_deducted_with_disassembly(
        app_data,
        manual,
        recalculate_kg=False,
        recalculate_kg_when_tai_changed=True,
    )


data_management_bp = Blueprint('data_management', __name__)


@data_management_bp.before_request
def _check_read_only_for_write():
    """基础数据写操作仅允许管理员或指定维护员"""
    if request.method in ('POST', 'PUT', 'DELETE'):
        user = get_current_user()
        if user and not is_base_data_maintainer(user):
            return jsonify({'success': False, 'message': '仅管理员或指定维护员可编辑基础数据'}), 403


# 添加全局错误处理器
@data_management_bp.errorhandler(Exception)
def handle_error(error):
    """全局错误处理器"""
    error_msg = str(error)
    print(f"数据管理API错误: {error_msg}")
    traceback.print_exc()
    return jsonify({
        'success': False,
        'error': f'服务器内部错误: {error_msg}'
    }), 500


@data_management_bp.route('/sync', methods=['POST'])
@login_required
@require_base_data_maintainer
def sync_data_to_users():
    """
    将当前用户会话数据同步到指定用户、用户组或全部用户。
    请求体: target_type in ['users','groups','all'], user_ids=[], group_ids=[], 可选 data_keys=[]
    """
    try:
        current_user_id = session.get('user_id')
        if not current_user_id:
            return jsonify({'success': False, 'message': '请先登录'}), 401
        current_user = User.query.get(current_user_id)
        if not current_user or getattr(current_user, 'is_read_only', False):
            return jsonify({'success': False, 'message': '只读用户无同步权限'}), 403

        data = request.get_json(silent=True) or {}
        target_type = data.get('target_type', 'users')
        if target_type not in ('users', 'groups', 'all'):
            return jsonify({'success': False, 'message': 'target_type 必须为 users、groups 或 all'}), 400
        user_ids = data.get('user_ids') or []
        group_ids = data.get('group_ids') or []
        data_keys = data.get('data_keys') or SYNC_DATA_KEYS
        if not isinstance(data_keys, list):
            data_keys = SYNC_DATA_KEYS
        if not data_keys:
            data_keys = SYNC_DATA_KEYS

        # 解析目标用户 ID 列表
        target_user_ids = set()
        if target_type == 'users':
            for uid in user_ids:
                try:
                    target_user_ids.add(int(uid))
                except (TypeError, ValueError):
                    pass
            if not target_user_ids:
                return jsonify({'success': False, 'message': '请选择至少一个目标用户'}), 400
            valid_users = User.query.filter(
                User.id.in_(target_user_ids),
                User.is_active == True
            ).all()
            target_user_ids = {u.id for u in valid_users}
            if not target_user_ids:
                return jsonify({'success': False, 'message': '所选用户无效或已禁用'}), 400
        elif target_type == 'groups':
            for gid in group_ids:
                try:
                    gid = int(gid)
                    for u in User.query.filter_by(group_id=gid, is_active=True).all():
                        target_user_ids.add(u.id)
                except (TypeError, ValueError):
                    pass
            if not target_user_ids:
                return jsonify({'success': False, 'message': '所选分组下没有用户或请选择分组'}), 400
        elif target_type == 'all':
            for u in User.query.filter_by(is_active=True).all():
                target_user_ids.add(u.id)
            if not target_user_ids:
                return jsonify({'success': False, 'message': '没有可同步的目标用户'}), 400

        # 排除自己
        target_user_ids.discard(current_user_id)
        if not target_user_ids:
            return jsonify({'success': True, 'message': '已同步到 0 个用户（仅自己或未选他人）', 'synced_count': 0})

        source_session_id = f'user_{current_user_id}'

        # 确保源会话存在
        if not UserSession.query.filter_by(session_id=source_session_id).first():
            return jsonify({'success': False, 'message': '当前用户暂无会话数据，请先上传或计算后再同步'}), 400

        ui_updates = {}
        raw_pp = data.get('prediction_period')
        if raw_pp is not None:
            try:
                pp = int(raw_pp)
                if 1 <= pp <= 120:
                    ui_updates['prediction_period'] = pp
            except (TypeError, ValueError):
                pass
        if not ui_updates:
            existing_ui = SessionDataManager.get_user_ui_settings(session_id=source_session_id)
            if isinstance(existing_ui.get('prediction_period'), int):
                ui_updates = {'prediction_period': existing_ui['prediction_period']}
        if ui_updates:
            SessionDataManager.upsert_user_ui_settings(source_session_id, ui_updates)
            if SessionDataManager.USER_UI_SETTINGS_KEY not in data_keys:
                data_keys = list(data_keys) + [SessionDataManager.USER_UI_SETTINGS_KEY]

        source_ui_settings = SessionDataManager.get_user_ui_settings(session_id=source_session_id)

        prod_snap_mem = None
        deep_snap_mem = None
        base_key_set = {'base_product_disassembly', 'base_deep_processing_coefficients'}
        if base_key_set.intersection(set(data_keys)):
            prod_snap_mem = copy.deepcopy(PRODUCT_DISASSEMBLY_DATA)
            deep_snap_mem = copy.deepcopy(dpd.DEEP_PROCESSING_DATA)
            if 'base_product_disassembly' in data_keys:
                row_p = SessionDataset.query.filter_by(
                    session_id=source_session_id, data_key='base_product_disassembly'
                ).first()
                if not row_p:
                    row_p = SessionDataset(session_id=source_session_id, data_key='base_product_disassembly')
                    db.session.add(row_p)
                row_p.set_json_data(prod_snap_mem)
            if 'base_deep_processing_coefficients' in data_keys:
                row_d = SessionDataset.query.filter_by(
                    session_id=source_session_id, data_key='base_deep_processing_coefficients'
                ).first()
                if not row_d:
                    row_d = SessionDataset(session_id=source_session_id, data_key='base_deep_processing_coefficients')
                    db.session.add(row_d)
                row_d.set_json_data(deep_snap_mem)
            db.session.flush()

        source_datasets = SessionDataset.query.filter(
            SessionDataset.session_id == source_session_id,
            SessionDataset.data_key.in_(data_keys)
        ).all()

        # 同时查询通配缓存键（如 production_cost_allocation_result_v2_1 等动态 period 键）
        cache_datasets = []
        for pattern in SYNC_CACHE_KEY_PATTERNS:
            matched = SessionDataset.query.filter(
                SessionDataset.session_id == source_session_id,
                SessionDataset.data_key.like(f'{pattern}%')
            ).all()
            cache_datasets.extend(matched)
        # 合并到 source_datasets，并将缓存键名加入 data_keys 以便 Redis 清理
        if cache_datasets:
            source_datasets = list(source_datasets) + cache_datasets
            for cd in cache_datasets:
                if cd.data_key not in data_keys:
                    data_keys = list(data_keys) + [cd.data_key]

        synced = 0
        synced_at_iso = datetime.now().isoformat()
        for target_user_id in target_user_ids:
            target_session_id = f'user_{target_user_id}'
            # 确保目标 UserSession 存在
            target_session = UserSession.query.filter_by(session_id=target_session_id).first()
            if not target_session:
                target_session = UserSession(
                    session_id=target_session_id,
                    user_ip=None,
                    user_agent=None,
                    expires_hours=24
                )
                db.session.add(target_session)
                db.session.flush()

            sync_id = None
            try:
                sync_id = save_sync_entry(
                    target_user_id,
                    source_datasets,
                    {
                        'from_user_id': current_user_id,
                        'from_name': (current_user.display_name or current_user.username or ''),
                        'synced_at': synced_at_iso,
                        'prediction_period': source_ui_settings.get('prediction_period'),
                        'data_keys': data_keys,
                    },
                )
            except Exception as hist_err:
                traceback.print_exc()
                print(f'保存同步历史失败(user_{target_user_id}): {hist_err}')

            # 仅投递同步历史与通知；接收方确认刷新时再通过 apply_sync_entry 写入会话
            try:
                SessionDataManager.set_sync_notification(
                    target_user_id,
                    {
                        'from_user_id': current_user_id,
                        'from_name': (current_user.display_name or current_user.username or ''),
                        'synced_at': synced_at_iso,
                        'prediction_period': source_ui_settings.get('prediction_period'),
                        'sync_id': sync_id,
                    },
                    redis_client=getattr(current_app, 'redis', None)
                )
            except Exception:
                pass
            synced += 1

        db.session.commit()

        if prod_snap_mem is not None and deep_snap_mem is not None:
            try:
                apply_product_disassembly_snapshot(copy.deepcopy(prod_snap_mem))
                dpd.apply_deep_processing_snapshot(copy.deepcopy(deep_snap_mem))
            except Exception as apply_err:
                traceback.print_exc()
                print(f"同步后应用基础数据快照到进程/磁盘失败: {apply_err}")

        return jsonify({
            'success': True,
            'message': f'已推送到 {synced} 个用户，待对方确认刷新后生效',
            'synced_count': synced,
            'target_type': target_type
        })
    except Exception as e:
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@data_management_bp.route('/sync-targets', methods=['GET'])
@login_required
@require_can_edit
def get_sync_targets():
    """获取可选的同步目标（全部活跃用户、全部分组）。"""
    try:
        current_user_id = session.get('user_id')
        current_user = User.query.get(current_user_id) if current_user_id else None
        if not current_user or getattr(current_user, 'is_read_only', False):
            return jsonify({'success': False, 'message': '无权限'}), 403
        users = User.query.filter_by(is_active=True).order_by(User.id).all()
        groups = Group.query.order_by(Group.id).all()
        return jsonify({
            'success': True,
            'data': {
                'users': [u.to_dict() for u in users],
                'groups': [g.to_dict() for g in groups],
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@data_management_bp.route('/info', methods=['GET'])
def get_data_info():
    """获取数据模块概览信息"""
    try:
        reload_product_disassembly_json_if_stale()
        dpd.reload_deep_processing_json_if_stale()
        # 获取映射表数据
        mapping_df = get_mapping_dataframe()
        mapping_count = len(mapping_df) if mapping_df is not None else 0
        
        # 获取减扣规则数据
        deduction_count = len(DEDUCTION_CODES)
        
        # 获取可销售量数据 - 优先使用计算结果数据，然后是手工数据，最后是原始数据
        app_data = get_session_data_manager()
        
        # 首先检查是否有计算结果数据
        calculated_saleable_data = app_data.get_data('saleable_data')
        if calculated_saleable_data is not None and not calculated_saleable_data.empty:
            saleable_data_count = len(calculated_saleable_data)
        else:
            # 其次检查手工数据
            manual_saleable_data = app_data.get_data('saleable_data_manual')
            if manual_saleable_data is not None and not manual_saleable_data.empty:
                saleable_data_count = len(manual_saleable_data)
            else:
                # 最后使用原始数据
                saleable_data_df = get_saleable_data_dataframe()
                saleable_data_count = len(saleable_data_df) if saleable_data_df is not None else 0
        
        # 检查是否有手工数据
        saleable_data_modified = app_data.get_data('saleable_data_modified') or False
        manual_saleable_data = app_data.get_data('saleable_data_manual')
        saleable_manual_count = len(manual_saleable_data) if manual_saleable_data is not None and not manual_saleable_data.empty else 0
        
        # 获取深加工数据
        deep_processing_df = dpd.get_deep_processing_dataframe()
        deep_processing_count = len(deep_processing_df) if deep_processing_df is not None else 0
        
        # 获取产品拆解系数数据
        product_count = len(PRODUCT_DISASSEMBLY_DATA)
        
        return jsonify({
            'success': True,
            'data': {
                'mapping': {
                    'count': mapping_count,
                    'name': '内置映射表',
                    'description': '电废拆解产物映射表数据'
                },
                'deduction': {
                    'count': deduction_count,
                    'name': '减扣规则',
                    'description': '一次拆解收发存及深加工投入产出数据'
                },
                'saleable_data': {
                    'count': saleable_data_count,
                    'name': '可销售量数据',
                    'description': '最终可销售量数据编辑和调整'
                },
                'saleable_manual': {
                    'count': saleable_manual_count,
                    'name': '可销售量数据(手工)',
                    'description': '手工编辑的可销售量数据',
                    'show': saleable_data_modified
                },
                'deep_processing': {
                    'count': deep_processing_count,
                    'name': '深加工数据',
                    'description': '深加工拆解系数表数据'
                },
                'product': {
                    'count': product_count,
                    'name': '产品拆解系数',
                    'description': '产品拆解系数数据字典'
                }
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/mapping', methods=['GET'])
def get_mapping_data():
    """获取映射表数据"""
    try:
        print(f"映射表API被调用 - 参数: {dict(request.args)}")
        
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        search = request.args.get('search', '')
        category_filter = request.args.get('category', '')
        
        print(f"解析参数 - page: {page}, per_page: {per_page}, search: '{search}', category: '{category_filter}'")
        
        df = get_mapping_dataframe()
        print(f"获取映射表数据 - 数据形状: {df.shape if df is not None else 'None'}")
        
        if df is None or df.empty:
            print("映射表数据为空，返回错误")
            return jsonify({'success': False, 'error': '无法获取映射表数据'}), 500
        
        # 搜索过滤
        if search:
            mask = (
                df['R3系统代码'].astype(str).str.contains(search, case=False, na=False) |
                df['系统名称'].astype(str).str.contains(search, case=False, na=False)
            )
            df = df[mask]
        
        # 类别过滤
        if category_filter and category_filter != '全部':
            df = df[df['类别'] == category_filter]
        
        # 分页
        total = len(df)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        df_page = df.iloc[start_idx:end_idx]
        
        # 获取所有类别用于筛选
        categories = ['全部'] + sorted(df['类别'].unique().tolist())
        
        return jsonify({
            'success': True,
            'data': safe_json_convert(df_page),
            'total': total,
            'current_page': page,
            'per_page': per_page,
            'pages': (total + per_page - 1) // per_page,
            'categories': categories
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/product', methods=['GET'])
def get_product_data():
    """获取产品拆解系数数据"""
    try:
        reload_product_disassembly_json_if_stale()
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        search = request.args.get('search', '').strip()
        record_type = request.args.get('record_type', '').strip()
        
        # 转换产品数据为DataFrame格式
        rows = []
        for product_code, product_info in PRODUCT_DISASSEMBLY_DATA.items():
            # 查找产品拆解系数名称
            product_name = ''
            for mapping in MAPPING_DATA:
                if mapping['R3系统代码'] == product_code:
                    product_name = mapping['系统名称']
                    break
            
            # 基本信息行
            base_row = {
                '产品代码': product_code,
                '产品拆解系数名称': product_name,
                '单台重量': product_info['单台重量'],
                '投入产出比例': product_info['一次拆解投入产出比例'],
                '拆解产物编码': '',
                '拆解产物名称': '',
                '拆解系数': '',
                '记录类型': '基本信息'
            }
            rows.append(base_row)
            
            # 拆解明细行
            for detail in product_info['拆解系数_明细']:
                # 修复拆解产物编码的小数点问题
                detail_product_code = str(detail['一次拆解产物编码'])
                if detail_product_code.endswith('.0'):
                    detail_product_code = detail_product_code[:-2]
                    
                detail_row = {
                    '产品代码': product_code,
                    '产品拆解系数名称': '',
                    '单台重量': '',
                    '投入产出比例': '',
                    '拆解产物编码': detail_product_code,
                    '拆解产物名称': detail['一次拆解产物名称'],
                    '拆解系数': detail['一次拆解系数'],
                    '记录类型': '拆解明细'
                }
                rows.append(detail_row)
        
        # 过滤数据
        if search:
            filtered_rows = []
            for row in rows:
                if (search.lower() in row['产品代码'].lower() or 
                    (row['拆解产物名称'] and search.lower() in row['拆解产物名称'].lower()) or
                    (row['产品拆解系数名称'] and search.lower() in row['产品拆解系数名称'].lower())):
                    filtered_rows.append(row)
            rows = filtered_rows
        
        if record_type:
            rows = [row for row in rows if row['记录类型'] == record_type]
        
        # 分页
        total = len(rows)
        start = (page - 1) * per_page
        end = start + per_page
        data = rows[start:end]
        
        # 统计信息
        product_count = len(PRODUCT_DISASSEMBLY_DATA)
        detail_count = sum(len(info['拆解系数_明细']) for info in PRODUCT_DISASSEMBLY_DATA.values())
        avg_weight = sum(info['单台重量'] for info in PRODUCT_DISASSEMBLY_DATA.values()) / product_count if product_count > 0 else 0
        
        statistics = {
            'product_count': product_count,
            'detail_count': detail_count,
            'avg_weight': avg_weight,
            'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        return jsonify({
            'success': True,
            'data': data,
            'pagination': {
                'current_page': page,
                'per_page': per_page,
                'total': total,
                'total_pages': (total + per_page - 1) // per_page
            },
            'statistics': statistics
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/product', methods=['POST'])
def create_product():
    """新增产品数据"""
    try:
        from data.base_data.product_data import add_product_data, backup_product_file, PRODUCT_DISASSEMBLY_DATA
        from data.base_data.mapping_data import MAPPING_TABLE_DATA as MAPPING_DATA
        
        # 获取请求数据
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'error': '未提供数据'}), 400
        
        # 验证必填字段
        product_code = data.get('产品代码', '').strip()
        if not product_code:
            return jsonify({'success': False, 'error': '产品代码不能为空'}), 400
        
        # 检查产品代码是否已存在
        if product_code in PRODUCT_DISASSEMBLY_DATA:
            return jsonify({'success': False, 'error': f'产品代码 {product_code} 已存在'}), 400
        
        # 验证其他必填字段
        if '单台重量' not in data:
            return jsonify({'success': False, 'error': '单台重量不能为空'}), 400
        if '投入产出比例' not in data:
            return jsonify({'success': False, 'error': '投入产出比例不能为空'}), 400
        if '拆解明细' not in data or not data['拆解明细']:
            return jsonify({'success': False, 'error': '至少需要一个拆解明细'}), 400
        
        # 转换数据格式：前端格式 -> 内部格式
        product_data = {
            '单台重量': float(data['单台重量']),
            '一次拆解投入产出比例': float(data['投入产出比例']),
            '拆解系数_明细': []
        }
        
        # 转换拆解明细
        for detail in data['拆解明细']:
            if not detail.get('拆解产物编码') or not detail.get('拆解产物名称') or '拆解系数' not in detail:
                continue
            
            product_data['拆解系数_明细'].append({
                '一次拆解产物编码': detail['拆解产物编码'],
                '一次拆解产物名称': detail['拆解产物名称'],
                '一次拆解系数': float(detail['拆解系数'])
            })
        
        if not product_data['拆解系数_明细']:
            return jsonify({'success': False, 'error': '拆解明细数据无效'}), 400
        
        # 备份文件
        backup_result = backup_product_file()
        if not backup_result:
            return jsonify({'success': False, 'error': '备份产品数据文件失败'}), 500
        
        # 添加产品数据
        if add_product_data(product_code, product_data):
            if not save_product_disassembly_to_disk():
                return jsonify({'success': False, 'error': '保存产品数据到磁盘失败'}), 500
            return jsonify({
                'success': True,
                'message': f'产品 {product_code} 添加成功',
                'product_code': product_code
            })
        else:
            return jsonify({'success': False, 'error': '添加产品数据失败'}), 500
        
    except ValueError as e:
        return jsonify({'success': False, 'error': f'数据格式错误: {str(e)}'}), 400
    except Exception as e:
        print(f"新增产品数据错误: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/product/<product_code>', methods=['GET'])
def get_single_product(product_code):
    """获取单个产品的详细信息"""
    try:
        reload_product_disassembly_json_if_stale()
        from data.base_data.product_data import get_product_data, PRODUCT_DISASSEMBLY_DATA
        from data.base_data.mapping_data import MAPPING_TABLE_DATA as MAPPING_DATA
        
        # 检查产品是否存在
        if product_code not in PRODUCT_DISASSEMBLY_DATA:
            return jsonify({'success': False, 'error': f'产品代码 {product_code} 不存在'}), 404
        
        # 获取产品数据
        product_info = PRODUCT_DISASSEMBLY_DATA[product_code]
        
        # 查找产品拆解系数名称
        product_name = ''
        for mapping in MAPPING_DATA:
            if mapping['R3系统代码'] == product_code:
                product_name = mapping['系统名称']
                break
        
        # 转换数据格式：内部格式 -> 前端格式
        product_data = {
            '产品代码': product_code,
            '产品拆解系数名称': product_name,
            '单台重量': product_info['单台重量'],
            '投入产出比例': product_info['一次拆解投入产出比例'],
            '拆解明细': []
        }
        
        # 转换拆解明细
        for detail in product_info['拆解系数_明细']:
            # 修复拆解产物编码的小数点问题
            detail_product_code = str(detail['一次拆解产物编码'])
            if detail_product_code.endswith('.0'):
                detail_product_code = detail_product_code[:-2]
            
            product_data['拆解明细'].append({
                '拆解产物编码': detail_product_code,
                '拆解产物名称': detail['一次拆解产物名称'],
                '拆解系数': detail['一次拆解系数']
            })
        
        return jsonify({
            'success': True,
            'data': product_data
        })
        
    except Exception as e:
        print(f"获取产品数据错误: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/product/<product_code>', methods=['PUT'])
def update_product(product_code):
    """更新产品数据"""
    try:
        from data.base_data.product_data import update_product_data, backup_product_file, PRODUCT_DISASSEMBLY_DATA
        
        # 检查产品是否存在
        if product_code not in PRODUCT_DISASSEMBLY_DATA:
            return jsonify({'success': False, 'error': f'产品代码 {product_code} 不存在'}), 404
        
        # 获取请求数据
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'error': '未提供数据'}), 400
        
        # 验证必填字段
        if '单台重量' not in data:
            return jsonify({'success': False, 'error': '单台重量不能为空'}), 400
        if '投入产出比例' not in data:
            return jsonify({'success': False, 'error': '投入产出比例不能为空'}), 400
        if '拆解明细' not in data or not data['拆解明细']:
            return jsonify({'success': False, 'error': '至少需要一个拆解明细'}), 400
        
        # 转换数据格式：前端格式 -> 内部格式
        product_data = {
            '单台重量': float(data['单台重量']),
            '一次拆解投入产出比例': float(data['投入产出比例']),
            '拆解系数_明细': []
        }
        
        # 转换拆解明细
        for detail in data['拆解明细']:
            if not detail.get('拆解产物编码') or not detail.get('拆解产物名称') or '拆解系数' not in detail:
                continue
            
            product_data['拆解系数_明细'].append({
                '一次拆解产物编码': detail['拆解产物编码'],
                '一次拆解产物名称': detail['拆解产物名称'],
                '一次拆解系数': float(detail['拆解系数'])
            })
        
        if not product_data['拆解系数_明细']:
            return jsonify({'success': False, 'error': '拆解明细数据无效'}), 400
        
        # 备份文件
        backup_result = backup_product_file()
        if not backup_result:
            return jsonify({'success': False, 'error': '备份产品数据文件失败'}), 500
        
        # 更新产品数据
        if update_product_data(product_code, product_data):
            if not save_product_disassembly_to_disk():
                return jsonify({'success': False, 'error': '保存产品数据到磁盘失败'}), 500
            return jsonify({
                'success': True,
                'message': f'产品 {product_code} 更新成功'
            })
        else:
            return jsonify({'success': False, 'error': '更新产品数据失败'}), 500
        
    except ValueError as e:
        return jsonify({'success': False, 'error': f'数据格式错误: {str(e)}'}), 400
    except Exception as e:
        print(f"更新产品数据错误: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/product/<product_code>', methods=['DELETE'])
def delete_product_record(product_code):
    """删除产品数据记录"""
    try:
        from data.base_data.product_data import delete_product_data, backup_product_file, PRODUCT_DISASSEMBLY_DATA
        
        # 检查产品是否存在
        if product_code not in PRODUCT_DISASSEMBLY_DATA:
            return jsonify({'success': False, 'error': f'产品代码 {product_code} 不存在'}), 404
        
        # 备份文件
        backup_result = backup_product_file()
        if not backup_result:
            return jsonify({'success': False, 'error': '备份产品数据文件失败'}), 500
        
        # 删除产品数据（内存 + JSON 侧车）
        if delete_product_data(product_code):
            if not save_product_disassembly_to_disk():
                return jsonify({'success': False, 'error': '保存产品数据到磁盘失败'}), 500
            return jsonify({
                'success': True,
                'message': f'产品 {product_code} 删除成功'
            })
        else:
            return jsonify({'success': False, 'error': '删除产品数据失败'}), 500
        
    except Exception as e:
        print(f"删除产品数据错误: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/deduction', methods=['GET'])
def get_deduction_rules():
    """获取减扣规则数据"""
    try:
        print(f"减扣规则API被调用 - 参数: {dict(request.args)}")
        
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        search = request.args.get('search', '')
        category_filter = request.args.get('category', '')
        
        print(f"解析参数 - page: {page}, per_page: {per_page}, search: '{search}', category: '{category_filter}'")
        print(f"减扣规则数据总数: {len(DEDUCTION_CODES)}")
        
        # 将字典转换为DataFrame
        data_list = []
        for code, info in DEDUCTION_CODES.items():
            data_list.append({
                '编码': code,
                '说明': info.get('说明', ''),
                '处置类别': info.get('处置类别', '未分类'),
                '来源': info.get('来源', ''),
                '备注': ''
            })
        
        print(f"转换为DataFrame - 记录数: {len(data_list)}")
        
        df = pd.DataFrame(data_list)
        
        # 搜索过滤
        if search:
            mask = (
                df['编码'].astype(str).str.contains(search, case=False, na=False) |
                df['说明'].astype(str).str.contains(search, case=False, na=False)
            )
            df = df[mask]
        
        # 类别过滤
        if category_filter and category_filter != '全部':
            df = df[df['处置类别'] == category_filter]
        
        # 分页
        total = len(df)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        df_page = df.iloc[start_idx:end_idx]
        
        # 获取所有类别用于筛选
        categories = ['全部'] + sorted(df['处置类别'].unique().tolist())
        
        return jsonify({
            'success': True,
            'data': safe_json_convert(df_page),
            'total': total,
            'current_page': page,
            'per_page': per_page,
            'pages': (total + per_page - 1) // per_page,
            'categories': categories
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/deduction/<deduction_code>', methods=['DELETE'])
def delete_deduction_rule(deduction_code):
    """删除减扣规则"""
    try:
        # 检查减扣规则是否存在
        if deduction_code not in DEDUCTION_CODES:
            print(f"尝试删除不存在的减扣规则: {deduction_code}")
            return jsonify({'success': False, 'error': f'减扣规则 {deduction_code} 不存在或已被删除'}), 404
        
        # 备份减扣数据文件
        backup_deduction_file()
        
        # 删除减扣规则
        if delete_deduction_code_from_file(deduction_code):
            print(f"✓ 已删除减扣规则: {deduction_code}")
            return jsonify({'success': True, 'message': '减扣规则删除成功'})
        else:
            return jsonify({'success': False, 'error': '删除减扣规则失败'}), 500
        
    except Exception as e:
        print(f"删除减扣规则错误: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/deduction/<deduction_code>', methods=['PUT'])
def update_deduction_rule(deduction_code):
    """更新减扣规则"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '请提供数据'}), 400
        
        # 验证必需字段
        description = data.get('说明', '').strip()
        disposal_category = data.get('处置类别', '').strip()
        source = data.get('来源', '').strip()
        
        if not description:
            return jsonify({'success': False, 'error': '说明不能为空'}), 400
        if not disposal_category:
            return jsonify({'success': False, 'error': '处置类别不能为空'}), 400
        
        # 检查记录是否存在
        if deduction_code not in DEDUCTION_CODES:
            print(f"尝试更新不存在的减扣规则: {deduction_code}")
            return jsonify({'success': False, 'error': f'减扣规则 {deduction_code} 不存在或已被删除'}), 404
        
        # 备份减扣数据文件
        backup_deduction_file()
        
        # 更新记录
        updated_record = {
            '说明': description,
            '处置类别': disposal_category,
            '来源': source or '一次拆解收发存及深加工投入产出.xlsx'
        }
        
        if update_deduction_code_in_file(deduction_code, updated_record):
            print(f"✓ 已更新减扣规则: {deduction_code}")
            return jsonify({'success': True, 'message': '减扣规则更新成功'})
        else:
            return jsonify({'success': False, 'error': '更新减扣规则失败'}), 500
        
    except Exception as e:
        print(f"更新减扣规则错误: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/deducted-data', methods=['GET'])
def get_deducted_data():
    """获取被减扣数据（只读）：始终返回最新系统计算数据（与可销售量模式一致）。"""
    try:
        app_data = get_session_data_manager()
        deducted_data_modified = bool(app_data.get_data('deducted_data_modified'))
        deducted_data = _build_deducted_readonly_dataframe(app_data)
        
        if deducted_data is None or deducted_data.empty:
            return jsonify({
                'success': True,
                'data': [],
                'total': 0,
                'message': '暂无被减扣数据',
                'deducted_data_modified': deducted_data_modified,
            })
        
        # 获取分页参数
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        search = request.args.get('search', '')
        disposal_category = (request.args.get('disposal_category') or '').strip()
        
        # 搜索过滤
        if search:
            mask = (
                deducted_data['拆解产物编码'].astype(str).str.contains(search, case=False, na=False) |
                deducted_data['拆解产物名称'].astype(str).str.contains(search, case=False, na=False) |
                deducted_data['原物料名称'].astype(str).str.contains(search, case=False, na=False)
            )
            filtered_data = deducted_data[mask]
        else:
            filtered_data = deducted_data

        if disposal_category and '处置类别' in filtered_data.columns:
            filtered_data = filtered_data[
                filtered_data['处置类别'].astype(str).str.strip() == disposal_category
            ]
        
        # 分页
        total = len(filtered_data)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        page_data = filtered_data.iloc[start_idx:end_idx]
        
        return jsonify({
            'success': True,
            'data': app_data.safe_json_convert(page_data),
            'total': total,
            'page': page,
            'per_page': per_page,
            'readonly': True,  # 标记为只读
            'deducted_data_modified': deducted_data_modified,
        })
        
    except Exception as e:
        print(f"获取被减扣数据失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/deducted-data-manual', methods=['GET'])
def get_deducted_data_manual():
    """获取被减扣数据(手工) - 可编辑；未编辑被减扣时返回空（无手工表）。"""
    try:
        app_data = get_session_data_manager()
        deducted_data_modified = bool(app_data.get_data('deducted_data_modified'))

        if not deducted_data_modified:
            return jsonify({
                'success': True,
                'data': [],
                'total': 0,
                'message': '未编辑被减扣数据，无被减扣数据(手工)',
                'deducted_data_modified': False,
                'readonly': False,
            })

        deducted_data_manual = _build_deducted_manual_dataframe(app_data)
        
        if deducted_data_manual is None or deducted_data_manual.empty:
            return jsonify({
                'success': True,
                'data': [],
                'total': 0,
                'message': '暂无被减扣数据(手工)',
                'deducted_data_modified': True,
                'readonly': False,
            })
        
        # 获取分页参数
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        search = request.args.get('search', '')
        disposal_category = (request.args.get('disposal_category') or '').strip()
        
        # 搜索过滤
        if search:
            mask = (
                deducted_data_manual['拆解产物编码'].astype(str).str.contains(search, case=False, na=False) |
                deducted_data_manual['拆解产物名称'].astype(str).str.contains(search, case=False, na=False) |
                deducted_data_manual['原物料名称'].astype(str).str.contains(search, case=False, na=False)
            )
            filtered_data = deducted_data_manual[mask]
        else:
            filtered_data = deducted_data_manual

        if disposal_category and '处置类别' in filtered_data.columns:
            filtered_data = filtered_data[
                filtered_data['处置类别'].astype(str).str.strip() == disposal_category
            ]
        
        # 分页
        total = len(filtered_data)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        page_data = filtered_data.iloc[start_idx:end_idx]
        
        return jsonify({
            'success': True,
            'data': app_data.safe_json_convert(page_data),
            'total': total,
            'page': page,
            'per_page': per_page,
            'readonly': False,  # 标记为可编辑
            'deducted_data_modified': True,
        })
        
    except Exception as e:
        print(f"获取被减扣数据(手工)失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/deducted-data', methods=['PUT'])
def update_deducted_data():
    """更新被减扣数据(手工)。"""
    try:
        import pandas as pd
        
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '请提供数据'}), 400
        
        app_data = get_session_data_manager()
        current_data = app_data.get_data('deducted_data_manual')
        
        if current_data is None or current_data.empty:
            return jsonify({'success': False, 'error': '没有被减扣数据(手工)'}), 400
        
        # 更新数据（批量上下文内合并多次 set_data 提交）
        modified_records = data.get('modified_records', [])
        updated_count = 0
        with app_data.batch_update():
            # 预构建序号索引，避免每条记录都全列扫描
            seq_index_map = {}
            try:
                seq_series = pd.to_numeric(current_data['序号'], errors='coerce')
                for idx, val in seq_series.items():
                    if pd.notna(val) and val not in seq_index_map:
                        seq_index_map[val] = idx
            except Exception:
                seq_index_map = {}

            # 首次修改时确保已备份原始数据
            original_data = app_data.get_data('original_deducted_data')
            if original_data is None or (isinstance(original_data, pd.DataFrame) and original_data.empty):
                app_data.backup_original_deducted_data()

            for record in modified_records:
                row_id = record.get('序号')
                if row_id is not None:
                    row_index = None
                    # 优先走预构建映射（增量更新时更快）
                    try:
                        rid_num = pd.to_numeric(row_id, errors='coerce')
                        if pd.notna(rid_num):
                            row_index = seq_index_map.get(rid_num)
                    except Exception:
                        row_index = None

                    # 映射未命中时回退旧逻辑
                    if row_index is None:
                        try:
                            mask = current_data['序号'] == row_id
                            if mask.any():
                                row_index = current_data[mask].index[0]
                        except Exception:
                            row_index = None

                    if row_index is not None:

                        # 直接更新计算结果字段
                        if '计算结果(KG)' in record:
                            old_value = current_data.loc[row_index, '计算结果(KG)']
                            new_value = record['计算结果(KG)']

                            try:
                                # 验证新值是数字
                                if new_value == '' or new_value is None:
                                    new_value = 0
                                else:
                                    new_value = float(new_value)

                                _ov = pd.to_numeric(old_value, errors='coerce')
                                old_f = 0.0 if pd.isna(_ov) else float(_ov)
                                if abs(old_f - new_value) > 1e-9:
                                    current_data.loc[row_index, '计算结果(KG)'] = new_value
                                    updated_count += 1
                                    print(f"✅ 修改被减扣数据(手工): 序号{row_id}, {old_value} -> {new_value}")
                            except (ValueError, TypeError) as e:
                                print(f"⚠️ 数值转换失败: 序号{row_id}, 值{new_value}, 错误{e}")
                                continue
        
        if updated_count <= 0:
            return jsonify({
                'success': True,
                'message': '未检测到有效修改',
                'updated_count': 0
            })

        # 标记数据已修改并保存
        with app_data.batch_update():
            app_data.mark_deducted_data_modified()
            app_data.set_data('deducted_data_manual', current_data)
            app_data.save_persistent_data()

        return jsonify({
            'success': True,
            'message': f'成功更新 {updated_count} 条记录，请返回首页点击"重新计算"使结果链路生效',
            'updated_count': updated_count
        })
        
    except Exception as e:
        print(f"更新被减扣数据(手工)失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/deducted-data/comparison', methods=['GET'])
def get_deducted_comparison():
    """获取被减扣数据对比统计"""
    try:
        app_data = get_session_data_manager()
        stats = app_data.get_deducted_comparison_stats()
        
        return jsonify({
            'success': True,
            'data': stats
        })
        
    except Exception as e:
        print(f"获取被减扣数据对比统计失败: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/deducted-data/update', methods=['POST'])
def update_deducted_data_bulk():
    """批量更新被减扣数据"""
    try:
        import pandas as pd
        from datetime import datetime
        
        data = request.get_json()
        if not data or 'data' not in data:
            return jsonify({
                'success': False,
                'error': '缺少数据参数'
            })
        
        # 获取数据管理器实例
        app_data = get_session_data_manager()
        
        # 将前端数据转换为DataFrame
        updated_data = pd.DataFrame(data['data'])
        
        # 🔧 架构重构：更新 deducted_data_manual，不再使用 deducted_data (只读)
        app_data.set_data('deducted_data_manual', updated_data)
        app_data.set_data('deducted_data_modified', True)
        app_data.set_data('modification_timestamp', datetime.now().isoformat())
        
        # 保存数据
        app_data.save_persistent_data()
        
        return jsonify({
            'success': True,
            'message': '被减扣数据已更新'
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'更新失败: {str(e)}'
        }), 500

@data_management_bp.route('/deducted-data/reset', methods=['POST'])
def reset_deducted_data():
    """重置被减扣数据到原始状态"""
    try:
                # 获取数据管理器实例（单例模式）
        app_data = get_session_data_manager()
        
        # 🔧 架构重构：检查 deducted_data_manual，不再使用 deducted_data (只读)
        deducted_data = app_data.get_data('deducted_data_manual')
        if deducted_data is None or deducted_data.empty:
            # 如果手工数据为空，尝试使用原始备份
            original_data = app_data.get_data('original_deducted_data')
            if original_data is not None and not original_data.empty:
                deducted_data = original_data
            else:
                return jsonify({
                    'success': False,
                    'error': '没有找到被减扣数据，请先进行营业收入预测计算'
                })
        
        # 检查是否有原始数据备份
        original_data = app_data.get_data('original_deducted_data')
        if original_data is None or original_data.empty:
            # 如果没有原始备份，创建备份
            success = app_data.backup_original_deducted_data()
            if not success:
                return jsonify({
                    'success': False,
                    'error': '无法创建原始数据备份'
                })
        
        # 使用AppDataManager的重置方法
        success = app_data.reset_deducted_data_to_original()
        
        if not success:
            return jsonify({
                'success': False,
                'error': '重置失败，请检查原始数据备份'
            })
        
        # 🔧 重要：恢复原始数据后，应该使用原始数据重新计算
        # 清除所有相关缓存，确保使用原始数据重新计算
        try:
            from flask import session
            prediction_period = 1
            try:
                session_prediction_period = session.get('prediction_period')
                if session_prediction_period:
                    prediction_period = int(session_prediction_period)
            except:
                pass
            
            # 清除所有可能依赖被减扣数据的缓存
            cache_keys_to_clear = [
                f'disassembly_product_cost_result_v2_{prediction_period}',
                f'screen_cost_allocation_result_v2_{prediction_period}',
                f'production_cost_allocation_result_v2_{prediction_period}',
                f'screen_cost_allocation_result_{prediction_period}',
                f'production_cost_allocation_result_{prediction_period}',
            ]
            for cache_key in cache_keys_to_clear:
                app_data.set_data(cache_key, None)  # 清除缓存
                print(f"🗑️ 已清除缓存: {cache_key}")
            
            # 重新计算深加工数据（使用原始数据，因为 deducted_data_modified 现在是 False）
            print("🔄 使用原始数据重新计算深加工数据...")
            from app.core.calculation_engine import CalculationEngine
            calculation_engine = CalculationEngine()
            success_deep = calculation_engine.calculate_deep_processing_auto()
            
            if success_deep:
                print("✅ 深加工数据重新计算完成（使用原始数据）")
                # 自动合并生成可销售量数据
                calculation_engine.merge_saleable_data()
                print("✅ 可销售量数据已更新（使用原始数据）")
                
                # 🔧 重要：恢复原始数据后，需要清除可销售量手工数据，确保使用系统计算的数据
                # 因为恢复原始数据意味着不再使用手工编辑的数据
                saleable_data = app_data.get_data('saleable_data')
                if saleable_data is not None and not saleable_data.empty:
                    # 清除手工数据，使用系统计算的数据
                    app_data.set_data('saleable_data_manual', saleable_data.copy())
                    app_data.set_data('saleable_data_modified', False)  # 标记为未修改
                    print("✅ 可销售量手工数据已重置为系统计算的数据（未标记为手工修改）")
            else:
                print("⚠️ 深加工数据重新计算失败")
            
            # 重新计算成本预测
            try:
                print("🔄 自动重新计算成本预测数据...")
                manual_extracted_data = app_data.get_data('extracted_data_manual')
                if manual_extracted_data is not None and not manual_extracted_data.empty:
                    from app.api.cost_forecast_api import calculate_material_cost
                    cost_data = calculate_material_cost(manual_extracted_data)
                    app_data.set_data('cost_forecast_data', cost_data)
                    print(f"✅ 成本预测数据（拆解物原料成本）重新计算完成: {len(cost_data)} 条记录")
                else:
                    print("⚠️ 没有提取结果手工数据，跳过成本预测计算")
            except Exception as cost_error:
                print(f"⚠️ 成本预测计算失败: {cost_error}")
                import traceback
                traceback.print_exc()
            
            # 重新计算成本计算（一次拆解产物成本等）
            try:
                print("🔄 自动重新计算成本计算数据（使用原始数据）...")
                from app.api.cost_forecast_api import calculate_disassembly_product_cost
                result_data = calculate_disassembly_product_cost(app_data, prediction_period)
                if result_data:
                    cache_key = f'disassembly_product_cost_result_v2_{prediction_period}'
                    app_data.set_data(cache_key, result_data)
                    print(f"✅ 成本计算数据（一次拆解产物成本）重新计算完成: {len(result_data)} 条记录")
                    print(f"   📊 所有计算已使用原始数据重新计算")
                else:
                    print("⚠️ 成本计算数据为空")
            except Exception as cost_calc_error:
                print(f"⚠️ 成本计算失败: {cost_calc_error}")
                import traceback
                traceback.print_exc()
        except Exception as calc_error:
            print(f"⚠️ 自动重新计算失败: {calc_error}")
            import traceback
            traceback.print_exc()
            # 不影响重置的成功返回
        
        # 保存数据
        app_data.save_persistent_data()
        
        return jsonify({
            'success': True,
            'message': '已成功重置被减扣数据到原始状态，并已使用原始数据重新计算'
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'重置失败: {str(e)}'
        }), 500

@data_management_bp.route('/deducted-data/export', methods=['GET'])
def export_deducted_data():
    """导出被减扣数据到Excel（支持 format=aggregated）"""
    try:
        from app.utils.excel_utils import create_deducted_data_excel
        import os
        from flask import send_file

        export_format = request.args.get('format', 'detail')

        app_data = get_session_data_manager()
        deducted_data = app_data.get_data('deducted_data_manual')

        if deducted_data is None or deducted_data.empty:
            deducted_data_modified = app_data.get_data('deducted_data_modified')
            if not deducted_data_modified:
                system_data = app_data.get_data('deducted_data')
                if system_data is not None and not system_data.empty:
                    deducted_data = system_data

        if deducted_data is None or deducted_data.empty:
            return jsonify({'success': False, 'error': '暂无被减扣数据可导出'}), 400

        # 辅助函数：从明细数据构建汇总DataFrame
        def build_deducted_aggregated_df(source_df):
            grouped = {}
            for _, row in source_df.iterrows():
                code = str(row.get('拆解产物编码', '')).strip()
                if not code: continue
                if code not in grouped:
                    grouped[code] = {
                        '拆解产物编码': code,
                        '拆解产物名称': str(row.get('拆解产物名称', '')).strip(),
                        '汇总计算结果(KG)': 0.0,
                        '原库存数量(TAI)': 0.0,
                        '明细行数': 0
                    }
                kg = pd.to_numeric(row.get('计算结果(KG)', 0), errors='coerce')
                kg = kg if pd.notna(kg) else 0.0
                tai = pd.to_numeric(row.get('原库存数量(TAI)', 0), errors='coerce')
                tai = tai if pd.notna(tai) else 0.0
                grouped[code]['汇总计算结果(KG)'] += kg
                grouped[code]['原库存数量(TAI)'] += tai
                grouped[code]['明细行数'] += 1
                if not grouped[code]['拆解产物名称'] or grouped[code]['拆解产物名称'] == 'nan':
                    grouped[code]['拆解产物名称'] = str(row.get('拆解产物名称', '')).strip()
            rows = []
            for code in sorted(grouped.keys()):
                g = grouped[code]
                rows.append({
                    '拆解产物编码': g['拆解产物编码'],
                    '拆解产物名称': g['拆解产物名称'],
                    '汇总计算结果(KG)': round(g['汇总计算结果(KG)'], 6),
                    '原库存数量(TAI)': round(g['原库存数量(TAI)'], 6),
                    '明细行数': g['明细行数']
                })
            return pd.DataFrame(rows)

        def auto_width_deducted(ws):
            for col_idx, col in enumerate(ws.columns, start=1):
                col_letter = chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)
                try:
                    ws.column_dimensions[col_letter].width = 18
                except: pass

        if export_format == 'all':
            # 导出明细 + 汇总（双 sheet）
            detail_df = deducted_data
            aggregated_df = build_deducted_aggregated_df(deducted_data)
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                detail_df.to_excel(writer, sheet_name='明细数据', index=False)
                auto_width_deducted(writer.sheets['明细数据'])
                aggregated_df.to_excel(writer, sheet_name='汇总编辑数据', index=False)
                auto_width_deducted(writer.sheets['汇总编辑数据'])
            output.seek(0)
            return send_file(output, as_attachment=True,
                download_name=f'被减扣数据_完整版_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        if export_format == 'aggregated':
            df = build_deducted_aggregated_df(deducted_data)
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='汇总编辑数据', index=False)
                auto_width_deducted(writer.sheets['汇总编辑数据'])
            output.seek(0)
            return send_file(output, as_attachment=True,
                download_name=f'被减扣数据_汇总编辑_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # 明细导出
        excel_path = create_deducted_data_excel(deducted_data)

        if excel_path and os.path.exists(excel_path):
            return send_file(excel_path, as_attachment=True,
                download_name='被减扣数据.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        else:
            return jsonify({'success': False, 'error': 'Excel文件创建失败'}), 500

    except Exception as e:
        print(f"导出被减扣数据失败: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/deducted-data/template', methods=['GET'])
def download_deducted_data_template():
    """下载被减扣数据导入模板（支持 format=aggregated）"""
    try:
        from app.utils.excel_utils import create_deducted_data_template
        import os
        from flask import send_file

        template_format = request.args.get('format', 'detail')

        if template_format == 'aggregated':
            # 汇总格式模板：导出当前实际汇总数据
            app_data = get_session_data_manager()
            source = app_data.get_data('deducted_data_manual')
            if source is None or (hasattr(source, 'empty') and source.empty):
                source = app_data.get_data('deducted_data')
            if source is None or (hasattr(source, 'empty') and source.empty):
                # 无数据时返回示例
                df = pd.DataFrame([{
                    '拆解产物编码': '811052988',
                    '拆解产物名称': '示例产物',
                    '汇总计算结果(KG)': 300.123456,
                    '原库存数量(TAI)': 50.000000,
                    '明细行数': 3
                }])
            else:
                grouped = {}
                for _, row in source.iterrows():
                    code = str(row.get('拆解产物编码', '')).strip()
                    if not code: continue
                    if code not in grouped:
                        grouped[code] = {
                            '拆解产物编码': code,
                            '拆解产物名称': str(row.get('拆解产物名称', '')).strip(),
                            '汇总计算结果(KG)': 0.0,
                            '原库存数量(TAI)': 0.0,
                            '明细行数': 0
                        }
                    kg = pd.to_numeric(row.get('计算结果(KG)', 0), errors='coerce')
                    kg = kg if pd.notna(kg) else 0.0
                    tai = pd.to_numeric(row.get('原库存数量(TAI)', 0), errors='coerce')
                    tai = tai if pd.notna(tai) else 0.0
                    grouped[code]['汇总计算结果(KG)'] += kg
                    grouped[code]['原库存数量(TAI)'] += tai
                    grouped[code]['明细行数'] += 1
                    if not grouped[code]['拆解产物名称'] or grouped[code]['拆解产物名称'] == 'nan':
                        grouped[code]['拆解产物名称'] = str(row.get('拆解产物名称', '')).strip()

                rows = []
                for code in sorted(grouped.keys()):
                    g = grouped[code]
                    rows.append({
                        '拆解产物编码': g['拆解产物编码'],
                        '拆解产物名称': g['拆解产物名称'],
                        '汇总计算结果(KG)': round(g['汇总计算结果(KG)'], 6),
                        '原库存数量(TAI)': round(g['原库存数量(TAI)'], 6),
                        '明细行数': g['明细行数']
                    })
                df = pd.DataFrame(rows)

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='汇总编辑数据', index=False)
                ws = writer.sheets['汇总编辑数据']
                for col_idx in range(len(df.columns)):
                    try:
                        ws.column_dimensions[chr(65 + col_idx)].width = 18
                    except: pass

            output.seek(0)
            return send_file(output, as_attachment=True,
                download_name=f'被减扣数据导入模板_汇总_{datetime.now().strftime("%Y%m%d")}.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # 明细格式模板
        template_path = create_deducted_data_template()

        if template_path and os.path.exists(template_path):
            return send_file(template_path, as_attachment=True,
                download_name='被减扣数据导入模板.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        else:
            return jsonify({'success': False, 'error': '模板文件创建失败'}), 500

    except Exception as e:
        print(f"创建模板失败: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/parse-deducted-excel', methods=['POST'])
def parse_deducted_excel():
    """解析被减扣数据Excel（自动识别明细/汇总格式）"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '没有文件'}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '没有选择文件'}), 400

        df = pd.read_excel(file)
        if df.empty:
            return jsonify({'success': False, 'error': 'Excel文件为空'}), 400

        is_aggregated = '汇总计算结果(KG)' in df.columns

        if is_aggregated:
            if '拆解产物编码' not in df.columns or '汇总计算结果(KG)' not in df.columns:
                return jsonify({'success': False, 'error': '汇总格式缺少必要列: 拆解产物编码, 汇总计算结果(KG)'}), 400
            df = df.dropna(subset=['拆解产物编码', '汇总计算结果(KG)'])
            df['汇总计算结果(KG)'] = pd.to_numeric(df['汇总计算结果(KG)'], errors='coerce')
            df = df.dropna(subset=['汇总计算结果(KG)'])
            if df.empty:
                return jsonify({'success': False, 'error': '没有有效的数据行'}), 400
            data_dict = safe_json_convert(df)
            return jsonify({'success': True, 'data': data_dict, 'count': len(data_dict), 'format': 'aggregated'})
        else:
            required = ['原物料代码', '原物料名称', '拆解产物编码', '拆解产物名称', '计算结果(KG)', '处置类别']
            missing = [c for c in required if c not in df.columns]
            if missing:
                return jsonify({'success': False, 'error': f'明细格式缺少必要列: {", ".join(missing)}'}), 400
            df = df.dropna(subset=['拆解产物编码', '计算结果(KG)'])
            df['计算结果(KG)'] = pd.to_numeric(df['计算结果(KG)'], errors='coerce')
            df = df.dropna(subset=['计算结果(KG)'])
            if df.empty:
                return jsonify({'success': False, 'error': '没有有效的数据行'}), 400
            data_dict = safe_json_convert(df)
            return jsonify({'success': True, 'data': data_dict, 'count': len(data_dict), 'format': 'detail'})

    except Exception as e:
        print(f"解析被减扣Excel失败: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/import-deducted-data', methods=['POST'])
def import_deducted_data():
    """导入被减扣数据（明细格式，写入手工表）"""
    try:
        import copy
        import pandas as pd

        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '请提供数据'}), 400

        import_data = data.get('data', [])
        import_mode = data.get('mode', 'replace')

        if not import_data:
            return jsonify({'success': False, 'error': '没有有效的导入数据'}), 400

        import_df = pd.DataFrame(import_data)
        app_data = get_session_data_manager()

        current_data = app_data.get_data('deducted_data_manual')
        if current_data is not None and not current_data.empty:
            app_data.set_data('original_deducted_data', copy.deepcopy(current_data))
        else:
            original_data = app_data.get_data('original_deducted_data')
            if original_data is None or (isinstance(original_data, pd.DataFrame) and original_data.empty):
                app_data.backup_original_deducted_data()

        if import_mode == 'replace':
            new_data = import_df
            imported_count = len(import_df)
        elif import_mode == 'merge':
            if current_data is not None and not current_data.empty:
                key_column = '拆解产物编码'
                if key_column in import_df.columns and key_column in current_data.columns:
                    mask = ~current_data[key_column].isin(import_df[key_column])
                    filtered_current = current_data[mask]
                    new_data = pd.concat([filtered_current, import_df], ignore_index=True)
                else:
                    new_data = pd.concat([current_data, import_df], ignore_index=True)
            else:
                new_data = import_df
            imported_count = len(import_df)
        elif import_mode == 'append':
            if current_data is not None and not current_data.empty:
                new_data = pd.concat([current_data, import_df], ignore_index=True)
            else:
                new_data = import_df
            imported_count = len(import_df)
        else:
            return jsonify({'success': False, 'error': '无效的导入模式'}), 400

        if not new_data.empty:
            new_data['序号'] = range(1, len(new_data) + 1)

        app_data.set_data('deducted_data_manual', new_data)
        app_data.mark_deducted_data_modified()
        app_data.save_persistent_data()

        print(f"✅ 被减扣数据导入成功: {imported_count} 条记录，总计 {len(new_data)} 条")

        return jsonify({
            'success': True,
            'message': f'数据导入成功，{import_mode}模式导入了 {imported_count} 条记录。请返回首页点击“重新计算”使结果链路生效',
            'imported_count': imported_count,
            'total_count': len(new_data)
        })

    except Exception as e:
        print(f"导入被减扣数据失败: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'导入失败: {str(e)}'}), 500

@data_management_bp.route('/deep-processing', methods=['GET'])
def get_deep_processing_data():
    """获取深加工数据"""
    try:
        dpd.reload_deep_processing_json_if_stale()
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        search = request.args.get('search', '')
        material_code = request.args.get('material_code', '')
        
        df = dpd.get_deep_processing_dataframe()
        if df is None or df.empty:
            return jsonify({'success': False, 'error': '无法获取深加工数据'}), 500
        
        # 为每条记录添加唯一ID（基于索引）
        df_data = df.to_dict('records')
        for i, item in enumerate(df_data):
            item['id'] = i
        
        # 搜索过滤
        if search:
            filtered_data = []
            for item in df_data:
                for key, value in item.items():
                    if search.lower() in str(value).lower():
                        filtered_data.append(item)
                        break
            df_data = filtered_data
        
        # 物料代码过滤
        if material_code:
            filtered_data = []
            for item in df_data:
                if material_code.lower() in str(item.get('拆解产物编码', '')).lower():
                    filtered_data.append(item)
            df_data = filtered_data
        
        # 分页
        total = len(df_data)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        page_data = df_data[start_idx:end_idx]
        
        return jsonify({
            'success': True,
            'data': page_data,
            'total': total,
            'current_page': page,
            'per_page': per_page,
            'pages': (total + per_page - 1) // per_page
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/deep-processing/statistics', methods=['GET'])
def get_deep_processing_statistics():
    """获取深加工数据统计信息"""
    try:
        dpd.reload_deep_processing_json_if_stale()
        df = dpd.get_deep_processing_dataframe()
        if df is None or df.empty:
            return jsonify({'success': False, 'error': '无法获取深加工数据'}), 500
        
        # 计算统计信息
        total_records = len(df)
        
        # 统计拆解产物编码数量（投入物料）
        input_material_count = 0
        if '拆解产物编码' in df.columns:
            input_material_count = df['拆解产物编码'].nunique()
        
        # 统计深加工产物编码数量（产出物料）
        output_material_count = 0
        if '深加工产物编码' in df.columns:
            output_material_count = df['深加工产物编码'].nunique()
        
        # 计算平均深加工拆解系数
        avg_coefficient = 0
        if '深加工拆解系数' in df.columns:
            numeric_coefficients = pd.to_numeric(df['深加工拆解系数'], errors='coerce')
            avg_coefficient = numeric_coefficients.mean() if not numeric_coefficients.isna().all() else 0
        
        statistics = {
            'total_records': total_records,
            'material_count': input_material_count,  # 前端期望的字段名
            'deep_product_count': output_material_count,  # 前端期望的字段名
            'avg_coefficient': float(avg_coefficient) if pd.notna(avg_coefficient) else 0,
            'last_updated': datetime.now().strftime('%Y-%m-%d')
        }
        
        return jsonify({
            'success': True,
            'data': statistics  # 前端期望的是data字段，不是statistics
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/deep-processing/<int:record_id>', methods=['GET'])
def get_deep_processing_record(record_id):
    """获取单个深加工数据记录"""
    try:
        dpd.reload_deep_processing_json_if_stale()
        from data.base_data.deep_processing_data import DEEP_PROCESSING_DATA
        
        if record_id < 0 or record_id >= len(DEEP_PROCESSING_DATA):
            return jsonify({'success': False, 'error': '记录不存在'}), 404
        
        record = DEEP_PROCESSING_DATA[record_id].copy()
        record['id'] = record_id
        
        return jsonify({
            'success': True,
            'data': record
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/deep-processing/<int:record_id>', methods=['PUT'])
def update_deep_processing_record(record_id):
    """更新深加工数据记录"""
    try:
        from data.base_data.deep_processing_data import DEEP_PROCESSING_DATA, write_deep_processing_data_to_file, backup_deep_processing_file
        
        if record_id < 0 or record_id >= len(DEEP_PROCESSING_DATA):
            return jsonify({'success': False, 'error': '记录不存在'}), 404
        
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '没有提供数据'}), 400
        
        # 备份文件
        backup_deep_processing_file()
        
        # 更新记录
        updated_record = {}
        required_fields = [
            '拆解产物编码', '拆解产物名称', '产品拆解系数名称',
            '深加工投入产出比例', '深加工拆解系数', '深加工产物编码', '深加工产物名称'
        ]
        
        for field in required_fields:
            if field in data:
                if field in ['深加工投入产出比例', '深加工拆解系数']:
                    try:
                        updated_record[field] = float(data[field])
                    except (ValueError, TypeError):
                        updated_record[field] = 1.0
                else:
                    updated_record[field] = str(data[field]).strip()
            else:
                updated_record[field] = DEEP_PROCESSING_DATA[record_id].get(field, '')
        
        # 更新数据
        DEEP_PROCESSING_DATA[record_id] = updated_record
        
        # 写入文件
        write_deep_processing_data_to_file()
        
        return jsonify({
            'success': True,
            'message': '深加工数据更新成功',
            'data': updated_record
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/deep-processing/<int:record_id>', methods=['DELETE'])
def delete_deep_processing_record(record_id):
    """删除深加工数据记录"""
    try:
        from data.base_data.deep_processing_data import DEEP_PROCESSING_DATA, write_deep_processing_data_to_file, backup_deep_processing_file
        
        if record_id < 0 or record_id >= len(DEEP_PROCESSING_DATA):
            return jsonify({'success': False, 'error': '记录不存在'}), 404
        
        # 备份文件
        backup_deep_processing_file()
        
        # 获取要删除的记录信息
        deleted_record = DEEP_PROCESSING_DATA[record_id]
        
        # 删除记录
        del DEEP_PROCESSING_DATA[record_id]
        
        # 写入文件
        write_deep_processing_data_to_file()
        
        return jsonify({
            'success': True,
            'message': f'已删除深加工数据记录: {deleted_record.get("拆解产物名称", "未知")} -> {deleted_record.get("深加工产物名称", "未知")}'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/deep-processing/import', methods=['POST'])
def import_deep_processing_data():
    """导入深加工数据"""
    try:
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'message': '没有上传文件'
            }), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({
                'success': False,
                'message': '没有选择文件'
            }), 400
        
        import_mode = request.form.get('import_mode', 'append')  # append 或 replace
        
        # 备份深加工数据文件
        from data.base_data.deep_processing_data import backup_deep_processing_file
        backup_deep_processing_file()
        
        # 读取Excel文件
        df = pd.read_excel(file)
        
        # 验证必要的列
        required_columns = [
            '拆解产物编码', '拆解产物名称', '产品拆解系数名称',
            '深加工投入产出比例', '深加工拆解系数', '深加工产物编码', '深加工产物名称'
        ]
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({
                'success': False,
                'message': f'缺少必要的列: {", ".join(missing_columns)}'
            }), 400
        
        # 数据验证和处理
        df = df.dropna(subset=['拆解产物编码'])  # 删除编码为空的行
        df['拆解产物编码'] = df['拆解产物编码'].astype(str)
        df = df.fillna('')  # 填充空值
        
        # 转换为列表格式
        new_data = []
        for _, row in df.iterrows():
            code = str(row['拆解产物编码']).strip()
            if code:
                record = {}
                for col in required_columns:
                    if col in ['深加工投入产出比例', '深加工拆解系数']:
                        # 数值类型处理
                        try:
                            record[col] = float(row[col]) if pd.notna(row[col]) else 1.0
                        except (ValueError, TypeError):
                            record[col] = 1.0
                    else:
                        # 字符串类型处理
                        record[col] = str(row[col]).strip() if pd.notna(row[col]) else ''
                new_data.append(record)
        
        if not new_data:
            return jsonify({
                'success': False,
                'message': '没有有效的数据记录'
            }), 400
        
        # 获取原有数据统计
        from data.base_data.deep_processing_data import DEEP_PROCESSING_DATA
        original_count = len(DEEP_PROCESSING_DATA)
        
        # 根据导入模式处理数据
        if import_mode == 'replace':
            # 覆盖模式：替换所有数据
            DEEP_PROCESSING_DATA.clear()
            DEEP_PROCESSING_DATA.extend(new_data)
            message = f"已覆盖导入深加工数据: {len(new_data)} 条记录"
        else:
            # 追加模式：合并数据（根据拆解产物编码去重）
            existing_codes = {item['拆解产物编码'] for item in DEEP_PROCESSING_DATA}
            added_count = 0
            updated_count = 0
            
            for new_item in new_data:
                code = new_item['拆解产物编码']
                if code in existing_codes:
                    # 更新现有记录
                    for i, existing_item in enumerate(DEEP_PROCESSING_DATA):
                        if existing_item['拆解产物编码'] == code:
                            DEEP_PROCESSING_DATA[i] = new_item
                            updated_count += 1
                            break
                else:
                    # 添加新记录
                    DEEP_PROCESSING_DATA.append(new_item)
                    added_count += 1
            
            message = f"已追加导入深加工数据: 新增 {added_count} 条，更新 {updated_count} 条，总计 {len(DEEP_PROCESSING_DATA)} 条记录"
        
        # 更新文件
        from data.base_data.deep_processing_data import write_deep_processing_data_to_file
        write_deep_processing_data_to_file()
        print(f"✓ {message}")
        
        return jsonify({
            'success': True,
            'message': message,
            'statistics': {
                'original_count': original_count,
                'imported_count': len(new_data),
                'final_count': len(DEEP_PROCESSING_DATA),
                'import_mode': import_mode
            }
        })
        
    except Exception as e:
        print(f"✗ 导入深加工数据失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'导入失败: {str(e)}'
        }), 500

@data_management_bp.route('/mapping/export', methods=['GET'])
def export_mapping_data():
    """导出映射表数据为Excel"""
    try:
        df = get_mapping_dataframe()
        if df is None or df.empty:
            return jsonify({'success': False, 'error': '无法获取映射表数据'}), 500
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='映射表数据', index=False)
        
        output.seek(0)
        filename = f'映射表数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/deduction/export', methods=['GET'])
def export_deduction_data():
    """导出减扣规则数据为Excel"""
    try:
        # 将字典转换为DataFrame
        data_list = []
        for code, info in DEDUCTION_CODES.items():
            data_list.append({
                '编码': code,
                '说明': info.get('说明', ''),
                '处置类别': info.get('处置类别', '未分类'),
                '来源': info.get('来源', ''),
                '备注': ''
            })
        
        df = pd.DataFrame(data_list)
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='减扣规则数据', index=False)
        
        output.seek(0)
        filename = f'减扣规则数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

def write_deduction_data_to_file():
    """将减扣数据写入文件（格式化输出）"""
    try:
        deduction_file_path = os.path.abspath('data/base_data/deduction_data.py')
        
        # 生成格式化的文件内容
        content = '''# -*- coding: utf-8 -*-
"""
减扣规则数据
用于存储需要从计算结果中减扣的编码规则
"""

# 减扣规则字典
# 格式: {编码: {说明, 来源, 处置类别}}
DEDUCTION_CODES = {
'''
        
        # 添加每个减扣规则，格式化输出
        for code, info in sorted(DEDUCTION_CODES.items()):
            content += f"    '{code}': {{\n"
            content += f"        '说明': '{info['说明']}',\n"
            content += f"        '处置类别': '{info['处置类别']}',\n"
            content += f"        '来源': '{info['来源']}'\n"
            content += f"    }},\n"
        
        content += '''}

def get_deduction_codes():
    """获取所有减扣规则"""
    return DEDUCTION_CODES

def is_deduction_code(code):
    """检查是否为减扣代码"""
    return str(code) in DEDUCTION_CODES

def should_deduct(code):
    """检查编码是否需要减扣"""
    return str(code) in DEDUCTION_CODES

def get_deduction_info(code):
    """获取减扣代码信息"""
    return DEDUCTION_CODES.get(str(code))

def get_code_description(code):
    """获取编码说明"""
    info = DEDUCTION_CODES.get(str(code))
    return info['说明'] if info else None

def get_disposal_category(code):
    """获取处置类别"""
    info = DEDUCTION_CODES.get(str(code))
    return info['处置类别'] if info else None

def add_deduction_code(code, description, source, category):
    """添加减扣代码"""
    DEDUCTION_CODES[str(code)] = {
        '说明': description,
        '来源': source,
        '处置类别': category
    }

def remove_deduction_code(code):
    """删除减扣代码"""
    if str(code) in DEDUCTION_CODES:
        del DEDUCTION_CODES[str(code)]
        return True
    return False

def update_deduction_code(code, description=None, source=None, category=None):
    """更新减扣代码信息"""
    code_str = str(code)
    if code_str in DEDUCTION_CODES:
        if description is not None:
            DEDUCTION_CODES[code_str]['说明'] = description
        if source is not None:
            DEDUCTION_CODES[code_str]['来源'] = source
        if category is not None:
            DEDUCTION_CODES[code_str]['处置类别'] = category
        return True
    return False

def backup_deduction_file():
    """备份减扣数据文件"""
    try:
        import os
        import shutil
        from datetime import datetime
        
        # 获取当前文件的绝对路径
        current_file = os.path.abspath('data/base_data/deduction_data.py')
        
        # 创建备份目录（在项目根目录下）
        project_root = os.path.dirname(os.path.dirname(current_file))
        backup_dir = os.path.join(project_root, 'backups')
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        # 生成备份文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f'deduction_data_backup_{timestamp}.py'
        backup_path = os.path.join(backup_dir, backup_filename)
        
        # 复制当前文件到备份目录
        if os.path.exists(current_file):
            shutil.copy2(current_file, backup_path)
            print(f"✓ 已备份减扣数据文件: {backup_path}")
            return backup_path
        else:
            print("✗ 原文件不存在，无法备份")
            return None
            
    except Exception as e:
        print(f"✗ 备份减扣数据文件失败: {str(e)}")
        return None

def update_deduction_data_file():
    """更新减扣数据文件（简化版本以避免递归调用）"""
    try:
        print("✓ 减扣数据已在内存中更新，文件将在下次重启时生效")
        return True
    except Exception as e:
        print(f"✗ 更新减扣数据失败: {str(e)}")
        return False
'''
        
        # 写入文件
        with open(deduction_file_path, 'w', encoding='utf-8') as f:
            f.write(content)
        
        print(f"✓ 已将减扣数据写入文件: {deduction_file_path}")
        return True
        
    except Exception as e:
        print(f"✗ 写入减扣数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

@data_management_bp.route('/deduction/import', methods=['POST'])
def import_deduction_data():
    """导入减扣规则数据"""
    try:
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'message': '没有上传文件'
            }), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({
                'success': False,
                'message': '没有选择文件'
            }), 400
        
        import_mode = request.form.get('mode', 'append')  # append 或 replace
        
        # 备份减扣数据文件
        from data.base_data.deduction_data import backup_deduction_file
        backup_deduction_file()
        
        # 读取Excel文件
        df = pd.read_excel(file)
        
        # 验证必要的列 - 支持两种列名格式
        required_columns_v1 = ['R3代码', '说明', '来源', '处置类别']
        required_columns_v2 = ['编码', '说明', '来源', '处置类别']
        
        # 检查哪种格式的列名存在
        if all(col in df.columns for col in required_columns_v1):
            code_col = 'R3代码'
        elif all(col in df.columns for col in required_columns_v2):
            code_col = '编码'
        else:
            # 检查缺少哪些列
            missing_v1 = [col for col in required_columns_v1 if col not in df.columns]
            missing_v2 = [col for col in required_columns_v2 if col not in df.columns]
            return jsonify({
                'success': False,
                'message': f'缺少必要的列。支持的格式：{required_columns_v1} 或 {required_columns_v2}'
            }), 400
        
        # 数据验证和处理
        df = df.dropna(subset=[code_col])  # 删除代码为空的行
        df[code_col] = df[code_col].astype(str)
        df = df.fillna('')  # 填充空值
        
        # 转换为字典格式
        new_data = {}
        for _, row in df.iterrows():
            code = str(row[code_col]).strip()
            if code:
                new_data[code] = {
                    '说明': str(row['说明']).strip(),
                    '来源': str(row['来源']).strip(),
                    '处置类别': str(row['处置类别']).strip()
                }
        
        if not new_data:
            return jsonify({
                'success': False,
                'message': '没有有效的数据记录'
            }), 400
        
        # 根据导入模式处理数据
        if import_mode == 'replace':
            # 覆盖模式：替换所有数据
            DEDUCTION_CODES.clear()
            DEDUCTION_CODES.update(new_data)
            message = f"已覆盖导入减扣规则数据: {len(new_data)} 条记录"
        else:
            # 追加模式：合并数据
            original_count = len(DEDUCTION_CODES)
            DEDUCTION_CODES.update(new_data)
            new_count = len(DEDUCTION_CODES)
            added_count = new_count - original_count
            message = f"已追加导入减扣规则数据: {added_count} 条新记录，总计 {new_count} 条记录"
        
        # 更新文件
        from data.base_data.deduction_data import update_deduction_data_file
        # 实际写入文件而不是仅在内存中更新
        write_deduction_data_to_file()
        print(f"✓ {message}")
        
        return jsonify({
            'success': True,
            'message': message,
            'total_count': len(DEDUCTION_CODES)
        })
        
    except Exception as e:
        print(f"✗ 导入减扣规则数据失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'导入失败: {str(e)}'
        }), 500

@data_management_bp.route('/deduction/template', methods=['GET'])
def download_deduction_template():
    """下载减扣规则数据导入模板"""
    try:
        # 创建模板数据
        template_data = {
            '编码': ['示例代码1', '示例代码2'],
            '说明': ['示例说明1', '示例说明2'],
            '来源': ['一次拆解收发存及深加工投入产出.xlsx', '一次拆解收发存及深加工投入产出.xlsx'],
            '处置类别': ['示例类别1', '示例类别2'],
            '备注': ['请填写实际数据，删除示例行', '支持批量导入']
        }
        
        df = pd.DataFrame(template_data)
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='减扣规则模板', index=False)
        
        output.seek(0)
        filename = '减扣规则数据导入模板.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/deep-processing/export', methods=['GET'])
def export_deep_processing_data():
    """导出深加工数据为Excel"""
    try:
        dpd.reload_deep_processing_json_if_stale()
        df = dpd.get_deep_processing_dataframe()
        if df is None or df.empty:
            return jsonify({'success': False, 'error': '无法获取深加工数据'}), 500
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='深加工数据', index=False)
        
        output.seek(0)
        filename = f'深加工数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/product/export', methods=['GET'])
def export_product_data():
    """导出产品数据为Excel"""
    try:
        reload_product_disassembly_json_if_stale()
        # 将字典转换为DataFrame
        data_list = []
        for product_code, product_info in PRODUCT_DISASSEMBLY_DATA.items():
            # 获取产品基本信息
            unit_weight = product_info.get('单台重量', '')
            io_ratio = product_info.get('一次拆解投入产出比例', '')
            
            # 添加基本信息行
            data_list.append({
                '产品代码': product_code,
                '单台重量': unit_weight,
                '投入产出比例': io_ratio,
                '拆解产物编码': '',
                '拆解产物名称': '',
                '拆解系数': '',
                '记录类型': '基本信息'
            })
            
            # 遍历拆解系数明细，添加拆解明细行
            disassembly_details = product_info.get('拆解系数_明细', [])
            if isinstance(disassembly_details, list):
                for detail in disassembly_details:
                    # 处理拆解产物编码，移除.0后缀
                    detail_code = str(detail.get('一次拆解产物编码', ''))
                    if detail_code.endswith('.0'):
                        detail_code = detail_code[:-2]
                    
                    data_list.append({
                        '产品代码': product_code,
                        '单台重量': '',
                        '投入产出比例': '',
                        '拆解产物编码': detail_code,
                        '拆解产物名称': detail.get('一次拆解产物名称', ''),
                        '拆解系数': detail.get('一次拆解系数', ''),
                        '记录类型': '拆解明细'
                    })
        
        df = pd.DataFrame(data_list)
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 写入数据
            df.to_excel(writer, sheet_name='产品拆解系数', index=False)
            
            # 设置列宽，使Excel更易读
            worksheet = writer.sheets['产品拆解系数']
            worksheet.column_dimensions['A'].width = 15  # 产品代码
            worksheet.column_dimensions['B'].width = 15  # 单台重量
            worksheet.column_dimensions['C'].width = 18  # 投入产出比例
            worksheet.column_dimensions['D'].width = 20  # 拆解产物编码
            worksheet.column_dimensions['E'].width = 40  # 拆解产物名称
            worksheet.column_dimensions['F'].width = 18  # 拆解系数
            worksheet.column_dimensions['G'].width = 15  # 记录类型
        
        output.seek(0)
        filename = f'产品拆解系数_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500 

@data_management_bp.route('/product/import', methods=['POST'])
def import_product_data():
    """导入产品拆解系数数据"""
    try:
        # 检查文件
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'message': '没有上传文件'
            }), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({
                'success': False,
                'message': '没有选择文件'
            }), 400
        
        # 检查文件类型
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({
                'success': False,
                'message': '只支持Excel文件格式(.xlsx, .xls)'
            }), 400
        
        # 获取导入模式
        import_mode = request.form.get('mode', 'overwrite')
        
        # 读取Excel文件
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            file.save(tmp_file.name)
            
            try:
                df = pd.read_excel(tmp_file.name)
            except Exception as e:
                return jsonify({
                    'success': False,
                    'message': f'读取Excel文件失败: {str(e)}'
                }), 400
        
        # 验证必需列
        required_columns = ['产品代码', '单台重量', '投入产出比例', '拆解产物编码', '拆解产物名称', '拆解系数', '记录类型']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({
                'success': False,
                'message': f'Excel文件缺少必需列: {", ".join(missing_columns)}'
            }), 400
        
        # 导入数据
        from data.base_data.product_data import PRODUCT_DISASSEMBLY_DATA, backup_product_file
        
        # 备份原数据
        backup_product_file()
        
        # 处理数据
        total_records = len(df)  # 总记录数
        success_count = 0
        error_count = 0
        errors = []
        
        # 按产品代码分组处理
        grouped = df.groupby('产品代码')
        
        for product_code, group in grouped:
            try:
                product_code = str(product_code).strip()
                if not product_code or product_code == 'nan':
                    continue
                
                # 获取基本信息
                base_info = group[group['记录类型'] == '基本信息']
                if base_info.empty:
                    error_count += 1
                    errors.append(f'产品 {product_code} 缺少基本信息记录')
                    continue
                
                base_row = base_info.iloc[0]
                
                # 验证基本信息
                try:
                    unit_weight = float(base_row['单台重量'])
                    input_output_ratio = float(base_row['投入产出比例'])
                except (ValueError, TypeError):
                    error_count += 1
                    errors.append(f'产品 {product_code} 的重量或投入产出比例格式不正确')
                    continue
                
                # 获取拆解明细
                details = group[group['记录类型'] == '拆解明细']
                breakdown_details = []
                
                for _, detail_row in details.iterrows():
                    try:
                        breakdown_coef = float(detail_row['拆解系数'])
                        product_code_detail = str(detail_row['拆解产物编码']).strip()
                        product_name_detail = str(detail_row['拆解产物名称']).strip()
                        
                        if product_code_detail and product_name_detail:
                            # 确保编码格式一致（加.0后缀）
                            if not product_code_detail.endswith('.0'):
                                product_code_detail += '.0'
                            
                            breakdown_details.append({
                                "一次拆解系数": breakdown_coef,
                                "一次拆解产物编码": product_code_detail,
                                "一次拆解产物名称": product_name_detail
                            })
                    except (ValueError, TypeError):
                        error_count += 1
                        errors.append(f'产品 {product_code} 的拆解明细数据格式不正确')
                        continue
                
                # 构建产品数据
                product_data = {
                    "单台重量": unit_weight,
                    "一次拆解投入产出比例": input_output_ratio,
                    "拆解系数_明细": breakdown_details
                }
                
                # 检查是否已存在
                if import_mode == 'append' and product_code in PRODUCT_DISASSEMBLY_DATA:
                    error_count += 1
                    errors.append(f'产品 {product_code} 已存在，追加模式下已跳过')
                    continue
                
                # 更新数据
                PRODUCT_DISASSEMBLY_DATA[product_code] = product_data
                success_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append(f'处理产品 {product_code} 时出错: {str(e)}')
        
        if success_count > 0 and not save_product_disassembly_to_disk():
            return jsonify({
                'success': False,
                'message': '导入已写入内存，但保存 JSON 侧车失败',
            }), 500
        
        # 返回结果
        result = {
            'success': True,
            'message': f'导入完成，成功: {success_count}，失败: {error_count}',
            'imported_count': success_count,
            'total_count': total_records,
            'mode': import_mode,
            'details': {
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors[:10]  # 只返回前10个错误
            }
        }
        
        if error_count > 0:
            result['message'] += f'（前10个错误详情见details）'
            result['warnings'] = [f"行{err.get('row', '?')}: {err.get('error', '未知错误')}" for err in errors[:10]]
        
        return jsonify(result)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/load-extracted', methods=['POST'])
def load_extracted():
    """加载提取数据 - 临时端点处理遗留请求"""
    try:
        # 这个端点主要用于处理遗留的前端请求，避免404错误
        # 实际的数据加载功能已经整合到其他API中
        return jsonify({
            'success': False, 
            'message': '此功能已迁移，请使用其他数据加载方式',
            'note': '建议使用 /api/upload 或 /api/auto_process 进行数据处理'
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500 

@data_management_bp.route('/mapping', methods=['POST'])
def create_mapping_record():
    """创建新的映射表记录"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '请提供数据'}), 400
        
        # 验证必需字段
        required_fields = ['category', 'code', 'name']
        for field in required_fields:
            if not data.get(field, '').strip():
                return jsonify({'success': False, 'error': f'字段 {field} 不能为空'}), 400
        
        category = data['category'].strip()
        code = data['code'].strip()
        name = data['name'].strip()
        
        # 检查代码是否已存在
        df = get_mapping_dataframe()
        if df is None:
            return jsonify({'success': False, 'error': '无法获取映射表数据'}), 500
            
        if not df.empty and code in df['R3系统代码'].astype(str).values:
            return jsonify({'success': False, 'error': f'代码 {code} 已存在'}), 400
        
        # 备份文件
        backup_mapping_file()
        
        # 创建新记录
        new_record = {
            '类别': category,
            'R3系统代码': code,
            '系统名称': name
        }
        
        if add_mapping_record_to_file(new_record):
            # 清除所有会话的映射数据缓存，确保使用最新的映射数据
            _clear_mapping_cache_all_sessions()
            return jsonify({'success': True, 'message': '记录创建成功'})
        else:
            return jsonify({'success': False, 'error': '创建记录失败'}), 500
        
    except Exception as e:
        print(f"创建映射记录错误: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/mapping/<record_code>', methods=['PUT'])
def update_mapping_record(record_code):
    """更新映射表记录"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '请提供数据'}), 400
        
        # 验证必需字段
        required_fields = ['category', 'code', 'name']
        for field in required_fields:
            if not data.get(field, '').strip():
                return jsonify({'success': False, 'error': f'字段 {field} 不能为空'}), 400
        
        category = data['category'].strip()
        new_code = data['code'].strip()
        name = data['name'].strip()
        
        # 检查记录是否存在
        df = get_mapping_dataframe()
        if df is None or df.empty:
            return jsonify({'success': False, 'error': '无法获取映射表数据'}), 500
            
        old_record_mask = df['R3系统代码'].astype(str) == record_code
        if not old_record_mask.any():
            return jsonify({'success': False, 'error': '记录不存在'}), 404
        
        # 如果代码改变了，检查新代码是否已存在
        if new_code != record_code and new_code in df['R3系统代码'].astype(str).values:
            return jsonify({'success': False, 'error': f'代码 {new_code} 已存在'}), 400
        
        # 备份文件
        backup_mapping_file()
        
        # 更新记录
        updated_record = {
            '类别': category,
            'R3系统代码': new_code,
            '系统名称': name
        }
        
        if update_mapping_record_in_file(record_code, updated_record):
            # 清除所有会话的映射数据缓存，确保使用最新的映射数据
            _clear_mapping_cache_all_sessions()
            return jsonify({'success': True, 'message': '记录更新成功'})
        else:
            return jsonify({'success': False, 'error': '更新记录失败'}), 500
        
    except Exception as e:
        print(f"更新映射记录错误: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/mapping/<record_code>', methods=['DELETE'])
def delete_mapping_record(record_code):
    """删除映射表记录"""
    try:
        # 检查记录是否存在
        df = get_mapping_dataframe()
        if df is None or df.empty:
            return jsonify({'success': False, 'error': '无法获取映射表数据'}), 500
            
        record_mask = df['R3系统代码'].astype(str) == record_code
        if not record_mask.any():
            return jsonify({'success': False, 'error': '记录不存在'}), 404
        
        # 备份文件
        backup_mapping_file()
        
        # 删除记录
        if delete_mapping_record_from_file(record_code):
            # 清除所有会话的映射数据缓存，确保使用最新的映射数据
            _clear_mapping_cache_all_sessions()
            return jsonify({'success': True, 'message': '记录删除成功'})
        else:
            return jsonify({'success': False, 'error': '删除记录失败'}), 500
        
    except Exception as e:
        print(f"删除映射记录错误: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

def backup_mapping_file():
    """备份映射表文件"""
    try:
        import shutil
        from datetime import datetime
        
        # 创建备份目录
        backup_dir = 'backups'
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        # 生成备份文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"mapping_data_backup_{timestamp}.py"
        backup_path = os.path.join(backup_dir, backup_filename)
        
        # 复制文件
        source_path = 'data/base_data/mapping_data.py'
        shutil.copy2(source_path, backup_path)
        print(f"✓ 映射表文件已备份到: {backup_path}")
        return True
        
    except Exception as e:
        print(f"✗ 备份映射表文件失败: {e}")
        return False

def delete_mapping_record_from_file(code):
    """从映射表文件中删除记录"""
    try:
        import re
        import importlib
        
        # 读取当前文件内容
        mapping_file_path = 'data/base_data/mapping_data.py'
        with open(mapping_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 找到并删除对应的记录
        # 匹配包含指定代码的完整记录行
        pattern = rf'    \{{"类别":\s*"[^"]*",\s*"R3系统代码":\s*"{re.escape(code)}",\s*"系统名称":\s*"[^"]*"\}},?\n'
        new_content = re.sub(pattern, '', content)
        
        # 检查是否找到并删除了记录
        if new_content == content:
            raise ValueError(f"未找到R3代码为 {code} 的记录")
        
        # 写回文件
        with open(mapping_file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        # 重新加载模块
        from data.base_data import mapping_data
        importlib.reload(mapping_data)
        
        print(f"✓ 已删除映射表记录: {code}")
        return True
        
    except Exception as e:
        print(f"✗ 删除映射表记录失败: {e}")
        return False 

def update_mapping_record_in_file(old_code, new_record):
    """更新映射表文件中的记录"""
    try:
        import re
        import importlib
        
        # 读取当前文件内容
        mapping_file_path = 'data/base_data/mapping_data.py'
        with open(mapping_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 转义特殊字符
        category = str(new_record["类别"]).replace('"', '\\"')
        code = str(new_record["R3系统代码"]).replace('"', '\\"')
        name = str(new_record["系统名称"]).replace('"', '\\"')
        
        # 找到并替换对应的记录
        old_pattern = rf'    \{{"类别":\s*"[^"]*",\s*"R3系统代码":\s*"{re.escape(old_code)}",\s*"系统名称":\s*"[^"]*"\}},?\n'
        new_line = f'    {{"类别": "{category}", "R3系统代码": "{code}", "系统名称": "{name}"}},\n'
        
        new_content = re.sub(old_pattern, new_line, content)
        
        # 检查是否找到并替换了记录
        if new_content == content:
            raise ValueError(f"未找到R3代码为 {old_code} 的记录")
        
        # 写回文件
        with open(mapping_file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        # 重新加载模块
        from data.base_data import mapping_data
        importlib.reload(mapping_data)
        
        print(f"✓ 已更新映射表记录: {old_code} -> {new_record['R3系统代码']}")
        return True
        
    except Exception as e:
        print(f"✗ 更新映射表记录失败: {e}")
        return False

def add_mapping_record_to_file(new_record):
    """将新的映射表记录添加到文件中"""
    try:
        import re
        import importlib
        
        # 读取当前文件内容
        mapping_file_path = 'data/base_data/mapping_data.py'
        with open(mapping_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 转义特殊字符
        category = str(new_record["类别"]).replace('"', '\\"')
        code = str(new_record["R3系统代码"]).replace('"', '\\"')
        name = str(new_record["系统名称"]).replace('"', '\\"')
        
        # 创建新记录行
        new_line = f'    {{"类别": "{category}", "R3系统代码": "{code}", "系统名称": "{name}"}},\n'
        
        # 找到列表的结束位置（最后一个 ] 之前）
        # 先找到最后一个记录，在它后面插入新记录
        pattern = r'(\s*\{"类别":[^}]+\},?\n)(\s*\])'
        
        def replace_func(match):
            last_record = match.group(1)
            closing_bracket = match.group(2)
            
            # 确保最后一个记录有逗号
            if not last_record.rstrip().endswith(','):
                last_record = last_record.rstrip() + ',\n'
            
            return last_record + new_line + closing_bracket
        
        new_content = re.sub(pattern, replace_func, content, flags=re.MULTILINE | re.DOTALL)
        
        # 检查是否成功替换
        if new_content == content:
            raise ValueError("无法找到合适的位置插入新记录")
        
        # 写回文件
        with open(mapping_file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        # 重新加载模块
        from data.base_data import mapping_data
        importlib.reload(mapping_data)
        
        print(f"✓ 已添加映射表记录: {new_record['R3系统代码']}")
        return True
        
    except Exception as e:
        print(f"✗ 添加映射表记录失败: {e}")
        return False 

@data_management_bp.route('/mapping/import', methods=['POST'])
def import_mapping_data():
    """导入映射表数据"""
    try:
        # 检查文件
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '请选择要导入的Excel文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '请选择要导入的Excel文件'}), 400
        
        # 检查文件格式
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': '只支持Excel文件格式(.xlsx, .xls)'}), 400
        
        # 获取导入模式
        import_mode = request.form.get('import_mode', 'append')
        
        # 获取原有数据统计
        original_df = get_mapping_dataframe()
        original_count = len(original_df) if original_df is not None else 0
        
        # 读取Excel文件
        try:
            df = pd.read_excel(file, engine='openpyxl' if file.filename.endswith('.xlsx') else 'xlrd')
        except Exception as e:
            return jsonify({'success': False, 'error': f'文件读取失败: {str(e)}'}), 400
        
        # 验证数据
        required_columns = ['类别', 'R3系统代码', '系统名称']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return jsonify({
                'success': False, 
                'error': f'Excel文件缺少必要的列: {", ".join(missing_columns)}'
            }), 400
        
        # 清理数据
        df = df[required_columns].copy()
        df = df.dropna(subset=required_columns)  # 删除关键字段为空的行
        
        # 数据验证
        validation_errors = []
        for index, row in df.iterrows():
            row_num = index + 2  # Excel行号（从2开始，因为有标题行）
            
            # 检查必填字段
            if not str(row['类别']).strip():
                validation_errors.append(f'第{row_num}行：类别不能为空')
            if not str(row['R3系统代码']).strip():
                validation_errors.append(f'第{row_num}行：R3系统代码不能为空')
            if not str(row['系统名称']).strip():
                validation_errors.append(f'第{row_num}行：系统名称不能为空')
        
        if validation_errors:
            return jsonify({
                'success': False,
                'error': '数据验证失败',
                'details': validation_errors[:10]  # 最多显示10个错误
            }), 400
        
        # 去重处理（基于R3系统代码）
        df = df.drop_duplicates(subset=['R3系统代码'], keep='first')
        
        if len(df) == 0:
            return jsonify({'success': False, 'error': '没有有效的数据可以导入'}), 400
        
        # 备份现有数据
        backup_mapping_file()
        
        # 执行导入
        try:
            if import_mode == 'replace':
                # 覆盖模式：清空现有数据，重新写入
                import_mapping_data_to_file(df.to_dict('records'), mode='replace')
                
                # 重新加载mapping_data模块以获取最新数据
                import importlib
                from data.base_data import mapping_data
                importlib.reload(mapping_data)
                
                message = f'数据覆盖导入成功！共导入 {len(df)} 条记录'
            else:
                # 追加模式：合并数据
                existing_codes = set(original_df['R3系统代码'].astype(str).tolist()) if original_df is not None else set()
                new_records = []
                updated_count = 0
                
                for _, row in df.iterrows():
                    if str(row['R3系统代码']) in existing_codes:
                        updated_count += 1
                    new_records.append(row.to_dict())
                
                import_mapping_data_to_file(new_records, mode='append')
                
                # 重新加载mapping_data模块以获取最新数据
                import importlib
                from data.base_data import mapping_data
                importlib.reload(mapping_data)
                
                new_count = len(df) - updated_count
                message = f'数据追加导入成功！新增 {new_count} 条记录，更新 {updated_count} 条记录'
            
            # 获取最终统计（使用重新加载后的数据）
            from data.base_data.mapping_data import get_mapping_dataframe as get_fresh_mapping_dataframe
            final_df = get_fresh_mapping_dataframe()
            final_count = len(final_df) if final_df is not None else 0
            
            # 清除所有会话的映射数据缓存，确保使用最新的映射数据
            _clear_mapping_cache_all_sessions()
            
            # 返回成功响应
            return jsonify({
                'success': True,
                'message': message,
                'statistics': {
                    'original_count': original_count,
                    'imported_count': len(df),
                    'final_count': final_count,
                    'import_mode': '覆盖' if import_mode == 'replace' else '追加'
                }
            })
            
        except Exception as e:
            return jsonify({'success': False, 'error': f'数据导入失败: {str(e)}'}), 500
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'导入失败: {str(e)}'}), 500

def import_mapping_data_to_file(records, mode='append'):
    """批量导入映射表数据到文件"""
    try:
        if mode == 'replace':
            # 覆盖模式：重新生成整个文件
            new_content = '''# 电废拆解产物映射表数据 - 内置数据
# 数据来源：电废拆解产物映射表(含旧机).xlsx

import pandas as pd

# 完整的映射表数据
MAPPING_TABLE_DATA = [
'''
            
            # 添加所有记录
            for record in records:
                # 转义特殊字符
                category = str(record['类别']).replace('"', '\\"')
                code = str(record['R3系统代码']).replace('"', '\\"')
                name = str(record['系统名称']).replace('"', '\\"')
                new_content += f'    {{"类别": "{category}", "R3系统代码": "{code}", "系统名称": "{name}"}},\n'
            
            new_content += ''']

def get_mapping_dataframe():
    """获取映射表DataFrame"""
    return pd.DataFrame(MAPPING_TABLE_DATA)

def filter_by_category(category):
    """根据类别筛选映射数据"""
    df = get_mapping_dataframe()
    return df[df['类别'] == category]

def get_all_categories():
    """获取所有类别列表"""
    df = get_mapping_dataframe()
    return df['类别'].unique().tolist()

def get_category_stats():
    """获取类别统计信息"""
    df = get_mapping_dataframe()
    return df['类别'].value_counts().to_dict()

# 常用的预定义数据集
def get_old_machine_mapping():
    """获取旧机映射数据"""
    return filter_by_category('旧机')

def get_disassembly_product_mapping():
    """获取一次拆解产物映射数据"""
    return filter_by_category('一次拆解产物')

def get_breaking_mapping():
    """获取一破映射数据"""
    return filter_by_category('一破')

def get_packing_iron_mapping():
    """获取打包铁映射数据"""
    return filter_by_category('打包铁')
'''
            
            # 写入文件
            mapping_file_path = 'data/base_data/mapping_data.py'
            with open(mapping_file_path, 'w', encoding='utf-8') as f:
                f.write(new_content)
            
            print(f"✓ 已覆盖导入映射表数据: {len(records)} 条记录")
            
        else:
            # 追加模式：读取现有数据，合并后重写
            current_df = get_mapping_dataframe()
            existing_codes = set(current_df['R3系统代码'].astype(str).tolist()) if current_df is not None else set()
            
            # 获取所有现有记录
            all_records = current_df.to_dict('records') if current_df is not None else []
            
            # 添加新记录或更新现有记录
            for record in records:
                code = str(record['R3系统代码'])
                if code in existing_codes:
                    # 更新现有记录
                    for i, existing_record in enumerate(all_records):
                        if str(existing_record['R3系统代码']) == code:
                            all_records[i] = record
                            break
                else:
                    # 添加新记录
                    all_records.append(record)
            
            # 使用覆盖模式重写整个文件
            import_mapping_data_to_file(all_records, mode='replace')
            print(f"✓ 已追加导入映射表数据: {len(records)} 条记录")
        
        return True
        
    except Exception as e:
        print(f"✗ 导入映射表数据失败: {e}")
        return False

def backup_deduction_file():
    """备份减扣数据文件"""
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f"backups/deduction_data_backup_{timestamp}.py"
        deduction_file_path = 'data/base_data/deduction_data.py'
        
        # 确保备份目录存在
        os.makedirs('backups', exist_ok=True)
        
        # 复制文件
        import shutil
        shutil.copy2(deduction_file_path, backup_filename)
        print(f"✓ 减扣数据文件已备份到: {backup_filename}")
        
    except Exception as e:
        print(f"✗ 备份减扣数据文件失败: {e}")

def delete_deduction_code_from_file(deduction_code):
    """从减扣数据文件中删除指定编码"""
    try:
        # 从内存中删除记录
        if deduction_code in DEDUCTION_CODES:
            del DEDUCTION_CODES[deduction_code]
            print(f"已从内存中删除编码: {deduction_code}")
        else:
            print(f"编码 {deduction_code} 在内存中不存在")
            return False
        
        # 将更新后的数据写入文件
        write_deduction_data_to_file()
        
        print(f"删除后剩余记录数: {len(DEDUCTION_CODES)}")
        return True
        
    except Exception as e:
        print(f"✗ 删除减扣编码失败: {e}")
        import traceback
        traceback.print_exc()
        return False 

def update_deduction_code_in_file(deduction_code, updated_record):
    """更新减扣数据文件中的记录"""
    try:
        # 检查记录是否存在
        if deduction_code not in DEDUCTION_CODES:
            raise ValueError(f"编码 {deduction_code} 不存在")
        
        # 在内存中更新记录
        DEDUCTION_CODES[deduction_code] = {
            '说明': updated_record['说明'],
            '处置类别': updated_record['处置类别'],
            '来源': updated_record['来源']
        }
        print(f"已在内存中更新编码: {deduction_code}")
        
        # 将更新后的数据写入文件
        write_deduction_data_to_file()
        
        print(f"✓ 已更新减扣规则: {deduction_code}")
        return True
        
    except Exception as e:
        print(f"✗ 更新减扣规则失败: {e}")
        import traceback
        traceback.print_exc()
        return False 

def add_deduction_code_to_file(deduction_code, new_record):
    """添加新的减扣规则到文件中"""
    try:
        # 检查编码是否已存在
        if deduction_code in DEDUCTION_CODES:
            raise ValueError(f"编码 {deduction_code} 已存在")
        
        # 在内存中添加新记录
        DEDUCTION_CODES[deduction_code] = {
            '说明': new_record['说明'],
            '处置类别': new_record['处置类别'],
            '来源': new_record['来源']
        }
        print(f"已在内存中添加编码: {deduction_code}")
        
        # 将更新后的数据写入文件
        write_deduction_data_to_file()
        
        print(f"✓ 已添加减扣规则: {deduction_code}")
        return True
        
    except Exception as e:
        print(f"✗ 添加减扣规则失败: {e}")
        import traceback
        traceback.print_exc()
        return False

@data_management_bp.route('/deduction', methods=['POST'])
def create_deduction_rule():
    """新增减扣规则"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '请提供数据'}), 400
        
        # 验证必需字段
        deduction_code = data.get('编码', '').strip()
        description = data.get('说明', '').strip()
        disposal_category = data.get('处置类别', '').strip()
        source = data.get('来源', '').strip()
        
        if not deduction_code:
            return jsonify({'success': False, 'error': '编码不能为空'}), 400
        if not description:
            return jsonify({'success': False, 'error': '说明不能为空'}), 400
        if not disposal_category:
            return jsonify({'success': False, 'error': '处置类别不能为空'}), 400
        
        # 检查编码是否已存在
        if deduction_code in DEDUCTION_CODES:
            return jsonify({'success': False, 'error': '编码已存在'}), 400
        
        # 备份减扣数据文件
        backup_deduction_file()
        
        # 新增记录
        new_record = {
            '说明': description,
            '处置类别': disposal_category,
            '来源': source or '一次拆解收发存及深加工投入产出.xlsx'
        }
        
        if add_deduction_code_to_file(deduction_code, new_record):
            print(f"✓ 已新增减扣规则: {deduction_code}")
            return jsonify({'success': True, 'message': '减扣规则新增成功'})
        else:
            return jsonify({'success': False, 'error': '新增减扣规则失败'}), 500
        
    except Exception as e:
        print(f"新增减扣规则错误: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500 

@data_management_bp.route('/product/template', methods=['GET'])
def download_product_template():
    """下载产品数据导入模板"""
    try:
        # 创建示例数据
        template_data = [
            {
                '产品代码': '810978870',
                '单台重量': 22.26,
                '投入产出比例': 0.9899,
                '拆解产物编码': '',
                '拆解产物名称': '',
                '拆解系数': '',
                '记录类型': '基本信息'
            },
            {
                '产品代码': '810978870',
                '单台重量': '',
                '投入产出比例': '',
                '拆解产物编码': '811052760',
                '拆解产物名称': '洗衣机-其他垃圾',
                '拆解系数': 0.00692453,
                '记录类型': '拆解明细'
            },
            {
                '产品代码': '810978870',
                '单台重量': '',
                '投入产出比例': '',
                '拆解产物编码': '811052955',
                '拆解产物名称': '铁及其合金洗衣机-铁外壳、内桶',
                '拆解系数': 0.02299160,
                '记录类型': '拆解明细'
            },
            {
                '产品代码': '811132341',
                '单台重量': 15.5,
                '投入产出比例': 0.995,
                '拆解产物编码': '',
                '拆解产物名称': '',
                '拆解系数': '',
                '记录类型': '基本信息'
            },
            {
                '产品代码': '811132341',
                '单台重量': '',
                '投入产出比例': '',
                '拆解产物编码': '811052763',
                '拆解产物名称': '塑料空调-塑料外壳',
                '拆解系数': 0.45621,
                '记录类型': '拆解明细'
            }
        ]
        
        df = pd.DataFrame(template_data)
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 添加数据工作表
            df.to_excel(writer, sheet_name='产品拆解系数', index=False)
            
            # 添加说明工作表
            instructions = [
                ['说明', '内容'],
                ['文件格式', 'Excel文件(.xlsx, .xls)'],
                ['必需列', '产品代码、单台重量、投入产出比例、拆解产物编码、拆解产物名称、拆解系数、记录类型'],
                ['记录类型', '基本信息：每个产品必须有一行基本信息记录，包含产品代码、单台重量、投入产出比例'],
                ['', '拆解明细：每个产品可以有多行拆解明细记录，包含拆解产物编码、拆解产物名称、拆解系数'],
                ['导入模式', '覆盖模式：清空现有数据，只保留导入数据'],
                ['', '追加模式：重复数据会更新，新数据会添加'],
                ['数据要求', '每个产品需要一行基本信息记录和多行拆解明细记录'],
                ['', '产品代码不能为空'],
                ['', '单台重量和投入产出比例必须为数字'],
                ['', '拆解系数必须为数字'],
                ['示例', '参考"产品拆解系数"工作表中的示例数据']
            ]
            
            instructions_df = pd.DataFrame(instructions)
            instructions_df.to_excel(writer, sheet_name='导入说明', index=False, header=False)
            
            # 美化工作表
            workbook = writer.book
            
            # 设置产品数据工作表格式
            product_ws = writer.sheets['产品拆解系数']
            for column in product_ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                product_ws.column_dimensions[column_letter].width = adjusted_width
            
            # 设置说明工作表格式
            instruction_ws = writer.sheets['导入说明']
            instruction_ws.column_dimensions['A'].width = 15
            instruction_ws.column_dimensions['B'].width = 50
        
        output.seek(0)
        filename = f'产品拆解系数导入模板_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/deducted-data/initialize-manual', methods=['POST'])
def initialize_manual_from_readonly():
    """从只读数据初始化手工数据（参考可销售量模式）"""
    try:
        from datetime import datetime

        app_data = get_session_data_manager()

        # 直接从 deducted_data（系统最新计算结果）复制，不再依赖旧快照
        deducted_data = app_data.get_data('deducted_data')
        if deducted_data is None or (hasattr(deducted_data, 'empty') and deducted_data.empty):
            return jsonify({
                'success': False,
                'error': '没有可用的被减扣数据，请先进行自动计算'
            }), 400

        manual_data = deducted_data.copy()

        app_data.set_data('deducted_data_manual', manual_data)
        app_data.set_data('deducted_data_modified', True)
        app_data.set_data('modification_timestamp', datetime.now().isoformat())

        app_data.save_persistent_data()

        print(f"手工数据初始化完成: 从 deducted_data 复制了 {len(manual_data)} 条记录")

        return jsonify({
            'success': True,
            'message': '手工数据初始化成功',
            'copied_count': len(manual_data)
        })
        
    except Exception as e:
        print(f"初始化手工数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'初始化失败: {str(e)}'
        }), 500

def _clear_mapping_cache_all_sessions():
    """清除所有会话的映射数据缓存"""
    try:
        # 获取所有活跃的会话管理器
        from app.models.session_data_manager import SessionDataManagerFactory
        
        # 获取所有会话实例
        active_sessions = SessionDataManagerFactory.get_all_sessions()
        
        # 清除每个会话的映射数据缓存
        for session_id, session_manager in active_sessions.items():
            try:
                # 清除映射数据缓存
                session_manager.clear_data('mapping_data')
                # 移除单个会话的频繁日志
            except Exception as e:
                print(f"⚠️ 清除会话 {session_id} 映射缓存失败: {e}")
        
        print(f"✅ 共清除了 {len(active_sessions)} 个会话的映射数据缓存")
        
    except Exception as e:
        print(f"⚠️ 清除映射缓存失败: {e}")
        import traceback
        traceback.print_exc()


# ==================== 可销售量数据管理 API ====================

@data_management_bp.route('/saleable-data', methods=['GET'])
def get_saleable_data():
    """获取可销售量数据（系统计算版本）"""
    try:
        app_data = get_session_data_manager()
        
        # 获取系统计算的可销售量数据
        saleable_data = app_data.get_data('saleable_data')
        
        if saleable_data is None or saleable_data.empty:
            return jsonify({
                'success': True,
                'data': [],
                'message': '暂无可销售量数据，请先完成数据处理流程'
            })
        
        # 检查是否需要匹配价格信息
        need_price_update = False
        if hasattr(saleable_data, 'columns'):
            # 检查价格列是否存在或为0
            if '销售单价(元/KG)' not in saleable_data.columns or '销售单价-不含税(元/KG)' not in saleable_data.columns:
                need_price_update = True
            elif '拆解产物编码' in saleable_data.columns:
                # 检查是否有价格列为0的情况
                price_with_tax_zero = (saleable_data['销售单价(元/KG)'] == 0).sum() if '销售单价(元/KG)' in saleable_data.columns else len(saleable_data)
                price_no_tax_zero = (saleable_data['销售单价-不含税(元/KG)'] == 0).sum() if '销售单价-不含税(元/KG)' in saleable_data.columns else len(saleable_data)
                # 如果大部分价格都是0，需要更新
                if price_with_tax_zero > len(saleable_data) * 0.5 or price_no_tax_zero > len(saleable_data) * 0.5:
                    need_price_update = True
        
        # 如果需要更新价格，从价格数据中匹配
        if need_price_update and hasattr(saleable_data, 'columns') and '拆解产物编码' in saleable_data.columns:
            print("检测到系统数据价格信息缺失或为0，开始匹配价格...")
            from data.base_data.price_data import load_price_data, get_price_mapping
            import pandas as pd
            
            # 获取价格数据
            price_df = load_price_data()
            price_mapping_no_tax = get_price_mapping()
            
            if price_df is not None and not price_df.empty and price_mapping_no_tax:
                # 创建含税价映射
                price_mapping_with_tax = {}
                for _, row in price_df.iterrows():
                    code = str(row['拆解产物编码']).strip()
                    price_with_tax = row.get('销售单价(元/KG)', 0)
                    if pd.notna(price_with_tax):
                        price_mapping_with_tax[code] = float(price_with_tax)
                
                # 更新或添加含税价列
                if '销售单价(元/KG)' not in saleable_data.columns:
                    saleable_data['销售单价(元/KG)'] = 0.0
                
                # 更新含税价：根据拆解产物编码匹配
                def update_price_with_tax(row):
                    code = str(row['拆解产物编码']).strip() if pd.notna(row['拆解产物编码']) else ''
                    matched_price = price_mapping_with_tax.get(code, 0.0)
                    # 如果当前价格不为0，保留原值；否则使用匹配的价格
                    current_price = row.get('销售单价(元/KG)', 0.0)
                    return matched_price if (matched_price > 0 and current_price == 0) else current_price
                
                saleable_data['销售单价(元/KG)'] = saleable_data.apply(update_price_with_tax, axis=1)
                
                # 更新或添加不含税价列
                if '销售单价-不含税(元/KG)' not in saleable_data.columns:
                    saleable_data['销售单价-不含税(元/KG)'] = 0.0
                
                # 更新不含税价：根据拆解产物编码匹配
                def update_price_no_tax(row):
                    code = str(row['拆解产物编码']).strip() if pd.notna(row['拆解产物编码']) else ''
                    matched_price = price_mapping_no_tax.get(code, 0.0)
                    # 如果当前价格不为0，保留原值；否则使用匹配的价格
                    current_price = row.get('销售单价-不含税(元/KG)', 0.0)
                    return matched_price if (matched_price > 0 and current_price == 0) else current_price
                
                saleable_data['销售单价-不含税(元/KG)'] = saleable_data.apply(update_price_no_tax, axis=1)
                
                # 重新计算销售收益（使用不含税价）
                if '计算结果(KG)' in saleable_data.columns:
                    if '销售收益(元)' not in saleable_data.columns:
                        saleable_data['销售收益(元)'] = 0
                    saleable_data['销售收益(元)'] = saleable_data.apply(
                        lambda row: round(float(row['计算结果(KG)']) * float(row['销售单价-不含税(元/KG)']), 2)
                        if pd.notna(row['计算结果(KG)']) and pd.notna(row['销售单价-不含税(元/KG)']) and float(row['计算结果(KG)']) > 0 and float(row['销售单价-不含税(元/KG)']) != 0 else 0,
                        axis=1
                    )
                
                # 保存更新后的数据
                app_data.set_data('saleable_data', saleable_data)
                app_data.save_persistent_data()
                
                matched_count = (saleable_data['销售单价-不含税(元/KG)'] > 0).sum() if '销售单价-不含税(元/KG)' in saleable_data.columns else 0
                print(f"系统数据价格匹配完成: 成功匹配 {matched_count} 条记录的价格信息")
        
        # 转换为字典格式
        data_dict = safe_json_convert(saleable_data)
        
        return jsonify({
            'success': True,
            'data': data_dict,
            'count': len(data_dict),
            'message': '可销售量数据获取成功'
        })
        
    except Exception as e:
        print(f"获取可销售量数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'获取数据失败: {str(e)}'
        }), 500

@data_management_bp.route('/saleable-data-manual', methods=['GET'])
def get_saleable_data_manual():
    """获取可销售量数据（手工编辑版本）"""
    try:
        app_data = get_session_data_manager()
        
        # 获取手工编辑的可销售量数据
        manual_data = app_data.get_data('saleable_data_manual')
        
        if manual_data is None or (hasattr(manual_data, 'empty') and manual_data.empty):
            return jsonify({
                'success': True,
                'data': [],
                'message': '暂无手工编辑数据，可从原始数据初始化'
            })
        
        # 检查是否需要匹配价格信息
        need_price_update = False
        if hasattr(manual_data, 'columns'):
            # 检查价格列是否存在或为0
            if '销售单价(元/KG)' not in manual_data.columns or '销售单价-不含税(元/KG)' not in manual_data.columns:
                need_price_update = True
            elif '拆解产物编码' in manual_data.columns:
                # 检查是否有价格列为0的情况
                price_with_tax_zero = (manual_data['销售单价(元/KG)'] == 0).sum() if '销售单价(元/KG)' in manual_data.columns else len(manual_data)
                price_no_tax_zero = (manual_data['销售单价-不含税(元/KG)'] == 0).sum() if '销售单价-不含税(元/KG)' in manual_data.columns else len(manual_data)
                # 如果大部分价格都是0，需要更新
                if price_with_tax_zero > len(manual_data) * 0.5 or price_no_tax_zero > len(manual_data) * 0.5:
                    need_price_update = True
        
        # 如果需要更新价格，从价格数据中匹配
        if need_price_update and hasattr(manual_data, 'columns') and '拆解产物编码' in manual_data.columns:
            print("检测到价格信息缺失或为0，开始匹配价格...")
            from data.base_data.price_data import load_price_data, get_price_mapping
            import pandas as pd
            
            # 获取价格数据
            price_df = load_price_data()
            price_mapping_no_tax = get_price_mapping()
            
            if price_df is not None and not price_df.empty and price_mapping_no_tax:
                # 创建含税价映射
                price_mapping_with_tax = {}
                for _, row in price_df.iterrows():
                    code = str(row['拆解产物编码']).strip()
                    price_with_tax = row.get('销售单价(元/KG)', 0)
                    if pd.notna(price_with_tax):
                        price_mapping_with_tax[code] = float(price_with_tax)
                
                # 更新或添加含税价列
                if '销售单价(元/KG)' not in manual_data.columns:
                    manual_data['销售单价(元/KG)'] = 0.0
                
                # 更新含税价：根据拆解产物编码匹配
                def update_price_with_tax(row):
                    code = str(row['拆解产物编码']).strip() if pd.notna(row['拆解产物编码']) else ''
                    matched_price = price_mapping_with_tax.get(code, 0.0)
                    # 如果当前价格不为0，保留原值；否则使用匹配的价格
                    current_price = row.get('销售单价(元/KG)', 0.0)
                    return matched_price if (matched_price > 0 and current_price == 0) else current_price
                
                manual_data['销售单价(元/KG)'] = manual_data.apply(update_price_with_tax, axis=1)
                
                # 更新或添加不含税价列
                if '销售单价-不含税(元/KG)' not in manual_data.columns:
                    manual_data['销售单价-不含税(元/KG)'] = 0.0
                
                # 更新不含税价：根据拆解产物编码匹配
                def update_price_no_tax(row):
                    code = str(row['拆解产物编码']).strip() if pd.notna(row['拆解产物编码']) else ''
                    matched_price = price_mapping_no_tax.get(code, 0.0)
                    # 如果当前价格不为0，保留原值；否则使用匹配的价格
                    current_price = row.get('销售单价-不含税(元/KG)', 0.0)
                    return matched_price if (matched_price > 0 and current_price == 0) else current_price
                
                manual_data['销售单价-不含税(元/KG)'] = manual_data.apply(update_price_no_tax, axis=1)
                
                # 重新计算销售收益（使用不含税价）
                if '计算结果(KG)' in manual_data.columns:
                    if '销售收益(元)' not in manual_data.columns:
                        manual_data['销售收益(元)'] = 0
                    manual_data['销售收益(元)'] = manual_data.apply(
                        lambda row: round(float(row['计算结果(KG)']) * float(row['销售单价-不含税(元/KG)']), 2)
                        if pd.notna(row['计算结果(KG)']) and pd.notna(row['销售单价-不含税(元/KG)']) and float(row['计算结果(KG)']) > 0 and float(row['销售单价-不含税(元/KG)']) != 0 else 0,
                        axis=1
                    )
                
                # 保存更新后的数据
                app_data.set_data('saleable_data_manual', manual_data)
                app_data.save_persistent_data()
                
                matched_count = (manual_data['销售单价-不含税(元/KG)'] > 0).sum() if '销售单价-不含税(元/KG)' in manual_data.columns else 0
                print(f"价格匹配完成: 成功匹配 {matched_count} 条记录的价格信息")
        
        # 如果是DataFrame，转换为字典
        if hasattr(manual_data, 'to_dict'):
            data_dict = safe_json_convert(manual_data)
        else:
            data_dict = manual_data if isinstance(manual_data, list) else []
        
        return jsonify({
            'success': True,
            'data': data_dict,
            'count': len(data_dict),
            'message': '手工编辑数据获取成功'
        })
        
    except Exception as e:
        print(f"获取手工可销售量数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'获取数据失败: {str(e)}'
        }), 500

@data_management_bp.route('/saleable-data', methods=['PUT'])
def update_saleable_data():
    """更新可销售量数据"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '没有接收到数据'}), 400
        
        app_data = get_session_data_manager()
        
        # 获取当前手工数据，如果没有则从原始数据初始化
        manual_data = app_data.get_data('saleable_data_manual')
        if manual_data is None or (hasattr(manual_data, 'empty') and manual_data.empty):
            # 从原始数据初始化
            original_data = app_data.get_data('saleable_data')
            if original_data is None or original_data.empty:
                return jsonify({'success': False, 'error': '没有原始数据可以编辑'}), 400
            manual_data = original_data.copy()
        
        # 如果传入了完整数据，直接更新
        if 'data' in data:
            if isinstance(data['data'], list):
                manual_df = pd.DataFrame(data['data'])
                
                # 确保计算结果(KG)列是数值类型
                if '计算结果(KG)' in manual_df.columns:
                    try:
                        manual_df['计算结果(KG)'] = pd.to_numeric(manual_df['计算结果(KG)'], errors='coerce')
                    except Exception as e:
                        print(f"⚠️ 转换计算结果(KG)列类型失败: {e}")
                
                app_data.set_data('saleable_data_manual', manual_df)
            else:
                return jsonify({'success': False, 'error': '数据格式错误'}), 400
        
        # 如果是单个记录更新
        elif 'index' in data:
            index = data['index']
            
            # 转换为DataFrame处理
            if isinstance(manual_data, list):
                manual_df = pd.DataFrame(manual_data)
            else:
                manual_df = manual_data.copy()
            
            if index < 0 or index >= len(manual_df):
                return jsonify({'success': False, 'error': '索引超出范围'}), 400
            
            # 更新重量值
            if 'weight' in data:
                try:
                    numeric_weight = float(data['weight'])
                    manual_df.iloc[index, manual_df.columns.get_loc('计算结果(KG)')] = numeric_weight
                except (ValueError, TypeError):
                    return jsonify({'success': False, 'error': '重量值必须是有效数字'}), 400
            
            # 更新销售单价（含税价）
            if 'price' in data:
                try:
                    numeric_price = float(data['price'])
                    if '销售单价(元/KG)' in manual_df.columns:
                        manual_df.iloc[index, manual_df.columns.get_loc('销售单价(元/KG)')] = numeric_price
                    else:
                        # 如果列不存在，添加列
                        manual_df['销售单价(元/KG)'] = 0
                        manual_df.iloc[index, manual_df.columns.get_loc('销售单价(元/KG)')] = numeric_price
                    
                    # 自动更新不含税价：使用税率计算
                    # 获取税率（从价格数据中获取，如果不存在则使用默认值13%）
                    from data.base_data.price_data import load_price_data, _calculate_price_no_tax, _get_default_tax_rate
                    price_df = load_price_data()
                    tax_rate = 13.0  # 默认税率
                    if price_df is not None and not price_df.empty:
                        product_code = str(manual_df.iloc[index]['拆解产物编码']) if '拆解产物编码' in manual_df.columns else None
                        if product_code:
                            matching_rows = price_df[price_df['拆解产物编码'].astype(str) == product_code]
                            if not matching_rows.empty:
                                tax_rate = float(matching_rows.iloc[0]['税率']) if '税率' in matching_rows.columns and pd.notna(matching_rows.iloc[0]['税率']) else _get_default_tax_rate(numeric_price)
                            else:
                                tax_rate = _get_default_tax_rate(numeric_price)
                        else:
                            tax_rate = _get_default_tax_rate(numeric_price)
                    else:
                        tax_rate = _get_default_tax_rate(numeric_price)
                    
                    price_no_tax = _calculate_price_no_tax(numeric_price, tax_rate)
                    if '销售单价-不含税(元/KG)' in manual_df.columns:
                        manual_df.iloc[index, manual_df.columns.get_loc('销售单价-不含税(元/KG)')] = price_no_tax
                    else:
                        manual_df['销售单价-不含税(元/KG)'] = 0
                        manual_df.iloc[index, manual_df.columns.get_loc('销售单价-不含税(元/KG)')] = price_no_tax
                except (ValueError, TypeError):
                    return jsonify({'success': False, 'error': '单价必须是有效数字'}), 400
            
            # 重新计算销售收益（使用不含税价）
            if '销售单价-不含税(元/KG)' in manual_df.columns:
                weight_kg = float(manual_df.iloc[index]['计算结果(KG)']) if manual_df.iloc[index]['计算结果(KG)'] else 0
                price_no_tax = float(manual_df.iloc[index]['销售单价-不含税(元/KG)']) if manual_df.iloc[index]['销售单价-不含税(元/KG)'] else 0
                revenue = round(weight_kg * price_no_tax, 2)
                
                if '销售收益(元)' in manual_df.columns:
                    manual_df.iloc[index, manual_df.columns.get_loc('销售收益(元)')] = revenue
                else:
                    manual_df['销售收益(元)'] = 0
                    manual_df.iloc[index, manual_df.columns.get_loc('销售收益(元)')] = revenue
            
            app_data.set_data('saleable_data_manual', manual_df)
        
        else:
            return jsonify({'success': False, 'error': '缺少必要参数'}), 400
        
        # 标记数据已修改
        app_data.set_data('saleable_data_modified', True)
        app_data.set_data('saleable_modification_timestamp', datetime.now().isoformat())
        
        # 保存持久化数据
        app_data.save_persistent_data()
        
        # 🔧 清除深加工产物成本计算缓存，确保编辑后的数据能同步到深加工产物成本计算页面
        try:
            print("🔄 清除深加工产物成本计算缓存...")
            # 清除所有预测期数的缓存（1-120个月）
            for period in range(1, 121):
                cache_key = f'deep_processing_product_cost_result_v1_{period}'
                app_data.set_data(cache_key, None)
            print("✅ 深加工产物成本计算缓存已清除")
        except Exception as cache_error:
            print(f"⚠️ 清除深加工产物成本计算缓存失败: {cache_error}")
        
        return jsonify({
            'success': True,
            'message': '可销售量数据更新成功'
        })
        
    except Exception as e:
        print(f"更新可销售量数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'更新失败: {str(e)}'
        }), 500

@data_management_bp.route('/saleable-data/initialize-manual', methods=['POST'])
def initialize_saleable_manual_data():
    """从原始数据初始化手工可销售量数据"""
    try:
        app_data = get_session_data_manager()
        
        # 获取原始可销售量数据
        original_data = app_data.get_data('saleable_data')
        if original_data is None or original_data.empty:
            return jsonify({
                'success': False,
                'error': '没有原始可销售量数据，请先完成数据处理流程'
            }), 400
        
        # 复制数据到手工数据
        manual_data = original_data.copy()
        
        # 保存手工数据
        app_data.set_data('saleable_data_manual', manual_data)
        app_data.set_data('saleable_data_modified', True)
        app_data.set_data('saleable_modification_timestamp', datetime.now().isoformat())
        
        # 保存持久化数据
        app_data.save_persistent_data()
        
        print(f"✅ 可销售量手工数据初始化完成: 从原始数据复制了 {len(manual_data)} 条记录")
        
        return jsonify({
            'success': True,
            'message': '可销售量手工数据初始化成功',
            'copied_count': len(manual_data)
        })
        
    except Exception as e:
        print(f"初始化可销售量手工数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'初始化失败: {str(e)}'
        }), 500

@data_management_bp.route('/saleable-data/initialize-manual-silent', methods=['POST'])
def initialize_saleable_manual_data_silent():
    """静默初始化手工可销售量数据（不设置修改标志）"""
    try:
        app_data = get_session_data_manager()
        
        # 获取原始可销售量数据
        original_data = app_data.get_data('saleable_data')
        if original_data is None or original_data.empty:
            return jsonify({
                'success': False,
                'error': '没有原始可销售量数据，请先完成数据处理流程'
            }), 400
        
        # 复制数据到手工数据
        manual_data = original_data.copy()
        
        # 确保价格信息已正确匹配（如果价格列为0或缺失，从价格数据中匹配）
        if hasattr(manual_data, 'columns') and '拆解产物编码' in manual_data.columns:
            from data.base_data.price_data import load_price_data, get_price_mapping
            import pandas as pd
            
            price_df = load_price_data()
            price_mapping_no_tax = get_price_mapping()
            
            if price_df is not None and not price_df.empty and price_mapping_no_tax:
                # 创建含税价映射
                price_mapping_with_tax = {}
                for _, row in price_df.iterrows():
                    code = str(row['拆解产物编码']).strip()
                    price_with_tax = row.get('销售单价(元/KG)', 0)
                    if pd.notna(price_with_tax):
                        price_mapping_with_tax[code] = float(price_with_tax)
                
                # 更新或添加含税价列
                if '销售单价(元/KG)' not in manual_data.columns:
                    manual_data['销售单价(元/KG)'] = 0.0
                
                def update_price_with_tax(row):
                    code = str(row['拆解产物编码']).strip() if pd.notna(row['拆解产物编码']) else ''
                    matched_price = price_mapping_with_tax.get(code, 0.0)
                    current_price = row.get('销售单价(元/KG)', 0.0)
                    return matched_price if (matched_price > 0 and current_price == 0) else current_price
                
                manual_data['销售单价(元/KG)'] = manual_data.apply(update_price_with_tax, axis=1)
                
                # 更新或添加不含税价列
                if '销售单价-不含税(元/KG)' not in manual_data.columns:
                    manual_data['销售单价-不含税(元/KG)'] = 0.0
                
                def update_price_no_tax(row):
                    code = str(row['拆解产物编码']).strip() if pd.notna(row['拆解产物编码']) else ''
                    matched_price = price_mapping_no_tax.get(code, 0.0)
                    current_price = row.get('销售单价-不含税(元/KG)', 0.0)
                    return matched_price if (matched_price > 0 and current_price == 0) else current_price
                
                manual_data['销售单价-不含税(元/KG)'] = manual_data.apply(update_price_no_tax, axis=1)
                
                # 重新计算销售收益（使用不含税价）
                if '计算结果(KG)' in manual_data.columns:
                    if '销售收益(元)' not in manual_data.columns:
                        manual_data['销售收益(元)'] = 0.0
                    manual_data['销售收益(元)'] = manual_data.apply(
                        lambda row: round(float(row['计算结果(KG)']) * float(row['销售单价-不含税(元/KG)']), 2)
                        if pd.notna(row['计算结果(KG)']) and pd.notna(row['销售单价-不含税(元/KG)']) and float(row['计算结果(KG)']) > 0 and float(row['销售单价-不含税(元/KG)']) != 0 else 0.0,
                        axis=1
                    )
        
        # 保存手工数据（但不设置修改标志）
        app_data.set_data('saleable_data_manual', manual_data)
        # 注意：这里不设置 saleable_data_modified，保持未修改状态
        from datetime import datetime
        app_data.set_data('saleable_silent_init_timestamp', datetime.now().isoformat())
        
        # 保存持久化数据
        app_data.save_persistent_data()
        
        print(f"✅ 可销售量手工数据静默初始化完成: 从原始数据复制了 {len(manual_data)} 条记录（未标记为修改）")
        
        return jsonify({
            'success': True,
            'message': '可销售量手工数据静默初始化成功',
            'copied_count': len(manual_data)
        })
        
    except Exception as e:
        print(f"静默初始化可销售量手工数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'静默初始化失败: {str(e)}'
        }), 500

@data_management_bp.route('/saleable-data/reset', methods=['POST'])
def reset_saleable_to_original():
    """重置可销售量数据到原始状态"""
    try:
        app_data = get_session_data_manager()
        
        # 清除手工数据
        app_data.set_data('saleable_data_manual', None)
        app_data.set_data('saleable_data_modified', False)
        app_data.set_data('saleable_modification_timestamp', None)
        
        # 保存持久化数据
        app_data.save_persistent_data()
        
        print("✅ 可销售量数据已重置到原始状态")
        
        return jsonify({
            'success': True,
            'message': '可销售量数据已重置到原始状态'
        })
        
    except Exception as e:
        print(f"重置可销售量数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'重置失败: {str(e)}'
        }), 500


# ==================== 可销售量数据 Excel 处理 ====================

@data_management_bp.route('/parse-saleable-excel', methods=['POST'])
def parse_saleable_excel():
    """解析可销售量Excel文件（自动识别明细/汇总格式）"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '没有文件'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '没有选择文件'}), 400

        # 读取Excel文件
        df = pd.read_excel(file)

        # 验证数据
        if df.empty:
            return jsonify({'success': False, 'error': 'Excel文件为空'}), 400

        # 自动识别格式：汇总格式包含「汇总计算结果(KG)」列
        is_aggregated = '汇总计算结果(KG)' in df.columns

        if is_aggregated:
            # 汇总格式：拆解产物编码 + 汇总计算结果(KG)
            required_columns = ['拆解产物编码', '汇总计算结果(KG)']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return jsonify({
                    'success': False,
                    'error': f'汇总格式Excel缺少必要列: {", ".join(missing_columns)}'
                }), 400

            df = df.dropna(subset=['拆解产物编码', '汇总计算结果(KG)'])
            try:
                df['汇总计算结果(KG)'] = pd.to_numeric(df['汇总计算结果(KG)'], errors='coerce')
                df = df.dropna(subset=['汇总计算结果(KG)'])
            except:
                return jsonify({'success': False, 'error': '汇总计算结果(KG)列包含无效数据'}), 400

            if df.empty:
                return jsonify({'success': False, 'error': '没有有效的数据行'}), 400

            data_dict = safe_json_convert(df)
            return jsonify({
                'success': True,
                'data': data_dict,
                'count': len(data_dict),
                'format': 'aggregated',
                'message': f'成功解析 {len(data_dict)} 条汇总编辑数据'
            })

        else:
            # 明细格式
            required_columns = ['原物料代码', '原物料名称', '拆解产物编码', '拆解产物名称', '计算结果(KG)', '类别', '期间']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                return jsonify({
                    'success': False,
                    'error': f'Excel文件缺少必要列: {", ".join(missing_columns)}'
                }), 400

            df = df.dropna(subset=['拆解产物编码', '计算结果(KG)'])
            try:
                df['计算结果(KG)'] = pd.to_numeric(df['计算结果(KG)'], errors='coerce')
                df = df.dropna(subset=['计算结果(KG)'])
            except:
                return jsonify({'success': False, 'error': '计算结果(KG)列包含无效数据'}), 400

            if df.empty:
                return jsonify({'success': False, 'error': '没有有效的数据行'}), 400

            data_dict = safe_json_convert(df)
            return jsonify({
                'success': True,
                'data': data_dict,
                'count': len(data_dict),
                'format': 'detail',
                'message': f'成功解析 {len(data_dict)} 条可销售量数据'
            })
        
    except Exception as e:
        print(f"解析可销售量Excel失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'文件解析失败: {str(e)}'
        }), 500

@data_management_bp.route('/import-saleable-data', methods=['POST'])
def import_saleable_data():
    """导入可销售量Excel数据"""
    try:
        data = request.get_json()
        if not data or 'data' not in data:
            return jsonify({'success': False, 'error': '没有接收到数据'}), 400
        
        import_data = data['data']
        if not isinstance(import_data, list) or len(import_data) == 0:
            return jsonify({'success': False, 'error': '导入数据为空'}), 400
        
        app_data = get_session_data_manager()
        
        # 转换为DataFrame
        df = pd.DataFrame(import_data)
        
        # 确保计算结果(KG)列是数值类型
        if '计算结果(KG)' in df.columns:
            try:
                df['计算结果(KG)'] = pd.to_numeric(df['计算结果(KG)'], errors='coerce')
                # 检查是否有无效数据
                if df['计算结果(KG)'].isna().any():
                    print("⚠️ 导入数据中存在无效的重量值，已转换为NaN")
            except Exception as e:
                print(f"⚠️ 转换计算结果(KG)列类型失败: {e}")
        
        # 验证必要列
        required_columns = ['原物料代码', '原物料名称', '拆解产物编码', '拆解产物名称', '计算结果(KG)', '类别', '期间']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return jsonify({
                'success': False,
                'error': f'导入数据缺少必要列: {", ".join(missing_columns)}'
            }), 400
        
        # 保存手工数据
        app_data.set_data('saleable_data_manual', df)
        app_data.set_data('saleable_data_modified', True)
        app_data.set_data('saleable_modification_timestamp', datetime.now().isoformat())
        
        # 保存持久化数据
        app_data.save_persistent_data()
        
        print(f"✅ 可销售量数据导入成功: {len(df)} 条记录")
        
        return jsonify({
            'success': True,
            'message': f'成功导入 {len(df)} 条可销售量数据',
            'imported_count': len(df)
        })
        
    except Exception as e:
        print(f"导入可销售量数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'导入失败: {str(e)}'
        }), 500

@data_management_bp.route('/export-saleable-data', methods=['GET'])
def export_saleable_data():
    """导出可销售量数据到Excel（支持 type=aggregated）"""
    try:
        data_type = request.args.get('type', 'manual')  # 'manual', 'readonly', 或 'aggregated'
        app_data = get_session_data_manager()

        # 辅助函数：构建汇总数据
        def build_aggregated_df(source_df):
            grouped = {}
            for _, row in source_df.iterrows():
                code = str(row.get('拆解产物编码', '')).strip()
                if not code:
                    continue
                if code not in grouped:
                    grouped[code] = {
                        '拆解产物编码': code,
                        '拆解产物名称': str(row.get('拆解产物名称', '')).strip(),
                        '汇总计算结果(KG)': 0.0,
                        '销售单价-含税(元/KG)': 0.0,
                        '销售单价-不含税(元/KG)': 0.0,
                        '明细行数': 0,
                        '_price_wt': [], '_price_nt': []
                    }
                kg = pd.to_numeric(row.get('计算结果(KG)', 0), errors='coerce')
                kg = kg if pd.notna(kg) else 0.0
                price_wt = pd.to_numeric(row.get('销售单价(元/KG)', 0), errors='coerce')
                price_nt = pd.to_numeric(row.get('销售单价-不含税(元/KG)', 0), errors='coerce')

                grouped[code]['汇总计算结果(KG)'] += kg
                grouped[code]['明细行数'] += 1
                if pd.notna(price_wt) and price_wt > 0:
                    grouped[code]['_price_wt'].append(price_wt)
                if pd.notna(price_nt) and price_nt > 0:
                    grouped[code]['_price_nt'].append(price_nt)
                if not grouped[code]['拆解产物名称'] or grouped[code]['拆解产物名称'] == 'nan':
                    grouped[code]['拆解产物名称'] = str(row.get('拆解产物名称', '')).strip()

            rows = []
            for code in sorted(grouped.keys()):
                g = grouped[code]
                wt_prices = g['_price_wt']
                nt_prices = g['_price_nt']
                avg_wt = sum(wt_prices) / len(wt_prices) if wt_prices else 0.0
                avg_nt = sum(nt_prices) / len(nt_prices) if nt_prices else 0.0
                total_kg = round(g['汇总计算结果(KG)'], 6)
                avg_nt_rounded = round(avg_nt, 2)
                rows.append({
                    '拆解产物编码': g['拆解产物编码'],
                    '拆解产物名称': g['拆解产物名称'],
                    '汇总计算结果(KG)': total_kg,
                    '销售单价-含税(元/KG)': round(avg_wt, 2),
                    '销售单价-不含税(元/KG)': avg_nt_rounded,
                    '销售收益(元)': round(total_kg * avg_nt_rounded, 2),
                    '明细行数': g['明细行数']
                })
            return pd.DataFrame(rows)

        def auto_width(ws):
            for col_idx, col in enumerate(ws.columns, start=1):
                col_letter = chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)
                max_len = max((len(str(c.value or '')) for c in col), default=10)
                try:
                    ws.column_dimensions[col_letter].width = min(max_len + 4, 30)
                except:
                    pass

        if data_type == 'all':
            # 导出明细 + 汇总（双 sheet）
            source = app_data.get_data('saleable_data_manual')
            if source is None or (hasattr(source, 'empty') and source.empty):
                source = app_data.get_data('saleable_data')
            if source is None or (hasattr(source, 'empty') and source.empty):
                return jsonify({'success': False, 'error': '没有可导出的数据'}), 400

            detail_df = source
            aggregated_df = build_aggregated_df(source)
            filename_suffix = '完整版'

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                detail_df.to_excel(writer, sheet_name='明细数据', index=False)
                auto_width(writer.sheets['明细数据'])
                aggregated_df.to_excel(writer, sheet_name='汇总编辑数据', index=False)
                auto_width(writer.sheets['汇总编辑数据'])

        elif data_type == 'aggregated':
            source = app_data.get_data('saleable_data_manual')
            if source is None or (hasattr(source, 'empty') and source.empty):
                source = app_data.get_data('saleable_data')
            if source is None or (hasattr(source, 'empty') and source.empty):
                return jsonify({'success': False, 'error': '没有可导出的汇总数据'}), 400

            df = build_aggregated_df(source)
            filename_suffix = '汇总编辑版'

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='汇总编辑数据', index=False)
                auto_width(writer.sheets['汇总编辑数据'])

        elif data_type == 'manual':
            data = app_data.get_data('saleable_data_manual')
            if data is None or (hasattr(data, 'empty') and data.empty):
                return jsonify({'success': False, 'error': '没有可导出的手工数据'}), 400
            df = data
            filename_suffix = '手工版'

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='可销售量数据', index=False)
                auto_width(writer.sheets['可销售量数据'])

        else:
            data = app_data.get_data('saleable_data')
            if data is None or (hasattr(data, 'empty') and data.empty):
                return jsonify({'success': False, 'error': '没有可导出的原始数据'}), 400
            df = data
            filename_suffix = '原始版'

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='可销售量数据', index=False)
                auto_width(writer.sheets['可销售量数据'])

        output.seek(0)
        filename = f'可销售量数据_{filename_suffix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"导出可销售量数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/saleable-data-template', methods=['GET'])
def get_saleable_data_template():
    """下载可销售量数据Excel模板（支持 format=aggregated 参数）"""
    try:
        template_format = request.args.get('format', 'detail')  # 'detail' 或 'aggregated'

        if template_format == 'aggregated':
            # 汇总格式模板 - 导出当前实际汇总数据（不含销售收益列）
            app_data = get_session_data_manager()
            source = app_data.get_data('saleable_data_manual')
            if source is None or (hasattr(source, 'empty') and source.empty):
                source = app_data.get_data('saleable_data')
            if source is None or (hasattr(source, 'empty') and source.empty):
                # 无数据时返回示例模板
                aggregated_rows = [{
                    '拆解产物编码': '产物编码1',
                    '拆解产物名称': '产物名称1',
                    '汇总计算结果(KG)': 300.123456,
                    '销售单价-含税(元/KG)': 12.50,
                    '销售单价-不含税(元/KG)': 11.06,
                    '明细行数': 3
                }]
            else:
                grouped = {}
                for _, row in source.iterrows():
                    code = str(row.get('拆解产物编码', '')).strip()
                    if not code: continue
                    if code not in grouped:
                        grouped[code] = {
                            '拆解产物编码': code,
                            '拆解产物名称': str(row.get('拆解产物名称', '')).strip(),
                            '汇总计算结果(KG)': 0.0,
                            '销售单价-含税(元/KG)': 0.0,
                            '销售单价-不含税(元/KG)': 0.0,
                            '明细行数': 0,
                            '_price_wt': [], '_price_nt': []
                        }
                    kg = pd.to_numeric(row.get('计算结果(KG)', 0), errors='coerce')
                    kg = kg if pd.notna(kg) else 0.0
                    price_wt = pd.to_numeric(row.get('销售单价(元/KG)', 0), errors='coerce')
                    price_nt = pd.to_numeric(row.get('销售单价-不含税(元/KG)', 0), errors='coerce')

                    grouped[code]['汇总计算结果(KG)'] += kg
                    grouped[code]['明细行数'] += 1
                    if pd.notna(price_wt) and price_wt > 0:
                        grouped[code]['_price_wt'].append(price_wt)
                    if pd.notna(price_nt) and price_nt > 0:
                        grouped[code]['_price_nt'].append(price_nt)
                    if not grouped[code]['拆解产物名称'] or grouped[code]['拆解产物名称'] == 'nan':
                        grouped[code]['拆解产物名称'] = str(row.get('拆解产物名称', '')).strip()

                aggregated_rows = []
                for code in sorted(grouped.keys()):
                    g = grouped[code]
                    wt_prices = g['_price_wt']
                    nt_prices = g['_price_nt']
                    avg_wt = sum(wt_prices) / len(wt_prices) if wt_prices else 0.0
                    avg_nt = sum(nt_prices) / len(nt_prices) if nt_prices else 0.0
                    aggregated_rows.append({
                        '拆解产物编码': g['拆解产物编码'],
                        '拆解产物名称': g['拆解产物名称'],
                        '汇总计算结果(KG)': round(g['汇总计算结果(KG)'], 6),
                        '销售单价-含税(元/KG)': round(avg_wt, 2),
                        '销售单价-不含税(元/KG)': round(avg_nt, 2),
                        '明细行数': g['明细行数']
                    })

            df = pd.DataFrame(aggregated_rows)
            sheet_name = '汇总编辑数据'
            instructions = [
                ['字段名', '说明'],
                ['拆解产物编码', '拆解后产物的代码（必填，用于匹配回填）'],
                ['拆解产物名称', '拆解后产物的名称'],
                ['汇总计算结果(KG)', '该编码下所有明细行的KG合计值（必填，数值型）'],
                ['销售单价-含税(元/KG)', '仅供参考，回填时不影响明细单价'],
                ['销售单价-不含税(元/KG)', '仅供参考，回填时不影响明细单价'],
                ['明细行数', '该编码下的明细行数（仅供参考，回填时自动计算）'],
                ['', ''],
                ['操作说明', ''],
                ['1', '修改「汇总计算结果(KG)」列后导入，系统会按原占比自动分配回明细行'],
                ['2', '导入时根据「拆解产物编码」匹配原有数据，按原占比回填'],
                ['3', '如果编码在原数据中不存在，该行将被忽略'],
                ['4', '销售收益(元) = 汇总计算结果(KG) * 销售单价-不含税(元/KG)，由系统自动计算'],
                ['5', '文件格式支持 .xlsx, .xls, .csv'],
            ]
        else:
            # 原有明细格式模板
            template_data = {
                '原物料代码': ['示例代码1', '示例代码2'],
                '原物料名称': ['示例原料1', '示例原料2'],
                '拆解产物编码': ['产物编码1', '产物编码2'],
                '拆解产物名称': ['产物名称1', '产物名称2'],
                '计算结果(KG)': [100.123456, 200.654321],
                '类别': ['示例类别1', '示例类别2'],
                '期间': ['2024-01', '2024-01']
            }
            sheet_name = '可销售量数据'
            instructions = [
                ['字段名', '说明'],
                ['原物料代码', '原始物料的代码标识'],
                ['原物料名称', '原始物料的名称'],
                ['拆解产物编码', '拆解后产物的代码'],
                ['拆解产物名称', '拆解后产物的名称'],
                ['计算结果(KG)', '可销售量重量值（必填，数值型）'],
                ['类别', '产物分类'],
                ['期间', '数据所属期间'],
                ['', ''],
                ['注意事项', ''],
                ['1', '计算结果(KG)列必须为数值，支持小数点后6位'],
                ['2', '导入时会覆盖现有手工数据'],
                ['3', '建议先导出现有数据作为备份'],
                ['4', '文件格式支持 .xlsx, .xls, .csv'],
            ]

            df = pd.DataFrame(template_data)

        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 写入模板数据
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # 写入导入说明
            instruction_df = pd.DataFrame(instructions)
            instruction_df.to_excel(writer, sheet_name='导入说明', index=False, header=False)

            # 设置列宽
            data_ws = writer.sheets[sheet_name]
            col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
            col_widths = [18, 22, 20, 20, 18, 18, 15, 12]
            for i, width in enumerate(col_widths):
                if i < len(col_letters):
                    try:
                        data_ws.column_dimensions[col_letters[i]].width = width
                    except:
                        pass

            instruction_ws = writer.sheets['导入说明']
            instruction_ws.column_dimensions['A'].width = 20
            instruction_ws.column_dimensions['B'].width = 55

        output.seek(0)
        suffix = '汇总编辑' if template_format == 'aggregated' else '明细'
        filename = f'可销售量数据导入模板_{suffix}_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 提取结果编辑数据管理 API ====================

def _ensure_manual_edit_columns(manual_data):
    """确保手工编辑列存在，新增旧机行的编辑列默认为 0"""
    manual_cols = ['初始数据', '本期计划采购数量', '计划采购单价', '本期计划投产数量']
    for col in manual_cols:
        if col not in manual_data.columns:
            manual_data[col] = 0.0
        else:
            manual_data[col] = pd.to_numeric(manual_data[col], errors='coerce').fillna(0)
    return manual_data


@data_management_bp.route('/extracted-data', methods=['GET'])
def get_extracted_data():
    """获取提取结果数据(只读) - 只显示旧机类别"""
    try:
        app_data = get_session_data_manager()
        extracted_data = app_data.get_data('extracted_data')
        
        if extracted_data is None or extracted_data.empty:
            return jsonify({
                'success': True,
                'data': [],
                'message': '暂无提取结果数据，请先进行数据提取'
            })

        extracted_data = extracted_data.copy()
        extracted_data, added = supplement_old_machine_from_mapping(extracted_data)
        if added > 0:
            app_data.set_data('extracted_data', extracted_data)
            print(f"📊 [懒补全] extracted_data 补全旧机 {added} 条")
        
        # 只返回类别为"旧机"的数据
        if '类别' in extracted_data.columns:
            extracted_data = extracted_data[extracted_data['类别'] == '旧机'].copy()
        
        # 分页参数
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        
        # 计算分页
        total = len(extracted_data)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        data_page = extracted_data.iloc[start_idx:end_idx]
        
        return jsonify({
            'success': True,
            'data': safe_json_convert(data_page),
            'total': total,
            'current_page': page,
            'per_page': per_page,
            'pages': (total + per_page - 1) // per_page
        })
        
    except Exception as e:
        print(f"获取提取结果数据失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/extracted-data-manual', methods=['GET'])
def get_extracted_data_manual():
    """获取提取结果数据(手工) - 只返回旧机类别供编辑页面显示"""
    try:
        app_data = get_session_data_manager()
        
        # 获取手工数据（不自动创建）
        manual_data = app_data.get_data('extracted_data_manual')
        
        if manual_data is None or manual_data.empty:
            return jsonify({
                'success': True,
                'data': [],
                'message': '暂无提取结果数据(手工)，请先从只读数据初始化'
            })

        manual_data = manual_data.copy()
        manual_data, added = supplement_old_machine_from_mapping(manual_data)
        if added > 0:
            manual_data = _ensure_manual_edit_columns(manual_data)
            app_data.set_data('extracted_data_manual', manual_data)
            print(f"📊 [懒补全] extracted_data_manual 补全旧机 {added} 条")
        
        # 如果计划采购单价列为0，从单价列复制（仅针对旧机类别）
        if '计划采购单价' in manual_data.columns and '单价' in manual_data.columns and '类别' in manual_data.columns:
            mask_old_machine = manual_data['类别'] == '旧机'
            mask_zero_price = (manual_data['计划采购单价'] == 0) | (manual_data['计划采购单价'].isna())
            mask_to_update = mask_old_machine & mask_zero_price
            if mask_to_update.any():
                price_values = pd.to_numeric(manual_data.loc[mask_to_update, '单价'], errors='coerce').fillna(0)
                manual_data.loc[mask_to_update, '计划采购单价'] = price_values
                print(f"[修复] 将 {mask_to_update.sum()} 条旧机记录的计划采购单价从单价列复制")
        
        # 只返回旧机类别供页面显示（但数据库中保存了所有类别）
        if '类别' in manual_data.columns:
            display_data = manual_data[manual_data['类别'] == '旧机'].copy()
            print(f"📊 手工数据总行数: {len(manual_data)}, 返回旧机行数: {len(display_data)}")
        else:
            display_data = manual_data
        
        # 分页参数
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        
        # 计算分页（基于过滤后的数据）
        total = len(display_data)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        data_page = display_data.iloc[start_idx:end_idx]
        
        return jsonify({
            'success': True,
            'data': safe_json_convert(data_page),
            'total': total,
            'current_page': page,
            'per_page': per_page,
            'pages': (total + per_page - 1) // per_page
        })
        
    except Exception as e:
        print(f"获取提取结果数据(手工)失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/extracted-data', methods=['PUT'])
def update_extracted_data():
    """更新提取结果数据(手工) - 编辑本期计划采购数量"""
    print("=" * 80)
    print("📝 开始更新提取结果数据")
    print("=" * 80)
    try:
        import pandas as pd
        from datetime import datetime
        
        app_data = get_session_data_manager()
        request_data = request.get_json()
        
        print(f"📥 接收到的请求数据: {request_data}")
        
        if not request_data:
            return jsonify({'success': False, 'error': '未接收到数据'}), 400
        
        modified_records = request_data.get('modified_records', [])
        print(f"📝 要修改的记录: {modified_records}")
        
        if not modified_records:
            return jsonify({'success': False, 'error': '没有要更新的记录'}), 400
        
        # 获取手工数据
        print("📥 正在获取手工数据...")
        manual_data = app_data.get_data('extracted_data_manual')
        print(f"📥 手工数据: {manual_data is not None}, 行数: {len(manual_data) if manual_data is not None else 0}")
        
        if manual_data is None or manual_data.empty:
            return jsonify({
                'success': False,
                'error': '没有找到手工数据，请先初始化'
            }), 400
        
        # 确保序号列存在
        if '序号' not in manual_data.columns:
            manual_data.insert(0, '序号', range(1, len(manual_data) + 1))
        
        # 确保必需的列存在
        if '初始数据' not in manual_data.columns:
            # 如果没有初始数据列，从非限制使用的库存创建
            manual_data['初始数据'] = pd.to_numeric(manual_data['非限制使用的库存'], errors='coerce').fillna(0)
        
        if '本期计划采购数量' not in manual_data.columns:
            manual_data['本期计划采购数量'] = 0.0
        
        if '计划采购单价' not in manual_data.columns:
            manual_data['计划采购单价'] = 0.0
        
        if '本期计划投产数量' not in manual_data.columns:
            manual_data['本期计划投产数量'] = manual_data['初始数据'].astype(float)
        
        # 打印序号列的信息
        print(f"[数据统计] 序号列数据类型: {manual_data['序号'].dtype}")
        print(f"[数据统计] 序号列前5个值: {manual_data['序号'].head().tolist()}")
        
        # 更新记录
        updated_count = 0
        for record in modified_records:
            row_no = record.get('序号')
            print(f"[处理记录] 处理记录序号: {row_no}, 类型: {type(row_no)}")
            
            if row_no is None:
                print("[警告] 序号为空，跳过")
                continue
            
            # 尝试类型转换以匹配
            try:
                # 如果序号列是int类型，确保row_no也是int
                if manual_data['序号'].dtype in ['int64', 'int32']:
                    row_no = int(row_no)
                # 如果是float类型，转换为float
                elif manual_data['序号'].dtype in ['float64', 'float32']:
                    row_no = float(row_no)
                print(f"[转换] 转换后的序号: {row_no}, 类型: {type(row_no)}")
            except Exception as e:
                print(f"[警告] 序号类型转换失败: {e}")
            
            # 查找对应行（必须是旧机类别）
            if '类别' in manual_data.columns:
                mask = (manual_data['序号'] == row_no) & (manual_data['类别'] == '旧机')
                print(f"[查找] 查找序号 {row_no}（旧机类别），找到: {mask.any()}, 匹配行数: {mask.sum()}")
            else:
                mask = manual_data['序号'] == row_no
                print(f"[查找] 查找序号 {row_no}，找到: {mask.any()}, 匹配行数: {mask.sum()}")
            
            if mask.any():
                # 更新本期计划采购数量
                if '本期计划采购数量' in record:
                    purchase_qty = float(record['本期计划采购数量'])
                    print(f"[更新] 更新采购数量: {purchase_qty}")
                    manual_data.loc[mask, '本期计划采购数量'] = purchase_qty
                    
                    # 更新计算列
                    initial_data = manual_data.loc[mask, '初始数据'].values[0]
                    if pd.isna(initial_data):
                        initial_data = 0
                    else:
                        initial_data = float(initial_data)
                    
                    print(f"[数据统计] 初始数据: {initial_data}")
                    
                    # 本期计划投产数量 = 初始数据 + 本期计划采购数量
                    production_qty = initial_data + purchase_qty
                    manual_data.loc[mask, '本期计划投产数量'] = production_qty
                    print(f"[成功] 投产数量: {production_qty}")
                    
                    # 非限制使用的库存 = 本期计划投产数量（如果没有单独设置实际投产数量）
                    manual_data.loc[mask, '非限制使用的库存'] = production_qty
                    print(f"[成功] 库存更新: {production_qty}")
                
                # 更新计划采购单价
                if '计划采购单价' in record:
                    purchase_price = float(record['计划采购单价']) if record['计划采购单价'] is not None else 0.0
                    print(f"[更新] 更新计划采购单价: {purchase_price}")
                    manual_data.loc[mask, '计划采购单价'] = purchase_price
                    print(f"[成功] 计划采购单价更新: {purchase_price}")
                
                # 更新本期实际投产数量（独立编辑，不反向计算）
                if '本期实际投产数量' in record:
                    actual_qty = float(record['本期实际投产数量'])
                    print(f"[更新] 更新实际投产数量: {actual_qty}")
                    # 直接更新非限制使用的库存字段，不影响其他字段
                    manual_data.loc[mask, '非限制使用的库存'] = actual_qty
                    print(f"[成功] 实际投产数量更新: {actual_qty}（不反向计算）")
                
                if '本期计划采购数量' in record or '计划采购单价' in record or '本期实际投产数量' in record:
                    updated_count += 1
                else:
                    print("[警告] 记录中没有可编辑字段")
        
        if updated_count > 0:
            # 保存更新后的数据
            app_data.set_data('extracted_data_manual', manual_data)
            app_data.mark_extracted_data_modified()
            
            # 不覆盖extracted_data，保持原始数据用于统计对比
            # 计算引擎会优先使用extracted_data_manual
            
            print(f"[成功] 成功更新 {updated_count} 条提取结果数据")
            
            return jsonify({
                'success': True,
                'message': f'成功更新 {updated_count} 条记录，请返回首页点击"重新计算"使结果链路生效',
                'updated_count': updated_count
            })
        else:
            return jsonify({
                'success': False,
                'error': '没有记录被更新'
            }), 400
        
    except Exception as e:
        print("=" * 80)
        print(f"[错误] 更新提取结果数据(手工)失败: {str(e)}")
        print("=" * 80)
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/extracted-data/row', methods=['POST'])
def append_extracted_data_row():
    """向提取结果手工表追加一条「旧机」物料行（不参与只读提取，仅手工表）。"""
    try:
        import pandas as pd

        app_data = get_session_data_manager()
        body = request.get_json() or {}

        material_code = (body.get('物料代码') or '').strip()
        material_desc = (body.get('物料描述') or '').strip()
        if not material_code:
            return jsonify({'success': False, 'error': '物料代码不能为空'}), 400
        if not material_desc:
            return jsonify({'success': False, 'error': '物料描述不能为空'}), 400

        manual_data = app_data.get_data('extracted_data_manual')
        if manual_data is None or manual_data.empty:
            return jsonify({
                'success': False,
                'error': '没有找到手工数据，请先点击「从只读数据初始化」'
            }), 400

        manual_data = manual_data.copy()

        if '序号' not in manual_data.columns:
            manual_data.insert(0, '序号', range(1, len(manual_data) + 1))
        if '初始数据' not in manual_data.columns:
            if '非限制使用的库存' in manual_data.columns:
                manual_data['初始数据'] = pd.to_numeric(
                    manual_data['非限制使用的库存'], errors='coerce'
                ).fillna(0)
            else:
                manual_data['初始数据'] = 0.0
        if '本期计划采购数量' not in manual_data.columns:
            manual_data['本期计划采购数量'] = 0.0
        if '计划采购单价' not in manual_data.columns:
            manual_data['计划采购单价'] = 0.0
        if '本期计划投产数量' not in manual_data.columns:
            manual_data['本期计划投产数量'] = pd.to_numeric(
                manual_data['初始数据'], errors='coerce'
            ).fillna(0)

        unit = (body.get('单位') or '').strip()
        initial_data = float(body.get('初始数据', 0) or 0)
        purchase_qty = float(body.get('本期计划采购数量', 0) or 0)
        purchase_price = float(body.get('计划采购单价', 0) or 0)
        production_qty = initial_data + purchase_qty

        actual_raw = body.get('本期实际投产数量')
        if actual_raw is not None and actual_raw != '':
            inventory_qty = float(actual_raw)
        else:
            inventory_qty = production_qty

        seq_series = pd.to_numeric(manual_data['序号'], errors='coerce')
        mx = seq_series.max()
        next_seq = int(mx) + 1 if pd.notna(mx) else 1

        period_default = ''
        if '期间' in manual_data.columns:
            s = manual_data['期间'].dropna()
            if len(s) > 0:
                period_default = s.iloc[0]

        new_row = {}
        for col in manual_data.columns:
            if col == '序号':
                new_row[col] = next_seq
            elif col == '类别':
                new_row[col] = '旧机'
            elif col == '物料代码':
                new_row[col] = material_code
            elif col == '物料描述':
                new_row[col] = material_desc
            elif col == '单位':
                new_row[col] = unit
            elif col == '期间':
                new_row[col] = period_default
            elif col == '初始数据':
                new_row[col] = initial_data
            elif col == '本期计划采购数量':
                new_row[col] = purchase_qty
            elif col == '计划采购单价':
                new_row[col] = purchase_price
            elif col == '本期计划投产数量':
                new_row[col] = production_qty
            elif col == '非限制使用的库存':
                new_row[col] = inventory_qty
            elif col == '单价':
                new_row[col] = purchase_price
            elif col in ('价值',):
                new_row[col] = 0.0
            elif col == '库位描述':
                new_row[col] = ''
            else:
                dtype = manual_data[col].dtype
                if pd.api.types.is_integer_dtype(dtype):
                    new_row[col] = 0
                elif pd.api.types.is_float_dtype(dtype):
                    new_row[col] = 0.0
                else:
                    new_row[col] = ''

        new_df = pd.DataFrame([new_row])
        new_df = new_df.reindex(columns=manual_data.columns)
        manual_data = pd.concat([manual_data, new_df], ignore_index=True)

        app_data.set_data('extracted_data_manual', manual_data)
        app_data.mark_extracted_data_modified()

        print(f"[成功] 追加提取结果旧机行: 序号={next_seq}, 物料代码={material_code}")

        return jsonify({
            'success': True,
            'message': '已新增物料行，请返回首页点击「重新计算」使结果链路生效',
            '序号': next_seq
        })

    except Exception as e:
        print(f"[错误] 追加提取结果行失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/extracted-data/row', methods=['DELETE'])
def delete_extracted_data_row():
    """从提取结果手工表中删除一条「旧机」行（按序号，与编辑页范围一致）。"""
    try:
        import pandas as pd

        app_data = get_session_data_manager()
        body = request.get_json() or {}
        row_no = body.get('序号')

        if row_no is None:
            return jsonify({'success': False, 'error': '缺少参数：序号'}), 400

        manual_data = app_data.get_data('extracted_data_manual')
        if manual_data is None or manual_data.empty:
            return jsonify({
                'success': False,
                'error': '没有找到手工数据，请先点击「从只读数据初始化」'
            }), 400

        if '序号' not in manual_data.columns:
            return jsonify({'success': False, 'error': '手工数据缺少序号列'}), 400

        manual_data = manual_data.copy()

        try:
            if manual_data['序号'].dtype in ['int64', 'int32']:
                row_no = int(row_no)
            elif manual_data['序号'].dtype in ['float64', 'float32']:
                row_no = float(row_no)
        except Exception:
            pass

        if '类别' in manual_data.columns:
            mask = (manual_data['序号'] == row_no) & (manual_data['类别'] == '旧机')
        else:
            mask = manual_data['序号'] == row_no

        if not mask.any():
            return jsonify({
                'success': False,
                'error': f'未找到序号为 {row_no} 的旧机记录，无法删除'
            }), 400

        removed = int(mask.sum())
        manual_data = manual_data.loc[~mask].reset_index(drop=True)

        app_data.set_data('extracted_data_manual', manual_data)
        app_data.mark_extracted_data_modified()

        print(f"[成功] 删除提取结果旧机行: 序号={row_no}, 删除行数={removed}")

        return jsonify({
            'success': True,
            'message': '已删除该物料行，请返回首页点击「重新计算」使结果链路生效',
            'removed_count': removed
        })

    except Exception as e:
        print(f"[错误] 删除提取结果行失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/extracted-data/comparison', methods=['GET'])
def get_extracted_comparison():
    """获取提取结果数据对比统计"""
    try:
        app_data = get_session_data_manager()
        stats = app_data.get_extracted_comparison_stats()
        
        return jsonify({
            'success': True,
            'data': stats
        })
    except Exception as e:
        print(f"获取提取结果数据对比统计失败: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/extracted-data/reset', methods=['POST'])
def reset_extracted_data():
    """清除提取结果手工数据，允许重新初始化"""
    try:
        import pandas as pd
        app_data = get_session_data_manager()
        
        print("🔄 开始清除提取结果手工数据...")
        
        # 清除手工数据（设置为空DataFrame）
        app_data.set_data('extracted_data_manual', pd.DataFrame())
        app_data.set_data('extracted_data_modified', False)
        app_data.set_data('extracted_modification_timestamp', None)
        
        # 不修改extracted_data，保持原始值用于统计对比
        # extracted_data始终是原始提取数据
        # extracted_data_manual是手工编辑数据
        
        print("✅ 提取结果手工数据已清除，extracted_data保持不变")
        
        return jsonify({
            'success': True,
            'message': '提取结果数据已清除，可以重新初始化；如需更新结果链路请返回首页点击"重新计算"'
        })
        
    except Exception as e:
        print(f"清除提取结果数据失败: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'清除失败: {str(e)}'
        }), 500

@data_management_bp.route('/extracted-data/initialize-manual', methods=['POST'])
def initialize_extracted_manual_from_readonly():
    """从只读数据初始化手工数据"""
    print("=" * 80)
    print("[初始化] 开始初始化提取结果手工数据")
    print("=" * 80)
    try:
        import pandas as pd
        from datetime import datetime
        
        app_data = get_session_data_manager()
        
        # 获取只读数据
        print("[获取数据] 正在获取只读数据...")
        readonly_data = app_data.get_data('extracted_data')
        print(f"[获取数据] 只读数据获取完成: {readonly_data is not None}, 行数: {len(readonly_data) if readonly_data is not None else 0}")

        if readonly_data is None or readonly_data.empty:
            readonly_data, added = supplement_old_machine_from_mapping(pd.DataFrame())
            if added == 0:
                return jsonify({
                    'success': False,
                    'error': '没有只读数据可以复制，请先进行数据提取'
                }), 400
            app_data.set_data('extracted_data', readonly_data)
            print(f"[补全] 只读数据为空，已从映射表补全 {added} 条旧机占位记录")
        else:
            readonly_data = readonly_data.copy()
            readonly_data, added = supplement_old_machine_from_mapping(readonly_data)
            if added > 0:
                app_data.set_data('extracted_data', readonly_data)
                print(f"[补全] 只读数据补全旧机 {added} 条")
        
        # 复制所有类别的数据（保留完整数据）
        print("[初始化] 开始初始化手工数据（保留所有类别）...")
        manual_data = readonly_data.copy()
        
        if '类别' in manual_data.columns:
            print(f"[数据统计] 类别列存在，唯一值: {manual_data['类别'].unique().tolist()}")
            print(f"[数据统计] 总数据行数: {len(manual_data)}")
            
            # 统计旧机数据
            old_machine_count = len(manual_data[manual_data['类别'] == '旧机'])
            other_count = len(manual_data[manual_data['类别'] != '旧机'])
            print(f"[数据统计] 旧机类别: {old_machine_count} 条")
            print(f"[数据统计] 其他类别: {other_count} 条")
            
            if old_machine_count == 0:
                print("[错误] 映射表中无旧机数据可初始化")
                return jsonify({
                    'success': False,
                    'error': '内置映射表中没有旧机类别数据，无法初始化'
                }), 400
        
        # 打印调试信息
        print(f"[数据统计] 原始数据列: {list(manual_data.columns)}")
        print(f"[数据统计] 数据行数: {len(manual_data)}")
        
        # 添加新列（所有行都添加，但只为旧机填充值）
        # 1. 初始数据列
        manual_data['初始数据'] = 0.0
        
        # 2. 本期计划采购数量列
        manual_data['本期计划采购数量'] = 0.0
        
        # 3. 计划采购单价列
        manual_data['计划采购单价'] = 0.0
        
        # 4. 本期计划投产数量列
        manual_data['本期计划投产数量'] = 0.0
        
        # 只为旧机类别填充数据
        if '类别' in manual_data.columns and '非限制使用的库存' in manual_data.columns:
            mask_old_machine = manual_data['类别'] == '旧机'
            
            # 初始数据 = 非限制使用的库存（只针对旧机）
            inventory_values = pd.to_numeric(manual_data.loc[mask_old_machine, '非限制使用的库存'], errors='coerce').fillna(0)
            manual_data.loc[mask_old_machine, '初始数据'] = inventory_values
            
            # 本期计划投产数量 = 初始数据（只针对旧机）
            manual_data.loc[mask_old_machine, '本期计划投产数量'] = manual_data.loc[mask_old_machine, '初始数据']
            
            # 计划采购单价 = 单价（只针对旧机，从Excel表内的单价列复制）
            if '单价' in manual_data.columns:
                price_values = pd.to_numeric(manual_data.loc[mask_old_machine, '单价'], errors='coerce').fillna(0)
                manual_data.loc[mask_old_machine, '计划采购单价'] = price_values
                print(f"[成功] 旧机类别的计划采购单价已从单价列复制，前5行: {manual_data[mask_old_machine]['计划采购单价'].head().tolist()}")
            else:
                print(f"[警告] 数据中缺少'单价'列，计划采购单价保持为0")
            
            print(f"[成功] 旧机类别的初始数据已填充，前5行: {manual_data[mask_old_machine]['初始数据'].head().tolist()}")
            print(f"[提示] 其他类别的三个编辑列保持为0")
        
        # 重新排列列顺序，将新列放在"非限制使用的库存"之后
        cols = list(manual_data.columns)
        # 找到"非限制使用的库存"的位置
        if '非限制使用的库存' in cols:
            insert_pos = cols.index('非限制使用的库存') + 1
            # 移除新增列
            new_cols = [c for c in cols if c not in ['初始数据', '本期计划采购数量', '计划采购单价', '本期计划投产数量']]
            # 在指定位置插入新列
            new_cols.insert(insert_pos, '初始数据')
            new_cols.insert(insert_pos + 1, '本期计划投产数量')
            new_cols.insert(insert_pos + 2, '本期计划采购数量')
            new_cols.insert(insert_pos + 3, '计划采购单价')
            manual_data = manual_data[new_cols]
            print(f"[成功] 列顺序已调整: {list(manual_data.columns)}")
        
        # 再次打印初始数据列的值以确认
        if '类别' in manual_data.columns:
            mask_old_machine = manual_data['类别'] == '旧机'
            print(f"[验证] 旧机类别最终初始数据前5行: {manual_data[mask_old_machine]['初始数据'].head().tolist()}")
        
        # 保存完整的手工数据（包含所有类别）
        app_data.set_data('extracted_data_manual', manual_data)
        app_data.set_data('original_extracted_data', readonly_data.copy())
        app_data.set_data('extracted_data_modified', False)
        app_data.set_data('extracted_modification_timestamp', datetime.now().isoformat())
        
        total_count = len(manual_data)
        old_machine_count = len(manual_data[manual_data['类别'] == '旧机']) if '类别' in manual_data.columns else total_count
        print(f"[成功] 提取结果手工数据初始化完成: 总共 {total_count} 条记录（其中旧机 {old_machine_count} 条可编辑）")
        
        return jsonify({
            'success': True,
            'message': f'手工数据初始化成功（总共{total_count}条，其中旧机{old_machine_count}条可编辑）。如需更新结果链路请返回首页点击"重新计算"',
            'copied_count': total_count,
            'old_machine_count': old_machine_count
        })
        
    except Exception as e:
        print("=" * 80)
        print(f"❌ 初始化手工数据失败: {str(e)}")
        print("=" * 80)
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'初始化失败: {str(e)}'
        }), 500

@data_management_bp.route('/extracted-data/template', methods=['GET'])
def download_extracted_data_template():
    """下载提取结果数据导入模板"""
    try:
        import pandas as pd
        import io
        from datetime import datetime
        from app.utils.excel_utils import EXTRACTED_MANUAL_EXPORT_COLUMNS, auto_width_excel_columns
        
        template_data = {
            '序号': [1, 2, 3],
            '类别': ['旧机', '旧机', '旧机'],
            '期间': ['202412', '202412', '202412'],
            '物料代码': ['示例代码1', '示例代码2', '示例代码3'],
            '物料描述': ['示例物料1', '示例物料2', '示例物料3'],
            '单位': ['TAI', 'TAI', 'TAI'],
            '本期实际投产数量': [120.0, 220.0, 320.0],
            '初始数据': [100.0, 200.0, 300.0],
            '本期计划投产数量': [120.0, 220.0, 320.0],
            '本期计划采购数量': [20.0, 20.0, 20.0],
            '计划采购单价': [42.0, 42.0, 42.0],
            '价值': [5000.0, 9000.0, 13000.0],
            '单价': [40.0, 40.0, 40.0],
            '库位描述': ['原料库', '原料库', '原料库'],
        }
        
        df = pd.DataFrame(template_data)[EXTRACTED_MANUAL_EXPORT_COLUMNS]
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='提取结果数据', index=False)
            
            instructions = [
                ['字段名', '说明'],
                ['序号', '数据行号（自动生成）'],
                ['类别', '物料分类，导入旧机数据时请保持为「旧机」'],
                ['期间', '数据所属期间'],
                ['物料代码', '物料代码标识'],
                ['物料描述', '物料名称描述'],
                ['单位', '数量单位'],
                ['本期实际投产数量', '实际投产数量（可编辑，数值型）'],
                ['初始数据', '原始库存值（只读参考）'],
                ['本期计划投产数量', '计划投产数量（自动计算：初始数据 + 本期计划采购数量）'],
                ['本期计划采购数量', '计划采购数量（可编辑，数值型）'],
                ['计划采购单价', '计划采购单价（可编辑，数值型）'],
                ['价值', '物料价值（只读参考）'],
                ['单价', '物料单价（只读参考）'],
                ['库位描述', '库位描述（只读参考）'],
                ['', ''],
                ['计算逻辑', ''],
                ['步骤1', '编辑"本期计划采购数量" → 自动计算：本期计划投产数量 = 初始数据 + 本期计划采购数量'],
                ['步骤2', '编辑"本期实际投产数量"（可选） → 手工调整最终投产数量'],
                ['', ''],
                ['注意事项', ''],
                ['1', '列顺序须与导出文件一致，共 14 列'],
                ['2', '导入时会覆盖现有手工旧机数据'],
                ['3', '建议先导出现有数据作为备份'],
                ['4', '文件格式支持 .xlsx, .xls'],
            ]
            
            instruction_df = pd.DataFrame(instructions)
            instruction_df.to_excel(writer, sheet_name='导入说明', index=False, header=False)
            
            auto_width_excel_columns(writer.sheets['提取结果数据'])
            instruction_ws = writer.sheets['导入说明']
            instruction_ws.column_dimensions['A'].width = 18
            instruction_ws.column_dimensions['B'].width = 60
        
        output.seek(0)
        filename = f'提取结果数据导入模板_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

def _read_extracted_import_sheet(file_like):
    """
    读取「提取结果」导入用工作表。导出手工数据时使用「提取结果(手工)」，
    模板/只读导出使用「提取结果数据」，此处需兼容多种 sheet 名。
    """
    import pandas as pd

    try:
        file_like.seek(0)
    except (AttributeError, OSError):
        pass

    xl = pd.ExcelFile(file_like)
    preferred = ('提取结果数据', '提取结果(手工)', '提取结果(只读)')
    for name in preferred:
        if name in xl.sheet_names:
            return xl.parse(sheet_name=name)

    if len(xl.sheet_names) == 1:
        return xl.parse(sheet_name=0)

    raise ValueError(
        '未找到工作表「提取结果数据」或「提取结果(手工)」。'
        f'当前工作簿中的工作表为: {", ".join(xl.sheet_names)}'
    )


@data_management_bp.route('/parse-extracted-excel', methods=['POST'])
def parse_extracted_excel():
    """解析提取结果Excel文件"""
    try:
        import pandas as pd
        
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '未上传文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '文件名为空'}), 400
        
        df = _read_extracted_import_sheet(file)
        
        # 验证必需列
        required_columns = ['序号', '物料代码', '本期计划采购数量']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({
                'success': False,
                'error': f'缺少必需列: {", ".join(missing_columns)}'
            }), 400
        
        # 转换数据并验证
        parsed_data = []
        errors = []
        
        for idx, row in df.iterrows():
            try:
                # 验证本期计划采购数量
                purchase_qty = row.get('本期计划采购数量', 0)
                try:
                    purchase_qty = float(purchase_qty) if pd.notna(purchase_qty) else 0
                except (ValueError, TypeError):
                    errors.append(f'第{idx+2}行: 本期计划采购数量必须是数值')
                    parsed_data.append({**row.to_dict(), 'error': '本期计划采购数量格式错误'})
                    continue
                
                # 验证本期实际投产数量（如果存在）
                if '本期实际投产数量' in row:
                    actual_qty = row.get('本期实际投产数量', 0)
                    try:
                        actual_qty = float(actual_qty) if pd.notna(actual_qty) else 0
                    except (ValueError, TypeError):
                        errors.append(f'第{idx+2}行: 本期实际投产数量必须是数值')
                        parsed_data.append({**row.to_dict(), 'error': '本期实际投产数量格式错误'})
                        continue
                
                parsed_data.append(row.to_dict())
                
            except Exception as e:
                errors.append(f'第{idx+2}行: {str(e)}')
                parsed_data.append({**row.to_dict(), 'error': str(e)})
        
        safe_data = safe_json_convert(pd.DataFrame(parsed_data)) if parsed_data else []
        return jsonify({
            'success': True,
            'data': safe_data,
            'errors': errors
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/export-extracted-data', methods=['GET'])
def export_extracted_data():
    """导出提取结果数据到Excel"""
    try:
        from app.utils.excel_utils import prepare_extracted_export_df, auto_width_excel_columns

        data_type = request.args.get('type', 'manual')  # 'manual' 或 'readonly'
        app_data = get_session_data_manager()
        
        if data_type == 'manual':
            data = app_data.get_data('extracted_data_manual')
            filename_suffix = '手工'
        else:
            data = app_data.get_data('extracted_data')
            filename_suffix = '只读'
        
        data = prepare_extracted_export_df(data, export_type=data_type)
        
        if data.empty:
            return jsonify({
                'success': False,
                'error': f'没有可导出的提取结果数据({filename_suffix})'
            }), 400
        
        sheet_name = '提取结果(手工)' if data_type == 'manual' else '提取结果数据'
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            data.to_excel(writer, sheet_name=sheet_name, index=False)
            auto_width_excel_columns(writer.sheets[sheet_name])
        
        output.seek(0)
        filename = f'提取结果数据({filename_suffix})_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"导出提取结果数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/import-extracted-data', methods=['POST'])
def import_extracted_data():
    """导入提取结果数据"""
    try:
        import pandas as pd
        from datetime import datetime
        
        request_data = request.get_json()
        import_data = request_data.get('data', [])
        import_mode = request_data.get('mode', 'replace')  # replace, merge, append
        
        if not import_data:
            return jsonify({'success': False, 'error': '没有要导入的数据'}), 400
        
        app_data = get_session_data_manager()
        
        # 获取当前手工数据
        current_data = app_data.get_data('extracted_data_manual')
        
        # 转换导入数据为DataFrame
        import_df = pd.DataFrame(import_data)
        
        # 列名映射：将用户友好的列名映射回内部列名
        # 用户界面使用"本期实际投产数量"，但内部存储使用"非限制使用的库存"
        if '本期实际投产数量' in import_df.columns:
            # 将"本期实际投产数量"映射到"非限制使用的库存"（空白保持 NaN，与显式 0 区分）
            import_df['非限制使用的库存'] = pd.to_numeric(import_df['本期实际投产数量'], errors='coerce')
            print(f"📋 列名映射: '本期实际投产数量' → '非限制使用的库存'")
        elif '非限制使用的库存' not in import_df.columns:
            # 如果两个列名都不存在，需要在后续计算中生成
            print(f"⚠️ 警告: 导入数据中缺少'本期实际投产数量'或'非限制使用的库存'列")
        
        # 根据导入模式处理
        # 注意：导入的Excel只包含旧机类别数据，需要保留其他类别数据
        if import_mode == 'replace':
            # "完全覆盖"模式：只覆盖旧机类别，保留其他类别
            if current_data is not None and not current_data.empty:
                # 保留非旧机类别的数据
                non_old_machine_data = current_data[current_data['类别'] != '旧机'].copy() if '类别' in current_data.columns else pd.DataFrame()
                # 合并：导入的旧机数据 + 保留的非旧机数据
                if not non_old_machine_data.empty:
                    final_data = pd.concat([import_df, non_old_machine_data], ignore_index=True)
                    # 重新编号
                    final_data['序号'] = range(1, len(final_data) + 1)
                    print(f"   保留了 {len(non_old_machine_data)} 条非旧机数据")
                else:
                    final_data = import_df
            else:
                final_data = import_df
        elif import_mode == 'merge':
            # 合并更新（根据物料代码，只更新旧机类别）
            if current_data is not None and not current_data.empty:
                final_data = current_data.copy()
                updated_count = 0
                # 只更新旧机类别的记录
                for idx, row in import_df.iterrows():
                    material_code = row.get('物料代码')
                    if not material_code:
                        continue
                    
                    # 查找对应的旧机记录
                    if '类别' in final_data.columns and '物料代码' in final_data.columns:
                        mask = (final_data['物料代码'] == material_code) & (final_data['类别'] == '旧机')
                    else:
                        mask = final_data['物料代码'] == material_code
                    
                    if mask.any():
                        # 更新现有旧机记录
                        for col in import_df.columns:
                            # 跳过"本期实际投产数量"列，因为已经映射到"非限制使用的库存"
                            if col == '本期实际投产数量':
                                continue
                            if col in final_data.columns:
                                final_data.loc[mask, col] = row[col]
                        updated_count += 1
                    else:
                        # 新增旧机记录
                        final_data = pd.concat([final_data, pd.DataFrame([row])], ignore_index=True)
                        updated_count += 1
                
                # 重新编号
                final_data['序号'] = range(1, len(final_data) + 1)
                print(f"   合并模式：更新/新增了 {updated_count} 条旧机记录，保留了其他类别数据")
            else:
                final_data = import_df
        else:  # append
            # 追加模式
            if current_data is not None and not current_data.empty:
                final_data = pd.concat([current_data, import_df], ignore_index=True)
                # 重新编号
                final_data['序号'] = range(1, len(final_data) + 1)
            else:
                final_data = import_df
        
        # 重新计算衍生列（与页面编辑逻辑保持一致）
        if '初始数据' in final_data.columns and '本期计划采购数量' in final_data.columns:
            # 确保数值类型
            final_data['初始数据'] = pd.to_numeric(final_data['初始数据'], errors='coerce').fillna(0)
            final_data['本期计划采购数量'] = pd.to_numeric(final_data['本期计划采购数量'], errors='coerce').fillna(0)
            
            # 计算本期计划投产数量 = 初始数据 + 本期计划采购数量
            final_data['本期计划投产数量'] = final_data['初始数据'] + final_data['本期计划采购数量']
            
            # 处理"非限制使用的库存"（即本期实际投产数量）
            if '非限制使用的库存' in final_data.columns:
                final_data['非限制使用的库存'] = pd.to_numeric(final_data['非限制使用的库存'], errors='coerce')
                # 仅未填写（NaN）时默认等于本期计划投产数量；显式填 0 表示实际为 0，不得覆盖
                mask = pd.isna(final_data['非限制使用的库存'])
                final_data.loc[mask, '非限制使用的库存'] = final_data.loc[mask, '本期计划投产数量']
                final_data['非限制使用的库存'] = final_data['非限制使用的库存'].fillna(0)
                print(f"   '非限制使用的库存'列已存在，仅缺省值已用本期计划投产数量填充")
            else:
                # 如果没有该列，默认使用本期计划投产数量
                final_data['非限制使用的库存'] = final_data['本期计划投产数量']
                print(f"   '非限制使用的库存'列不存在，已使用计划投产数量创建")
        else:
            print(f"⚠️ 警告: 缺少'初始数据'或'本期计划采购数量'列，无法计算衍生列")
        
        # 打印导入后的数据信息用于调试
        print(f"📊 导入后数据统计:")
        print(f"  - 总行数: {len(final_data)}")
        if '类别' in final_data.columns:
            category_counts = final_data['类别'].value_counts().to_dict()
            print(f"  - 各类别记录数: {category_counts}")
            print(f"  - 旧机类别记录数: {len(final_data[final_data['类别'] == '旧机'])}")
        if '非限制使用的库存' in final_data.columns:
            old_machine_data = final_data[final_data['类别'] == '旧机'] if '类别' in final_data.columns else final_data
            print(f"  - 旧机'非限制使用的库存'列前5个值: {old_machine_data['非限制使用的库存'].head().tolist()}")
        
        # 保存数据
        app_data.set_data('extracted_data_manual', final_data)
        app_data.mark_extracted_data_modified()
        
        # 不覆盖extracted_data，保持原始数据用于统计对比
        # 计算引擎会优先使用extracted_data_manual
        
        print(f"✅ 成功导入 {len(import_df)} 条提取结果数据")
        
        return jsonify({
            'success': True,
            'message': '数据导入成功，请返回首页点击"重新计算"使结果链路生效',
            'imported_count': len(import_df)
        })
        
    except Exception as e:
        print(f"导入提取结果数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 价格数据管理 API ====================

def safe_save_price_data(df, operation_name="保存"):
    """
    安全保存价格数据到Excel文件
    
    Args:
        df: 要保存的DataFrame
        operation_name: 操作名称，用于错误提示
    
    Returns:
        tuple: (success: bool, message: str, backup_path: str or None)
    """
    from data.base_data.price_data import PRICE_FILE_PATH, refresh_price_data, backup_price_excel
    import time
    
    backup_path = None
    
    try:
        # 备份原文件到 data/backups/
        if os.path.exists(PRICE_FILE_PATH):
            try:
                backup_path = backup_price_excel(PRICE_FILE_PATH)
            except PermissionError:
                return False, '无法备份原文件，请确保文件未被其他程序打开（如Excel）', None
        
        # 保存新数据 - 添加重试机制
        max_retries = 3
        retry_delay = 0.5
        
        # 移除不含税价列（这是动态计算的，不需要保存到Excel）
        df_to_save = df.copy()
        if '销售单价-不含税(元/KG)' in df_to_save.columns:
            df_to_save = df_to_save.drop(columns=['销售单价-不含税(元/KG)'])
        
        for attempt in range(max_retries):
            try:
                # 尝试保存（只保存含税价）
                df_to_save.to_excel(PRICE_FILE_PATH, index=False, sheet_name='销售价格')
                print(f"✅ 成功{operation_name}价格数据到: {PRICE_FILE_PATH}")
                
                # 刷新缓存
                refresh_price_data()
                
                return True, f'{operation_name}成功', backup_path
                
            except PermissionError as e:
                print(f"⚠️ 第 {attempt + 1} 次保存失败，文件可能被占用")
                
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
                    retry_delay *= 2  # 指数退避
                else:
                    # 最后一次尝试失败，返回友好的错误信息
                    error_msg = (
                        f'无法{operation_name}到文件 "销售价格.xlsx"。\n'
                        f'可能原因：\n'
                        f'1. 文件正在被Excel或其他程序打开\n'
                        f'2. 文件所在目录没有写入权限\n'
                        f'3. 文件被设置为只读\n\n'
                        f'请关闭所有打开该文件的程序后重试。'
                    )
                    
                    if backup_path:
                        error_msg += f'\n\n💾 备份文件已保存: {os.path.basename(backup_path)}'
                    
                    return False, error_msg, backup_path
        
        return False, f'{operation_name}失败', backup_path
        
    except Exception as e:
        error_msg = f'{operation_name}失败: {str(e)}'
        print(f"❌ {error_msg}")
        import traceback
        traceback.print_exc()
        return False, error_msg, backup_path


@data_management_bp.route('/price', methods=['GET'])
def get_price_data():
    """获取销售价格数据"""
    try:
        from data.base_data.price_data import load_price_data, get_data_source
        
        price_df = load_price_data()
        data_source = get_data_source()
        
        if price_df is None or price_df.empty:
            return jsonify({
                'success': True,
                'data': [],
                'total': 0,
                'data_source': data_source,
                'message': '暂无价格数据'
            })
        
        # 确保有序号列
        if '序号' not in price_df.columns:
            price_df.insert(0, '序号', range(1, len(price_df) + 1))
        
        # 转换为JSON格式
        records = []
        for _, row in price_df.iterrows():
            # 获取税率，如果不存在则根据含税单价计算默认值
            tax_rate = row.get('税率')
            if pd.isna(tax_rate) or tax_rate == '':
                price_with_tax = row.get('销售单价(元/KG)', 0)
                tax_rate = 13.0 if (pd.notna(price_with_tax) and float(price_with_tax) > 0) else 0.0
            
            record = {
                '序号': int(row['序号']) if pd.notna(row['序号']) else 0,
                '销售产物名称': str(row.get('销售产物名称', '')) if pd.notna(row.get('销售产物名称')) else '',
                '拆解产物编码': str(row['拆解产物编码']) if pd.notna(row['拆解产物编码']) else '',
                '销售单价(元/KG)': float(row['销售单价(元/KG)']) if pd.notna(row['销售单价(元/KG)']) else 0.0,
                '销售单价-不含税(元/KG)': float(row['销售单价-不含税(元/KG)']) if pd.notna(row['销售单价-不含税(元/KG)']) else 0.0,
                '税率': float(tax_rate) if pd.notna(tax_rate) else 13.0
            }
            records.append(record)
        
        return jsonify({
            'success': True,
            'data': records,
            'total': len(records),
            'data_source': data_source  # 'excel' 或 'builtin'
        })
        
    except Exception as e:
        print(f"获取价格数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/price', methods=['PUT'])
def update_price_data():
    """更新价格数据（同时更新Excel和.py文件）"""
    try:
        from data.base_data.price_data import load_price_data, refresh_price_data, PRICE_FILE_PATH
        import os
        
        data = request.get_json()
        
        if not data or 'modifications' not in data:
            return jsonify({'success': False, 'error': '缺少修改数据'}), 400
        
        modifications = data['modifications']
        
        if not modifications:
            return jsonify({'success': False, 'error': '修改数据为空'}), 400
        
        # 备份.py文件
        backup_price_file()
        
        # 加载当前价格数据
        price_df = load_price_data()
        
        if price_df is None:
            return jsonify({'success': False, 'error': '无法加载价格数据'}), 500
        
        # 确保有序号列
        if '序号' not in price_df.columns:
            price_df.insert(0, '序号', range(1, len(price_df) + 1))
        
        # 应用修改
        updated_count = 0
        updated_records = []
        for mod in modifications:
            row_id = mod.get('序号')
            if row_id is None:
                continue
            
            # 找到对应的行
            mask = price_df['序号'] == row_id
            if not mask.any():
                continue
            
            # 获取原始产物编码
            original_code = str(price_df.loc[mask, '拆解产物编码'].iloc[0])
            
            # 更新可编辑的字段
            if '销售产物名称' in mod:
                price_df.loc[mask, '销售产物名称'] = mod['销售产物名称']
            if '拆解产物编码' in mod:
                price_df.loc[mask, '拆解产物编码'] = str(mod['拆解产物编码'])
            
            # 处理税率更新
            tax_rate_changed = False
            if '税率' in mod:
                new_tax_rate = float(mod['税率']) if mod['税率'] != '' else None
                if new_tax_rate is not None:
                    price_df.loc[mask, '税率'] = new_tax_rate
                    tax_rate_changed = True
            
            # 处理含税价更新
            price_changed = False
            if '销售单价(元/KG)' in mod:
                new_price = float(mod['销售单价(元/KG)'])
                price_df.loc[mask, '销售单价(元/KG)'] = new_price
                price_changed = True
                # 如果税率未更新，根据含税单价设置默认税率
                if not tax_rate_changed:
                    default_tax_rate = 13.0 if new_price > 0 else 0.0
                    if '税率' not in price_df.columns or price_df.loc[mask, '税率'].isna().any():
                        price_df.loc[mask, '税率'] = default_tax_rate
            
            # 重新计算不含税价（如果税率或含税价发生变化）
            if tax_rate_changed or price_changed:
                from data.base_data.price_data import _calculate_price_no_tax
                current_price = price_df.loc[mask, '销售单价(元/KG)'].iloc[0]
                current_tax_rate = price_df.loc[mask, '税率'].iloc[0] if '税率' in price_df.columns else 13.0
                price_df.loc[mask, '销售单价-不含税(元/KG)'] = _calculate_price_no_tax(current_price, current_tax_rate)
            
            # 准备更新.py文件的记录
            row_data = price_df.loc[mask].iloc[0]
            tax_rate = row_data.get('税率', 13.0)
            if pd.isna(tax_rate):
                tax_rate = 13.0
            updated_record = {
                '序号': int(row_data['序号']),
                '销售产物名称': str(row_data.get('销售产物名称', '')),
                '拆解产物编码': str(row_data['拆解产物编码']),
                '销售单价(元/KG)': float(row_data['销售单价(元/KG)']),
                '税率': float(tax_rate)
            }
            updated_records.append((original_code, updated_record))
            updated_count += 1
        
        # 保存到Excel文件（如果存在）
        if os.path.exists(PRICE_FILE_PATH):
            success, message, backup_path = safe_save_price_data(price_df, "更新")
            if not success:
                return jsonify({'success': False, 'error': f'Excel保存失败: {message}'}), 500
        
        # 更新.py文件中的内置数据
        file_update_success = True
        for original_code, updated_record in updated_records:
            if not update_price_record_in_file(original_code, updated_record):
                file_update_success = False
                print(f"警告: 更新.py文件中的记录失败: {original_code}")
        
        # 刷新缓存
        refresh_price_data()
        
        # 自动触发可销售量数据重新计算（因为价格变化会影响收益计算）
        recalc_success = False
        recalc_message = ""
        if updated_count > 0:
            try:
                from app.core.calculation_engine import CalculationEngine
                calculation_engine = CalculationEngine()
                
                print("🔄 价格数据已更新，自动重新计算可销售量数据...")
                calculation_engine.merge_saleable_data()
                print("✅ 可销售量数据已更新（包含新的价格和收益信息）")
                
                recalc_success = True
                recalc_message = "，可销售量数据（含收益）已自动重新计算"
            except Exception as calc_error:
                print(f"⚠️ 自动重新计算可销售量数据失败: {calc_error}")
                import traceback
                traceback.print_exc()
                recalc_message = "，但可销售量数据重新计算失败"
        
        if file_update_success:
            return jsonify({
                'success': True,
                'message': f'成功更新 {updated_count} 条记录（Excel和内置数据）{recalc_message}',
                'recalc_success': recalc_success
            })
        else:
            return jsonify({
                'success': True,
                'message': f'成功更新 {updated_count} 条记录（Excel），但部分内置数据更新失败{recalc_message}',
                'recalc_success': recalc_success
            })
        
    except Exception as e:
        print(f"更新价格数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/price', methods=['POST'])
def create_price_record():
    """新增价格记录（同时添加到Excel和.py文件）"""
    try:
        from data.base_data.price_data import load_price_data, refresh_price_data, PRICE_FILE_PATH
        import os
        
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'error': '缺少数据'}), 400
        
        # 备份.py文件
        backup_price_file()
        
        # 加载当前价格数据
        price_df = load_price_data()
        
        if price_df is None:
            # 如果没有数据，创建新的DataFrame
            price_df = pd.DataFrame(columns=['序号', '销售产物名称', '拆解产物编码', '销售单价(元/KG)', '税率'])
        
        # 确保有序号列
        if '序号' not in price_df.columns:
            price_df.insert(0, '序号', range(1, len(price_df) + 1))
        
        # 创建新记录
        price_with_tax = float(data.get('销售单价(元/KG)', 0))
        # 获取税率，如果没有提供则根据含税单价设置默认值
        tax_rate = data.get('税率')
        if tax_rate is None or tax_rate == '':
            tax_rate = 13.0 if price_with_tax > 0 else 0.0
        else:
            tax_rate = float(tax_rate)
        
        # 计算不含税价格
        from data.base_data.price_data import _calculate_price_no_tax
        price_no_tax = _calculate_price_no_tax(price_with_tax, tax_rate)
        
        new_row = {
            '序号': len(price_df) + 1,
            '销售产物名称': data.get('销售产物名称', ''),
            '拆解产物编码': str(data.get('拆解产物编码', '')),
            '销售单价(元/KG)': price_with_tax,
            '销售单价-不含税(元/KG)': price_no_tax,
            '税率': tax_rate
        }
        
        # 检查产物编码是否已存在
        if not price_df.empty and str(new_row['拆解产物编码']) in price_df['拆解产物编码'].astype(str).values:
            return jsonify({'success': False, 'error': '该产物编码已存在'}), 400
        
        # 添加新行
        price_df = pd.concat([price_df, pd.DataFrame([new_row])], ignore_index=True)
        
        # 重新编号
        price_df['序号'] = range(1, len(price_df) + 1)
        
        # 更新 new_row 的序号
        new_row['序号'] = len(price_df)
        
        # 保存到Excel文件（如果路径存在或需要创建）
        try:
            success, message, backup_path = safe_save_price_data(price_df, "新增")
            if not success:
                return jsonify({'success': False, 'error': f'Excel保存失败: {message}'}), 500
        except Exception as e:
            print(f"Excel保存失败: {str(e)}")
        
        # 添加到.py文件中的内置数据
        file_record = {
            '序号': new_row['序号'],
            '销售产物名称': new_row['销售产物名称'],
            '拆解产物编码': new_row['拆解产物编码'],
            '销售单价(元/KG)': new_row['销售单价(元/KG)'],
            '税率': new_row['税率']
        }
        
        file_success = add_price_record_to_file(file_record)
        
        # 刷新缓存
        refresh_price_data()
        
        # 自动触发可销售量数据重新计算（因为新增价格会影响收益计算）
        recalc_success = False
        recalc_message = ""
        try:
            from app.core.calculation_engine import CalculationEngine
            calculation_engine = CalculationEngine()
            
            print("🔄 价格数据已新增，自动重新计算可销售量数据...")
            calculation_engine.merge_saleable_data()
            print("✅ 可销售量数据已更新（包含新的价格和收益信息）")
            
            recalc_success = True
            recalc_message = "，可销售量数据（含收益）已自动重新计算"
        except Exception as calc_error:
            print(f"⚠️ 自动重新计算可销售量数据失败: {calc_error}")
            import traceback
            traceback.print_exc()
            recalc_message = "，但可销售量数据重新计算失败"
        
        if file_success:
            return jsonify({
                'success': True,
                'message': f'新增价格记录成功（Excel和内置数据）{recalc_message}',
                'record': new_row,
                'recalc_success': recalc_success
            })
        else:
            return jsonify({
                'success': True,
                'message': f'新增价格记录成功（Excel），但内置数据更新失败{recalc_message}',
                'record': new_row,
                'recalc_success': recalc_success
            })
        
    except Exception as e:
        print(f"新增价格记录失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/price/<product_code>', methods=['DELETE'])
def delete_price_record(product_code):
    """删除价格记录（同时从Excel和.py文件删除）"""
    try:
        from data.base_data.price_data import load_price_data, refresh_price_data, PRICE_FILE_PATH
        import os
        
        # 备份.py文件
        backup_price_file()
        
        # 加载当前价格数据
        price_df = load_price_data()
        
        if price_df is None or price_df.empty:
            return jsonify({'success': False, 'error': '没有价格数据'}), 404
        
        # 查找要删除的记录
        mask = price_df['拆解产物编码'].astype(str) == str(product_code)
        
        if not mask.any():
            return jsonify({'success': False, 'error': '未找到该产物编码'}), 404
        
        # 删除记录
        price_df = price_df[~mask].copy()
        
        # 重新编号
        if not price_df.empty:
            price_df['序号'] = range(1, len(price_df) + 1)
        
        # 从Excel保存（如果存在）
        if os.path.exists(PRICE_FILE_PATH):
            success, message, backup_path = safe_save_price_data(price_df, "删除")
            if not success:
                return jsonify({'success': False, 'error': f'Excel保存失败: {message}'}), 500
        
        # 从.py文件删除
        file_success = delete_price_record_from_file(str(product_code))
        
        # 刷新缓存
        refresh_price_data()
        
        # 自动触发可销售量数据重新计算（因为删除价格会影响收益计算）
        recalc_success = False
        recalc_message = ""
        try:
            from app.core.calculation_engine import CalculationEngine
            calculation_engine = CalculationEngine()
            
            print("🔄 价格数据已删除，自动重新计算可销售量数据...")
            calculation_engine.merge_saleable_data()
            print("✅ 可销售量数据已更新（包含新的价格和收益信息）")
            
            recalc_success = True
            recalc_message = "，可销售量数据（含收益）已自动重新计算"
        except Exception as calc_error:
            print(f"⚠️ 自动重新计算可销售量数据失败: {calc_error}")
            import traceback
            traceback.print_exc()
            recalc_message = "，但可销售量数据重新计算失败"
        
        if file_success:
            return jsonify({
                'success': True,
                'message': f'删除价格记录成功（Excel和内置数据）{recalc_message}',
                'recalc_success': recalc_success
            })
        else:
            return jsonify({
                'success': True,
                'message': f'删除价格记录成功（Excel），但内置数据删除失败{recalc_message}',
                'recalc_success': recalc_success
            })
        
    except Exception as e:
        print(f"删除价格记录失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def backup_price_file():
    """备份价格数据文件（price_data.py）"""
    try:
        import shutil
        from datetime import datetime
        
        # 创建备份目录
        backup_dir = 'backups'
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        # 生成备份文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"price_data_backup_{timestamp}.py"
        backup_path = os.path.join(backup_dir, backup_filename)
        
        # 复制文件
        source_path = 'data/base_data/price_data.py'
        shutil.copy2(source_path, backup_path)
        print(f"✓ 价格数据文件已备份到: {backup_path}")
        return True, backup_path
        
    except Exception as e:
        print(f"✗ 备份价格数据文件失败: {e}")
        return False, None


def update_price_record_in_file(product_code, new_record):
    """更新价格数据文件中的记录"""
    try:
        import re
        import importlib
        
        # 读取当前文件内容
        price_file_path = 'data/base_data/price_data.py'
        with open(price_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 转义特殊字符
        name = str(new_record.get("销售产物名称", "")).replace('"', '\\"')
        code = str(new_record["拆解产物编码"]).replace('"', '\\"')
        price = float(new_record["销售单价(元/KG)"])
        tax_rate = float(new_record.get("税率", 13.0))
        seq = int(new_record.get("序号", 0))
        
        # 找到并替换对应的记录
        # 匹配模式：{"序号": X, "销售产物名称": "...", "拆解产物编码": "PRODUCT_CODE", "销售单价(元/KG)": X.X, "税率": X.X}
        old_pattern = rf'    \{{"序号":\s*\d+,\s*"销售产物名称":\s*"[^"]*",\s*"拆解产物编码":\s*"{re.escape(product_code)}",\s*"销售单价\(元/KG\)":\s*[0-9.\-]+,\s*"税率":\s*[0-9.]+\}},?\n'
        new_line = f'    {{"序号": {seq}, "销售产物名称": "{name}", "拆解产物编码": "{code}", "销售单价(元/KG)": {price}, "税率": {tax_rate}}},\n'
        
        new_content = re.sub(old_pattern, new_line, content)
        
        # 检查是否找到并替换了记录
        if new_content == content:
            print(f"警告: 未找到产物编码为 {product_code} 的记录，尝试添加新记录")
            return add_price_record_to_file(new_record)
        
        # 写回文件
        with open(price_file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        # 重新加载模块
        from data.base_data import price_data
        importlib.reload(price_data)
        
        print(f"✓ 已更新价格记录: {product_code} -> {new_record['拆解产物编码']}")
        return True
        
    except Exception as e:
        print(f"✗ 更新价格记录失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def add_price_record_to_file(new_record):
    """将新的价格记录添加到文件中"""
    try:
        import re
        import importlib
        
        # 读取当前文件内容
        price_file_path = 'data/base_data/price_data.py'
        with open(price_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 转义特殊字符
        name = str(new_record.get("销售产物名称", "")).replace('"', '\\"')
        code = str(new_record["拆解产物编码"]).replace('"', '\\"')
        price = float(new_record["销售单价(元/KG)"])
        tax_rate = float(new_record.get("税率", 13.0))
        seq = int(new_record.get("序号", 1))
        
        # 创建新记录行
        new_line = f'    {{"序号": {seq}, "销售产物名称": "{name}", "拆解产物编码": "{code}", "销售单价(元/KG)": {price}, "税率": {tax_rate}}},\n'
        
        # 找到 BUILTIN_PRICE_DATA 列表的结束位置（最后一个 ] 之前）
        pattern = r'(    \{"序号":[^}]+\},?\n)(\s*\])'
        
        def replace_func(match):
            last_record = match.group(1)
            closing_bracket = match.group(2)
            
            # 确保最后一个记录有逗号
            if not last_record.rstrip().endswith(','):
                last_record = last_record.rstrip() + ',\n'
            
            return last_record + new_line + closing_bracket
        
        new_content = re.sub(pattern, replace_func, content, flags=re.MULTILINE | re.DOTALL)
        
        # 检查是否成功替换
        if new_content == content:
            raise ValueError("无法找到合适的位置插入新记录")
        
        # 写回文件
        with open(price_file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        # 重新加载模块
        from data.base_data import price_data
        importlib.reload(price_data)
        
        print(f"✓ 已添加价格记录: {new_record['拆解产物编码']}")
        return True
        
    except Exception as e:
        print(f"✗ 添加价格记录失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def delete_price_record_from_file(product_code):
    """从价格数据文件中删除记录"""
    try:
        import re
        import importlib
        
        # 读取当前文件内容
        price_file_path = 'data/base_data/price_data.py'
        with open(price_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 找到并删除对应的记录
        pattern = rf'    \{{"序号":\s*\d+,\s*"销售产物名称":\s*"[^"]*",\s*"拆解产物编码":\s*"{re.escape(product_code)}",\s*"销售单价\(元/KG\)":\s*[0-9.]+,\s*"备注":\s*"[^"]*"\}},?\n'
        new_content = re.sub(pattern, '', content)
        
        # 检查是否找到并删除了记录
        if new_content == content:
            raise ValueError(f"未找到产物编码为 {product_code} 的记录")
        
        # 写回文件
        with open(price_file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        # 重新加载模块
        from data.base_data import price_data
        importlib.reload(price_data)
        
        print(f"✓ 已删除价格记录: {product_code}")
        return True
        
    except Exception as e:
        print(f"✗ 删除价格记录失败: {e}")
        import traceback
        traceback.print_exc()
        return False


@data_management_bp.route('/price/export', methods=['GET'])
def export_price_data():
    """导出价格数据为Excel"""
    try:
        from data.base_data.price_data import load_price_data
        
        price_df = load_price_data()
        
        if price_df is None or price_df.empty:
            return jsonify({'success': False, 'error': '没有价格数据可导出'}), 400
        
        # 创建Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            price_df.to_excel(writer, index=False, sheet_name='销售价格')
        
        output.seek(0)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'销售价格_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"导出价格数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/price/export-builtin-to-excel', methods=['POST'])
def export_builtin_to_excel():
    """将内置价格数据导出为Excel文件（保存到默认路径）"""
    try:
        from data.base_data.price_data import export_builtin_data_to_excel, PRICE_FILE_PATH, refresh_price_data
        
        if export_builtin_data_to_excel():
            # 刷新缓存以使用新的Excel文件
            refresh_price_data()
            return jsonify({
                'success': True,
                'message': f'成功将内置数据导出到Excel: {PRICE_FILE_PATH}'
            })
        else:
            return jsonify({'success': False, 'error': '导出失败'}), 500
        
    except Exception as e:
        print(f"导出内置数据到Excel失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/price/import', methods=['POST'])
def import_price_data():
    """导入价格数据"""
    try:
        from data.base_data.price_data import refresh_price_data, PRICE_FILE_PATH
        import os
        
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '没有上传文件'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'error': '文件名为空'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': '只支持Excel文件'}), 400
        
        # 读取上传的文件
        df = pd.read_excel(file)
        
        # 验证必需的列
        required_columns = ['拆解产物编码', '销售单价(元/KG)']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return jsonify({
                'success': False,
                'error': f'文件缺少必需的列: {", ".join(missing_columns)}'
            }), 400
        
        # 确保有序号列
        if '序号' not in df.columns:
            df.insert(0, '序号', range(1, len(df) + 1))
        
        # 确保数据类型正确
        df['拆解产物编码'] = df['拆解产物编码'].astype(str)
        df['销售单价(元/KG)'] = pd.to_numeric(df['销售单价(元/KG)'], errors='coerce').fillna(0)
        
        # 如果没有销售产物名称列，添加空列
        if '销售产物名称' not in df.columns:
            df['销售产物名称'] = ''
        
        # 处理税率列：如果存在"备注"列，转换为"税率"列
        if '备注' in df.columns and '税率' not in df.columns:
            from data.base_data.price_data import _get_default_tax_rate
            df['税率'] = df['销售单价(元/KG)'].apply(_get_default_tax_rate)
            df = df.drop(columns=['备注'])
        elif '税率' not in df.columns:
            # 如果没有税率列，根据含税单价设置默认税率
            from data.base_data.price_data import _get_default_tax_rate
            df['税率'] = df['销售单价(元/KG)'].apply(_get_default_tax_rate)
        else:
            # 确保税率列为数值类型，并处理缺失值
            from data.base_data.price_data import _get_default_tax_rate
            df['税率'] = pd.to_numeric(df['税率'], errors='coerce')
            mask = df['税率'].isna()
            df.loc[mask, '税率'] = df.loc[mask, '销售单价(元/KG)'].apply(_get_default_tax_rate)
        
        # 计算不含税价格
        from data.base_data.price_data import _calculate_price_no_tax
        df['销售单价-不含税(元/KG)'] = df.apply(
            lambda row: _calculate_price_no_tax(row['销售单价(元/KG)'], row['税率']), axis=1
        )
        
        # 保存新数据
        success, message, backup_path = safe_save_price_data(df, "导入")
        
        if success:
            # 刷新缓存
            refresh_price_data()
            
            # 自动触发可销售量数据重新计算（因为导入价格会影响收益计算）
            recalc_success = False
            recalc_message = ""
            try:
                from app.core.calculation_engine import CalculationEngine
                calculation_engine = CalculationEngine()
                
                print("🔄 价格数据已导入，自动重新计算可销售量数据...")
                calculation_engine.merge_saleable_data()
                print("✅ 可销售量数据已更新（包含新的价格和收益信息）")
                
                recalc_success = True
                recalc_message = "，可销售量数据（含收益）已自动重新计算"
            except Exception as calc_error:
                print(f"⚠️ 自动重新计算可销售量数据失败: {calc_error}")
                import traceback
                traceback.print_exc()
                recalc_message = "，但可销售量数据重新计算失败"
            
            return jsonify({
                'success': True,
                'message': f'成功导入 {len(df)} 条价格记录{recalc_message}',
                'recalc_success': recalc_success
            })
        else:
            return jsonify({'success': False, 'error': message}), 500
        
    except Exception as e:
        print(f"导入价格数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'导入失败: {str(e)}'}), 500


@data_management_bp.route('/price/statistics', methods=['GET'])
def get_price_statistics():
    """获取价格数据统计信息"""
    try:
        from data.base_data.price_data import get_price_statistics
        
        stats = get_price_statistics()
        
        return jsonify({
            'success': True,
            'statistics': stats
        })
        
    except Exception as e:
        print(f"获取价格统计信息失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 基金补贴单价管理 API ====================

@data_management_bp.route('/subsidy', methods=['GET'])
def get_subsidy_data():
    """获取基金补贴单价数据"""
    try:
        from data.base_data.subsidy_data import load_subsidy_data, get_data_source
        
        subsidy_df = load_subsidy_data()
        
        if subsidy_df is None:
            return jsonify({
                'success': False,
                'error': '加载补贴单价数据失败'
            }), 500
        
        # 转换为字典格式
        data_dict = safe_json_convert(subsidy_df)
        
        return jsonify({
            'success': True,
            'data': data_dict,
            'data_source': get_data_source(),
            'message': '补贴单价数据获取成功'
        })
        
    except Exception as e:
        print(f"获取补贴单价数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/subsidy', methods=['PUT'])
def update_subsidy_data():
    """更新补贴单价数据"""
    try:
        from data.base_data.subsidy_data import load_subsidy_data, SUBSIDY_FILE_PATH, refresh_subsidy_data
        
        data = request.get_json()
        modifications = data.get('modifications', [])
        
        if not modifications:
            return jsonify({
                'success': False,
                'error': '没有提供修改数据'
            }), 400
        
        # 加载当前数据
        subsidy_df = load_subsidy_data()
        
        if subsidy_df is None:
            return jsonify({
                'success': False,
                'error': '加载补贴单价数据失败'
            }), 500
        
        # 应用修改
        for mod in modifications:
            row_id = mod.get('序号')
            if row_id is None:
                continue
            
            # 查找对应的行
            mask = subsidy_df['序号'] == row_id
            if not mask.any():
                continue
            
            # 更新数据
            for key, value in mod.items():
                if key in subsidy_df.columns and key != '序号':
                    subsidy_df.loc[mask, key] = value
        
        # 保存到Excel文件
        subsidy_df.to_excel(SUBSIDY_FILE_PATH, index=False)
        
        # 刷新缓存
        refresh_subsidy_data()
        
        return jsonify({
            'success': True,
            'message': f'成功更新 {len(modifications)} 条补贴单价数据'
        })
        
    except Exception as e:
        print(f"更新补贴单价数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/subsidy/import', methods=['POST'])
def import_subsidy_data():
    """导入补贴单价数据"""
    try:
        from data.base_data.subsidy_data import SUBSIDY_FILE_PATH, refresh_subsidy_data
        
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': '没有上传文件'
            }), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({
                'success': False,
                'error': '文件名为空'
            }), 400
        
        # 读取Excel文件
        df = pd.read_excel(file)
        
        # 验证必需的列
        required_columns = ['类别', '补贴单价(元/台)']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return jsonify({
                'success': False,
                'error': f'Excel文件缺少必需的列: {", ".join(missing_columns)}'
            }), 400
        
        # 确保有序号列
        if '序号' not in df.columns:
            df.insert(0, '序号', range(1, len(df) + 1))
        
        # 确保有备注列
        if '备注' not in df.columns:
            df['备注'] = ''
        
        # 保存到文件
        df.to_excel(SUBSIDY_FILE_PATH, index=False)
        
        # 刷新缓存
        refresh_subsidy_data()
        
        return jsonify({
            'success': True,
            'message': f'成功导入 {len(df)} 条补贴单价数据'
        })
        
    except Exception as e:
        print(f"导入补贴单价数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/subsidy/export', methods=['GET'])
def export_subsidy_data():
    """导出补贴单价数据"""
    try:
        from data.base_data.subsidy_data import load_subsidy_data
        
        subsidy_df = load_subsidy_data()
        
        if subsidy_df is None or subsidy_df.empty:
            return jsonify({
                'success': False,
                'error': '没有可导出的补贴单价数据'
            }), 400
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            subsidy_df.to_excel(writer, sheet_name='基金补贴单价', index=False)
            
            # 设置列宽
            worksheet = writer.sheets['基金补贴单价']
            worksheet.column_dimensions['A'].width = 12  # 序号
            worksheet.column_dimensions['B'].width = 20  # 类别
            worksheet.column_dimensions['C'].width = 20  # 补贴单价(元/台)
            worksheet.column_dimensions['D'].width = 30  # 备注
        
        output.seek(0)
        filename = f'基金补贴单价_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"导出补贴单价数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/subsidy/export-builtin-to-excel', methods=['POST'])
def export_builtin_subsidy_to_excel():
    """将内置补贴数据导出为Excel文件"""
    try:
        from data.base_data.subsidy_data import export_builtin_data_to_excel
        
        success = export_builtin_data_to_excel()
        
        if success:
            return jsonify({
                'success': True,
                'message': '内置补贴数据已成功保存为Excel文件'
            })
        else:
            return jsonify({
                'success': False,
                'error': '保存失败，请查看服务器日志'
            }), 500
            
    except Exception as e:
        print(f"导出内置补贴数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/subsidy/statistics', methods=['GET'])
def get_subsidy_statistics():
    """获取补贴数据统计信息"""
    try:
        from data.base_data.subsidy_data import get_subsidy_statistics
        
        stats = get_subsidy_statistics()
        
        return jsonify({
            'success': True,
            'statistics': stats
        })
        
    except Exception as e:
        print(f"获取补贴统计信息失败: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 旧机拆解人工提成单价管理 ====================

def get_labor_cost_dataframe():
    """获取计件人工标准DataFrame"""
    try:
        # 导入模块并重新加载，确保获取最新数据
        import importlib
        from data.base_data import labor_cost_data
        
        # 强制重新加载模块，清除Python的模块缓存
        importlib.reload(labor_cost_data)
        
        # 调用函数获取数据
        df = labor_cost_data.get_labor_cost_dataframe()
        
        # 验证DataFrame
        if df is None:
            print("错误: get_labor_cost_dataframe() 返回 None")
            return pd.DataFrame()
        
        if not isinstance(df, pd.DataFrame):
            print(f"错误: get_labor_cost_dataframe() 返回了非DataFrame类型: {type(df)}")
            return pd.DataFrame()
        
        print(f"成功加载数据: {len(df)} 条记录")
        return df
        
    except ImportError as e:
        print(f"导入错误: {e}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame()
    except Exception as e:
        print(f"获取计件人工标准数据失败: {e}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame()

def backup_labor_cost_file():
    """备份人工提成单价文件"""
    try:
        import shutil
        from datetime import datetime
        
        # 创建备份目录
        backup_dir = 'backups'
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        # 生成备份文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"labor_cost_data_backup_{timestamp}.py"
        backup_path = os.path.join(backup_dir, backup_filename)
        
        # 复制文件
        source_path = 'data/base_data/labor_cost_data.py'
        shutil.copy2(source_path, backup_path)
        print(f"✓ 人工提成单价文件已备份到: {backup_path}")
        return True
        
    except Exception as e:
        print(f"✗ 备份人工提成单价文件失败: {e}")
        return False

def add_labor_cost_record_to_file(new_record):
    """将新的计件人工标准记录添加到文件中"""
    try:
        import re
        import importlib
        
        # 读取当前文件内容
        labor_cost_file_path = 'data/base_data/labor_cost_data.py'
        with open(labor_cost_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 转义特殊字符
        category = str(new_record["类别"]).replace('"', '\\"').replace('\\', '\\\\')
        code = str(new_record["R3系统代码"]).replace('"', '\\"').replace('\\', '\\\\')
        name = str(new_record["系统名称"]).replace('"', '\\"').replace('\\', '\\\\')
        
        # 获取所有单价字段
        price_fields = [
            "生产计件单价", "品管提成单价", "物流主管提成单价", "物流卸货提成单价",
            "班组长提成单价", "生产主管提成单价", "维修班长提成单价", "维修员提成单价",
            "冰箱维修主管提成单价", "叉车司磅库管等提成单价"
        ]
        
        # 构建记录字典
        record_parts = [
            f'"类别": "{category}"',
            f'"R3系统代码": "{code}"',
            f'"系统名称": "{name}"'
        ]
        
        for price_field in price_fields:
            price_val = float(new_record.get(price_field, 0.0))
            record_parts.append(f'"{price_field}": {price_val}')
        
        record_str = '{' + ', '.join(record_parts) + '}'
        new_line = f'    {record_str},\n'
        
        # 找到列表的结束位置（最后一个 ] 之前）
        pattern = r'(\s*\{"类别":[^}]+\},?\n)(\s*\])'
        
        def replace_func(match):
            last_record = match.group(1)
            closing_bracket = match.group(2)
            
            # 确保最后一个记录有逗号
            if not last_record.rstrip().endswith(','):
                last_record = last_record.rstrip() + ',\n'
            
            return last_record + new_line + closing_bracket
        
        new_content = re.sub(pattern, replace_func, content, flags=re.MULTILINE | re.DOTALL)
        
        # 检查是否成功替换
        if new_content == content:
            # 如果列表为空，需要在列表开始后插入
            pattern2 = r'(LABOR_COST_DATA = \[)(\s*\])'
            new_content = re.sub(pattern2, r'\1\n' + new_line + r'\2', content)
            if new_content == content:
                raise ValueError("无法找到合适的位置插入新记录")
        
        # 写回文件
        with open(labor_cost_file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        # 重新加载模块
        from data.base_data import labor_cost_data
        importlib.reload(labor_cost_data)
        
        print(f"✓ 已添加计件人工标准记录: {new_record['R3系统代码']}")
        return True
        
    except Exception as e:
        print(f"✗ 添加计件人工标准记录失败: {e}")
        traceback.print_exc()
        return False

def update_labor_cost_record_in_file(old_code, new_record):
    """更新计件人工标准文件中的记录"""
    try:
        import re
        import importlib
        
        # 读取当前文件内容
        labor_cost_file_path = 'data/base_data/labor_cost_data.py'
        with open(labor_cost_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 转义特殊字符
        category = str(new_record["类别"]).replace('"', '\\"').replace('\\', '\\\\')
        code = str(new_record["R3系统代码"]).replace('"', '\\"').replace('\\', '\\\\')
        name = str(new_record["系统名称"]).replace('"', '\\"').replace('\\', '\\\\')
        
        # 获取所有单价字段
        price_fields = [
            "生产计件单价", "品管提成单价", "物流主管提成单价", "物流卸货提成单价",
            "班组长提成单价", "生产主管提成单价", "维修班长提成单价", "维修员提成单价",
            "冰箱维修主管提成单价", "叉车司磅库管等提成单价"
        ]
        
        # 构建记录字典
        record_parts = [
            f'"类别": "{category}"',
            f'"R3系统代码": "{code}"',
            f'"系统名称": "{name}"'
        ]
        
        for price_field in price_fields:
            price_val = float(new_record.get(price_field, 0.0))
            record_parts.append(f'"{price_field}": {price_val}')
        
        record_str = '{' + ', '.join(record_parts) + '}'
        new_line = f'    {record_str},\n'
        
        # 找到并替换对应的记录（匹配包含R3系统代码的记录）
        old_pattern = rf'    \{{[^}}]*"R3系统代码":\s*"{re.escape(old_code)}"[^}}]*}},?\n'
        
        new_content = re.sub(old_pattern, new_line, content)
        
        # 检查是否找到并替换了记录
        if new_content == content:
            raise ValueError(f"未找到R3代码为 {old_code} 的记录")
        
        # 写回文件
        with open(labor_cost_file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        # 重新加载模块
        from data.base_data import labor_cost_data
        importlib.reload(labor_cost_data)
        
        print(f"✓ 已更新计件人工标准记录: {old_code} -> {new_record['R3系统代码']}")
        return True
        
    except Exception as e:
        print(f"✗ 更新计件人工标准记录失败: {e}")
        traceback.print_exc()
        return False

def delete_labor_cost_record_from_file(code):
    """从计件人工标准文件中删除记录"""
    try:
        import re
        import importlib
        
        # 读取当前文件内容
        labor_cost_file_path = 'data/base_data/labor_cost_data.py'
        with open(labor_cost_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 找到并删除对应的记录（匹配包含R3系统代码的记录）
        pattern = rf'    \{{[^}}]*"R3系统代码":\s*"{re.escape(code)}"[^}}]*}},?\n'
        new_content = re.sub(pattern, '', content)
        
        # 检查是否找到并删除了记录
        if new_content == content:
            raise ValueError(f"未找到R3代码为 {code} 的记录")
        
        # 写回文件
        with open(labor_cost_file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        # 重新加载模块
        from data.base_data import labor_cost_data
        importlib.reload(labor_cost_data)
        
        print(f"✓ 已删除计件人工标准记录: {code}")
        return True
        
    except Exception as e:
        print(f"✗ 删除计件人工标准记录失败: {e}")
        traceback.print_exc()
        return False

def import_labor_cost_data_to_file(records, mode='append'):
    """批量导入计件人工标准数据到文件"""
    try:
        price_fields = [
            "生产计件单价", "品管提成单价", "物流主管提成单价", "物流卸货提成单价",
            "班组长提成单价", "生产主管提成单价", "维修班长提成单价", "维修员提成单价",
            "冰箱维修主管提成单价", "叉车司磅库管等提成单价"
        ]
        
        if mode == 'replace':
            # 覆盖模式：重新生成整个文件
            new_content = '''# 计件人工标准数据 - 内置数据
# 数据来源：计件人工标准.xlsx
# 单价单位：元/台（TAI）

import pandas as pd
import os

# 完整的计件人工标准数据
LABOR_COST_DATA = [
'''
            
            # 添加所有记录
            for record in records:
                category = str(record['类别']).replace('"', '\\"').replace('\\', '\\\\')
                code = str(record['R3系统代码']).replace('"', '\\"').replace('\\', '\\\\')
                name = str(record['系统名称']).replace('"', '\\"').replace('\\', '\\\\')
                
                record_parts = [
                    f'"类别": "{category}"',
                    f'"R3系统代码": "{code}"',
                    f'"系统名称": "{name}"'
                ]
                
                for price_field in price_fields:
                    price_val = float(record.get(price_field, 0.0))
                    record_parts.append(f'"{price_field}": {price_val}')
                
                record_str = '{' + ', '.join(record_parts) + '}'
                new_content += f'    {record_str},\n'
            
            new_content += ''']

def get_labor_cost_dataframe():
    """获取计件人工标准DataFrame"""
    return pd.DataFrame(LABOR_COST_DATA)

def filter_by_category(category):
    """根据类别筛选数据"""
    df = get_labor_cost_dataframe()
    if df.empty:
        return df
    return df[df['类别'] == category]

def get_all_categories():
    """获取所有类别列表"""
    df = get_labor_cost_dataframe()
    if df.empty:
        return []
    return df['类别'].unique().tolist()

def get_category_stats():
    """获取类别统计信息"""
    df = get_labor_cost_dataframe()
    if df.empty:
        return {}
    return df['类别'].value_counts().to_dict()

def get_labor_cost_by_code(code):
    """根据R3系统代码获取计件人工标准"""
    df = get_labor_cost_dataframe()
    if df.empty:
        return None
    result = df[df['R3系统代码'].astype(str) == str(code)]
    if len(result) > 0:
        return result.iloc[0].to_dict()
    return None
'''
            
            # 写入文件
            labor_cost_file_path = 'data/base_data/labor_cost_data.py'
            print(f"[导入] 开始写入文件(覆盖模式): {labor_cost_file_path}")
            with open(labor_cost_file_path, 'w', encoding='utf-8') as f:
                f.write(new_content)
            
            # 验证文件是否写入成功
            import os
            if os.path.exists(labor_cost_file_path):
                file_size = os.path.getsize(labor_cost_file_path)
                print(f"[导入] 文件写入成功，文件大小: {file_size} 字节")
            else:
                print(f"[导入] 错误: 文件不存在: {labor_cost_file_path}")
                return False
            
            # 重新加载模块
            import importlib
            from data.base_data import labor_cost_data
            importlib.reload(labor_cost_data)
            
            # 验证数据是否正确加载
            test_df = labor_cost_data.get_labor_cost_dataframe()
            print(f"[导入] 重新加载后验证: {len(test_df)} 条记录")
            
            print(f"✓ 已覆盖导入 {len(records)} 条计件人工标准记录")
            return True
        else:
            # 追加模式：合并数据，一次性写入文件
            existing_df = get_labor_cost_dataframe()
            existing_codes = set(existing_df['R3系统代码'].astype(str).tolist()) if existing_df is not None and not existing_df.empty else set()
            
            # 构建合并后的记录字典（以R3系统代码为键）
            merged_records = {}
            
            # 先添加现有记录
            if existing_df is not None and not existing_df.empty:
                for _, row in existing_df.iterrows():
                    code = str(row['R3系统代码'])
                    merged_records[code] = row.to_dict()
            
            # 然后添加或更新导入的记录
            for record in records:
                code = str(record['R3系统代码'])
                merged_records[code] = record
            
            # 重新生成整个文件
            new_content = '''# 计件人工标准数据 - 内置数据
# 数据来源：计件人工标准.xlsx
# 单价单位：元/台（TAI）

import pandas as pd
import os

# 完整的计件人工标准数据
LABOR_COST_DATA = [
'''
            
            # 添加所有合并后的记录
            for code, record in merged_records.items():
                category = str(record['类别']).replace('"', '\\"').replace('\\', '\\\\')
                code_str = str(record['R3系统代码']).replace('"', '\\"').replace('\\', '\\\\')
                name = str(record['系统名称']).replace('"', '\\"').replace('\\', '\\\\')
                
                record_parts = [
                    f'"类别": "{category}"',
                    f'"R3系统代码": "{code_str}"',
                    f'"系统名称": "{name}"'
                ]
                
                for price_field in price_fields:
                    price_val = float(record.get(price_field, 0.0))
                    record_parts.append(f'"{price_field}": {price_val}')
                
                record_str = '{' + ', '.join(record_parts) + '}'
                new_content += f'    {record_str},\n'
            
            new_content += ''']

def get_labor_cost_dataframe():
    """获取计件人工标准DataFrame"""
    return pd.DataFrame(LABOR_COST_DATA)

def filter_by_category(category):
    """根据类别筛选数据"""
    df = get_labor_cost_dataframe()
    if df.empty:
        return df
    return df[df['类别'] == category]

def get_all_categories():
    """获取所有类别列表"""
    df = get_labor_cost_dataframe()
    if df.empty:
        return []
    return df['类别'].unique().tolist()

def get_category_stats():
    """获取类别统计信息"""
    df = get_labor_cost_dataframe()
    if df.empty:
        return {}
    return df['类别'].value_counts().to_dict()

def get_labor_cost_by_code(code):
    """根据R3系统代码获取计件人工标准"""
    df = get_labor_cost_dataframe()
    if df.empty:
        return None
    result = df[df['R3系统代码'].astype(str) == str(code)]
    if len(result) > 0:
        return result.iloc[0].to_dict()
    return None
'''
            
            # 写入文件
            labor_cost_file_path = 'data/base_data/labor_cost_data.py'
            print(f"[导入] 开始写入文件: {labor_cost_file_path}")
            with open(labor_cost_file_path, 'w', encoding='utf-8') as f:
                f.write(new_content)
            
            # 验证文件是否写入成功
            import os
            if os.path.exists(labor_cost_file_path):
                file_size = os.path.getsize(labor_cost_file_path)
                print(f"[导入] 文件写入成功，文件大小: {file_size} 字节")
            else:
                print(f"[导入] 错误: 文件不存在: {labor_cost_file_path}")
                return False
            
            # 重新加载模块
            import importlib
            from data.base_data import labor_cost_data
            importlib.reload(labor_cost_data)
            
            # 验证数据是否正确加载
            test_df = labor_cost_data.get_labor_cost_dataframe()
            print(f"[导入] 重新加载后验证: {len(test_df)} 条记录")
            
            print(f"✓ 已追加导入 {len(records)} 条计件人工标准记录，共 {len(merged_records)} 条记录")
            return True
            
    except Exception as e:
        print(f"✗ 导入计件人工标准数据失败: {e}")
        traceback.print_exc()
        return False

@data_management_bp.route('/labor-cost', methods=['GET'])
def get_labor_cost_data():
    """获取计件人工标准数据"""
    try:
        print(f"[API] 收到获取计件人工标准数据请求: page={request.args.get('page', 1)}, per_page={request.args.get('per_page', 50)}")
        
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        search = request.args.get('search', '')
        category_filter = request.args.get('category', '')
        
        print(f"[API] 开始调用 get_labor_cost_dataframe()")
        df = get_labor_cost_dataframe()
        print(f"[API] get_labor_cost_dataframe() 返回: type={type(df)}, empty={df.empty if hasattr(df, 'empty') else 'N/A'}, len={len(df) if hasattr(df, '__len__') else 'N/A'}")
        
        # 添加调试信息
        if df is None:
            print("警告: get_labor_cost_dataframe() 返回 None")
            return jsonify({
                'success': False,
                'error': '无法获取数据，DataFrame为None'
            }), 500
        
        if df.empty:
            print(f"警告: DataFrame为空，列名: {list(df.columns) if hasattr(df, 'columns') else 'N/A'}")
            return jsonify({
                'success': True,
                'data': [],
                'total': 0,
                'current_page': page,
                'per_page': per_page,
                'pages': 0,
                'categories': ['全部']
            })
        
        print(f"成功获取数据: {len(df)} 条记录，列名: {list(df.columns)}")
        
        # 获取所有类别用于筛选（在过滤前获取）
        if '类别' in df.columns:
            all_categories = ['全部'] + sorted(df['类别'].unique().tolist())
        else:
            all_categories = ['全部']
        
        # 搜索过滤
        if search:
            mask = (
                df['R3系统代码'].astype(str).str.contains(search, case=False, na=False) |
                df['系统名称'].astype(str).str.contains(search, case=False, na=False)
            )
            df = df[mask]
        
        # 类别过滤
        if category_filter and category_filter != '全部':
            df = df[df['类别'] == category_filter]
        
        # 分页
        total = len(df)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        df_page = df.iloc[start_idx:end_idx] if total > 0 else pd.DataFrame()
        
        # 使用之前获取的所有类别
        categories = all_categories
        
        # 转换数据
        data_list = safe_json_convert(df_page)
        print(f"[API] 返回数据: total={total}, page={page}, per_page={per_page}, data_count={len(data_list) if isinstance(data_list, list) else 'N/A'}")
        if isinstance(data_list, list) and len(data_list) > 0:
            print(f"[API] 第一条数据示例: {data_list[0]}")
        
        return jsonify({
            'success': True,
            'data': data_list,
            'total': total,
            'current_page': page,
            'per_page': per_page,
            'pages': (total + per_page - 1) // per_page,
            'categories': categories
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/labor-cost', methods=['POST'])
def create_labor_cost_record():
    """创建新的人工提成单价记录"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '请提供数据'}), 400
        
        # 验证必需字段
        required_fields = ['category', 'code', 'name']
        for field in required_fields:
            if not data.get(field, ''):
                return jsonify({'success': False, 'error': f'字段 {field} 不能为空'}), 400
        
        category = data['category'].strip()
        code = data['code'].strip()
        name = data['name'].strip()
        
        # 获取所有单价字段
        price_fields = [
            "生产计件单价", "品管提成单价", "物流主管提成单价", "物流卸货提成单价",
            "班组长提成单价", "生产主管提成单价", "维修班长提成单价", "维修员提成单价",
            "冰箱维修主管提成单价", "叉车司磅库管等提成单价"
        ]
        
        # 验证并获取所有单价字段的值
        new_record = {
            '类别': category,
            'R3系统代码': code,
            '系统名称': name
        }
        
        for price_field in price_fields:
            field_key = price_field.lower().replace('/', '_').replace('等', '')
            try:
                price_val = float(data.get(field_key, 0.0))
                if price_val < 0:
                    return jsonify({'success': False, 'error': f'{price_field} 不能为负数'}), 400
                new_record[price_field] = price_val
            except (ValueError, TypeError):
                return jsonify({'success': False, 'error': f'{price_field} 必须是有效数字'}), 400
        
        # 检查代码是否已存在
        df = get_labor_cost_dataframe()
        if df is not None and not df.empty and code in df['R3系统代码'].astype(str).values:
            return jsonify({'success': False, 'error': f'代码 {code} 已存在'}), 400
        
        # 备份文件
        backup_labor_cost_file()
        
        if add_labor_cost_record_to_file(new_record):
            return jsonify({'success': True, 'message': '记录创建成功'})
        else:
            return jsonify({'success': False, 'error': '创建记录失败'}), 500
        
    except Exception as e:
        print(f"创建人工提成单价记录错误: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/labor-cost/<record_code>', methods=['PUT'])
def update_labor_cost_record(record_code):
    """更新计件人工标准记录"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '请提供数据'}), 400
        
        # 验证必需字段
        required_fields = ['category', 'code', 'name']
        for field in required_fields:
            if not data.get(field, ''):
                return jsonify({'success': False, 'error': f'字段 {field} 不能为空'}), 400
        
        category = data['category'].strip()
        new_code = data['code'].strip()
        name = data['name'].strip()
        
        # 获取所有单价字段
        price_fields = [
            "生产计件单价", "品管提成单价", "物流主管提成单价", "物流卸货提成单价",
            "班组长提成单价", "生产主管提成单价", "维修班长提成单价", "维修员提成单价",
            "冰箱维修主管提成单价", "叉车司磅库管等提成单价"
        ]
        
        # 验证单价字段
        for price_field in price_fields:
            field_key = price_field.lower().replace('/', '_').replace('等', '')
            try:
                price_val = float(data.get(field_key, 0.0))
            except (ValueError, TypeError):
                return jsonify({'success': False, 'error': f'{price_field}必须是有效数字'}), 400
        
        # 检查记录是否存在
        df = get_labor_cost_dataframe()
        if df is None or df.empty:
            return jsonify({'success': False, 'error': '无法获取计件人工标准数据'}), 500
            
        old_record_mask = df['R3系统代码'].astype(str) == record_code
        if not old_record_mask.any():
            return jsonify({'success': False, 'error': '记录不存在'}), 404
        
        # 如果代码改变了，检查新代码是否已存在
        if new_code != record_code and new_code in df['R3系统代码'].astype(str).values:
            return jsonify({'success': False, 'error': f'代码 {new_code} 已存在'}), 400
        
        # 备份文件
        backup_labor_cost_file()
        
        # 更新记录
        updated_record = {
            '类别': category,
            'R3系统代码': new_code,
            '系统名称': name
        }
        
        # 添加所有单价字段
        for price_field in price_fields:
            field_key = price_field.lower().replace('/', '_').replace('等', '')
            updated_record[price_field] = float(data.get(field_key, 0.0))
        
        if update_labor_cost_record_in_file(record_code, updated_record):
            return jsonify({'success': True, 'message': '记录更新成功'})
        else:
            return jsonify({'success': False, 'error': '更新记录失败'}), 500
        
    except Exception as e:
        print(f"更新人工提成单价记录错误: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/labor-cost/<record_code>', methods=['DELETE'])
def delete_labor_cost_record(record_code):
    """删除人工提成单价记录"""
    try:
        # 检查记录是否存在
        df = get_labor_cost_dataframe()
        if df is None or df.empty:
            return jsonify({'success': False, 'error': '无法获取人工提成单价数据'}), 500
            
        record_mask = df['R3系统代码'].astype(str) == record_code
        if not record_mask.any():
            return jsonify({'success': False, 'error': '记录不存在'}), 404
        
        # 备份文件
        backup_labor_cost_file()
        
        # 删除记录
        if delete_labor_cost_record_from_file(record_code):
            return jsonify({'success': True, 'message': '记录删除成功'})
        else:
            return jsonify({'success': False, 'error': '删除记录失败'}), 500
        
    except Exception as e:
        print(f"删除人工提成单价记录错误: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/labor-cost/import', methods=['POST'])
def import_labor_cost_data():
    """导入人工提成单价数据"""
    try:
        # 检查文件
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '请选择要导入的Excel文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '请选择要导入的Excel文件'}), 400
        
        # 检查文件格式
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': '只支持Excel文件格式(.xlsx, .xls)'}), 400
        
        # 获取导入模式
        import_mode = request.form.get('import_mode', 'append')
        
        # 获取原有数据统计
        original_df = get_labor_cost_dataframe()
        original_count = len(original_df) if original_df is not None and not original_df.empty else 0
        
        # 读取Excel文件
        try:
            df = pd.read_excel(file, engine='openpyxl' if file.filename.endswith('.xlsx') else 'xlrd')
        except Exception as e:
            return jsonify({'success': False, 'error': f'文件读取失败: {str(e)}'}), 400
        
        # 验证数据
        required_columns = ['类别', 'R3系统代码', '系统名称']
        price_columns = [
            '生产计件单价', '品管提成单价', '物流主管提成单价', '物流卸货提成单价',
            '班组长提成单价', '生产主管提成单价', '维修班长提成单价', '维修员提成单价',
            '冰箱维修主管提成单价', '叉车司磅库管等提成单价'
        ]
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return jsonify({
                'success': False, 
                'error': f'Excel文件缺少必要的列: {", ".join(missing_columns)}'
            }), 400
        
        # 清理数据
        df = df.dropna(subset=['R3系统代码', '系统名称'])
        
        # 处理单价列：转换为数值类型，缺失值填充为0
        for col in price_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            else:
                df[col] = 0
        
        # 填充类别空值
        df['类别'] = df['类别'].fillna('')
        
        # 数据验证
        validation_errors = []
        for index, row in df.iterrows():
            row_num = index + 2
            if not str(row['R3系统代码']).strip():
                validation_errors.append(f'第{row_num}行：R3系统代码不能为空')
            if not str(row['系统名称']).strip():
                validation_errors.append(f'第{row_num}行：系统名称不能为空')
        
        if validation_errors:
            return jsonify({
                'success': False,
                'error': '数据验证失败',
                'details': validation_errors[:10]
            }), 400
        
        # 去重处理（基于R3系统代码）
        df = df.drop_duplicates(subset=['R3系统代码'], keep='first')
        
        if len(df) == 0:
            return jsonify({'success': False, 'error': '没有有效的数据可以导入'}), 400
        
        # 备份现有数据
        backup_labor_cost_file()
        
        # 执行导入
        try:
            records = df.to_dict('records')
            if import_labor_cost_data_to_file(records, mode=import_mode):
                # 重新加载模块
                import importlib
                from data.base_data import labor_cost_data
                importlib.reload(labor_cost_data)
                
                # 获取最终统计
                final_df = get_labor_cost_dataframe()
                final_count = len(final_df) if final_df is not None and not final_df.empty else 0
                
                if import_mode == 'replace':
                    message = f'数据覆盖导入成功！共导入 {len(df)} 条记录'
                else:
                    # 计算新增和更新的记录数
                    existing_codes = set(original_df['R3系统代码'].astype(str).tolist()) if original_df is not None and not original_df.empty else set()
                    imported_codes = set(str(r['R3系统代码']) for r in records)
                    new_count = len(imported_codes - existing_codes)
                    updated_count = len(imported_codes & existing_codes)
                    message = f'数据追加导入成功！新增 {new_count} 条记录，更新 {updated_count} 条记录'
                
                return jsonify({
                    'success': True,
                    'message': message,
                    'statistics': {
                        'original_count': original_count,
                        'imported_count': len(df),
                        'final_count': final_count,
                        'import_mode': '覆盖' if import_mode == 'replace' else '追加'
                    }
                })
            else:
                return jsonify({'success': False, 'error': '数据导入失败'}), 500
                
        except Exception as e:
            return jsonify({'success': False, 'error': f'数据导入失败: {str(e)}'}), 500
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'导入失败: {str(e)}'}), 500

@data_management_bp.route('/labor-cost/export', methods=['GET'])
def export_labor_cost_data():
    """导出人工提成单价数据"""
    try:
        df = get_labor_cost_dataframe()
        
        if df is None or df.empty:
            return jsonify({
                'success': False,
                'error': '没有可导出的人工提成单价数据'
            }), 400
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='旧机拆解人工提成单价', index=False)
            
            # 设置列宽
            worksheet = writer.sheets['旧机拆解人工提成单价']
            worksheet.column_dimensions['A'].width = 15  # 类别
            worksheet.column_dimensions['B'].width = 15  # R3系统代码
            worksheet.column_dimensions['C'].width = 40  # 系统名称
            worksheet.column_dimensions['D'].width = 20  # 计件型号
            worksheet.column_dimensions['E'].width = 15  # 计件单价
        
        output.seek(0)
        filename = f'旧机拆解人工提成单价_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"导出人工提成单价数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 薪酬核算基础数据 API ====================

def get_salary_accounting_dataframe():
    """获取薪酬核算基础数据DataFrame"""
    try:
        # 导入模块并重新加载，确保获取最新数据
        import importlib
        from data.base_data import salary_accounting_data
        
        # 强制重新加载模块，清除Python的模块缓存
        importlib.reload(salary_accounting_data)
        
        # 调用函数获取数据
        df = salary_accounting_data.get_salary_accounting_dataframe()
        
        # 验证DataFrame
        if df is None:
            print("错误: get_salary_accounting_dataframe() 返回 None")
            return pd.DataFrame()
        
        if not isinstance(df, pd.DataFrame):
            print(f"错误: get_salary_accounting_dataframe() 返回了非DataFrame类型: {type(df)}")
            return pd.DataFrame()
        
        print(f"成功加载薪酬核算基础数据: {len(df)} 条记录")
        return df
        
    except ImportError as e:
        print(f"导入错误: {e}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame()
    except Exception as e:
        print(f"获取薪酬核算基础数据失败: {e}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame()

def backup_salary_accounting_file():
    """备份薪酬核算基础数据文件"""
    try:
        import shutil
        from datetime import datetime
        
        # 创建备份目录
        backup_dir = 'backups'
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        # 生成备份文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"salary_accounting_data_backup_{timestamp}.py"
        backup_path = os.path.join(backup_dir, backup_filename)
        
        # 复制文件
        source_path = 'data/base_data/salary_accounting_data.py'
        if os.path.exists(source_path):
            shutil.copy2(source_path, backup_path)
            print(f"✓ 薪酬核算基础数据文件已备份到: {backup_path}")
            return True
        else:
            print(f"⚠️ 源文件不存在: {source_path}")
            return False
        
    except Exception as e:
        print(f"✗ 备份薪酬核算基础数据文件失败: {e}")
        return False

def add_salary_accounting_record_to_file(new_record):
    """将新的薪酬核算基础数据记录添加到文件中"""
    try:
        import re
        import importlib
        
        # 读取当前文件内容
        salary_accounting_file_path = 'data/base_data/salary_accounting_data.py'
        if not os.path.exists(salary_accounting_file_path):
            # 如果文件不存在，创建新文件
            with open(salary_accounting_file_path, 'w', encoding='utf-8') as f:
                f.write('''# 薪酬核算基础数据 - 内置数据
# 数据来源：薪酬核算基础数据.xlsx
# 说明：此文件由 scripts/init_salary_accounting_from_excel.py 自动生成

import pandas as pd
import os

# 完整的薪酬核算基础数据
SALARY_ACCOUNTING_DATA = [
]
''')
        
        with open(salary_accounting_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 构建记录字典（处理所有字段）
        record_parts = []
        for key, value in new_record.items():
            if pd.isna(value) or value is None:
                record_parts.append(f'"{key}": None')
            elif isinstance(value, (int, float)):
                record_parts.append(f'"{key}": {value}')
            else:
                # 字符串类型，需要转义
                str_value = str(value).replace('"', '\\"').replace('\\', '\\\\')
                record_parts.append(f'"{key}": "{str_value}"')
        
        record_str = '{' + ', '.join(record_parts) + '}'
        new_line = f'    {record_str},\n'
        
        # 找到列表的结束位置（最后一个 ] 之前）
        pattern = r'(\s*\{[^}]+\},?\n)(\s*\])'
        
        def replace_func(match):
            last_record = match.group(1)
            closing_bracket = match.group(2)
            
            # 确保最后一个记录有逗号
            if not last_record.rstrip().endswith(','):
                last_record = last_record.rstrip() + ',\n'
            
            return last_record + new_line + closing_bracket
        
        new_content = re.sub(pattern, replace_func, content, flags=re.MULTILINE | re.DOTALL)
        
        # 检查是否成功替换
        if new_content == content:
            # 如果列表为空，需要在列表开始后插入
            pattern2 = r'(SALARY_ACCOUNTING_DATA = \[)(\s*\])'
            new_content = re.sub(pattern2, r'\1\n' + new_line + r'\2', content)
            if new_content == content:
                raise ValueError("无法找到合适的位置插入新记录")
        
        # 写回文件
        with open(salary_accounting_file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        # 重新加载模块
        from data.base_data import salary_accounting_data
        importlib.reload(salary_accounting_data)
        
        print(f"✓ 已添加薪酬核算基础数据记录")
        return True
        
    except Exception as e:
        print(f"✗ 添加薪酬核算基础数据记录失败: {e}")
        traceback.print_exc()
        return False

def update_salary_accounting_record_in_file(record_id, new_record):
    """更新薪酬核算基础数据文件中的记录"""
    try:
        import re
        import importlib
        
        # 读取当前文件内容
        salary_accounting_file_path = 'data/base_data/salary_accounting_data.py'
        with open(salary_accounting_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 构建记录字典（处理所有字段）
        record_parts = []
        for key, value in new_record.items():
            if pd.isna(value) or value is None:
                record_parts.append(f'"{key}": None')
            elif isinstance(value, (int, float)):
                record_parts.append(f'"{key}": {value}')
            else:
                # 字符串类型，需要转义
                str_value = str(value).replace('"', '\\"').replace('\\', '\\\\')
                record_parts.append(f'"{key}": "{str_value}"')
        
        record_str = '{' + ', '.join(record_parts) + '}'
        new_line = f'    {record_str},\n'
        
        # 获取第一列作为ID字段（用于匹配记录）
        df = get_salary_accounting_dataframe()
        if df.empty:
            raise ValueError("数据为空，无法更新")
        
        # 明确使用"岗位"作为主键字段
        if '岗位' in df.columns:
            first_col = '岗位'
            print(f"[更新记录] 使用'岗位'作为主键字段，record_id={record_id}")
        else:
            first_col = df.columns[0]
            print(f"[更新记录] 未找到'岗位'字段，使用第一列: {first_col}")
        
        # 找到并替换对应的记录（匹配第一列的值）
        old_value = str(record_id)
        # 转义特殊字符用于正则表达式
        escaped_value = re.escape(old_value)
        old_pattern = rf'    \{{[^}}]*"{re.escape(first_col)}":\s*"{escaped_value}"[^}}]*}},?\n'
        
        new_content = re.sub(old_pattern, new_line, content)
        
        # 检查是否找到并替换了记录
        if new_content == content:
            raise ValueError(f"未找到ID为 {record_id} 的记录")
        
        # 写回文件
        with open(salary_accounting_file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        # 重新加载模块
        from data.base_data import salary_accounting_data
        importlib.reload(salary_accounting_data)
        
        print(f"✓ 已更新薪酬核算基础数据记录: {record_id}")
        return True
        
    except Exception as e:
        print(f"✗ 更新薪酬核算基础数据记录失败: {e}")
        traceback.print_exc()
        return False

def delete_salary_accounting_record_from_file(record_id):
    """从薪酬核算基础数据文件中删除记录"""
    try:
        import re
        import importlib
        
        # 读取当前文件内容
        salary_accounting_file_path = 'data/base_data/salary_accounting_data.py'
        with open(salary_accounting_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 获取第一列作为ID字段（用于匹配记录）
        df = get_salary_accounting_dataframe()
        if df.empty:
            raise ValueError("数据为空，无法删除")
        
        # 明确使用"岗位"作为主键字段
        if '岗位' in df.columns:
            first_col = '岗位'
        else:
            first_col = df.columns[0]
        
        # 找到并删除对应的记录
        old_value = str(record_id)
        escaped_value = re.escape(old_value)
        pattern = rf'    \{{[^}}]*"{re.escape(first_col)}":\s*"{escaped_value}"[^}}]*}},?\n'
        new_content = re.sub(pattern, '', content)
        
        # 检查是否找到并删除了记录
        if new_content == content:
            raise ValueError(f"未找到ID为 {record_id} 的记录")
        
        # 写回文件
        with open(salary_accounting_file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        # 重新加载模块
        from data.base_data import salary_accounting_data
        importlib.reload(salary_accounting_data)
        
        print(f"✓ 已删除薪酬核算基础数据记录: {record_id}")
        return True
        
    except Exception as e:
        print(f"✗ 删除薪酬核算基础数据记录失败: {e}")
        traceback.print_exc()
        return False

@data_management_bp.route('/salary-accounting', methods=['GET'])
def get_salary_accounting_data():
    """获取薪酬核算基础数据"""
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        search = request.args.get('search', '')
        expense_type_filter = request.args.get('expense_type', '')
        
        df = get_salary_accounting_dataframe()
        
        if df is None:
            return jsonify({
                'success': False,
                'error': '无法获取数据，DataFrame为None'
            }), 500
        
        if df.empty:
            return jsonify({
                'success': True,
                'data': [],
                'total': 0,
                'current_page': page,
                'per_page': per_page,
                'pages': 0,
                'categories': ['全部']
            })
        
        # 定义列的正确顺序（按照截图要求）
        column_order = [
            '岗位',
            '人员基础配置',
            '提成工资',
            '平均工资（元/月/人）',
            '奖励/补助（元/月）',
            '餐补（元/月/人）',
            '年终奖（元/人）',
            '养老保险费（元/月/人）',
            '失业保险费（元/月/人）',
            '医疗/生育保险费（元/月/人）',
            '工伤保险费（元/月/人）',
            '住房公积金（元/月/人）',
            '费用类型'
        ]
        
        # 重新排列DataFrame的列顺序
        # 先获取存在的列，按照指定顺序排列，然后添加其他未指定的列
        existing_ordered_cols = [col for col in column_order if col in df.columns]
        other_cols = [col for col in df.columns if col not in column_order]
        df = df[existing_ordered_cols + other_cols]
        
        # 获取所有费用类型用于筛选（在过滤前获取）
        if '费用类型' in df.columns:
            all_expense_types = ['全部'] + sorted(df['费用类型'].dropna().unique().tolist())
        else:
            all_expense_types = ['全部']
        
        # 搜索过滤（在所有文本列中搜索）
        if search:
            mask = pd.Series([False] * len(df))
            for col in df.columns:
                if df[col].dtype == 'object':  # 文本列
                    mask |= df[col].astype(str).str.contains(search, case=False, na=False)
            df = df[mask]
        
        # 费用类型过滤
        if expense_type_filter and expense_type_filter != '全部' and '费用类型' in df.columns:
            df = df[df['费用类型'] == expense_type_filter]
        
        # 分页
        total = len(df)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        df_page = df.iloc[start_idx:end_idx] if total > 0 else pd.DataFrame()
        
        # 转换数据
        data_list = safe_json_convert(df_page)
        
        return jsonify({
            'success': True,
            'data': data_list,
            'total': total,
            'current_page': page,
            'per_page': per_page,
            'pages': (total + per_page - 1) // per_page,
            'expense_types': all_expense_types
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/salary-accounting', methods=['POST'])
def create_salary_accounting_record():
    """创建新的薪酬核算基础数据记录"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '请提供数据'}), 400
        
        # 获取现有数据以确定必需字段
        df = get_salary_accounting_dataframe()
        if not df.empty:
            # 使用第一列作为必需字段
            first_col = df.columns[0]
            if first_col not in data or not data.get(first_col, ''):
                return jsonify({'success': False, 'error': f'字段 {first_col} 不能为空'}), 400
        
        # 构建新记录（包含所有提供的字段）
        new_record = {}
        for key, value in data.items():
            new_record[key] = value
        
        # 检查记录是否已存在（基于第一列）
        if not df.empty:
            first_col = df.columns[0]
            if first_col in new_record:
                existing_value = str(new_record[first_col])
                if existing_value in df[first_col].astype(str).values:
                    return jsonify({'success': False, 'error': f'{first_col} {existing_value} 已存在'}), 400
        
        # 备份文件
        backup_salary_accounting_file()
        
        if add_salary_accounting_record_to_file(new_record):
            return jsonify({'success': True, 'message': '记录创建成功'})
        else:
            return jsonify({'success': False, 'error': '创建记录失败'}), 500
        
    except Exception as e:
        print(f"创建薪酬核算基础数据记录错误: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/salary-accounting/<record_id>', methods=['PUT'])
def update_salary_accounting_record(record_id):
    """更新薪酬核算基础数据记录"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '请提供数据'}), 400
        
        # 检查记录是否存在
        df = get_salary_accounting_dataframe()
        if df is None or df.empty:
            return jsonify({'success': False, 'error': '无法获取薪酬核算基础数据'}), 500
        
        # 明确使用"岗位"作为主键字段
        if '岗位' in df.columns:
            first_col = '岗位'
        else:
            first_col = df.columns[0]
        
        print(f"[更新API] 主键字段: {first_col}, record_id: {record_id}")
        print(f"[更新API] 数据列: {df.columns.tolist()}")
        print(f"[更新API] {first_col}列的值: {df[first_col].tolist()[:5]}")
        
        record_mask = df[first_col].astype(str) == str(record_id)
        if not record_mask.any():
            print(f"[更新API] 未找到记录，搜索的值: {record_id}")
            return jsonify({'success': False, 'error': '记录不存在'}), 404
        
        # 构建更新记录
        updated_record = {}
        for key, value in data.items():
            updated_record[key] = value
        
        # 如果第一列的值改变了，检查新值是否已存在
        if first_col in updated_record:
            new_value = str(updated_record[first_col])
            if new_value != str(record_id) and new_value in df[first_col].astype(str).values:
                return jsonify({'success': False, 'error': f'{first_col} {new_value} 已存在'}), 400
        
        # 备份文件
        backup_salary_accounting_file()
        
        if update_salary_accounting_record_in_file(record_id, updated_record):
            return jsonify({'success': True, 'message': '记录更新成功'})
        else:
            return jsonify({'success': False, 'error': '更新记录失败'}), 500
        
    except Exception as e:
        print(f"更新薪酬核算基础数据记录错误: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/salary-accounting/<record_id>', methods=['DELETE'])
def delete_salary_accounting_record(record_id):
    """删除薪酬核算基础数据记录"""
    try:
        # 检查记录是否存在
        df = get_salary_accounting_dataframe()
        if df is None or df.empty:
            return jsonify({'success': False, 'error': '无法获取薪酬核算基础数据'}), 500
        
        # 明确使用"岗位"作为主键字段
        if '岗位' in df.columns:
            first_col = '岗位'
        else:
            first_col = df.columns[0]
        
        record_mask = df[first_col].astype(str) == str(record_id)
        if not record_mask.any():
            return jsonify({'success': False, 'error': '记录不存在'}), 404
        
        # 备份文件
        backup_salary_accounting_file()
        
        # 删除记录
        if delete_salary_accounting_record_from_file(record_id):
            return jsonify({'success': True, 'message': '记录删除成功'})
        else:
            return jsonify({'success': False, 'error': '删除记录失败'}), 500
        
    except Exception as e:
        print(f"删除薪酬核算基础数据记录错误: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/salary-accounting/import', methods=['POST'])
def import_salary_accounting_data():
    """导入薪酬核算基础数据"""
    try:
        # 检查文件
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '请选择要导入的Excel文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '请选择要导入的Excel文件'}), 400
        
        # 检查文件格式
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': '只支持Excel文件格式(.xlsx, .xls)'}), 400
        
        # 获取导入模式
        import_mode = request.form.get('import_mode', 'append')
        
        # 获取原有数据统计
        original_df = get_salary_accounting_dataframe()
        original_count = len(original_df) if original_df is not None and not original_df.empty else 0
        
        # 读取Excel文件
        try:
            df = pd.read_excel(file, engine='openpyxl' if file.filename.endswith('.xlsx') else 'xlrd')
        except Exception as e:
            return jsonify({'success': False, 'error': f'文件读取失败: {str(e)}'}), 400
        
        # 删除完全为空的行
        df = df.dropna(how='all')
        
        if df.empty:
            return jsonify({'success': False, 'error': 'Excel文件中没有有效数据'}), 400
        
        # 备份文件
        backup_salary_accounting_file()
        
        # 根据导入模式处理数据
        if import_mode == 'replace':
            # 覆盖模式：重新生成整个文件
            new_content = '''# 薪酬核算基础数据 - 内置数据
# 数据来源：薪酬核算基础数据.xlsx
# 说明：此文件由 scripts/init_salary_accounting_from_excel.py 自动生成

import pandas as pd
import os

# 完整的薪酬核算基础数据
SALARY_ACCOUNTING_DATA = [
'''
            
            # 添加所有记录
            for index, row in df.iterrows():
                record_parts = []
                for col in df.columns:
                    value = row[col]
                    if pd.isna(value):
                        record_parts.append(f'"{col}": None')
                    elif isinstance(value, (int, float)):
                        record_parts.append(f'"{col}": {value}')
                    else:
                        str_value = str(value).replace('"', '\\"').replace('\\', '\\\\')
                        record_parts.append(f'"{col}": "{str_value}"')
                
                record_str = '{' + ', '.join(record_parts) + '}'
                new_content += f'    {record_str},\n'
            
            new_content += ''']

def get_salary_accounting_dataframe():
    """获取薪酬核算基础数据DataFrame"""
    return pd.DataFrame(SALARY_ACCOUNTING_DATA)

def filter_by_category(category):
    """根据类别筛选数据（如果有类别字段）"""
    df = get_salary_accounting_dataframe()
    if df.empty:
        return df
    if '类别' in df.columns:
        return df[df['类别'] == category]
    return df

def get_all_categories():
    """获取所有类别列表（如果有类别字段）"""
    df = get_salary_accounting_dataframe()
    if df.empty:
        return []
    if '类别' in df.columns:
        return df['类别'].unique().tolist()
    return []

def get_category_stats():
    """获取类别统计信息（如果有类别字段）"""
    df = get_salary_accounting_dataframe()
    if df.empty:
        return {}
    if '类别' in df.columns:
        return df['类别'].value_counts().to_dict()
    return {}

def get_salary_accounting_by_id(record_id):
    """根据记录ID获取薪酬核算基础数据"""
    df = get_salary_accounting_dataframe()
    if df.empty:
        return None
    if len(df) > 0:
        first_col = df.columns[0]
        if 'ID' in first_col.upper() or '代码' in first_col or '编号' in first_col:
            result = df[df[first_col].astype(str) == str(record_id)]
        else:
            try:
                idx = int(record_id)
                if 0 <= idx < len(df):
                    result = df.iloc[[idx]]
                else:
                    return None
            except (ValueError, TypeError):
                return None
        
        if len(result) > 0:
            return result.iloc[0].to_dict()
    return None
'''
            
            # 写入文件
            with open('data/base_data/salary_accounting_data.py', 'w', encoding='utf-8') as f:
                f.write(new_content)
            
            # 重新加载模块
            import importlib
            from data.base_data import salary_accounting_data
            importlib.reload(salary_accounting_data)
            
            new_count = len(df)
            return jsonify({
                'success': True,
                'message': f'成功导入 {new_count} 条记录（覆盖模式）',
                'original_count': original_count,
                'new_count': new_count,
                'imported_count': new_count
            })
        else:
            # 追加模式：添加新记录
            records = df.to_dict('records')
            added_count = 0
            skipped_count = 0
            
            for record in records:
                # 检查是否已存在（基于第一列）
                if not original_df.empty:
                    first_col = original_df.columns[0]
                    if first_col in record:
                        existing_value = str(record[first_col])
                        if existing_value in original_df[first_col].astype(str).values:
                            skipped_count += 1
                            continue
                
                if add_salary_accounting_record_to_file(record):
                    added_count += 1
            
            return jsonify({
                'success': True,
                'message': f'成功导入 {added_count} 条记录，跳过 {skipped_count} 条重复记录',
                'original_count': original_count,
                'new_count': len(get_salary_accounting_dataframe()),
                'imported_count': added_count,
                'skipped_count': skipped_count
            })
        
    except Exception as e:
        print(f"导入薪酬核算基础数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'数据导入失败: {str(e)}'}), 500

@data_management_bp.route('/salary-accounting/export', methods=['GET'])
def export_salary_accounting_data():
    """导出薪酬核算基础数据"""
    try:
        df = get_salary_accounting_dataframe()
        
        if df is None or df.empty:
            return jsonify({
                'success': False,
                'error': '没有可导出的薪酬核算基础数据'
            }), 400
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='薪酬核算基础数据', index=False)
            
            # 设置列宽
            worksheet = writer.sheets['薪酬核算基础数据']
            for idx, col in enumerate(df.columns, start=1):
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(idx)
                # 根据列名长度设置列宽
                col_width = max(len(str(col)), 15)
                worksheet.column_dimensions[col_letter].width = min(col_width, 50)
        
        output.seek(0)
        filename = f'薪酬核算基础数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"导出薪酬核算基础数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 制造费用基础数据 API ====================

def get_manufacturing_cost_dataframe():
    """获取制造费用基础数据DataFrame"""
    try:
        # 导入模块并重新加载，确保获取最新数据
        import importlib
        from data.base_data import manufacturing_cost_data
        
        # 强制重新加载模块，清除Python的模块缓存
        importlib.reload(manufacturing_cost_data)
        
        # 调用函数获取数据
        df = manufacturing_cost_data.get_manufacturing_cost_dataframe()
        
        # 验证DataFrame
        if df is None:
            print("错误: get_manufacturing_cost_dataframe() 返回 None")
            return pd.DataFrame()
        
        if not isinstance(df, pd.DataFrame):
            print(f"错误: get_manufacturing_cost_dataframe() 返回了非DataFrame类型: {type(df)}")
            return pd.DataFrame()
        
        print(f"成功加载制造费用基础数据: {len(df)} 条记录")
        return df
        
    except ImportError as e:
        print(f"导入错误: {e}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame()
    except Exception as e:
        print(f"获取制造费用基础数据失败: {e}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame()

def backup_manufacturing_cost_file():
    """备份制造费用基础数据文件"""
    try:
        import shutil
        from datetime import datetime
        
        # 创建备份目录
        backup_dir = 'backups'
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        # 生成备份文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"manufacturing_cost_data_backup_{timestamp}.py"
        backup_path = os.path.join(backup_dir, backup_filename)
        
        # 复制文件
        source_path = 'data/base_data/manufacturing_cost_data.py'
        if os.path.exists(source_path):
            shutil.copy2(source_path, backup_path)
            print(f"✓ 制造费用基础数据文件已备份到: {backup_path}")
            return True
        else:
            print(f"⚠️ 源文件不存在: {source_path}")
            return False
        
    except Exception as e:
        print(f"✗ 备份制造费用基础数据文件失败: {e}")
        return False

def add_manufacturing_cost_record_to_file(new_record):
    """将新的制造费用基础数据记录添加到文件中"""
    try:
        import re
        import importlib
        
        # 读取当前文件内容
        manufacturing_cost_file_path = 'data/base_data/manufacturing_cost_data.py'
        if not os.path.exists(manufacturing_cost_file_path):
            # 如果文件不存在，创建新文件
            with open(manufacturing_cost_file_path, 'w', encoding='utf-8') as f:
                f.write('''# 制造费用基础数据 - 内置数据
# 数据来源：制造费用基础数据.xlsx
# 说明：此文件由 scripts/init_manufacturing_cost_from_excel.py 自动生成

import pandas as pd
import os

# 完整的制造费用基础数据
MANUFACTURING_COST_DATA = [
]
''')
        
        with open(manufacturing_cost_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 构建记录字典（处理所有字段）
        record_parts = []
        for key, value in new_record.items():
            if pd.isna(value) or value is None:
                record_parts.append(f'"{key}": None')
            elif isinstance(value, (int, float)):
                record_parts.append(f'"{key}": {value}')
            else:
                # 字符串类型，需要转义
                str_value = str(value).replace('"', '\\"').replace('\\', '\\\\')
                record_parts.append(f'"{key}": "{str_value}"')
        
        record_str = '{' + ', '.join(record_parts) + '}'
        new_line = f'    {record_str},\n'
        
        # 找到列表的结束位置（最后一个 ] 之前）
        pattern = r'(\s*\{[^}]+\},?\n)(\s*\])'
        
        def replace_func(match):
            last_record = match.group(1)
            closing_bracket = match.group(2)
            
            # 确保最后一个记录有逗号
            if not last_record.rstrip().endswith(','):
                last_record = last_record.rstrip() + ',\n'
            
            return last_record + new_line + closing_bracket
        
        new_content = re.sub(pattern, replace_func, content, flags=re.MULTILINE | re.DOTALL)
        
        # 检查是否成功替换
        if new_content == content:
            # 如果列表为空，需要在列表开始后插入
            pattern2 = r'(MANUFACTURING_COST_DATA = \[)(\s*\])'
            new_content = re.sub(pattern2, r'\1\n' + new_line + r'\2', content)
            if new_content == content:
                raise ValueError("无法找到合适的位置插入新记录")
        
        # 写回文件
        with open(manufacturing_cost_file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        # 重新加载模块
        from data.base_data import manufacturing_cost_data
        importlib.reload(manufacturing_cost_data)
        
        print(f"✓ 已添加制造费用基础数据记录")
        return True
        
    except Exception as e:
        print(f"✗ 添加制造费用基础数据记录失败: {e}")
        traceback.print_exc()
        return False

def update_manufacturing_cost_record_in_file(record_id, new_record):
    """更新制造费用基础数据文件中的记录"""
    try:
        import re
        import importlib
        
        # 读取当前文件内容
        manufacturing_cost_file_path = 'data/base_data/manufacturing_cost_data.py'
        with open(manufacturing_cost_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 构建记录字典（处理所有字段）
        record_parts = []
        for key, value in new_record.items():
            if pd.isna(value) or value is None:
                record_parts.append(f'"{key}": None')
            elif isinstance(value, (int, float)):
                record_parts.append(f'"{key}": {value}')
            else:
                # 字符串类型，需要转义
                str_value = str(value).replace('"', '\\"').replace('\\', '\\\\')
                record_parts.append(f'"{key}": "{str_value}"')
        
        record_str = '{' + ', '.join(record_parts) + '}'
        new_line = f'    {record_str},\n'
        
        # 获取"费用名称"作为主键字段（用于匹配记录）
        df = get_manufacturing_cost_dataframe()
        if df.empty:
            raise ValueError("数据为空，无法更新")
        
        # 使用"费用名称"作为主键字段
        if '费用名称' in df.columns:
            first_col = '费用名称'
            print(f"[更新记录] 使用'费用名称'作为主键字段，record_id={record_id}")
        else:
            first_col = df.columns[0]
            print(f"[更新记录] 未找到'费用名称'字段，使用第一列: {first_col}")
        
        # 找到并替换对应的记录（匹配费用名称的值）
        old_value = str(record_id)
        # 转义特殊字符用于正则表达式
        escaped_value = re.escape(old_value)
        old_pattern = rf'    \{{[^}}]*"{re.escape(first_col)}":\s*"{escaped_value}"[^}}]*}},?\n'
        
        new_content = re.sub(old_pattern, new_line, content)
        
        # 检查是否找到并替换了记录
        if new_content == content:
            raise ValueError(f"未找到ID为 {record_id} 的记录")
        
        # 写回文件
        with open(manufacturing_cost_file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        # 重新加载模块
        from data.base_data import manufacturing_cost_data
        importlib.reload(manufacturing_cost_data)
        
        print(f"✓ 已更新制造费用基础数据记录: {record_id}")
        return True
        
    except Exception as e:
        print(f"✗ 更新制造费用基础数据记录失败: {e}")
        traceback.print_exc()
        return False

def delete_manufacturing_cost_record_from_file(record_id):
    """从制造费用基础数据文件中删除记录"""
    try:
        import re
        import importlib
        
        # 读取当前文件内容
        manufacturing_cost_file_path = 'data/base_data/manufacturing_cost_data.py'
        with open(manufacturing_cost_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 获取"费用名称"作为主键字段（用于匹配记录）
        df = get_manufacturing_cost_dataframe()
        if df.empty:
            raise ValueError("数据为空，无法删除")
        
        # 使用"费用名称"作为主键字段
        if '费用名称' in df.columns:
            first_col = '费用名称'
        else:
            first_col = df.columns[0]
        
        # 找到并删除对应的记录
        old_value = str(record_id)
        escaped_value = re.escape(old_value)
        pattern = rf'    \{{[^}}]*"{re.escape(first_col)}":\s*"{escaped_value}"[^}}]*}},?\n'
        new_content = re.sub(pattern, '', content)
        
        # 检查是否找到并删除了记录
        if new_content == content:
            raise ValueError(f"未找到ID为 {record_id} 的记录")
        
        # 写回文件
        with open(manufacturing_cost_file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        # 重新加载模块
        from data.base_data import manufacturing_cost_data
        importlib.reload(manufacturing_cost_data)
        
        print(f"✓ 已删除制造费用基础数据记录: {record_id}")
        return True
        
    except Exception as e:
        print(f"✗ 删除制造费用基础数据记录失败: {e}")
        traceback.print_exc()
        return False

@data_management_bp.route('/manufacturing-cost', methods=['GET'])
def get_manufacturing_cost_data():
    """获取制造费用基础数据"""
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        search = request.args.get('search', '')
        expense_type_filter = request.args.get('expense_type', '')
        
        df = get_manufacturing_cost_dataframe()
        
        if df is None:
            return jsonify({
                'success': False,
                'error': '无法获取数据，DataFrame为None'
            }), 500
        
        if df.empty:
            return jsonify({
                'success': True,
                'data': [],
                'total': 0,
                'current_page': page,
                'per_page': per_page,
                'pages': 0,
                'expense_types': ['全部']
            })
        
        # 定义正确的列顺序（按照Excel文件顺序）
        column_order = [
            '费用类型',
            '费用种类',
            '费用名称',
            '公共',
            '冰箱',
            '电脑',
            '电视',
            '空调',
            '洗衣机',
            '屏',
            '备注'
        ]
        
        # 重新排列DataFrame的列顺序
        # 先获取存在的列，按照指定顺序排列，然后添加其他未指定的列
        existing_ordered_cols = [col for col in column_order if col in df.columns]
        other_cols = [col for col in df.columns if col not in column_order]
        df = df[existing_ordered_cols + other_cols]
        
        # 获取所有费用类型用于筛选（在过滤前获取）
        if '费用类型' in df.columns:
            all_expense_types = ['全部'] + sorted(df['费用类型'].dropna().unique().tolist())
        else:
            all_expense_types = ['全部']
        
        # 搜索过滤（在所有文本列中搜索）
        if search:
            mask = pd.Series([False] * len(df))
            for col in df.columns:
                if df[col].dtype == 'object':  # 文本列
                    mask |= df[col].astype(str).str.contains(search, case=False, na=False)
            df = df[mask]
        
        # 费用类型过滤
        if expense_type_filter and expense_type_filter != '全部' and '费用类型' in df.columns:
            df = df[df['费用类型'] == expense_type_filter]
        
        # 将备注列为空（None或NaN）的记录填充为"预计月均费用"
        if not df.empty and '备注' in df.columns:
            df['备注'] = df['备注'].fillna('预计月均费用')
            # 处理字符串类型的空值
            df.loc[df['备注'].astype(str).str.strip() == '', '备注'] = '预计月均费用'
            df.loc[df['备注'].astype(str).str.strip() == 'nan', '备注'] = '预计月均费用'
        
        # 分页
        total = len(df)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        df_page = df.iloc[start_idx:end_idx] if total > 0 else pd.DataFrame()
        
        # 转换数据
        data_list = safe_json_convert(df_page)
        
        return jsonify({
            'success': True,
            'data': data_list,
            'total': total,
            'current_page': page,
            'per_page': per_page,
            'pages': (total + per_page - 1) // per_page,
            'expense_types': all_expense_types
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/manufacturing-cost', methods=['POST'])
def create_manufacturing_cost_record():
    """创建新的制造费用基础数据记录"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '请提供数据'}), 400
        
        # 获取现有数据以确定必需字段
        df = get_manufacturing_cost_dataframe()
        if not df.empty:
            # 使用"费用名称"作为必需字段
            if '费用名称' in df.columns:
                required_col = '费用名称'
            else:
                required_col = df.columns[0]
            if required_col not in data or not data.get(required_col, ''):
                return jsonify({'success': False, 'error': f'字段 {required_col} 不能为空'}), 400
        
        # 构建新记录（包含所有提供的字段）
        new_record = {}
        for key, value in data.items():
            new_record[key] = value
        
        # 检查记录是否已存在（基于费用名称）
        if not df.empty:
            if '费用名称' in df.columns:
                key_col = '费用名称'
            else:
                key_col = df.columns[0]
            if key_col in new_record:
                existing_value = str(new_record[key_col])
                if existing_value in df[key_col].astype(str).values:
                    return jsonify({'success': False, 'error': f'{key_col} {existing_value} 已存在'}), 400
        
        # 备份文件
        backup_manufacturing_cost_file()
        
        if add_manufacturing_cost_record_to_file(new_record):
            return jsonify({'success': True, 'message': '记录创建成功'})
        else:
            return jsonify({'success': False, 'error': '创建记录失败'}), 500
        
    except Exception as e:
        print(f"创建制造费用基础数据记录错误: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/manufacturing-cost/<record_id>', methods=['PUT'])
def update_manufacturing_cost_record(record_id):
    """更新制造费用基础数据记录"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '请提供数据'}), 400
        
        # 检查记录是否存在
        df = get_manufacturing_cost_dataframe()
        if df is None or df.empty:
            return jsonify({'success': False, 'error': '无法获取制造费用基础数据'}), 500
        
        # 使用"费用名称"作为主键字段
        if '费用名称' in df.columns:
            first_col = '费用名称'
        else:
            first_col = df.columns[0]
        
        print(f"[更新API] 主键字段: {first_col}, record_id: {record_id}")
        
        record_mask = df[first_col].astype(str) == str(record_id)
        if not record_mask.any():
            print(f"[更新API] 未找到记录，搜索的值: {record_id}")
            return jsonify({'success': False, 'error': '记录不存在'}), 404
        
        # 构建更新记录
        updated_record = {}
        for key, value in data.items():
            updated_record[key] = value
        
        # 如果费用名称的值改变了，检查新值是否已存在
        if first_col in updated_record:
            new_value = str(updated_record[first_col])
            if new_value != str(record_id) and new_value in df[first_col].astype(str).values:
                return jsonify({'success': False, 'error': f'{first_col} {new_value} 已存在'}), 400
        
        # 备份文件
        backup_manufacturing_cost_file()
        
        if update_manufacturing_cost_record_in_file(record_id, updated_record):
            return jsonify({'success': True, 'message': '记录更新成功'})
        else:
            return jsonify({'success': False, 'error': '更新记录失败'}), 500
        
    except Exception as e:
        print(f"更新制造费用基础数据记录错误: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/manufacturing-cost/<record_id>', methods=['DELETE'])
def delete_manufacturing_cost_record(record_id):
    """删除制造费用基础数据记录"""
    try:
        # 检查记录是否存在
        df = get_manufacturing_cost_dataframe()
        if df is None or df.empty:
            return jsonify({'success': False, 'error': '无法获取制造费用基础数据'}), 500
        
        # 使用"费用名称"作为主键字段
        if '费用名称' in df.columns:
            first_col = '费用名称'
        else:
            first_col = df.columns[0]
        
        record_mask = df[first_col].astype(str) == str(record_id)
        if not record_mask.any():
            return jsonify({'success': False, 'error': '记录不存在'}), 404
        
        # 备份文件
        backup_manufacturing_cost_file()
        
        # 删除记录
        if delete_manufacturing_cost_record_from_file(record_id):
            return jsonify({'success': True, 'message': '记录删除成功'})
        else:
            return jsonify({'success': False, 'error': '删除记录失败'}), 500
        
    except Exception as e:
        print(f"删除制造费用基础数据记录错误: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/manufacturing-cost/import', methods=['POST'])
def import_manufacturing_cost_data():
    """导入制造费用基础数据"""
    try:
        # 检查文件
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '请选择要导入的Excel文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '请选择要导入的Excel文件'}), 400
        
        # 检查文件格式
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': '只支持Excel文件格式(.xlsx, .xls)'}), 400
        
        # 获取导入模式
        import_mode = request.form.get('import_mode', 'append')
        
        # 获取原有数据统计
        original_df = get_manufacturing_cost_dataframe()
        original_count = len(original_df) if original_df is not None and not original_df.empty else 0
        
        # 读取Excel文件
        try:
            df = pd.read_excel(file, engine='openpyxl' if file.filename.endswith('.xlsx') else 'xlrd')
        except Exception as e:
            return jsonify({'success': False, 'error': f'文件读取失败: {str(e)}'}), 400
        
        # 删除完全为空的行
        df = df.dropna(how='all')
        
        if df.empty:
            return jsonify({'success': False, 'error': 'Excel文件中没有有效数据'}), 400
        
        # 备份文件
        backup_manufacturing_cost_file()
        
        # 根据导入模式处理数据
        if import_mode == 'replace':
            # 覆盖模式：重新生成整个文件
            new_content = '''# 制造费用基础数据 - 内置数据
# 数据来源：制造费用基础数据.xlsx
# 说明：此文件由 scripts/init_manufacturing_cost_from_excel.py 自动生成

import pandas as pd
import os

# 完整的制造费用基础数据
MANUFACTURING_COST_DATA = [
'''
            
            # 添加所有记录
            for index, row in df.iterrows():
                record_parts = []
                for col in df.columns:
                    value = row[col]
                    if pd.isna(value):
                        record_parts.append(f'"{col}": None')
                    elif isinstance(value, (int, float)):
                        record_parts.append(f'"{col}": {value}')
                    else:
                        str_value = str(value).replace('"', '\\"').replace('\\', '\\\\')
                        record_parts.append(f'"{col}": "{str_value}"')
                
                record_str = '{' + ', '.join(record_parts) + '}'
                new_content += f'    {record_str},\n'
            
            new_content += ''']

def get_manufacturing_cost_dataframe():
    """获取制造费用基础数据DataFrame"""
    return pd.DataFrame(MANUFACTURING_COST_DATA)

def filter_by_category(category):
    """根据类别筛选数据（如果有类别字段）"""
    df = get_manufacturing_cost_dataframe()
    if df.empty:
        return df
    if '类别' in df.columns:
        return df[df['类别'] == category]
    return df

def get_all_categories():
    """获取所有类别列表（如果有类别字段）"""
    df = get_manufacturing_cost_dataframe()
    if df.empty:
        return []
    if '类别' in df.columns:
        return df['类别'].unique().tolist()
    return []

def get_category_stats():
    """获取类别统计信息（如果有类别字段）"""
    df = get_manufacturing_cost_dataframe()
    if df.empty:
        return {}
    if '类别' in df.columns:
        return df['类别'].value_counts().to_dict()
    return {}

def get_manufacturing_cost_by_id(record_id):
    """根据记录ID获取制造费用基础数据"""
    df = get_manufacturing_cost_dataframe()
    if df.empty:
        return None
    # 使用"费用名称"作为主键
    if len(df) > 0:
        if '费用名称' in df.columns:
            result = df[df['费用名称'].astype(str) == str(record_id)]
            if len(result) > 0:
                return result.iloc[0].to_dict()
        # 如果找不到费用名称字段，尝试使用第一列
        first_col = df.columns[0]
        if 'ID' in first_col.upper() or '代码' in first_col or '编号' in first_col:
            result = df[df[first_col].astype(str) == str(record_id)]
        else:
            try:
                idx = int(record_id)
                if 0 <= idx < len(df):
                    result = df.iloc[[idx]]
                else:
                    return None
            except (ValueError, TypeError):
                return None
        
        if len(result) > 0:
            return result.iloc[0].to_dict()
    return None
'''
            
            # 写入文件
            with open('data/base_data/manufacturing_cost_data.py', 'w', encoding='utf-8') as f:
                f.write(new_content)
            
            # 重新加载模块
            import importlib
            from data.base_data import manufacturing_cost_data
            importlib.reload(manufacturing_cost_data)
            
            new_count = len(df)
            return jsonify({
                'success': True,
                'message': f'成功导入 {new_count} 条记录（覆盖模式）',
                'original_count': original_count,
                'new_count': new_count,
                'imported_count': new_count
            })
        else:
            # 追加模式：添加新记录
            records = df.to_dict('records')
            added_count = 0
            skipped_count = 0
            
            for record in records:
                # 检查是否已存在（基于费用名称）
                if not original_df.empty:
                    if '费用名称' in original_df.columns:
                        key_col = '费用名称'
                    else:
                        key_col = original_df.columns[0]
                    if key_col in record:
                        existing_value = str(record[key_col])
                        if existing_value in original_df[key_col].astype(str).values:
                            skipped_count += 1
                            continue
                
                if add_manufacturing_cost_record_to_file(record):
                    added_count += 1
            
            return jsonify({
                'success': True,
                'message': f'成功导入 {added_count} 条记录，跳过 {skipped_count} 条重复记录',
                'original_count': original_count,
                'new_count': len(get_manufacturing_cost_dataframe()),
                'imported_count': added_count,
                'skipped_count': skipped_count
            })
        
    except Exception as e:
        print(f"导入制造费用基础数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'数据导入失败: {str(e)}'}), 500

@data_management_bp.route('/manufacturing-cost/export', methods=['GET'])
def export_manufacturing_cost_data():
    """导出制造费用基础数据"""
    try:
        df = get_manufacturing_cost_dataframe()
        
        if df is None or df.empty:
            return jsonify({
                'success': False,
                'error': '没有可导出的制造费用基础数据'
            }), 400
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='制造费用基础数据', index=False)
            
            # 设置列宽
            worksheet = writer.sheets['制造费用基础数据']
            for idx, col in enumerate(df.columns, start=1):
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(idx)
                # 根据列名长度设置列宽
                col_width = max(len(str(col)), 15)
                worksheet.column_dimensions[col_letter].width = min(col_width, 50)
        
        output.seek(0)
        filename = f'制造费用基础数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"导出制造费用基础数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 期间费用基础数据 ====================

def get_period_cost_dataframe():
    """获取期间费用基础数据DataFrame"""
    try:
        from data.base_data import period_cost_data
        return period_cost_data.get_period_cost_dataframe()
    except Exception as e:
        print(f"获取期间费用基础数据失败: {e}")
        traceback.print_exc()
        return pd.DataFrame()

def backup_period_cost_file():
    """备份期间费用基础数据文件"""
    try:
        import shutil
        from datetime import datetime
        
        # 创建备份目录
        backup_dir = 'backups'
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        # 生成备份文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"period_cost_data_backup_{timestamp}.py"
        backup_path = os.path.join(backup_dir, backup_filename)
        
        # 复制文件
        source_path = 'data/base_data/period_cost_data.py'
        if os.path.exists(source_path):
            shutil.copy2(source_path, backup_path)
            print(f"✓ 期间费用基础数据文件已备份到: {backup_path}")
            return True
        else:
            print(f"⚠️ 源文件不存在: {source_path}")
            return False
        
    except Exception as e:
        print(f"✗ 备份期间费用基础数据文件失败: {e}")
        return False

def add_period_cost_record_to_file(new_record):
    """将新的期间费用基础数据记录添加到文件中"""
    try:
        import re
        import importlib
        
        # 读取当前文件内容
        period_cost_file_path = 'data/base_data/period_cost_data.py'
        if not os.path.exists(period_cost_file_path):
            # 如果文件不存在，创建新文件
            with open(period_cost_file_path, 'w', encoding='utf-8') as f:
                f.write('''# 期间费用基础数据 - 内置数据
# 数据来源：期间费用基础数据.xlsx
# 说明：此文件由 scripts/init_period_cost_from_excel.py 自动生成

import pandas as pd
import os

# 完整的期间费用基础数据
PERIOD_COST_DATA = [
]
''')
        
        with open(period_cost_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 构建记录字典（处理所有字段）
        record_parts = []
        for key, value in new_record.items():
            if pd.isna(value) or value is None:
                record_parts.append(f'"{key}": None')
            elif isinstance(value, (int, float)):
                record_parts.append(f'"{key}": {value}')
            else:
                # 字符串类型，需要转义
                str_value = str(value).replace('"', '\\"').replace('\\', '\\\\')
                record_parts.append(f'"{key}": "{str_value}"')
        
        record_str = '{' + ', '.join(record_parts) + '}'
        new_line = f'    {record_str},\n'
        
        # 找到列表的结束位置（最后一个 ] 之前）
        pattern = r'(\s*\{[^}]+\},?\n)(\s*\])'
        
        def replace_func(match):
            last_record = match.group(1)
            closing_bracket = match.group(2)
            
            # 确保最后一个记录有逗号
            if not last_record.rstrip().endswith(','):
                last_record = last_record.rstrip() + ',\n'
            
            return last_record + new_line + closing_bracket
        
        new_content = re.sub(pattern, replace_func, content, flags=re.MULTILINE | re.DOTALL)
        
        # 检查是否成功替换
        if new_content == content:
            # 如果列表为空，需要在列表开始后插入
            pattern2 = r'(PERIOD_COST_DATA = \[)(\s*\])'
            new_content = re.sub(pattern2, r'\1\n' + new_line + r'\2', content)
            if new_content == content:
                raise ValueError("无法找到合适的位置插入新记录")
        
        # 写回文件
        with open(period_cost_file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        # 重新加载模块
        from data.base_data import period_cost_data
        importlib.reload(period_cost_data)
        
        print(f"✓ 已添加期间费用基础数据记录")
        return True
        
    except Exception as e:
        print(f"✗ 添加期间费用基础数据记录失败: {e}")
        traceback.print_exc()
        return False

def update_period_cost_record_in_file(record_id, new_record):
    """更新期间费用基础数据文件中的记录"""
    try:
        import re
        import importlib
        
        # 读取当前文件内容
        period_cost_file_path = 'data/base_data/period_cost_data.py'
        with open(period_cost_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 构建记录字典（处理所有字段）
        record_parts = []
        for key, value in new_record.items():
            if pd.isna(value) or value is None:
                record_parts.append(f'"{key}": None')
            elif isinstance(value, (int, float)):
                record_parts.append(f'"{key}": {value}')
            else:
                # 字符串类型，需要转义
                str_value = str(value).replace('"', '\\"').replace('\\', '\\\\')
                record_parts.append(f'"{key}": "{str_value}"')
        
        record_str = '{' + ', '.join(record_parts) + '}'
        new_line = f'    {record_str},\n'
        
        # 获取"费用明细"作为主键字段（用于匹配记录）
        df = get_period_cost_dataframe()
        if df.empty:
            raise ValueError("数据为空，无法更新")
        
        # 使用"费用明细"作为主键字段
        if '费用明细' in df.columns:
            first_col = '费用明细'
            print(f"[更新记录] 使用'费用明细'作为主键字段，record_id={record_id}")
        else:
            first_col = df.columns[0]
            print(f"[更新记录] 未找到'费用明细'字段，使用第一列: {first_col}")
        
        # 找到并替换对应的记录（匹配费用明细的值）
        old_value = str(record_id)
        # 转义特殊字符用于正则表达式
        escaped_value = re.escape(old_value)
        old_pattern = rf'    \{{[^}}]*"{re.escape(first_col)}":\s*"{escaped_value}"[^}}]*}},?\n'
        
        new_content = re.sub(old_pattern, new_line, content)
        
        # 检查是否找到并替换了记录
        if new_content == content:
            raise ValueError(f"未找到ID为 {record_id} 的记录")
        
        # 写回文件
        with open(period_cost_file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        # 重新加载模块
        from data.base_data import period_cost_data
        importlib.reload(period_cost_data)
        
        print(f"✓ 已更新期间费用基础数据记录: {record_id}")
        return True
        
    except Exception as e:
        print(f"✗ 更新期间费用基础数据记录失败: {e}")
        traceback.print_exc()
        return False

def delete_period_cost_record_from_file(record_id):
    """从期间费用基础数据文件中删除记录"""
    try:
        import re
        import importlib
        
        # 读取当前文件内容
        period_cost_file_path = 'data/base_data/period_cost_data.py'
        with open(period_cost_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 获取"费用明细"作为主键字段（用于匹配记录）
        df = get_period_cost_dataframe()
        if df.empty:
            raise ValueError("数据为空，无法删除")
        
        # 使用"费用明细"作为主键字段
        if '费用明细' in df.columns:
            first_col = '费用明细'
        else:
            first_col = df.columns[0]
        
        # 找到并删除对应的记录
        old_value = str(record_id)
        escaped_value = re.escape(old_value)
        pattern = rf'    \{{[^}}]*"{re.escape(first_col)}":\s*"{escaped_value}"[^}}]*}},?\n'
        new_content = re.sub(pattern, '', content)
        
        # 检查是否找到并删除了记录
        if new_content == content:
            raise ValueError(f"未找到ID为 {record_id} 的记录")
        
        # 写回文件
        with open(period_cost_file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
        
        # 重新加载模块
        from data.base_data import period_cost_data
        importlib.reload(period_cost_data)
        
        print(f"✓ 已删除期间费用基础数据记录: {record_id}")
        return True
        
    except Exception as e:
        print(f"✗ 删除期间费用基础数据记录失败: {e}")
        traceback.print_exc()
        return False

@data_management_bp.route('/period-cost', methods=['GET'])
def get_period_cost_data():
    """获取期间费用基础数据"""
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        search = request.args.get('search', '')
        
        df = get_period_cost_dataframe()
        
        if df is None:
            return jsonify({
                'success': False,
                'error': '无法获取数据，DataFrame为None'
            }), 500
        
        if df.empty:
            return jsonify({
                'success': True,
                'data': [],
                'total': 0,
                'current_page': page,
                'per_page': per_page,
                'pages': 0
            })
        
        # 定义正确的列顺序（按照Excel文件顺序）
        column_order = [
            '费用明细',
            '管理-电废事业部-质量管理',
            '管理-电废事业部-库房',
            '管理-电废事业部-平台',
            '管理-电废事业部-基金管理项目组',
            '管理-电废事业部-回收经营项目组',
            '管理-消电项目组',
            '销售-平台',
            '销售-消电项目组',
            '屏',
        ]
        
        # 重新排列DataFrame的列顺序
        # 先获取存在的列，按照指定顺序排列，然后添加其他未指定的列
        existing_ordered_cols = [col for col in column_order if col in df.columns]
        other_cols = [col for col in df.columns if col not in column_order]
        df = df[existing_ordered_cols + other_cols]
        
        # 搜索过滤（在所有文本列中搜索）
        if search:
            mask = pd.Series([False] * len(df))
            for col in df.columns:
                if df[col].dtype == 'object':  # 文本列
                    mask |= df[col].astype(str).str.contains(search, case=False, na=False)
            df = df[mask]
        
        # 分页
        total = len(df)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        df_page = df.iloc[start_idx:end_idx] if total > 0 else pd.DataFrame()
        
        # 转换数据
        data_list = safe_json_convert(df_page)
        
        return jsonify({
            'success': True,
            'data': data_list,
            'total': total,
            'current_page': page,
            'per_page': per_page,
            'pages': (total + per_page - 1) // per_page
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/period-cost', methods=['POST'])
def create_period_cost_record():
    """创建新的期间费用基础数据记录"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '请提供数据'}), 400
        
        # 获取现有数据以确定必需字段
        df = get_period_cost_dataframe()
        if not df.empty:
            # 使用"费用明细"作为必需字段
            if '费用明细' in df.columns:
                required_col = '费用明细'
            else:
                required_col = df.columns[0]
            if required_col not in data or not data.get(required_col, ''):
                return jsonify({'success': False, 'error': f'字段 {required_col} 不能为空'}), 400
        
        # 构建新记录（包含所有提供的字段）
        new_record = {}
        for key, value in data.items():
            new_record[key] = value
        
        # 检查记录是否已存在（基于费用明细）
        if not df.empty:
            if '费用明细' in df.columns:
                key_col = '费用明细'
            else:
                key_col = df.columns[0]
            if key_col in new_record:
                existing_value = str(new_record[key_col])
                if existing_value in df[key_col].astype(str).values:
                    return jsonify({'success': False, 'error': f'{key_col} {existing_value} 已存在'}), 400
        
        # 备份文件
        backup_period_cost_file()
        
        if add_period_cost_record_to_file(new_record):
            return jsonify({'success': True, 'message': '记录创建成功'})
        else:
            return jsonify({'success': False, 'error': '创建记录失败'}), 500
        
    except Exception as e:
        print(f"创建期间费用基础数据记录错误: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/period-cost/<record_id>', methods=['PUT'])
def update_period_cost_record(record_id):
    """更新期间费用基础数据记录"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '请提供数据'}), 400
        
        # 检查记录是否存在
        df = get_period_cost_dataframe()
        if df is None or df.empty:
            return jsonify({'success': False, 'error': '无法获取期间费用基础数据'}), 500
        
        # 使用"费用明细"作为主键字段
        if '费用明细' in df.columns:
            first_col = '费用明细'
        else:
            first_col = df.columns[0]
        
        print(f"[更新API] 主键字段: {first_col}, record_id: {record_id}")
        
        record_mask = df[first_col].astype(str) == str(record_id)
        if not record_mask.any():
            print(f"[更新API] 未找到记录，搜索的值: {record_id}")
            return jsonify({'success': False, 'error': '记录不存在'}), 404
        
        # 构建更新记录
        updated_record = {}
        for key, value in data.items():
            updated_record[key] = value
        
        # 如果费用明细的值改变了，检查新值是否已存在
        if first_col in updated_record:
            new_value = str(updated_record[first_col])
            if new_value != str(record_id) and new_value in df[first_col].astype(str).values:
                return jsonify({'success': False, 'error': f'{first_col} {new_value} 已存在'}), 400
        
        # 备份文件
        backup_period_cost_file()
        
        if update_period_cost_record_in_file(record_id, updated_record):
            return jsonify({'success': True, 'message': '记录更新成功'})
        else:
            return jsonify({'success': False, 'error': '更新记录失败'}), 500
        
    except Exception as e:
        print(f"更新期间费用基础数据记录错误: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/period-cost/<record_id>', methods=['DELETE'])
def delete_period_cost_record(record_id):
    """删除期间费用基础数据记录"""
    try:
        # 检查记录是否存在
        df = get_period_cost_dataframe()
        if df is None or df.empty:
            return jsonify({'success': False, 'error': '无法获取期间费用基础数据'}), 500
        
        # 使用"费用明细"作为主键字段
        if '费用明细' in df.columns:
            first_col = '费用明细'
        else:
            first_col = df.columns[0]
        
        record_mask = df[first_col].astype(str) == str(record_id)
        if not record_mask.any():
            return jsonify({'success': False, 'error': '记录不存在'}), 404
        
        # 备份文件
        backup_period_cost_file()
        
        # 删除记录
        if delete_period_cost_record_from_file(record_id):
            return jsonify({'success': True, 'message': '记录删除成功'})
        else:
            return jsonify({'success': False, 'error': '删除记录失败'}), 500
        
    except Exception as e:
        print(f"删除期间费用基础数据记录错误: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@data_management_bp.route('/period-cost/import', methods=['POST'])
def import_period_cost_data():
    """导入期间费用基础数据"""
    try:
        # 检查文件
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '请选择要导入的Excel文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '请选择要导入的Excel文件'}), 400
        
        # 检查文件格式
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': '只支持Excel文件格式(.xlsx, .xls)'}), 400
        
        # 获取导入模式
        import_mode = request.form.get('import_mode', 'append')
        
        # 获取原有数据统计
        original_df = get_period_cost_dataframe()
        original_count = len(original_df) if original_df is not None and not original_df.empty else 0
        
        # 读取Excel文件
        try:
            df = pd.read_excel(file, engine='openpyxl' if file.filename.endswith('.xlsx') else 'xlrd')
        except Exception as e:
            return jsonify({'success': False, 'error': f'文件读取失败: {str(e)}'}), 400
        
        # 删除完全为空的行
        df = df.dropna(how='all')
        
        if df.empty:
            return jsonify({'success': False, 'error': 'Excel文件中没有有效数据'}), 400
        
        # 备份文件
        backup_period_cost_file()
        
        # 根据导入模式处理数据
        if import_mode == 'replace':
            # 覆盖模式：重新生成整个文件
            new_content = '''# 期间费用基础数据 - 内置数据
# 数据来源：期间费用基础数据.xlsx
# 说明：此文件由 scripts/init_period_cost_from_excel.py 自动生成

import pandas as pd
import os

# 完整的期间费用基础数据
PERIOD_COST_DATA = [
'''
            
            # 添加所有记录
            for index, row in df.iterrows():
                record_parts = []
                for col in df.columns:
                    value = row[col]
                    if pd.isna(value):
                        record_parts.append(f'"{col}": None')
                    elif isinstance(value, (int, float)):
                        record_parts.append(f'"{col}": {value}')
                    else:
                        str_value = str(value).replace('"', '\\"').replace('\\', '\\\\')
                        record_parts.append(f'"{col}": "{str_value}"')
                
                record_str = '{' + ', '.join(record_parts) + '}'
                new_content += f'    {record_str},\n'
            
            new_content += ''']

def get_period_cost_dataframe():
    """获取期间费用基础数据DataFrame"""
    return pd.DataFrame(PERIOD_COST_DATA)

def filter_by_category(category):
    """根据类别筛选数据（如果有类别字段）"""
    df = get_period_cost_dataframe()
    if df.empty:
        return df
    if '类别' in df.columns:
        return df[df['类别'] == category]
    return df

def get_all_categories():
    """获取所有类别列表（如果有类别字段）"""
    df = get_period_cost_dataframe()
    if df.empty:
        return []
    if '类别' in df.columns:
        return df['类别'].unique().tolist()
    return []

def get_category_stats():
    """获取类别统计信息（如果有类别字段）"""
    df = get_period_cost_dataframe()
    if df.empty:
        return {}
    if '类别' in df.columns:
        return df['类别'].value_counts().to_dict()
    return {}

def get_period_cost_by_id(record_id):
    """根据记录ID获取期间费用基础数据"""
    df = get_period_cost_dataframe()
    if df.empty:
        return None
    # 使用"费用明细"作为主键
    if len(df) > 0:
        if '费用明细' in df.columns:
            result = df[df['费用明细'].astype(str) == str(record_id)]
            if len(result) > 0:
                return result.iloc[0].to_dict()
        # 如果找不到费用明细字段，尝试使用第一列
        first_col = df.columns[0]
        if 'ID' in first_col.upper() or '代码' in first_col or '编号' in first_col:
            result = df[df[first_col].astype(str) == str(record_id)]
        else:
            try:
                idx = int(record_id)
                if 0 <= idx < len(df):
                    result = df.iloc[[idx]]
                else:
                    return None
            except (ValueError, TypeError):
                return None
        
        if len(result) > 0:
            return result.iloc[0].to_dict()
    return None
'''
            
            # 写入文件
            with open('data/base_data/period_cost_data.py', 'w', encoding='utf-8') as f:
                f.write(new_content)
            
            # 重新加载模块
            import importlib
            from data.base_data import period_cost_data
            importlib.reload(period_cost_data)
            
            new_count = len(df)
            return jsonify({
                'success': True,
                'message': f'成功导入 {new_count} 条记录（覆盖模式）',
                'original_count': original_count,
                'new_count': new_count,
                'imported_count': new_count
            })
        else:
            # 追加模式：添加新记录
            records = df.to_dict('records')
            added_count = 0
            skipped_count = 0
            
            for record in records:
                # 检查是否已存在（基于费用明细）
                if not original_df.empty:
                    if '费用明细' in original_df.columns:
                        key_col = '费用明细'
                    else:
                        key_col = original_df.columns[0]
                    if key_col in record:
                        existing_value = str(record[key_col])
                        if existing_value in original_df[key_col].astype(str).values:
                            skipped_count += 1
                            continue
                
                if add_period_cost_record_to_file(record):
                    added_count += 1
            
            return jsonify({
                'success': True,
                'message': f'成功导入 {added_count} 条记录，跳过 {skipped_count} 条重复记录',
                'original_count': original_count,
                'new_count': len(get_period_cost_dataframe()),
                'imported_count': added_count,
                'skipped_count': skipped_count
            })
        
    except Exception as e:
        print(f"导入期间费用基础数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'数据导入失败: {str(e)}'}), 500

@data_management_bp.route('/period-cost/export', methods=['GET'])
def export_period_cost_data():
    """导出期间费用基础数据"""
    try:
        df = get_period_cost_dataframe()
        
        if df is None or df.empty:
            return jsonify({
                'success': False,
                'error': '没有可导出的期间费用基础数据'
            }), 400
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 写入数据
            df.to_excel(writer, sheet_name='期间费用基础数据', index=False)
            
            # 设置列宽和样式
            worksheet = writer.sheets['期间费用基础数据']
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF", name="仿宋")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # 列宽设置
            column_widths = {
                '费用明细': 30,
                '管理-电废事业部-质量管理': 22,
                '管理-电废事业部-库房': 22,
                '管理-电废事业部-平台': 22,
                '管理-电废事业部-基金管理项目组': 28,
                '管理-电废事业部-回收经营项目组': 28,
                '管理-消电项目组': 20,
                '销售-平台': 18,
                '销售-消电项目组': 20,
                '屏': 14,
            }
            
            # 设置表头样式
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                
                # 设置列宽
                col_letter = get_column_letter(col)
                col_name = df.columns[col - 1]
                if col_name in column_widths:
                    worksheet.column_dimensions[col_letter].width = column_widths[col_name]
                else:
                    worksheet.column_dimensions[col_letter].width = 18
            
            # 设置数据行样式
            data_font = Font(name="仿宋")
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = data_font
                    cell.alignment = center_alignment
                    
                    # 费用明细列高亮
                    col_name = df.columns[col - 1]
                    if col_name == '费用明细':
                        cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                        cell.font = Font(name="仿宋", bold=True)
        
        output.seek(0)
        filename = f'期间费用基础数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"导出期间费用基础数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 税金及附加基础数据 ====================

def get_tax_surcharge_dataframe():
    """获取税金及附加基础数据DataFrame"""
    try:
        from data.base_data import tax_surcharge_data
        import importlib
        importlib.reload(tax_surcharge_data)
        return tax_surcharge_data.get_tax_surcharge_dataframe()
    except Exception as e:
        print(f"获取税金及附加基础数据失败: {e}")
        traceback.print_exc()
        return pd.DataFrame()


def backup_tax_surcharge_file():
    """备份税金及附加基础数据文件"""
    try:
        import shutil
        from datetime import datetime

        backup_dir = 'backups'
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"tax_surcharge_data_backup_{timestamp}.py"
        backup_path = os.path.join(backup_dir, backup_filename)

        source_path = 'data/base_data/tax_surcharge_data.py'
        if os.path.exists(source_path):
            shutil.copy2(source_path, backup_path)
            print(f"✓ 税金及附加基础数据文件已备份到: {backup_path}")
            return True
        print(f"⚠️ 源文件不存在: {source_path}")
        return False
    except Exception as e:
        print(f"✗ 备份税金及附加基础数据文件失败: {e}")
        return False


def _format_tax_surcharge_record_value(value):
    """格式化记录字段值为 Python 字面量字符串"""
    if pd.isna(value) or value is None:
        return 'None'
    if isinstance(value, (int, float)):
        return str(value)
    str_value = str(value).replace('"', '\\"').replace('\\', '\\\\')
    return f'"{str_value}"'


def add_tax_surcharge_record_to_file(new_record):
    """将新的税金及附加基础数据记录添加到文件中"""
    try:
        import re
        import importlib

        tax_surcharge_file_path = 'data/base_data/tax_surcharge_data.py'
        if not os.path.exists(tax_surcharge_file_path):
            with open(tax_surcharge_file_path, 'w', encoding='utf-8') as f:
                f.write('''# 税金及附加基础数据 - 内置数据

import pandas as pd

TAX_SURCHARGE_DATA = [
]
''')

        with open(tax_surcharge_file_path, 'r', encoding='utf-8') as f:
            content = f.read()

        record_parts = [
            f'"{key}": {_format_tax_surcharge_record_value(value)}'
            for key, value in new_record.items()
        ]
        record_str = '{' + ', '.join(record_parts) + '}'
        new_line = f'    {record_str},\n'

        pattern = r'(\s*\{[^}]+\},?\n)(\s*\])'

        def replace_func(match):
            last_record = match.group(1)
            closing_bracket = match.group(2)
            if not last_record.rstrip().endswith(','):
                last_record = last_record.rstrip() + ',\n'
            return last_record + new_line + closing_bracket

        new_content = re.sub(pattern, replace_func, content, flags=re.MULTILINE | re.DOTALL)
        if new_content == content:
            pattern2 = r'(TAX_SURCHARGE_DATA = \[)(\s*\])'
            new_content = re.sub(pattern2, r'\1\n' + new_line + r'\2', content)
            if new_content == content:
                raise ValueError("无法找到合适的位置插入新记录")

        with open(tax_surcharge_file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)

        from data.base_data import tax_surcharge_data
        importlib.reload(tax_surcharge_data)
        print("✓ 已添加税金及附加基础数据记录")
        return True
    except Exception as e:
        print(f"✗ 添加税金及附加基础数据记录失败: {e}")
        traceback.print_exc()
        return False


def update_tax_surcharge_record_in_file(record_id, new_record):
    """更新税金及附加基础数据文件中的记录"""
    try:
        import re
        import importlib

        tax_surcharge_file_path = 'data/base_data/tax_surcharge_data.py'
        with open(tax_surcharge_file_path, 'r', encoding='utf-8') as f:
            content = f.read()

        record_parts = [
            f'"{key}": {_format_tax_surcharge_record_value(value)}'
            for key, value in new_record.items()
        ]
        record_str = '{' + ', '.join(record_parts) + '}'
        new_line = f'    {record_str},\n'

        df = get_tax_surcharge_dataframe()
        if df.empty:
            raise ValueError("数据为空，无法更新")

        first_col = '项目' if '项目' in df.columns else df.columns[0]
        escaped_value = re.escape(str(record_id))
        old_pattern = rf'    \{{[^}}]*"{re.escape(first_col)}":\s*"{escaped_value}"[^}}]*}},?\n'
        new_content = re.sub(old_pattern, new_line, content)
        if new_content == content:
            raise ValueError(f"未找到ID为 {record_id} 的记录")

        with open(tax_surcharge_file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)

        from data.base_data import tax_surcharge_data
        importlib.reload(tax_surcharge_data)
        print(f"✓ 已更新税金及附加基础数据记录: {record_id}")
        return True
    except Exception as e:
        print(f"✗ 更新税金及附加基础数据记录失败: {e}")
        traceback.print_exc()
        return False


def delete_tax_surcharge_record_from_file(record_id):
    """从税金及附加基础数据文件中删除记录"""
    try:
        import re
        import importlib

        tax_surcharge_file_path = 'data/base_data/tax_surcharge_data.py'
        with open(tax_surcharge_file_path, 'r', encoding='utf-8') as f:
            content = f.read()

        df = get_tax_surcharge_dataframe()
        if df.empty:
            raise ValueError("数据为空，无法删除")

        first_col = '项目' if '项目' in df.columns else df.columns[0]
        escaped_value = re.escape(str(record_id))
        pattern = rf'    \{{[^}}]*"{re.escape(first_col)}":\s*"{escaped_value}"[^}}]*}},?\n'
        new_content = re.sub(pattern, '', content)
        if new_content == content:
            raise ValueError(f"未找到ID为 {record_id} 的记录")

        with open(tax_surcharge_file_path, 'w', encoding='utf-8') as f:
            f.write(new_content)

        from data.base_data import tax_surcharge_data
        importlib.reload(tax_surcharge_data)
        print(f"✓ 已删除税金及附加基础数据记录: {record_id}")
        return True
    except Exception as e:
        print(f"✗ 删除税金及附加基础数据记录失败: {e}")
        traceback.print_exc()
        return False


@data_management_bp.route('/tax-surcharge', methods=['GET'])
def get_tax_surcharge_data():
    """获取税金及附加基础数据"""
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        search = request.args.get('search', '')

        df = get_tax_surcharge_dataframe()
        if df is None:
            return jsonify({'success': False, 'error': '无法获取数据，DataFrame为None'}), 500

        if df.empty:
            return jsonify({
                'success': True,
                'data': [],
                'total': 0,
                'current_page': page,
                'per_page': per_page,
                'pages': 0
            })

        column_order = ['项目', '金额', '备注']
        existing_ordered_cols = [col for col in column_order if col in df.columns]
        other_cols = [col for col in df.columns if col not in column_order]
        df = df[existing_ordered_cols + other_cols]

        if search:
            mask = pd.Series([False] * len(df))
            for col in df.columns:
                mask |= df[col].astype(str).str.contains(search, case=False, na=False)
            df = df[mask]

        total = len(df)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        df_page = df.iloc[start_idx:end_idx] if total > 0 else pd.DataFrame()
        data_list = safe_json_convert(df_page)

        return jsonify({
            'success': True,
            'data': data_list,
            'total': total,
            'current_page': page,
            'per_page': per_page,
            'pages': (total + per_page - 1) // per_page
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/tax-surcharge', methods=['POST'])
def create_tax_surcharge_record():
    """创建新的税金及附加基础数据记录"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '请提供数据'}), 400

        df = get_tax_surcharge_dataframe()
        required_col = '项目'
        if required_col not in data or not str(data.get(required_col, '')).strip():
            return jsonify({'success': False, 'error': f'字段 {required_col} 不能为空'}), 400

        new_record = {key: value for key, value in data.items()}

        if not df.empty and '项目' in df.columns:
            existing_value = str(new_record['项目'])
            if existing_value in df['项目'].astype(str).values:
                return jsonify({'success': False, 'error': f'项目 {existing_value} 已存在'}), 400

        backup_tax_surcharge_file()
        if add_tax_surcharge_record_to_file(new_record):
            return jsonify({'success': True, 'message': '记录创建成功'})
        return jsonify({'success': False, 'error': '创建记录失败'}), 500
    except Exception as e:
        print(f"创建税金及附加基础数据记录错误: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/tax-surcharge/<record_id>', methods=['PUT'])
def update_tax_surcharge_record(record_id):
    """更新税金及附加基础数据记录"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '请提供数据'}), 400

        df = get_tax_surcharge_dataframe()
        if df is None or df.empty:
            return jsonify({'success': False, 'error': '无法获取税金及附加基础数据'}), 500

        first_col = '项目'
        record_mask = df[first_col].astype(str) == str(record_id)
        if not record_mask.any():
            return jsonify({'success': False, 'error': '记录不存在'}), 404

        updated_record = {key: value for key, value in data.items()}
        if first_col in updated_record:
            new_value = str(updated_record[first_col])
            if new_value != str(record_id) and new_value in df[first_col].astype(str).values:
                return jsonify({'success': False, 'error': f'项目 {new_value} 已存在'}), 400

        backup_tax_surcharge_file()
        if update_tax_surcharge_record_in_file(record_id, updated_record):
            return jsonify({'success': True, 'message': '记录更新成功'})
        return jsonify({'success': False, 'error': '更新记录失败'}), 500
    except Exception as e:
        print(f"更新税金及附加基础数据记录错误: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/tax-surcharge/<record_id>', methods=['DELETE'])
def delete_tax_surcharge_record(record_id):
    """删除税金及附加基础数据记录"""
    try:
        df = get_tax_surcharge_dataframe()
        if df is None or df.empty:
            return jsonify({'success': False, 'error': '无法获取税金及附加基础数据'}), 500

        first_col = '项目'
        record_mask = df[first_col].astype(str) == str(record_id)
        if not record_mask.any():
            return jsonify({'success': False, 'error': '记录不存在'}), 404

        backup_tax_surcharge_file()
        if delete_tax_surcharge_record_from_file(record_id):
            return jsonify({'success': True, 'message': '记录删除成功'})
        return jsonify({'success': False, 'error': '删除记录失败'}), 500
    except Exception as e:
        print(f"删除税金及附加基础数据记录错误: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/tax-surcharge/import', methods=['POST'])
def import_tax_surcharge_data():
    """导入税金及附加基础数据"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '请选择要导入的Excel文件'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '请选择要导入的Excel文件'}), 400

        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': '只支持Excel文件格式(.xlsx, .xls)'}), 400

        import_mode = request.form.get('import_mode', 'append')
        original_df = get_tax_surcharge_dataframe()
        original_count = len(original_df) if original_df is not None and not original_df.empty else 0

        try:
            df = pd.read_excel(file, engine='openpyxl' if file.filename.endswith('.xlsx') else 'xlrd')
        except Exception as e:
            return jsonify({'success': False, 'error': f'文件读取失败: {str(e)}'}), 400

        df = df.dropna(how='all')
        if df.empty:
            return jsonify({'success': False, 'error': 'Excel文件中没有有效数据'}), 400

        backup_tax_surcharge_file()

        if import_mode == 'replace':
            new_content = '''# 税金及附加基础数据 - 内置数据

import pandas as pd

# 完整的税金及附加基础数据
TAX_SURCHARGE_DATA = [
'''
            for _, row in df.iterrows():
                record_parts = []
                for col in df.columns:
                    value = row[col]
                    record_parts.append(f'"{col}": {_format_tax_surcharge_record_value(value)}')
                record_str = '{' + ', '.join(record_parts) + '}'
                new_content += f'    {record_str},\n'

            new_content += '''
]


def get_tax_surcharge_dataframe():
    """获取税金及附加基础数据DataFrame"""
    return pd.DataFrame(TAX_SURCHARGE_DATA)
'''
            with open('data/base_data/tax_surcharge_data.py', 'w', encoding='utf-8') as f:
                f.write(new_content)

            import importlib
            from data.base_data import tax_surcharge_data
            importlib.reload(tax_surcharge_data)

            new_count = len(df)
            return jsonify({
                'success': True,
                'message': f'成功导入 {new_count} 条记录（覆盖模式）',
                'original_count': original_count,
                'new_count': new_count,
                'imported_count': new_count
            })

        records = df.to_dict('records')
        added_count = 0
        skipped_count = 0
        for record in records:
            if not original_df.empty and '项目' in original_df.columns and '项目' in record:
                if str(record['项目']) in original_df['项目'].astype(str).values:
                    skipped_count += 1
                    continue
            if add_tax_surcharge_record_to_file(record):
                added_count += 1

        return jsonify({
            'success': True,
            'message': f'成功导入 {added_count} 条记录，跳过 {skipped_count} 条重复记录',
            'original_count': original_count,
            'new_count': len(get_tax_surcharge_dataframe()),
            'imported_count': added_count,
            'skipped_count': skipped_count
        })
    except Exception as e:
        print(f"导入税金及附加基础数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'数据导入失败: {str(e)}'}), 500


@data_management_bp.route('/tax-surcharge/export', methods=['GET'])
def export_tax_surcharge_data():
    """导出税金及附加基础数据"""
    try:
        df = get_tax_surcharge_dataframe()
        if df is None or df.empty:
            return jsonify({'success': False, 'error': '没有可导出的税金及附加基础数据'}), 400

        column_order = ['项目', '金额', '备注']
        existing_ordered_cols = [col for col in column_order if col in df.columns]
        other_cols = [col for col in df.columns if col not in column_order]
        df = df[existing_ordered_cols + other_cols]

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='税金及附加基础数据', index=False)
            worksheet = writer.sheets['税金及附加基础数据']
            from openpyxl.utils import get_column_letter
            for idx, col in enumerate(df.columns, start=1):
                col_letter = get_column_letter(idx)
                col_width = max(len(str(col)), 15)
                worksheet.column_dimensions[col_letter].width = min(col_width, 50)

        output.seek(0)
        filename = f'税金及附加基础数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"导出税金及附加基础数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/revenue-forecast/subsidy-income', methods=['GET'])
def get_subsidy_income():
    """获取基金补贴收入数据"""
    try:
        app_data = get_session_data_manager()
        
        # 检查数据是否已被清除
        data_cleared = app_data.get_data('__data_cleared__')
        if data_cleared:
            # 数据已被清除，直接返回空数据
            return jsonify({
                'success': True,
                'data': {
                    'total_count': 0,
                    'category_count': 0,
                    'total_income': 0.0,
                    'category_stats': [],
                    'details': []
                },
                'message': '数据已清除'
            })
        
        # 获取基金补贴收入数据
        subsidy_income_data = app_data.get_data('subsidy_income_data')
        
        if subsidy_income_data is None or subsidy_income_data.empty:
            return jsonify({
                'success': True,
                'data': {
                    'total_count': 0,
                    'category_count': 0,
                    'total_income': 0.0,
                    'category_stats': [],
                    'details': []
                },
                'message': '暂无基金补贴收入数据'
            })
        
        # 确保数值类型
        if '基金补贴收入(元)' in subsidy_income_data.columns:
            subsidy_income_data['基金补贴收入(元)'] = pd.to_numeric(
                subsidy_income_data['基金补贴收入(元)'], errors='coerce'
            ).fillna(0)
        
        # 处理列名映射：支持实际数据中的列名（与EXCEL一致）
        # 实际数据使用 '当期拆解量(台)' 和 '补贴单价(元/台)'
        # 前端显示使用 '本期实际投产数量' 和 '基金补贴单价(元)'
        quantity_col = None
        price_col = None
        
        # 优先使用实际数据中的列名
        if '当期拆解量(台)' in subsidy_income_data.columns:
            quantity_col = '当期拆解量(台)'
            subsidy_income_data[quantity_col] = pd.to_numeric(
                subsidy_income_data[quantity_col], errors='coerce'
            ).fillna(0)
        elif '本期实际投产数量' in subsidy_income_data.columns:
            quantity_col = '本期实际投产数量'
            subsidy_income_data[quantity_col] = pd.to_numeric(
                subsidy_income_data[quantity_col], errors='coerce'
            ).fillna(0)
        
        if '补贴单价(元/台)' in subsidy_income_data.columns:
            price_col = '补贴单价(元/台)'
            subsidy_income_data[price_col] = pd.to_numeric(
                subsidy_income_data[price_col], errors='coerce'
            ).fillna(0)
        elif '基金补贴单价(元)' in subsidy_income_data.columns:
            price_col = '基金补贴单价(元)'
            subsidy_income_data[price_col] = pd.to_numeric(
                subsidy_income_data[price_col], errors='coerce'
            ).fillna(0)
        
        # 将补贴大类映射到产品类型（五大类：电视机、冰箱、洗衣机、空调、电脑）
        def map_subsidy_category_to_product_type(subsidy_category):
            """将补贴大类映射到产品类型"""
            subsidy_category = str(subsidy_category).strip()
            # 空调的细分补贴大类
            if subsidy_category in ['整机', '内机', '外机']:
                return '空调'
            # 电脑的细分补贴大类
            elif subsidy_category in ['笔记本', '显示器', '主机']:
                return '电脑'
            # 其他直接返回（电视机、冰箱、洗衣机）
            elif subsidy_category in ['电视机', '冰箱', '洗衣机']:
                return subsidy_category
            else:
                # 默认返回原值（兼容旧数据）
                return subsidy_category
        
        # 按产品类型（五大类）统计
        category_stats = []
        if '补贴大类' in subsidy_income_data.columns:
            # 添加产品类型列用于统计
            subsidy_income_data['产品类型'] = subsidy_income_data['补贴大类'].apply(map_subsidy_category_to_product_type)
            
            # 按产品类型分组统计
            product_type_groups = subsidy_income_data.groupby('产品类型').agg({
                '基金补贴收入(元)': 'sum'
            }).reset_index()
            
            for _, row in product_type_groups.iterrows():
                product_type = row['产品类型']
                income = float(row['基金补贴收入(元)'])
                count = len(subsidy_income_data[subsidy_income_data['产品类型'] == product_type])
                category_stats.append({
                    'category': product_type,
                    'count': count,
                    'income': round(income, 2)
                })
            
            # 按收入排序
            category_stats.sort(key=lambda x: x['income'], reverse=True)
        
        # 构建明细数据
        details = []
        for _, row in subsidy_income_data.iterrows():
            # 使用正确的列名读取数据（与EXCEL一致）
            unit_price = float(row.get(price_col, 0)) if price_col else 0
            quantity = float(row.get(quantity_col, 0)) if quantity_col else 0
            
            # 获取补贴大类（小类）
            subsidy_category = str(row.get('补贴大类', ''))
            # 将补贴大类映射为五大类（用于显示）
            mapped_category = map_subsidy_category_to_product_type(subsidy_category)
            
            details.append({
                'material_code': str(row.get('物料代码', '')),
                'material_desc': str(row.get('物料描述', '')),
                'category': mapped_category,  # 映射后的五大类（用于显示）
                'sub_category': subsidy_category,  # 原始补贴大类（小类显示，如"整机"、"内机"、"外机"等）
                'unit_price': unit_price,
                'quantity': quantity,
                'income': float(row.get('基金补贴收入(元)', 0))
            })
        
        # 计算汇总
        total_income = float(subsidy_income_data['基金补贴收入(元)'].sum()) if '基金补贴收入(元)' in subsidy_income_data.columns else 0.0
        category_count = len(category_stats)
        
        return jsonify({
            'success': True,
            'data': {
                'total_count': len(subsidy_income_data),
                'category_count': category_count,
                'total_income': round(total_income, 2),
                'category_stats': category_stats,
                'details': details
            }
        })
        
    except Exception as e:
        print(f"获取基金补贴收入数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/revenue-forecast/sales-revenue', methods=['GET'])
def get_sales_revenue():
    """获取销售收益汇总数据"""
    try:
        app_data = get_session_data_manager()
        
        # 检查数据是否已被清除
        data_cleared = app_data.get_data('__data_cleared__')
        if data_cleared:
            # 数据已被清除，直接返回空数据
            return jsonify({
                'success': True,
                'data': {
                    'total_count': 0,
                    'total_revenue': 0.0
                },
                'message': '数据已清除'
            })
        
        # 优先使用手工数据，否则使用系统数据
        saleable_data_manual = app_data.get_data('saleable_data_manual')
        saleable_data = app_data.get_data('saleable_data')
        saleable_data_modified = app_data.get_data('saleable_data_modified')
        
        print(f"📊 获取销售收益数据 - 手工数据: {saleable_data_manual is not None and not saleable_data_manual.empty if saleable_data_manual is not None else False}, 系统数据: {saleable_data is not None and not saleable_data.empty if saleable_data is not None else False}, 已修改: {saleable_data_modified}")
        
        revenue_data = None
        data_source = None
        
        # 检查是否有手工数据且包含收益列，并且已标记为修改（与导出结果逻辑一致）
        if saleable_data_manual is not None and not saleable_data_manual.empty and saleable_data_modified:
            print(f"  手工数据列: {list(saleable_data_manual.columns)}")
            if '销售收益(元)' in saleable_data_manual.columns:
                revenue_data = saleable_data_manual
                data_source = '手工'
                print(f"  ✅ 使用手工数据，记录数: {len(revenue_data)}")
            else:
                print(f"  ⚠️ 手工数据缺少'销售收益(元)'列")
        
        # 如果没有手工数据或未修改，使用系统数据
        if revenue_data is None and saleable_data is not None and not saleable_data.empty:
            print(f"  系统数据列: {list(saleable_data.columns)}")
            if '销售收益(元)' in saleable_data.columns:
                revenue_data = saleable_data
                data_source = '系统'
                print(f"  ✅ 使用系统数据，记录数: {len(revenue_data)}")
            else:
                print(f"  ⚠️ 系统数据缺少'销售收益(元)'列")
        
        if revenue_data is None or revenue_data.empty:
            print("  ❌ 没有可用的销售收益数据")
            return jsonify({
                'success': True,
                'data': {
                    'total_count': 0,
                    'total_revenue': 0.0
                },
                'message': '暂无销售收益数据，请先完成数据处理流程'
            })
        
        # 确保数值类型
        if '销售收益(元)' in revenue_data.columns:
            revenue_data['销售收益(元)'] = pd.to_numeric(
                revenue_data['销售收益(元)'], errors='coerce'
            ).fillna(0)
        
        # 计算汇总
        total_revenue = float(revenue_data['销售收益(元)'].sum()) if '销售收益(元)' in revenue_data.columns else 0.0
        
        print(f"  ✅ 销售收益汇总完成 - 数据源: {data_source}, 记录数: {len(revenue_data)}, 总收益: {total_revenue:.2f} 元")
        
        return jsonify({
            'success': True,
            'data': {
                'total_count': len(revenue_data),
                'total_revenue': round(total_revenue, 2)
            }
        })
        
    except Exception as e:
        print(f"获取销售收益数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/revenue-forecast/sales-revenue/detail', methods=['GET'])
def get_sales_revenue_detail():
    """获取销售收益详细数据（排除是否减扣、减扣说明、处置类别三列）"""
    try:
        app_data = get_session_data_manager()
        
        # 优先使用手工数据，否则使用系统数据
        saleable_data_manual = app_data.get_data('saleable_data_manual')
        saleable_data = app_data.get_data('saleable_data')
        saleable_data_modified = app_data.get_data('saleable_data_modified')
        
        print(f"📊 获取销售收益详细数据 - 手工数据: {saleable_data_manual is not None and not saleable_data_manual.empty if saleable_data_manual is not None else False}, 系统数据: {saleable_data is not None and not saleable_data.empty if saleable_data is not None else False}, 已修改: {saleable_data_modified}")
        
        revenue_data = None
        data_source = None
        
        # 检查是否有手工数据且包含收益列，并且已标记为修改（与导出结果逻辑一致）
        if saleable_data_manual is not None and not saleable_data_manual.empty and saleable_data_modified:
            print(f"  手工数据列: {list(saleable_data_manual.columns)}")
            if '销售收益(元)' in saleable_data_manual.columns:
                revenue_data = saleable_data_manual.copy()
                data_source = '手工'
                print(f"  ✅ 使用手工数据，记录数: {len(revenue_data)}")
            else:
                print(f"  ⚠️ 手工数据缺少'销售收益(元)'列")
        
        # 如果没有手工数据或未修改，使用系统数据
        if revenue_data is None and saleable_data is not None and not saleable_data.empty:
            print(f"  系统数据列: {list(saleable_data.columns)}")
            if '销售收益(元)' in saleable_data.columns:
                revenue_data = saleable_data.copy()
                data_source = '系统'
                print(f"  ✅ 使用系统数据，记录数: {len(revenue_data)}")
            else:
                print(f"  ⚠️ 系统数据缺少'销售收益(元)'列")
        
        if revenue_data is None or revenue_data.empty:
            print("  ❌ 没有可用的销售收益数据")
            return jsonify({
                'success': True,
                'data': {
                    'total_count': 0,
                    'total_revenue': 0.0,
                    'details': []
                },
                'message': '暂无销售收益数据，请先完成数据处理流程'
            })
        
        # 排除指定三列
        columns_to_exclude = ['是否减扣', '减扣说明', '处置类别']
        filtered_columns = [col for col in revenue_data.columns if col not in columns_to_exclude]
        revenue_data_filtered = revenue_data[filtered_columns].copy()  # 使用.copy()避免SettingWithCopyWarning
        
        print(f"  过滤后列数: {len(filtered_columns)} (原始: {len(revenue_data.columns)})")
        
        # 确保数值类型
        if '销售收益(元)' in revenue_data_filtered.columns:
            revenue_data_filtered['销售收益(元)'] = pd.to_numeric(
                revenue_data_filtered['销售收益(元)'], errors='coerce'
            ).fillna(0)
        
        # 转换为字典列表，并处理NaN值
        import numpy as np
        
        # 先替换DataFrame中的NaN值（使用replace方法，不改变数据类型）
        revenue_data_filtered = revenue_data_filtered.replace([np.nan, np.inf, -np.inf], None)
        
        # 转换为字典列表
        details = revenue_data_filtered.to_dict('records')
        
        # 清理字典中的NaN值（双重保险）
        for record in details:
            for key, value in record.items():
                # 检查各种NaN情况
                try:
                    if value is None:
                        continue
                    elif pd.isna(value):
                        record[key] = None
                    elif isinstance(value, float) and (value != value):  # NaN检查 (NaN != NaN)
                        record[key] = None
                    elif isinstance(value, (pd.Timestamp, pd.Timedelta)):
                        record[key] = str(value)
                    elif hasattr(value, 'item'):  # numpy标量类型
                        try:
                            py_value = value.item()
                            if pd.isna(py_value) or (isinstance(py_value, float) and (py_value != py_value)):
                                record[key] = None
                            else:
                                record[key] = py_value
                        except (ValueError, AttributeError):
                            record[key] = None
                except Exception as e:
                    # 如果处理失败，设置为None
                    print(f"  警告: 处理字段 {key} 的值时出错: {e}")
                    record[key] = None
        
        # 计算汇总
        total_revenue = float(revenue_data_filtered['销售收益(元)'].sum()) if '销售收益(元)' in revenue_data_filtered.columns else 0.0
        
        print(f"  ✅ 销售收益详细数据获取完成 - 数据源: {data_source}, 记录数: {len(revenue_data_filtered)}, 总收益: {total_revenue:.2f} 元")
        
        return jsonify({
            'success': True,
            'data': {
                'total_count': len(revenue_data_filtered),
                'total_revenue': round(total_revenue, 2),
                'details': details
            }
        })
        
    except Exception as e:
        print(f"获取销售收益详细数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/revenue-forecast/sales-revenue/by-category', methods=['GET'])
def get_sales_revenue_by_category():
    """获取按分类汇总的销售收益数据（与首页和销售收益页共用）"""
    try:
        app_data = get_session_data_manager()
        
        # 分类映射函数（与statistics_api.py中的map_category_py保持一致）
        def map_category_py(material_name):
            """映射物料名称到四机一脑分类（Python版本）"""
            if pd.isna(material_name) or not material_name:
                return None
            name = str(material_name)
            
            # 电视映射规则（按优先级顺序匹配）
            if ('CRT其它机壳破碎塑料' in name or '线路板边框破碎塑料' in name or 
                '废旧玻璃电子枪' in name or '废旧金属荫罩压块铁' in name or 
                '黑白' in name or '电视' in name or '彩电' in name or '等离子' in name):
                return '电视'
            
            # 电脑映射规则
            if ('电脑' in name or '显示器' in name or '笔记本' in name or 
                '主机' in name or '废旧金属黑色金属-铁及其合金-电子枪' in name):
                return '电脑'
            
            # 冰箱
            if '冰箱' in name or '冰柜' in name:
                return '冰箱'
            
            # 空调
            if '空调' in name:
                return '空调'
            
            # 洗衣机
            if '洗衣机' in name or '双缸' in name:
                return '洗衣机'
            
            return None
        
        # 优先使用手工数据，否则使用系统数据（与get_sales_revenue_detail逻辑一致）
        saleable_data_manual = app_data.get_data('saleable_data_manual')
        saleable_data = app_data.get_data('saleable_data')
        saleable_data_modified = app_data.get_data('saleable_data_modified')
        
        print(f"📊 获取按分类汇总的销售收益数据 - 手工数据: {saleable_data_manual is not None and not saleable_data_manual.empty if saleable_data_manual is not None else False}, 系统数据: {saleable_data is not None and not saleable_data.empty if saleable_data is not None else False}, 已修改: {saleable_data_modified}")
        
        revenue_data = None
        data_source = None
        
        # 检查是否有手工数据且包含收益列，并且已标记为修改（与导出结果逻辑一致）
        if saleable_data_manual is not None and not saleable_data_manual.empty and saleable_data_modified:
            print(f"  手工数据列: {list(saleable_data_manual.columns)}")
            if '销售收益(元)' in saleable_data_manual.columns:
                revenue_data = saleable_data_manual.copy()
                data_source = '手工'
                print(f"  ✅ 使用手工数据，记录数: {len(revenue_data)}")
            else:
                print(f"  ⚠️ 手工数据缺少'销售收益(元)'列")
        
        # 如果没有手工数据或未修改，使用系统数据
        if revenue_data is None and saleable_data is not None and not saleable_data.empty:
            print(f"  系统数据列: {list(saleable_data.columns)}")
            if '销售收益(元)' in saleable_data.columns:
                revenue_data = saleable_data.copy()
                data_source = '系统'
                print(f"  ✅ 使用系统数据，记录数: {len(revenue_data)}")
            else:
                print(f"  ⚠️ 系统数据缺少'销售收益(元)'列")
        
        if revenue_data is None or revenue_data.empty:
            print("  ❌ 没有可用的销售收益数据")
            categories = ['冰箱', '空调', '电脑', '电视', '洗衣机']
            return jsonify({
                'success': True,
                'data': {
                    'by_category': {cat: 0.0 for cat in categories},
                    'total': 0.0,
                    'data_source': None
                },
                'message': '暂无销售收益数据，请先完成数据处理流程'
            })
        
        # 确保数值类型
        if '销售收益(元)' in revenue_data.columns:
            revenue_data['销售收益(元)'] = pd.to_numeric(
                revenue_data['销售收益(元)'], errors='coerce'
            ).fillna(0)
        
        # 按分类汇总销售收益
        categories = ['冰箱', '空调', '电脑', '电视', '洗衣机']
        by_category = {cat: 0.0 for cat in categories}
        
        if '原物料名称' in revenue_data.columns:
            revenue_data['分类'] = revenue_data['原物料名称'].apply(map_category_py)
            for category in categories:
                mask = revenue_data['分类'] == category
                by_category[category] = float(revenue_data.loc[mask, '销售收益(元)'].sum())
        
        # 计算总计
        total = sum(by_category.values())
        
        print(f"  ✅ 按分类汇总完成 - 数据源: {data_source}, 总收益: {total:.2f} 元")
        print(f"  分类明细: {by_category}")
        
        return jsonify({
            'success': True,
            'data': {
                'by_category': by_category,
                'total': round(total, 2),
                'data_source': data_source
            }
        })
        
    except Exception as e:
        print(f"获取按分类汇总的销售收益数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/revenue-forecast/sales-revenue/export', methods=['GET'])
def export_sales_revenue():
    """导出销售收益数据为Excel"""
    try:
        print("📊 开始导出销售收益数据...")
        app_data = get_session_data_manager()
        
        # 优先使用手工数据，否则使用系统数据
        saleable_data_manual = app_data.get_data('saleable_data_manual')
        saleable_data = app_data.get_data('saleable_data')
        saleable_data_modified = app_data.get_data('saleable_data_modified')
        
        print(f"  手工数据: {saleable_data_manual is not None and not saleable_data_manual.empty if saleable_data_manual is not None else False}")
        print(f"  系统数据: {saleable_data is not None and not saleable_data.empty if saleable_data is not None else False}")
        print(f"  已修改: {saleable_data_modified}")
        
        revenue_data = None
        data_source = None
        # 检查是否有手工数据且包含收益列，并且已标记为修改（与导出结果逻辑一致）
        if saleable_data_manual is not None and not saleable_data_manual.empty and saleable_data_modified:
            print(f"  手工数据列: {list(saleable_data_manual.columns)}")
            if '销售收益(元)' in saleable_data_manual.columns:
                revenue_data = saleable_data_manual.copy()
                data_source = '手工'
                print(f"  ✅ 使用手工数据，记录数: {len(revenue_data)}")
            else:
                print(f"  ⚠️ 手工数据缺少'销售收益(元)'列")
        
        # 如果没有手工数据或未修改，使用系统数据
        if revenue_data is None and saleable_data is not None and not saleable_data.empty:
            print(f"  系统数据列: {list(saleable_data.columns)}")
            if '销售收益(元)' in saleable_data.columns:
                revenue_data = saleable_data.copy()
                data_source = '系统'
                print(f"  ✅ 使用系统数据，记录数: {len(revenue_data)}")
            else:
                print(f"  ⚠️ 系统数据缺少'销售收益(元)'列")
        
        if revenue_data is None or revenue_data.empty:
            print("  ❌ 没有可用的销售收益数据")
            return jsonify({
                'success': False,
                'error': '暂无销售收益数据'
            }), 400
        
        print(f"  原始数据列: {list(revenue_data.columns)}")
        
        # 排除指定三列
        columns_to_exclude = ['是否减扣', '减扣说明', '处置类别']
        filtered_columns = [col for col in revenue_data.columns if col not in columns_to_exclude]
        print(f"  过滤后列: {filtered_columns}")
        
        # 定义列顺序
        ordered_columns = [
            '序号',
            '原物料代码',
            '原物料名称',
            '拆解产物编码',
            '拆解产物名称',
            '拆解系数',
            '原物料重量(KG)',
            '计算结果(KG)',
            '类别',
            '期间',
            '销售单价(元/KG)',
            '销售单价-不含税(元/KG)',
            '销售收益(元)'
        ]
        
        # 按照指定顺序重新排列列（只包含实际存在的列）
        final_columns = [col for col in ordered_columns if col in filtered_columns]
        # 添加其他未在ordered_columns中的列（如果有）
        other_columns = [col for col in filtered_columns if col not in ordered_columns]
        final_columns.extend(other_columns)
        
        print(f"  最终列顺序: {final_columns}")
        
        if not final_columns:
            print("  ❌ 没有可导出的列")
            return jsonify({
                'success': False,
                'error': '没有可导出的列'
            }), 400
        
        revenue_data_filtered = revenue_data[final_columns].copy()
        print(f"  ✅ 准备导出 {len(revenue_data_filtered)} 条记录，{len(final_columns)} 列")
        
        # 重命名列名用于显示（不影响原始数据）
        revenue_data_filtered_renamed = revenue_data_filtered.rename(columns={'销售收益(元)': '销售收入（元）'})
        
        # 追加合计行：数值列求和（拆解系数/销售单价/期间/序号等除外）
        _sum_cols_sr = {'原物料重量(KG)', '计算结果(KG)', '销售收入（元）'}
        _dash_cols_sr = {'拆解系数', '销售单价(元/KG)', '销售单价-不含税(元/KG)', '期间'}
        _total_row_sr = {c: '' for c in revenue_data_filtered_renamed.columns}
        _label_placed_sr = False
        for _c in revenue_data_filtered_renamed.columns:
            if _c == '序号':
                _total_row_sr[_c] = '合计'
                _label_placed_sr = True
            elif _c in _sum_cols_sr:
                _s = pd.to_numeric(revenue_data_filtered_renamed[_c], errors='coerce').fillna(0).sum()
                _total_row_sr[_c] = round(float(_s), 2)
            elif _c in _dash_cols_sr:
                _total_row_sr[_c] = '-'
            else:
                if not _label_placed_sr:
                    _total_row_sr[_c] = '合计'
                    _label_placed_sr = True
                else:
                    _total_row_sr[_c] = ''
        revenue_data_filtered_renamed = pd.concat(
            [revenue_data_filtered_renamed, pd.DataFrame([_total_row_sr])],
            ignore_index=True
        )
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Sheet 1: 明细数据（导出原始数据，不受格式化限制）
            revenue_data_filtered_renamed.to_excel(writer, sheet_name='销售收入', index=False)
            worksheet = writer.sheets['销售收入']
            
            # 设置表头样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", name="仿宋")
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 设置列宽（使用重命名后的列名）
            for idx, col in enumerate(revenue_data_filtered_renamed.columns, start=1):
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(idx)
                col_width = max(len(str(col)), 15)
                worksheet.column_dimensions[col_letter].width = min(col_width, 50)
            
            # 合计行样式：浅灰底、加粗
            total_row_excel_idx = len(revenue_data_filtered_renamed) + 1
            total_fill_sr = PatternFill(start_color="E8E6DC", end_color="E8E6DC", fill_type="solid")
            total_font_sr = Font(bold=True, name="仿宋")
            for _col_i in range(1, len(revenue_data_filtered_renamed.columns) + 1):
                _cell = worksheet.cell(row=total_row_excel_idx, column=_col_i)
                _cell.fill = total_fill_sr
                _cell.font = total_font_sr
                _cell.alignment = Alignment(horizontal="center", vertical="center")
        
        output.seek(0)
        filename = f'销售收益_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        print(f"  ✅ Excel文件创建成功: {filename}")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"  ❌ 导出销售收益数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/revenue-forecast/subsidy-income/export', methods=['GET'])
def export_subsidy_income():
    """导出基金补贴收入数据为Excel"""
    try:
        app_data = get_session_data_manager()
        
        # 获取基金补贴收入数据
        subsidy_income_data = app_data.get_data('subsidy_income_data')
        
        if subsidy_income_data is None or subsidy_income_data.empty:
            return jsonify({
                'success': False,
                'error': '暂无基金补贴收入数据'
            }), 400
        
        # 确保数值类型
        if '基金补贴收入(元)' in subsidy_income_data.columns:
            subsidy_income_data['基金补贴收入(元)'] = pd.to_numeric(
                subsidy_income_data['基金补贴收入(元)'], errors='coerce'
            ).fillna(0)
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Sheet 1: 明细数据
            # 将列名重命名为与前端显示一致的名称
            export_df = subsidy_income_data.copy()
            
            # 将补贴大类映射到产品类型（五大类：电视机、冰箱、洗衣机、空调、电脑）
            def map_subsidy_category_to_product_type(subsidy_category):
                """将补贴大类映射到产品类型"""
                subsidy_category = str(subsidy_category).strip()
                # 空调的细分补贴大类
                if subsidy_category in ['整机', '内机', '外机']:
                    return '空调'
                # 电脑的细分补贴大类
                elif subsidy_category in ['笔记本', '显示器', '主机']:
                    return '电脑'
                # 其他直接返回（电视机、冰箱、洗衣机）
                elif subsidy_category in ['电视机', '冰箱', '洗衣机']:
                    return subsidy_category
                else:
                    # 默认返回原值（兼容旧数据）
                    return subsidy_category
            
            # 将补贴大类列映射为五大类（用于显示）
            if '补贴大类' in export_df.columns:
                export_df['补贴大类'] = export_df['补贴大类'].apply(map_subsidy_category_to_product_type)
            
            # 列名映射：将实际数据列名映射为前端显示的列名
            column_mapping = {}
            if '当期拆解量(台)' in export_df.columns:
                column_mapping['当期拆解量(台)'] = '本期实际投产数量'
            if '补贴单价(元/台)' in export_df.columns:
                column_mapping['补贴单价(元/台)'] = '基金补贴单价(元)'
            
            if column_mapping:
                export_df = export_df.rename(columns=column_mapping)
            
            # 追加合计行：本期实际投产数量 / 基金补贴收入(元) 求和；基金补贴单价(元) 等比率列显示 "-"
            _sum_cols_si = {'本期实际投产数量', '基金补贴收入(元)'}
            _dash_cols_si = {'基金补贴单价(元)', '补贴单价(元/台)'}
            _total_row_si = {c: '' for c in export_df.columns}
            _label_placed_si = False
            for _c in export_df.columns:
                if _c in _sum_cols_si:
                    _s = pd.to_numeric(export_df[_c], errors='coerce').fillna(0).sum()
                    _total_row_si[_c] = round(float(_s), 2)
                elif _c in _dash_cols_si:
                    _total_row_si[_c] = '-'
                else:
                    if not _label_placed_si:
                        _total_row_si[_c] = '合计'
                        _label_placed_si = True
                    else:
                        _total_row_si[_c] = ''
            export_df = pd.concat([export_df, pd.DataFrame([_total_row_si])], ignore_index=True)
            
            export_df.to_excel(writer, sheet_name='基金补贴收入明细', index=False)
            
            # 设置样式
            worksheet = writer.sheets['基金补贴收入明细']
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", name="仿宋")
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 合计行样式
            total_row_excel_idx_si = len(export_df) + 1
            total_fill_si = PatternFill(start_color="E8E6DC", end_color="E8E6DC", fill_type="solid")
            total_font_si = Font(bold=True, name="仿宋")
            for _col_i in range(1, len(export_df.columns) + 1):
                _cell = worksheet.cell(row=total_row_excel_idx_si, column=_col_i)
                _cell.fill = total_fill_si
                _cell.font = total_font_si
                _cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Sheet 2: 分类统计
            if '补贴大类' in subsidy_income_data.columns:
                # 创建副本并映射补贴大类为五大类
                stats_df = subsidy_income_data.copy()
                stats_df['补贴大类'] = stats_df['补贴大类'].apply(map_subsidy_category_to_product_type)
                
                category_groups = stats_df.groupby('补贴大类').agg({
                    '基金补贴收入(元)': ['sum', 'count']
                }).reset_index()
                category_groups.columns = ['补贴大类', '基金补贴收入合计(元)', '记录数']
                
                # 添加合计行
                total_row = pd.DataFrame([{
                    '补贴大类': '合计',
                    '基金补贴收入合计(元)': category_groups['基金补贴收入合计(元)'].sum(),
                    '记录数': category_groups['记录数'].sum()
                }])
                category_groups = pd.concat([category_groups, total_row], ignore_index=True)
                
                category_groups.to_excel(writer, sheet_name='分类统计', index=False)
                
                stats_worksheet = writer.sheets['分类统计']
                for cell in stats_worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        output.seek(0)
        
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"基金补贴收入_{timestamp}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"导出基金补贴收入数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def calculate_disassembly_product_output_value_data(app_data):
    """
    计算一次拆解产物产值数据（公共函数）
    
    Args:
        app_data: 应用数据管理器
        
    Returns:
        tuple: (success: bool, result_data: list, error_message: str)
        如果success为True，result_data包含计算结果列表；如果为False，error_message包含错误信息
    """
    try:
        # 检查数据是否已被清除
        data_cleared = app_data.get_data('__data_cleared__')
        if data_cleared:
            return (True, [], '数据已清除')
        
        # 获取原始数据（未减扣）
        disassembly_data = app_data.get_data('disassembly_data')
        
        if disassembly_data is None or disassembly_data.empty:
            return (True, [], '暂无原始数据（未减扣），请先完成数据处理')
        
        # 筛选类别为"拆解产物"的记录
        if '类别' not in disassembly_data.columns:
            return (False, [], '数据中缺少"类别"列')
        
        product_data = disassembly_data[disassembly_data['类别'] == '拆解产物'].copy()
        
        if product_data.empty:
            return (True, [], '暂无拆解产物数据')
        
        # 获取价格数据
        from data.base_data.price_data import load_price_data
        price_df = load_price_data()
        
        if price_df is None or price_df.empty:
            return (False, [], '暂无销售价格数据，请先在销售价格管理页面配置价格')
        
        # 创建价格映射（使用拆解产物编码作为键）
        price_mapping = {}
        for _, row in price_df.iterrows():
            code = str(row['拆解产物编码']).strip()
            price_no_tax = row.get('销售单价-不含税(元/KG)', 0)
            if pd.notna(price_no_tax):
                price_mapping[code] = float(price_no_tax)
        
        # 分类关键词映射规则
        # 电视：电视、彩电、CRT其它机壳破碎塑料、线路板边框破碎塑料、等离子、废旧玻璃电子枪、废旧金属荫罩压块铁、黑白
        # 电脑：电脑、显示器、笔记本、主机、废旧金属黑色金属-铁及其合金-电子枪
        # 冰箱：冰箱、冰柜
        # 空调：空调
        # 洗衣机：洗衣机、双缸
        category_keyword_mapping = {
            '电视': ['电视', '彩电', 'CRT其它机壳破碎塑料', '线路板边框破碎塑料', '等离子', '废旧玻璃电子枪', '废旧金属荫罩压块铁', '黑白'],
            '电脑': ['电脑', '显示器', '笔记本', '主机', '废旧金属黑色金属-铁及其合金-电子枪'],
            '冰箱': ['冰箱', '冰柜'],
            '空调': ['空调'],
            '洗衣机': ['洗衣机', '双缸']
        }
        
        # 处理数据：添加分类、匹配价格、计算产值
        result_data = []
        for idx, row in product_data.iterrows():
            # 获取原物料名称
            material_name = str(row.get('原物料名称', '')).strip()
            
            # 根据原物料名称进行模糊匹配分类
            category = None
            for cat, keywords in category_keyword_mapping.items():
                for keyword in keywords:
                    if keyword in material_name:
                        category = cat
                        break
                if category:
                    break
            
            # 如果没有匹配到分类，跳过该记录（不产生"其他"分类）
            if not category:
                continue
            
            # 获取拆解产物编码
            product_code = str(row.get('拆解产物编码', '')).strip()
            
            # 获取计算结果(KG)
            calculated_weight = row.get('计算结果(KG)', 0)
            try:
                calculated_weight = float(calculated_weight) if pd.notna(calculated_weight) else 0
            except (ValueError, TypeError):
                calculated_weight = 0
            
            # 匹配价格
            price_no_tax = price_mapping.get(product_code, 0)
            
            # 计算物料产值
            # 如果销售单价-不含税(元/KG) < 0，则物料产值 = 计算结果(KG) × 0
            if price_no_tax < 0:
                material_value = calculated_weight * 0
            else:
                material_value = calculated_weight * price_no_tax
            
            result_data.append({
                '原物料名称': material_name,
                '分类': category,
                '拆解产物编码': product_code,
                '拆解产物名称': str(row.get('拆解产物名称', '')).strip(),
                '计算结果(KG)': calculated_weight,
                '销售单价-不含税(元/KG)': price_no_tax,
                '物料产值（元）': material_value
            })
        
        return (True, result_data, None)
        
    except Exception as e:
        error_msg = f"计算一次拆解产物产值数据失败: {str(e)}"
        print(error_msg)
        traceback.print_exc()
        return (False, [], error_msg)


@data_management_bp.route('/revenue-forecast/disassembly-product-output-value', methods=['GET'])
def get_disassembly_product_output_value():
    """获取一次拆解产物产值数据"""
    try:
        app_data = get_session_data_manager()
        success, result_data, error_message = calculate_disassembly_product_output_value_data(app_data)
        
        if not success:
            return jsonify({
                'success': False,
                'error': error_message
            }), 500
        
        return jsonify({
            'success': True,
            'data': result_data,
            'message': error_message if error_message else None
        })
        
    except Exception as e:
        print(f"获取一次拆解产物产值数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def calculate_deep_processing_product_output_value_data(app_data):
    """计算深加工拆解产物产值数据（公共函数）"""
    try:
        # 检查数据是否已被清除
        data_cleared = app_data.get_data('__data_cleared__')
        if data_cleared:
            return (True, [], '数据已清除')

        # 获取深加工数据
        deep_processing_data = app_data.get_data('deep_processing_data')

        if deep_processing_data is None or deep_processing_data.empty:
            return (True, [], '暂无深加工数据，请先完成数据处理')

        # 筛选是否减扣 == '否'的记录（只统计非减扣的深加工产物）
        if '是否减扣' not in deep_processing_data.columns:
            return (False, [], '数据中缺少"是否减扣"列')

        # 筛选非减扣记录
        non_deducted_data = deep_processing_data[deep_processing_data['是否减扣'] == '否'].copy()

        if non_deducted_data.empty:
            return (True, [], '暂无非减扣的深加工产物数据')

        # 获取价格数据
        from data.base_data.price_data import load_price_data
        price_df = load_price_data()

        if price_df is None or price_df.empty:
            return (False, [], '暂无销售价格数据，请先在销售价格管理页面配置价格')

        # 创建价格映射（使用深加工产物编码作为键）
        price_mapping = {}
        for _, row in price_df.iterrows():
            code = str(row['拆解产物编码']).strip()
            price_no_tax = row.get('销售单价-不含税(元/KG)', 0)
            if pd.notna(price_no_tax):
                price_mapping[code] = float(price_no_tax)

        # 创建R3代码到类别的映射
        r3_to_category = {}
        for mapping_item in MAPPING_DATA:
            r3_code = str(mapping_item.get('R3系统代码', '')).strip()
            category = str(mapping_item.get('类别', '')).strip()
            if r3_code and category:
                r3_to_category[r3_code] = category

        # 分类关键词映射规则（四机一脑类别）
        category_keyword_mapping = {
            '电视': ['电视', '彩电', 'CRT其它机壳破碎塑料', '线路板边框破碎塑料', '等离子', '废旧玻璃电子枪', '废旧金属荫罩压块铁', '黑白'],
            '电脑': ['电脑', '显示器', '笔记本', '主机', '废旧金属黑色金属-铁及其合金-电子枪'],
            '冰箱': ['冰箱', '冰柜'],
            '空调': ['空调'],
            '洗衣机': ['洗衣机', '双缸']
        }

        # 处理数据
        result_data = []
        for _, row in non_deducted_data.iterrows():
            # 提取基础字段
            material_code = str(row.get('原物料代码', '')).strip()
            material_name = str(row.get('原物料名称', '')).strip()
            first_product_code = str(row.get('一次拆解产物编码', '')).strip()
            first_product_name = str(row.get('一次拆解产物名称', '')).strip()
            deep_product_code = str(row.get('深加工产物编码', '')).strip()
            deep_product_name = str(row.get('深加工产物名称', '')).strip()

            # 获取深加工结果(KG)
            deep_result_kg = row.get('深加工结果(KG)', 0)
            try:
                deep_result_kg = float(deep_result_kg) if pd.notna(deep_result_kg) else 0
            except (ValueError, TypeError):
                deep_result_kg = 0

            # 根据原物料名称进行模糊匹配分类（四机一脑类别）
            four_category = None
            for cat, keywords in category_keyword_mapping.items():
                for keyword in keywords:
                    if keyword in material_name:
                        four_category = cat
                        break
                if four_category:
                    break

            # 如果没有匹配到四机一脑类别，跳过该记录
            if not four_category:
                continue

            # 使用深加工产物编码匹配价格
            price_no_tax = price_mapping.get(deep_product_code, 0)

            # 使用深加工产物编码（作为R3代码）匹配映射表获取类别
            mapping_category = r3_to_category.get(deep_product_code, '')

            # 计算物料产值（元）= 深加工结果(KG) × 销售单价-不含税(元/KG)
            # 如果销售单价-不含税(元/KG) < 0，则物料产值 = 深加工结果(KG) × 0
            if price_no_tax < 0:
                material_value = deep_result_kg * 0
            else:
                material_value = deep_result_kg * price_no_tax

            result_data.append({
                '原物料代码': material_code,
                '原物料名称': material_name,
                '四机一脑类别': four_category,
                '类别': mapping_category,
                '一次拆解产物编码': first_product_code,
                '一次拆解产物名称': first_product_name,
                '深加工产物编码': deep_product_code,
                '深加工产物名称': deep_product_name,
                '深加工结果(KG)': deep_result_kg,
                '销售单价-不含税(元/KG)': price_no_tax,
                '物料产值（元）': material_value
            })

        return (True, result_data, None)

    except Exception as e:
        print(f"计算深加工拆解产物产值数据失败: {str(e)}")
        traceback.print_exc()
        return (False, [], str(e))


def calculate_deep_processing_product_output_value_without_stock_data(app_data):
    """计算深加工拆解产物产值（不考虑期初库存和库存结余）

    说明：
    - 始终使用"被减扣数据"只读视图 (_build_deducted_readonly_dataframe)，避免在用户手工编辑后
      误用"被减扣数据（手工）"sheet。
    - 按"拆解物原料成本"页类别为"旧机"且"本期实际投产数量(非限制使用的库存) > 0"的物料代码
      构造 valid_material_codes，过滤"原物料代码"不在清单中的行。
    """
    try:
        data_cleared = app_data.get_data('__data_cleared__')
        if data_cleared:
            return (True, [], '数据已清除')

        deducted_data = _build_deducted_readonly_dataframe(app_data)
        if deducted_data is None or deducted_data.empty:
            return (True, [], '暂无被减扣数据，请先完成数据处理')

        required_columns = ['类别', '处置类别', '原物料代码', '拆解产物编码', '计算结果(KG)']
        missing_columns = [col for col in required_columns if col not in deducted_data.columns]
        if missing_columns:
            return (False, [], f'被减扣数据缺少必要字段: {", ".join(missing_columns)}')

        target_disposal_types = ['内转屏处置', '内转印制板处置', '深加工-打包铁', '深加工-塑料一破']
        filtered_data = deducted_data[
            (deducted_data['类别'] == '拆解产物') &
            (deducted_data['处置类别'].isin(target_disposal_types))
        ].copy()

        if filtered_data.empty:
            return (True, [], '筛选后暂无符合条件的数据')

        # 构造有效物料代码集合：类别=旧机 且 本期实际投产数量(非限制使用的库存) > 0
        from app.api.cost_forecast_api import calculate_material_cost

        extracted_data = app_data.get_data('extracted_data_manual')
        if extracted_data is None or getattr(extracted_data, 'empty', True):
            return (True, [], '暂无提取数据，无法获取拆解数量')

        cost_data = calculate_material_cost(extracted_data)
        if cost_data is None or cost_data.empty:
            return (True, [], '暂无有效物料代码（拆解物原料成本为空）')

        valid_material_codes = set()
        if '类别' in cost_data.columns and '物料代码' in cost_data.columns \
                and '非限制使用的库存' in cost_data.columns:
            old_machine_data = cost_data[cost_data['类别'] == '旧机'].copy()
            old_machine_data['非限制使用的库存'] = pd.to_numeric(
                old_machine_data['非限制使用的库存'], errors='coerce'
            ).fillna(0)
            valid_data = old_machine_data[old_machine_data['非限制使用的库存'] > 0]
            for _, row in valid_data.iterrows():
                code = str(row.get('物料代码', '')).strip()
                if code:
                    valid_material_codes.add(code)

        if not valid_material_codes:
            return (True, [], '暂无有效物料代码（拆解数量均为 0）')

        deep_processing_df = dpd.get_deep_processing_dataframe()
        if deep_processing_df is None or deep_processing_df.empty:
            return (False, [], '暂无深加工数据管理配置，请先维护拆解系数')

        from data.base_data.price_data import load_price_data
        price_df = load_price_data()
        if price_df is None or price_df.empty:
            return (False, [], '暂无销售价格数据，请先在销售价格管理页面配置价格')

        price_mapping = {}
        for _, row in price_df.iterrows():
            code = str(row.get('拆解产物编码', '')).strip()
            price_no_tax = row.get('销售单价-不含税(元/KG)', 0)
            if code and pd.notna(price_no_tax):
                price_mapping[code] = float(price_no_tax)

        r3_to_category = {}
        for mapping_item in MAPPING_DATA:
            r3_code = str(mapping_item.get('R3系统代码', '')).strip()
            category = str(mapping_item.get('类别', '')).strip()
            if r3_code and category:
                r3_to_category[r3_code] = category

        category_keyword_mapping = {
            '电视': ['电视', '彩电', 'CRT其它机壳破碎塑料', '线路板边框破碎塑料', '等离子', '废旧玻璃电子枪', '废旧金属荫罩压块铁', '黑白'],
            '电脑': ['电脑', '显示器', '笔记本', '主机', '废旧金属黑色金属-铁及其合金-电子枪'],
            '冰箱': ['冰箱', '冰柜'],
            '空调': ['空调'],
            '洗衣机': ['洗衣机', '双缸']
        }

        aggregation = {}
        for _, row in filtered_data.iterrows():
            raw_material_code = str(row.get('原物料代码', '')).strip()
            # 按"拆解数量"过滤：原物料代码不在 valid_material_codes 中则跳过
            if raw_material_code not in valid_material_codes:
                continue
            raw_material_name = str(row.get('原物料名称', '')).strip()
            first_product_code = str(row.get('拆解产物编码', '')).strip()
            first_product_name = str(row.get('拆解产物名称', '')).strip()
            try:
                input_kg = float(row.get('计算结果(KG)', 0) or 0)
            except (ValueError, TypeError):
                input_kg = 0.0

            matching_records = deep_processing_df[deep_processing_df['拆解产物编码'].astype(str).str.strip() == first_product_code]
            if matching_records.empty or input_kg == 0:
                continue

            for _, deep_row in matching_records.iterrows():
                try:
                    coefficient = float(deep_row.get('深加工拆解系数', 0) or 0)
                except (ValueError, TypeError):
                    coefficient = 0.0
                if coefficient <= 0:
                    continue
                try:
                    input_output_ratio = float(deep_row.get('深加工投入产出比例', 1) or 1)
                except (ValueError, TypeError):
                    input_output_ratio = 1.0

                deep_product_code = str(deep_row.get('深加工产物编码', '')).strip()
                deep_product_name = str(deep_row.get('深加工产物名称', '')).strip()
                deep_result_kg = input_kg * coefficient * input_output_ratio
                price_no_tax = float(price_mapping.get(deep_product_code, 0) or 0)
                material_value = deep_result_kg * (0 if price_no_tax < 0 else price_no_tax)

                # 必须包含一次拆解产物编码：同一原物料下不同一次拆解产物可能对应同一深加工产物编码，
                # 仅用 (原物料, 深加工) 会错误合并并保留先写入的一次拆解编码，造成漏算/展示错位。
                key = (raw_material_code, first_product_code, deep_product_code)
                if key not in aggregation:
                    four_category = ''
                    for cat, keywords in category_keyword_mapping.items():
                        if any(keyword in raw_material_name for keyword in keywords):
                            four_category = cat
                            break
                    aggregation[key] = {
                        '原物料代码': raw_material_code,
                        '原物料名称': raw_material_name,
                        '四机一脑类别': four_category,
                        '类别': r3_to_category.get(deep_product_code, ''),
                        '一次拆解产物编码': first_product_code,
                        '一次拆解产物名称': first_product_name,
                        '深加工产物编码': deep_product_code,
                        '深加工产物名称': deep_product_name,
                        '深加工结果(KG)': 0.0,
                        '销售单价-不含税(元/KG)': price_no_tax,
                        '物料产值（元）': 0.0
                    }
                aggregation[key]['深加工结果(KG)'] += deep_result_kg
                aggregation[key]['物料产值（元）'] += material_value

        result_data = sorted(aggregation.values(), key=lambda item: (
            str(item.get('原物料代码', '')),
            str(item.get('一次拆解产物编码', '')),
            str(item.get('深加工产物编码', ''))
        ))
        return (True, result_data, None)

    except Exception as e:
        print(f"计算深加工拆解产物产值（不考虑期初库存和库存结余）失败: {str(e)}")
        traceback.print_exc()
        return (False, [], str(e))


@data_management_bp.route('/revenue-forecast/deep-processing-product-output-value', methods=['GET'])
def get_deep_processing_product_output_value():
    """获取深加工拆解产物产值数据"""
    app_data = get_session_data_manager()
    success, result_data, message = calculate_deep_processing_product_output_value_data(app_data)
    without_stock_success, without_stock_data, without_stock_message = calculate_deep_processing_product_output_value_without_stock_data(app_data)

    if not success:
        return jsonify({
            'success': False,
            'error': message
        })
    if not without_stock_success:
        return jsonify({
            'success': False,
            'error': without_stock_message
        })

    return jsonify({
        'success': True,
        'data': result_data,
        'without_stock_data': without_stock_data,
        'message': message,
        'without_stock_message': without_stock_message
    })


@data_management_bp.route('/revenue-forecast/disassembly-product-output-value/export', methods=['GET'])
def export_disassembly_product_output_value():
    """导出一拆解产物产值数据为Excel"""
    try:
        app_data = get_session_data_manager()
        
        # 获取原始数据（未减扣）
        disassembly_data = app_data.get_data('disassembly_data')
        
        if disassembly_data is None or disassembly_data.empty:
            return jsonify({
                'success': False,
                'error': '暂无原始数据（未减扣），请先完成数据处理'
            }), 400
        
        # 筛选类别为"拆解产物"的记录
        if '类别' not in disassembly_data.columns:
            return jsonify({
                'success': False,
                'error': '数据中缺少"类别"列'
            }), 400
        
        product_data = disassembly_data[disassembly_data['类别'] == '拆解产物'].copy()
        
        if product_data.empty:
            return jsonify({
                'success': False,
                'error': '暂无拆解产物数据'
            }), 400
        
        # 获取价格数据
        from data.base_data.price_data import load_price_data
        price_df = load_price_data()
        
        if price_df is None or price_df.empty:
            return jsonify({
                'success': False,
                'error': '暂无销售价格数据，请先在销售价格管理页面配置价格'
            }), 400
        
        # 创建价格映射
        price_mapping = {}
        for _, row in price_df.iterrows():
            code = str(row['拆解产物编码']).strip()
            price_no_tax = row.get('销售单价-不含税(元/KG)', 0)
            if pd.notna(price_no_tax):
                price_mapping[code] = float(price_no_tax)
        
        # 分类关键词映射规则
        # 电视：电视、彩电、CRT其它机壳破碎塑料、线路板边框破碎塑料、等离子、废旧玻璃电子枪、废旧金属荫罩压块铁、黑白
        # 电脑：电脑、显示器、笔记本、主机、废旧金属黑色金属-铁及其合金-电子枪
        # 冰箱：冰箱、冰柜
        # 空调：空调
        # 洗衣机：洗衣机、双缸
        category_keyword_mapping = {
            '电视': ['电视', '彩电', 'CRT其它机壳破碎塑料', '线路板边框破碎塑料', '等离子', '废旧玻璃电子枪', '废旧金属荫罩压块铁', '黑白'],
            '电脑': ['电脑', '显示器', '笔记本', '主机', '废旧金属黑色金属-铁及其合金-电子枪'],
            '冰箱': ['冰箱', '冰柜'],
            '空调': ['空调'],
            '洗衣机': ['洗衣机', '双缸']
        }
        
        # 处理数据
        result_rows = []
        for idx, row in product_data.iterrows():
            material_name = str(row.get('原物料名称', '')).strip()
            
            # 分类匹配
            category = None
            for cat, keywords in category_keyword_mapping.items():
                for keyword in keywords:
                    if keyword in material_name:
                        category = cat
                        break
                if category:
                    break
            
            # 如果没有匹配到分类，跳过该记录（不产生"其他"分类）
            if not category:
                continue
            
            product_code = str(row.get('拆解产物编码', '')).strip()
            calculated_weight = row.get('计算结果(KG)', 0)
            try:
                calculated_weight = float(calculated_weight) if pd.notna(calculated_weight) else 0
            except (ValueError, TypeError):
                calculated_weight = 0
            
            price_no_tax = price_mapping.get(product_code, 0)
            
            # 计算物料产值
            if price_no_tax < 0:
                material_value = calculated_weight * 0
            else:
                material_value = calculated_weight * price_no_tax
            
            result_rows.append({
                '序号': len(result_rows) + 1,
                '原物料名称': material_name,
                '分类': category,
                '拆解产物编码': product_code,
                '拆解产物名称': str(row.get('拆解产物名称', '')).strip(),
                '计算结果(KG)': calculated_weight,
                '销售单价-不含税(元/KG)': price_no_tax,
                '物料产值（元）': material_value
            })
        
        # 创建详细数据DataFrame
        detail_df = pd.DataFrame(result_rows)
        
        # 创建分类汇总DataFrame
        category_summary = {}
        for row in result_rows:
            category = row['分类']
            if category not in category_summary:
                category_summary[category] = {
                    '分类': category,
                    '记录数': 0,
                    '总重量(KG)': 0,
                    '总产值（元）': 0
                }
            category_summary[category]['记录数'] += 1
            category_summary[category]['总重量(KG)'] += row['计算结果(KG)']
            category_summary[category]['总产值（元）'] += row['物料产值（元）']
        
        summary_rows = []
        for category, data in sorted(category_summary.items()):
            summary_rows.append({
                '分类': data['分类'],
                '记录数': data['记录数'],
                '总重量(KG)': round(data['总重量(KG)'], 6),
                '总产值（元）': round(data['总产值（元）'], 2)
            })
        
        # 添加总计行
        total_count = len(result_rows)
        total_weight = sum(row['计算结果(KG)'] for row in result_rows)
        total_value = sum(row['物料产值（元）'] for row in result_rows)
        summary_rows.append({
            '分类': '总计',
            '记录数': total_count,
            '总重量(KG)': round(total_weight, 6),
            '总产值（元）': round(total_value, 2)
        })
        
        summary_df = pd.DataFrame(summary_rows)
        
        # 创建Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 详细数据
            detail_df.to_excel(writer, sheet_name='详细数据', index=False)
            
            # 分类汇总
            summary_df.to_excel(writer, sheet_name='分类汇总', index=False)
            
            # 设置列宽和样式
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # 详细数据工作表
            detail_worksheet = writer.sheets['详细数据']
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", name="仿宋")
            
            for col_idx, col_name in enumerate(detail_df.columns, start=1):
                col_letter = get_column_letter(col_idx)
                cell = detail_worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # 设置列宽
                col_width = max(len(str(col_name)), 15)
                detail_worksheet.column_dimensions[col_letter].width = min(col_width, 30)
            
            # 分类汇总工作表
            summary_worksheet = writer.sheets['分类汇总']
            for col_idx, col_name in enumerate(summary_df.columns, start=1):
                col_letter = get_column_letter(col_idx)
                cell = summary_worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # 设置列宽
                col_width = max(len(str(col_name)), 15)
                summary_worksheet.column_dimensions[col_letter].width = min(col_width, 20)
            
            # 总计行加粗
            if len(summary_rows) > 0:
                total_row = len(summary_rows)
                for col_idx in range(1, len(summary_df.columns) + 1):
                    cell = summary_worksheet.cell(row=total_row + 1, column=col_idx)
                    cell.font = Font(bold=True, name="仿宋")
        
        output.seek(0)
        filename = f'一次拆解产物产值_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"导出一拆解产物产值数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@data_management_bp.route('/revenue-forecast/deep-processing-product-output-value/export', methods=['GET'])
def export_deep_processing_product_output_value():
    """导出深加工拆解产物产值数据为Excel"""
    try:
        app_data = get_session_data_manager()

        success, with_stock_rows, with_stock_message = calculate_deep_processing_product_output_value_data(app_data)
        if not success:
            return jsonify({'success': False, 'error': with_stock_message}), 400
        without_stock_success, without_stock_rows, without_stock_message = calculate_deep_processing_product_output_value_without_stock_data(app_data)
        if not without_stock_success:
            return jsonify({'success': False, 'error': without_stock_message}), 400

        if not with_stock_rows and not without_stock_rows:
            return jsonify({
                'success': False,
                'error': '没有符合条件的数据可导出'
            }), 400

        with_stock_df = pd.DataFrame(with_stock_rows)
        without_stock_df = pd.DataFrame(without_stock_rows)

        if with_stock_df.empty:
            with_stock_df = pd.DataFrame(columns=[
                '原物料代码', '原物料名称', '四机一脑类别', '类别',
                '一次拆解产物编码', '一次拆解产物名称', '深加工产物编码', '深加工产物名称',
                '深加工结果(KG)', '销售单价-不含税(元/KG)', '物料产值（元）'
            ])
        if without_stock_df.empty:
            without_stock_df = pd.DataFrame(columns=with_stock_df.columns.tolist())

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            column_widths = {
                '原物料代码': 18, '原物料名称': 30, '四机一脑类别': 15, '类别': 15,
                '一次拆解产物编码': 18, '一次拆解产物名称': 30, '深加工产物编码': 18,
                '深加工产物名称': 30, '深加工结果(KG)': 18, '销售单价-不含税(元/KG)': 22, '物料产值（元）': 18
            }

            def _append_total_row_dpo(df):
                """在深加工拆解产物产值DataFrame末尾追加合计行"""
                if df is None or df.empty:
                    return df
                _sum_cols = {'深加工结果(KG)', '物料产值（元）'}
                _dash_cols = {'销售单价-不含税(元/KG)'}
                _total_row = {c: '' for c in df.columns}
                _label_placed = False
                for _c in df.columns:
                    if _c in _sum_cols:
                        _s = pd.to_numeric(df[_c], errors='coerce').fillna(0).sum()
                        _total_row[_c] = round(float(_s), 6 if _c == '深加工结果(KG)' else 2)
                    elif _c in _dash_cols:
                        _total_row[_c] = '-'
                    else:
                        if not _label_placed:
                            _total_row[_c] = '合计'
                            _label_placed = True
                        else:
                            _total_row[_c] = ''
                return pd.concat([df, pd.DataFrame([_total_row])], ignore_index=True)
            
            def write_detail_sheet(sheet_name, df):
                df = _append_total_row_dpo(df)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                header_font = Font(bold=True, color="FFFFFF", name="仿宋")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                center_alignment = Alignment(horizontal="center", vertical="center")
                data_font = Font(name="仿宋")
                total_fill = PatternFill(start_color="E8E6DC", end_color="E8E6DC", fill_type="solid")
                total_font = Font(name="仿宋", bold=True)
                total_row_idx = len(df) + 1

                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    col_letter = get_column_letter(col)
                    col_name = df.columns[col - 1]
                    worksheet.column_dimensions[col_letter].width = column_widths.get(col_name, 18)

                for row in range(2, len(df) + 2):
                    is_total_row = (row == total_row_idx)
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        col_name = df.columns[col - 1]
                        cell.font = data_font
                        cell.alignment = center_alignment
                        if col_name == '物料产值（元）':
                            cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                            cell.font = Font(name="仿宋", bold=True)
                        if is_total_row:
                            cell.fill = total_fill
                            cell.font = total_font

            write_detail_sheet('深加工拆解产物产值(考虑库存)', with_stock_df)
            write_detail_sheet('深加工拆解产物产值(不考虑库存)', without_stock_df)
        
        output.seek(0)
        
        # 生成文件名
        from datetime import datetime
        filename = f'深加工拆解产物产值_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"导出深加工拆解产物产值数据失败: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500



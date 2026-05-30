"""
成本预测相关API
"""
from flask import Blueprint, jsonify, request, session, send_file
import pandas as pd
import io
from datetime import datetime
from app.models.compatibility import AppDataManagerAdapter
from app.utils.auth_utils import login_required, require_can_edit
from data.base_data.labor_cost_data import get_labor_cost_dataframe
from app.api.data_management_api import get_salary_accounting_dataframe, get_period_cost_dataframe, get_tax_surcharge_dataframe

def get_session_data_manager():
    """获取会话数据管理器的便利函数"""
    from flask import session
    return AppDataManagerAdapter.get_instance(session.get('session_id'))

cost_forecast_bp = Blueprint('cost_forecast', __name__, url_prefix='/api/cost-forecast')


def _safe_str_for_maps(value):
    """将可能为 NaN/None 的值安全转为字符串（计件/间接人工映射共用）。"""
    if value is None:
        return ''
    try:
        if pd.isna(value):
            return ''
    except Exception:
        pass
    s = str(value).strip()
    if s.lower() == 'nan':
        return ''
    return s


DEEP_PART3_CATEGORIES = ('一破', '打包铁', '屏')


def _ingest_deep_output_value_items(items, kg_map, meta_map):
    """将深加工拆解产物产值明细写入 KG 映射（含 0 值）及展示元数据映射。"""
    if not items:
        return
    for item in items:
        category = item.get('类别', '')
        if category not in DEEP_PART3_CATEGORIES:
            continue
        key = (
            _safe_str_for_maps(item.get('原物料代码', '')),
            _safe_str_for_maps(item.get('一次拆解产物编码', '')),
            _safe_str_for_maps(item.get('深加工产物编码', ''))
        )
        kg = float(item.get('深加工结果(KG)', 0) or 0)
        kg_map[key] = kg_map.get(key, 0.0) + kg
        meta_map[key] = item


def _build_no_opening_deep_result_kg_map(app_data):
    """
    不考虑期初库存和库存结余口径下，深加工结果(KG) 按
    (原物料代码, 一次拆解/拆解产物编码, 深加工产物编码) 聚合。
    与计件工资、深加工拆解产物价值（不考虑期初）同源。

    口径要求（与 calculate_deep_processing_product_output_value_without_stock_data 保持一致）：
    - 被减扣数据始终使用"被减扣数据"只读视图，避免用户手工编辑后取到"被减扣数据（手工）"。
    - 仅统计"原物料代码"在 拆解物原料成本页(类别=旧机 且 本期实际投产数量>0) 的清单内的行。
    """
    result = {}
    try:
        from data.base_data.deep_processing_data import DEEP_PROCESSING_DATA
        # 局部 import 避免与 data_management_api 的模块级循环依赖
        from app.api.data_management_api import _build_deducted_readonly_dataframe

        deducted_data = _build_deducted_readonly_dataframe(app_data)
        if deducted_data is None or deducted_data.empty:
            return result
        required_cols = ['类别', '处置类别', '原物料代码', '拆解产物编码', '计算结果(KG)']
        if not all(col in deducted_data.columns for col in required_cols):
            return result
        deep_deducted_disposal = ['内转屏处置', '内转印制板处置', '深加工-打包铁', '深加工-塑料一破']
        filtered = deducted_data[
            (deducted_data['类别'] == '拆解产物')
            & (deducted_data['处置类别'].astype(str).str.strip().isin(deep_deducted_disposal))
        ].copy()
        filtered['计算结果(KG)'] = pd.to_numeric(filtered['计算结果(KG)'], errors='coerce').fillna(0)

        # 构造 valid_material_codes：类别=旧机 且 本期实际投产数量(非限制使用的库存) > 0
        valid_material_codes = set()
        extracted_data = app_data.get_data('extracted_data_manual')
        if extracted_data is None or getattr(extracted_data, 'empty', True):
            return result
        cost_data = calculate_material_cost(extracted_data)
        if cost_data is None or cost_data.empty:
            return result
        if '类别' in cost_data.columns and '物料代码' in cost_data.columns \
                and '非限制使用的库存' in cost_data.columns:
            old_machine_data = cost_data[cost_data['类别'] == '旧机'].copy()
            old_machine_data['非限制使用的库存'] = pd.to_numeric(
                old_machine_data['非限制使用的库存'], errors='coerce'
            ).fillna(0)
            for _, mrow in old_machine_data[old_machine_data['非限制使用的库存'] > 0].iterrows():
                code = _safe_str_for_maps(mrow.get('物料代码', ''))
                if code:
                    valid_material_codes.add(code)
        if not valid_material_codes:
            return result

        coeff_by_first = {}
        for coeff in DEEP_PROCESSING_DATA:
            first_code = _safe_str_for_maps(coeff.get('拆解产物编码', ''))
            if first_code:
                coeff_by_first.setdefault(first_code, []).append(coeff)

        for _, row in filtered.iterrows():
            origin_code = _safe_str_for_maps(row.get('原物料代码', ''))
            # 按"拆解数量"过滤：原物料代码不在 valid_material_codes 中则跳过
            if origin_code not in valid_material_codes:
                continue
            first_product_code = _safe_str_for_maps(row.get('拆解产物编码', ''))
            input_kg = float(row.get('计算结果(KG)', 0) or 0)
            if input_kg <= 0 or not first_product_code:
                continue

            for coeff in coeff_by_first.get(first_product_code, []):
                deep_product_code = _safe_str_for_maps(coeff.get('深加工产物编码', ''))
                if not deep_product_code:
                    continue
                try:
                    io_ratio = float(coeff.get('深加工投入产出比例', 1) or 1)
                except (ValueError, TypeError):
                    io_ratio = 1.0
                try:
                    deep_coeff = float(coeff.get('深加工拆解系数', 0) or 0)
                except (ValueError, TypeError):
                    deep_coeff = 0.0
                if deep_coeff <= 0:
                    continue

                deep_result_kg = input_kg * io_ratio * deep_coeff
                if deep_result_kg <= 0:
                    continue
                map_key = (origin_code, first_product_code, deep_product_code)
                result[map_key] = result.get(map_key, 0.0) + deep_result_kg
    except Exception:
        return {}
    return result


def calculate_material_cost(manual_data):
    """
    计算拆解物原料成本
    
    计算规则：
    - 如果编辑过数据（本期计划采购数量 > 0）:
      单位投料成本 = (价值 + 本期计划采购数量 × 计划采购单价) ÷ (初始数据 + 本期计划采购数量)
      拆解物原料成本 = 本期实际投产数量 × 单位投料成本
    - 如果没有编辑过数据（本期计划采购数量 = 0）:
      单位投料成本 = 提取结果（单价列）
      拆解物原料成本 = 本期实际投产数量 × 单位投料成本
    
    注意：本期实际投产数量对应"非限制使用的库存"字段
    
    Args:
        manual_data: DataFrame，包含提取结果手工数据
        
    Returns:
        DataFrame，包含计算后的成本数据
    """
    if manual_data is None or manual_data.empty:
        return pd.DataFrame()
    
    # 创建结果DataFrame的副本
    result_df = manual_data.copy()
    
    # 确保必需的列存在
    required_cols = ['价值', '单价', '初始数据', '本期计划采购数量', '计划采购单价', '非限制使用的库存']
    for col in required_cols:
        if col not in result_df.columns:
            if col in ['价值', '单价']:
                # 如果缺少价值或单价，填充为0
                result_df[col] = 0.0
            else:
                # 其他列填充为0
                result_df[col] = 0.0
    
    # 确保数值类型
    numeric_cols = ['价值', '单价', '初始数据', '本期计划采购数量', '计划采购单价', '非限制使用的库存']
    for col in numeric_cols:
        result_df[col] = pd.to_numeric(result_df[col], errors='coerce').fillna(0)
    
    # 初始化计算列
    result_df['单位投料成本'] = 0.0
    result_df['拆解物原料成本'] = 0.0
    
    # 只处理旧机类别
    if '类别' in result_df.columns:
        mask_old_machine = result_df['类别'] == '旧机'
        
        # 判断是否编辑过数据
        # 如果"本期计划采购数量" > 0，说明用户编辑过采购数量，需要使用加权平均计算单位投料成本
        # 如果"本期计划采购数量" = 0 且"计划采购单价"也被编辑过（不等于单价），也认为是编辑过
        # 但为了简化判断，我们主要看"本期计划采购数量"是否 > 0
        # 因为如果用户编辑了"计划采购单价"，通常也会编辑"本期计划采购数量"
        has_edited = result_df.loc[mask_old_machine, '本期计划采购数量'] > 0
        
        # 编辑过数据的情况
        edited_mask = mask_old_machine & has_edited
        if edited_mask.any():
            # 计算单位投料成本 = (价值 + 本期计划采购数量 × 计划采购单价) ÷ (初始数据 + 本期计划采购数量)
            numerator = result_df.loc[edited_mask, '价值'] + \
                       (result_df.loc[edited_mask, '本期计划采购数量'] * 
                        result_df.loc[edited_mask, '计划采购单价'])
            denominator = result_df.loc[edited_mask, '初始数据'] + \
                         result_df.loc[edited_mask, '本期计划采购数量']
            
            # 避免除零
            denominator = denominator.replace(0, 1)
            
            result_df.loc[edited_mask, '单位投料成本'] = numerator / denominator
            
            # 计算拆解物原料成本 = 本期实际投产数量 × 单位投料成本
            # 本期实际投产数量对应"非限制使用的库存"字段
            result_df.loc[edited_mask, '拆解物原料成本'] = \
                result_df.loc[edited_mask, '非限制使用的库存'] * \
                result_df.loc[edited_mask, '单位投料成本']
        
        # 没有编辑过数据的情况
        not_edited_mask = mask_old_machine & ~has_edited
        if not_edited_mask.any():
            # 单位投料成本 = 提取结果（单价列）
            result_df.loc[not_edited_mask, '单位投料成本'] = \
                result_df.loc[not_edited_mask, '单价']
            
            # 拆解物原料成本 = 本期实际投产数量 × 单位投料成本
            # 本期实际投产数量对应"非限制使用的库存"字段
            result_df.loc[not_edited_mask, '拆解物原料成本'] = \
                result_df.loc[not_edited_mask, '非限制使用的库存'] * \
                result_df.loc[not_edited_mask, '单位投料成本']
        
        print(f"[成本计算] 已编辑数据: {edited_mask.sum()} 条")
        print(f"[成本计算] 未编辑数据: {not_edited_mask.sum()} 条")
    
    return result_df


def calculate_piece_rate_wage(app_data):
    """
    计算生产工人计件工资
    
    计算公式：
    1. 提取结果数据编辑表中"本期实际投产数量"（旧机类别） × 计件人工标准管理表中旧机类别"生产计件单价"
    2. + 原始数据（未扣减）表中拆解产物"计算结果" × 计件人工标准管理表中一次拆解产物类别"生产计件单价"
    3. + 深加工数据表中"深加工结果"（一破、打包铁、屏类别） × 计件人工标准管理表中对应类别"生产计件单价"
    
    Args:
        app_data: AppDataManager实例
        
    Returns:
        dict: 包含总工资、各部分明细的字典
    """
    try:
        def _safe_str(value):
            """将可能为 NaN/None 的值安全转为字符串。"""
            if value is None:
                return ''
            try:
                if pd.isna(value):
                    return ''
            except Exception:
                pass
            s = str(value).strip()
            if s.lower() == 'nan':
                return ''
            return s

        # 构建“原物料代码 -> 原物料名称”的回填映射（尽量从旧机/提取结果里拿到名称）
        # 目的：当原始数据/深加工数据缺失“原物料名称”时，页面新增列不至于为空。
        origin_material_name_map = {}
        try:
            extracted_for_origin = app_data.get_data('extracted_data_manual')
            if extracted_for_origin is None or extracted_for_origin.empty:
                extracted_for_origin = app_data.get_data('extracted_data')

            if extracted_for_origin is not None and not extracted_for_origin.empty:
                # 物料代码列兼容
                origin_code_col = None
                for col in ['物料代码', 'R3系统代码', '代码']:
                    if col in extracted_for_origin.columns:
                        origin_code_col = col
                        break

                # 名称列兼容（优先“物料描述”，其次“物料名称”）
                origin_name_col = None
                for col in ['物料描述', '物料名称', '原物料名称']:
                    if col in extracted_for_origin.columns:
                        origin_name_col = col
                        break

                if origin_code_col and origin_name_col:
                    for _, r in extracted_for_origin[[origin_code_col, origin_name_col]].dropna().iterrows():
                        code = str(r.get(origin_code_col, '')).strip()
                        if code.endswith('.0'):
                            code = code[:-2]
                        name = str(r.get(origin_name_col, '')).strip()
                        if code and name and code not in origin_material_name_map:
                            origin_material_name_map[code] = name
        except Exception:
            # 回填映射失败不影响主流程
            pass

        # 获取计件人工标准数据
        labor_cost_df = get_labor_cost_dataframe()
        if labor_cost_df is None or labor_cost_df.empty:
            return {
                'total_wage': 0.0,
                'total_wage_no_opening': 0.0,
                'part1_wage': 0.0,
                'part2_wage': 0.0,
                'part3_wage': 0.0,
                'part1_wage_no_opening': 0.0,
                'part2_wage_no_opening': 0.0,
                'part3_wage_no_opening': 0.0,
                'part1_details': [],
                'part2_details': [],
                'part3_details': [],
                'error': '无法获取计件人工标准数据'
            }
        
        # 创建物料代码到单价的映射字典（按类别）
        labor_cost_dict = {}
        for _, row in labor_cost_df.iterrows():
            code = str(row['R3系统代码'])
            category = str(row['类别'])
            price = float(row['生产计件单价']) if pd.notna(row['生产计件单价']) else 0.0
            key = (code, category)
            labor_cost_dict[key] = price
        
        total_wage = 0.0
        total_wage_no_opening = 0.0
        part1_wage = 0.0  # 旧机类别
        part2_wage = 0.0  # 一次拆解产物类别
        part3_wage = 0.0  # 一破、打包铁、屏类别
        part1_wage_no_opening = 0.0
        part2_wage_no_opening = 0.0
        part3_wage_no_opening = 0.0
        part1_details = []
        part2_details = []
        part3_details = []

        # 使用与深加工拆解产物产值页完全一致的数据源，避免两页数据不一致
        from app.api.data_management_api import (
            calculate_deep_processing_product_output_value_without_stock_data,
            calculate_deep_processing_product_output_value_data,
        )
        _output_value_ok, _output_value_data, _msg = calculate_deep_processing_product_output_value_without_stock_data(app_data)
        no_opening_deep_result_map = {}
        no_opening_meta_map = {}
        if _output_value_ok and _output_value_data:
            _ingest_deep_output_value_items(
                _output_value_data, no_opening_deep_result_map, no_opening_meta_map
            )

        # 考虑期初库存口径的深加工结果KG映射（用于数量/重量列）
        _with_stock_ok, _with_stock_data, _ = calculate_deep_processing_product_output_value_data(app_data)
        with_stock_deep_result_map = {}
        with_stock_meta_map = {}
        if _with_stock_ok and _with_stock_data:
            _ingest_deep_output_value_items(
                _with_stock_data, with_stock_deep_result_map, with_stock_meta_map
            )

        # 第一部分：提取结果数据编辑表（旧机类别）
        extracted_data = app_data.get_data('extracted_data_manual')
        if extracted_data is not None and not extracted_data.empty:
            if '类别' in extracted_data.columns and '非限制使用的库存' in extracted_data.columns:
                old_machine_data = extracted_data[extracted_data['类别'] == '旧机'].copy()
                if not old_machine_data.empty:
                    # 确保数值类型
                    old_machine_data['非限制使用的库存'] = pd.to_numeric(
                        old_machine_data['非限制使用的库存'], errors='coerce'
                    ).fillna(0)
                    
                    # 获取物料代码字段（可能是'物料代码'或其他名称）
                    material_code_col = None
                    for col in ['物料代码', 'R3系统代码', '代码']:
                        if col in old_machine_data.columns:
                            material_code_col = col
                            break
                    
                    if material_code_col:
                        for _, row in old_machine_data.iterrows():
                            material_code = _safe_str(row.get(material_code_col, ''))
                            if material_code.endswith('.0'):
                                material_code = material_code[:-2]
                            quantity = float(row['非限制使用的库存'])
                            
                            if quantity > 0:
                                # 查找旧机类别的单价
                                key = (material_code, '旧机')
                                unit_price = labor_cost_dict.get(key, 0.0)
                                
                                if unit_price > 0:
                                    wage = quantity * unit_price
                                    part1_wage += wage
                                    material_name = _safe_str(row.get('物料描述', ''))
                                    part1_details.append({
                                        '物料代码': material_code,
                                        '物料名称': material_name,
                                        # 旧机类别：原物料即本身
                                        '原物料代码': material_code,
                                        '原物料名称': material_name,
                                        '数量': float(quantity),
                                        '数量(不考虑期初库存和库存结余)': float(quantity),
                                        '单价': float(unit_price),
                                        '工资': float(wage),
                                        '工资(不考虑期初库存和库存结余)': float(wage)
                                    })
                                else:
                                    print(f"[计件工资] 未找到物料代码 {material_code} 的旧机类别单价")
        
        # 第二部分：原始数据（未扣减）表（拆解产物类别）
        disassembly_data = app_data.get_data('disassembly_data')
        if disassembly_data is not None and not disassembly_data.empty:
            if '类别' in disassembly_data.columns and '计算结果(KG)' in disassembly_data.columns:
                disassembly_product_data = disassembly_data[disassembly_data['类别'] == '拆解产物'].copy()
                if not disassembly_product_data.empty:
                    # 确保数值类型
                    disassembly_product_data['计算结果(KG)'] = pd.to_numeric(
                        disassembly_product_data['计算结果(KG)'], errors='coerce'
                    ).fillna(0)
                    
                    if '拆解产物编码' in disassembly_product_data.columns:
                        for _, row in disassembly_product_data.iterrows():
                            product_code = _safe_str(row.get('拆解产物编码', ''))
                            # 处理可能的.0后缀
                            if product_code.endswith('.0'):
                                product_code = product_code[:-2]
                            result_kg = float(row['计算结果(KG)'])
                            
                            if result_kg > 0:
                                # 查找一次拆解产物类别的单价
                                key = (product_code, '一次拆解产物')
                                unit_price = labor_cost_dict.get(key, 0.0)
                                
                                if unit_price > 0:
                                    wage = result_kg * unit_price
                                    part2_wage += wage
                                    origin_code = _safe_str(row.get('原物料代码', ''))
                                    if origin_code.endswith('.0'):
                                        origin_code = origin_code[:-2]
                                    origin_name = _safe_str(row.get('原物料名称', ''))
                                    if not origin_name and origin_code:
                                        origin_name = origin_material_name_map.get(origin_code, '')
                                    part2_details.append({
                                        '拆解产物编码': product_code,
                                        '拆解产物名称': row.get('拆解产物名称', ''),
                                        '原物料代码': origin_code,
                                        '原物料名称': origin_name,
                                        '计算结果(KG)': float(result_kg),
                                        '数量(不考虑期初库存和库存结余)': float(result_kg),
                                        '单价': float(unit_price),
                                        '工资': float(wage),
                                        '工资(不考虑期初库存和库存结余)': float(wage)
                                    })
                                else:
                                    print(f"[计件工资] 未找到拆解产物编码 {product_code} 的一次拆解产物类别单价")
        
        # 第三部分：深加工（一破、打包铁、屏），行集 = 两口径产值 sheet 的并集
        all_part3_keys = set(with_stock_deep_result_map.keys()) | set(no_opening_deep_result_map.keys())
        for result_key in sorted(all_part3_keys):
            origin_code, first_product_code, deep_product_code = result_key
            with_stock_kg = float(with_stock_deep_result_map.get(result_key, 0.0) or 0.0)
            no_opening_kg = float(no_opening_deep_result_map.get(result_key, 0.0) or 0.0)

            item = with_stock_meta_map.get(result_key) or no_opening_meta_map.get(result_key) or {}
            category = item.get('类别', '')
            if category not in DEEP_PART3_CATEGORIES:
                continue

            unit_price = labor_cost_dict.get((deep_product_code, category), 0.0)
            wage = with_stock_kg * unit_price
            wage_no_opening = no_opening_kg * unit_price
            part3_wage += wage
            part3_wage_no_opening += wage_no_opening

            origin_name = _safe_str_for_maps(item.get('原物料名称', ''))
            if not origin_name and origin_code:
                origin_name = origin_material_name_map.get(origin_code, '')

            part3_details.append({
                '深加工产物编码': deep_product_code,
                '深加工产物名称': item.get('深加工产物名称', ''),
                '原物料代码': origin_code,
                '原物料名称': origin_name,
                '一次拆解产物编码': first_product_code,
                '深加工结果(KG)': float(with_stock_kg),
                '数量(不考虑期初库存和库存结余)': float(no_opening_kg),
                '类别': category,
                '单价': float(unit_price),
                '工资': float(wage),
                '工资(不考虑期初库存和库存结余)': float(wage_no_opening)
            })
        
        # 计算总工资
        part1_wage_no_opening = part1_wage
        part2_wage_no_opening = part2_wage
        total_wage = part1_wage + part2_wage + part3_wage
        total_wage_no_opening = part1_wage_no_opening + part2_wage_no_opening + part3_wage_no_opening
        
        return {
            'total_wage': float(total_wage),
            'total_wage_no_opening': float(total_wage_no_opening),
            'part1_wage': float(part1_wage),
            'part2_wage': float(part2_wage),
            'part3_wage': float(part3_wage),
            'part1_wage_no_opening': float(part1_wage_no_opening),
            'part2_wage_no_opening': float(part2_wage_no_opening),
            'part3_wage_no_opening': float(part3_wage_no_opening),
            'part1_details': part1_details,
            'part2_details': part2_details,
            'part3_details': part3_details,
            'part1_count': len(part1_details),
            'part2_count': len(part2_details),
            'part3_count': len(part3_details)
        }
        
    except Exception as e:
        print(f"[计件工资计算] 计算失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'total_wage': 0.0,
            'total_wage_no_opening': 0.0,
            'part1_wage': 0.0,
            'part2_wage': 0.0,
            'part3_wage': 0.0,
            'part1_wage_no_opening': 0.0,
            'part2_wage_no_opening': 0.0,
            'part3_wage_no_opening': 0.0,
            'part1_details': [],
            'part2_details': [],
            'part3_details': [],
            'error': str(e)
        }


def classify_by_product_name(product_name):
    """
    根据物料/产物名称进行模糊匹配分类
    
    匹配顺序：电视、电脑、冰箱、空调、洗衣机
    按第一个匹配的关键词分类
    
    Args:
        product_name: 物料/产物名称字符串
        
    Returns:
        str: 分类名称（'电视'、'电脑'、'冰箱'、'空调'、'洗衣机'），如果没有匹配则返回 None
    """
    if not product_name:
        return None
    
    product_name_str = str(product_name)
    
    # 特殊物料名称映射（优先匹配）
    special_tv_keywords = ['CRT其它机壳破碎塑料', '线路板边框破碎塑料', '废旧玻璃电子枪', '废旧金属荫罩压块铁', '黑白']
    for keyword in special_tv_keywords:
        if keyword in product_name_str:
            return '电视'
    
    # 电视归类关键词：电视、彩电、等离子
    tv_keywords = ['电视', '彩电', '等离子']
    for keyword in tv_keywords:
        if keyword in product_name_str:
            return '电视'
    
    # 电脑归类关键词：电脑、显示器、笔记本、主机
    pc_keywords = ['电脑', '显示器', '笔记本', '主机', '废旧金属黑色金属-铁及其合金-电子枪']
    for keyword in pc_keywords:
        if keyword in product_name_str:
            return '电脑'
    
    # 其他分类：冰箱、空调、洗衣机
    if '屏' in product_name_str:
        return '电视'

    other_keywords = ['冰箱', '冰柜', '空调', '洗衣机', '双缸']
    for keyword in other_keywords:
        if keyword in product_name_str:
            if keyword in ['冰箱', '冰柜']:
                return '冰箱'
            elif keyword in ['洗衣机', '双缸']:
                return '洗衣机'
            else:
                return keyword
    
    return None


def calculate_direct_labor_cost(app_data, prediction_period=1):
    """
    计算直接人工成本（包含计件工资和分摊的固定工资、社保、公积金）
    
    Args:
        app_data: AppDataManager实例
        prediction_period: 预测期数（月），默认1
        
    Returns:
        dict: 包含计件工资、分摊固定成本和直接人工成本的完整数据
    """
    try:
        # 先计算计件工资
        piece_rate_result = calculate_piece_rate_wage(app_data)
        
        if piece_rate_result.get('error'):
            return {
                **piece_rate_result,
                'total_fixed_cost': 0.0,
                'direct_labor_cost': piece_rate_result.get('total_wage', 0.0),
                'direct_labor_cost_no_opening': piece_rate_result.get('total_wage_no_opening', piece_rate_result.get('total_wage', 0.0)),
                'category_details': {},
                'product_category_stats': {
                    '电视': {'wage': 0.0, 'fixed_cost': 0.0},
                    '电脑': {'wage': 0.0, 'fixed_cost': 0.0},
                    '冰箱': {'wage': 0.0, 'fixed_cost': 0.0},
                    '空调': {'wage': 0.0, 'fixed_cost': 0.0},
                    '洗衣机': {'wage': 0.0, 'fixed_cost': 0.0}
                },
                'product_category_stats_display': {
                    '电视': {'wage': 0.0, 'fixed_cost': 0.0},
                    '电脑': {'wage': 0.0, 'fixed_cost': 0.0},
                    '冰箱': {'wage': 0.0, 'fixed_cost': 0.0},
                    '空调': {'wage': 0.0, 'fixed_cost': 0.0},
                    '洗衣机': {'wage': 0.0, 'fixed_cost': 0.0}
                },
                'product_category_stats_no_opening': {
                    '电视': {'wage': 0.0, 'fixed_cost': 0.0},
                    '电脑': {'wage': 0.0, 'fixed_cost': 0.0},
                    '冰箱': {'wage': 0.0, 'fixed_cost': 0.0},
                    '空调': {'wage': 0.0, 'fixed_cost': 0.0},
                    '洗衣机': {'wage': 0.0, 'fixed_cost': 0.0}
                }
            }
        
        # 获取薪酬核算基础数据
        salary_df = get_salary_accounting_dataframe()
        if salary_df is None or salary_df.empty:
            return {
                **piece_rate_result,
                'total_fixed_cost': 0.0,
                'direct_labor_cost': piece_rate_result.get('total_wage', 0.0),
                'direct_labor_cost_no_opening': piece_rate_result.get('total_wage_no_opening', piece_rate_result.get('total_wage', 0.0)),
                'category_details': {},
                'product_category_stats': {
                    '电视': {'wage': 0.0, 'fixed_cost': 0.0},
                    '电脑': {'wage': 0.0, 'fixed_cost': 0.0},
                    '冰箱': {'wage': 0.0, 'fixed_cost': 0.0},
                    '空调': {'wage': 0.0, 'fixed_cost': 0.0},
                    '洗衣机': {'wage': 0.0, 'fixed_cost': 0.0}
                },
                'product_category_stats_display': {
                    '电视': {'wage': 0.0, 'fixed_cost': 0.0},
                    '电脑': {'wage': 0.0, 'fixed_cost': 0.0},
                    '冰箱': {'wage': 0.0, 'fixed_cost': 0.0},
                    '空调': {'wage': 0.0, 'fixed_cost': 0.0},
                    '洗衣机': {'wage': 0.0, 'fixed_cost': 0.0}
                },
                'product_category_stats_no_opening': {
                    '电视': {'wage': 0.0, 'fixed_cost': 0.0},
                    '电脑': {'wage': 0.0, 'fixed_cost': 0.0},
                    '冰箱': {'wage': 0.0, 'fixed_cost': 0.0},
                    '空调': {'wage': 0.0, 'fixed_cost': 0.0},
                    '洗衣机': {'wage': 0.0, 'fixed_cost': 0.0}
                },
                'error': '无法获取薪酬核算基础数据'
            }
        
        # 创建岗位到薪酬数据的映射
        salary_dict = {}
        for _, row in salary_df.iterrows():
            position = str(row.get('岗位', ''))
            if position:
                salary_dict[position] = {
                    '人员基础配置': float(row.get('人员基础配置', 0)) if pd.notna(row.get('人员基础配置')) else 0.0,
                    '平均工资（元/月/人）': float(row.get('平均工资（元/月/人）', 0)) if pd.notna(row.get('平均工资（元/月/人）')) else 0.0,
                    '奖励/补助（元/月）': float(row.get('奖励/补助（元/月）', 0)) if pd.notna(row.get('奖励/补助（元/月）')) else 0.0,
                    '餐补（元/月/人）': float(row.get('餐补（元/月/人）', 0)) if pd.notna(row.get('餐补（元/月/人）')) else 0.0,
                    '年终奖（元/人）': float(row.get('年终奖（元/人）', 0)) if pd.notna(row.get('年终奖（元/人）')) else 0.0,
                    '养老保险费（元/月/人）': float(row.get('养老保险费（元/月/人）', 0)) if pd.notna(row.get('养老保险费（元/月/人）')) else 0.0,
                    '失业保险费（元/月/人）': float(row.get('失业保险费（元/月/人）', 0)) if pd.notna(row.get('失业保险费（元/月/人）')) else 0.0,
                    '医疗/生育保险费（元/月/人）': float(row.get('医疗/生育保险费（元/月/人）', 0)) if pd.notna(row.get('医疗/生育保险费（元/月/人）')) else 0.0,
                    '工伤保险费（元/月/人）': float(row.get('工伤保险费（元/月/人）', 0)) if pd.notna(row.get('工伤保险费（元/月/人）')) else 0.0,
                    '住房公积金（元/月/人）': float(row.get('住房公积金（元/月/人）', 0)) if pd.notna(row.get('住房公积金（元/月/人）')) else 0.0,
                }
        
        # 定义类别映射
        category_mapping = {
            '白电': '白电拆解',
            '黑电': '黑电拆解',
            '冰箱': '冰箱拆解',
            '金属打包': '金属打包',
            '塑料': '塑料破碎分选',
            '屏': '辅助车间-屏'
        }
        
        def _get_wage_no_opening(item, default_wage):
            """读取不考虑口径工资：仅在缺失/空值时回退，保留0值。"""
            raw_value = item.get('工资(不考虑期初库存和库存结余)', None)
            if raw_value is None or raw_value == '':
                return float(default_wage or 0.0)
            return float(raw_value)

        # 收集所有计件工资明细
        all_wage_details = []
        
        # 第一部分：旧机类别（需要按物料名称筛选）
        if piece_rate_result.get('part1_details'):
            for item in piece_rate_result['part1_details']:
                material_name = str(item.get('物料名称', '')).upper()
                category = None
                
                # 白电：包含"空调"或"洗衣机"
                if '空调' in material_name or '洗衣机' in material_name:
                    category = '白电'
                # 黑电：包含"电视"或"电脑"
                elif '电视' in material_name or '电脑' in material_name:
                    category = '黑电'
                # 冰箱：包含"冰箱"
                elif '冰箱' in material_name:
                    category = '冰箱'
                
                if category:
                    wage = float(item.get('工资', 0))
                    wage_no_opening = _get_wage_no_opening(item, wage)
                    all_wage_details.append({
                        'category': category,
                        'wage': wage,
                        'wage_no_opening': wage_no_opening,
                        'item': item
                    })
        
        # 第二部分：一次拆解产物（塑料）
        if piece_rate_result.get('part2_details'):
            for item in piece_rate_result['part2_details']:
                wage = float(item.get('工资', 0))
                wage_no_opening = _get_wage_no_opening(item, wage)
                all_wage_details.append({
                    'category': '塑料',
                    'wage': wage,
                    'wage_no_opening': wage_no_opening,
                    'item': item
                })
        
        # 第三部分：深加工（一破、打包铁、屏）
        if piece_rate_result.get('part3_details'):
            for item in piece_rate_result['part3_details']:
                item_category = str(item.get('类别', ''))
                if item_category == '打包铁':
                    wage = float(item.get('工资', 0))
                    wage_no_opening = _get_wage_no_opening(item, wage)
                    all_wage_details.append({
                        'category': '金属打包',
                        'wage': wage,
                        'wage_no_opening': wage_no_opening,
                        'item': item
                    })
                elif item_category == '一破':
                    wage = float(item.get('工资', 0))
                    wage_no_opening = _get_wage_no_opening(item, wage)
                    all_wage_details.append({
                        'category': '塑料',
                        'wage': wage,
                        'wage_no_opening': wage_no_opening,
                        'item': item
                    })
                elif item_category == '屏':
                    wage = float(item.get('工资', 0))
                    wage_no_opening = _get_wage_no_opening(item, wage)
                    all_wage_details.append({
                        'category': '屏',
                        'wage': wage,
                        'wage_no_opening': wage_no_opening,
                        'item': item
                    })
        
        # 按类别汇总工资（考虑/不考虑期初库存和库存结余）
        category_wage_sum = {}
        category_wage_sum_no_opening = {}
        category_items = {}
        for detail in all_wage_details:
            cat = detail['category']
            if cat not in category_wage_sum:
                category_wage_sum[cat] = 0.0
                category_wage_sum_no_opening[cat] = 0.0
                category_items[cat] = []
            category_wage_sum[cat] += detail['wage']
            wage_no_opening = float(detail.get('wage_no_opening', detail['wage']))
            category_wage_sum_no_opening[cat] += wage_no_opening
            category_items[cat].append(detail)
        
        # 计算每个类别的分摊固定成本
        category_details = {}
        total_fixed_cost = 0.0
        
        no_opening_fixed_cost_recalc_categories = {'金属打包', '塑料', '屏'}
        for category, total_wage in category_wage_sum.items():
            if total_wage == 0:
                continue
            
            # 获取对应岗位的薪酬数据
            position = category_mapping.get(category)
            if not position or position not in salary_dict:
                print(f"[直接人工成本] 未找到类别 {category} 对应的岗位 {position} 的薪酬数据")
                continue
            
            salary_data = salary_dict[position]
            personnel_base = salary_data['人员基础配置']
            
            # 计算月均固定成本（年终奖已经是月平均数，不需要除以12）
            monthly_fixed_cost_per_person = (
                salary_data['平均工资（元/月/人）'] +
                salary_data['奖励/补助（元/月）'] +
                salary_data['餐补（元/月/人）'] +
                salary_data['年终奖（元/人）'] +
                salary_data['养老保险费（元/月/人）'] +
                salary_data['失业保险费（元/月/人）'] +
                salary_data['医疗/生育保险费（元/月/人）'] +
                salary_data['工伤保险费（元/月/人）'] +
                salary_data['住房公积金（元/月/人）']
            )
            
            # 计算每个物料的分摊
            item_allocations = []
            total_wage_no_opening_by_category = category_wage_sum_no_opening.get(category, total_wage)
            for detail in category_items[category]:
                detail_item = detail['item']
                wage = float(detail.get('wage', 0))
                wage_no_opening = float(detail.get('wage_no_opening', wage))
                allocation_ratio = detail['wage'] / total_wage if total_wage > 0 else 0
                allocation_ratio_no_opening = wage_no_opening / total_wage_no_opening_by_category if total_wage_no_opening_by_category > 0 else allocation_ratio
                item_fixed_cost = allocation_ratio * personnel_base * monthly_fixed_cost_per_person * prediction_period
                if category in no_opening_fixed_cost_recalc_categories:
                    item_fixed_cost_no_opening = allocation_ratio_no_opening * personnel_base * monthly_fixed_cost_per_person * prediction_period
                else:
                    item_fixed_cost_no_opening = item_fixed_cost
                
                item_allocations.append({
                    'item': detail_item,
                    'wage_no_opening': float(wage_no_opening),
                    'allocation_ratio': float(allocation_ratio),
                    'allocation_ratio_no_opening': float(allocation_ratio_no_opening),
                    'fixed_cost': float(item_fixed_cost),
                    'fixed_cost_no_opening': float(item_fixed_cost_no_opening)
                })
            
            category_total_fixed_cost = sum(item['fixed_cost'] for item in item_allocations)
            total_fixed_cost += category_total_fixed_cost
            
            category_details[category] = {
                'total_wage': float(total_wage),
                'total_fixed_cost': float(category_total_fixed_cost),
                'personnel_base': float(personnel_base),
                'monthly_fixed_cost_per_person': float(monthly_fixed_cost_per_person),
                'item_allocations': item_allocations
            }
        
        # 计算直接人工成本
        total_wage = piece_rate_result.get('total_wage', 0.0)
        total_wage_no_opening = piece_rate_result.get('total_wage_no_opening', total_wage)
        direct_labor_cost = total_wage + total_fixed_cost
        direct_labor_cost_no_opening = total_wage_no_opening + total_fixed_cost
        
        # 按四机一脑分类统计（电视、电脑、冰箱、空调、洗衣机）
        product_category_stats = {
            '电视': {'wage': 0.0, 'fixed_cost': 0.0},
            '电脑': {'wage': 0.0, 'fixed_cost': 0.0},
            '冰箱': {'wage': 0.0, 'fixed_cost': 0.0},
            '空调': {'wage': 0.0, 'fixed_cost': 0.0},
            '洗衣机': {'wage': 0.0, 'fixed_cost': 0.0}
        }
        product_category_stats_no_opening = {
            '电视': {'wage': 0.0, 'fixed_cost': 0.0},
            '电脑': {'wage': 0.0, 'fixed_cost': 0.0},
            '冰箱': {'wage': 0.0, 'fixed_cost': 0.0},
            '空调': {'wage': 0.0, 'fixed_cost': 0.0},
            '洗衣机': {'wage': 0.0, 'fixed_cost': 0.0}
        }
        # 仅用于「直接人工成本页/导出」展示：按原物料名称等优化归类；不影响生产成本分摊缓存链路
        product_category_stats_display = {
            '电视': {'wage': 0.0, 'fixed_cost': 0.0},
            '电脑': {'wage': 0.0, 'fixed_cost': 0.0},
            '冰箱': {'wage': 0.0, 'fixed_cost': 0.0},
            '空调': {'wage': 0.0, 'fixed_cost': 0.0},
            '洗衣机': {'wage': 0.0, 'fixed_cost': 0.0}
        }
        
        # 方法1a：统计计件工资（历史口径，供生产成本分摊等下游保持与改造前一致）
        if piece_rate_result.get('part1_details'):
            for item in piece_rate_result['part1_details']:
                legacy_name = str(item.get('物料名称', '') or '')
                product_category = classify_by_product_name(legacy_name)
                wage = float(item.get('工资', 0))
                if product_category and product_category in product_category_stats:
                    product_category_stats[product_category]['wage'] += wage
        
        if piece_rate_result.get('part2_details'):
            for item in piece_rate_result['part2_details']:
                legacy_name = str(item.get('拆解产物名称', '') or '')
                product_category = classify_by_product_name(legacy_name)
                wage = float(item.get('工资', 0))
                if product_category and product_category in product_category_stats:
                    product_category_stats[product_category]['wage'] += wage
        
        if piece_rate_result.get('part3_details'):
            for item in piece_rate_result['part3_details']:
                legacy_name = str(item.get('深加工产物名称', '') or '')
                product_category = classify_by_product_name(legacy_name)
                wage = float(item.get('工资', 0))
                if product_category and product_category in product_category_stats:
                    product_category_stats[product_category]['wage'] += wage

        # 方法1b：统计计件工资（展示口径 + 不考虑口径工资）
        if piece_rate_result.get('part1_details'):
            for item in piece_rate_result['part1_details']:
                classify_name = item.get('原物料名称', '') or item.get('物料名称', '') or ''
                product_category = classify_by_product_name(classify_name)
                wage = float(item.get('工资', 0))
                wage_no_opening = _get_wage_no_opening(item, wage)
                if product_category and product_category in product_category_stats_display:
                    product_category_stats_display[product_category]['wage'] += wage
                    product_category_stats_no_opening[product_category]['wage'] += wage_no_opening
        
        if piece_rate_result.get('part2_details'):
            for item in piece_rate_result['part2_details']:
                classify_name = item.get('原物料名称', '') or item.get('拆解产物名称', '') or ''
                product_category = classify_by_product_name(classify_name)
                wage = float(item.get('工资', 0))
                wage_no_opening = _get_wage_no_opening(item, wage)
                if product_category and product_category in product_category_stats_display:
                    product_category_stats_display[product_category]['wage'] += wage
                    product_category_stats_no_opening[product_category]['wage'] += wage_no_opening
        
        if piece_rate_result.get('part3_details'):
            for item in piece_rate_result['part3_details']:
                classify_name = item.get('原物料名称', '') or item.get('深加工产物名称', '') or ''
                product_category = classify_by_product_name(classify_name)
                wage = float(item.get('工资', 0))
                wage_no_opening = _get_wage_no_opening(item, wage)
                if product_category and product_category in product_category_stats_display:
                    product_category_stats_display[product_category]['wage'] += wage
                    product_category_stats_no_opening[product_category]['wage'] += wage_no_opening
        
        # 方法2：统计分摊固定成本 - 直接遍历category_details中的item_allocations
        # 这样可以确保匹配到所有有分摊固定成本的物料
        for category, category_data in category_details.items():
            if not category_data or not category_data.get('item_allocations'):
                continue
            
            for allocation in category_data['item_allocations']:
                allocation_item = allocation.get('item', {})
                fixed_cost = allocation.get('fixed_cost', 0)
                fixed_cost_no_opening = allocation.get('fixed_cost_no_opening', fixed_cost)
                
                # 历史口径：与改造前一致（用于生产成本分摊缓存，避免首页利润汇总等联动变化）
                legacy_product_name = ''
                if allocation_item.get('物料名称'):
                    legacy_product_name = allocation_item.get('物料名称', '')
                elif allocation_item.get('拆解产物名称'):
                    legacy_product_name = allocation_item.get('拆解产物名称', '')
                elif allocation_item.get('深加工产物名称'):
                    legacy_product_name = allocation_item.get('深加工产物名称', '')

                # 展示口径：优先原物料名称（直接人工成本页）
                display_product_name = (
                    allocation_item.get('原物料名称')
                    or allocation_item.get('物料名称')
                    or allocation_item.get('拆解产物名称')
                    or allocation_item.get('深加工产物名称')
                    or ''
                )
                
                legacy_category = classify_by_product_name(legacy_product_name)
                if legacy_category and legacy_category in product_category_stats:
                    product_category_stats[legacy_category]['fixed_cost'] += fixed_cost

                display_category = classify_by_product_name(display_product_name)
                if display_category and display_category in product_category_stats_display:
                    product_category_stats_display[display_category]['fixed_cost'] += fixed_cost
                    product_category_stats_no_opening[display_category]['fixed_cost'] += fixed_cost_no_opening
        
        return {
            **piece_rate_result,
            'total_fixed_cost': float(total_fixed_cost),
            'direct_labor_cost': float(direct_labor_cost),
            'direct_labor_cost_no_opening': float(direct_labor_cost_no_opening),
            'prediction_period': int(prediction_period),
            'category_details': category_details,
            'product_category_stats': product_category_stats,
            'product_category_stats_display': product_category_stats_display,
            'product_category_stats_no_opening': product_category_stats_no_opening
        }
        
    except Exception as e:
        print(f"[直接人工成本计算] 计算失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'total_wage': 0.0,
            'total_wage_no_opening': 0.0,
            'part1_wage': 0.0,
            'part2_wage': 0.0,
            'part3_wage': 0.0,
            'part1_wage_no_opening': 0.0,
            'part2_wage_no_opening': 0.0,
            'part3_wage_no_opening': 0.0,
            'part1_details': [],
            'part2_details': [],
            'part3_details': [],
            'total_fixed_cost': 0.0,
            'direct_labor_cost': 0.0,
            'direct_labor_cost_no_opening': 0.0,
            'prediction_period': int(prediction_period),
            'category_details': {},
            'product_category_stats': {
                '电视': {'wage': 0.0, 'fixed_cost': 0.0},
                '电脑': {'wage': 0.0, 'fixed_cost': 0.0},
                '冰箱': {'wage': 0.0, 'fixed_cost': 0.0},
                '空调': {'wage': 0.0, 'fixed_cost': 0.0},
                '洗衣机': {'wage': 0.0, 'fixed_cost': 0.0}
            },
            'product_category_stats_display': {
                '电视': {'wage': 0.0, 'fixed_cost': 0.0},
                '电脑': {'wage': 0.0, 'fixed_cost': 0.0},
                '冰箱': {'wage': 0.0, 'fixed_cost': 0.0},
                '空调': {'wage': 0.0, 'fixed_cost': 0.0},
                '洗衣机': {'wage': 0.0, 'fixed_cost': 0.0}
            },
            'product_category_stats_no_opening': {
                '电视': {'wage': 0.0, 'fixed_cost': 0.0},
                '电脑': {'wage': 0.0, 'fixed_cost': 0.0},
                '冰箱': {'wage': 0.0, 'fixed_cost': 0.0},
                '空调': {'wage': 0.0, 'fixed_cost': 0.0},
                '洗衣机': {'wage': 0.0, 'fixed_cost': 0.0}
            },
            'error': str(e)
        }


@cost_forecast_bp.route('/material-cost', methods=['GET'])
def get_material_cost():
    """获取拆解物原料成本数据"""
    try:
        app_data = get_session_data_manager()
        
        # 获取手工数据
        manual_data = app_data.get_data('extracted_data_manual')
        
        # 检查数据是否已被清除（通过检查清除标志）
        data_cleared = app_data.get_data('__data_cleared__')
        if data_cleared:
            # 数据已被清除，直接返回空数据，不自动初始化
            return jsonify({
                'success': True,
                'data': [],
                'message': '数据已清除'
            })
        
        # 如果手工数据不存在，尝试从只读数据自动初始化
        if manual_data is None or manual_data.empty:
            readonly_data = app_data.get_data('extracted_data')
            if readonly_data is not None and not readonly_data.empty:
                # 有只读数据，自动初始化手工数据
                print("[自动初始化] 检测到只读数据，自动初始化手工数据...")
                try:
                    import pandas as pd
                    from datetime import datetime
                    
                    # 复制所有类别的数据
                    manual_data = readonly_data.copy()
                    
                    # 检查是否有旧机类别
                    if '类别' in manual_data.columns:
                        old_machine_count = len(manual_data[manual_data['类别'] == '旧机'])
                        if old_machine_count == 0:
                            print("[自动初始化] 没有找到旧机类别数据，跳过初始化")
                            return jsonify({
                                'success': True,
                                'data': [],
                                'message': '暂无旧机类别数据'
                            })
                    
                    # 添加新列
                    manual_data['初始数据'] = 0.0
                    manual_data['本期计划采购数量'] = 0.0
                    manual_data['计划采购单价'] = 0.0
                    manual_data['本期计划投产数量'] = 0.0
                    
                    # 只为旧机类别填充数据
                    if '类别' in manual_data.columns and '非限制使用的库存' in manual_data.columns:
                        mask_old_machine = manual_data['类别'] == '旧机'
                        
                        # 初始数据 = 非限制使用的库存
                        inventory_values = pd.to_numeric(manual_data.loc[mask_old_machine, '非限制使用的库存'], errors='coerce').fillna(0)
                        manual_data.loc[mask_old_machine, '初始数据'] = inventory_values
                        
                        # 本期计划投产数量 = 初始数据
                        manual_data.loc[mask_old_machine, '本期计划投产数量'] = manual_data.loc[mask_old_machine, '初始数据']
                        
                        # 计划采购单价 = 单价（从Excel表内的单价列复制）
                        if '单价' in manual_data.columns:
                            price_values = pd.to_numeric(manual_data.loc[mask_old_machine, '单价'], errors='coerce').fillna(0)
                            manual_data.loc[mask_old_machine, '计划采购单价'] = price_values
                    
                    # 重新排列列顺序
                    cols = list(manual_data.columns)
                    if '非限制使用的库存' in cols:
                        insert_pos = cols.index('非限制使用的库存') + 1
                        new_cols = [c for c in cols if c not in ['初始数据', '本期计划采购数量', '计划采购单价', '本期计划投产数量']]
                        new_cols.insert(insert_pos, '初始数据')
                        new_cols.insert(insert_pos + 1, '本期计划投产数量')
                        new_cols.insert(insert_pos + 2, '本期计划采购数量')
                        new_cols.insert(insert_pos + 3, '计划采购单价')
                        manual_data = manual_data[new_cols]
                    
                    # 保存手工数据
                    app_data.set_data('extracted_data_manual', manual_data)
                    app_data.set_data('original_extracted_data', readonly_data.copy())
                    app_data.set_data('extracted_data_modified', False)
                    app_data.set_data('extracted_modification_timestamp', datetime.now().isoformat())
                    
                    print("[自动初始化] 手工数据初始化成功")
                except Exception as init_error:
                    print(f"[自动初始化] 初始化失败: {init_error}")
                    import traceback
                    traceback.print_exc()
        
        if manual_data is None or manual_data.empty:
            return jsonify({
                'success': True,
                'data': [],
                'message': '暂无提取结果数据'
            })
        
        # 计算成本（每次调用都重新计算，确保数据最新）
        cost_data = calculate_material_cost(manual_data)
        
        # 保存计算结果（可选，用于缓存）
        app_data.set_data('cost_forecast_data', cost_data)
        
        # 只返回旧机类别
        if '类别' in cost_data.columns:
            display_data = cost_data[cost_data['类别'] == '旧机'].copy()
        else:
            display_data = cost_data
        
        # 转换为JSON
        from app.utils.data_utils import safe_json_convert
        result_data = safe_json_convert(display_data)
        
        return jsonify({
            'success': True,
            'data': result_data
        })
        
    except Exception as e:
        print(f"获取拆解物原料成本失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@cost_forecast_bp.route('/calculate', methods=['POST'])
@login_required
@require_can_edit
def calculate_cost():
    """触发成本计算"""
    try:
        app_data = get_session_data_manager()
        
        # 获取手工数据
        manual_data = app_data.get_data('extracted_data_manual')
        
        if manual_data is None or manual_data.empty:
            return jsonify({
                'success': False,
                'error': '没有找到提取结果数据，请先初始化数据'
            }), 400
        
        # 计算成本
        cost_data = calculate_material_cost(manual_data)
        
        # 保存计算结果
        app_data.set_data('cost_forecast_data', cost_data)
        
        # 统计信息
        if '类别' in cost_data.columns:
            display_data = cost_data[cost_data['类别'] == '旧机'].copy()
        else:
            display_data = cost_data
        
        total_cost = display_data['拆解物原料成本'].sum() if '拆解物原料成本' in display_data.columns else 0
        total_records = len(display_data)
        
        return jsonify({
            'success': True,
            'message': f'成本计算完成，共 {total_records} 条记录',
            'total_cost': float(total_cost),
            'total_records': total_records
        })
        
    except Exception as e:
        print(f"成本计算失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


def calculate_indirect_labor_cost(app_data, prediction_period=1, include_no_opening_columns=True):
    """
    计算间接人工成本（包含9种提成单价的成本和分摊的固定工资、社保、公积金）
    
    计算公式：
    1. 旧机：物料代码匹配的"本期实际投产数量" × 各种提成单价
    2. 一次拆解产物：拆解产物编码匹配的"计算结果(KG)" × 各种提成单价
    3. 一破：深加工产物编码匹配的"深加工结果(KG)" × 各种提成单价
    4. 分摊固定成本 = 物料提成成本/提成总数 × 人员基础配置 × 月均固定成本 × 预测期数
    
    Args:
        app_data: AppDataManager实例
        prediction_period: 预测期数（月），默认1
        include_no_opening_columns: 为 True 时计算「不考虑期初库存和库存结余」扩展列及塑料破碎分摊
            对应列（仅间接人工成本页/导出需要）。期间费用、利润汇总等链路应传 False，避免与历史汇总口径产生任何差异。
        
    Returns:
        dict: 包含总成本和分类明细的字典
    """
    try:
        def _safe_str(value):
            """将可能为 NaN/None 的值安全转为字符串。"""
            if value is None:
                return ''
            try:
                if pd.isna(value):
                    return ''
            except Exception:
                pass
            s = str(value).strip()
            if s.lower() == 'nan':
                return ''
            return s

        # 构建"原物料代码 -> 原物料名称"的回填映射（尽量从旧机/提取结果里拿到名称）
        # 目的：当原始数据/深加工数据缺失"原物料名称"时，页面新增列不至于为空。
        origin_material_name_map = {}
        try:
            extracted_for_origin = app_data.get_data('extracted_data_manual')
            if extracted_for_origin is None or extracted_for_origin.empty:
                extracted_for_origin = app_data.get_data('extracted_data')

            if extracted_for_origin is not None and not extracted_for_origin.empty:
                # 物料代码列兼容
                origin_code_col = None
                for col in ['物料代码', 'R3系统代码', '代码']:
                    if col in extracted_for_origin.columns:
                        origin_code_col = col
                        break

                # 名称列兼容（优先"物料描述"，其次"物料名称"）
                origin_name_col = None
                for col in ['物料描述', '物料名称', '原物料名称']:
                    if col in extracted_for_origin.columns:
                        origin_name_col = col
                        break

                if origin_code_col and origin_name_col:
                    for _, r in extracted_for_origin[[origin_code_col, origin_name_col]].dropna().iterrows():
                        code = str(r.get(origin_code_col, '')).strip()
                        if code.endswith('.0'):
                            code = code[:-2]
                        name = str(r.get(origin_name_col, '')).strip()
                        if code and name and code not in origin_material_name_map:
                            origin_material_name_map[code] = name
        except Exception:
            # 回填映射失败不影响主流程
            pass

        # 获取计件人工标准数据
        labor_cost_df = get_labor_cost_dataframe()
        if labor_cost_df is None or labor_cost_df.empty:
            return {
                'total_cost': 0.0,
                'part1_cost': 0.0,
                'part2_cost': 0.0,
                'part3_cost': 0.0,
                'part1_details': [],
                'part2_details': [],
                'part3_details': [],
                'category_totals': {
                    '品管提成': 0.0,
                    '物流主管提成': 0.0,
                    '物流卸货提成': 0.0,
                    '班组长提成': 0.0,
                    '生产主管提成': 0.0,
                    '维修班长提成': 0.0,
                    '维修员提成': 0.0,
                    '冰箱维修主管提成': 0.0,
                    '叉车司磅库管等提成': 0.0
                },
                'total_fixed_cost': 0.0,
                'indirect_labor_cost': 0.0,
                'category_fixed_costs': {},
                'prediction_period': int(prediction_period),
                'error': '无法获取计件人工标准数据'
            }
        
        # 定义9种提成单价字段名
        commission_fields = [
            '品管提成单价',
            '物流主管提成单价',
            '物流卸货提成单价',
            '班组长提成单价',
            '生产主管提成单价',
            '维修班长提成单价',
            '维修员提成单价',
            '冰箱维修主管提成单价',
            '叉车司磅库管等提成单价'
        ]
        
        # 创建物料代码到提成单价的映射字典（按类别）
        labor_cost_dict = {}
        for _, row in labor_cost_df.iterrows():
            code = str(row['R3系统代码'])
            category = str(row['类别'])
            key = (code, category)
            labor_cost_dict[key] = {}
            for field in commission_fields:
                labor_cost_dict[key][field] = float(row[field]) if pd.notna(row[field]) else 0.0
        
        no_opening_deep_result_map = {}
        no_opening_meta_map = {}
        with_stock_deep_result_map2 = {}
        with_stock_meta_map2 = {}
        if include_no_opening_columns:
            from app.api.data_management_api import (
                calculate_deep_processing_product_output_value_without_stock_data,
                calculate_deep_processing_product_output_value_data,
            )
            _output_value_ok2, _output_value_data2, _msg2 = calculate_deep_processing_product_output_value_without_stock_data(app_data)
            if _output_value_ok2 and _output_value_data2:
                _ingest_deep_output_value_items(
                    _output_value_data2, no_opening_deep_result_map, no_opening_meta_map
                )
            _ws_ok, _ws_data, _ = calculate_deep_processing_product_output_value_data(app_data)
            if _ws_ok and _ws_data:
                _ingest_deep_output_value_items(
                    _ws_data, with_stock_deep_result_map2, with_stock_meta_map2
                )

        total_cost = 0.0
        part1_cost = 0.0  # 旧机类别
        part2_cost = 0.0  # 一次拆解产物类别
        part3_cost = 0.0  # 一破、打包铁、屏类别
        part1_details = []
        part2_details = []
        part3_details = []
        category_totals = {
            '品管提成': 0.0,
            '物流主管提成': 0.0,
            '物流卸货提成': 0.0,
            '班组长提成': 0.0,
            '生产主管提成': 0.0,
            '维修班长提成': 0.0,
            '维修员提成': 0.0,
            '冰箱维修主管提成': 0.0,
            '叉车司磅库管等提成': 0.0
        }
        
        # 第一部分：提取结果数据编辑表（旧机类别）
        extracted_data = app_data.get_data('extracted_data_manual')
        if extracted_data is not None and not extracted_data.empty:
            if '类别' in extracted_data.columns and '非限制使用的库存' in extracted_data.columns:
                old_machine_data = extracted_data[extracted_data['类别'] == '旧机'].copy()
                if not old_machine_data.empty:
                    # 确保数值类型
                    old_machine_data['非限制使用的库存'] = pd.to_numeric(
                        old_machine_data['非限制使用的库存'], errors='coerce'
                    ).fillna(0)
                    
                    # 获取物料代码字段
                    material_code_col = None
                    for col in ['物料代码', 'R3系统代码', '代码']:
                        if col in old_machine_data.columns:
                            material_code_col = col
                            break
                    
                    if material_code_col:
                        for _, row in old_machine_data.iterrows():
                            material_code = str(row[material_code_col])
                            quantity = float(row['非限制使用的库存'])
                            
                            if quantity > 0:
                                # 查找旧机类别的提成单价
                                key = (material_code, '旧机')
                                commission_prices = labor_cost_dict.get(key, {})
                                
                                if commission_prices:
                                    material_name = _safe_str(row.get('物料描述', ''))
                                    detail = {
                                        '物料代码': material_code,
                                        '物料名称': material_name,
                                        # 旧机类别：原物料即本身
                                        '原物料代码': material_code,
                                        '原物料名称': material_name,
                                        '数量': float(quantity),
                                        '类别': '旧机',
                                        '品管提成单价': commission_prices.get('品管提成单价', 0.0),
                                        '物流主管提成单价': commission_prices.get('物流主管提成单价', 0.0),
                                        '物流卸货提成单价': commission_prices.get('物流卸货提成单价', 0.0),
                                        '班组长提成单价': commission_prices.get('班组长提成单价', 0.0),
                                        '生产主管提成单价': commission_prices.get('生产主管提成单价', 0.0),
                                        '维修班长提成单价': commission_prices.get('维修班长提成单价', 0.0),
                                        '维修员提成单价': commission_prices.get('维修员提成单价', 0.0),
                                        '冰箱维修主管提成单价': commission_prices.get('冰箱维修主管提成单价', 0.0),
                                        '叉车司磅库管等提成单价': commission_prices.get('叉车司磅库管等提成单价', 0.0)
                                    }
                                    
                                    # 计算各项提成成本
                                    item_total = 0.0
                                    for field in commission_fields:
                                        price = detail[field]
                                        cost = quantity * price
                                        cost_field = field.replace('单价', '成本')
                                        detail[cost_field] = float(cost)
                                        item_total += cost
                                        # 更新分类汇总
                                        category_name = field.replace('提成单价', '提成')
                                        category_totals[category_name] += cost
                                    
                                    detail['总成本'] = float(item_total)
                                    if include_no_opening_columns:
                                        detail['数量/重量(不考虑期初库存和库存结余)'] = float(quantity)
                                        detail['班组长提成成本(不考虑期初库存和库存结余)'] = float(
                                            detail.get('班组长提成成本', 0.0)
                                        )
                                    part1_cost += item_total
                                    total_cost += item_total
                                    part1_details.append(detail)
        
        # 第二部分：原始数据（未扣减）表（拆解产物类别）
        disassembly_data = app_data.get_data('disassembly_data')
        if disassembly_data is not None and not disassembly_data.empty:
            if '类别' in disassembly_data.columns and '计算结果(KG)' in disassembly_data.columns:
                disassembly_product_data = disassembly_data[disassembly_data['类别'] == '拆解产物'].copy()
                if not disassembly_product_data.empty:
                    # 确保数值类型
                    disassembly_product_data['计算结果(KG)'] = pd.to_numeric(
                        disassembly_product_data['计算结果(KG)'], errors='coerce'
                    ).fillna(0)
                    
                    if '拆解产物编码' in disassembly_product_data.columns:
                        for _, row in disassembly_product_data.iterrows():
                            product_code = str(row['拆解产物编码'])
                            # 处理可能的.0后缀
                            if product_code.endswith('.0'):
                                product_code = product_code[:-2]
                            result_kg = float(row['计算结果(KG)'])
                            
                            if result_kg > 0:
                                # 查找一次拆解产物类别的提成单价
                                key = (product_code, '一次拆解产物')
                                commission_prices = labor_cost_dict.get(key, {})
                                
                                if commission_prices:
                                    origin_code = _safe_str(row.get('原物料代码', ''))
                                    if origin_code.endswith('.0'):
                                        origin_code = origin_code[:-2]
                                    origin_name = _safe_str(row.get('原物料名称', ''))
                                    if not origin_name and origin_code:
                                        origin_name = origin_material_name_map.get(origin_code, '')
                                    detail = {
                                        '拆解产物编码': product_code,
                                        '拆解产物名称': row.get('拆解产物名称', ''),
                                        '原物料代码': origin_code,
                                        '原物料名称': origin_name,
                                        '计算结果(KG)': float(result_kg),
                                        '类别': '一次拆解产物',
                                        '品管提成单价': commission_prices.get('品管提成单价', 0.0),
                                        '物流主管提成单价': commission_prices.get('物流主管提成单价', 0.0),
                                        '物流卸货提成单价': commission_prices.get('物流卸货提成单价', 0.0),
                                        '班组长提成单价': commission_prices.get('班组长提成单价', 0.0),
                                        '生产主管提成单价': commission_prices.get('生产主管提成单价', 0.0),
                                        '维修班长提成单价': commission_prices.get('维修班长提成单价', 0.0),
                                        '维修员提成单价': commission_prices.get('维修员提成单价', 0.0),
                                        '冰箱维修主管提成单价': commission_prices.get('冰箱维修主管提成单价', 0.0),
                                        '叉车司磅库管等提成单价': commission_prices.get('叉车司磅库管等提成单价', 0.0)
                                    }
                                    
                                    # 计算各项提成成本
                                    item_total = 0.0
                                    for field in commission_fields:
                                        price = detail[field]
                                        cost = result_kg * price
                                        cost_field = field.replace('单价', '成本')
                                        detail[cost_field] = float(cost)
                                        item_total += cost
                                        # 更新分类汇总
                                        category_name = field.replace('提成单价', '提成')
                                        category_totals[category_name] += cost
                                    
                                    detail['总成本'] = float(item_total)
                                    if include_no_opening_columns:
                                        detail['数量/重量(不考虑期初库存和库存结余)'] = float(result_kg)
                                        detail['班组长提成成本(不考虑期初库存和库存结余)'] = float(
                                            detail.get('班组长提成成本', 0.0)
                                        )
                                    part2_cost += item_total
                                    total_cost += item_total
                                    part2_details.append(detail)
        
        # 第三部分：深加工（一破、打包铁、屏），行集 = 两口径产值 sheet 的并集
        if include_no_opening_columns:
            all_part3_keys = set(with_stock_deep_result_map2.keys()) | set(no_opening_deep_result_map.keys())
            for result_key in sorted(all_part3_keys):
                origin_code, first_product_code, deep_product_code = result_key
                with_stock_kg = float(with_stock_deep_result_map2.get(result_key, 0.0) or 0.0)
                no_opening_kg = float(no_opening_deep_result_map.get(result_key, 0.0) or 0.0)

                item = with_stock_meta_map2.get(result_key) or no_opening_meta_map.get(result_key) or {}
                category = item.get('类别', '')
                if category not in DEEP_PART3_CATEGORIES:
                    continue

                commission_prices = labor_cost_dict.get((deep_product_code, category), {})
                deep_result_kg = with_stock_kg

                origin_name = _safe_str_for_maps(item.get('原物料名称', ''))
                if not origin_name and origin_code:
                    origin_name = origin_material_name_map.get(origin_code, '')

                detail = {
                    '深加工产物编码': deep_product_code,
                    '深加工产物名称': item.get('深加工产物名称', ''),
                    '原物料代码': origin_code,
                    '原物料名称': origin_name,
                    '一次拆解产物编码': first_product_code,
                    '深加工结果(KG)': float(with_stock_kg),
                    '类别': category,
                    '品管提成单价': commission_prices.get('品管提成单价', 0.0),
                    '物流主管提成单价': commission_prices.get('物流主管提成单价', 0.0),
                    '物流卸货提成单价': commission_prices.get('物流卸货提成单价', 0.0),
                    '班组长提成单价': commission_prices.get('班组长提成单价', 0.0),
                    '生产主管提成单价': commission_prices.get('生产主管提成单价', 0.0),
                    '维修班长提成单价': commission_prices.get('维修班长提成单价', 0.0),
                    '维修员提成单价': commission_prices.get('维修员提成单价', 0.0),
                    '冰箱维修主管提成单价': commission_prices.get('冰箱维修主管提成单价', 0.0),
                    '叉车司磅库管等提成单价': commission_prices.get('叉车司磅库管等提成单价', 0.0)
                }

                item_total = 0.0
                for field in commission_fields:
                    price = detail[field]
                    cost = deep_result_kg * price
                    cost_field = field.replace('单价', '成本')
                    detail[cost_field] = float(cost)
                    item_total += cost
                    category_name = field.replace('提成单价', '提成')
                    category_totals[category_name] += cost

                detail['总成本'] = float(item_total)
                tl_price = float(detail.get('班组长提成单价', 0.0) or 0.0)
                detail['数量/重量(不考虑期初库存和库存结余)'] = float(no_opening_kg)
                detail['班组长提成成本(不考虑期初库存和库存结余)'] = float(no_opening_kg * tl_price)
                part3_cost += item_total
                total_cost += item_total
                part3_details.append(detail)
        
        # 确保所有分类汇总都是浮点数
        for key in category_totals:
            category_totals[key] = float(category_totals[key])
        
        # 计算人员分摊的固定工资、社保、公积金
        total_fixed_cost = 0.0
        category_fixed_costs = {}
        
        try:
            # 获取薪酬核算基础数据
            salary_df = get_salary_accounting_dataframe()
            if salary_df is not None and not salary_df.empty:
                # 创建岗位到薪酬数据的映射
                salary_dict = {}
                for _, row in salary_df.iterrows():
                    position = str(row.get('岗位', ''))
                    if position:
                        salary_dict[position] = {
                            '人员基础配置': float(row.get('人员基础配置', 0)) if pd.notna(row.get('人员基础配置')) else 0.0,
                            '平均工资（元/月/人）': float(row.get('平均工资（元/月/人）', 0)) if pd.notna(row.get('平均工资（元/月/人）')) else 0.0,
                            '奖励/补助（元/月）': float(row.get('奖励/补助（元/月）', 0)) if pd.notna(row.get('奖励/补助（元/月）')) else 0.0,
                            '餐补（元/月/人）': float(row.get('餐补（元/月/人）', 0)) if pd.notna(row.get('餐补（元/月/人）')) else 0.0,
                            '年终奖（元/人）': float(row.get('年终奖（元/人）', 0)) if pd.notna(row.get('年终奖（元/人）')) else 0.0,
                            '养老保险费（元/月/人）': float(row.get('养老保险费（元/月/人）', 0)) if pd.notna(row.get('养老保险费（元/月/人）')) else 0.0,
                            '失业保险费（元/月/人）': float(row.get('失业保险费（元/月/人）', 0)) if pd.notna(row.get('失业保险费（元/月/人）')) else 0.0,
                            '医疗/生育保险费（元/月/人）': float(row.get('医疗/生育保险费（元/月/人）', 0)) if pd.notna(row.get('医疗/生育保险费（元/月/人）')) else 0.0,
                            '工伤保险费（元/月/人）': float(row.get('工伤保险费（元/月/人）', 0)) if pd.notna(row.get('工伤保险费（元/月/人）')) else 0.0,
                            '住房公积金（元/月/人）': float(row.get('住房公积金（元/月/人）', 0)) if pd.notna(row.get('住房公积金（元/月/人）')) else 0.0,
                        }
                
                # 定义提成类型到岗位的映射
                commission_to_position = {
                    '品管提成': '品管',
                    '物流主管提成': '物流主管',
                    '生产主管提成': '生产/安全环保主管',
                    '物流卸货提成': '物流卸货人员',
                    '维修班长提成': '机修班长',
                    '维修员提成': '机修工',
                    '冰箱维修主管提成': '冰箱维修主管',
                    '叉车司磅库管等提成': '叉车/司磅/库管等'
                }
                
                # 处理非班组长的提成类型
                for commission_type, position in commission_to_position.items():
                    if position not in salary_dict:
                        print(f"[间接人工成本] 未找到岗位 {position} 的薪酬数据")
                        continue
                    
                    total_commission_cost = category_totals.get(commission_type, 0.0)
                    if total_commission_cost == 0:
                        category_fixed_costs[commission_type] = 0.0
                        continue
                    
                    salary_data = salary_dict[position]
                    personnel_base = salary_data['人员基础配置']
                    
                    # 计算月均固定成本
                    monthly_fixed_cost_per_person = (
                        salary_data['平均工资（元/月/人）'] +
                        salary_data['奖励/补助（元/月）'] +
                        salary_data['餐补（元/月/人）'] +
                        salary_data['年终奖（元/人）'] +
                        salary_data['养老保险费（元/月/人）'] +
                        salary_data['失业保险费（元/月/人）'] +
                        salary_data['医疗/生育保险费（元/月/人）'] +
                        salary_data['工伤保险费（元/月/人）'] +
                        salary_data['住房公积金（元/月/人）']
                    )
                    
                    # 计算分摊固定成本（按提成成本比例分摊）
                    # 公式：物料分摊固定成本 = (物料提成成本 / 总提成成本) × 人员基础配置 × 月均固定成本 × 预测期数
                    category_fixed_cost_sum = 0.0
                    commission_cost_field = commission_type.replace('提成', '提成成本')
                    
                    # 为每个物料明细计算并添加分摊固定成本
                    for detail in part1_details + part2_details + part3_details:
                        if commission_cost_field in detail:
                            item_commission_cost = detail.get(commission_cost_field, 0.0)
                            if total_commission_cost > 0:
                                allocation_ratio = item_commission_cost / total_commission_cost
                            else:
                                allocation_ratio = 0.0
                            item_fixed_cost = allocation_ratio * personnel_base * monthly_fixed_cost_per_person * prediction_period
                            fixed_cost_field = commission_type.replace('提成', '分摊固定成本')
                            detail[fixed_cost_field] = float(item_fixed_cost)
                            category_fixed_cost_sum += item_fixed_cost
                    
                    category_fixed_costs[commission_type] = float(category_fixed_cost_sum)
                    total_fixed_cost += category_fixed_cost_sum
                
                # 处理班组长提成（需要分类处理）
                team_leader_mapping = {
                    '白电小组长': {
                        'position': '白电小组长',
                        'filter': lambda detail: detail.get('类别') == '旧机' and (
                            '空调' in str(detail.get('物料名称', '')).upper() or 
                            '洗衣机' in str(detail.get('物料名称', '')).upper()
                        )
                    },
                    '生产班组长(黑电)': {
                        'position': '生产班组长(黑电)',
                        'filter': lambda detail: detail.get('类别') == '旧机' and (
                            '电视' in str(detail.get('物料名称', '')).upper() or 
                            '电脑' in str(detail.get('物料名称', '')).upper()
                        )
                    },
                    '生产班组长(冰箱)': {
                        'position': '生产班组长(冰箱)',
                        'filter': lambda detail: detail.get('类别') == '旧机' and (
                            '冰箱' in str(detail.get('物料名称', '')).upper()
                        )
                    },
                    '生产班组长(塑料破碎)': {
                        'position': '生产班组长(塑料破碎)',
                        'filter': lambda detail: detail.get('类别') in ['一次拆解产物', '一破']
                    }
                }
                
                for leader_type, config in team_leader_mapping.items():
                    position = config['position']
                    filter_func = config['filter']
                    
                    if position not in salary_dict:
                        print(f"[间接人工成本] 未找到岗位 {position} 的薪酬数据")
                        category_fixed_costs[f'{leader_type}提成'] = 0.0
                        continue
                    
                    # 筛选出该班组长类型的提成明细
                    leader_details = []
                    leader_total_commission = 0.0
                    tl_no_opening_field = '班组长提成成本(不考虑期初库存和库存结余)'
                    is_plastic_leader = leader_type == '生产班组长(塑料破碎)'
                    leader_details_no_opening = []
                    leader_total_no_opening = 0.0

                    for detail in part1_details + part2_details + part3_details:
                        if filter_func(detail):
                            commission_cost = float(detail.get('班组长提成成本', 0.0) or 0)
                            if commission_cost > 0:
                                leader_details.append(detail)
                                leader_total_commission += commission_cost
                            if include_no_opening_columns and is_plastic_leader:
                                tl_no = float(detail.get(tl_no_opening_field, 0) or 0)
                                if tl_no > 0:
                                    leader_details_no_opening.append(detail)
                                    leader_total_no_opening += tl_no

                    if leader_total_commission == 0:
                        if not (
                            include_no_opening_columns
                            and is_plastic_leader
                            and leader_total_no_opening > 0
                        ):
                            category_fixed_costs[f'{leader_type}提成'] = 0.0
                            continue

                    salary_data = salary_dict[position]
                    personnel_base = salary_data['人员基础配置']

                    # 计算月均固定成本
                    monthly_fixed_cost_per_person = (
                        salary_data['平均工资（元/月/人）'] +
                        salary_data['奖励/补助（元/月）'] +
                        salary_data['餐补（元/月/人）'] +
                        salary_data['年终奖（元/人）'] +
                        salary_data['养老保险费（元/月/人）'] +
                        salary_data['失业保险费（元/月/人）'] +
                        salary_data['医疗/生育保险费（元/月/人）'] +
                        salary_data['工伤保险费（元/月/人）'] +
                        salary_data['住房公积金（元/月/人）']
                    )

                    # 计算该班组长的分摊固定成本（按提成成本比例分摊）
                    # 公式：物料分摊固定成本 = (物料提成成本 / 总提成成本) × 人员基础配置 × 月均固定成本 × 预测期数
                    leader_fixed_cost_sum = 0.0
                    pool_leader = personnel_base * monthly_fixed_cost_per_person * prediction_period

                    # 为每个物料明细计算并添加分摊固定成本（考虑库存口径）
                    if leader_total_commission > 0:
                        for detail in leader_details:
                            item_commission_cost = float(detail.get('班组长提成成本', 0.0) or 0)
                            allocation_ratio = item_commission_cost / leader_total_commission
                            item_fixed_cost = allocation_ratio * pool_leader
                            fixed_cost_field = f'{leader_type}分摊固定成本'
                            detail[fixed_cost_field] = float(item_fixed_cost)
                            leader_fixed_cost_sum += item_fixed_cost

                    if include_no_opening_columns and is_plastic_leader:
                        fc_plastic_no = '生产班组长(塑料破碎)分摊固定成本(不考虑期初库存和库存结余)'
                        if leader_total_no_opening > 0:
                            for detail in leader_details_no_opening:
                                item_tl_no = float(detail.get(tl_no_opening_field, 0) or 0)
                                detail[fc_plastic_no] = float(
                                    (item_tl_no / leader_total_no_opening) * pool_leader
                                )

                    category_fixed_costs[f'{leader_type}提成'] = float(leader_fixed_cost_sum)
                    total_fixed_cost += leader_fixed_cost_sum
                        
        except Exception as e:
            print(f"[间接人工成本分摊计算] 计算失败: {str(e)}")
            import traceback
            traceback.print_exc()
        
        if include_no_opening_columns:
            _fc_plastic_no_key = '生产班组长(塑料破碎)分摊固定成本(不考虑期初库存和库存结余)'
            for _detail in part1_details + part2_details + part3_details:
                if _fc_plastic_no_key not in _detail:
                    if _detail.get('类别') in ('一次拆解产物', '一破'):
                        _detail[_fc_plastic_no_key] = 0.0
                    else:
                        _detail[_fc_plastic_no_key] = float(
                            _detail.get('生产班组长(塑料破碎)分摊固定成本', 0) or 0
                        )
        
        # 计算其他岗位成本
        other_positions_cost = 0.0
        other_positions_details = []
        
        # 定义需要计算的其他岗位列表（19个岗位）
        other_positions_list = [
            '旧机回收',
            '生产工艺管理主管',
            '客服专员',
            '设备维修副经理',
            '生产管理副高级经理',
            '质量管理经理',
            '库房',
            '电废综合管理人员',
            '销售组人员',
            '基金管理项目组',
            '视频监控员',
            '库管（兼叉车）',  # 注意：数据中是全角括号
            '计数员兼库房协管',
            '原材料库房管理员',
            '数据统计',
            '黑电保洁',
            '白电保洁',
            '物流保洁',
            '设备维护储备岗'
        ]
        
        try:
            # 获取薪酬核算基础数据
            salary_df = get_salary_accounting_dataframe()
            if salary_df is not None and not salary_df.empty:
                # 创建岗位到薪酬数据的映射
                salary_dict = {}
                for _, row in salary_df.iterrows():
                    position = str(row.get('岗位', ''))
                    if position:
                        salary_dict[position] = {
                            '人员基础配置': float(row.get('人员基础配置', 0)) if pd.notna(row.get('人员基础配置')) else 0.0,
                            '平均工资（元/月/人）': float(row.get('平均工资（元/月/人）', 0)) if pd.notna(row.get('平均工资（元/月/人）')) else 0.0,
                            '奖励/补助（元/月）': float(row.get('奖励/补助（元/月）', 0)) if pd.notna(row.get('奖励/补助（元/月）')) else 0.0,
                            '餐补（元/月/人）': float(row.get('餐补（元/月/人）', 0)) if pd.notna(row.get('餐补（元/月/人）')) else 0.0,
                            '年终奖（元/人）': float(row.get('年终奖（元/人）', 0)) if pd.notna(row.get('年终奖（元/人）')) else 0.0,
                            '养老保险费（元/月/人）': float(row.get('养老保险费（元/月/人）', 0)) if pd.notna(row.get('养老保险费（元/月/人）')) else 0.0,
                            '失业保险费（元/月/人）': float(row.get('失业保险费（元/月/人）', 0)) if pd.notna(row.get('失业保险费（元/月/人）')) else 0.0,
                            '医疗/生育保险费（元/月/人）': float(row.get('医疗/生育保险费（元/月/人）', 0)) if pd.notna(row.get('医疗/生育保险费（元/月/人）')) else 0.0,
                            '工伤保险费（元/月/人）': float(row.get('工伤保险费（元/月/人）', 0)) if pd.notna(row.get('工伤保险费（元/月/人）')) else 0.0,
                            '住房公积金（元/月/人）': float(row.get('住房公积金（元/月/人）', 0)) if pd.notna(row.get('住房公积金（元/月/人）')) else 0.0,
                        }
                
                # 计算每个其他岗位的成本
                for position in other_positions_list:
                    if position not in salary_dict:
                        print(f"[间接人工成本-其他岗位] 未找到岗位 {position} 的薪酬数据")
                        continue
                    
                    salary_data = salary_dict[position]
                    personnel_base = salary_data['人员基础配置']
                    
                    # 计算月均固定成本（各项费用之和）
                    monthly_fixed_cost_per_person = (
                        salary_data['平均工资（元/月/人）'] +
                        salary_data['奖励/补助（元/月）'] +
                        salary_data['餐补（元/月/人）'] +
                        salary_data['年终奖（元/人）'] +
                        salary_data['养老保险费（元/月/人）'] +
                        salary_data['失业保险费（元/月/人）'] +
                        salary_data['医疗/生育保险费（元/月/人）'] +
                        salary_data['工伤保险费（元/月/人）'] +
                        salary_data['住房公积金（元/月/人）']
                    )
                    
                    # 计算岗位总成本 = 人员基础配置 × 月均固定成本 × 预测期数
                    position_cost = personnel_base * monthly_fixed_cost_per_person * prediction_period
                    
                    other_positions_details.append({
                        '岗位': position,
                        '人员基础配置': float(personnel_base),
                        '平均工资（元/月/人）': float(salary_data['平均工资（元/月/人）']),
                        '奖励/补助（元/月）': float(salary_data['奖励/补助（元/月）']),
                        '餐补（元/月/人）': float(salary_data['餐补（元/月/人）']),
                        '年终奖（元/人）': float(salary_data['年终奖（元/人）']),
                        '养老保险费（元/月/人）': float(salary_data['养老保险费（元/月/人）']),
                        '失业保险费（元/月/人）': float(salary_data['失业保险费（元/月/人）']),
                        '医疗/生育保险费（元/月/人）': float(salary_data['医疗/生育保险费（元/月/人）']),
                        '工伤保险费（元/月/人）': float(salary_data['工伤保险费（元/月/人）']),
                        '住房公积金（元/月/人）': float(salary_data['住房公积金（元/月/人）']),
                        '月均固定成本（元/月/人）': float(monthly_fixed_cost_per_person),
                        '岗位成本': float(position_cost)
                    })
                    
                    other_positions_cost += position_cost
                    
        except Exception as e:
            print(f"[间接人工成本-其他岗位] 计算失败: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # 计算总间接人工成本（提成成本 + 分摊固定成本 + 其他岗位成本）
        indirect_labor_cost = total_cost + total_fixed_cost + other_positions_cost
        
        return {
            'total_cost': float(total_cost),
            'part1_cost': float(part1_cost),
            'part2_cost': float(part2_cost),
            'part3_cost': float(part3_cost),
            'part1_details': part1_details,
            'part2_details': part2_details,
            'part3_details': part3_details,
            'part1_count': len(part1_details),
            'part2_count': len(part2_details),
            'part3_count': len(part3_details),
            'category_totals': category_totals,
            'total_fixed_cost': float(total_fixed_cost),
            'indirect_labor_cost': float(indirect_labor_cost),
            'category_fixed_costs': category_fixed_costs,
            'other_positions_cost': float(other_positions_cost),
            'other_positions_details': other_positions_details,
            'prediction_period': int(prediction_period)
        }
        
    except Exception as e:
        print(f"[间接人工成本计算] 计算失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'total_cost': 0.0,
            'part1_cost': 0.0,
            'part2_cost': 0.0,
            'part3_cost': 0.0,
            'part1_details': [],
            'part2_details': [],
            'part3_details': [],
            'part1_count': 0,
            'part2_count': 0,
            'part3_count': 0,
            'category_totals': {
                '品管提成': 0.0,
                '物流主管提成': 0.0,
                '物流卸货提成': 0.0,
                '班组长提成': 0.0,
                '生产主管提成': 0.0,
                '维修班长提成': 0.0,
                '维修员提成': 0.0,
                '冰箱维修主管提成': 0.0,
                '叉车司磅库管等提成': 0.0
            },
            'total_fixed_cost': 0.0,
            'indirect_labor_cost': 0.0,
            'category_fixed_costs': {},
            'other_positions_cost': 0.0,
            'other_positions_details': [],
            'prediction_period': int(prediction_period),
            'error': str(e)
        }


@cost_forecast_bp.route('/piece-rate-wage', methods=['GET'])
def get_piece_rate_wage():
    """获取直接人工成本数据（包含计件工资和分摊固定成本）"""
    try:
        app_data = get_session_data_manager()
        
        # 获取预测期数参数
        prediction_period = request.args.get('prediction_period', 1, type=int)
        if prediction_period <= 0:
            prediction_period = 1
        
        # 检查数据是否已被清除
        data_cleared = app_data.get_data('__data_cleared__')
        if data_cleared:
            # 数据已被清除，直接返回空数据
            return jsonify({
                'success': True,
                'data': {
                    'total_wage': 0.0,
                    'part1_wage': 0.0,
                    'part2_wage': 0.0,
                    'part3_wage': 0.0,
                    'part1_details': [],
                    'part2_details': [],
                    'part3_details': [],
                    'part1_count': 0,
                    'part2_count': 0,
                    'part3_count': 0,
                    'total_fixed_cost': 0.0,
                    'direct_labor_cost': 0.0,
                    'prediction_period': prediction_period,
                    'category_details': {}
                }
            })
        
        # 计算直接人工成本（包含计件工资和分摊固定成本）
        result = calculate_direct_labor_cost(app_data, prediction_period)
        
        return jsonify({
            'success': True,
            'data': result
        })
        
    except Exception as e:
        print(f"获取直接人工成本失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@cost_forecast_bp.route('/indirect-labor-cost', methods=['GET'])
def get_indirect_labor_cost():
    """获取间接人工成本数据（包含提成成本和分摊固定成本）"""
    try:
        app_data = get_session_data_manager()
        
        # 获取预测期数参数
        prediction_period = request.args.get('prediction_period', 1, type=int)
        if prediction_period <= 0:
            prediction_period = 1
        
        # 检查数据是否已被清除
        data_cleared = app_data.get_data('__data_cleared__')
        if data_cleared:
            # 数据已被清除，直接返回空数据
            return jsonify({
                'success': True,
                'data': {
                    'total_cost': 0.0,
                    'part1_cost': 0.0,
                    'part2_cost': 0.0,
                    'part3_cost': 0.0,
                    'part1_details': [],
                    'part2_details': [],
                    'part3_details': [],
                    'part1_count': 0,
                    'part2_count': 0,
                    'part3_count': 0,
                    'category_totals': {
                        '品管提成': 0.0,
                        '物流主管提成': 0.0,
                        '物流卸货提成': 0.0,
                        '班组长提成': 0.0,
                        '生产主管提成': 0.0,
                        '维修班长提成': 0.0,
                        '维修员提成': 0.0,
                        '冰箱维修主管提成': 0.0,
                        '叉车司磅库管等提成': 0.0
                    },
                    'total_fixed_cost': 0.0,
                    'indirect_labor_cost': 0.0,
                    'category_fixed_costs': {},
                    'other_positions_cost': 0.0,
                    'other_positions_details': [],
                    'prediction_period': int(prediction_period)
                }
            })
        
        # 计算间接人工成本
        result = calculate_indirect_labor_cost(
            app_data,
            prediction_period,
            include_no_opening_columns=INDIRECT_LABOR_PAGE_INCLUDE_NO_OPENING,
        )
        result['manufacturing_summary_by_category'] = (
            summarize_indirect_labor_manufacturing_stats_by_category(result)
        )
        result['manufacturing_summary_totals'] = (
            summarize_indirect_labor_manufacturing_by_category(result)
        )

        return jsonify({
            'success': True,
            'data': result
        })
        
    except Exception as e:
        print(f"获取间接人工成本失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@cost_forecast_bp.route('/material-cost/export', methods=['GET'])
def export_material_cost():
    """导出拆解物原料成本数据到Excel"""
    try:
        app_data = get_session_data_manager()
        
        # 获取手工数据
        manual_data = app_data.get_data('extracted_data_manual')
        
        if manual_data is None or manual_data.empty:
            return jsonify({
                'success': False,
                'error': '没有可导出的拆解物原料成本数据'
            }), 400
        
        # 计算成本数据
        cost_data = calculate_material_cost(manual_data)
        
        # 只导出旧机类别
        if '类别' in cost_data.columns:
            export_data = cost_data[cost_data['类别'] == '旧机'].copy()
        else:
            export_data = cost_data
        
        if export_data.empty:
            return jsonify({
                'success': False,
                'error': '没有可导出的拆解物原料成本数据'
            }), 400
        
        # 选择要导出的列（排除一些内部列）
        export_columns = [
            '序号', '物料代码', '物料描述', '初始数据', '本期计划采购数量',
            '计划采购单价', '非限制使用的库存', '单位投料成本', '拆解物原料成本'
        ]
        
        # 只保留存在的列
        available_columns = [col for col in export_columns if col in export_data.columns]
        export_df = export_data[available_columns].copy()
        
        # 重命名"非限制使用的库存"为"本期实际投产数量"
        if '非限制使用的库存' in export_df.columns:
            export_df = export_df.rename(columns={'非限制使用的库存': '本期实际投产数量'})
        
        # 追加合计行：单价/单位成本等比率列显示 "-"，其余数值列求和
        _sum_cols_mc = {'初始数据', '本期计划采购数量', '本期实际投产数量', '拆解物原料成本'}
        _dash_cols_mc = {'计划采购单价', '单位投料成本'}
        _total_row_mc = {c: '' for c in export_df.columns}
        _label_placed_mc = False
        for _c in export_df.columns:
            if _c == '序号':
                _total_row_mc[_c] = '合计'
                _label_placed_mc = True
            elif _c in _sum_cols_mc:
                _s = pd.to_numeric(export_df[_c], errors='coerce').fillna(0).sum()
                _total_row_mc[_c] = round(float(_s), 2)
            elif _c in _dash_cols_mc:
                _total_row_mc[_c] = '-'
            else:
                if not _label_placed_mc:
                    _total_row_mc[_c] = '合计'
                    _label_placed_mc = True
                else:
                    _total_row_mc[_c] = ''
        export_df = pd.concat([export_df, pd.DataFrame([_total_row_mc])], ignore_index=True)
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 写入数据
            export_df.to_excel(writer, sheet_name='拆解物原料成本', index=False)
            
            # 设置列宽和样式
            worksheet = writer.sheets['拆解物原料成本']
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF", name="仿宋")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            for col in range(1, len(export_df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                # 设置列宽
                col_letter = worksheet.cell(row=1, column=col).column_letter
                if col == 1:  # 序号
                    worksheet.column_dimensions[col_letter].width = 10
                elif col == 2:  # 物料代码
                    worksheet.column_dimensions[col_letter].width = 18
                elif col == 3:  # 物料描述
                    worksheet.column_dimensions[col_letter].width = 30
                else:
                    worksheet.column_dimensions[col_letter].width = 18
            
            # 设置数据行样式
            data_font = Font(name="仿宋")
            total_row_idx_mc = len(export_df) + 1
            total_fill_mc = PatternFill(start_color="E8E6DC", end_color="E8E6DC", fill_type="solid")
            total_font_mc = Font(name="仿宋", bold=True)
            for row in range(2, len(export_df) + 2):
                is_total_row = (row == total_row_idx_mc)
                for col in range(1, len(export_df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = data_font
                    cell.alignment = center_alignment
                    if is_total_row:
                        cell.fill = total_fill_mc
                        cell.font = total_font_mc
        
        output.seek(0)
        filename = f'拆解物原料成本_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"导出拆解物原料成本失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@cost_forecast_bp.route('/piece-rate-wage/export', methods=['GET'])
def export_piece_rate_wage():
    """导出直接人工成本数据到Excel"""
    try:
        app_data = get_session_data_manager()
        
        # 获取预测期数参数
        prediction_period = request.args.get('prediction_period', 1, type=int)
        if prediction_period <= 0:
            prediction_period = 1
        
        # 计算直接人工成本
        result = calculate_direct_labor_cost(app_data, prediction_period)
        
        if not result or (result.get('total_wage', 0) == 0 and result.get('total_fixed_cost', 0) == 0):
            return jsonify({
                'success': False,
                'error': '没有可导出的直接人工成本数据'
            }), 400
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # 定义样式
            header_font = Font(bold=True, color="FFFFFF", name="仿宋")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            data_font = Font(name="仿宋")
            
            # 1. 直接人工成本汇总统计表
            summary_data = {
                '统计项目': [
                    '计件总工资',
                    '旧机类别工资',
                    '一次拆解产物工资',
                    '深加工工资（一破/打包铁/屏）',
                    '分摊固定成本（白电）',
                    '分摊固定成本（黑电）',
                    '分摊固定成本（冰箱）',
                    '分摊固定成本（金属打包）',
                    '分摊固定成本（塑料）',
                    '分摊固定成本（屏）',
                    '总分摊固定成本',
                    '直接人工成本'
                ],
                '金额(元)': [
                    result.get('total_wage', 0),
                    result.get('part1_wage', 0),
                    result.get('part2_wage', 0),
                    result.get('part3_wage', 0),
                    result.get('category_details', {}).get('白电', {}).get('total_fixed_cost', 0),
                    result.get('category_details', {}).get('黑电', {}).get('total_fixed_cost', 0),
                    result.get('category_details', {}).get('冰箱', {}).get('total_fixed_cost', 0),
                    result.get('category_details', {}).get('金属打包', {}).get('total_fixed_cost', 0),
                    result.get('category_details', {}).get('塑料', {}).get('total_fixed_cost', 0),
                    result.get('category_details', {}).get('屏', {}).get('total_fixed_cost', 0),
                    result.get('total_fixed_cost', 0),
                    result.get('direct_labor_cost', 0)
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='汇总统计', index=False)
            
            # 设置汇总统计表样式
            summary_ws = writer.sheets['汇总统计']
            for col in range(1, len(summary_df.columns) + 1):
                cell = summary_ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                summary_ws.column_dimensions[get_column_letter(col)].width = 25
            
            for row in range(2, len(summary_df) + 2):
                for col in range(1, len(summary_df.columns) + 1):
                    cell = summary_ws.cell(row=row, column=col)
                    cell.font = data_font
                    cell.alignment = center_alignment
            
            # 2. 详细数据表（合并所有部分）
            all_details = []
            
            # 第一部分：旧机类别
            if result.get('part1_details'):
                for item in result['part1_details']:
                    all_details.append({
                        '序号': len(all_details) + 1,
                        '原物料代码': item.get('原物料代码', '') or item.get('物料代码', ''),
                        '原物料名称': item.get('原物料名称', '') or item.get('物料名称', ''),
                        '物料/产物编码': item.get('物料代码', ''),
                        '物料/产物名称': item.get('物料名称', ''),
                        '数量/重量': item.get('数量', 0),
                        '数量/重量(不考虑期初库存和库存结余)': item.get('数量(不考虑期初库存和库存结余)', item.get('数量', 0)),
                        '类别': '旧机',
                        '单价': item.get('单价', 0),
                        '工资': item.get('工资', 0),
                        '工资(不考虑期初库存和库存结余)': item.get('工资(不考虑期初库存和库存结余)', item.get('工资', 0))
                    })
            
            # 第二部分：一次拆解产物
            if result.get('part2_details'):
                for item in result['part2_details']:
                    all_details.append({
                        '序号': len(all_details) + 1,
                        '原物料代码': item.get('原物料代码', ''),
                        '原物料名称': item.get('原物料名称', ''),
                        '物料/产物编码': item.get('拆解产物编码', ''),
                        '物料/产物名称': item.get('拆解产物名称', ''),
                        '数量/重量': item.get('计算结果(KG)', 0),
                        '数量/重量(不考虑期初库存和库存结余)': item.get('数量(不考虑期初库存和库存结余)', item.get('计算结果(KG)', 0)),
                        '类别': '一次拆解产物',
                        '单价': item.get('单价', 0),
                        '工资': item.get('工资', 0),
                        '工资(不考虑期初库存和库存结余)': item.get('工资(不考虑期初库存和库存结余)', item.get('工资', 0))
                    })
            
            # 第三部分：深加工
            if result.get('part3_details'):
                for item in result['part3_details']:
                    all_details.append({
                        '序号': len(all_details) + 1,
                        '原物料代码': item.get('原物料代码', ''),
                        '原物料名称': item.get('原物料名称', ''),
                        '物料/产物编码': item.get('深加工产物编码', ''),
                        '物料/产物名称': item.get('深加工产物名称', ''),
                        '数量/重量': item.get('深加工结果(KG)', 0),
                        '数量/重量(不考虑期初库存和库存结余)': item.get('数量(不考虑期初库存和库存结余)', item.get('深加工结果(KG)', 0)),
                        '类别': item.get('类别', ''),
                        '单价': item.get('单价', 0),
                        '工资': item.get('工资', 0),
                        '工资(不考虑期初库存和库存结余)': item.get('工资(不考虑期初库存和库存结余)', item.get('工资', 0))
                    })
            
            if all_details:
                details_df = pd.DataFrame(all_details)
                details_df.to_excel(writer, sheet_name='计件工资', index=False)
                
                # 设置详细数据表样式
                details_ws = writer.sheets['计件工资']
                for col in range(1, len(details_df.columns) + 1):
                    cell = details_ws.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    col_letter = get_column_letter(col)
                    if col == 1:  # 序号
                        details_ws.column_dimensions[col_letter].width = 10
                    elif col in [2, 4]:  # 原物料代码、物料/产物编码
                        details_ws.column_dimensions[col_letter].width = 18
                    elif col in [3, 5]:  # 原物料名称、物料/产物名称
                        details_ws.column_dimensions[col_letter].width = 30
                    else:
                        details_ws.column_dimensions[col_letter].width = 18
                
                for row in range(2, len(details_df) + 2):
                    for col in range(1, len(details_df.columns) + 1):
                        cell = details_ws.cell(row=row, column=col)
                        cell.font = data_font
                        cell.alignment = center_alignment
            
            # 3. 分摊固定成本明细表
            fixed_cost_details = []
            category_details = result.get('category_details', {})
            category_order = ['白电', '黑电', '冰箱', '金属打包', '塑料', '屏']
            index = 1
            
            for category in category_order:
                category_data = category_details.get(category)
                if not category_data or not category_data.get('item_allocations'):
                    continue
                
                category_wage_sum = 0
                category_wage_sum_no_opening = 0
                category_fixed_cost_sum = 0
                category_fixed_cost_sum_no_opening = 0
                
                for allocation in category_data['item_allocations']:
                    item = allocation.get('item', {})
                    
                    # 提取物料信息
                    origin_code = ''
                    origin_name = ''
                    code = ''
                    name = ''
                    wage = 0
                    wage_no_opening = 0
                    
                    if item.get('物料代码'):
                        code = item.get('物料代码', '')
                        name = item.get('物料名称', '')
                        origin_code = item.get('原物料代码', '') or code
                        origin_name = item.get('原物料名称', '') or name
                        wage = item.get('工资', 0)
                        wage_no_opening = item.get('工资(不考虑期初库存和库存结余)', wage)
                    elif item.get('拆解产物编码'):
                        code = item.get('拆解产物编码', '')
                        name = item.get('拆解产物名称', '')
                        origin_code = item.get('原物料代码', '')
                        origin_name = item.get('原物料名称', '')
                        wage = item.get('工资', 0)
                        wage_no_opening = item.get('工资(不考虑期初库存和库存结余)', wage)
                    elif item.get('深加工产物编码'):
                        code = item.get('深加工产物编码', '')
                        name = item.get('深加工产物名称', '')
                        origin_code = item.get('原物料代码', '')
                        origin_name = item.get('原物料名称', '')
                        wage = item.get('工资', 0)
                        wage_no_opening = item.get('工资(不考虑期初库存和库存结余)', wage)
                    
                    category_wage_sum += wage
                    category_wage_sum_no_opening += (wage_no_opening or 0)
                    category_fixed_cost_sum += allocation.get('fixed_cost', 0)
                    category_fixed_cost_sum_no_opening += allocation.get('fixed_cost_no_opening', allocation.get('fixed_cost', 0))
                    
                    fixed_cost_details.append({
                        '序号': index,
                        '原物料代码': origin_code,
                        '原物料名称': origin_name,
                        '类别': category,
                        '物料/产物编码': code,
                        '物料/产物名称': name,
                        '计件工资（元）': wage,
                        '计件工资（不考虑期初库存和库存结余）（元）': wage_no_opening,
                        '分摊比例': f"{allocation.get('allocation_ratio', 0) * 100:.2f}%",
                        '分摊比例（不考虑期初库存和库存结余）': f"{allocation.get('allocation_ratio_no_opening', allocation.get('allocation_ratio', 0)) * 100:.2f}%",
                        '分摊固定成本（元）': allocation.get('fixed_cost', 0),
                        '分摊固定成本（不考虑期初库存和库存结余）（元）': allocation.get('fixed_cost_no_opening', allocation.get('fixed_cost', 0))
                    })
                    index += 1
                
                # 添加类别小计行
                if category_data['item_allocations']:
                    fixed_cost_details.append({
                        '序号': '',
                        '原物料代码': '',
                        '原物料名称': '',
                        '类别': category,
                        '物料/产物编码': '',
                        '物料/产物名称': f'{category}小计',
                        '计件工资（元）': category_wage_sum,
                        '计件工资（不考虑期初库存和库存结余）（元）': category_wage_sum_no_opening,
                        '分摊比例': '',
                        '分摊比例（不考虑期初库存和库存结余）': '',
                        '分摊固定成本（元）': category_fixed_cost_sum,
                        '分摊固定成本（不考虑期初库存和库存结余）（元）': category_fixed_cost_sum_no_opening
                    })
            
            if fixed_cost_details:
                fixed_cost_df = pd.DataFrame(fixed_cost_details)
                fixed_cost_df.to_excel(writer, sheet_name='分摊固定成本明细', index=False)
                
                # 设置分摊固定成本明细表样式
                fixed_cost_ws = writer.sheets['分摊固定成本明细']
                for col in range(1, len(fixed_cost_df.columns) + 1):
                    cell = fixed_cost_ws.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    col_letter = get_column_letter(col)
                    if col == 1:  # 序号
                        fixed_cost_ws.column_dimensions[col_letter].width = 10
                    elif col == 4:  # 类别
                        fixed_cost_ws.column_dimensions[col_letter].width = 12
                    elif col in [2, 5]:  # 原物料代码、物料/产物编码
                        fixed_cost_ws.column_dimensions[col_letter].width = 18
                    elif col in [3, 6]:  # 原物料名称、物料/产物名称
                        fixed_cost_ws.column_dimensions[col_letter].width = 30
                    else:
                        fixed_cost_ws.column_dimensions[col_letter].width = 18
                
                # 设置数据行样式，小计行特殊处理
                for row in range(2, len(fixed_cost_df) + 2):
                    # 小计行：物料/产物名称列（第6列）包含“小计”
                    is_subtotal = fixed_cost_ws.cell(row=row, column=6).value and '小计' in str(fixed_cost_ws.cell(row=row, column=6).value)
                    
                    for col in range(1, len(fixed_cost_df.columns) + 1):
                        cell = fixed_cost_ws.cell(row=row, column=col)
                        cell.font = data_font
                        cell.alignment = center_alignment
                        
                        # 小计行特殊样式
                        if is_subtotal:
                            cell.font = Font(bold=True, name="仿宋")
                            if col == 6:  # 名称列
                                cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                            elif col in [7, 8, 11, 12]:  # 工资和分摊成本列（考虑/不考虑）
                                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                        # 分摊固定成本列高亮（考虑/不考虑）
                        elif col in [11, 12]:
                            cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                            cell.font = Font(bold=True, name="仿宋", color="2E7D32")
            
            # 4. 分类统计表（按四机一脑分类）——考虑口径与页面一致，使用展示统计
            product_category_stats = result.get('product_category_stats_display') or result.get('product_category_stats', {})
            product_category_stats_no_opening = result.get('product_category_stats_no_opening', {})
            category_statistics = []
            
            categories_order = ['电视', '电脑', '冰箱', '空调', '洗衣机']
            total_category_wage = 0.0
            total_category_fixed_cost = 0.0
            total_category_fixed_cost_no_opening = 0.0
            
            for category in categories_order:
                stats = product_category_stats.get(category, {'wage': 0.0, 'fixed_cost': 0.0})
                wage = stats.get('wage', 0.0)
                fixed_cost = stats.get('fixed_cost', 0.0)
                total = wage + fixed_cost
                stats_no_opening = product_category_stats_no_opening.get(category, {'wage': wage, 'fixed_cost': fixed_cost})
                wage_no_opening = stats_no_opening.get('wage', wage)
                fixed_cost_no_opening_row = stats_no_opening.get('fixed_cost', fixed_cost)
                total_no_opening = wage_no_opening + fixed_cost_no_opening_row
                
                category_statistics.append({
                    '分类': category,
                    '生产工人计件工资（元）': wage,
                    '生产工人计件工资（不考虑期初库存和库存结余）（元）': wage_no_opening,
                    '分摊固定成本（元）': fixed_cost,
                    '分摊固定成本（不考虑期初库存和库存结余）（元）': fixed_cost_no_opening_row,
                    '合计（元）': total,
                    '合计（不考虑期初库存和库存结余）（元）': total_no_opening
                })
                
                total_category_wage += wage
                total_category_fixed_cost += fixed_cost
                total_category_fixed_cost_no_opening += fixed_cost_no_opening_row
            
            # 添加汇总行
            category_statistics.append({
                '分类': '汇总',
                '生产工人计件工资（元）': total_category_wage,
                '生产工人计件工资（不考虑期初库存和库存结余）（元）': sum(
                    (product_category_stats_no_opening.get(cat, {}).get('wage', product_category_stats.get(cat, {}).get('wage', 0.0)) or 0.0)
                    for cat in categories_order
                ),
                '分摊固定成本（元）': total_category_fixed_cost,
                '分摊固定成本（不考虑期初库存和库存结余）（元）': total_category_fixed_cost_no_opening,
                '合计（元）': total_category_wage + total_category_fixed_cost,
                '合计（不考虑期初库存和库存结余）（元）': sum(
                    (product_category_stats_no_opening.get(cat, {}).get('wage', product_category_stats.get(cat, {}).get('wage', 0.0)) or 0.0)
                    + (product_category_stats_no_opening.get(cat, {}).get('fixed_cost', product_category_stats.get(cat, {}).get('fixed_cost', 0.0)) or 0.0)
                    for cat in categories_order
                )
            })
            
            if category_statistics:
                category_stats_df = pd.DataFrame(category_statistics)
                category_stats_df.to_excel(writer, sheet_name='分类统计', index=False)
                
                # 设置分类统计表样式
                category_stats_ws = writer.sheets['分类统计']
                for col in range(1, len(category_stats_df.columns) + 1):
                    cell = category_stats_ws.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    col_letter = get_column_letter(col)
                    if col == 1:  # 分类
                        category_stats_ws.column_dimensions[col_letter].width = 15
                    else:
                        category_stats_ws.column_dimensions[col_letter].width = 25
                
                # 设置数据行样式
                for row in range(2, len(category_stats_df) + 2):
                    is_total = category_stats_ws.cell(row=row, column=1).value == '汇总'
                    
                    for col in range(1, len(category_stats_df.columns) + 1):
                        cell = category_stats_ws.cell(row=row, column=col)
                        cell.font = data_font
                        cell.alignment = center_alignment
                        
                        # 汇总行特殊样式
                        if is_total:
                            cell.font = Font(bold=True, name="仿宋")
                            if col == 1:  # 分类列
                                cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                            else:  # 数值列
                                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                        # 数值列高亮
                        elif col > 1:
                            cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                            if col in [6, 7]:  # 合计列（考虑/不考虑）
                                cell.font = Font(bold=True, name="仿宋", color="2E7D32")
        
        output.seek(0)
        filename = f'直接人工成本_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"导出直接人工成本失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@cost_forecast_bp.route('/indirect-labor-cost/export', methods=['GET'])
def export_indirect_labor_cost():
    """导出间接人工成本数据到Excel"""
    try:
        app_data = get_session_data_manager()
        
        # 获取预测期数参数
        prediction_period = request.args.get('prediction_period', 1, type=int)
        if prediction_period <= 0:
            prediction_period = 1
        
        # 计算间接人工成本
        result = calculate_indirect_labor_cost(app_data, prediction_period, include_no_opening_columns=True)
        
        if not result or result.get('total_cost', 0) == 0:
            return jsonify({
                'success': False,
                'error': '没有可导出的间接人工成本数据'
            }), 400
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # 定义样式
            header_font = Font(bold=True, color="FFFFFF", name="仿宋", size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            left_alignment = Alignment(horizontal="left", vertical="center")
            data_font = Font(name="仿宋", size=11)
            title_font = Font(bold=True, name="仿宋", size=14, color="FFFFFF")
            title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            highlight_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            highlight_font = Font(bold=True, name="仿宋", size=11, color="856404")
            total_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            total_font = Font(bold=True, name="仿宋", size=12, color="2E7D32")
            
            # 1. 先构建人工提成明细数据，以便计算总成本
            category_totals = result.get('category_totals', {})
            category_fixed_costs = result.get('category_fixed_costs', {})
            
            # 详细数据表（合并所有部分）
            all_details = []
            
            # 第一部分：旧机类别
            if result.get('part1_details'):
                for item in result['part1_details']:
                    all_details.append({
                        '序号': len(all_details) + 1,
                        '原物料代码': item.get('原物料代码', '') or item.get('物料代码', ''),
                        '原物料名称': item.get('原物料名称', '') or item.get('物料名称', ''),
                        '物料/产物编码': item.get('物料代码', ''),
                        '物料/产物名称': item.get('物料名称', ''),
                        '数量/重量': item.get('数量', 0),
                        '数量/重量(不考虑期初库存和库存结余)': item.get(
                            '数量/重量(不考虑期初库存和库存结余)', item.get('数量', 0)
                        ),
                        '类别': '旧机',
                        '品管提成成本': item.get('品管提成成本', 0),
                        '物流主管提成成本': item.get('物流主管提成成本', 0),
                        '物流卸货提成成本': item.get('物流卸货提成成本', 0),
                        '班组长提成成本': item.get('班组长提成成本', 0),
                        '班组长提成成本(不考虑期初库存和库存结余)': item.get(
                            '班组长提成成本(不考虑期初库存和库存结余)', item.get('班组长提成成本', 0)
                        ),
                        '生产主管提成成本': item.get('生产主管提成成本', 0),
                        '维修班长提成成本': item.get('维修班长提成成本', 0),
                        '维修员提成成本': item.get('维修员提成成本', 0),
                        '冰箱维修主管提成成本': item.get('冰箱维修主管提成成本', 0),
                        '叉车司磅库管等提成成本': item.get('叉车司磅库管等提成成本', 0),
                        '总成本': item.get('总成本', 0)
                    })
            
            # 第二部分：一次拆解产物
            if result.get('part2_details'):
                for item in result['part2_details']:
                    all_details.append({
                        '序号': len(all_details) + 1,
                        '原物料代码': item.get('原物料代码', ''),
                        '原物料名称': item.get('原物料名称', ''),
                        '物料/产物编码': item.get('拆解产物编码', ''),
                        '物料/产物名称': item.get('拆解产物名称', ''),
                        '数量/重量': item.get('计算结果(KG)', 0),
                        '数量/重量(不考虑期初库存和库存结余)': item.get(
                            '数量/重量(不考虑期初库存和库存结余)', item.get('计算结果(KG)', 0)
                        ),
                        '类别': '一次拆解产物',
                        '品管提成成本': item.get('品管提成成本', 0),
                        '物流主管提成成本': item.get('物流主管提成成本', 0),
                        '物流卸货提成成本': item.get('物流卸货提成成本', 0),
                        '班组长提成成本': item.get('班组长提成成本', 0),
                        '班组长提成成本(不考虑期初库存和库存结余)': item.get(
                            '班组长提成成本(不考虑期初库存和库存结余)', item.get('班组长提成成本', 0)
                        ),
                        '生产主管提成成本': item.get('生产主管提成成本', 0),
                        '维修班长提成成本': item.get('维修班长提成成本', 0),
                        '维修员提成成本': item.get('维修员提成成本', 0),
                        '冰箱维修主管提成成本': item.get('冰箱维修主管提成成本', 0),
                        '叉车司磅库管等提成成本': item.get('叉车司磅库管等提成成本', 0),
                        '总成本': item.get('总成本', 0)
                    })
            
            # 第三部分：深加工
            if result.get('part3_details'):
                for item in result['part3_details']:
                    all_details.append({
                        '序号': len(all_details) + 1,
                        '原物料代码': item.get('原物料代码', ''),
                        '原物料名称': item.get('原物料名称', ''),
                        '物料/产物编码': item.get('深加工产物编码', ''),
                        '物料/产物名称': item.get('深加工产物名称', ''),
                        '数量/重量': item.get('深加工结果(KG)', 0),
                        '数量/重量(不考虑期初库存和库存结余)': item.get(
                            '数量/重量(不考虑期初库存和库存结余)', item.get('深加工结果(KG)', 0)
                        ),
                        '类别': item.get('类别', ''),
                        '品管提成成本': item.get('品管提成成本', 0),
                        '物流主管提成成本': item.get('物流主管提成成本', 0),
                        '物流卸货提成成本': item.get('物流卸货提成成本', 0),
                        '班组长提成成本': item.get('班组长提成成本', 0),
                        '班组长提成成本(不考虑期初库存和库存结余)': item.get(
                            '班组长提成成本(不考虑期初库存和库存结余)', item.get('班组长提成成本', 0)
                        ),
                        '生产主管提成成本': item.get('生产主管提成成本', 0),
                        '维修班长提成成本': item.get('维修班长提成成本', 0),
                        '维修员提成成本': item.get('维修员提成成本', 0),
                        '冰箱维修主管提成成本': item.get('冰箱维修主管提成成本', 0),
                        '叉车司磅库管等提成成本': item.get('叉车司磅库管等提成成本', 0),
                        '总成本': item.get('总成本', 0)
                    })
            
            # 计算人工提成明细的总成本
            commission_detail_total = sum(item.get('总成本', 0) for item in all_details)
            
            # 2. 间接人工成本汇总统计表
            summary_rows = []
            
            # 总体汇总
            summary_rows.append({
                '类别': '总体汇总',
                '项目': '间接人工提成汇总',
                '金额(元)': result.get('total_cost', 0)
            })
            summary_rows.append({
                '类别': '',
                '项目': '人工提成成本',
                '金额(元)': commission_detail_total
            })
            summary_rows.append({
                '类别': '',
                '项目': '分摊固定成本',
                '金额(元)': result.get('total_fixed_cost', 0)
            })
            summary_rows.append({
                '类别': '',
                '项目': '其他岗位成本',
                '金额(元)': result.get('other_positions_cost', 0)
            })
            summary_rows.append({
                '类别': '',
                '项目': '间接人工总成本',
                '金额(元)': result.get('indirect_labor_cost', 0)
            })
            
            # 先写入数据到Excel
            summary_df = pd.DataFrame(summary_rows)
            summary_df.to_excel(writer, sheet_name='汇总统计', index=False, startrow=1)
            
            # 设置汇总统计表样式
            summary_ws = writer.sheets['汇总统计']
            
            # 设置列宽
            summary_ws.column_dimensions['A'].width = 20  # 类别
            summary_ws.column_dimensions['B'].width = 35  # 项目
            summary_ws.column_dimensions['C'].width = 20  # 金额
            
            # 添加标题行（预测期数信息）
            title_row = 1
            summary_ws.merge_cells(f'A{title_row}:C{title_row}')
            title_cell = summary_ws.cell(row=title_row, column=1)
            title_cell.value = f'间接人工成本汇总统计（预测期数：{prediction_period}个月）'
            title_cell.font = Font(bold=True, name="仿宋", size=16, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            summary_ws.row_dimensions[title_row].height = 30
            
            # 设置表头（第2行）
            header_row = 2
            for col in range(1, 4):
                cell = summary_ws.cell(row=header_row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            summary_ws.row_dimensions[header_row].height = 25
            
            # 设置数据行样式（从第3行开始，因为startrow=1，所以数据从第3行开始）
            current_category = ''
            data_start_row = 3
            for idx, (_, row_data) in enumerate(summary_df.iterrows()):
                row = data_start_row + idx
                category = row_data['类别']
                item = row_data['项目']
                amount = row_data['金额(元)']
                
                # 如果是新的类别标题行
                if category and category != current_category:
                    current_category = category
                    # 类别标题行样式
                    for col in range(1, 4):
                        cell = summary_ws.cell(row=row, column=col)
                        if col == 1:
                            cell.value = category
                        elif col == 2:
                            cell.value = ''
                        else:
                            cell.value = ''
                        cell.font = title_font
                        cell.fill = title_fill
                        cell.alignment = left_alignment
                    summary_ws.row_dimensions[row].height = 22
                else:
                    # 普通数据行（数据已经通过to_excel写入，这里只设置样式和格式化）
                    for col in range(1, 4):
                        cell = summary_ws.cell(row=row, column=col)
                        cell.font = data_font
                        cell.alignment = left_alignment if col <= 2 else center_alignment
                        # 格式化金额列
                        if col == 3:
                            if cell.value is None:
                                cell.value = float(amount) if amount else 0.0
                            cell.number_format = '#,##0.00'
                    
                    # 特殊项目高亮
                    if item in ['间接人工总成本']:
                        for col in range(1, 4):
                            cell = summary_ws.cell(row=row, column=col)
                            cell.fill = total_fill
                            cell.font = total_font
                    elif item in ['间接人工提成汇总', '人工提成成本', '分摊固定成本', '其他岗位成本']:
                        for col in range(1, 4):
                            cell = summary_ws.cell(row=row, column=col)
                            cell.fill = highlight_fill
                            cell.font = highlight_font
                    summary_ws.row_dimensions[row].height = 20
            
            # 3. 人工提成明细表
            if all_details:
                details_df = pd.DataFrame(all_details)
                details_df.to_excel(writer, sheet_name='人工提成明细', index=False)
                
                # 设置详细数据表样式
                details_ws = writer.sheets['人工提成明细']
                for col in range(1, len(details_df.columns) + 1):
                    cell = details_ws.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    col_letter = get_column_letter(col)
                    if col == 1:  # 序号
                        details_ws.column_dimensions[col_letter].width = 10
                    elif col == 2:  # 原物料代码
                        details_ws.column_dimensions[col_letter].width = 18
                    elif col == 3:  # 原物料名称
                        details_ws.column_dimensions[col_letter].width = 30
                    elif col == 4:  # 物料/产物编码
                        details_ws.column_dimensions[col_letter].width = 18
                    elif col == 5:  # 物料/产物名称
                        details_ws.column_dimensions[col_letter].width = 30
                    else:
                        details_ws.column_dimensions[col_letter].width = 18
                
                for row in range(2, len(details_df) + 2):
                    for col in range(1, len(details_df.columns) + 1):
                        cell = details_ws.cell(row=row, column=col)
                        cell.font = data_font
                        cell.alignment = center_alignment
                        # 总成本列高亮
                        if col == len(details_df.columns):  # 最后一列（总成本）
                            cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                            cell.font = Font(bold=True, name="仿宋", color="2E7D32")
            
            # 3. 分摊固定成本明细表
            fixed_cost_details = []
            fixed_cost_index = 1
            
            def collect_fixed_cost_details(items, get_code, get_name, get_category, get_origin_code, get_origin_name):
                nonlocal fixed_cost_index
                for item in items:
                    # 计算该物料的总分摊固定成本
                    fixed_costs = {
                        '品管分摊': item.get('品管分摊固定成本', 0) or 0,
                        '物流主管分摊': item.get('物流主管分摊固定成本', 0) or 0,
                        '生产主管分摊': item.get('生产主管分摊固定成本', 0) or 0,
                        '物流卸货分摊': item.get('物流卸货分摊固定成本', 0) or 0,
                        '维修班长分摊': item.get('维修班长分摊固定成本', 0) or 0,
                        '维修员分摊': item.get('维修员分摊固定成本', 0) or 0,
                        '冰箱维修主管分摊': item.get('冰箱维修主管分摊固定成本', 0) or 0,
                        '叉车司磅库管等分摊': item.get('叉车司磅库管等分摊固定成本', 0) or 0,
                        '白电小组长分摊': 0,
                        '生产班组长(黑电)分摊': 0,
                        '生产班组长(冰箱)分摊': 0,
                        '生产班组长(塑料破碎)分摊': 0
                    }
                    
                    # 班组长分摊固定成本（根据物料名称和类别分类）
                    material_name = str(get_name(item) or '').upper()
                    category = get_category(item)
                    
                    if category == '旧机':
                        if '空调' in material_name or '洗衣机' in material_name:
                            fixed_costs['白电小组长分摊'] = item.get('白电小组长分摊固定成本', 0) or 0
                        elif '电视' in material_name or '电脑' in material_name:
                            fixed_costs['生产班组长(黑电)分摊'] = item.get('生产班组长(黑电)分摊固定成本', 0) or 0
                        elif '冰箱' in material_name:
                            fixed_costs['生产班组长(冰箱)分摊'] = item.get('生产班组长(冰箱)分摊固定成本', 0) or 0
                    elif category in ['一次拆解产物', '一破']:
                        fixed_costs['生产班组长(塑料破碎)分摊'] = item.get('生产班组长(塑料破碎)分摊固定成本', 0) or 0
                    
                    plastic_fc_no = float(
                        item.get('生产班组长(塑料破碎)分摊固定成本(不考虑期初库存和库存结余)', 0) or 0
                    )
                    total_fixed_cost = sum(fixed_costs.values())

                    # 有常规分摊或仅有塑料破碎(不考虑期初)分摊时均展示
                    if total_fixed_cost > 0 or plastic_fc_no > 0:
                        fixed_cost_details.append({
                            '序号': fixed_cost_index,
                            '原物料代码': get_origin_code(item) or '',
                            '原物料名称': get_origin_name(item) or '',
                            '物料/产物编码': get_code(item) or '',
                            '物料/产物名称': get_name(item) or '',
                            '类别': category,
                            '品管分摊': fixed_costs['品管分摊'],
                            '物流主管分摊': fixed_costs['物流主管分摊'],
                            '生产主管分摊': fixed_costs['生产主管分摊'],
                            '物流卸货分摊': fixed_costs['物流卸货分摊'],
                            '维修班长分摊': fixed_costs['维修班长分摊'],
                            '维修员分摊': fixed_costs['维修员分摊'],
                            '冰箱维修主管分摊': fixed_costs['冰箱维修主管分摊'],
                            '叉车司磅库管等分摊': fixed_costs['叉车司磅库管等分摊'],
                            '白电小组长分摊': fixed_costs['白电小组长分摊'],
                            '生产班组长(黑电)分摊': fixed_costs['生产班组长(黑电)分摊'],
                            '生产班组长(冰箱)分摊': fixed_costs['生产班组长(冰箱)分摊'],
                            '生产班组长(塑料破碎)分摊': fixed_costs['生产班组长(塑料破碎)分摊'],
                            '生产班组长(塑料破碎)分摊（不考虑期初库存和库存结余）': plastic_fc_no,
                            '分摊固定成本合计': total_fixed_cost
                        })
                        fixed_cost_index += 1
            
            # 收集所有明细数据
            if result.get('part1_details'):
                collect_fixed_cost_details(
                    result['part1_details'],
                    lambda item: item.get('物料代码', ''),
                    lambda item: item.get('物料名称', ''),
                    lambda item: '旧机',
                    lambda item: item.get('原物料代码', '') or item.get('物料代码', ''),
                    lambda item: item.get('原物料名称', '') or item.get('物料名称', '')
                )
            
            if result.get('part2_details'):
                collect_fixed_cost_details(
                    result['part2_details'],
                    lambda item: item.get('拆解产物编码', ''),
                    lambda item: item.get('拆解产物名称', ''),
                    lambda item: '一次拆解产物',
                    lambda item: item.get('原物料代码', ''),
                    lambda item: item.get('原物料名称', '')
                )
            
            if result.get('part3_details'):
                collect_fixed_cost_details(
                    result['part3_details'],
                    lambda item: item.get('深加工产物编码', ''),
                    lambda item: item.get('深加工产物名称', ''),
                    lambda item: item.get('类别', ''),
                    lambda item: item.get('原物料代码', ''),
                    lambda item: item.get('原物料名称', '')
                )
            
            if fixed_cost_details:
                fixed_cost_df = pd.DataFrame(fixed_cost_details)
                fixed_cost_df.to_excel(writer, sheet_name='分摊固定成本明细', index=False)
                
                # 设置分摊固定成本明细表样式
                fixed_cost_ws = writer.sheets['分摊固定成本明细']
                for col in range(1, len(fixed_cost_df.columns) + 1):
                    cell = fixed_cost_ws.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    col_letter = get_column_letter(col)
                    if col == 1:  # 序号
                        fixed_cost_ws.column_dimensions[col_letter].width = 10
                    elif col == 2:  # 原物料代码
                        fixed_cost_ws.column_dimensions[col_letter].width = 18
                    elif col == 3:  # 原物料名称
                        fixed_cost_ws.column_dimensions[col_letter].width = 30
                    elif col == 4:  # 物料/产物编码
                        fixed_cost_ws.column_dimensions[col_letter].width = 18
                    elif col == 5:  # 物料/产物名称
                        fixed_cost_ws.column_dimensions[col_letter].width = 30
                    else:
                        fixed_cost_ws.column_dimensions[col_letter].width = 18
                
                for row in range(2, len(fixed_cost_df) + 2):
                    for col in range(1, len(fixed_cost_df.columns) + 1):
                        cell = fixed_cost_ws.cell(row=row, column=col)
                        cell.font = data_font
                        cell.alignment = center_alignment
                        # 分摊固定成本合计列高亮
                        if col == len(fixed_cost_df.columns):  # 最后一列（分摊固定成本合计）
                            cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                            cell.font = Font(bold=True, name="仿宋", color="856404")
            
            # 4. 其他岗位成本明细表
            other_positions_details = result.get('other_positions_details', [])
            if other_positions_details:
                other_positions_data = []
                for item in other_positions_details:
                    other_positions_data.append({
                        '序号': len(other_positions_data) + 1,
                        '岗位': item.get('岗位', ''),
                        '人员基础配置': item.get('人员基础配置', 0),
                        '平均工资（元/月/人）': item.get('平均工资（元/月/人）', 0),
                        '奖励/补助（元/月）': item.get('奖励/补助（元/月）', 0),
                        '餐补（元/月/人）': item.get('餐补（元/月/人）', 0),
                        '年终奖（元/人）': item.get('年终奖（元/人）', 0),
                        '养老保险费（元/月/人）': item.get('养老保险费（元/月/人）', 0),
                        '失业保险费（元/月/人）': item.get('失业保险费（元/月/人）', 0),
                        '医疗/生育保险费（元/月/人）': item.get('医疗/生育保险费（元/月/人）', 0),
                        '工伤保险费（元/月/人）': item.get('工伤保险费（元/月/人）', 0),
                        '住房公积金（元/月/人）': item.get('住房公积金（元/月/人）', 0),
                        '月均固定成本（元/月/人）': item.get('月均固定成本（元/月/人）', 0),
                        '岗位成本': item.get('岗位成本', 0)
                    })
                
                # 添加合计行
                if other_positions_data:
                    total_personnel = sum(item.get('人员基础配置', 0) for item in other_positions_data)
                    total_cost = sum(item.get('岗位成本', 0) for item in other_positions_data)
                    other_positions_data.append({
                        '序号': '',
                        '岗位': '合计',
                        '人员基础配置': total_personnel,
                        '平均工资（元/月/人）': '',
                        '奖励/补助（元/月）': '',
                        '餐补（元/月/人）': '',
                        '年终奖（元/人）': '',
                        '养老保险费（元/月/人）': '',
                        '失业保险费（元/月/人）': '',
                        '医疗/生育保险费（元/月/人）': '',
                        '工伤保险费（元/月/人）': '',
                        '住房公积金（元/月/人）': '',
                        '月均固定成本（元/月/人）': '',
                        '岗位成本': total_cost
                    })
                
                other_positions_df = pd.DataFrame(other_positions_data)
                other_positions_df.to_excel(writer, sheet_name='其他岗位成本明细', index=False)
                
                # 设置其他岗位成本明细表样式
                other_positions_ws = writer.sheets['其他岗位成本明细']
                for col in range(1, len(other_positions_df.columns) + 1):
                    cell = other_positions_ws.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    col_letter = get_column_letter(col)
                    if col == 1:  # 序号
                        other_positions_ws.column_dimensions[col_letter].width = 10
                    elif col == 2:  # 岗位
                        other_positions_ws.column_dimensions[col_letter].width = 25
                    else:
                        other_positions_ws.column_dimensions[col_letter].width = 18
                
                for row in range(2, len(other_positions_df) + 2):
                    is_total_row = other_positions_ws.cell(row=row, column=2).value == '合计'
                    
                    for col in range(1, len(other_positions_df.columns) + 1):
                        cell = other_positions_ws.cell(row=row, column=col)
                        cell.font = data_font
                        cell.alignment = center_alignment
                        
                        # 合计行特殊样式
                        if is_total_row:
                            cell.font = Font(bold=True, name="仿宋")
                            if col == 2:  # 岗位列
                                cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                            elif col == len(other_positions_df.columns):  # 岗位成本列
                                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                                cell.font = Font(bold=True, name="仿宋", color="2E7D32")
                        # 岗位成本列高亮
                        elif col == len(other_positions_df.columns):  # 岗位成本列
                            cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                            cell.font = Font(bold=True, name="仿宋", color="2E7D32")
            
            # 5. 分类费用汇总表（按电视、电脑、冰箱、空调、洗衣机统计）
            # 产品类别映射规则
            PRODUCT_CATEGORY_MAPPING = {
                '电视': ['电视', '彩电', 'CRT其它机壳破碎塑料', '线路板边框破碎塑料', '等离子', '废旧玻璃电子枪', '废旧金属荫罩压块铁', '黑白'],
                '电脑': ['电脑', '显示器', '笔记本', '主机', '废旧金属黑色金属-铁及其合金-电子枪'],
                '冰箱': ['冰箱', '冰柜'],
                '空调': ['空调'],
                '洗衣机': ['洗衣机', '双缸']
            }
            
            def get_product_category(material_name):
                """根据物料名称判断归属类别"""
                if not material_name:
                    return None
                name = str(material_name)
                for category, keywords in PRODUCT_CATEGORY_MAPPING.items():
                    for keyword in keywords:
                        if keyword in name:
                            return category
                return None
            
            # 初始化统计数据
            category_stats = {
                '电视': {'间接人工成本': 0, '分摊固定成本': 0},
                '电脑': {'间接人工成本': 0, '分摊固定成本': 0},
                '冰箱': {'间接人工成本': 0, '分摊固定成本': 0},
                '空调': {'间接人工成本': 0, '分摊固定成本': 0},
                '洗衣机': {'间接人工成本': 0, '分摊固定成本': 0}
            }
            
            # 计算分摊固定成本合计的辅助函数
            def calc_fixed_cost_total(item):
                return (float(item.get('品管分摊固定成本', 0) or 0) +
                        float(item.get('物流主管分摊固定成本', 0) or 0) +
                        float(item.get('生产主管分摊固定成本', 0) or 0) +
                        float(item.get('物流卸货分摊固定成本', 0) or 0) +
                        float(item.get('维修班长分摊固定成本', 0) or 0) +
                        float(item.get('维修员分摊固定成本', 0) or 0) +
                        float(item.get('冰箱维修主管分摊固定成本', 0) or 0) +
                        float(item.get('叉车司磅库管等分摊固定成本', 0) or 0) +
                        float(item.get('白电小组长分摊固定成本', 0) or 0) +
                        float(item.get('生产班组长(黑电)分摊固定成本', 0) or 0) +
                        float(item.get('生产班组长(冰箱)分摊固定成本', 0) or 0) +
                        float(item.get('生产班组长(塑料破碎)分摊固定成本', 0) or 0))
            
            # 处理第一部分：旧机类别
            if result.get('part1_details'):
                for item in result['part1_details']:
                    name = item.get('物料名称', '')
                    category = get_product_category(name)
                    if category and category in category_stats:
                        category_stats[category]['间接人工成本'] += float(item.get('总成本', 0) or 0)
                        category_stats[category]['分摊固定成本'] += calc_fixed_cost_total(item)
            
            # 处理第二部分：一次拆解产物
            if result.get('part2_details'):
                for item in result['part2_details']:
                    name = item.get('拆解产物名称', '')
                    category = get_product_category(name)
                    if category and category in category_stats:
                        category_stats[category]['间接人工成本'] += float(item.get('总成本', 0) or 0)
                        category_stats[category]['分摊固定成本'] += calc_fixed_cost_total(item)
            
            # 处理第三部分：深加工
            if result.get('part3_details'):
                for item in result['part3_details']:
                    name = item.get('深加工产物名称', '')
                    category = get_product_category(name)
                    if category and category in category_stats:
                        category_stats[category]['间接人工成本'] += float(item.get('总成本', 0) or 0)
                        category_stats[category]['分摊固定成本'] += calc_fixed_cost_total(item)
            
            # 构建分类费用汇总数据
            category_summary_rows = []
            for category in ['电视', '电脑', '冰箱', '空调', '洗衣机']:
                labor_cost = category_stats[category]['间接人工成本']
                fixed_cost = category_stats[category]['分摊固定成本']
                total = labor_cost + fixed_cost
                category_summary_rows.append({
                    '产品类别': category,
                    '间接人工成本': labor_cost,
                    '分摊固定成本': fixed_cost,
                    '合计': total
                })
            
            # 添加总计行
            total_labor = sum(row['间接人工成本'] for row in category_summary_rows)
            total_fixed = sum(row['分摊固定成本'] for row in category_summary_rows)
            total_all = total_labor + total_fixed
            category_summary_rows.append({
                '产品类别': '总计',
                '间接人工成本': total_labor,
                '分摊固定成本': total_fixed,
                '合计': total_all
            })
            
            category_summary_df = pd.DataFrame(category_summary_rows)
            category_summary_df.to_excel(writer, sheet_name='分类费用汇总', index=False)
            
            # 设置分类费用汇总表样式
            category_summary_ws = writer.sheets['分类费用汇总']
            category_summary_ws.column_dimensions['A'].width = 15
            category_summary_ws.column_dimensions['B'].width = 18
            category_summary_ws.column_dimensions['C'].width = 18
            category_summary_ws.column_dimensions['D'].width = 18
            
            for col in range(1, 5):
                cell = category_summary_ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            for row in range(2, len(category_summary_df) + 2):
                is_total_row = category_summary_ws.cell(row=row, column=1).value == '总计'
                for col in range(1, 5):
                    cell = category_summary_ws.cell(row=row, column=col)
                    cell.font = data_font
                    cell.alignment = center_alignment
                    if col > 1:
                        cell.number_format = '#,##0.00'
                    
                    if is_total_row:
                        cell.font = Font(bold=True, name="仿宋", size=12)
                        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    elif col == 4:  # 合计列
                        cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                        cell.font = Font(bold=True, name="仿宋", color="2E7D32")
            
            # 6. 制造费用汇总（与间接人工成本页「制造费用汇总」卡片一致）
            sm_category_stats = summarize_indirect_labor_manufacturing_stats_by_category(result)
            sm_summary_rows = []
            for category in PRODUCTION_COST_CATEGORIES:
                labor_cost = sm_category_stats[category]['间接人工成本']
                fixed_cost = sm_category_stats[category]['分摊固定成本']
                total = labor_cost + fixed_cost
                sm_summary_rows.append({
                    '产品类别': category,
                    '间接人工成本': labor_cost,
                    '分摊固定成本': fixed_cost,
                    '合计': total
                })
            
            # 添加总计行
            sm_total_labor = sum(row['间接人工成本'] for row in sm_summary_rows)
            sm_total_fixed = sum(row['分摊固定成本'] for row in sm_summary_rows)
            sm_total_all = sm_total_labor + sm_total_fixed
            sm_summary_rows.append({
                '产品类别': '总计',
                '间接人工成本': sm_total_labor,
                '分摊固定成本': sm_total_fixed,
                '合计': sm_total_all
            })
            
            sm_summary_df = pd.DataFrame(sm_summary_rows)
            sm_summary_df.to_excel(writer, sheet_name='制造费用汇总', index=False)
            
            # 设置制造费用汇总表样式
            sm_summary_ws = writer.sheets['制造费用汇总']
            sm_summary_ws.column_dimensions['A'].width = 15
            sm_summary_ws.column_dimensions['B'].width = 18
            sm_summary_ws.column_dimensions['C'].width = 18
            sm_summary_ws.column_dimensions['D'].width = 18
            
            for col in range(1, 5):
                cell = sm_summary_ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            for row in range(2, len(sm_summary_df) + 2):
                is_total_row = sm_summary_ws.cell(row=row, column=1).value == '总计'
                for col in range(1, 5):
                    cell = sm_summary_ws.cell(row=row, column=col)
                    cell.font = data_font
                    cell.alignment = center_alignment
                    if col > 1:
                        cell.number_format = '#,##0.00'
                    
                    if is_total_row:
                        cell.font = Font(bold=True, name="仿宋", size=12)
                        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                    elif col == 4:  # 合计列
                        cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                        cell.font = Font(bold=True, name="仿宋", color="2E7D32")
        
        output.seek(0)
        filename = f'间接人工成本_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"导出间接人工成本失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def resolve_disassembly_category_unit_price(cost_row, category_col):
    """
    与拆解量相关费用：品类列有非零单价时优先使用，否则使用「公共」列。
    cost_row: pandas Series（如 manufacturing_cost_df.iterrows() 的一行）。
    """
    if category_col in cost_row.index:
        v = cost_row.get(category_col)
        if pd.notna(v):
            try:
                fv = float(v)
                if fv != 0:
                    return fv
            except (TypeError, ValueError):
                pass
    if '公共' in cost_row.index:
        v = cost_row.get('公共')
        if pd.notna(v):
            try:
                fv = float(v)
                if fv != 0:
                    return fv
            except (TypeError, ValueError):
                pass
    return 0.0


def _disassembly_appliance_columns_use_only_public(cost_row):
    """
    五大家电竞品列均未维护非零单价、且「公共」有非零单价时，
    制造费用成本页按「公共」一条汇总展示（数量=五类拆解量合计）。
    """
    for category in ['冰箱', '电脑', '电视', '空调', '洗衣机']:
        if category not in cost_row.index:
            continue
        v = cost_row.get(category)
        if pd.notna(v):
            try:
                if float(v) != 0:
                    return False
            except (TypeError, ValueError):
                pass
    if '公共' not in cost_row.index:
        return False
    v = cost_row.get('公共')
    if pd.notna(v):
        try:
            return float(v) != 0
        except (TypeError, ValueError):
            return False
    return False


PRODUCTION_COST_CATEGORIES = ['电视', '电脑', '冰箱', '空调', '洗衣机']

# 间接人工成本页与生产成本分摊「制造费用」第三来源口径一致（含不考虑期初列）
INDIRECT_LABOR_PAGE_INCLUDE_NO_OPENING = True


def _calc_indirect_labor_fixed_cost_total(item):
    """间接人工成本：分摊固定成本合计（含品管、叉车司磅库管等）。"""
    return (
        float(item.get('品管分摊固定成本', 0) or 0) +
        float(item.get('物流主管分摊固定成本', 0) or 0) +
        float(item.get('生产主管分摊固定成本', 0) or 0) +
        float(item.get('物流卸货分摊固定成本', 0) or 0) +
        float(item.get('维修班长分摊固定成本', 0) or 0) +
        float(item.get('维修员分摊固定成本', 0) or 0) +
        float(item.get('冰箱维修主管分摊固定成本', 0) or 0) +
        float(item.get('叉车司磅库管等分摊固定成本', 0) or 0) +
        float(item.get('白电小组长分摊固定成本', 0) or 0) +
        float(item.get('生产班组长(黑电)分摊固定成本', 0) or 0) +
        float(item.get('生产班组长(冰箱)分摊固定成本', 0) or 0) +
        float(item.get('生产班组长(塑料破碎)分摊固定成本', 0) or 0)
    )


def _calc_indirect_labor_manufacturing_fixed_cost(item):
    """制造费用汇总口径：分摊固定成本合计（排除品管、叉车司磅库管等）。"""
    return (
        float(item.get('物流主管分摊固定成本', 0) or 0) +
        float(item.get('生产主管分摊固定成本', 0) or 0) +
        float(item.get('物流卸货分摊固定成本', 0) or 0) +
        float(item.get('维修班长分摊固定成本', 0) or 0) +
        float(item.get('维修员分摊固定成本', 0) or 0) +
        float(item.get('冰箱维修主管分摊固定成本', 0) or 0) +
        float(item.get('白电小组长分摊固定成本', 0) or 0) +
        float(item.get('生产班组长(黑电)分摊固定成本', 0) or 0) +
        float(item.get('生产班组长(冰箱)分摊固定成本', 0) or 0) +
        float(item.get('生产班组长(塑料破碎)分摊固定成本', 0) or 0)
    )


def classify_environmental_fee_by_product_name(product_name):
    """
    环保费按拆解产物名称归类（与制造费用成本页 calculateAndRenderCategoryStats 一致）。
    返回产线名称，无法匹配时返回 None。
    """
    if not product_name:
        return None
    name = str(product_name)
    if (
        '彩电' in name or '电视' in name or '废旧玻璃电子枪' in name
        or '废旧金属荫罩压块铁' in name or '黑白' in name
    ):
        return '电视'
    if (
        '电脑' in name or '显示器' in name or '主机' in name or '笔记本' in name
        or '废旧金属黑色金属-铁及其合金-电子枪' in name
    ):
        return '电脑'
    if '冰箱' in name or '冰柜' in name:
        return '冰箱'
    if '空调' in name:
        return '空调'
    if '洗衣机' in name or '双缸' in name:
        return '洗衣机'
    if '屏' in name:
        return '电视'
    return None


def summarize_manufacturing_cost_category_totals(manufacturing_result):
    """
    制造费用成本页「分类费用汇总」按产线合计。
    仅统计五类产线明细；「公共」「屏」由公共费用分摊页处理。
    """
    categories = PRODUCTION_COST_CATEGORIES
    totals = {cat: 0.0 for cat in categories}
    if not manufacturing_result or manufacturing_result.get('success') is False:
        return totals

    for key in ('disassembly_related', 'motor_inventory_related', 'monthly_average'):
        for item in manufacturing_result.get(key) or []:
            for detail in item.get('明细', []):
                detail_cat = detail.get('category', '')
                if detail_cat in categories:
                    totals[detail_cat] += float(detail.get('cost', 0) or 0)

    env_fee = manufacturing_result.get('environmental_fee') or {}
    for detail in env_fee.get('明细', []):
        product_name = detail.get('拆解产物名称', '')
        fee = float(detail.get('费用', 0) or 0)
        cat = classify_environmental_fee_by_product_name(product_name)
        if cat and cat in totals:
            totals[cat] += fee

    return totals


def summarize_manufacturing_cost_category_breakdown(manufacturing_result):
    """制造费用成本页「分类费用汇总」分项明细（供导出/前端）。"""
    categories = PRODUCTION_COST_CATEGORIES
    stats = {
        cat: {'disassembly': 0.0, 'motor': 0.0, 'monthly': 0.0, 'environmental': 0.0}
        for cat in categories
    }
    if not manufacturing_result or manufacturing_result.get('success') is False:
        return stats

    for item in manufacturing_result.get('disassembly_related') or []:
        for detail in item.get('明细', []):
            cat = detail.get('category', '')
            if cat in stats:
                stats[cat]['disassembly'] += float(detail.get('cost', 0) or 0)

    for item in manufacturing_result.get('motor_inventory_related') or []:
        for detail in item.get('明细', []):
            cat = detail.get('category', '')
            if cat in stats:
                stats[cat]['motor'] += float(detail.get('cost', 0) or 0)

    for item in manufacturing_result.get('monthly_average') or []:
        for detail in item.get('明细', []):
            cat = detail.get('category', '')
            if cat in stats:
                stats[cat]['monthly'] += float(detail.get('cost', 0) or 0)

    env_fee = manufacturing_result.get('environmental_fee') or {}
    for detail in env_fee.get('明细', []):
        product_name = detail.get('拆解产物名称', '')
        fee = float(detail.get('费用', 0) or 0)
        cat = classify_environmental_fee_by_product_name(product_name)
        if cat and cat in stats:
            stats[cat]['environmental'] += fee

    return stats


def summarize_indirect_labor_manufacturing_stats_by_category(indirect_labor_result):
    """
    间接人工成本页「制造费用汇总」分项（间接人工成本、分摊固定成本）。
    """
    categories = PRODUCTION_COST_CATEGORIES
    stats = {cat: {'间接人工成本': 0.0, '分摊固定成本': 0.0} for cat in categories}

    if not indirect_labor_result or indirect_labor_result.get('error'):
        return stats

    def _accumulate_part(items, name_field):
        for item in items or []:
            name = item.get(name_field, '')
            category = classify_by_product_name(name)
            if not category or category not in stats:
                continue
            total_cost = float(item.get('总成本', 0) or 0)
            qc_cost = float(item.get('品管提成成本', 0) or 0)
            forklift_cost = float(item.get('叉车司磅库管等提成成本', 0) or 0)
            stats[category]['间接人工成本'] += (total_cost - qc_cost - forklift_cost)
            full_fixed = _calc_indirect_labor_fixed_cost_total(item)
            qc_fixed = float(item.get('品管分摊固定成本', 0) or 0)
            forklift_fixed = float(item.get('叉车司磅库管等分摊固定成本', 0) or 0)
            stats[category]['分摊固定成本'] += (full_fixed - qc_fixed - forklift_fixed)

    _accumulate_part(indirect_labor_result.get('part1_details'), '物料名称')
    _accumulate_part(indirect_labor_result.get('part2_details'), '拆解产物名称')
    _accumulate_part(indirect_labor_result.get('part3_details'), '深加工产物名称')

    return stats


def summarize_indirect_labor_manufacturing_by_category(indirect_labor_result):
    """间接人工成本页「制造费用汇总」按产线合计。"""
    stats = summarize_indirect_labor_manufacturing_stats_by_category(indirect_labor_result)
    return {
        cat: stats[cat]['间接人工成本'] + stats[cat]['分摊固定成本']
        for cat in PRODUCTION_COST_CATEGORIES
    }


def collect_production_manufacturing_cost_by_category(app_data, prediction_period=1):
    """
    生产成本分摊页「制造费用」列：三来源按产线汇总。
    1. 制造费用成本页分类费用汇总
    2. 公共费用分摊页按类别分摊
    3. 间接人工成本页制造费用汇总
    """
    categories = PRODUCTION_COST_CATEGORIES
    totals = {cat: 0.0 for cat in categories}

    try:
        manufacturing_result = calculate_manufacturing_cost(app_data, prediction_period)
        mfg_totals = summarize_manufacturing_cost_category_totals(manufacturing_result)
        for cat in categories:
            totals[cat] += mfg_totals.get(cat, 0.0)
    except Exception as e:
        print(f"收集制造费用成本分类汇总失败: {str(e)}")

    try:
        screen_allocation_result = calculate_screen_cost_allocation(app_data, prediction_period)
        category_allocation = (
            screen_allocation_result.get('allocation', {}) or {}
        ).get('category_allocation') or {}
        for cat in categories:
            totals[cat] += float(category_allocation.get(cat, 0) or 0)
    except Exception as e:
        print(f"收集公共费用分摊失败: {str(e)}")

    try:
        indirect_labor_result = calculate_indirect_labor_cost(
            app_data,
            prediction_period,
            include_no_opening_columns=INDIRECT_LABOR_PAGE_INCLUDE_NO_OPENING,
        )
        ind_totals = summarize_indirect_labor_manufacturing_by_category(indirect_labor_result)
        for cat in categories:
            totals[cat] += ind_totals.get(cat, 0.0)
    except Exception as e:
        print(f"收集间接人工制造费用汇总失败: {str(e)}")

    return totals


def calculate_manufacturing_cost(app_data, prediction_period=1):
    """
    计算制造费用成本
    
    包括三类费用：
    1. 与拆解量相关的费用
    2. 与电机入库量相关的费用
    3. 预计月均费用
    
    Args:
        app_data: 应用数据管理器
        prediction_period: 预测期数（月）
        
    Returns:
        dict: 包含所有制造费用成本数据的字典
    """
    try:
        from app.api.data_management_api import get_manufacturing_cost_dataframe
        
        # 获取制造费用基础数据
        manufacturing_cost_df = get_manufacturing_cost_dataframe()
        if manufacturing_cost_df.empty:
            return {
                'success': False,
                'error': '无法获取制造费用基础数据'
            }
        
        # 确保备注列存在
        if '备注' not in manufacturing_cost_df.columns:
            manufacturing_cost_df['备注'] = ''
        
        # 填充空备注为"预计月均费用"
        manufacturing_cost_df['备注'] = manufacturing_cost_df['备注'].fillna('预计月均费用')
        manufacturing_cost_df.loc[manufacturing_cost_df['备注'].astype(str).str.strip() == '', '备注'] = '预计月均费用'
        manufacturing_cost_df.loc[manufacturing_cost_df['备注'].astype(str).str.strip() == 'nan', '备注'] = '预计月均费用'
        
        result = {
            'disassembly_related': [],  # 与拆解量相关的费用
            'motor_inventory_related': [],  # 与电机入库量相关的费用
            'monthly_average': [],  # 预计月均费用
            'environmental_fee': [],  # 环保费
            'total_cost': 0.0
        }
        
        # ========== 一、与拆解量相关的费用计算 ==========
        
        # 1.1 从拆解物原料成本页获取数据
        extracted_data = app_data.get_data('extracted_data_manual')
        category_totals = {
            '冰箱': 0.0,
            '电脑': 0.0,
            '电视': 0.0,
            '空调': 0.0,
            '洗衣机': 0.0
        }
        
        if extracted_data is not None and not extracted_data.empty:
            # 确保有物料描述和本期实际投产数量列
            if '物料描述' in extracted_data.columns and '非限制使用的库存' in extracted_data.columns:
                # 只处理旧机类别
                if '类别' in extracted_data.columns:
                    old_machine_data = extracted_data[extracted_data['类别'] == '旧机'].copy()
                else:
                    old_machine_data = extracted_data.copy()
                
                # 确保数值类型
                old_machine_data['非限制使用的库存'] = pd.to_numeric(
                    old_machine_data['非限制使用的库存'], errors='coerce'
                ).fillna(0)
                
                # 按物料描述分类汇总
                for category in category_totals.keys():
                    mask = old_machine_data['物料描述'].astype(str).str.contains(category, case=False, na=False)
                    category_totals[category] = old_machine_data.loc[mask, '非限制使用的库存'].sum()
        
        # 1.2 从被减扣数据获取屏的数据
        deducted_data = app_data.get_data('deducted_data_manual')
        screen_total = 0.0
        
        if deducted_data is not None and not deducted_data.empty:
            if '处置类别' in deducted_data.columns and '计算结果(KG)' in deducted_data.columns:
                # 筛选处置类别为"内转屏处置"的记录
                screen_mask = deducted_data['处置类别'].astype(str).str.contains('内转屏处置', case=False, na=False)
                screen_data = deducted_data[screen_mask].copy()
                if not screen_data.empty:
                    screen_data['计算结果(KG)'] = pd.to_numeric(
                        screen_data['计算结果(KG)'], errors='coerce'
                    ).fillna(0)
                    screen_total = screen_data['计算结果(KG)'].sum()
        
        # 1.3 从制造费用基础数据筛选备注为"与拆解量相关"的记录
        disassembly_related_df = manufacturing_cost_df[
            manufacturing_cost_df['备注'].astype(str).str.contains('与拆解量相关', case=False, na=False)
        ].copy()
        
        # 1.4 计算与拆解量相关的费用
        for _, cost_row in disassembly_related_df.iterrows():
            cost_name = cost_row.get('费用名称', '')
            expense_type = cost_row.get('费用类型', '')
            expense_category = cost_row.get('费用种类', '')
            
            total_cost = 0.0
            details = []
            appliance_categories = ['冰箱', '电脑', '电视', '空调', '洗衣机']

            # 仅维护「公共」、未维护各品类单价时：一条明细「公共」× 五类拆解量合计
            if _disassembly_appliance_columns_use_only_public(cost_row):
                try:
                    public_val = float(cost_row.get('公共'))
                except (TypeError, ValueError):
                    public_val = 0.0
                total_qty = sum(category_totals.get(c, 0.0) for c in appliance_categories)
                if total_qty > 0 and public_val != 0:
                    cost_ap = total_qty * public_val
                    total_cost += cost_ap
                    details.append({
                        'category': '公共',
                        'quantity': total_qty,
                        'unit_price': public_val,
                        'cost': cost_ap,
                    })
            else:
                # 品类列优先，否则回退「公共」；按品类分行
                for category in appliance_categories:
                    unit_price = resolve_disassembly_category_unit_price(cost_row, category)
                    if unit_price == 0:
                        continue
                    quantity = category_totals.get(category, 0.0)
                    cost = quantity * unit_price
                    total_cost += cost

                    if cost > 0:
                        details.append({
                            'category': category,
                            'quantity': quantity,
                            'unit_price': unit_price,
                            'cost': cost
                        })
            
            # 计算屏的费用
            if '屏' in cost_row and pd.notna(cost_row['屏']) and cost_row['屏'] != 0:
                unit_price = float(cost_row['屏'])
                cost = screen_total * unit_price
                total_cost += cost
                
                if cost > 0:
                    details.append({
                        'category': '屏',
                        'quantity': screen_total,
                        'unit_price': unit_price,
                        'cost': cost
                    })
            
            if total_cost > 0 or len(details) > 0:
                result['disassembly_related'].append({
                    '费用类型': expense_type,
                    '费用种类': expense_category,
                    '费用名称': cost_name,
                    '总成本': total_cost,
                    '明细': details
                })
                result['total_cost'] += total_cost
        
        # ========== 二、与电机入库量相关的费用计算 ==========
        
        motor_codes = ['811053046', '811053050', '811304664', '811437999']
        motor_totals = {
            '空调': 0.0,
            '洗衣机': 0.0
        }
        
        if deducted_data is not None and not deducted_data.empty:
            if '拆解产物编码' in deducted_data.columns and '计算结果(KG)' in deducted_data.columns:
                # 筛选拆解产物编码（去除空格后比较）
                motor_mask = deducted_data['拆解产物编码'].astype(str).str.strip().isin([code.strip() for code in motor_codes])
                motor_data = deducted_data[motor_mask].copy()
                
                if not motor_data.empty:
                    # 确保数值类型
                    motor_data['计算结果(KG)'] = pd.to_numeric(
                        motor_data['计算结果(KG)'], errors='coerce'
                    ).fillna(0)
                    
                    # 按物料名称分类汇总（优先使用拆解产物名称，如果没有则使用原物料名称）
                    name_column = None
                    if '拆解产物名称' in motor_data.columns:
                        name_column = '拆解产物名称'
                    elif '原物料名称' in motor_data.columns:
                        name_column = '原物料名称'
                    elif '物料名称' in motor_data.columns:
                        name_column = '物料名称'
                    
                    if name_column:
                        for category in motor_totals.keys():
                            name_mask = motor_data[name_column].astype(str).str.contains(category, case=False, na=False)
                            motor_totals[category] = motor_data.loc[name_mask, '计算结果(KG)'].sum()
        
        # 从制造费用基础数据筛选备注为"与电机入库量相关"的记录
        motor_related_df = manufacturing_cost_df[
            manufacturing_cost_df['备注'].astype(str).str.contains('与电机入库量相关', case=False, na=False)
        ].copy()
        
        # 计算与电机入库量相关的费用
        for _, cost_row in motor_related_df.iterrows():
            cost_name = cost_row.get('费用名称', '')
            expense_type = cost_row.get('费用类型', '')
            expense_category = cost_row.get('费用种类', '')
            
            total_cost = 0.0
            details = []
            
            # 计算空调和洗衣机的费用
            for category in ['空调', '洗衣机']:
                if category in cost_row and pd.notna(cost_row[category]) and cost_row[category] != 0:
                    unit_price = float(cost_row[category])
                    quantity = motor_totals.get(category, 0.0)
                    cost = quantity * unit_price
                    total_cost += cost
                    
                    if cost > 0:
                        details.append({
                            'category': category,
                            'quantity': quantity,
                            'unit_price': unit_price,
                            'cost': cost
                        })
            
            if total_cost > 0 or len(details) > 0:
                result['motor_inventory_related'].append({
                    '费用类型': expense_type,
                    '费用种类': expense_category,
                    '费用名称': cost_name,
                    '总成本': total_cost,
                    '明细': details
                })
                result['total_cost'] += total_cost
        
        # ========== 三、预计月均费用计算 ==========
        
        # 从制造费用基础数据筛选备注为"预计月均费用"的记录
        monthly_average_df = manufacturing_cost_df[
            manufacturing_cost_df['备注'].astype(str).str.contains('预计月均费用', case=False, na=False)
        ].copy()
        
        # 计算预计月均费用
        for _, cost_row in monthly_average_df.iterrows():
            cost_name = cost_row.get('费用名称', '')
            expense_type = cost_row.get('费用类型', '')
            expense_category = cost_row.get('费用种类', '')
            
            total_cost = 0.0
            details = []
            
            # 计算公共费用
            if '公共' in cost_row and pd.notna(cost_row['公共']) and cost_row['公共'] != 0:
                monthly_cost = float(cost_row['公共'])
                cost = monthly_cost * prediction_period
                total_cost += cost
                details.append({
                    'category': '公共',
                    'monthly_cost': monthly_cost,
                    'periods': prediction_period,
                    'cost': cost
                })
            
            # 计算各类别的费用
            for category in ['冰箱', '电脑', '电视', '空调', '洗衣机', '屏']:
                if category in cost_row and pd.notna(cost_row[category]) and cost_row[category] != 0:
                    monthly_cost = float(cost_row[category])
                    cost = monthly_cost * prediction_period
                    total_cost += cost
                    details.append({
                        'category': category,
                        'monthly_cost': monthly_cost,
                        'periods': prediction_period,
                        'cost': cost
                    })
            
            if total_cost > 0 or len(details) > 0:
                result['monthly_average'].append({
                    '费用类型': expense_type,
                    '费用种类': expense_category,
                    '费用名称': cost_name,
                    '总成本': total_cost,
                    '明细': details
                })
                result['total_cost'] += total_cost
        
        # ========== 四、环保费计算 ==========
        
        # 4.1 加载价格数据
        from data.base_data.price_data import load_price_data
        price_df = load_price_data()
        
        if price_df is not None and not price_df.empty:
            # 4.2 筛选环保费候选编码：与 statistics_api 2.5.6 一致，以「销售单价-不含税(元/KG) < 0」识别环保费单价，
            # 不再按销售产物名称关键字缩小范围（避免笔记本/显示器/主机等被整类漏算）。
            if '销售产物名称' not in price_df.columns:
                price_df['销售产物名称'] = ''
            price_df['销售单价(元/KG)'] = pd.to_numeric(
                price_df['销售单价(元/KG)'], errors='coerce'
            ).fillna(0)
            if '销售单价-不含税(元/KG)' in price_df.columns:
                price_df['销售单价-不含税(元/KG)'] = pd.to_numeric(
                    price_df['销售单价-不含税(元/KG)'], errors='coerce'
                ).fillna(0)
                negative_price_mask = price_df['销售单价-不含税(元/KG)'] < 0
            else:
                negative_price_mask = price_df['销售单价(元/KG)'] < 0
                price_df['销售单价-不含税(元/KG)'] = price_df['销售单价(元/KG)']
            
            filtered_price_df = price_df[negative_price_mask].copy()
            
            if not filtered_price_df.empty:
                # 确保编码列为字符串类型并去除空格
                filtered_price_df['拆解产物编码'] = filtered_price_df['拆解产物编码'].astype(str).str.strip()
                
                # 确保不含税单价列为数值类型
                filtered_price_df['销售单价-不含税(元/KG)'] = pd.to_numeric(
                    filtered_price_df['销售单价-不含税(元/KG)'], errors='coerce'
                ).fillna(0)
                
                # 同编码多行时保留最后一条（与 statistics_api 逐行覆盖映射行为一致）
                filtered_price_df = filtered_price_df.drop_duplicates(subset=['拆解产物编码'], keep='last')
                
                # 处理"彩电"映射为"电视"（与品类说明一致）
                filtered_price_df['销售产物名称'] = filtered_price_df['销售产物名称'].astype(str).str.replace('彩电', '电视', case=False)
                
                # 4.3 计算环保费第一部分（被减扣数据，排除屏）
                # 重量取自「被减扣数据（只读）」与数据管理页被减扣数据表 /deducted-data 同源，
                # 不使用「被减扣数据（手工）」deducted_data_manual，避免编辑手工表后环保费重量与只读表不一致。
                deducted_environmental_fee = 0.0
                deducted_details = []
                
                from app.api.data_management_api import _build_deducted_readonly_dataframe
                deducted_data = _build_deducted_readonly_dataframe(app_data)
                
                if deducted_data is not None and not deducted_data.empty:
                    # 筛选非屏相关的编码
                    non_screen_mask = ~filtered_price_df['销售产物名称'].astype(str).str.contains('屏', case=False, na=False)
                    non_screen_codes = filtered_price_df[non_screen_mask]['拆解产物编码'].unique().tolist()  # 使用unique()确保去重
                    
                    if non_screen_codes:
                        # 确保被减扣数据的编码列为字符串类型并去除空格
                        deducted_data = deducted_data.copy()
                        deducted_data['拆解产物编码'] = deducted_data['拆解产物编码'].astype(str).str.strip()
                        
                        # 仅统计「类别」为拆解产物的行（与页面说明一致；strip 避免 Excel 首尾空格导致筛不掉）
                        if '类别' not in deducted_data.columns:
                            deducted_data_filtered = deducted_data.iloc[0:0].copy()
                        else:
                            cat_norm = (
                                deducted_data['类别']
                                .astype(str)
                                .str.strip()
                                .str.replace('\u00a0', '', regex=False)
                                .str.replace('\u3000', '', regex=False)
                            )
                            deducted_data_filtered = deducted_data[cat_norm == '拆解产物'].copy()
                        
                        # 确保重量列为数值类型
                        if '计算结果(KG)' in deducted_data_filtered.columns:
                            deducted_data_filtered['计算结果(KG)'] = pd.to_numeric(
                                deducted_data_filtered['计算结果(KG)'], errors='coerce'
                            ).fillna(0)
                            
                            # 匹配编码
                            for code in non_screen_codes:
                                code_mask = deducted_data_filtered['拆解产物编码'] == code
                                matched_data = deducted_data_filtered[code_mask]
                                
                                if not matched_data.empty:
                                    # 获取价格信息
                                    price_info = filtered_price_df[filtered_price_df['拆解产物编码'] == code]
                                    if not price_info.empty:
                                        unit_price = float(price_info.iloc[0]['销售单价-不含税(元/KG)'])
                                        product_name = str(price_info.iloc[0]['销售产物名称'])
                                        
                                        # 统计重量
                                        total_weight = matched_data['计算结果(KG)'].sum()
                                        
                                        if total_weight > 0:
                                            # 计算费用：取单价绝对值进行计算
                                            unit_price_abs = abs(unit_price)
                                            fee = total_weight * unit_price_abs
                                            deducted_environmental_fee += fee
                                            
                                            deducted_details.append({
                                                '拆解产物编码': code,
                                                '拆解产物名称': str(matched_data.iloc[0].get('拆解产物名称', '')),
                                                '类别': '被减扣',
                                                '重量': float(total_weight),
                                                '单价': float(unit_price_abs),
                                                '费用': float(fee)
                                            })
                
                # 4.4 计算环保费第二部分（深加工数据，只使用屏）
                deep_processing_environmental_fee = 0.0
                deep_processing_details = []
                
                deep_processing_data = app_data.get_data('deep_processing_data')
                if deep_processing_data is not None and not deep_processing_data.empty:
                    # 筛选屏相关的编码
                    screen_mask = filtered_price_df['销售产物名称'].astype(str).str.contains('屏', case=False, na=False)
                    screen_codes = filtered_price_df[screen_mask]['拆解产物编码'].unique().tolist()  # 使用unique()确保去重
                    
                    if screen_codes:
                        # 确保深加工数据的编码列为字符串类型并去除空格
                        if '深加工产物编码' in deep_processing_data.columns:
                            deep_processing_data['深加工产物编码'] = deep_processing_data['深加工产物编码'].astype(str).str.strip()
                            
                            # 确保重量列为数值类型
                            if '深加工结果(KG)' in deep_processing_data.columns:
                                deep_processing_data['深加工结果(KG)'] = pd.to_numeric(
                                    deep_processing_data['深加工结果(KG)'], errors='coerce'
                                ).fillna(0)
                                
                                # 匹配编码（注意：深加工数据中是深加工产物编码）
                                for code in screen_codes:
                                    code_mask = deep_processing_data['深加工产物编码'] == code
                                    matched_data = deep_processing_data[code_mask]
                                    
                                    if not matched_data.empty:
                                        # 获取价格信息
                                        price_info = filtered_price_df[filtered_price_df['拆解产物编码'] == code]
                                        if not price_info.empty:
                                            unit_price = float(price_info.iloc[0]['销售单价-不含税(元/KG)'])
                                            product_name = str(price_info.iloc[0]['销售产物名称'])
                                            
                                            # 统计重量
                                            total_weight = matched_data['深加工结果(KG)'].sum()
                                            
                                            if total_weight > 0:
                                                # 计算费用：取单价绝对值进行计算
                                                unit_price_abs = abs(unit_price)
                                                fee = total_weight * unit_price_abs
                                                deep_processing_environmental_fee += fee
                                                
                                                deep_processing_details.append({
                                                    '拆解产物编码': code,
                                                    '拆解产物名称': str(matched_data.iloc[0].get('深加工产物名称', '')),
                                                    '类别': '深加工',
                                                    '重量': float(total_weight),
                                                    '单价': float(unit_price_abs),
                                                    '费用': float(fee)
                                                })
                
                # 4.5 合并环保费明细
                total_environmental_fee = deducted_environmental_fee + deep_processing_environmental_fee
                all_environmental_details = deducted_details + deep_processing_details
                
                if total_environmental_fee != 0 or len(all_environmental_details) > 0:
                    result['environmental_fee'] = {
                        '总成本': float(total_environmental_fee),
                        '明细': all_environmental_details
                    }
                    result['total_cost'] += total_environmental_fee
        
        result['success'] = True
        result['category_summary'] = summarize_manufacturing_cost_category_totals(result)
        result['category_breakdown'] = summarize_manufacturing_cost_category_breakdown(result)
        return result
        
    except Exception as e:
        print(f"计算制造费用成本失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'error': str(e)
        }


@cost_forecast_bp.route('/manufacturing-cost', methods=['GET'])
def get_manufacturing_cost():
    """获取制造费用成本数据"""
    try:
        app_data = get_session_data_manager()
        prediction_period = int(request.args.get('prediction_period', 1))
        
        # 检查数据是否已被清除
        data_cleared = app_data.get_data('__data_cleared__')
        if data_cleared:
            # 数据已被清除，直接返回空数据
            return jsonify({
                'success': True,
                'data': {
                    'disassembly_related': [],
                    'motor_inventory_related': [],
                    'monthly_average': [],
                    'environmental_fee': [],
                    'total_cost': 0.0
                }
            })
        
        result = calculate_manufacturing_cost(app_data, prediction_period)
        
        if not result.get('success', False):
            return jsonify({
                'success': False,
                'error': result.get('error', '计算失败')
            }), 500
        
        return jsonify({
            'success': True,
            'data': result
        })
        
    except Exception as e:
        print(f"获取制造费用成本失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@cost_forecast_bp.route('/manufacturing-cost/export', methods=['GET'])
def export_manufacturing_cost():
    """导出制造费用成本数据到Excel"""
    try:
        app_data = get_session_data_manager()
        prediction_period = int(request.args.get('prediction_period', 1))
        
        result = calculate_manufacturing_cost(app_data, prediction_period)
        
        if not result.get('success', False):
            return jsonify({
                'success': False,
                'error': result.get('error', '计算失败')
            }), 500
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: 与拆解量相关的费用
            disassembly_data = []
            for item in result.get('disassembly_related', []):
                for detail in item.get('明细', []):
                    disassembly_data.append({
                        '费用类型': item.get('费用类型', ''),
                        '费用种类': item.get('费用种类', ''),
                        '费用名称': item.get('费用名称', ''),
                        '类别': detail.get('category', ''),
                        '数量/重量': detail.get('quantity', 0),
                        '单价': detail.get('unit_price', 0),
                        '成本': detail.get('cost', 0)
                    })
                # 添加小计行
                if item.get('明细'):
                    disassembly_data.append({
                        '费用类型': item.get('费用类型', ''),
                        '费用种类': item.get('费用种类', ''),
                        '费用名称': f"{item.get('费用名称', '')}小计",
                        '类别': '',
                        '数量/重量': '',
                        '单价': '',
                        '成本': item.get('总成本', 0)
                    })
            
            if disassembly_data:
                disassembly_df = pd.DataFrame(disassembly_data)
                disassembly_df.to_excel(writer, sheet_name='与拆解量相关', index=False)
                worksheet = writer.sheets['与拆解量相关']
                from openpyxl.styles import Font, PatternFill, Alignment
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", name="仿宋")
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Sheet 2: 与电机入库量相关的费用
            motor_data = []
            for item in result.get('motor_inventory_related', []):
                for detail in item.get('明细', []):
                    motor_data.append({
                        '费用类型': item.get('费用类型', ''),
                        '费用种类': item.get('费用种类', ''),
                        '费用名称': item.get('费用名称', ''),
                        '类别': detail.get('category', ''),
                        '数量/重量': detail.get('quantity', 0),
                        '单价': detail.get('unit_price', 0),
                        '成本': detail.get('cost', 0)
                    })
                if item.get('明细'):
                    motor_data.append({
                        '费用类型': item.get('费用类型', ''),
                        '费用种类': item.get('费用种类', ''),
                        '费用名称': f"{item.get('费用名称', '')}小计",
                        '类别': '',
                        '数量/重量': '',
                        '单价': '',
                        '成本': item.get('总成本', 0)
                    })
            
            if motor_data:
                motor_df = pd.DataFrame(motor_data)
                motor_df.to_excel(writer, sheet_name='与电机入库量相关', index=False)
                worksheet = writer.sheets['与电机入库量相关']
                from openpyxl.styles import Font, PatternFill, Alignment
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", name="仿宋")
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Sheet 3: 预计月均费用
            monthly_data = []
            for item in result.get('monthly_average', []):
                for detail in item.get('明细', []):
                    monthly_data.append({
                        '费用类型': item.get('费用类型', ''),
                        '费用种类': item.get('费用种类', ''),
                        '费用名称': item.get('费用名称', ''),
                        '类别': detail.get('category', ''),
                        '月均费用': detail.get('monthly_cost', 0),
                        '期数': detail.get('periods', 0),
                        '成本': detail.get('cost', 0)
                    })
                if item.get('明细'):
                    monthly_data.append({
                        '费用类型': item.get('费用类型', ''),
                        '费用种类': item.get('费用种类', ''),
                        '费用名称': f"{item.get('费用名称', '')}小计",
                        '类别': '',
                        '月均费用': '',
                        '期数': '',
                        '成本': item.get('总成本', 0)
                    })
            
            if monthly_data:
                monthly_df = pd.DataFrame(monthly_data)
                monthly_df.to_excel(writer, sheet_name='预计月均费用', index=False)
                worksheet = writer.sheets['预计月均费用']
                from openpyxl.styles import Font, PatternFill, Alignment
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", name="仿宋")
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Sheet 4: 环保费
            environmental_fee_data = []
            environmental_fee_info = result.get('environmental_fee', {})
            if environmental_fee_info and environmental_fee_info.get('明细'):
                for detail in environmental_fee_info.get('明细', []):
                    environmental_fee_data.append({
                        '拆解产物编码': detail.get('拆解产物编码', ''),
                        '拆解产物名称': detail.get('拆解产物名称', ''),
                        '类别': detail.get('类别', ''),
                        '重量(KG)': detail.get('重量', 0),
                        '单价(元/KG)': detail.get('单价', 0),
                        '费用(元)': detail.get('费用', 0)
                    })
                # 添加小计行
                if environmental_fee_data:
                    environmental_fee_data.append({
                        '拆解产物编码': '',
                        '拆解产物名称': '环保费小计',
                        '类别': '',
                        '重量(KG)': '',
                        '单价(元/KG)': '',
                        '费用(元)': environmental_fee_info.get('总成本', 0)
                    })
            
            if environmental_fee_data:
                environmental_fee_df = pd.DataFrame(environmental_fee_data)
                environmental_fee_df.to_excel(writer, sheet_name='环保费', index=False)
                worksheet = writer.sheets['环保费']
                from openpyxl.styles import Font, PatternFill, Alignment
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", name="仿宋")
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Sheet 5: 分类费用汇总
            category_stats = summarize_manufacturing_cost_category_breakdown(result)
            category_summary_data = []
            for cat in PRODUCTION_COST_CATEGORIES:
                stats = category_stats[cat]
                total = stats['disassembly'] + stats['motor'] + stats['monthly'] + stats['environmental']
                category_summary_data.append({
                    '类别': cat,
                    '与拆解量相关': stats['disassembly'],
                    '与电机入库量相关': stats['motor'],
                    '预计月均费用': stats['monthly'],
                    '环保费': stats['environmental'],
                    '合计': total
                })
            
            if category_summary_data:
                category_summary_df = pd.DataFrame(category_summary_data)
                category_summary_df.to_excel(writer, sheet_name='分类费用汇总', index=False)
                worksheet = writer.sheets['分类费用汇总']
                from openpyxl.styles import Font, PatternFill, Alignment
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", name="仿宋")
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Sheet 6: 汇总
            environmental_fee_total = environmental_fee_info.get('总成本', 0) if environmental_fee_info else 0
            summary_data = [{
                '费用类别': '与拆解量相关',
                '总成本': sum(item.get('总成本', 0) for item in result.get('disassembly_related', []))
            }, {
                '费用类别': '与电机入库量相关',
                '总成本': sum(item.get('总成本', 0) for item in result.get('motor_inventory_related', []))
            }, {
                '费用类别': '预计月均费用',
                '总成本': sum(item.get('总成本', 0) for item in result.get('monthly_average', []))
            }, {
                '费用类别': '环保费',
                '总成本': environmental_fee_total
            }, {
                '费用类别': '制造费用成本合计',
                '总成本': result.get('total_cost', 0)
            }]
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='汇总', index=False)
            worksheet = writer.sheets['汇总']
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", name="仿宋")
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 设置最后一行为粗体
            if len(summary_df) > 0:
                last_row = worksheet[len(summary_df) + 1]
                for cell in last_row:
                    cell.font = Font(bold=True, name="仿宋", color="2E7D32")
        
        output.seek(0)
        filename = f'制造费用成本_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"导出制造费用成本失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def calculate_screen_cost_allocation(app_data, prediction_period=1):
    """
    计算公共费用分摊明细（屏相关费用分摊到电视和电脑）
    
    Args:
        app_data: 应用数据管理器
        prediction_period: 预测期数（月）
        
    Returns:
        dict: 包含所有屏相关费用和分摊计算结果的字典
    """
    try:
        result = {
            'direct_labor': {
                'screen_wage': 0.0,  # 屏工资（计件工资）
                'screen_allocation': 0.0,  # 屏分摊（固定工资分摊）
                'screen_wage_details': [],
                'screen_allocation_details': []
            },
            'manufacturing_cost': {
                'disassembly_related_screen': 0.0,  # 与拆解量相关的费用（屏）
                'monthly_average_screen': 0.0,  # 预计月均费用（屏）
                'environmental_fee_screen': 0.0,  # 环保费（屏）
                'disassembly_related_details': [],
                'monthly_average_details': [],
                'environmental_fee_details': []
            },
            'deducted_data': {
                'tv_weight': 0.0,  # 电视重量（KG）
                'computer_weight': 0.0,  # 电脑重量（KG）
                'total_weight': 0.0,  # 总重量（KG）
                'tv_details': [],
                'computer_details': []
            },
            'allocation': {
                'total_cost': 0.0,  # 总费用
                'tv_allocation_cost': 0.0,  # 电视（屏分摊费用）
                'computer_allocation_cost': 0.0  # 电脑（屏分摊费用）
            },
            'success': True
        }
        
        # ========== 一、收集直接人工数据 ==========
        
        # 1.1 获取计件工资数据，筛选类别为"屏"的
        piece_rate_result = calculate_piece_rate_wage(app_data)
        if piece_rate_result.get('part3_details'):
            for item in piece_rate_result['part3_details']:
                if item.get('类别') == '屏':
                    screen_wage = item.get('工资', 0.0)
                    result['direct_labor']['screen_wage'] += screen_wage
                    result['direct_labor']['screen_wage_details'].append(item)
        
        # 1.2 获取直接人工成本数据，筛选类别为"屏"的固定工资分摊
        direct_labor_result = calculate_direct_labor_cost(app_data, prediction_period)
        if direct_labor_result.get('category_details'):
            screen_category = direct_labor_result['category_details'].get('屏')
            if screen_category:
                screen_allocation = screen_category.get('total_fixed_cost', 0.0)
                result['direct_labor']['screen_allocation'] = screen_allocation
                result['direct_labor']['screen_allocation_details'] = screen_category.get('item_allocations', [])
        
        # ========== 二、收集制造费用数据 ==========
        
        manufacturing_result = calculate_manufacturing_cost(app_data, prediction_period)
        
        # 2.1 从"与拆解量相关的费用"中筛选类别为"屏"的成本
        if manufacturing_result.get('disassembly_related'):
            for item in manufacturing_result['disassembly_related']:
                for detail in item.get('明细', []):
                    if detail.get('category') == '屏':
                        cost = detail.get('cost', 0.0)
                        result['manufacturing_cost']['disassembly_related_screen'] += cost
                        result['manufacturing_cost']['disassembly_related_details'].append({
                            '费用类型': item.get('费用类型', ''),
                            '费用种类': item.get('费用种类', ''),
                            '费用名称': item.get('费用名称', ''),
                            '数量': detail.get('quantity', 0.0),
                            '单价': detail.get('unit_price', 0.0),
                            '成本': cost
                        })
        
        # 2.2 从"预计月均费用"中筛选类别为"屏"的成本
        if manufacturing_result.get('monthly_average'):
            for item in manufacturing_result['monthly_average']:
                for detail in item.get('明细', []):
                    if detail.get('category') == '屏':
                        cost = detail.get('cost', 0.0)
                        result['manufacturing_cost']['monthly_average_screen'] += cost
                        result['manufacturing_cost']['monthly_average_details'].append({
                            '费用类型': item.get('费用类型', ''),
                            '费用种类': item.get('费用种类', ''),
                            '费用名称': item.get('费用名称', ''),
                            '月均费用': detail.get('monthly_cost', 0.0),
                            '期数': detail.get('periods', 0),
                            '成本': cost
                        })
        
        # 2.3 从"环保费"中筛选拆解产物名称含"屏"的费用
        if manufacturing_result.get('environmental_fee'):
            environmental_fee_info = manufacturing_result['environmental_fee']
            if environmental_fee_info.get('明细'):
                for detail in environmental_fee_info['明细']:
                    product_name = str(detail.get('拆解产物名称', ''))
                    if '屏' in product_name:
                        fee = detail.get('费用', 0.0)
                        result['manufacturing_cost']['environmental_fee_screen'] += fee
                        result['manufacturing_cost']['environmental_fee_details'].append(detail)
        
        # ========== 三、处理被减扣数据 ==========
        
        deducted_data = app_data.get_data('deducted_data_manual')
        if deducted_data is not None and not deducted_data.empty:
            if '处置类别' in deducted_data.columns and '计算结果(KG)' in deducted_data.columns:
                # 筛选处置类别为"内转屏处置"的记录
                screen_disposal_mask = deducted_data['处置类别'].astype(str).str.contains('内转屏处置', case=False, na=False)
                screen_disposal_data = deducted_data[screen_disposal_mask].copy()
                
                if not screen_disposal_data.empty:
                    # 确保数值类型
                    screen_disposal_data['计算结果(KG)'] = pd.to_numeric(
                        screen_disposal_data['计算结果(KG)'], errors='coerce'
                    ).fillna(0)
                    
                    # 根据拆解产物名称进行分类（按匹配顺序）
                    if '拆解产物名称' in screen_disposal_data.columns:
                        for _, row in screen_disposal_data.iterrows():
                            product_name = str(row.get('拆解产物名称', ''))
                            weight = float(row['计算结果(KG)'])
                            
                            if weight > 0:
                                # 按匹配顺序判断：先检查是否含"电视"，再检查是否含"笔记本"或"显示器"
                                if '电视' in product_name:
                                    result['deducted_data']['tv_weight'] += weight
                                    result['deducted_data']['tv_details'].append({
                                        '拆解产物编码': row.get('拆解产物编码', ''),
                                        '拆解产物名称': product_name,
                                        '重量(KG)': weight
                                    })
                                elif '笔记本' in product_name or '显示器' in product_name:
                                    result['deducted_data']['computer_weight'] += weight
                                    result['deducted_data']['computer_details'].append({
                                        '拆解产物编码': row.get('拆解产物编码', ''),
                                        '拆解产物名称': product_name,
                                        '重量(KG)': weight
                                    })
        
        # 计算总重量
        result['deducted_data']['total_weight'] = (
            result['deducted_data']['tv_weight'] + 
            result['deducted_data']['computer_weight']
        )
        
        # ========== 四、制造费用间接人工分摊 ==========
        
        # 初始化制造费用间接人工分摊结果
        result['indirect_labor_allocation'] = {
            'total_cost': 0.0,
            'category_totals': {
                '冰箱': 0.0,
                '空调': 0.0,
                '电脑': 0.0,
                '电视': 0.0,
                '洗衣机': 0.0
            },
            'details': []
        }
        
        try:
            # 获取一次拆解产物产值数据
            disassembly_data = app_data.get_data('disassembly_data')
            if disassembly_data is not None and not disassembly_data.empty:
                # 筛选类别为"拆解产物"的记录
                if '类别' in disassembly_data.columns:
                    product_data = disassembly_data[disassembly_data['类别'] == '拆解产物'].copy()
                    
                    if not product_data.empty:
                        # 获取价格数据
                        from data.base_data.price_data import load_price_data
                        price_df = load_price_data()
                        
                        if price_df is not None and not price_df.empty:
                            # 创建价格映射
                            price_mapping = {}
                            for _, price_row in price_df.iterrows():
                                code = str(price_row['拆解产物编码']).strip()
                                price_no_tax = price_row.get('销售单价-不含税(元/KG)', 0)
                                if pd.notna(price_no_tax):
                                    price_mapping[code] = float(price_no_tax)
                            
                            # 分类关键词映射规则
                            # 电视：电视、彩电、CRT其它机壳破碎塑料、线路板边框破碎塑料、等离子、废旧玻璃电子枪、废旧金属荫罩压块铁、黑白
                            # 电脑：电脑、显示器、笔记本、主机、废旧金属黑色金属-铁及其合金-电子枪
                            # 冰箱：冰箱、冰柜
                            # 空调：空调
                            # 洗衣机：洗衣机、双缸
                            category_keyword_mapping = {
                                '电视': ['电视', '彩电', 'CRT其它机壳破碎塑料', '线路板边框破碎塑料', '等离子', '废旧玻璃电子枪', '废旧金属荫罩压块铁', '黑白'],
                                '电脑': ['电脑', '显示器', '笔记本', '主机', '废旧金属黑色金属-铁及其合金-电子枪'],
                                '冰箱': ['冰箱', '冰柜'],
                                '空调': ['空调'],
                                '洗衣机': ['洗衣机', '双缸']
                            }
                            
                            # 处理数据：添加分类、匹配价格、计算产值
                            output_value_data = []
                            for idx, row in product_data.iterrows():
                                # 获取原物料名称
                                material_name = str(row.get('原物料名称', '')).strip()
                                
                                # 根据原物料名称进行模糊匹配分类
                                category = None
                                for cat, keywords in category_keyword_mapping.items():
                                    for keyword in keywords:
                                        if keyword in material_name:
                                            category = cat
                                            break
                                    if category:
                                        break
                                
                                # 如果没有匹配到任何分类，跳过该记录
                                if not category:
                                    continue
                                
                                # 获取拆解产物编码
                                product_code = str(row.get('拆解产物编码', '')).strip()
                                
                                # 获取计算结果(KG)
                                calculated_weight = row.get('计算结果(KG)', 0)
                                try:
                                    calculated_weight = float(calculated_weight) if pd.notna(calculated_weight) else 0
                                except (ValueError, TypeError):
                                    calculated_weight = 0
                                
                                # 匹配价格
                                price_no_tax = price_mapping.get(product_code, 0)
                                
                                # 计算物料产值
                                if price_no_tax < 0:
                                    material_value = calculated_weight * 0
                                else:
                                    material_value = calculated_weight * price_no_tax
                                
                                output_value_data.append({
                                    '原物料名称': material_name,
                                    '分类': category,
                                    '拆解产物编码': product_code,
                                    '拆解产物名称': str(row.get('拆解产物名称', '')).strip(),
                                    '计算结果(KG)': calculated_weight,
                                    '销售单价-不含税(元/KG)': price_no_tax,
                                    '物料产值（元）': material_value
                                })
                            
                            # 计算总产值和各类别总产值
                            total_output_value = sum(item['物料产值（元）'] for item in output_value_data)
                            tv_computer_total = sum(item['物料产值（元）'] for item in output_value_data 
                                                   if item['分类'] in ['电视', '电脑'])
                            ac_wm_total = sum(item['物料产值（元）'] for item in output_value_data 
                                             if item['分类'] in ['空调', '洗衣机'])
                            
                            # 获取间接人工成本数据
                            indirect_labor_result = calculate_indirect_labor_cost(app_data, prediction_period, include_no_opening_columns=False)
                            other_positions_details = indirect_labor_result.get('other_positions_details', [])
                            
                            # 定义需要分摊的岗位列表（11个特定岗位 + 黑电保洁 + 白电保洁）
                            specific_positions = [
                                '生产工艺管理主管',
                                '客服专员',
                                '设备维修副经理',
                                '生产管理副高级经理',
                                '视频监控员',
                                '库管(兼叉车)',
                                '库管（兼叉车）',  # 兼容全角括号
                                '计数员兼库房协管',
                                '原材料库房管理员',
                                '数据统计',
                                '物流保洁',
                                '设备维护储备岗'
                            ]
                            
                            # 处理11个特定岗位的分摊
                            for position_detail in other_positions_details:
                                position = position_detail.get('岗位', '')
                                position_cost = position_detail.get('岗位成本', 0.0)
                                
                                if position_cost <= 0:
                                    continue
                                
                                # 检查是否是11个特定岗位之一（兼容库管岗位的两种括号格式）
                                is_specific_position = False
                                for spec_pos in specific_positions:
                                    if position == spec_pos:
                                        is_specific_position = True
                                        break
                                    # 特殊处理：库管(兼叉车)和库管（兼叉车）视为同一岗位
                                    if (spec_pos in ['库管(兼叉车)', '库管（兼叉车）'] and 
                                        position in ['库管(兼叉车)', '库管（兼叉车）']):
                                        is_specific_position = True
                                        break
                                
                                if is_specific_position and total_output_value > 0:
                                    # 对于11个特定岗位：每个物料的物料产值 / 总产值 × 岗位成本
                                    for item in output_value_data:
                                        material_value = item['物料产值（元）']
                                        allocation_ratio = material_value / total_output_value if total_output_value > 0 else 0
                                        allocation_cost = allocation_ratio * position_cost
                                        
                                        category = item['分类']
                                        if category in result['indirect_labor_allocation']['category_totals']:
                                            result['indirect_labor_allocation']['category_totals'][category] += allocation_cost
                                        
                                        result['indirect_labor_allocation']['details'].append({
                                            '岗位': position,
                                            '拆解产物编码': item['拆解产物编码'],
                                            '拆解产物名称': item['拆解产物名称'],
                                            '分类': category,
                                            '物料产值（元）': material_value,
                                            '分摊比例': allocation_ratio,
                                            '分摊成本（元）': allocation_cost
                                        })
                                
                                # 处理黑电保洁：仅限电视、电脑类别
                                elif position == '黑电保洁' and tv_computer_total > 0:
                                    for item in output_value_data:
                                        category = item['分类']
                                        if category in ['电视', '电脑']:
                                            material_value = item['物料产值（元）']
                                            allocation_ratio = material_value / tv_computer_total if tv_computer_total > 0 else 0
                                            allocation_cost = allocation_ratio * position_cost
                                            
                                            result['indirect_labor_allocation']['category_totals'][category] += allocation_cost
                                            
                                            result['indirect_labor_allocation']['details'].append({
                                                '岗位': position,
                                                '拆解产物编码': item['拆解产物编码'],
                                                '拆解产物名称': item['拆解产物名称'],
                                                '分类': category,
                                                '物料产值（元）': material_value,
                                                '分摊比例': allocation_ratio,
                                                '分摊成本（元）': allocation_cost
                                            })
                                
                                # 处理白电保洁：仅限空调、洗衣机类别
                                elif position == '白电保洁' and ac_wm_total > 0:
                                    for item in output_value_data:
                                        category = item['分类']
                                        if category in ['空调', '洗衣机']:
                                            material_value = item['物料产值（元）']
                                            allocation_ratio = material_value / ac_wm_total if ac_wm_total > 0 else 0
                                            allocation_cost = allocation_ratio * position_cost
                                            
                                            result['indirect_labor_allocation']['category_totals'][category] += allocation_cost
                                            
                                            result['indirect_labor_allocation']['details'].append({
                                                '岗位': position,
                                                '拆解产物编码': item['拆解产物编码'],
                                                '拆解产物名称': item['拆解产物名称'],
                                                '分类': category,
                                                '物料产值（元）': material_value,
                                                '分摊比例': allocation_ratio,
                                                '分摊成本（元）': allocation_cost
                                            })
                            
                            # 计算总分摊成本
                            result['indirect_labor_allocation']['total_cost'] = sum(
                                result['indirect_labor_allocation']['category_totals'].values()
                            )
                            
        except Exception as e:
            print(f"[制造费用间接人工分摊] 计算失败: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # ========== 五、制造费用公共成本分摊 ==========
        
        # 初始化制造费用公共成本分摊结果
        result['public_cost_allocation'] = {
            'total_cost': 0.0,
            'category_totals': {
                '冰箱': 0.0,
                '空调': 0.0,
                '电脑': 0.0,
                '电视': 0.0,
                '洗衣机': 0.0
            },
            'details': []
        }
        
        try:
            # 获取一次拆解产物产值数据（复用上面的逻辑）
            disassembly_data = app_data.get_data('disassembly_data')
            if disassembly_data is not None and not disassembly_data.empty:
                if '类别' in disassembly_data.columns:
                    product_data = disassembly_data[disassembly_data['类别'] == '拆解产物'].copy()
                    
                    if not product_data.empty:
                        # 获取价格数据
                        from data.base_data.price_data import load_price_data
                        price_df = load_price_data()
                        
                        if price_df is not None and not price_df.empty:
                            # 创建价格映射
                            price_mapping = {}
                            for _, price_row in price_df.iterrows():
                                code = str(price_row['拆解产物编码']).strip()
                                price_no_tax = price_row.get('销售单价-不含税(元/KG)', 0)
                                if pd.notna(price_no_tax):
                                    price_mapping[code] = float(price_no_tax)
                            
                            # 分类关键词映射规则
                            # 电视：电视、彩电、CRT其它机壳破碎塑料、线路板边框破碎塑料、等离子、废旧玻璃电子枪、废旧金属荫罩压块铁、黑白
                            # 电脑：电脑、显示器、笔记本、主机、废旧金属黑色金属-铁及其合金-电子枪
                            # 冰箱：冰箱、冰柜
                            # 空调：空调
                            # 洗衣机：洗衣机、双缸
                            category_keyword_mapping = {
                                '电视': ['电视', '彩电', 'CRT其它机壳破碎塑料', '线路板边框破碎塑料', '等离子', '废旧玻璃电子枪', '废旧金属荫罩压块铁', '黑白'],
                                '电脑': ['电脑', '显示器', '笔记本', '主机', '废旧金属黑色金属-铁及其合金-电子枪'],
                                '冰箱': ['冰箱', '冰柜'],
                                '空调': ['空调'],
                                '洗衣机': ['洗衣机', '双缸']
                            }
                            
                            # 处理数据：添加分类、匹配价格、计算产值
                            output_value_data = []
                            for idx, row in product_data.iterrows():
                                # 获取原物料名称
                                material_name = str(row.get('原物料名称', '')).strip()
                                
                                # 根据原物料名称进行模糊匹配分类
                                category = None
                                for cat, keywords in category_keyword_mapping.items():
                                    for keyword in keywords:
                                        if keyword in material_name:
                                            category = cat
                                            break
                                    if category:
                                        break
                                
                                # 如果没有匹配到任何分类，跳过该记录
                                if not category:
                                    continue
                                
                                # 获取拆解产物编码
                                product_code = str(row.get('拆解产物编码', '')).strip()
                                
                                # 获取计算结果(KG)
                                calculated_weight = row.get('计算结果(KG)', 0)
                                try:
                                    calculated_weight = float(calculated_weight) if pd.notna(calculated_weight) else 0
                                except (ValueError, TypeError):
                                    calculated_weight = 0
                                
                                # 匹配价格
                                price_no_tax = price_mapping.get(product_code, 0)
                                
                                # 计算物料产值
                                if price_no_tax < 0:
                                    material_value = calculated_weight * 0
                                else:
                                    material_value = calculated_weight * price_no_tax
                                
                                output_value_data.append({
                                    '原物料名称': material_name,
                                    '分类': category,
                                    '拆解产物编码': product_code,
                                    '拆解产物名称': str(row.get('拆解产物名称', '')).strip(),
                                    '计算结果(KG)': calculated_weight,
                                    '销售单价-不含税(元/KG)': price_no_tax,
                                    '物料产值（元）': material_value
                                })
                            
                            # 计算总产值
                            total_output_value = sum(item['物料产值（元）'] for item in output_value_data)

                            def _allocate_public_manufacturing_pool(public_cost, meta_item):
                                """按一次拆解产物产值占比分摊「公共」池（预计月均或拆解量相关明细 category=公共）。"""
                                try:
                                    pc = float(public_cost or 0)
                                except (TypeError, ValueError):
                                    pc = 0.0
                                if pc <= 0 or total_output_value <= 0:
                                    return
                                for output_item in output_value_data:
                                    material_value = output_item['物料产值（元）']
                                    allocation_ratio = material_value / total_output_value if total_output_value > 0 else 0
                                    allocation_cost = allocation_ratio * pc
                                    category = output_item['分类']
                                    if category in result['public_cost_allocation']['category_totals']:
                                        result['public_cost_allocation']['category_totals'][category] += allocation_cost
                                    result['public_cost_allocation']['details'].append({
                                        '费用类型': meta_item.get('费用类型', ''),
                                        '费用种类': meta_item.get('费用种类', ''),
                                        '费用名称': meta_item.get('费用名称', ''),
                                        '拆解产物编码': output_item['拆解产物编码'],
                                        '拆解产物名称': output_item['拆解产物名称'],
                                        '分类': category,
                                        '物料产值（元）': material_value,
                                        '分摊比例': allocation_ratio,
                                        '分摊成本（元）': allocation_cost
                                    })

                            # 预计月均费用中明细类别为「公共」的
                            if manufacturing_result.get('monthly_average'):
                                for item in manufacturing_result['monthly_average']:
                                    for detail in item.get('明细', []):
                                        if detail.get('category') == '公共':
                                            _allocate_public_manufacturing_pool(detail.get('cost', 0.0), item)

                            # 与拆解量相关费用中明细类别为「公共」的（如叉车费仅维护公共单价）
                            if manufacturing_result.get('disassembly_related'):
                                for item in manufacturing_result['disassembly_related']:
                                    for detail in item.get('明细', []):
                                        if detail.get('category') == '公共':
                                            _allocate_public_manufacturing_pool(detail.get('cost', 0.0), item)
                            
                            # 计算总分摊成本
                            result['public_cost_allocation']['total_cost'] = sum(
                                result['public_cost_allocation']['category_totals'].values()
                            )
                            
        except Exception as e:
            print(f"[制造费用公共成本分摊] 计算失败: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # ========== 六、计算分摊费用 ==========
        
        # 计算各项费用
        direct_labor_total = (
            result['direct_labor']['screen_wage'] +
            result['direct_labor']['screen_allocation']
        )
        
        manufacturing_cost_screen_total = (
            result['manufacturing_cost']['disassembly_related_screen'] +
            result['manufacturing_cost']['monthly_average_screen'] +
            result['manufacturing_cost']['environmental_fee_screen']
        )
        
        indirect_labor_allocation_total = result['indirect_labor_allocation'].get('total_cost', 0.0)
        public_cost_allocation_total = result['public_cost_allocation'].get('total_cost', 0.0)
        
        # 计算屏费用分摊结果（直接人工 + 制造费用（屏）按重量比例分摊到电视和电脑）
        screen_cost_total = direct_labor_total + manufacturing_cost_screen_total
        
        # 总费用 = 制造费用间接人工分摊 + 制造费用公共成本分摊 + 屏费用分摊结果
        total_cost = (
            indirect_labor_allocation_total +
            public_cost_allocation_total +
            screen_cost_total
        )
        total_weight = result['deducted_data']['total_weight']
        tv_screen_allocation = 0.0
        computer_screen_allocation = 0.0
        
        if total_weight > 0:
            tv_weight = result['deducted_data']['tv_weight']
            computer_weight = result['deducted_data']['computer_weight']
            
            tv_ratio = tv_weight / total_weight
            computer_ratio = computer_weight / total_weight
            
            tv_screen_allocation = screen_cost_total * tv_ratio
            computer_screen_allocation = screen_cost_total * computer_ratio
        
        # 重新设计分摊计算结果数据结构
        result['allocation'] = {
            'total_cost': total_cost,
            'cost_details': {
                '直接人工': {
                    '屏工资': result['direct_labor']['screen_wage'],
                    '屏分摊': result['direct_labor']['screen_allocation'],
                    '小计': direct_labor_total
                },
                '制造费用（屏）': {
                    '与拆解量相关的费用': result['manufacturing_cost']['disassembly_related_screen'],
                    '预计月均费用': result['manufacturing_cost']['monthly_average_screen'],
                    '环保费': result['manufacturing_cost']['environmental_fee_screen'],
                    '小计': manufacturing_cost_screen_total
                },
                '制造费用间接人工分摊': indirect_labor_allocation_total,
                '制造费用公共成本分摊': public_cost_allocation_total
            },
            'screen_allocation': {
                'total_cost': screen_cost_total,
                'tv_allocation': tv_screen_allocation,
                'computer_allocation': computer_screen_allocation
            },
            'category_allocation': {
                '冰箱': 0.0,
                '空调': 0.0,
                '电脑': 0.0,
                '电视': 0.0,
                '洗衣机': 0.0
            }
        }
        
        # 计算按类别分摊（基于制造费用间接人工分摊和公共成本分摊的类别汇总 + 屏费用分摊结果）
        indirect_category_totals = result['indirect_labor_allocation'].get('category_totals', {})
        public_category_totals = result['public_cost_allocation'].get('category_totals', {})
        
        for category in ['冰箱', '空调', '电脑', '电视', '洗衣机']:
            # 按类别分摊 = 间接人工分摊类别汇总 + 公共成本分摊类别汇总 + 屏费用分摊结果（仅限电视和电脑）
            category_total = 0.0
            
            # 间接人工分摊和公共成本分摊已经按类别汇总
            category_total += indirect_category_totals.get(category, 0.0)
            category_total += public_category_totals.get(category, 0.0)
            
            # 屏费用分摊结果只分摊到电视和电脑
            if category == '电视':
                category_total += tv_screen_allocation
            elif category == '电脑':
                category_total += computer_screen_allocation
            
            result['allocation']['category_allocation'][category] = category_total
        
        return result
        
    except Exception as e:
        print(f"计算公共费用分摊明细失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'error': str(e)
        }


@cost_forecast_bp.route('/screen-cost-allocation', methods=['GET'])
def get_screen_cost_allocation():
    """获取公共费用分摊明细数据"""
    try:
        app_data = get_session_data_manager()
        prediction_period = int(request.args.get('prediction_period', 1))
        force_recalculate = request.args.get('force_recalculate', 'false').lower() == 'true'
        
        # 检查数据是否已被清除
        data_cleared = app_data.get_data('__data_cleared__')
        if data_cleared:
            # 数据已被清除，直接返回空数据
            return jsonify({
                'success': True,
                'data': {
                    'allocation': {
                        'total_cost': 0.0,
                        'cost_details': {},
                        'screen_allocation': {
                            'total_cost': 0.0
                        }
                    }
                }
            })
        
        # 检查是否有缓存的计算结果
        cache_key = f'screen_cost_allocation_result_v2_{prediction_period}'
        if not force_recalculate:
            cached_result = app_data.get_data(cache_key)
            if cached_result is not None:
                print(f"[公共费用分摊明细] 使用缓存数据（预测期数: {prediction_period}）")
                return jsonify({
                    'success': True,
                    'data': cached_result,
                    'from_cache': True
                })
        
        # 没有缓存或强制重新计算，执行计算
        print(f"[公共费用分摊明细] 开始计算（预测期数: {prediction_period}, 强制重新计算: {force_recalculate}）")
        result = calculate_screen_cost_allocation(app_data, prediction_period)
        
        if not result.get('success', False):
            return jsonify({
                'success': False,
                'error': result.get('error', '计算失败')
            }), 500
        
        # 缓存计算结果
        app_data.set_data(cache_key, result)
        print(f"[公共费用分摊明细] 计算结果已缓存（预测期数: {prediction_period}）")
        
        return jsonify({
            'success': True,
            'data': result,
            'from_cache': False
        })
        
    except Exception as e:
        print(f"获取公共费用分摊明细失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


def calculate_period_cost(app_data, prediction_period=1, quality_manager_ratio=None, quality_group_ratio=None, warehouse_group_ratio=None):
    """
    计算期间费用
    
    Args:
        app_data: 应用数据管理器
        prediction_period: 预测期数
        quality_manager_ratio: 质量管理经理薪酬分摊比例（默认从localStorage或0.7）
        quality_group_ratio: 质量组费用分摊比例（默认从localStorage或0.9）
        warehouse_group_ratio: 库房组费用分摊比例（默认从localStorage或0.9）
    
    Returns:
        dict: 包含期间费用数据的字典
    """
    try:
        # 默认分摊比例
        if quality_manager_ratio is None:
            quality_manager_ratio = 0.7
        if quality_group_ratio is None:
            quality_group_ratio = 0.9
        if warehouse_group_ratio is None:
            warehouse_group_ratio = 0.9
        
        # 获取间接人工成本数据
        indirect_labor_result = calculate_indirect_labor_cost(app_data, prediction_period, include_no_opening_columns=False)
        
        # 获取其他岗位成本明细
        other_positions_details = indirect_labor_result.get('other_positions_details', [])
        
        # 创建岗位到成本的映射
        position_cost_map = {}
        for detail in other_positions_details:
            position = detail.get('岗位', '')
            cost = detail.get('岗位成本', 0.0)
            position_cost_map[position] = cost
        
        # 获取提成和分摊数据
        category_totals = indirect_labor_result.get('category_totals', {})
        category_fixed_costs = indirect_labor_result.get('category_fixed_costs', {})
        
        qc_commission = category_totals.get('品管提成', 0.0)
        forklift_commission = category_totals.get('叉车司磅库管等提成', 0.0)
        qc_fixed_cost = category_fixed_costs.get('品管提成', 0.0)
        forklift_fixed_cost = category_fixed_costs.get('叉车司磅库管等提成', 0.0)
        
        # 计算薪酬费用行
        salary_row = {
            '费用明细': '薪酬费用',
            '管理-电废事业部-质量管理': 0.0,
            '管理-电废事业部-库房': 0.0,
            '管理-电废事业部-平台': 0.0,
            '管理-电废事业部-基金管理项目组': 0.0,
            '管理-电废事业部-回收经营项目组': 0.0,
            '管理-消电项目组': 0.0,
            '销售-平台': 0.0,
            '销售-消电项目组': 0.0,
            '屏': 0.0,
        }
        
        # 管理-电废事业部-质量管理（薪酬费用）
        quality_manager_cost = position_cost_map.get('质量管理经理', 0.0)
        salary_row['管理-电废事业部-质量管理'] = (
            (quality_manager_cost * quality_manager_ratio + qc_commission + qc_fixed_cost) * quality_group_ratio
        )
        
        # 管理-电废事业部-库房（薪酬费用）
        warehouse_cost = position_cost_map.get('库房', 0.0)
        salary_row['管理-电废事业部-库房'] = (
            (warehouse_cost + forklift_commission + forklift_fixed_cost) * warehouse_group_ratio
        )
        
        # 管理-电废事业部-平台（薪酬费用）
        salary_row['管理-电废事业部-平台'] = position_cost_map.get('电废综合管理人员', 0.0)
        
        # 管理-电废事业部-基金管理项目组（薪酬费用）
        salary_row['管理-电废事业部-基金管理项目组'] = position_cost_map.get('基金管理项目组', 0.0)
        
        # 管理-电废事业部-回收经营项目组（薪酬费用）
        salary_row['管理-电废事业部-回收经营项目组'] = position_cost_map.get('旧机回收', 0.0)
        
        # 销售-平台（薪酬费用）
        salary_row['销售-平台'] = position_cost_map.get('销售组人员', 0.0)
        
        # 获取期间费用基础数据
        period_cost_df = get_period_cost_dataframe()
        
        # 定义列名
        columns = [
            '费用明细',
            '管理-电废事业部-质量管理',
            '管理-电废事业部-库房',
            '管理-电废事业部-平台',
            '管理-电废事业部-基金管理项目组',
            '管理-电废事业部-回收经营项目组',
            '管理-消电项目组',
            '销售-平台',
            '销售-消电项目组',
            '屏',
        ]
        
        # 处理期间费用基础数据，应用分摊比例并乘以预测期数
        period_cost_data = []
        if period_cost_df is not None and not period_cost_df.empty:
            for _, row in period_cost_df.iterrows():
                row_data = {'费用明细': row.get('费用明细', '')}
                
                for col in columns[1:]:  # 跳过费用明细列
                    value = row.get(col)
                    if pd.isna(value) or value is None:
                        row_data[col] = None
                    else:
                        value = float(value)
                        # 对"管理-电废事业部-质量管理"列（除薪酬费用外）应用质量组费用分摊比例
                        if col == '管理-电废事业部-质量管理':
                            row_data[col] = value * quality_group_ratio * prediction_period
                        # 对"管理-电废事业部-库房"列（除薪酬费用外）应用库房组费用分摊比例
                        elif col == '管理-电废事业部-库房':
                            row_data[col] = value * warehouse_group_ratio * prediction_period
                        else:
                            row_data[col] = value * prediction_period
                
                period_cost_data.append(row_data)
        
        # 计算各列总计（包括薪酬费用行和期间费用基础数据）
        totals = {}
        for col in columns[1:]:
            total = salary_row.get(col, 0.0) or 0.0
            for row_data in period_cost_data:
                value = row_data.get(col)
                if value is not None and not pd.isna(value):
                    total += float(value)
            totals[col] = float(total)
        
        # 计算总费用（所有列的总和）
        total_cost = sum(totals.values())
        
        return {
            'success': True,
            'salary_row': salary_row,
            'period_cost_data': period_cost_data,
            'totals': totals,
            'total_cost': float(total_cost)
        }
        
    except Exception as e:
        print(f"[期间费用计算] 计算失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'error': str(e),
            'salary_row': {},
            'period_cost_data': [],
            'totals': {},
            'total_cost': 0.0
        }


TAX_SURCHARGE_CATEGORIES = ['冰箱', '空调', '电脑', '电视', '洗衣机']
TAX_SURCHARGE_ROW_PROPERTY = '房产税'
TAX_SURCHARGE_ROW_LAND = '土地使用税'
TAX_SURCHARGE_ROW_RATE = '印花税、环保税、城建税及教育费附加'
TAX_SURCHARGE_ROW_TOTAL = '合计'


def parse_tax_rate(value):
    """解析税金及附加费率：0.77% -> 0.0077，数值型原样返回"""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0.0
    if isinstance(value, (int, float)):
        v = float(value)
        return v / 100.0 if v > 1 else v
    s = str(value).strip()
    if not s:
        return 0.0
    if s.endswith('%'):
        try:
            return float(s[:-1].strip()) / 100.0
        except ValueError:
            return 0.0
    try:
        v = float(s)
        return v / 100.0 if v > 1 else v
    except ValueError:
        return 0.0


def _get_tax_surcharge_base_amounts():
    """从基础数据读取房产税、土地使用税月均金额及综合费率"""
    property_tax_monthly = 0.0
    land_tax_monthly = 0.0
    rate = 0.0

    df = get_tax_surcharge_dataframe()
    if df is None or df.empty:
        return property_tax_monthly, land_tax_monthly, rate

    for _, row in df.iterrows():
        item = str(row.get('项目', '') or '').strip()
        amount = row.get('金额')
        if item == TAX_SURCHARGE_ROW_PROPERTY:
            if isinstance(amount, (int, float)) and not pd.isna(amount):
                property_tax_monthly = float(amount)
        elif item == TAX_SURCHARGE_ROW_LAND:
            if isinstance(amount, (int, float)) and not pd.isna(amount):
                land_tax_monthly = float(amount)
        elif item == TAX_SURCHARGE_ROW_RATE:
            rate = parse_tax_rate(amount)

    return property_tax_monthly, land_tax_monthly, rate


def calculate_tax_surcharge(app_data, prediction_period=1):
    """
    计算税金及附加（按分类列示）
    房产税/土地使用税：月均 × 预测期数 × 产值占比
    印花税费率行：营业收入 × 费率
    """
    try:
        from app.api.statistics_api import (
            get_disassembly_output_value_ratios,
            get_category_operating_revenue,
        )

        prediction_period = max(1, int(prediction_period or 1))
        categories = TAX_SURCHARGE_CATEGORIES

        property_tax_monthly, land_tax_monthly, rate = _get_tax_surcharge_base_amounts()
        ratios, category_values, total_output_value = get_disassembly_output_value_ratios(
            app_data, categories
        )
        operating_revenue, _, _ = get_category_operating_revenue(app_data, categories)

        def _fixed_tax_row(monthly_amount):
            values = {}
            for cat in categories:
                values[cat] = round(monthly_amount * prediction_period * ratios.get(cat, 0.0), 2)
            row_total = round(sum(values.values()), 2)
            return values, row_total

        property_values, property_row_total = _fixed_tax_row(property_tax_monthly)
        land_values, land_row_total = _fixed_tax_row(land_tax_monthly)

        rate_values = {}
        for cat in categories:
            rate_values[cat] = round(operating_revenue.get(cat, 0.0) * rate, 2)
        rate_row_total = round(sum(rate_values.values()), 2)

        total_values = {}
        for cat in categories:
            total_values[cat] = round(
                property_values[cat] + land_values[cat] + rate_values[cat], 2
            )
        total_row_total = round(sum(total_values.values()), 2)
        grand_total = total_row_total

        rows = [
            {'项目': TAX_SURCHARGE_ROW_PROPERTY, 'values': property_values, 'row_total': property_row_total},
            {'项目': TAX_SURCHARGE_ROW_LAND, 'values': land_values, 'row_total': land_row_total},
            {'项目': TAX_SURCHARGE_ROW_RATE, 'values': rate_values, 'row_total': rate_row_total},
            {'项目': TAX_SURCHARGE_ROW_TOTAL, 'values': total_values, 'row_total': total_row_total},
        ]

        return {
            'success': True,
            'categories': categories,
            'rows': rows,
            'grand_total': grand_total,
            'allocation_ratios': {k: round(v, 6) for k, v in ratios.items()},
            'category_output_values': {k: round(v, 2) for k, v in category_values.items()},
            'total_output_value': round(total_output_value, 2),
            'operating_revenue': {k: round(v, 2) for k, v in operating_revenue.items()},
            'prediction_period': prediction_period,
            'base': {
                'property_tax_monthly': property_tax_monthly,
                'land_tax_monthly': land_tax_monthly,
                'rate': rate,
            },
        }
    except Exception as e:
        print(f"[税金及附加计算] 失败: {e}")
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'error': str(e),
            'categories': TAX_SURCHARGE_CATEGORIES,
            'rows': [],
            'grand_total': 0.0,
        }


@cost_forecast_bp.route('/tax-surcharge', methods=['GET'])
def get_tax_surcharge():
    """获取税金及附加计算结果（首页卡片）"""
    try:
        app_data = get_session_data_manager()
        prediction_period = int(request.args.get('prediction_period', 1))

        data_cleared = app_data.get_data('__data_cleared__')
        if data_cleared:
            return jsonify({
                'success': True,
                'data': {
                    'categories': TAX_SURCHARGE_CATEGORIES,
                    'rows': [],
                    'grand_total': 0.0,
                    'allocation_ratios': {},
                },
                'message': '数据已清除',
            })

        result = calculate_tax_surcharge(app_data, prediction_period)
        if not result.get('success', False):
            return jsonify({
                'success': False,
                'error': result.get('error', '计算失败'),
            }), 500

        return jsonify({
            'success': True,
            'data': {
                'categories': result.get('categories', []),
                'rows': result.get('rows', []),
                'grand_total': result.get('grand_total', 0.0),
                'allocation_ratios': result.get('allocation_ratios', {}),
                'category_output_values': result.get('category_output_values', {}),
                'total_output_value': result.get('total_output_value', 0.0),
                'operating_revenue': result.get('operating_revenue', {}),
                'prediction_period': result.get('prediction_period', 1),
                'base': result.get('base', {}),
            },
        })
    except Exception as e:
        print(f"获取税金及附加失败: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@cost_forecast_bp.route('/period-cost', methods=['GET'])
def get_period_cost():
    """获取期间费用数据"""
    try:
        app_data = get_session_data_manager()
        prediction_period = int(request.args.get('prediction_period', 1))
        
        # 从请求参数获取分摊比例（如果前端传递了的话）
        quality_manager_ratio = request.args.get('quality_manager_ratio')
        quality_group_ratio = request.args.get('quality_group_ratio')
        warehouse_group_ratio = request.args.get('warehouse_group_ratio')
        
        if quality_manager_ratio is not None:
            quality_manager_ratio = float(quality_manager_ratio)
        if quality_group_ratio is not None:
            quality_group_ratio = float(quality_group_ratio)
        if warehouse_group_ratio is not None:
            warehouse_group_ratio = float(warehouse_group_ratio)
        
        result = calculate_period_cost(
            app_data, 
            prediction_period,
            quality_manager_ratio,
            quality_group_ratio,
            warehouse_group_ratio
        )
        
        if not result.get('success', False):
            return jsonify({
                'success': False,
                'error': result.get('error', '计算失败')
            }), 500
        
        return jsonify({
            'success': True,
            'data': {
                'salary_row': result.get('salary_row', {}),
                'period_cost_data': result.get('period_cost_data', []),
                'totals': result.get('totals', {}),
                'total_cost': result.get('total_cost', 0.0)
            }
        })
        
    except Exception as e:
        print(f"获取期间费用失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@cost_forecast_bp.route('/period-cost/export', methods=['GET'])
def export_period_cost():
    """导出期间费用数据到Excel"""
    try:
        app_data = get_session_data_manager()
        prediction_period = int(request.args.get('prediction_period', 1))
        
        # 从请求参数获取分摊比例
        quality_manager_ratio = request.args.get('quality_manager_ratio')
        quality_group_ratio = request.args.get('quality_group_ratio')
        warehouse_group_ratio = request.args.get('warehouse_group_ratio')
        
        if quality_manager_ratio is not None:
            quality_manager_ratio = float(quality_manager_ratio)
        if quality_group_ratio is not None:
            quality_group_ratio = float(quality_group_ratio)
        if warehouse_group_ratio is not None:
            warehouse_group_ratio = float(warehouse_group_ratio)
        
        # 计算期间费用数据
        result = calculate_period_cost(
            app_data, 
            prediction_period,
            quality_manager_ratio,
            quality_group_ratio,
            warehouse_group_ratio
        )
        
        if not result.get('success', False):
            return jsonify({
                'success': False,
                'error': result.get('error', '计算失败')
            }), 500
        
        # 构建导出数据
        export_data = []
        
        # 定义列名（合计前为金额列）
        columns = [
            '费用明细',
            '管理-电废事业部-质量管理',
            '管理-电废事业部-库房',
            '管理-电废事业部-平台',
            '管理-电废事业部-基金管理项目组',
            '管理-电废事业部-回收经营项目组',
            '管理-消电项目组',
            '销售-平台',
            '销售-消电项目组',
            '屏',
            '合计',
        ]
        value_columns = columns[1:-1]

        # 添加薪酬费用行
        if result.get('salary_row'):
            salary_row = result['salary_row'].copy()
            salary_row['费用明细'] = '薪酬费用'
            row_total = sum(float(salary_row.get(col, 0) or 0) for col in value_columns)
            salary_row['合计'] = row_total
            export_data.append(salary_row)

        # 添加期间费用基础数据
        if result.get('period_cost_data'):
            for row in result['period_cost_data']:
                row_data = row.copy()
                row_total = sum(
                    float(row_data.get(col, 0) or 0) if row_data.get(col) is not None else 0
                    for col in value_columns
                )
                row_data['合计'] = row_total
                export_data.append(row_data)

        # 添加财务费用行
        finance_row = {'费用明细': '财务费用'}
        for col in value_columns:
            finance_row[col] = None
        finance_row['合计'] = 0
        export_data.append(finance_row)

        # 添加合计行
        if result.get('totals'):
            total_row = {'费用明细': '合计'}
            column_total_sum = 0
            for col in value_columns:
                total_value = result['totals'].get(col, 0) or 0
                total_row[col] = float(total_value)
                column_total_sum += float(total_value)
            total_row['合计'] = column_total_sum
            export_data.append(total_row)
        
        if not export_data:
            return jsonify({
                'success': False,
                'error': '没有可导出的期间费用数据'
            }), 400
        
        # 转换为DataFrame
        export_df = pd.DataFrame(export_data)
        
        # 确保列顺序
        export_df = export_df[columns]
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 写入数据
            export_df.to_excel(writer, sheet_name='期间费用', index=False)
            
            # 设置列宽和样式
            worksheet = writer.sheets['期间费用']
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF", name="仿宋")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # 设置表头
            for col in range(1, len(export_df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                
                # 设置列宽
                col_letter = get_column_letter(col)
                col_name = export_df.columns[col - 1]
                if col_name == '费用明细':
                    worksheet.column_dimensions[col_letter].width = 30
                elif col_name == '合计':
                    worksheet.column_dimensions[col_letter].width = 18
                else:
                    worksheet.column_dimensions[col_letter].width = 25
            
            # 设置数据行样式
            data_font = Font(name="仿宋")
            right_alignment = Alignment(horizontal="right", vertical="center")
            left_alignment = Alignment(horizontal="left", vertical="center")
            
            for row_idx in range(2, len(export_df) + 2):
                row_data = export_df.iloc[row_idx - 2]
                is_salary_row = row_data['费用明细'] == '薪酬费用'
                is_finance_row = row_data['费用明细'] == '财务费用'
                is_total_row = row_data['费用明细'] == '合计'
                
                for col in range(1, len(export_df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col)
                    col_name = export_df.columns[col - 1]
                    
                    if col_name == '费用明细':
                        cell.font = Font(name="仿宋", bold=is_salary_row or is_finance_row or is_total_row)
                        cell.alignment = left_alignment
                    else:
                        cell.font = data_font
                        cell.alignment = right_alignment
                        
                        # 合计列高亮
                        if col_name == '合计':
                            if is_total_row:
                                cell.fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
                                cell.font = Font(name="仿宋", bold=True)
                            else:
                                cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                                cell.font = Font(name="仿宋", bold=True)
                        
                        # 薪酬费用行和合计行背景色
                        if is_salary_row or is_finance_row:
                            if col_name != '合计':
                                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                        elif is_total_row:
                            if col_name != '合计':
                                cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
        
        output.seek(0)
        filename = f'期间费用_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"导出期间费用失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@cost_forecast_bp.route('/screen-cost-allocation/export', methods=['GET'])
def export_screen_cost_allocation():
    """导出公共费用分摊明细数据到Excel"""
    try:
        app_data = get_session_data_manager()
        prediction_period = int(request.args.get('prediction_period', 1))
        
        result = calculate_screen_cost_allocation(app_data, prediction_period)
        
        if not result.get('success', False):
            return jsonify({
                'success': False,
                'error': result.get('error', '计算失败')
            }), 500
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # 定义样式
            header_font = Font(bold=True, color="FFFFFF", name="仿宋")
            data_font = Font(name="仿宋")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # Sheet 1: 直接人工
            direct_labor_data = []
            direct_labor = result.get('direct_labor', {})
            
            # 屏工资明细
            for item in direct_labor.get('screen_wage_details', []):
                code = item.get('物料代码') or item.get('拆解产物编码') or item.get('深加工产物编码') or ''
                name = item.get('物料名称') or item.get('拆解产物名称') or item.get('深加工产物名称') or ''
                wage = item.get('工资', 0.0)
                direct_labor_data.append({
                    '类型': '屏工资（计件工资）',
                    '物料代码': code,
                    '物料名称': name,
                    '金额（元）': wage
                })
            
            # 屏分摊明细
            for item in direct_labor.get('screen_allocation_details', []):
                item_data = item.get('item', {}) if isinstance(item.get('item'), dict) else {}
                code = item_data.get('物料代码') or item_data.get('拆解产物编码') or item_data.get('深加工产物编码') or ''
                name = item_data.get('物料名称') or item_data.get('拆解产物名称') or item_data.get('深加工产物名称') or ''
                fixed_cost = item.get('fixed_cost', 0.0)
                direct_labor_data.append({
                    '类型': '屏分摊（固定工资分摊）',
                    '物料代码': code,
                    '物料名称': name,
                    '金额（元）': fixed_cost
                })
            
            # 添加汇总行
            direct_labor_data.append({
                '类型': '屏工资小计',
                '物料代码': '',
                '物料名称': '',
                '金额（元）': direct_labor.get('screen_wage', 0.0)
            })
            direct_labor_data.append({
                '类型': '屏分摊小计',
                '物料代码': '',
                '物料名称': '',
                '金额（元）': direct_labor.get('screen_allocation', 0.0)
            })
            
            # 确保至少有一个sheet，即使没有数据
            if not direct_labor_data:
                direct_labor_data = [{
                    '类型': '暂无数据',
                    '物料代码': '',
                    '物料名称': '',
                    '金额（元）': 0.0
                }]
            
            direct_labor_df = pd.DataFrame(direct_labor_data)
            direct_labor_df.to_excel(writer, sheet_name='直接人工', index=False)
            worksheet = writer.sheets['直接人工']
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            for col in range(1, len(direct_labor_df.columns) + 1):
                worksheet.column_dimensions[get_column_letter(col)].width = 20
            
            # Sheet 2: 制造费用 - 与拆解量相关的费用
            manufacturing_cost = result.get('manufacturing_cost', {})
            disassembly_data = manufacturing_cost.get('disassembly_related_details', [])
            if not disassembly_data:
                disassembly_data = [{
                    '费用类型': '暂无数据',
                    '费用种类': '',
                    '费用名称': '',
                    '数量': 0.0,
                    '单价': 0.0,
                    '成本': 0.0
                }]
            
            disassembly_df = pd.DataFrame(disassembly_data)
            disassembly_df.to_excel(writer, sheet_name='与拆解量相关', index=False)
            worksheet = writer.sheets['与拆解量相关']
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            for col in range(1, len(disassembly_df.columns) + 1):
                worksheet.column_dimensions[get_column_letter(col)].width = 20
            
            # Sheet 3: 制造费用 - 预计月均费用
            monthly_data = manufacturing_cost.get('monthly_average_details', [])
            if not monthly_data:
                monthly_data = [{
                    '费用类型': '暂无数据',
                    '费用种类': '',
                    '费用名称': '',
                    '月均费用': 0.0,
                    '期数': 0,
                    '成本': 0.0
                }]
            
            monthly_df = pd.DataFrame(monthly_data)
            monthly_df.to_excel(writer, sheet_name='预计月均费用', index=False)
            worksheet = writer.sheets['预计月均费用']
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            for col in range(1, len(monthly_df.columns) + 1):
                worksheet.column_dimensions[get_column_letter(col)].width = 20
            
            # Sheet 4: 制造费用 - 环保费
            environmental_data = manufacturing_cost.get('environmental_fee_details', [])
            if not environmental_data:
                environmental_data = [{
                    '拆解产物编码': '暂无数据',
                    '拆解产物名称': '',
                    '类别': '',
                    '重量': 0.0,
                    '单价': 0.0,
                    '费用': 0.0
                }]
            
            environmental_df = pd.DataFrame(environmental_data)
            environmental_df.to_excel(writer, sheet_name='环保费', index=False)
            worksheet = writer.sheets['环保费']
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            for col in range(1, len(environmental_df.columns) + 1):
                worksheet.column_dimensions[get_column_letter(col)].width = 20
            
            # Sheet 5: 被减扣数据统计
            deducted_data = result.get('deducted_data', {})
            deducted_list = []
            
            # 电视明细
            for item in deducted_data.get('tv_details', []):
                deducted_list.append({
                    '类别': '电视',
                    '拆解产物编码': item.get('拆解产物编码', ''),
                    '拆解产物名称': item.get('拆解产物名称', ''),
                    '重量（KG）': item.get('重量(KG)', 0.0)
                })
            
            # 电脑明细
            for item in deducted_data.get('computer_details', []):
                deducted_list.append({
                    '类别': '电脑',
                    '拆解产物编码': item.get('拆解产物编码', ''),
                    '拆解产物名称': item.get('拆解产物名称', ''),
                    '重量（KG）': item.get('重量(KG)', 0.0)
                })
            
            # 添加汇总行
            deducted_list.append({
                '类别': '电视小计',
                '拆解产物编码': '',
                '拆解产物名称': '',
                '重量（KG）': deducted_data.get('tv_weight', 0.0)
            })
            deducted_list.append({
                '类别': '电脑小计',
                '拆解产物编码': '',
                '拆解产物名称': '',
                '重量（KG）': deducted_data.get('computer_weight', 0.0)
            })
            deducted_list.append({
                '类别': '总重量',
                '拆解产物编码': '',
                '拆解产物名称': '',
                '重量（KG）': deducted_data.get('total_weight', 0.0)
            })
            
            if not deducted_list:
                deducted_list = [{
                    '类别': '暂无数据',
                    '拆解产物编码': '',
                    '拆解产物名称': '',
                    '重量（KG）': 0.0
                }]
            
            deducted_df = pd.DataFrame(deducted_list)
            deducted_df.to_excel(writer, sheet_name='被减扣数据统计', index=False)
            worksheet = writer.sheets['被减扣数据统计']
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            for col in range(1, len(deducted_df.columns) + 1):
                worksheet.column_dimensions[get_column_letter(col)].width = 20
            
            # Sheet 6: 制造费用间接人工分摊
            indirect_labor_allocation = result.get('indirect_labor_allocation', {})
            indirect_labor_data = []
            
            # 添加明细数据
            for detail in indirect_labor_allocation.get('details', []):
                indirect_labor_data.append({
                    '岗位': detail.get('岗位', ''),
                    '拆解产物编码': detail.get('拆解产物编码', ''),
                    '拆解产物名称': detail.get('拆解产物名称', ''),
                    '分类': detail.get('分类', ''),
                    '物料产值（元）': detail.get('物料产值（元）', 0.0),
                    '分摊比例': detail.get('分摊比例', 0.0),
                    '分摊成本（元）': detail.get('分摊成本（元）', 0.0)
                })
            
            # 添加分类汇总行
            category_totals = indirect_labor_allocation.get('category_totals', {})
            for category, total in category_totals.items():
                if total > 0:
                    indirect_labor_data.append({
                        '岗位': f'{category}小计',
                        '拆解产物编码': '',
                        '拆解产物名称': '',
                        '分类': category,
                        '物料产值（元）': 0.0,
                        '分摊比例': 0.0,
                        '分摊成本（元）': total
                    })
            
            # 添加总合计行
            total_cost = indirect_labor_allocation.get('total_cost', 0.0)
            if total_cost > 0 or len(indirect_labor_data) > 0:
                indirect_labor_data.append({
                    '岗位': '总合计',
                    '拆解产物编码': '',
                    '拆解产物名称': '',
                    '分类': '',
                    '物料产值（元）': 0.0,
                    '分摊比例': 0.0,
                    '分摊成本（元）': total_cost
                })
            
            if not indirect_labor_data:
                indirect_labor_data = [{
                    '岗位': '暂无数据',
                    '拆解产物编码': '',
                    '拆解产物名称': '',
                    '分类': '',
                    '物料产值（元）': 0.0,
                    '分摊比例': 0.0,
                    '分摊成本（元）': 0.0
                }]
            
            indirect_labor_df = pd.DataFrame(indirect_labor_data)
            indirect_labor_df.to_excel(writer, sheet_name='制造费用间接人工分摊', index=False)
            worksheet = writer.sheets['制造费用间接人工分摊']
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            for col in range(1, len(indirect_labor_df.columns) + 1):
                worksheet.column_dimensions[get_column_letter(col)].width = 20
            
            # Sheet 7: 制造费用公共成本分摊
            public_cost_allocation = result.get('public_cost_allocation', {})
            public_cost_data = []
            
            # 添加明细数据
            for detail in public_cost_allocation.get('details', []):
                public_cost_data.append({
                    '费用类型': detail.get('费用类型', ''),
                    '费用种类': detail.get('费用种类', ''),
                    '费用名称': detail.get('费用名称', ''),
                    '拆解产物编码': detail.get('拆解产物编码', ''),
                    '拆解产物名称': detail.get('拆解产物名称', ''),
                    '分类': detail.get('分类', ''),
                    '物料产值（元）': detail.get('物料产值（元）', 0.0),
                    '分摊比例': detail.get('分摊比例', 0.0),
                    '分摊成本（元）': detail.get('分摊成本（元）', 0.0)
                })
            
            # 添加分类汇总行
            category_totals = public_cost_allocation.get('category_totals', {})
            for category, total in category_totals.items():
                if total > 0:
                    public_cost_data.append({
                        '费用类型': f'{category}小计',
                        '费用种类': '',
                        '费用名称': '',
                        '拆解产物编码': '',
                        '拆解产物名称': '',
                        '分类': category,
                        '物料产值（元）': 0.0,
                        '分摊比例': 0.0,
                        '分摊成本（元）': total
                    })
            
            # 添加总合计行
            total_cost = public_cost_allocation.get('total_cost', 0.0)
            if total_cost > 0 or len(public_cost_data) > 0:
                public_cost_data.append({
                    '费用类型': '总合计',
                    '费用种类': '',
                    '费用名称': '',
                    '拆解产物编码': '',
                    '拆解产物名称': '',
                    '分类': '',
                    '物料产值（元）': 0.0,
                    '分摊比例': 0.0,
                    '分摊成本（元）': total_cost
                })
            
            if not public_cost_data:
                public_cost_data = [{
                    '费用类型': '暂无数据',
                    '费用种类': '',
                    '费用名称': '',
                    '拆解产物编码': '',
                    '拆解产物名称': '',
                    '分类': '',
                    '物料产值（元）': 0.0,
                    '分摊比例': 0.0,
                    '分摊成本（元）': 0.0
                }]
            
            public_cost_df = pd.DataFrame(public_cost_data)
            public_cost_df.to_excel(writer, sheet_name='制造费用公共成本分摊', index=False)
            worksheet = writer.sheets['制造费用公共成本分摊']
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            for col in range(1, len(public_cost_df.columns) + 1):
                worksheet.column_dimensions[get_column_letter(col)].width = 20
            
            # Sheet 8: 分摊计算结果
            allocation = result.get('allocation', {})
            allocation_data = []
            
            # 添加费用明细（按照当前展示的费用统计）
            cost_details = allocation.get('cost_details', {})
            screen_allocation = allocation.get('screen_allocation', {})
            
            # 制造费用间接人工分摊
            allocation_data.append({
                '项目': '制造费用间接人工分摊',
                '明细': '',
                '金额（元）': cost_details.get('制造费用间接人工分摊', 0.0)
            })
            
            # 制造费用公共成本分摊
            allocation_data.append({
                '项目': '制造费用公共成本分摊',
                '明细': '',
                '金额（元）': cost_details.get('制造费用公共成本分摊', 0.0)
            })
            
            # 屏费用分摊结果
            allocation_data.append({
                '项目': '屏费用分摊结果',
                '明细': '',
                '金额（元）': screen_allocation.get('total_cost', 0.0)
            })
            
            # 屏费用分摊结果明细（直接人工 + 制造费用（屏））
            direct_labor_details = cost_details.get('直接人工', {})
            allocation_data.append({
                '项目': '  └─ 直接人工',
                '明细': '',
                '金额（元）': direct_labor_details.get('小计', 0.0)
            })
            allocation_data.append({
                '项目': '    └─ 屏工资',
                '明细': '',
                '金额（元）': direct_labor_details.get('屏工资', 0.0)
            })
            allocation_data.append({
                '项目': '    └─ 屏分摊',
                '明细': '',
                '金额（元）': direct_labor_details.get('屏分摊', 0.0)
            })
            
            manufacturing_details = cost_details.get('制造费用（屏）', {})
            allocation_data.append({
                '项目': '  └─ 制造费用（屏）',
                '明细': '',
                '金额（元）': manufacturing_details.get('小计', 0.0)
            })
            allocation_data.append({
                '项目': '    └─ 与拆解量相关的费用',
                '明细': '',
                '金额（元）': manufacturing_details.get('与拆解量相关的费用', 0.0)
            })
            allocation_data.append({
                '项目': '    └─ 预计月均费用',
                '明细': '',
                '金额（元）': manufacturing_details.get('预计月均费用', 0.0)
            })
            allocation_data.append({
                '项目': '    └─ 环保费',
                '明细': '',
                '金额（元）': manufacturing_details.get('环保费', 0.0)
            })
            
            # 屏费用分摊结果按重量比例分摊
            allocation_data.append({
                '项目': '  └─ 电视（屏）',
                '明细': '',
                '金额（元）': screen_allocation.get('tv_allocation', 0.0)
            })
            allocation_data.append({
                '项目': '  └─ 电脑（屏）',
                '明细': '',
                '金额（元）': screen_allocation.get('computer_allocation', 0.0)
            })
            
            # 总费用
            allocation_data.append({
                '项目': '总费用',
                '明细': '',
                '金额（元）': allocation.get('total_cost', 0.0)
            })
            
            # 空行
            allocation_data.append({
                '项目': '',
                '明细': '',
                '金额（元）': 0.0
            })
            
            # 按类别分摊
            category_allocation = allocation.get('category_allocation', {})
            allocation_data.append({
                '项目': '按类别分摊',
                '明细': '',
                '金额（元）': 0.0
            })
            for category in ['冰箱', '空调', '电脑', '电视', '洗衣机']:
                allocation_data.append({
                    '项目': f'  └─ {category}',
                    '明细': '',
                    '金额（元）': category_allocation.get(category, 0.0)
                })
            
            
            if not allocation_data:
                allocation_data = [{
                    '项目': '暂无数据',
                    '明细': '',
                    '金额（元）': 0.0
                }]
            
            allocation_df = pd.DataFrame(allocation_data)
            allocation_df.to_excel(writer, sheet_name='分摊计算结果', index=False)
            worksheet = writer.sheets['分摊计算结果']
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            for col in range(1, len(allocation_df.columns) + 1):
                worksheet.column_dimensions[get_column_letter(col)].width = 25
        
        output.seek(0)
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"公共费用分摊明细_{timestamp}.xlsx"
        
        # 使用send_file自动处理中文文件名编码
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"导出公共费用分摊明细失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


def calculate_production_cost_allocation(app_data, prediction_period=1):
    """计算生产成本分摊数据（供路由与服务层复用）"""
    try:
        categories = list(PRODUCTION_COST_CATEGORIES)

        # 初始化结果数据
        result_data = []
        category_totals = {
            'direct_material': {cat: 0.0 for cat in categories},
            'direct_labor': {cat: 0.0 for cat in categories},
            'manufacturing_cost': {cat: 0.0 for cat in categories},
            'product_value': {cat: 0.0 for cat in categories},
            'subsidy_income': {cat: 0.0 for cat in categories}
        }
        
        # ========== 1. 收集直接材料数据 ==========
        try:
            manual_data = app_data.get_data('extracted_data_manual')
            if manual_data is not None and not manual_data.empty:
                # 计算拆解物原料成本
                cost_data = calculate_material_cost(manual_data)
                
                # 只处理旧机类别
                if '类别' in cost_data.columns:
                    old_machine_data = cost_data[cost_data['类别'] == '旧机'].copy()
                else:
                    old_machine_data = cost_data.copy()
                
                # 确保有物料描述和拆解物原料成本列
                if '物料描述' in old_machine_data.columns and '拆解物原料成本' in old_machine_data.columns:
                    old_machine_data['拆解物原料成本'] = pd.to_numeric(
                        old_machine_data['拆解物原料成本'], errors='coerce'
                    ).fillna(0)
                    
                    # 按物料描述分类汇总
                    for category in categories:
                        mask = old_machine_data['物料描述'].astype(str).str.contains(category, case=False, na=False)
                        category_totals['direct_material'][category] = float(
                            old_machine_data.loc[mask, '拆解物原料成本'].sum()
                        )
        except Exception as e:
            print(f"收集直接材料数据失败: {str(e)}")
        
        # ========== 2. 收集直接人工数据（使用四机一脑分类统计数据） ==========
        try:
            direct_labor_result = calculate_direct_labor_cost(app_data, prediction_period)
            # 直接使用四机一脑分类统计数据（product_category_stats）
            product_category_stats = direct_labor_result.get('product_category_stats', {})
            for cat in categories:
                stats = product_category_stats.get(cat, {'wage': 0.0, 'fixed_cost': 0.0})
                category_totals['direct_labor'][cat] = stats.get('wage', 0.0) + stats.get('fixed_cost', 0.0)
        except Exception as e:
            print(f"收集直接人工数据失败: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # ========== 3. 收集制造费用数据 ==========
        try:
            mfg_by_category = collect_production_manufacturing_cost_by_category(
                app_data, prediction_period
            )
            for cat in categories:
                category_totals['manufacturing_cost'][cat] = mfg_by_category.get(cat, 0.0)
        except Exception as e:
            print(f"收集制造费用数据失败: {str(e)}")
        
        # ========== 4. 收集拆解产物价值数据 ==========
        try:
            disassembly_data = app_data.get_data('disassembly_data')
            if disassembly_data is not None and not disassembly_data.empty:
                if '类别' in disassembly_data.columns:
                    product_data = disassembly_data[disassembly_data['类别'] == '拆解产物'].copy()
                    if not product_data.empty:
                        from data.base_data.price_data import load_price_data
                        price_df = load_price_data()
                        if price_df is not None and not price_df.empty:
                            price_mapping = {}
                            for _, row in price_df.iterrows():
                                code = str(row['拆解产物编码']).strip()
                                price_no_tax = row.get('销售单价-不含税(元/KG)', 0)
                                if pd.notna(price_no_tax):
                                    price_mapping[code] = float(price_no_tax)
                            
                            # 分类关键词映射规则
                            # 电视：电视、彩电、CRT其它机壳破碎塑料、线路板边框破碎塑料、等离子、废旧玻璃电子枪、废旧金属荫罩压块铁、黑白
                            # 电脑：电脑、显示器、笔记本、主机、废旧金属黑色金属-铁及其合金-电子枪
                            # 冰箱：冰箱、冰柜
                            # 空调：空调
                            # 洗衣机：洗衣机、双缸
                            category_keyword_mapping = {
                                '电视': ['电视', '彩电', 'CRT其它机壳破碎塑料', '线路板边框破碎塑料', '等离子', '废旧玻璃电子枪', '废旧金属荫罩压块铁', '黑白'],
                                '电脑': ['电脑', '显示器', '笔记本', '主机', '废旧金属黑色金属-铁及其合金-电子枪'],
                                '冰箱': ['冰箱', '冰柜'],
                                '空调': ['空调'],
                                '洗衣机': ['洗衣机', '双缸']
                            }
                            
                            for idx, row in product_data.iterrows():
                                material_name = str(row.get('原物料名称', '')).strip()
                                category = None
                                for cat, keywords in category_keyword_mapping.items():
                                    for keyword in keywords:
                                        if keyword in material_name:
                                            category = cat
                                            break
                                    if category:
                                        break
                                
                                # 如果没有匹配到分类，跳过该记录
                                if not category or category not in categories:
                                    continue
                                
                                # 计算物料价值
                                product_code = str(row.get('拆解产物编码', '')).strip()
                                calculated_weight = row.get('计算结果(KG)', 0)
                                try:
                                    calculated_weight = float(calculated_weight) if pd.notna(calculated_weight) else 0
                                except (ValueError, TypeError):
                                    calculated_weight = 0
                                
                                price_no_tax = price_mapping.get(product_code, 0)
                                if price_no_tax < 0:
                                    material_value = 0
                                else:
                                    material_value = calculated_weight * price_no_tax
                                
                                category_totals['product_value'][category] += material_value
        except Exception as e:
            print(f"收集拆解产物价值数据失败: {str(e)}")
        
        # ========== 5. 收集基金补贴收入数据 ==========
        try:
            subsidy_income_data = app_data.get_data('subsidy_income_data')
            if subsidy_income_data is not None and not subsidy_income_data.empty:
                if '补贴大类' in subsidy_income_data.columns and '基金补贴收入(元)' in subsidy_income_data.columns:
                    subsidy_income_data['基金补贴收入(元)'] = pd.to_numeric(
                        subsidy_income_data['基金补贴收入(元)'], errors='coerce'
                    ).fillna(0)
                    
                    # 将补贴大类映射到产品类型（五大类：电视机、冰箱、洗衣机、空调、电脑）
                    def map_subsidy_category_to_product_type(subsidy_category):
                        """将补贴大类映射到产品类型"""
                        subsidy_category = str(subsidy_category).strip()
                        # 空调的细分补贴大类
                        if subsidy_category in ['整机', '内机', '外机']:
                            return '空调'
                        # 电脑的细分补贴大类
                        elif subsidy_category in ['笔记本', '显示器', '主机']:
                            return '电脑'
                        # 其他直接返回（电视机、冰箱、洗衣机）
                        elif subsidy_category in ['电视机', '冰箱', '洗衣机']:
                            return subsidy_category
                        else:
                            # 默认返回原值（兼容旧数据）
                            return subsidy_category
                    
                    # 添加产品类型列
                    subsidy_income_data['产品类型'] = subsidy_income_data['补贴大类'].apply(map_subsidy_category_to_product_type)
                    
                    # 按产品类型分组汇总
                    product_type_groups = subsidy_income_data.groupby('产品类型')['基金补贴收入(元)'].sum()
                    
                    # 映射产品类型到四机一脑分类
                    for product_type, total_subsidy in product_type_groups.items():
                        product_type = str(product_type).strip()
                        # 映射：电视机 -> 电视
                        if product_type == '电视机':
                            product_type = '电视'
                        
                        if product_type in categories:
                            category_totals['subsidy_income'][product_type] += float(total_subsidy)
        except Exception as e:
            print(f"收集基金补贴收入数据失败: {str(e)}")
        
        # ========== 6. 计算各项指标 ==========
        total_row = {
            'direct_material': 0.0,
            'direct_labor': 0.0,
            'manufacturing_cost': 0.0,
            'product_value': 0.0,
            'subsidy_income': 0.0,
            'production_cost_subtotal': 0.0,
            'subsidy_allocation_cost': 0.0,
            'product_allocation_cost': 0.0
        }
        
        for category in categories:
            direct_material = category_totals['direct_material'][category]
            direct_labor = category_totals['direct_labor'][category]
            manufacturing_cost = category_totals['manufacturing_cost'][category]
            product_value = category_totals['product_value'][category]
            subsidy_income = category_totals['subsidy_income'][category]
            
            # 计算生产成本小计
            production_cost_subtotal = direct_material + direct_labor + manufacturing_cost
            
            # 计算分摊比例
            total_revenue = subsidy_income + product_value
            if total_revenue > 0:
                subsidy_ratio = subsidy_income / total_revenue
                product_ratio = product_value / total_revenue
            else:
                subsidy_ratio = 0.0
                product_ratio = 0.0
            
            # 计算分摊成本
            subsidy_allocation_cost = production_cost_subtotal * subsidy_ratio
            product_allocation_cost = production_cost_subtotal * product_ratio
            
            # 构建行数据
            row_data = {
                '产线': category,
                '基金补贴收入': round(subsidy_income, 2),
                '拆解产物价值': round(product_value, 2),
                '基金补贴成本分摊比例': round(subsidy_ratio, 4),
                '拆解产物成本分摊比例': round(product_ratio, 4),
                '直接材料': round(direct_material, 2),
                '直接人工': round(direct_labor, 2),
                '制造费用': round(manufacturing_cost, 2),
                '生产成本小计': round(production_cost_subtotal, 2),
                '基金补贴收入分摊成本': round(subsidy_allocation_cost, 2),
                '拆解产物分摊成本': round(product_allocation_cost, 2)
            }
            result_data.append(row_data)
            
            # 累计合计
            total_row['direct_material'] += direct_material
            total_row['direct_labor'] += direct_labor
            total_row['manufacturing_cost'] += manufacturing_cost
            total_row['product_value'] += product_value
            total_row['subsidy_income'] += subsidy_income
            total_row['production_cost_subtotal'] += production_cost_subtotal
            total_row['subsidy_allocation_cost'] += subsidy_allocation_cost
            total_row['product_allocation_cost'] += product_allocation_cost
        
        # 计算合计行的分摊比例
        total_revenue = total_row['subsidy_income'] + total_row['product_value']
        if total_revenue > 0:
            total_subsidy_ratio = total_row['subsidy_income'] / total_revenue
            total_product_ratio = total_row['product_value'] / total_revenue
        else:
            total_subsidy_ratio = 0.0
            total_product_ratio = 0.0
        
        # 添加合计行
        total_row_data = {
            '产线': '合计',
            '基金补贴收入': round(total_row['subsidy_income'], 2),
            '拆解产物价值': round(total_row['product_value'], 2),
            '基金补贴成本分摊比例': round(total_subsidy_ratio, 4),
            '拆解产物成本分摊比例': round(total_product_ratio, 4),
            '直接材料': round(total_row['direct_material'], 2),
            '直接人工': round(total_row['direct_labor'], 2),
            '制造费用': round(total_row['manufacturing_cost'], 2),
            '生产成本小计': round(total_row['production_cost_subtotal'], 2),
            '基金补贴收入分摊成本': round(total_row['subsidy_allocation_cost'], 2),
            '拆解产物分摊成本': round(total_row['product_allocation_cost'], 2)
        }
        result_data.append(total_row_data)
        return result_data

    except Exception as e:
        print(f"计算生产成本分摊数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return []


@cost_forecast_bp.route('/production-cost-allocation', methods=['GET'])
def get_production_cost_allocation():
    """获取生产成本分摊计算数据"""
    try:
        app_data = get_session_data_manager()
        prediction_period = int(request.args.get('prediction_period', 1))
        force_refresh = request.args.get('force_refresh', 'false').lower() == 'true'

        data_cleared = app_data.get_data('__data_cleared__')
        if data_cleared:
            return jsonify({
                'success': True,
                'data': []
            })

        cache_key = f'production_cost_allocation_result_v2_{prediction_period}'
        if not force_refresh:
            cached_result = app_data.get_data(cache_key)
            if cached_result is not None and len(cached_result) > 0:
                return jsonify({
                    'success': True,
                    'data': cached_result,
                    'from_cache': True
                })

        result_data = calculate_production_cost_allocation(app_data, prediction_period)
        app_data.set_data(cache_key, result_data)

        return jsonify({
            'success': True,
            'data': result_data,
            'from_cache': False
        })

    except Exception as e:
        print(f"获取生产成本分摊数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@cost_forecast_bp.route('/production-cost-allocation/export', methods=['GET'])
def export_production_cost_allocation():
    """导出生产成本分摊数据到Excel"""
    try:
        app_data = get_session_data_manager()
        prediction_period = int(request.args.get('prediction_period', 1))
        
        categories = list(PRODUCTION_COST_CATEGORIES)
        
        # 初始化结果数据
        result_data = []
        category_totals = {
            'direct_material': {cat: 0.0 for cat in categories},
            'direct_labor': {cat: 0.0 for cat in categories},
            'manufacturing_cost': {cat: 0.0 for cat in categories},
            'product_value': {cat: 0.0 for cat in categories},
            'subsidy_income': {cat: 0.0 for cat in categories}
        }
        
        # ========== 数据收集（与get_production_cost_allocation相同的逻辑）==========
        # 1. 收集直接材料数据
        try:
            manual_data = app_data.get_data('extracted_data_manual')
            if manual_data is not None and not manual_data.empty:
                cost_data = calculate_material_cost(manual_data)
                if '类别' in cost_data.columns:
                    old_machine_data = cost_data[cost_data['类别'] == '旧机'].copy()
                else:
                    old_machine_data = cost_data.copy()
                
                if '物料描述' in old_machine_data.columns and '拆解物原料成本' in old_machine_data.columns:
                    old_machine_data['拆解物原料成本'] = pd.to_numeric(
                        old_machine_data['拆解物原料成本'], errors='coerce'
                    ).fillna(0)
                    
                    for category in categories:
                        mask = old_machine_data['物料描述'].astype(str).str.contains(category, case=False, na=False)
                        category_totals['direct_material'][category] = float(
                            old_machine_data.loc[mask, '拆解物原料成本'].sum()
                        )
        except Exception as e:
            print(f"收集直接材料数据失败: {str(e)}")
        
        # 2. 收集直接人工数据（使用四机一脑分类统计数据）
        try:
            direct_labor_result = calculate_direct_labor_cost(app_data, prediction_period)
            # 直接使用四机一脑分类统计数据（product_category_stats）
            product_category_stats = direct_labor_result.get('product_category_stats', {})
            for cat in categories:
                stats = product_category_stats.get(cat, {'wage': 0.0, 'fixed_cost': 0.0})
                category_totals['direct_labor'][cat] = stats.get('wage', 0.0) + stats.get('fixed_cost', 0.0)
        except Exception as e:
            print(f"收集直接人工数据失败: {str(e)}")
        
        # 3. 收集制造费用数据
        try:
            mfg_by_category = collect_production_manufacturing_cost_by_category(
                app_data, prediction_period
            )
            for cat in categories:
                category_totals['manufacturing_cost'][cat] = mfg_by_category.get(cat, 0.0)
        except Exception as e:
            print(f"收集制造费用数据失败: {str(e)}")
        
        # 4. 收集拆解产物价值数据
        try:
            disassembly_data = app_data.get_data('disassembly_data')
            if disassembly_data is not None and not disassembly_data.empty:
                if '类别' in disassembly_data.columns:
                    product_data = disassembly_data[disassembly_data['类别'] == '拆解产物'].copy()
                    if not product_data.empty:
                        from data.base_data.price_data import load_price_data
                        price_df = load_price_data()
                        if price_df is not None and not price_df.empty:
                            price_mapping = {}
                            for _, row in price_df.iterrows():
                                code = str(row['拆解产物编码']).strip()
                                price_no_tax = row.get('销售单价-不含税(元/KG)', 0)
                                if pd.notna(price_no_tax):
                                    price_mapping[code] = float(price_no_tax)
                            
                            # 分类关键词映射规则
                            # 电视：电视、彩电、CRT其它机壳破碎塑料、线路板边框破碎塑料、等离子、废旧玻璃电子枪、废旧金属荫罩压块铁、黑白
                            # 电脑：电脑、显示器、笔记本、主机、废旧金属黑色金属-铁及其合金-电子枪
                            # 冰箱：冰箱、冰柜
                            # 空调：空调
                            # 洗衣机：洗衣机、双缸
                            category_keyword_mapping = {
                                '电视': ['电视', '彩电', 'CRT其它机壳破碎塑料', '线路板边框破碎塑料', '等离子', '废旧玻璃电子枪', '废旧金属荫罩压块铁', '黑白'],
                                '电脑': ['电脑', '显示器', '笔记本', '主机', '废旧金属黑色金属-铁及其合金-电子枪'],
                                '冰箱': ['冰箱', '冰柜'],
                                '空调': ['空调'],
                                '洗衣机': ['洗衣机', '双缸']
                            }
                            
                            for idx, row in product_data.iterrows():
                                material_name = str(row.get('原物料名称', '')).strip()
                                category = None
                                for cat, keywords in category_keyword_mapping.items():
                                    for keyword in keywords:
                                        if keyword in material_name:
                                            category = cat
                                            break
                                    if category:
                                        break
                                
                                # 如果没有匹配到分类，跳过该记录
                                if not category or category not in categories:
                                    continue
                                
                                # 计算物料价值
                                product_code = str(row.get('拆解产物编码', '')).strip()
                                calculated_weight = row.get('计算结果(KG)', 0)
                                try:
                                    calculated_weight = float(calculated_weight) if pd.notna(calculated_weight) else 0
                                except (ValueError, TypeError):
                                    calculated_weight = 0
                                
                                price_no_tax = price_mapping.get(product_code, 0)
                                if price_no_tax < 0:
                                    material_value = 0
                                else:
                                    material_value = calculated_weight * price_no_tax
                                
                                category_totals['product_value'][category] += material_value
        except Exception as e:
            print(f"收集拆解产物价值数据失败: {str(e)}")
        
        # 5. 收集基金补贴收入数据
        try:
            subsidy_income_data = app_data.get_data('subsidy_income_data')
            if subsidy_income_data is not None and not subsidy_income_data.empty:
                if '补贴大类' in subsidy_income_data.columns and '基金补贴收入(元)' in subsidy_income_data.columns:
                    subsidy_income_data['基金补贴收入(元)'] = pd.to_numeric(
                        subsidy_income_data['基金补贴收入(元)'], errors='coerce'
                    ).fillna(0)
                    
                    # 将补贴大类映射到产品类型（五大类：电视机、冰箱、洗衣机、空调、电脑）
                    def map_subsidy_category_to_product_type(subsidy_category):
                        """将补贴大类映射到产品类型"""
                        subsidy_category = str(subsidy_category).strip()
                        # 空调的细分补贴大类
                        if subsidy_category in ['整机', '内机', '外机']:
                            return '空调'
                        # 电脑的细分补贴大类
                        elif subsidy_category in ['笔记本', '显示器', '主机']:
                            return '电脑'
                        # 其他直接返回（电视机、冰箱、洗衣机）
                        elif subsidy_category in ['电视机', '冰箱', '洗衣机']:
                            return subsidy_category
                        else:
                            # 默认返回原值（兼容旧数据）
                            return subsidy_category
                    
                    # 添加产品类型列
                    subsidy_income_data['产品类型'] = subsidy_income_data['补贴大类'].apply(map_subsidy_category_to_product_type)
                    
                    # 按产品类型分组汇总
                    product_type_groups = subsidy_income_data.groupby('产品类型')['基金补贴收入(元)'].sum()
                    
                    # 映射产品类型到四机一脑分类
                    for product_type, total_subsidy in product_type_groups.items():
                        product_type = str(product_type).strip()
                        # 映射：电视机 -> 电视
                        if product_type == '电视机':
                            product_type = '电视'
                        
                        if product_type in categories:
                            category_totals['subsidy_income'][product_type] += float(total_subsidy)
        except Exception as e:
            print(f"收集基金补贴收入数据失败: {str(e)}")
        
        # ========== 计算各项指标 ==========
        total_row = {
            'direct_material': 0.0,
            'direct_labor': 0.0,
            'manufacturing_cost': 0.0,
            'product_value': 0.0,
            'subsidy_income': 0.0,
            'production_cost_subtotal': 0.0,
            'subsidy_allocation_cost': 0.0,
            'product_allocation_cost': 0.0
        }
        
        for category in categories:
            direct_material = category_totals['direct_material'][category]
            direct_labor = category_totals['direct_labor'][category]
            manufacturing_cost = category_totals['manufacturing_cost'][category]
            product_value = category_totals['product_value'][category]
            subsidy_income = category_totals['subsidy_income'][category]
            
            production_cost_subtotal = direct_material + direct_labor + manufacturing_cost
            
            total_revenue = subsidy_income + product_value
            if total_revenue > 0:
                subsidy_ratio = subsidy_income / total_revenue
                product_ratio = product_value / total_revenue
            else:
                subsidy_ratio = 0.0
                product_ratio = 0.0
            
            subsidy_allocation_cost = production_cost_subtotal * subsidy_ratio
            product_allocation_cost = production_cost_subtotal * product_ratio
            
            row_data = {
                '产线': category,
                '基金补贴收入': round(subsidy_income, 2),
                '拆解产物价值': round(product_value, 2),
                '基金补贴成本分摊比例': round(subsidy_ratio, 4),
                '拆解产物成本分摊比例': round(product_ratio, 4),
                '直接材料': round(direct_material, 2),
                '直接人工': round(direct_labor, 2),
                '制造费用': round(manufacturing_cost, 2),
                '生产成本小计': round(production_cost_subtotal, 2),
                '基金补贴收入分摊成本': round(subsidy_allocation_cost, 2),
                '拆解产物分摊成本': round(product_allocation_cost, 2)
            }
            result_data.append(row_data)
            
            total_row['direct_material'] += direct_material
            total_row['direct_labor'] += direct_labor
            total_row['manufacturing_cost'] += manufacturing_cost
            total_row['product_value'] += product_value
            total_row['subsidy_income'] += subsidy_income
            total_row['production_cost_subtotal'] += production_cost_subtotal
            total_row['subsidy_allocation_cost'] += subsidy_allocation_cost
            total_row['product_allocation_cost'] += product_allocation_cost
        
        total_revenue = total_row['subsidy_income'] + total_row['product_value']
        if total_revenue > 0:
            total_subsidy_ratio = total_row['subsidy_income'] / total_revenue
            total_product_ratio = total_row['product_value'] / total_revenue
        else:
            total_subsidy_ratio = 0.0
            total_product_ratio = 0.0
        
        total_row_data = {
            '产线': '合计',
            '基金补贴收入': round(total_row['subsidy_income'], 2),
            '拆解产物价值': round(total_row['product_value'], 2),
            '基金补贴成本分摊比例': round(total_subsidy_ratio, 4),
            '拆解产物成本分摊比例': round(total_product_ratio, 4),
            '直接材料': round(total_row['direct_material'], 2),
            '直接人工': round(total_row['direct_labor'], 2),
            '制造费用': round(total_row['manufacturing_cost'], 2),
            '生产成本小计': round(total_row['production_cost_subtotal'], 2),
            '基金补贴收入分摊成本': round(total_row['subsidy_allocation_cost'], 2),
            '拆解产物分摊成本': round(total_row['product_allocation_cost'], 2)
        }
        result_data.append(total_row_data)
        
        if not result_data:
            return jsonify({
                'success': False,
                'error': '没有可导出的生产成本分摊数据'
            }), 400
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # 定义样式（与直接人工成本一致）
            header_font = Font(bold=True, color="FFFFFF", name="仿宋", size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            data_font = Font(name="仿宋", size=11)
            
            # 准备导出数据
            export_data = []
            for row in result_data:
                export_data.append({
                    '产线': row.get('产线', ''),
                    '基金补贴收入': row.get('基金补贴收入', 0),
                    '拆解产物价值': row.get('拆解产物价值', 0),
                    '基金补贴成本分摊比例': row.get('基金补贴成本分摊比例', 0),
                    '拆解产物成本分摊比例': row.get('拆解产物成本分摊比例', 0),
                    '直接材料': row.get('直接材料', 0),
                    '直接人工': row.get('直接人工', 0),
                    '制造费用': row.get('制造费用', 0),
                    '生产成本小计': row.get('生产成本小计', 0),
                    '基金补贴收入分摊成本': row.get('基金补贴收入分摊成本', 0),
                    '拆解产物分摊成本': row.get('拆解产物分摊成本', 0)
                })
            
            df = pd.DataFrame(export_data)
            df.to_excel(writer, sheet_name='生产成本分摊', index=False)
            
            # 设置样式
            ws = writer.sheets['生产成本分摊']
            
            # 设置列宽
            column_widths = {
                'A': 12,  # 产线
                'B': 18,  # 基金补贴收入
                'C': 18,  # 拆解产物价值
                'D': 22,  # 基金补贴成本分摊比例
                'E': 22,  # 拆解产物成本分摊比例
                'F': 15,  # 直接材料
                'G': 15,  # 直接人工
                'H': 15,  # 制造费用
                'I': 18,  # 生产成本小计
                'J': 22,  # 基金补贴收入分摊成本
                'K': 22   # 拆解产物分摊成本
            }
            
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # 设置表头样式
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # 设置数据行样式
            for row in range(2, len(df) + 2):
                is_total = ws.cell(row=row, column=1).value == '合计'
                
                for col in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = data_font
                    cell.alignment = center_alignment
                    
                    # 合计行特殊样式
                    if is_total:
                        cell.font = Font(bold=True, name="仿宋", size=11)
                        if col == 1:  # 产线列
                            cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                        elif col in [6, 7, 8, 9, 10, 11]:  # 成本相关列
                            cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                    # 比例列高亮（基金补贴成本分摊比例和拆解产物成本分摊比例）
                    elif col in [4, 5]:  # 比例列
                        cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
        
        output.seek(0)
        filename = f'生产成本分摊_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"导出生产成本分摊失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


def calculate_disassembly_product_cost(app_data, prediction_period=1):
    """
    计算一次拆解产物成本
    
    计算公式：
    一次拆解产物期末单位成本 =（一次拆解产物期初库存金额 + 一次拆解产物本期生产成本） / 
                                （一次拆解产物期初库存数量 + 当期生产一次拆解产物重量）
    
    数据来源：
    1. 一次拆解产物期初库存金额：提取结果，类别为"一次拆解产物"，按物料代码统计"价值"列
    2. 一次拆解产物期初库存数量：提取结果，类别为"一次拆解产物"，按物料代码统计"非限制使用的库存"
    3. 一次拆解产物本期生产成本：每个拆解产物编码对应的物料产值/四机一脑分类的物料产值总数 × 生产成本分摊页（拆解产物分摊成本）
    4. 当期生产一次拆解产物重量：一次拆解产物产值页面，按"拆解产物编码"统计"计算结果(KG)"列
    
    Args:
        app_data: 数据管理器实例
        prediction_period: 预测期数（默认1）
        
    Returns:
        list: 包含每个拆解产物编码的详细计算结果
    """
    try:
        result_data = []
        
        # 分类关键词映射规则（在整个函数中可用）
        # 电视：电视、彩电、CRT其它机壳破碎塑料、线路板边框破碎塑料、等离子、废旧玻璃电子枪、废旧金属荫罩压块铁、黑白
        # 电脑：电脑、显示器、笔记本、主机、废旧金属黑色金属-铁及其合金-电子枪
        # 冰箱：冰箱、冰柜
        # 空调：空调
        # 洗衣机：洗衣机、双缸
        # 🔧 修复：调整顺序，让"显示器"优先匹配为"电脑"类
        # 使用OrderedDict确保匹配顺序：电脑（优先）、电视、冰箱、空调、洗衣机
        from collections import OrderedDict
        category_keyword_mapping = OrderedDict([
            ('电脑', ['显示器', '电脑', '笔记本', '主机', '废旧金属黑色金属-铁及其合金-电子枪']),  # 显示器优先
            ('电视', ['电视', '彩电', 'CRT其它机壳破碎塑料', '线路板边框破碎塑料', '等离子', '废旧玻璃电子枪', '废旧金属荫罩压块铁', '黑白']),
            ('冰箱', ['冰箱', '冰柜']),
            ('空调', ['空调']),
            ('洗衣机', ['洗衣机', '双缸'])
        ])
        
        # ========== 1. 获取一次拆解产物期初库存数据（从提取结果） ==========
        extracted_data = app_data.get_data('extracted_data')
        initial_inventory = {}  # {物料代码: {'金额': 0, '数量': 0}}
        extracted_code_to_name = {}  # {物料代码: 物料描述} - 用于匹配拆解产物名称
        
        if extracted_data is not None and not extracted_data.empty:
            if '类别' in extracted_data.columns:
                # 筛选类别为"一次拆解产物"的数据
                disassembly_product_data = extracted_data[extracted_data['类别'] == '一次拆解产物'].copy()
                
                if not disassembly_product_data.empty:
                    # 确保必要的列存在
                    if '物料代码' in disassembly_product_data.columns:
                        # 按物料代码分组统计
                        for _, row in disassembly_product_data.iterrows():
                            material_code = str(row.get('物料代码', '')).strip()
                            if not material_code:
                                continue
                            
                            if material_code not in initial_inventory:
                                initial_inventory[material_code] = {
                                    '金额': 0.0,
                                    '数量': 0.0
                                }
                            
                            # 统计价值列
                            if '价值' in row:
                                value = pd.to_numeric(row['价值'], errors='coerce')
                                if pd.notna(value):
                                    initial_inventory[material_code]['金额'] += float(value)
                            
                            # 统计非限制使用的库存列
                            if '非限制使用的库存' in row:
                                stock = pd.to_numeric(row['非限制使用的库存'], errors='coerce')
                                if pd.notna(stock):
                                    initial_inventory[material_code]['数量'] += float(stock)
                            
                            # 建立物料代码到物料描述的映射（用于后续匹配名称）
                            if '物料描述' in row:
                                material_desc = str(row.get('物料描述', '')).strip()
                                if material_desc and material_code not in extracted_code_to_name:
                                    extracted_code_to_name[material_code] = material_desc
        
        # ========== 2. 从产品拆解系数数据中建立拆解产物编码到名称的映射 ==========
        # 用于补充缺失的拆解产物名称
        product_code_to_name = {}  # {拆解产物编码: 拆解产物名称}
        try:
            from data.base_data.product_data import PRODUCT_DISASSEMBLY_DATA
            for product_code, product_info in PRODUCT_DISASSEMBLY_DATA.items():
                if '拆解系数_明细' in product_info:
                    for detail in product_info['拆解系数_明细']:
                        detail_code = str(detail.get('一次拆解产物编码', '')).strip()
                        if detail_code.endswith('.0'):
                            detail_code = detail_code[:-2]
                        detail_name = str(detail.get('一次拆解产物名称', '')).strip()
                        if detail_code and detail_name:
                            product_code_to_name[detail_code] = detail_name
        except Exception as e:
            print(f"从产品数据获取拆解产物名称映射失败: {str(e)}")
        
        # ========== 3. 获取一次拆解产物产值数据 ==========
        disassembly_data = app_data.get_data('disassembly_data')
        # 按(分类, 拆解产物编码)分组统计物料产值
        output_value_data = {}  # {(分类, 拆解产物编码): {'产值': 0, '重量': 0, '名称': ''}}
        
        if disassembly_data is not None and not disassembly_data.empty:
            if '类别' in disassembly_data.columns:
                # 筛选类别为"拆解产物"的记录
                product_data = disassembly_data[disassembly_data['类别'] == '拆解产物'].copy()
                
                if not product_data.empty:
                    # 获取价格数据
                    from data.base_data.price_data import load_price_data
                    price_df = load_price_data()
                    
                    price_mapping = {}
                    if price_df is not None and not price_df.empty:
                        for _, price_row in price_df.iterrows():
                            code = str(price_row['拆解产物编码']).strip()
                            price_no_tax = price_row.get('销售单价-不含税(元/KG)', 0)
                            if pd.notna(price_no_tax):
                                price_mapping[code] = float(price_no_tax)
                    
                    # 处理数据：添加分类、匹配价格、计算产值
                    for idx, row in product_data.iterrows():
                        # 获取原物料名称
                        material_name = str(row.get('原物料名称', '')).strip()
                        
                        # 根据原物料名称进行模糊匹配分类
                        category = None
                        # 🔧 修复：按优先级匹配，先匹配电脑（特别是"显示器"），再匹配电视（"彩电"）
                        # 按顺序检查：电脑（显示器优先）、电视、冰箱、空调、洗衣机
                        for cat, keywords in category_keyword_mapping.items():
                            for keyword in keywords:
                                if keyword in material_name:
                                    category = cat
                                    break
                            if category:
                                break
                        
                        # 如果没有匹配到任何分类，跳过该记录
                        if not category:
                            continue
                        
                        # 获取拆解产物编码
                        product_code = str(row.get('拆解产物编码', '')).strip()
                        if not product_code:
                            continue
                        
                        # 获取计算结果(KG)
                        calculated_weight = row.get('计算结果(KG)', 0)
                        try:
                            calculated_weight = float(calculated_weight) if pd.notna(calculated_weight) else 0
                        except (ValueError, TypeError):
                            calculated_weight = 0
                        
                        # 匹配价格
                        price_no_tax = price_mapping.get(product_code, 0)
                        
                        # 计算物料产值
                        if price_no_tax < 0:
                            material_value = 0
                        else:
                            material_value = calculated_weight * price_no_tax
                        
                        # 使用(分类, 拆解产物编码)作为key
                        key = (category, product_code)
                        if key not in output_value_data:
                            # 优先使用disassembly_data中的名称，如果为空则从产品数据中查找
                            product_name = str(row.get('拆解产物名称', '')).strip()
                            if not product_name and product_code in product_code_to_name:
                                product_name = product_code_to_name[product_code]
                            
                            output_value_data[key] = {
                                '产值': 0.0,
                                '重量': 0.0,
                                '名称': product_name
                            }
                        
                        output_value_data[key]['产值'] += material_value
                        output_value_data[key]['重量'] += calculated_weight
        
        # ========== 4. 获取生产成本分摊数据（拆解产物分摊成本） ==========
        # 直接计算生产成本分摊的核心部分
        categories = ['电视', '电脑', '冰箱', '空调', '洗衣机']
        category_allocation_cost = {cat: 0.0 for cat in categories}
        
        try:
            # 获取拆解产物价值（按分类汇总）
            category_product_value = {cat: 0.0 for cat in categories}
            for (category, product_code), data in output_value_data.items():
                if category in categories:
                    category_product_value[category] += data.get('产值', 0.0)
            
            # 获取基金补贴收入（按分类汇总）- 使用groupby精确匹配，与生产成本分摊页面一致
            category_subsidy_income = {cat: 0.0 for cat in categories}
            subsidy_income_data = app_data.get_data('subsidy_income_data')
            if subsidy_income_data is not None and not subsidy_income_data.empty:
                if '补贴大类' in subsidy_income_data.columns and '基金补贴收入(元)' in subsidy_income_data.columns:
                    subsidy_income_data['基金补贴收入(元)'] = pd.to_numeric(
                        subsidy_income_data['基金补贴收入(元)'], errors='coerce'
                    ).fillna(0)
                    
                    # 将补贴大类映射到产品类型（五大类：电视机、冰箱、洗衣机、空调、电脑）
                    def map_subsidy_category_to_product_type(subsidy_category):
                        """将补贴大类映射到产品类型"""
                        subsidy_category = str(subsidy_category).strip()
                        # 空调的细分补贴大类
                        if subsidy_category in ['整机', '内机', '外机']:
                            return '空调'
                        # 电脑的细分补贴大类
                        elif subsidy_category in ['笔记本', '显示器', '主机']:
                            return '电脑'
                        # 其他直接返回（电视机、冰箱、洗衣机）
                        elif subsidy_category in ['电视机', '冰箱', '洗衣机']:
                            return subsidy_category
                        else:
                            # 默认返回原值（兼容旧数据）
                            return subsidy_category
                    
                    # 添加产品类型列
                    subsidy_income_data['产品类型'] = subsidy_income_data['补贴大类'].apply(map_subsidy_category_to_product_type)
                    
                    # 按产品类型分组汇总
                    product_type_groups = subsidy_income_data.groupby('产品类型')['基金补贴收入(元)'].sum()
                    
                    # 映射产品类型到四机一脑分类
                    for product_type, total_subsidy in product_type_groups.items():
                        product_type = str(product_type).strip()
                        # 映射：电视机 -> 电视
                        if product_type == '电视机':
                            product_type = '电视'
                        
                        if product_type in categories:
                            category_subsidy_income[product_type] += float(total_subsidy)
            
            # 获取直接材料
            category_direct_material = {cat: 0.0 for cat in categories}
            manual_data = app_data.get_data('extracted_data_manual')
            if manual_data is not None and not manual_data.empty:
                cost_data = calculate_material_cost(manual_data)
                if '类别' in cost_data.columns:
                    old_machine_data = cost_data[cost_data['类别'] == '旧机'].copy()
                    if '物料描述' in old_machine_data.columns and '拆解物原料成本' in old_machine_data.columns:
                        old_machine_data['拆解物原料成本'] = pd.to_numeric(
                            old_machine_data['拆解物原料成本'], errors='coerce'
                        ).fillna(0)
                        for category in categories:
                            mask = old_machine_data['物料描述'].astype(str).str.contains(category, case=False, na=False)
                            category_direct_material[category] = float(
                                old_machine_data.loc[mask, '拆解物原料成本'].sum()
                            )
            
            # 获取直接人工
            category_direct_labor = {cat: 0.0 for cat in categories}
            try:
                direct_labor_result = calculate_direct_labor_cost(app_data, prediction_period)
                product_category_stats = direct_labor_result.get('product_category_stats', {})
                for cat in categories:
                    stats = product_category_stats.get(cat, {'wage': 0.0, 'fixed_cost': 0.0})
                    category_direct_labor[cat] = stats.get('wage', 0.0) + stats.get('fixed_cost', 0.0)
            except Exception as e:
                print(f"获取直接人工数据失败: {str(e)}")
            
            category_manufacturing_cost = {cat: 0.0 for cat in categories}
            try:
                mfg_by_category = collect_production_manufacturing_cost_by_category(
                    app_data, prediction_period
                )
                for cat in categories:
                    category_manufacturing_cost[cat] = mfg_by_category.get(cat, 0.0)
            except Exception as e:
                print(f"获取制造费用数据失败: {str(e)}")
            
            # 计算每个分类的生产成本小计和拆解产物分摊成本
            for category in categories:
                direct_material = category_direct_material[category]
                direct_labor = category_direct_labor[category]
                manufacturing_cost = category_manufacturing_cost[category]
                product_value = category_product_value[category]
                subsidy_income = category_subsidy_income[category]
                
                # 计算生产成本小计
                production_cost_subtotal = direct_material + direct_labor + manufacturing_cost
                
                # 计算分摊比例
                total_revenue = subsidy_income + product_value
                if total_revenue > 0:
                    product_ratio = product_value / total_revenue
                else:
                    product_ratio = 0.0
                
                # 计算拆解产物分摊成本
                product_allocation_cost = production_cost_subtotal * product_ratio
                category_allocation_cost[category] = product_allocation_cost
                
        except Exception as e:
            print(f"获取生产成本分摊数据失败: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # ========== 5. 计算每个拆解产物编码的期末单位成本 ==========
        # 收集所有拆解产物编码（从期初库存和产值数据中）
        all_product_codes = set(initial_inventory.keys())
        for (category, product_code) in output_value_data.keys():
            all_product_codes.add(product_code)
        
        # 按拆解产物编码分组，处理同一编码多个分类的情况
        product_code_groups = {}  # {拆解产物编码: [(分类, 产值数据), ...]}
        for (category, product_code), data in output_value_data.items():
            if product_code not in product_code_groups:
                product_code_groups[product_code] = []
            product_code_groups[product_code].append((category, data))
        
        for product_code in all_product_codes:
            # 期初库存数据（同一编码的所有分类共享期初库存）
            initial_amount = initial_inventory.get(product_code, {}).get('金额', 0.0)
            initial_quantity = initial_inventory.get(product_code, {}).get('数量', 0.0)
            
            # 获取该编码的所有分类数据
            category_data_list = product_code_groups.get(product_code, [])
            
            # 如果同一拆解产物编码存在多个分类，需要汇总所有分类的数据
            if len(category_data_list) > 1:
                # 多个分类情况：汇总所有分类的数据
                total_material_value = 0.0
                total_production_weight = 0.0
                total_production_cost = 0.0
                categories_list = []
                product_name = ''
                
                for category, data in category_data_list:
                    material_value = data.get('产值', 0.0)
                    production_weight = data.get('重量', 0.0)
                    name = data.get('名称', '').strip()
                    
                    total_material_value += material_value
                    total_production_weight += production_weight
                    categories_list.append(category)
                    
                    if not product_name and name:
                        product_name = name
                    
                    # 计算每个分类的本期生产成本
                    category_total = category_product_value.get(category, 0.0)
                    category_cost = category_allocation_cost.get(category, 0.0)
                    
                    if category_total > 0 and material_value > 0:
                        production_cost = (material_value / category_total) * category_cost
                        total_production_cost += production_cost
                
                # 如果名称为空，尝试从其他来源匹配
                if not product_name and product_code in product_code_to_name:
                    product_name = product_code_to_name[product_code]
                elif not product_name and product_code in extracted_code_to_name:
                    product_name = extracted_code_to_name[product_code]
                
                # 计算期末单位成本（汇总所有分类）
                total_amount = initial_amount + total_production_cost
                total_qty = initial_quantity + total_production_weight
                
                unit_cost = 0.0
                if total_qty > 0:
                    unit_cost = total_amount / total_qty
                
                # 对于多个分类的情况，为每个分类创建一条记录，但使用汇总后的数据
                for category in categories_list:
                    category_total = category_product_value.get(category, 0.0)
                    category_cost = category_allocation_cost.get(category, 0.0)
                    category_material_value = next((d.get('产值', 0.0) for c, d in category_data_list if c == category), 0.0)
                    
                    # 计算该分类的本期生产成本
                    category_production_cost = 0.0
                    if category_total > 0 and category_material_value > 0:
                        category_production_cost = (category_material_value / category_total) * category_cost
                    
                    result_data.append({
                        '拆解产物编码': product_code,
                        '拆解产物名称': product_name,
                        '分类': category,
                        '期初库存数量': round(initial_quantity, 6) if category == categories_list[0] else 0.0,  # 只在第一个分类显示期初库存
                        '期初库存金额': round(initial_amount, 2) if category == categories_list[0] else 0.0,  # 只在第一个分类显示期初库存
                        '当期生产重量': round(next((d.get('重量', 0.0) for c, d in category_data_list if c == category), 0.0), 6),
                        '物料产值': round(category_material_value, 2),
                        '分类产值总数': round(category_total, 2),
                        '分类拆解产物分摊成本': round(category_cost, 2),
                        '本期生产成本': round(category_production_cost, 2),
                        '期末单位成本': round(unit_cost, 6)  # 使用汇总后的单位成本
                    })
            else:
                # 单个分类情况：正常处理
                if not category_data_list:
                    # 如果没有产值数据，只有期初库存，尝试从产品名称匹配分类
                    product_name = ''
                    if product_code in product_code_to_name:
                        product_name = product_code_to_name[product_code]
                    elif product_code in extracted_code_to_name:
                        product_name = extracted_code_to_name[product_code]
                    
                    # 尝试从产品名称匹配分类
                    category = None
                    if product_name:
                        for cat, keywords in category_keyword_mapping.items():
                            for keyword in keywords:
                                if keyword in product_name:
                                    category = cat
                                    break
                            if category:
                                break
                    
                    # 如果仍然无法匹配分类，跳过该记录（不产生"其他"分类）
                    if not category:
                        continue
                    
                    # 只有期初库存，没有产值数据
                    material_value = 0.0
                    production_weight = 0.0
                    production_cost = 0.0
                    category_total = category_product_value.get(category, 0.0)
                    category_cost = category_allocation_cost.get(category, 0.0)
                    
                    # 计算期末单位成本（只有期初库存）
                    total_amount = initial_amount
                    total_qty = initial_quantity
                    unit_cost = 0.0
                    if total_qty > 0:
                        unit_cost = total_amount / total_qty
                    
                    result_data.append({
                        '拆解产物编码': product_code,
                        '拆解产物名称': product_name,
                        '分类': category,
                        '期初库存数量': round(initial_quantity, 6),
                        '期初库存金额': round(initial_amount, 2),
                        '当期生产重量': round(production_weight, 6),
                        '物料产值': round(material_value, 2),
                        '分类产值总数': round(category_total, 2),
                        '分类拆解产物分摊成本': round(category_cost, 2),
                        '本期生产成本': round(production_cost, 2),
                        '期末单位成本': round(unit_cost, 6)
                    })
                else:
                    # 有产值数据的情况
                    category = category_data_list[0][0]
                    output_info = category_data_list[0][1]
                    material_value = output_info.get('产值', 0.0)
                    production_weight = output_info.get('重量', 0.0)
                    product_name = output_info.get('名称', '').strip()
                    
                    # 如果名称为空，按优先级尝试匹配
                    if not product_name and product_code in product_code_to_name:
                        product_name = product_code_to_name[product_code]
                    elif not product_name and product_code in extracted_code_to_name:
                        product_name = extracted_code_to_name[product_code]
                    
                    # 计算本期生产成本
                    production_cost = 0.0
                    category_total = category_product_value.get(category, 0.0)
                    category_cost = category_allocation_cost.get(category, 0.0)
                    
                    if category_total > 0 and material_value > 0:
                        production_cost = (material_value / category_total) * category_cost
                    
                    # 计算期末单位成本
                    total_amount = initial_amount + production_cost
                    total_qty = initial_quantity + production_weight
                    
                    unit_cost = 0.0
                    if total_qty > 0:
                        unit_cost = total_amount / total_qty
                    
                    result_data.append({
                        '拆解产物编码': product_code,
                        '拆解产物名称': product_name,
                        '分类': category,
                        '期初库存数量': round(initial_quantity, 6),
                        '期初库存金额': round(initial_amount, 2),
                        '当期生产重量': round(production_weight, 6),
                        '物料产值': round(material_value, 2),
                        '分类产值总数': round(category_total, 2),
                        '分类拆解产物分摊成本': round(category_cost, 2),
                        '本期生产成本': round(production_cost, 2),
                        '期末单位成本': round(unit_cost, 6)
                    })
        
        # ========== 6. 获取销售数量（从销售收益数据中统计） ==========
        # 优先使用手工数据，否则使用系统数据
        saleable_data_manual = app_data.get_data('saleable_data_manual')
        saleable_data = app_data.get_data('saleable_data')
        
        # 如果遇到重复"拆解产物编码"，需根据"原物料名称"，进行"四机一脑分类"统计"计算结果(KG)"
        # 使用(拆解产物编码, 分类)作为key
        sales_quantity_by_code_category = {}  # {(拆解产物编码, 分类): 销售数量(KG)}
        
        # 四机一脑分类映射函数（与calculate_disassembly_product_cost函数中的规则一致）
        def map_to_category_for_sales(material_name):
            """根据原物料名称映射到四机一脑分类"""
            if not material_name or pd.isna(material_name):
                return None
            name = str(material_name)
            
            # 🔧 修复：优先匹配"显示器"（电脑类），避免与"彩电"冲突
            # 电脑映射规则（优先检查，特别是"显示器"）
            if ('显示器' in name):
                return '电脑'
            if ('电脑' in name or '笔记本' in name or 
                '主机' in name or '废旧金属黑色金属-铁及其合金-电子枪' in name):
                return '电脑'
            
            # 电视映射规则（按优先级顺序匹配）
            if ('CRT其它机壳破碎塑料' in name or '线路板边框破碎塑料' in name or 
                '废旧玻璃电子枪' in name or '废旧金属荫罩压块铁' in name or 
                '黑白' in name):
                return '电视'
            if ('电视' in name or '彩电' in name or '等离子' in name):
                return '电视'
            
            # 冰箱
            if '冰箱' in name or '冰柜' in name:
                return '冰箱'
            
            # 空调
            if '空调' in name:
                return '空调'
            
            # 洗衣机
            if '洗衣机' in name or '双缸' in name:
                return '洗衣机'
            
            return None
        
        # 确定使用的数据源
        revenue_data = None
        if saleable_data_manual is not None and not saleable_data_manual.empty:
            if '类别' in saleable_data_manual.columns and '拆解产物编码' in saleable_data_manual.columns and '计算结果(KG)' in saleable_data_manual.columns:
                revenue_data = saleable_data_manual
        elif saleable_data is not None and not saleable_data.empty:
            if '类别' in saleable_data.columns and '拆解产物编码' in saleable_data.columns and '计算结果(KG)' in saleable_data.columns:
                revenue_data = saleable_data
        
        if revenue_data is not None and not revenue_data.empty:
            # 筛选类别为"拆解产物"或"一次拆解产物"的数据
            filtered_data = revenue_data[
                (revenue_data['类别'] == '拆解产物') | (revenue_data['类别'] == '一次拆解产物')
            ].copy()
            
            if not filtered_data.empty:
                # 确保计算结果(KG)是数值类型
                filtered_data['计算结果(KG)'] = pd.to_numeric(
                    filtered_data['计算结果(KG)'], errors='coerce'
                ).fillna(0)
                
                # 如果有原物料名称列，按(拆解产物编码, 四机一脑分类)分组统计
                if '原物料名称' in filtered_data.columns:
                    for _, row in filtered_data.iterrows():
                        product_code = str(row.get('拆解产物编码', '')).strip()
                        if not product_code:
                            continue
                        
                        row_category = str(row.get('类别', '')).strip()
                        category = None
                        
                        # 🔧 修复：对于类别为"一次拆解产物"的数据，优先使用拆解产物名称进行分类
                        # 因为原物料名称可能不准确（如"R-废旧玻璃 彩电-CRT锥玻璃-显示器锥玻璃"同时包含"彩电"和"显示器"）
                        if row_category == '一次拆解产物':
                            # 优先从拆解产物名称推断分类
                            product_name = str(row.get('拆解产物名称', '')).strip()
                            if product_name:
                                category = map_to_category_for_sales(product_name)
                            
                            # 如果拆解产物名称无法分类，再尝试原物料名称
                            if not category:
                                material_name = row.get('原物料名称', '')
                                if material_name:
                                    category = map_to_category_for_sales(material_name)
                        else:
                            # 对于类别为"拆解产物"的数据，使用原物料名称分类
                            material_name = row.get('原物料名称', '')
                            category = map_to_category_for_sales(material_name)
                        
                        # 如果仍然无法分类，跳过该记录（不统计未分类的数据）
                        if not category:
                            continue
                        
                        calculated_weight = float(row.get('计算结果(KG)', 0) or 0)
                        
                        key = (product_code, category)
                        if key not in sales_quantity_by_code_category:
                            sales_quantity_by_code_category[key] = 0.0
                        
                        sales_quantity_by_code_category[key] += calculated_weight
                    
                    # 调试输出：打印统计结果
                    print(f"  销售数量统计完成，共 {len(sales_quantity_by_code_category)} 组")
                    if len(sales_quantity_by_code_category) > 0:
                        print(f"  统计示例（前10组）:")
                        count = 0
                        for (code, cat), qty in sales_quantity_by_code_category.items():
                            if count < 10:
                                print(f"    ({code}, {cat}): {qty:.6f} KG")
                                count += 1
                        # 特别输出编码811053082的统计信息
                        for (code, cat), qty in sales_quantity_by_code_category.items():
                            if code == '811053082':
                                print(f"    🔍 编码811053082 ({cat}): {qty:.6f} KG")
                else:
                    # 如果没有原物料名称列，按拆解产物编码统计
                    for _, row in filtered_data.iterrows():
                        product_code = str(row.get('拆解产物编码', '')).strip()
                        if not product_code:
                            continue
                        
                        calculated_weight = float(row.get('计算结果(KG)', 0) or 0)
                        
                        key = (product_code, '未分类')
                        if key not in sales_quantity_by_code_category:
                            sales_quantity_by_code_category[key] = 0.0
                        
                        sales_quantity_by_code_category[key] += calculated_weight
        
        # 为每条记录添加销售数量（根据拆解产物编码和分类精确匹配）
        for record in result_data:
            product_code = str(record.get('拆解产物编码', '')).strip()
            category = str(record.get('分类', '')).strip()
            
            # 按(拆解产物编码, 分类)精确匹配
            key = (product_code, category)
            sales_quantity = sales_quantity_by_code_category.get(key, 0.0)
            
            record['销售数量'] = round(sales_quantity, 6)
            
            # 计算一次拆解产物销售成本 = 销售数量 × 期末单位成本
            unit_cost = record.get('期末单位成本', 0) or 0
            sales_cost = sales_quantity * unit_cost
            record['一次拆解产物销售成本'] = round(sales_cost, 6)
            
            # 调试输出：对于特定编码，打印匹配信息
            if product_code == '811053065' or product_code == '811053082':
                print(f"  匹配销售数量: 拆解产物编码={product_code}, 分类={category}, 销售数量={sales_quantity:.6f}")
        
        # ========== 7. 获取被减扣数据中各处置类别的数量统计 ==========
        # 按拆解产物编码和处置类别统计"计算结果(KG)"
        # 🔧 重要修复：使用手工编辑的数据（deducted_data_manual），而不是只读数据（deducted_data）
        # 确保编辑后的数据参与计算
        disposal_quantity_by_code = {}  # {拆解产物编码: {'付费处置': 0, '内转屏处置': 0, ...}}
        
        # 🔧 架构重构：使用 deducted_data_manual，不再使用 deducted_data (只读)
        deducted_data = app_data.get_data('deducted_data_manual')
        if deducted_data is None or deducted_data.empty:
            # 如果手工数据为空且未修改，使用原始备份作为后备
            deducted_data_modified = app_data.get_data('deducted_data_modified')
            if not deducted_data_modified:
                original_data = app_data.get_data('original_deducted_data')
                if original_data is not None and not original_data.empty:
                    deducted_data = original_data
                    print("⚠️ 使用原始备份被减扣数据（手工数据为空）")
                else:
                    print("⚠️ 没有可用的被减扣数据")
            else:
                print("⚠️ 手工数据为空但已标记为修改")
        else:
            print("✅ 使用被减扣数据(手工)进行成本计算")
        
        if deducted_data is not None and not deducted_data.empty:
            # 检查必要的列是否存在
            required_columns = ['拆解产物编码', '处置类别', '计算结果(KG)']
            if all(col in deducted_data.columns for col in required_columns):
                # 确保计算结果(KG)是数值类型
                deducted_data['计算结果(KG)'] = pd.to_numeric(
                    deducted_data['计算结果(KG)'], errors='coerce'
                ).fillna(0)
                
                # 按拆解产物编码和处置类别分组统计
                for _, row in deducted_data.iterrows():
                    product_code = str(row.get('拆解产物编码', '')).strip()
                    if not product_code:
                        continue
                    
                    disposal_category = str(row.get('处置类别', '')).strip()
                    if not disposal_category:
                        continue
                    
                    calculated_weight = float(row.get('计算结果(KG)', 0) or 0)
                    
                    # 初始化该编码的统计字典
                    if product_code not in disposal_quantity_by_code:
                        disposal_quantity_by_code[product_code] = {
                            '付费处置': 0.0,
                            '内转屏处置': 0.0,
                            '内转印制板处置': 0.0,
                            '内转荧光灯处置': 0.0,
                            '深加工-打包铁': 0.0,
                            '深加工-塑料一破': 0.0
                        }
                    
                    # 根据处置类别累加数量
                    if disposal_category == '付费处置':
                        disposal_quantity_by_code[product_code]['付费处置'] += calculated_weight
                    elif disposal_category == '内转屏处置':
                        disposal_quantity_by_code[product_code]['内转屏处置'] += calculated_weight
                    elif disposal_category == '内转印制板处置':
                        disposal_quantity_by_code[product_code]['内转印制板处置'] += calculated_weight
                    elif disposal_category == '内转荧光灯处置':
                        disposal_quantity_by_code[product_code]['内转荧光灯处置'] += calculated_weight
                    elif disposal_category == '深加工-打包铁':
                        disposal_quantity_by_code[product_code]['深加工-打包铁'] += calculated_weight
                    elif disposal_category == '深加工-塑料一破':
                        disposal_quantity_by_code[product_code]['深加工-塑料一破'] += calculated_weight
        
        # 为每条记录添加各处置类别的数量
        for record in result_data:
            product_code = str(record.get('拆解产物编码', '')).strip()
            
            # 获取该编码的处置类别统计
            disposal_stats = disposal_quantity_by_code.get(product_code, {
                '付费处置': 0.0,
                '内转屏处置': 0.0,
                '内转印制板处置': 0.0,
                '内转荧光灯处置': 0.0,
                '深加工-打包铁': 0.0,
                '深加工-塑料一破': 0.0
            })
            
            # 添加6个数量字段
            record['付费处置（数量）'] = round(disposal_stats['付费处置'], 6)
            record['内转屏处置（数量）'] = round(disposal_stats['内转屏处置'], 6)
            record['内转印制板处置（数量）'] = round(disposal_stats['内转印制板处置'], 6)
            record['内转荧光灯处置（数量）'] = round(disposal_stats['内转荧光灯处置'], 6)
            record['深加工-打包铁（数量）'] = round(disposal_stats['深加工-打包铁'], 6)
            record['深加工-塑料一破（数量）'] = round(disposal_stats['深加工-塑料一破'], 6)
            
            # 计算并添加6个成本字段：成本 = 数量 × 期末单位成本
            unit_cost = record.get('期末单位成本', 0) or 0
            record['付费处置（成本）'] = round(record['付费处置（数量）'] * unit_cost, 2)
            record['内转屏处置（成本）'] = round(record['内转屏处置（数量）'] * unit_cost, 2)
            record['内转印制板处置（成本）'] = round(record['内转印制板处置（数量）'] * unit_cost, 2)
            record['内转荧光灯处置（成本）'] = round(record['内转荧光灯处置（数量）'] * unit_cost, 2)
            record['深加工-打包铁（成本）'] = round(record['深加工-打包铁（数量）'] * unit_cost, 2)
            record['深加工-塑料一破（成本）'] = round(record['深加工-塑料一破（数量）'] * unit_cost, 2)
            
            # 计算期末库存数量 = 期初库存数量 + 当期生产重量 - 销售数量 - 所有处置数量
            initial_quantity = record.get('期初库存数量', 0) or 0
            production_weight = record.get('当期生产重量', 0) or 0
            sales_quantity = record.get('销售数量', 0) or 0
            paid_disposal_qty = record.get('付费处置（数量）', 0) or 0
            screen_disposal_qty = record.get('内转屏处置（数量）', 0) or 0
            pcb_disposal_qty = record.get('内转印制板处置（数量）', 0) or 0
            lamp_disposal_qty = record.get('内转荧光灯处置（数量）', 0) or 0
            iron_processing_qty = record.get('深加工-打包铁（数量）', 0) or 0
            plastic_processing_qty = record.get('深加工-塑料一破（数量）', 0) or 0
            
            ending_inventory_quantity = (initial_quantity + production_weight - sales_quantity - 
                                        paid_disposal_qty - screen_disposal_qty - pcb_disposal_qty - 
                                        lamp_disposal_qty - iron_processing_qty - plastic_processing_qty)
            record['期末库存数量'] = round(ending_inventory_quantity, 6)
            
            # 计算期末库存成本 = 期末库存数量 × 期末单位成本
            ending_inventory_cost = ending_inventory_quantity * unit_cost
            record['期末库存成本'] = round(ending_inventory_cost, 2)
        
        # 按拆解产物编码和分类排序
        result_data.sort(key=lambda x: (x['拆解产物编码'], x['分类']))
        
        return result_data
        
    except Exception as e:
        print(f"计算一次拆解产物成本失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return []


@cost_forecast_bp.route('/disassembly-product-cost', methods=['GET'])
def get_disassembly_product_cost():
    """获取一次拆解产物成本计算数据"""
    try:
        app_data = get_session_data_manager()
        prediction_period = int(request.args.get('prediction_period', 1))
        force_refresh = request.args.get('force_refresh', 'false').lower() == 'true'
        
        # 检查数据是否已被清除
        data_cleared = app_data.get_data('__data_cleared__')
        if data_cleared:
            # 数据已被清除，直接返回空数据
            return jsonify({
                'success': True,
                'data': []
            })
        
        # 检查缓存（使用版本号确保使用新的分类映射规则）
        # 版本号：v2 - 使用新的分类映射规则，不产生"其他"分类
        cache_key = f'disassembly_product_cost_result_v2_{prediction_period}'
        if not force_refresh:
            cached_result = app_data.get_data(cache_key)
            if cached_result is not None and len(cached_result) > 0:
                return jsonify({
                    'success': True,
                    'data': cached_result,
                    'from_cache': True
                })
        
        # 计算数据
        result_data = calculate_disassembly_product_cost(app_data, prediction_period)
        
        # 缓存结果
        app_data.set_data(cache_key, result_data)
        
        return jsonify({
            'success': True,
            'data': result_data
        })
        
    except Exception as e:
        print(f"获取一次拆解产物成本数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@cost_forecast_bp.route('/disassembly-product-cost/export', methods=['GET'])
def export_disassembly_product_cost():
    """导出一拆解产物成本计算数据到Excel"""
    try:
        app_data = get_session_data_manager()
        prediction_period = int(request.args.get('prediction_period', 1))
        
        # 计算数据
        result_data = calculate_disassembly_product_cost(app_data, prediction_period)
        
        if not result_data or len(result_data) == 0:
            return jsonify({
                'success': False,
                'error': '没有可导出的一次拆解产物成本数据'
            }), 400
        
        # 转换为DataFrame
        export_df = pd.DataFrame(result_data)
        
        # 追加合计行（期末单位成本显示 "-"，其余数值列求和）
        _decimals_6_cols = {
            '期初库存数量', '当期生产重量', '销售数量',
            '付费处置（数量）', '内转屏处置（数量）',
            '内转印制板处置（数量）', '内转荧光灯处置（数量）',
            '深加工-打包铁（数量）', '深加工-塑料一破（数量）',
            '期末库存数量'
        }
        _text_cols = {'拆解产物编码', '拆解产物名称', '分类', '期末单位成本'}
        _total_row = {c: '' for c in export_df.columns}
        if '拆解产物编码' in export_df.columns:
            _total_row['拆解产物编码'] = '合计'
        if '期末单位成本' in export_df.columns:
            _total_row['期末单位成本'] = '-'
        for _c in export_df.columns:
            if _c in _text_cols:
                continue
            _sum_val = pd.to_numeric(export_df[_c], errors='coerce').fillna(0).sum()
            _total_row[_c] = round(float(_sum_val), 6 if _c in _decimals_6_cols else 2)
        export_df = pd.concat([export_df, pd.DataFrame([_total_row])], ignore_index=True)
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 写入数据
            export_df.to_excel(writer, sheet_name='一次拆解产物成本计算', index=False)
            
            # 设置列宽和样式
            worksheet = writer.sheets['一次拆解产物成本计算']
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF", name="仿宋")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # 列宽设置
            column_widths = {
                '拆解产物编码': 18,
                '拆解产物名称': 30,
                '分类': 12,
                '期初库存金额': 18,
                '期初库存数量': 18,
                '物料产值': 18,
                '分类产值总数': 18,
                '分类拆解产物分摊成本': 22,
                '本期生产成本': 18,
                '当期生产重量': 18,
                '期末单位成本': 18,
                '销售数量': 18,
                '一次拆解产物销售成本': 22,
                '付费处置（数量）': 18,
                '内转屏处置（数量）': 18,
                '内转印制板处置（数量）': 20,
                '内转荧光灯处置（数量）': 20,
                '深加工-打包铁（数量）': 20,
                '深加工-塑料一破（数量）': 20,
                '付费处置（成本）': 18,
                '内转屏处置（成本）': 18,
                '内转印制板处置（成本）': 20,
                '内转荧光灯处置（成本）': 20,
                '深加工-打包铁（成本）': 20,
                '深加工-塑料一破（成本）': 20,
                '期末库存数量': 18,
                '期末库存成本': 18
            }
            
            for col in range(1, len(export_df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                
                # 设置列宽
                col_letter = get_column_letter(col)
                col_name = export_df.columns[col - 1]
                if col_name in column_widths:
                    worksheet.column_dimensions[col_letter].width = column_widths[col_name]
                else:
                    worksheet.column_dimensions[col_letter].width = 18
            
            # 设置数据行样式
            data_font = Font(name="仿宋")
            total_row_idx = len(export_df) + 1  # 合计行在Excel中的行号（含表头）
            total_fill = PatternFill(start_color="E8E6DC", end_color="E8E6DC", fill_type="solid")
            total_font = Font(name="仿宋", bold=True)
            for row in range(2, len(export_df) + 2):
                is_total_row = (row == total_row_idx)
                for col in range(1, len(export_df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = data_font
                    cell.alignment = center_alignment
                    
                    # 期末单位成本列高亮
                    col_name = export_df.columns[col - 1]
                    if col_name == '期末单位成本':
                        cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                        cell.font = Font(name="仿宋", bold=True)
                    
                    # 合计行样式：浅灰底、加粗（覆盖期末单位成本高亮之外的字体）
                    if is_total_row:
                        cell.fill = total_fill
                        cell.font = total_font
        
        output.seek(0)
        filename = f'一次拆解产物成本计算_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"导出一拆解产物成本计算失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def calculate_deep_processing_product_cost(app_data, prediction_period=1):
    """
    计算深加工产物成本
    
    计算公式：
    深加工产物期末单位成本 =（期初库存金额 + 本期生产成本） / 
                            （期初库存数量 + 当期生产重量）
    
    数据来源：
    1. 深加工产物期初库存金额：提取结果，类别为"打包铁"、"一破"、"屏"或"印制板"，按物料代码统计"价值"列
    2. 深加工产物期初库存数量：提取结果，类别为"打包铁"、"一破"、"屏"或"印制板"，按物料代码统计"非限制使用的库存"
    3. 深加工产物本期生产成本：每个深加工产物编码对应的物料产值/分类产值总数 × 分类深加工产物分摊成本
    4. 当期生产重量：深加工拆解产物产值页面，按"四机一脑类别"和"深加工产物编码"汇总"深加工结果(KG)"
    
    Args:
        app_data: 数据管理器实例
        prediction_period: 预测期数（默认1）
        
    Returns:
        list: 包含每个深加工产物编码的详细计算结果
    """
    try:
        result_data = []
        
        # 分类关键词映射规则（四机一脑类别）
        from collections import OrderedDict
        category_keyword_mapping = OrderedDict([
            ('电脑', ['显示器', '电脑', '笔记本', '主机', '废旧金属黑色金属-铁及其合金-电子枪']),
            ('电视', ['电视', '彩电', 'CRT其它机壳破碎塑料', '线路板边框破碎塑料', '等离子', '废旧玻璃电子枪', '废旧金属荫罩压块铁', '黑白']),
            ('冰箱', ['冰箱', '冰柜']),
            ('空调', ['空调']),
            ('洗衣机', ['洗衣机', '双缸'])
        ])
        
        # 根据产品名称映射四机一脑类别的函数
        def map_to_four_category_by_name(product_name):
            """
            根据产品名称映射到四机一脑类别
            
            Args:
                product_name: 产品名称字符串
                
            Returns:
                str: 四机一脑类别（'电视'、'电脑'、'冰箱'、'空调'、'洗衣机'），如果没有匹配则返回 None
            """
            if not product_name or pd.isna(product_name):
                return None
            
            name_str = str(product_name).strip()
            if not name_str:
                return None
            
            # 按优先级匹配：电脑（显示器优先）、电视、冰箱、空调、洗衣机
            for category, keywords in category_keyword_mapping.items():
                for keyword in keywords:
                    if keyword in name_str:
                        return category
            
            return None
        
        # ========== 1. 获取深加工产物期初库存数据（从提取结果） ==========
        # 辅助函数：统一物料代码格式（移除.0后缀，统一字符串格式）
        def normalize_material_code(code):
            """统一物料代码格式，移除.0后缀，去除空格"""
            if pd.isna(code) or code == '':
                return ''
            code_str = str(code).strip()
            # 移除.0后缀（如果存在）
            if code_str.endswith('.0'):
                code_str = code_str[:-2]
            return code_str
        
        # 获取提取结果数据（从Excel的"提取结果"sheet读取，或从自动提取的数据）
        # 注意：extracted_data 是从源数据自动提取的，如果源数据是从"提取结果"sheet读取的，
        # 那么 extracted_data 就包含了"提取结果"sheet的数据
        extracted_data = app_data.get_data('extracted_data')
        initial_inventory = {}  # {物料代码: {'金额': 0, '数量': 0}}
        extracted_code_to_name = {}  # {物料代码: 物料描述} - 用于匹配深加工产物名称
        
        # 调试信息：检查提取结果数据
        if extracted_data is None or extracted_data.empty:
            print("⚠️ 警告：提取结果数据为空或不存在")
        else:
            print(f"📊 提取结果数据总行数: {len(extracted_data)}")
            if '类别' in extracted_data.columns:
                # 检查所有类别
                unique_categories = extracted_data['类别'].unique()
                print(f"📋 提取结果中的类别: {list(unique_categories)}")
                print(f"📋 提取结果中的类别数量: {len(unique_categories)}")
                
                # 检查类别列的数据类型和实际值
                category_counts = extracted_data['类别'].value_counts()
                print(f"📋 各类别的记录数:")
                for cat, count in category_counts.items():
                    print(f"   - {cat}: {count} 条")
                
                # 筛选类别为"打包铁"、"一破"、"屏"和"印制板"的数据（用于深加工产物期初库存）
                deep_processing_categories = ['打包铁', '一破', '屏', '印制板']
                deep_processing_product_data = extracted_data[extracted_data['类别'].isin(deep_processing_categories)].copy()
                print(f"🔍 类别为'打包铁'、'一破'、'屏'或'印制板'的记录数: {len(deep_processing_product_data)}")
                
                if not deep_processing_product_data.empty:
                    # 显示各类别的记录数
                    category_breakdown = deep_processing_product_data['类别'].value_counts()
                    print(f"   各类别记录数: {dict(category_breakdown)}")
                
                if not deep_processing_product_data.empty:
                    # 确保必要的列存在
                    if '物料代码' in deep_processing_product_data.columns:
                        # 按物料代码分组统计
                        for _, row in deep_processing_product_data.iterrows():
                            material_code = normalize_material_code(row.get('物料代码', ''))
                            if not material_code:
                                continue
                            
                            if material_code not in initial_inventory:
                                initial_inventory[material_code] = {
                                    '金额': 0.0,
                                    '数量': 0.0
                                }
                            
                            # 统计价值列
                            if '价值' in row:
                                value = pd.to_numeric(row['价值'], errors='coerce')
                                if pd.notna(value):
                                    initial_inventory[material_code]['金额'] += float(value)
                            
                            # 统计非限制使用的库存列
                            if '非限制使用的库存' in row:
                                stock = pd.to_numeric(row['非限制使用的库存'], errors='coerce')
                                if pd.notna(stock):
                                    initial_inventory[material_code]['数量'] += float(stock)
                            
                            # 建立物料代码到物料描述的映射（用于后续匹配名称）
                            if '物料描述' in row:
                                material_desc = str(row.get('物料描述', '')).strip()
                                if material_desc and material_code not in extracted_code_to_name:
                                    extracted_code_to_name[material_code] = material_desc
                        
                        print(f"✅ 期初库存数据统计完成，共 {len(initial_inventory)} 个物料代码")
                        if len(initial_inventory) > 0:
                            # 显示前几个物料代码的统计信息
                            sample_codes = list(initial_inventory.keys())[:5]
                            for code in sample_codes:
                                info = initial_inventory[code]
                                print(f"   物料代码 {code}: 金额={info['金额']:.2f}, 数量={info['数量']:.6f}")
                    else:
                        print("⚠️ 警告：深加工产物数据中缺少'物料代码'列")
                else:
                    print("⚠️ 警告：提取结果中没有类别为'深加工产物'的数据")
            else:
                print("⚠️ 警告：提取结果数据中缺少'类别'列")
        
        # ========== 1.5. 构建深加工产物编码到拆解产物编码的逆向映射（用于期初库存回退匹配） ==========
        # 深加工产物编码（如811215806）与提取结果中的物料代码（可能是拆解产物编码如811052970）
        # 可能不一致，需要建立映射关系作为回退查找
        deep_processing_data = app_data.get_data('deep_processing_data')
        deep_to_original_map = {}  # {深加工产物编码: {拆解产物编码1, 拆解产物编码2, ...}}

        if deep_processing_data is not None and not deep_processing_data.empty:
            for _, row in deep_processing_data.iterrows():
                deep_code = normalize_material_code(row.get('深加工产物编码', ''))
                orig_code = normalize_material_code(row.get('拆解产物编码', ''))
                if deep_code and orig_code:
                    if deep_code not in deep_to_original_map:
                        deep_to_original_map[deep_code] = set()
                    deep_to_original_map[deep_code].add(orig_code)
            print(f"📋 构建深加工产物编码→拆解产物编码映射，共 {len(deep_to_original_map)} 个深加工产物编码")

        # 构建全量提取结果库存（不按类别过滤，用于回退查找）
        # 某些深加工产物的期初库存可能以拆解产物编码的形式存在于提取结果中
        full_extracted_inventory = {}  # {物料代码: {'金额': 0, '数量': 0}}
        if extracted_data is not None and not extracted_data.empty and '物料代码' in extracted_data.columns:
            for _, row in extracted_data.iterrows():
                material_code = normalize_material_code(row.get('物料代码', ''))
                if not material_code:
                    continue
                if material_code not in full_extracted_inventory:
                    full_extracted_inventory[material_code] = {
                        '金额': 0.0,
                        '数量': 0.0
                    }
                if '价值' in row:
                    value = pd.to_numeric(row['价值'], errors='coerce')
                    if pd.notna(value):
                        full_extracted_inventory[material_code]['金额'] += float(value)
                if '非限制使用的库存' in row:
                    stock = pd.to_numeric(row['非限制使用的库存'], errors='coerce')
                    if pd.notna(stock):
                        full_extracted_inventory[material_code]['数量'] += float(stock)
            # 同时补充 extracted_code_to_name（从全量提取结果中建立物料代码到物料描述的映射）
            if '物料描述' in extracted_data.columns:
                for _, row in extracted_data.iterrows():
                    material_code = normalize_material_code(row.get('物料代码', ''))
                    if material_code and material_code not in extracted_code_to_name:
                        material_desc = str(row.get('物料描述', '')).strip()
                        if material_desc:
                            extracted_code_to_name[material_code] = material_desc
            print(f"📊 全量提取结果库存构建完成，共 {len(full_extracted_inventory)} 个物料代码")

        # ========== 1.6. 构建R3系统代码到类别的映射（从内置映射表，无条件构建） ==========
        r3_to_category = {}
        try:
            from data.base_data.mapping_data import get_mapping_dataframe
            mapping_df = get_mapping_dataframe()
            if mapping_df is not None and not mapping_df.empty:
                r3_column = 'R3系统代码' if 'R3系统代码' in mapping_df.columns else ('R3代码' if 'R3代码' in mapping_df.columns else None)
                if r3_column and '类别' in mapping_df.columns:
                    for _, mapping_row in mapping_df.iterrows():
                        r3_code = str(mapping_row[r3_column]).strip()
                        category = str(mapping_row['类别']).strip()
                        if r3_code and category:
                            r3_to_category[r3_code] = category
                    print(f"📋 构建R3代码→类别映射完成，共 {len(r3_to_category)} 个映射")
        except Exception as e:
            print(f"⚠️ 构建R3代码→类别映射失败: {e}")

        # ========== 2. 获取深加工拆解产物产值数据 ==========
        # 从深加工数据中获取产值信息
        # deep_processing_data 已在上方加载

        # 按(四机一脑类别, 深加工产物编码)分组统计
        output_value_data = {}  # {(四机一脑类别, 深加工产物编码): {'产值': 0, '重量': 0, '名称': '', '类别': ''}}
        
        if deep_processing_data is not None and not deep_processing_data.empty:
            # 筛选是否减扣 == '否'的记录（只统计非减扣的深加工产物）
            if '是否减扣' in deep_processing_data.columns:
                non_deducted_data = deep_processing_data[deep_processing_data['是否减扣'] == '否'].copy()
                
                if not non_deducted_data.empty:
                    # 获取价格数据
                    from data.base_data.price_data import load_price_data
                    price_df = load_price_data()
                    
                    price_mapping = {}
                    if price_df is not None and not price_df.empty:
                        for _, price_row in price_df.iterrows():
                            code = str(price_row['拆解产物编码']).strip()
                            price_no_tax = price_row.get('销售单价-不含税(元/KG)', 0)
                            if pd.notna(price_no_tax):
                                price_mapping[code] = float(price_no_tax)
                    
                    # 处理数据：添加四机一脑类别、匹配价格、计算产值
                    # r3_to_category 已在上方 1.6 节无条件构建
                    for idx, row in non_deducted_data.iterrows():
                        # 获取原物料名称
                        material_name = str(row.get('原物料名称', '')).strip()
                        
                        # 根据原物料名称进行模糊匹配分类（四机一脑类别）
                        four_category = None
                        for cat, keywords in category_keyword_mapping.items():
                            for keyword in keywords:
                                if keyword in material_name:
                                    four_category = cat
                                    break
                            if four_category:
                                break
                        
                        # 如果根据原物料名称匹配不到，尝试根据深加工产物名称匹配
                        if not four_category:
                            deep_product_name = str(row.get('深加工产物名称', '')).strip()
                            if deep_product_name:
                                four_category = map_to_four_category_by_name(deep_product_name)
                        
                        # 如果仍然没有匹配到四机一脑类别，跳过该记录
                        if not four_category:
                            continue
                        
                        # 获取深加工产物编码（统一格式处理）
                        deep_product_code = normalize_material_code(row.get('深加工产物编码', ''))
                        if not deep_product_code:
                            continue
                        
                        # 获取深加工结果(KG)
                        deep_result_kg = row.get('深加工结果(KG)', 0)
                        try:
                            deep_result_kg = float(deep_result_kg) if pd.notna(deep_result_kg) else 0
                        except (ValueError, TypeError):
                            deep_result_kg = 0
                        
                        # 匹配价格（使用深加工产物编码）
                        price_no_tax = price_mapping.get(deep_product_code, 0)
                        
                        # 计算物料产值
                        if price_no_tax < 0:
                            material_value = 0
                        else:
                            material_value = deep_result_kg * price_no_tax
                        
                        # 获取类别（从映射表，优先深加工产物编码，回退到拆解产物编码）
                        # 只接受四个有效成本类别：打包铁、一破、屏、印制板
                        VALID_COST_CATEGORIES = {'打包铁', '一破', '屏', '印制板'}
                        mapping_category = r3_to_category.get(deep_product_code, '')
                        if mapping_category not in VALID_COST_CATEGORIES:
                            mapping_category = ''
                        if not mapping_category and deep_product_code in deep_to_original_map:
                            for orig_code in deep_to_original_map[deep_product_code]:
                                orig_cat = r3_to_category.get(orig_code, '')
                                if orig_cat in VALID_COST_CATEGORIES:
                                    mapping_category = orig_cat
                                    break
                        
                        # 使用(四机一脑类别, 深加工产物编码)作为key
                        key = (four_category, deep_product_code)
                        if key not in output_value_data:
                            deep_product_name = str(row.get('深加工产物名称', '')).strip()
                            if not deep_product_name and deep_product_code in extracted_code_to_name:
                                deep_product_name = extracted_code_to_name[deep_product_code]
                            
                            output_value_data[key] = {
                                '产值': 0.0,
                                '重量': 0.0,
                                '名称': deep_product_name,
                                '类别': mapping_category
                            }
                        
                        output_value_data[key]['产值'] += material_value
                        output_value_data[key]['重量'] += deep_result_kg
        
        # ========== 3. 按类别汇总分类产值总数 ==========
        category_total_value = {}  # {类别: 产值总数}
        for (four_category, deep_product_code), data in output_value_data.items():
            mapping_category = data.get('类别', '')
            if mapping_category:
                if mapping_category not in category_total_value:
                    category_total_value[mapping_category] = 0.0
                category_total_value[mapping_category] += data.get('产值', 0.0)
        
        # ========== 4. 获取分类深加工产物分摊成本（从一次拆解产物成本计算页面） ==========
        # 获取一次拆解产物成本计算数据
        disassembly_cost_data = calculate_disassembly_product_cost(app_data, prediction_period)
        
        # 按类别汇总对应的成本
        category_allocation_cost = {}  # {类别: 分摊成本}
        
        # 类别到成本字段的映射
        category_to_cost_field = {
            '打包铁': '深加工-打包铁（成本）',
            '一破': '深加工-塑料一破（成本）',
            '屏': '内转屏处置（成本）',
            '印制板': '内转印制板处置（成本）'
        }
        
        # 汇总一次拆解产物成本计算中各类别的成本
        for record in disassembly_cost_data:
            for category, cost_field in category_to_cost_field.items():
                if cost_field in record:
                    cost = float(record.get(cost_field, 0) or 0)
                    if category not in category_allocation_cost:
                        category_allocation_cost[category] = 0.0
                    category_allocation_cost[category] += cost
        
        # ========== 5. 计算每个深加工产物编码的期末单位成本 ==========
        # 收集所有深加工产物编码（从期初库存和产值数据中）
        all_product_codes = set(initial_inventory.keys())
        for (four_category, product_code) in output_value_data.keys():
            # 确保格式统一
            normalized_code = normalize_material_code(product_code)
            if normalized_code:
                all_product_codes.add(normalized_code)
        
        # 按深加工产物编码分组，处理同一编码多个四机一脑类别的情况
        product_code_groups = {}  # {深加工产物编码: [(四机一脑类别, 产值数据), ...]}
        for (four_category, product_code), data in output_value_data.items():
            # 确保格式统一
            normalized_code = normalize_material_code(product_code)
            if normalized_code:
                if normalized_code not in product_code_groups:
                    product_code_groups[normalized_code] = []
                product_code_groups[normalized_code].append((four_category, data))
        
        # 调试信息：显示匹配情况
        print(f"🔍 开始匹配期初库存，共有 {len(all_product_codes)} 个深加工产物编码需要匹配")
        matched_count = 0
        unmatched_codes = []
        
        # 辅助函数：获取期初库存（含回退匹配）
        def get_initial_inventory_with_fallback(product_code, inventory, fallback_inventory, deep_to_orig_map):
            """获取期初库存，优先直接匹配，未找到时回退到拆解产物编码匹配"""
            # 优先直接用深加工产物编码匹配
            result = inventory.get(product_code, {})
            amount = result.get('金额', 0.0)
            quantity = result.get('数量', 0.0)
            matched_via = 'direct'

            # 回退：通过拆解产物编码匹配
            if (amount == 0.0 and quantity == 0.0) and product_code in deep_to_orig_map:
                for orig_code in deep_to_orig_map[product_code]:
                    orig_result = fallback_inventory.get(orig_code, {})
                    orig_amount = orig_result.get('金额', 0.0)
                    orig_quantity = orig_result.get('数量', 0.0)
                    if orig_amount > 0 or orig_quantity > 0:
                        amount = orig_amount
                        quantity = orig_quantity
                        matched_via = f'fallback (via 拆解产物编码 {orig_code})'
                        break

            return amount, quantity, matched_via

        for product_code in all_product_codes:
            # 确保格式统一后再匹配
            normalized_product_code = normalize_material_code(product_code)
            # 期初库存数据（同一编码的所有四机一脑类别共享期初库存）
            initial_amount, initial_quantity, matched_via = get_initial_inventory_with_fallback(
                normalized_product_code, initial_inventory, full_extracted_inventory, deep_to_original_map
            )

            # 调试信息：记录匹配情况
            if initial_amount > 0 or initial_quantity > 0:
                matched_count += 1
                if matched_count <= 5:  # 只显示前5个匹配成功的
                    print(f"   ✅ 匹配成功({matched_via}): 深加工产物编码 {normalized_product_code} -> 金额={initial_amount:.2f}, 数量={initial_quantity:.6f}")
            else:
                unmatched_codes.append(normalized_product_code)
                if len(unmatched_codes) <= 5:  # 只显示前5个未匹配的
                    print(f"   ❌ 未匹配: 深加工产物编码 {normalized_product_code} (在期初库存和回退库存中均未找到)")
        
        # 输出匹配统计
        print(f"📊 匹配结果: 成功匹配 {matched_count} 个，未匹配 {len(unmatched_codes)} 个")
        if len(unmatched_codes) > 5:
            print(f"   ... 还有 {len(unmatched_codes) - 5} 个未匹配的编码")
        
        # 显示期初库存中的物料代码（用于对比）
        if len(initial_inventory) > 0:
            print(f"📋 期初库存中的物料代码总数: {len(initial_inventory)}")
            print(f"📋 期初库存中的物料代码示例（前20个）: {list(initial_inventory.keys())[:20]}")
            if len(unmatched_codes) > 0:
                print(f"📋 未匹配的深加工产物编码示例（前20个）: {unmatched_codes[:20]}")
                
                # 尝试找出可能的匹配（检查是否有部分匹配或格式问题，以及回退映射）
                print("🔍 尝试查找可能的匹配关系...")
                # 先检查回退映射
                codes_with_fallback = []
                for unmatched_code in unmatched_codes[:20]:
                    if unmatched_code in deep_to_original_map:
                        orig_codes = list(deep_to_original_map[unmatched_code])
                        codes_with_fallback.append((unmatched_code, orig_codes))

                if codes_with_fallback:
                    print(f"🔗 以下 {len(codes_with_fallback)} 个未匹配编码存在回退映射到拆解产物编码:")
                    for deep_code, orig_codes in codes_with_fallback[:5]:
                        # 检查回退编码在库存中是否存在
                        found_orig = []
                        for oc in orig_codes:
                            if oc in full_extracted_inventory:
                                info = full_extracted_inventory[oc]
                                found_orig.append(f"{oc}(金额={info['金额']:.2f},数量={info['数量']:.6f})")
                            else:
                                found_orig.append(f"{oc}(库存中不存在)")
                        print(f"   深加工产物编码 {deep_code} → 拆解产物编码: {', '.join(found_orig)}")

                # 再检查传统匹配（部分匹配等）
                for unmatched_code in unmatched_codes[:10]:  # 只检查前10个
                    # 检查是否有相似的编码（移除前导0、检查大小写等）
                    possible_matches = []
                    for inv_code in list(initial_inventory.keys())[:50]:  # 只检查前50个
                        # 检查是否完全相同（忽略大小写）
                        if unmatched_code.upper() == inv_code.upper():
                            possible_matches.append(f"{inv_code} (大小写不同)")
                        # 检查是否一个包含另一个
                        elif unmatched_code in inv_code or inv_code in unmatched_code:
                            possible_matches.append(f"{inv_code} (部分匹配)")
                    
                    if possible_matches:
                        print(f"   {unmatched_code} 可能的匹配: {possible_matches[:3]}")
        else:
            print("⚠️ 警告：期初库存为空，无法进行匹配")
        
        # 重新遍历以处理数据
        for product_code in all_product_codes:
            # 确保格式统一后再匹配
            normalized_product_code = normalize_material_code(product_code)
            # 期初库存数据（同一编码的所有四机一脑类别共享期初库存）
            # 使用回退匹配：优先直接匹配深加工产物编码，未找到时回退到拆解产物编码
            initial_amount, initial_quantity, _ = get_initial_inventory_with_fallback(
                normalized_product_code, initial_inventory, full_extracted_inventory, deep_to_original_map
            )
            
            # 获取该编码的所有四机一脑类别数据（使用统一格式的编码）
            category_data_list = product_code_groups.get(normalized_product_code, [])
            
            # 如果同一深加工产物编码存在多个四机一脑类别，需要分成多行
            if len(category_data_list) > 1:
                # 多个四机一脑类别情况：为每个类别创建一行，但期初库存需要合并计算
                # 先汇总所有类别的数据用于计算期末单位成本
                total_material_value = 0.0
                total_production_weight = 0.0
                total_production_cost = 0.0
                product_name = ''
                
                for four_category, data in category_data_list:
                    material_value = data.get('产值', 0.0)
                    production_weight = data.get('重量', 0.0)
                    name = data.get('名称', '').strip()
                    mapping_category = data.get('类别', '')
                    
                    total_material_value += material_value
                    total_production_weight += production_weight
                    
                    if not product_name and name:
                        product_name = name
                    
                    # 计算每个类别的本期生产成本
                    category_total = category_total_value.get(mapping_category, 0.0)
                    category_cost = category_allocation_cost.get(mapping_category, 0.0)
                    
                    if category_total > 0 and material_value > 0:
                        production_cost = (material_value / category_total) * category_cost
                        total_production_cost += production_cost
                
                # 如果名称为空，尝试从其他来源匹配（使用统一格式的编码）
                if not product_name and normalized_product_code in extracted_code_to_name:
                    product_name = extracted_code_to_name[normalized_product_code]
                
                # 计算期末单位成本（汇总所有类别）
                total_amount = initial_amount + total_production_cost
                total_qty = initial_quantity + total_production_weight
                
                unit_cost = 0.0
                if total_qty > 0:
                    unit_cost = total_amount / total_qty
                
                # 为每个四机一脑类别创建一条记录
                for four_category, data in category_data_list:
                    mapping_category = data.get('类别', '')
                    category_material_value = data.get('产值', 0.0)
                    category_production_weight = data.get('重量', 0.0)
                    category_total = category_total_value.get(mapping_category, 0.0)
                    category_cost = category_allocation_cost.get(mapping_category, 0.0)
                    
                    # 计算该类别的本期生产成本
                    category_production_cost = 0.0
                    if category_total > 0 and category_material_value > 0:
                        category_production_cost = (category_material_value / category_total) * category_cost
                    
                    # 期初库存只在第一个类别显示，其他类别显示0
                    is_first_category = (four_category == category_data_list[0][0])
                    
                    result_data.append({
                        '深加工产物编码': normalized_product_code,
                        '深加工产物名称': product_name,
                        '四机一脑类别': four_category,
                        '类别': mapping_category,
                        '期初库存数量': round(initial_quantity, 6) if is_first_category else 0.0,
                        '期初库存金额': round(initial_amount, 2) if is_first_category else 0.0,
                        '当期生产重量': round(category_production_weight, 6),
                        '物料产值': round(category_material_value, 2),
                        '分类产值总数': round(category_total, 2),
                        '分类深加工产物分摊成本': round(category_cost, 2),
                        '本期生产成本': round(category_production_cost, 2),
                        '期末单位成本': round(unit_cost, 6)  # 使用汇总后的单位成本
                    })
            else:
                # 单个四机一脑类别情况：正常处理
                if not category_data_list:
                    # 如果没有产值数据，只有期初库存，根据深加工产物名称匹配四机一脑类别
                    product_name = ''
                    if normalized_product_code in extracted_code_to_name:
                        product_name = extracted_code_to_name[normalized_product_code]
                    
                    # 根据深加工产物名称匹配四机一脑类别
                    four_category = None
                    if product_name:
                        four_category = map_to_four_category_by_name(product_name)
                    
                    # 如果仍然匹配不到，默认归类为"电视"
                    if not four_category:
                        four_category = '电视'
                    
                    # 从内置映射表查找类别（只接受四个有效成本类别）
                    VALID_COST_CATEGORIES = {'打包铁', '一破', '屏', '印制板'}
                    mapping_category = r3_to_category.get(normalized_product_code, '')
                    if mapping_category not in VALID_COST_CATEGORIES:
                        mapping_category = ''
                    # 回退：如果直接匹配不到，尝试通过拆解产物编码查找
                    if not mapping_category and normalized_product_code in deep_to_original_map:
                        for orig_code in deep_to_original_map[normalized_product_code]:
                            orig_cat = r3_to_category.get(orig_code, '')
                            if orig_cat in VALID_COST_CATEGORIES:
                                mapping_category = orig_cat
                                break

                    # 只有期初库存，没有产值数据
                    material_value = 0.0
                    production_weight = 0.0
                    production_cost = 0.0
                    # 从汇总数据中获取分类产值总数和分摊成本（而非硬编码为0）
                    category_total = category_total_value.get(mapping_category, 0.0) if mapping_category else 0.0
                    category_cost = category_allocation_cost.get(mapping_category, 0.0) if mapping_category else 0.0
                    
                    # 计算期末单位成本（只有期初库存）
                    total_amount = initial_amount
                    total_qty = initial_quantity
                    unit_cost = 0.0
                    if total_qty > 0:
                        unit_cost = total_amount / total_qty
                    
                    result_data.append({
                        '深加工产物编码': product_code,
                        '深加工产物名称': product_name,
                        '四机一脑类别': four_category,
                        '类别': mapping_category,
                        '期初库存数量': round(initial_quantity, 6),
                        '期初库存金额': round(initial_amount, 2),
                        '当期生产重量': round(production_weight, 6),
                        '物料产值': round(material_value, 2),
                        '分类产值总数': round(category_total, 2),
                        '分类深加工产物分摊成本': round(category_cost, 2),
                        '本期生产成本': round(production_cost, 2),
                        '期末单位成本': round(unit_cost, 6)
                    })
                else:
                    # 有产值数据的情况
                    four_category = category_data_list[0][0]
                    output_info = category_data_list[0][1]
                    material_value = output_info.get('产值', 0.0)
                    production_weight = output_info.get('重量', 0.0)
                    product_name = output_info.get('名称', '').strip()
                    mapping_category = output_info.get('类别', '')
                    
                    # 如果名称为空，按优先级尝试匹配（使用统一格式的编码）
                    if not product_name and normalized_product_code in extracted_code_to_name:
                        product_name = extracted_code_to_name[normalized_product_code]
                    
                    # 如果四机一脑类别为空或无效，尝试根据深加工产物名称匹配
                    if not four_category or four_category == '':
                        if product_name:
                            four_category = map_to_four_category_by_name(product_name)
                        # 如果仍然匹配不到，尝试从extracted_code_to_name中获取名称再匹配
                        if not four_category and normalized_product_code in extracted_code_to_name:
                            extracted_name = extracted_code_to_name[normalized_product_code]
                            four_category = map_to_four_category_by_name(extracted_name)
                    
                    # 计算本期生产成本
                    production_cost = 0.0
                    category_total = category_total_value.get(mapping_category, 0.0)
                    category_cost = category_allocation_cost.get(mapping_category, 0.0)
                    
                    if category_total > 0 and material_value > 0:
                        production_cost = (material_value / category_total) * category_cost
                    
                    # 计算期末单位成本
                    total_amount = initial_amount + production_cost
                    total_qty = initial_quantity + production_weight
                    
                    unit_cost = 0.0
                    if total_qty > 0:
                        unit_cost = total_amount / total_qty
                    
                    result_data.append({
                        '深加工产物编码': normalized_product_code,
                        '深加工产物名称': product_name,
                        '四机一脑类别': four_category,
                        '类别': mapping_category,
                        '期初库存数量': round(initial_quantity, 6),
                        '期初库存金额': round(initial_amount, 2),
                        '当期生产重量': round(production_weight, 6),
                        '物料产值': round(material_value, 2),
                        '分类产值总数': round(category_total, 2),
                        '分类深加工产物分摊成本': round(category_cost, 2),
                        '本期生产成本': round(production_cost, 2),
                        '期末单位成本': round(unit_cost, 6)
                    })
        
        # ========== 6. 获取销售数量（从销售收益数据中统计） ==========
        # 🔧 重要：与销售收益页面使用完全相同的数据源选择逻辑，确保数据一致性
        # 优先使用手工数据（如果存在且已标记为修改），否则使用系统数据
        saleable_data_manual = app_data.get_data('saleable_data_manual')
        saleable_data = app_data.get_data('saleable_data')
        saleable_data_modified = app_data.get_data('saleable_data_modified')
        
        print(f"📊 深加工产物成本计算 - 获取销售数量数据源 - 手工数据: {saleable_data_manual is not None and not saleable_data_manual.empty if saleable_data_manual is not None else False}, 系统数据: {saleable_data is not None and not saleable_data.empty if saleable_data is not None else False}, 已修改: {saleable_data_modified}")
        
        # 使用(深加工产物编码, 四机一脑类别)作为key
        sales_quantity_by_code_category = {}  # {(深加工产物编码, 四机一脑类别): 销售数量(KG)}
        
        # 四机一脑分类映射函数
        def map_to_category_for_sales(material_name):
            """根据原物料名称映射到四机一脑分类"""
            if not material_name or pd.isna(material_name):
                return None
            name = str(material_name)
            
            # 优先匹配"显示器"（电脑类）
            if '显示器' in name:
                return '电脑'
            if ('电脑' in name or '笔记本' in name or 
                '主机' in name or '废旧金属黑色金属-铁及其合金-电子枪' in name):
                return '电脑'
            
            # 电视映射规则
            if ('CRT其它机壳破碎塑料' in name or '线路板边框破碎塑料' in name or 
                '废旧玻璃电子枪' in name or '废旧金属荫罩压块铁' in name or 
                '黑白' in name):
                return '电视'
            if ('电视' in name or '彩电' in name or '等离子' in name):
                return '电视'
            
            # 冰箱
            if '冰箱' in name or '冰柜' in name:
                return '冰箱'
            
            # 空调
            if '空调' in name:
                return '空调'
            
            # 洗衣机
            if '洗衣机' in name or '双缸' in name:
                return '洗衣机'
            
            return None
        
        # 确定使用的数据源（与销售收益页面逻辑完全一致）
        revenue_data = None
        data_source = None
        
        # 检查是否有手工数据且包含收益列，并且已标记为修改（与销售收益页面逻辑一致）
        if saleable_data_manual is not None and not saleable_data_manual.empty and saleable_data_modified:
            if '计算结果(KG)' in saleable_data_manual.columns:
                revenue_data = saleable_data_manual
                data_source = '手工'
                print(f"  ✅ 使用手工数据（已修改），记录数: {len(revenue_data)}")
            else:
                print(f"  ⚠️ 手工数据缺少'计算结果(KG)'列")
        
        # 如果没有手工数据或未修改，使用系统数据（与销售收益页面逻辑一致）
        if revenue_data is None and saleable_data is not None and not saleable_data.empty:
            if '计算结果(KG)' in saleable_data.columns:
                revenue_data = saleable_data
                data_source = '系统'
                print(f"  ✅ 使用系统数据，记录数: {len(revenue_data)}")
            else:
                print(f"  ⚠️ 系统数据缺少'计算结果(KG)'列")
        
        if revenue_data is not None and not revenue_data.empty:
            if '拆解产物编码' in revenue_data.columns and '计算结果(KG)' in revenue_data.columns:
                # 确保计算结果(KG)是数值类型
                revenue_data['计算结果(KG)'] = pd.to_numeric(
                    revenue_data['计算结果(KG)'], errors='coerce'
                ).fillna(0)
                
                for _, row in revenue_data.iterrows():
                    product_code = normalize_material_code(row.get('拆解产物编码', ''))
                    if not product_code:
                        continue
                    
                    # 获取原物料名称
                    material_name = str(row.get('原物料名称', '')).strip()
                    
                    # 根据原物料名称映射到四机一脑类别
                    four_category = map_to_category_for_sales(material_name)
                    if not four_category:
                        continue
                    
                    calculated_weight = float(row.get('计算结果(KG)', 0) or 0)
                    
                    # 使用(深加工产物编码, 四机一脑类别)作为key（使用统一格式）
                    key = (product_code, four_category)
                    if key not in sales_quantity_by_code_category:
                        sales_quantity_by_code_category[key] = 0.0
                    sales_quantity_by_code_category[key] += calculated_weight
        
        # 为每条记录添加销售数量和其他字段
        for record in result_data:
            product_code = normalize_material_code(record.get('深加工产物编码', ''))
            four_category = str(record.get('四机一脑类别', '')).strip()
            
            # 获取销售数量（使用统一格式）
            key = (product_code, four_category)
            sales_quantity = sales_quantity_by_code_category.get(key, 0.0)
            record['销售数量'] = round(sales_quantity, 6)
            
            # 计算深加工产物销售成本
            unit_cost = record.get('期末单位成本', 0) or 0
            sales_cost = sales_quantity * unit_cost
            record['深加工产物销售成本'] = round(sales_cost, 2)
            
            # 付费处置（数量）和付费处置（成本）默认为0
            record['付费处置（数量）'] = 0.0
            record['付费处置（成本）'] = 0.0
            
            # 计算期末库存数量
            initial_quantity = record.get('期初库存数量', 0) or 0
            production_weight = record.get('当期生产重量', 0) or 0
            paid_disposal_qty = record.get('付费处置（数量）', 0) or 0
            
            ending_inventory_quantity = initial_quantity + production_weight - sales_quantity - paid_disposal_qty
            record['期末库存数量'] = round(ending_inventory_quantity, 6)
            
            # 计算期末库存成本
            ending_inventory_cost = ending_inventory_quantity * unit_cost
            record['期末库存成本'] = round(ending_inventory_cost, 2)
        
        # 按深加工产物编码和四机一脑类别排序
        result_data.sort(key=lambda x: (x['深加工产物编码'], x['四机一脑类别']))
        
        return result_data
        
    except Exception as e:
        print(f"计算深加工产物成本失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return []


@cost_forecast_bp.route('/deep-processing-product-cost', methods=['GET'])
def get_deep_processing_product_cost():
    """获取深加工产物成本计算数据"""
    try:
        app_data = get_session_data_manager()
        prediction_period = int(request.args.get('prediction_period', 1))
        force_refresh = request.args.get('force_refresh', 'false').lower() == 'true'
        
        # 检查数据是否已被清除
        data_cleared = app_data.get_data('__data_cleared__')
        if data_cleared:
            # 数据已被清除，直接返回空数据
            return jsonify({
                'success': True,
                'data': []
            })
        
        # 检查缓存
        cache_key = f'deep_processing_product_cost_result_v1_{prediction_period}'
        if not force_refresh:
            cached_result = app_data.get_data(cache_key)
            if cached_result is not None and len(cached_result) > 0:
                return jsonify({
                    'success': True,
                    'data': cached_result,
                    'from_cache': True
                })
        
        # 计算数据
        result_data = calculate_deep_processing_product_cost(app_data, prediction_period)
        
        # 缓存结果
        app_data.set_data(cache_key, result_data)
        
        return jsonify({
            'success': True,
            'data': result_data
        })
        
    except Exception as e:
        print(f"获取深加工产物成本数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@cost_forecast_bp.route('/deep-processing-product-cost/export', methods=['GET'])
def export_deep_processing_product_cost():
    """导出深加工产物成本计算数据到Excel"""
    try:
        app_data = get_session_data_manager()
        prediction_period = int(request.args.get('prediction_period', 1))
        
        # 计算数据
        result_data = calculate_deep_processing_product_cost(app_data, prediction_period)
        
        if not result_data or len(result_data) == 0:
            return jsonify({
                'success': False,
                'error': '没有可导出的深加工产物成本数据'
            }), 400
        
        # 转换为DataFrame
        export_df = pd.DataFrame(result_data)
        
        # 追加合计行（期末单位成本显示 "-"，其余数值列求和）
        _text_cols_dp = {'深加工产物编码', '深加工产物名称', '四机一脑类别', '类别', '期末单位成本'}
        _decimals_6_cols_dp = {
            '期初库存数量', '当期生产重量', '销售数量',
            '付费处置（数量）', '期末库存数量'
        }
        _total_row_dp = {c: '' for c in export_df.columns}
        if '深加工产物编码' in export_df.columns:
            _total_row_dp['深加工产物编码'] = '合计'
        if '期末单位成本' in export_df.columns:
            _total_row_dp['期末单位成本'] = '-'
        for _c in export_df.columns:
            if _c in _text_cols_dp:
                continue
            _sum_val = pd.to_numeric(export_df[_c], errors='coerce').fillna(0).sum()
            _total_row_dp[_c] = round(float(_sum_val), 6 if _c in _decimals_6_cols_dp else 2)
        export_df = pd.concat([export_df, pd.DataFrame([_total_row_dp])], ignore_index=True)
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 写入数据
            export_df.to_excel(writer, sheet_name='深加工产物成本计算', index=False)
            
            # 设置列宽和样式
            worksheet = writer.sheets['深加工产物成本计算']
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF", name="仿宋")
            header_fill = PatternFill(start_color="8b5cf6", end_color="8b5cf6", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # 列宽设置
            column_widths = {
                '深加工产物编码': 18,
                '深加工产物名称': 30,
                '四机一脑类别': 15,
                '类别': 12,
                '期初库存数量': 18,
                '期初库存金额': 18,
                '当期生产重量': 18,
                '物料产值': 18,
                '分类产值总数': 18,
                '分类深加工产物分摊成本': 22,
                '本期生产成本': 18,
                '期末单位成本': 18,
                '销售数量': 18,
                '深加工产物销售成本': 22,
                '付费处置（数量）': 18,
                '付费处置（成本）': 18,
                '期末库存数量': 18,
                '期末库存成本': 18
            }
            
            for col in range(1, len(export_df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                
                # 设置列宽
                col_letter = get_column_letter(col)
                col_name = export_df.columns[col - 1]
                if col_name in column_widths:
                    worksheet.column_dimensions[col_letter].width = column_widths[col_name]
                else:
                    worksheet.column_dimensions[col_letter].width = 18
            
            # 设置数据行样式
            data_font = Font(name="仿宋")
            total_row_idx = len(export_df) + 1  # 合计行在Excel中的行号（含表头）
            total_fill = PatternFill(start_color="E8E6DC", end_color="E8E6DC", fill_type="solid")
            total_font = Font(name="仿宋", bold=True)
            for row in range(2, len(export_df) + 2):
                is_total_row = (row == total_row_idx)
                for col in range(1, len(export_df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = data_font
                    cell.alignment = center_alignment
                    
                    # 期末单位成本列高亮
                    col_name = export_df.columns[col - 1]
                    if col_name == '期末单位成本':
                        cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                        cell.font = Font(name="仿宋", bold=True)
                    
                    # 合计行样式：浅灰底、加粗
                    if is_total_row:
                        cell.fill = total_fill
                        cell.font = total_font
        
        output.seek(0)
        filename = f'深加工产物成本计算_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"导出深加工产物成本计算失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@cost_forecast_bp.route('/cost-calculation/export-all', methods=['GET'])
def export_cost_calculation_all():
    """导出成本计算部分所有卡片详情页数据到一个Excel文件"""
    try:
        app_data = get_session_data_manager()
        prediction_period = int(request.args.get('prediction_period', 1))
        
        # 创建Excel文件
        output = io.BytesIO()
        errors = []  # 记录错误信息
        sheets_created = 0  # 跟踪创建的sheet数量
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # 定义统一的样式
            header_font = Font(bold=True, color="FFFFFF", name="仿宋")
            data_font = Font(name="仿宋")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # ========== Sheet 1: 生产成本分摊 ==========
            try:
                # 产品分类列表
                categories = ['电视', '电脑', '冰箱', '空调', '洗衣机']
                
                # 初始化结果数据
                result_data = []
                category_totals = {
                    'direct_material': {cat: 0.0 for cat in categories},
                    'direct_labor': {cat: 0.0 for cat in categories},
                    'manufacturing_cost': {cat: 0.0 for cat in categories},
                    'product_value': {cat: 0.0 for cat in categories},
                    'subsidy_income': {cat: 0.0 for cat in categories}
                }
                
                # 1. 收集直接材料数据
                try:
                    manual_data = app_data.get_data('extracted_data_manual')
                    if manual_data is not None and not manual_data.empty:
                        cost_data = calculate_material_cost(manual_data)
                        if '类别' in cost_data.columns:
                            old_machine_data = cost_data[cost_data['类别'] == '旧机'].copy()
                        else:
                            old_machine_data = cost_data.copy()
                        
                        if '物料描述' in old_machine_data.columns and '拆解物原料成本' in old_machine_data.columns:
                            old_machine_data['拆解物原料成本'] = pd.to_numeric(
                                old_machine_data['拆解物原料成本'], errors='coerce'
                            ).fillna(0)
                            
                            for category in categories:
                                mask = old_machine_data['物料描述'].astype(str).str.contains(category, case=False, na=False)
                                category_totals['direct_material'][category] = float(
                                    old_machine_data.loc[mask, '拆解物原料成本'].sum()
                                )
                except Exception as e:
                    print(f"收集直接材料数据失败: {str(e)}")
                
                # 2. 收集直接人工数据
                try:
                    direct_labor_result = calculate_direct_labor_cost(app_data, prediction_period)
                    product_category_stats = direct_labor_result.get('product_category_stats', {})
                    for cat in categories:
                        stats = product_category_stats.get(cat, {'wage': 0.0, 'fixed_cost': 0.0})
                        category_totals['direct_labor'][cat] = stats.get('wage', 0.0) + stats.get('fixed_cost', 0.0)
                except Exception as e:
                    print(f"收集直接人工数据失败: {str(e)}")
                
                # 3. 收集制造费用数据
                try:
                    mfg_by_category = collect_production_manufacturing_cost_by_category(
                        app_data, prediction_period
                    )
                    for cat in categories:
                        category_totals['manufacturing_cost'][cat] = mfg_by_category.get(cat, 0.0)
                except Exception as e:
                    print(f"收集制造费用数据失败: {str(e)}")
                
                # 4. 收集拆解产物价值数据
                try:
                    disassembly_data = app_data.get_data('disassembly_data')
                    if disassembly_data is not None and not disassembly_data.empty:
                        if '类别' in disassembly_data.columns:
                            product_data = disassembly_data[disassembly_data['类别'] == '拆解产物'].copy()
                            if not product_data.empty:
                                from data.base_data.price_data import load_price_data
                                price_df = load_price_data()
                                if price_df is not None and not price_df.empty:
                                    price_mapping = {}
                                    for _, row in price_df.iterrows():
                                        code = str(row['拆解产物编码']).strip()
                                        price_no_tax = row.get('销售单价-不含税(元/KG)', 0)
                                        if pd.notna(price_no_tax):
                                            price_mapping[code] = float(price_no_tax)
                                    
                                    category_keyword_mapping = {
                                        '电视': ['电视', '彩电', 'CRT其它机壳破碎塑料', '线路板边框破碎塑料', '等离子', '废旧玻璃电子枪', '废旧金属荫罩压块铁', '黑白'],
                                        '电脑': ['电脑', '显示器', '笔记本', '主机', '废旧金属黑色金属-铁及其合金-电子枪'],
                                        '冰箱': ['冰箱', '冰柜'],
                                        '空调': ['空调'],
                                        '洗衣机': ['洗衣机', '双缸']
                                    }
                                    
                                    for idx, row in product_data.iterrows():
                                        material_name = str(row.get('原物料名称', '')).strip()
                                        category = None
                                        for cat, keywords in category_keyword_mapping.items():
                                            for keyword in keywords:
                                                if keyword in material_name:
                                                    category = cat
                                                    break
                                            if category:
                                                break
                                        
                                        if not category or category not in categories:
                                            continue
                                        
                                        product_code = str(row.get('拆解产物编码', '')).strip()
                                        calculated_weight = row.get('计算结果(KG)', 0)
                                        try:
                                            calculated_weight = float(calculated_weight) if pd.notna(calculated_weight) else 0
                                        except (ValueError, TypeError):
                                            calculated_weight = 0
                                        
                                        price_no_tax = price_mapping.get(product_code, 0)
                                        if price_no_tax < 0:
                                            material_value = 0
                                        else:
                                            material_value = calculated_weight * price_no_tax
                                        
                                        category_totals['product_value'][category] += material_value
                except Exception as e:
                    print(f"收集拆解产物价值数据失败: {str(e)}")
                
                # 5. 收集基金补贴收入数据
                try:
                    subsidy_income_data = app_data.get_data('subsidy_income_data')
                    if subsidy_income_data is not None and not subsidy_income_data.empty:
                        if '补贴大类' in subsidy_income_data.columns and '基金补贴收入(元)' in subsidy_income_data.columns:
                            subsidy_income_data['基金补贴收入(元)'] = pd.to_numeric(
                                subsidy_income_data['基金补贴收入(元)'], errors='coerce'
                            ).fillna(0)
                            
                            # 将补贴大类映射到产品类型（五大类：电视机、冰箱、洗衣机、空调、电脑）
                            def map_subsidy_category_to_product_type(subsidy_category):
                                """将补贴大类映射到产品类型"""
                                subsidy_category = str(subsidy_category).strip()
                                # 空调的细分补贴大类
                                if subsidy_category in ['整机', '内机', '外机']:
                                    return '空调'
                                # 电脑的细分补贴大类
                                elif subsidy_category in ['笔记本', '显示器', '主机']:
                                    return '电脑'
                                # 其他直接返回（电视机、冰箱、洗衣机）
                                elif subsidy_category in ['电视机', '冰箱', '洗衣机']:
                                    return subsidy_category
                                else:
                                    # 默认返回原值（兼容旧数据）
                                    return subsidy_category
                            
                            # 添加产品类型列
                            subsidy_income_data['产品类型'] = subsidy_income_data['补贴大类'].apply(map_subsidy_category_to_product_type)
                            
                            # 按产品类型分组汇总
                            product_type_groups = subsidy_income_data.groupby('产品类型')['基金补贴收入(元)'].sum()
                            
                            # 映射产品类型到四机一脑分类
                            for product_type, total_subsidy in product_type_groups.items():
                                product_type = str(product_type).strip()
                                # 映射：电视机 -> 电视
                                if product_type == '电视机':
                                    product_type = '电视'
                                
                                if product_type in categories:
                                    category_totals['subsidy_income'][product_type] += float(total_subsidy)
                except Exception as e:
                    print(f"收集基金补贴收入数据失败: {str(e)}")
                
                # 计算各项指标
                total_row = {
                    'direct_material': 0.0,
                    'direct_labor': 0.0,
                    'manufacturing_cost': 0.0,
                    'product_value': 0.0,
                    'subsidy_income': 0.0,
                    'production_cost_subtotal': 0.0,
                    'subsidy_allocation_cost': 0.0,
                    'product_allocation_cost': 0.0
                }
                
                for category in categories:
                    direct_material = category_totals['direct_material'][category]
                    direct_labor = category_totals['direct_labor'][category]
                    manufacturing_cost = category_totals['manufacturing_cost'][category]
                    product_value = category_totals['product_value'][category]
                    subsidy_income = category_totals['subsidy_income'][category]
                    
                    production_cost_subtotal = direct_material + direct_labor + manufacturing_cost
                    
                    total_revenue = subsidy_income + product_value
                    if total_revenue > 0:
                        subsidy_ratio = subsidy_income / total_revenue
                        product_ratio = product_value / total_revenue
                    else:
                        subsidy_ratio = 0.0
                        product_ratio = 0.0
                    
                    subsidy_allocation_cost = production_cost_subtotal * subsidy_ratio
                    product_allocation_cost = production_cost_subtotal * product_ratio
                    
                    row_data = {
                        '产线': category,
                        '基金补贴收入': round(subsidy_income, 2),
                        '拆解产物价值': round(product_value, 2),
                        '基金补贴成本分摊比例': round(subsidy_ratio, 4),
                        '拆解产物成本分摊比例': round(product_ratio, 4),
                        '直接材料': round(direct_material, 2),
                        '直接人工': round(direct_labor, 2),
                        '制造费用': round(manufacturing_cost, 2),
                        '生产成本小计': round(production_cost_subtotal, 2),
                        '基金补贴收入分摊成本': round(subsidy_allocation_cost, 2),
                        '拆解产物分摊成本': round(product_allocation_cost, 2)
                    }
                    result_data.append(row_data)
                    
                    total_row['direct_material'] += direct_material
                    total_row['direct_labor'] += direct_labor
                    total_row['manufacturing_cost'] += manufacturing_cost
                    total_row['product_value'] += product_value
                    total_row['subsidy_income'] += subsidy_income
                    total_row['production_cost_subtotal'] += production_cost_subtotal
                    total_row['subsidy_allocation_cost'] += subsidy_allocation_cost
                    total_row['product_allocation_cost'] += product_allocation_cost
                
                total_revenue = total_row['subsidy_income'] + total_row['product_value']
                if total_revenue > 0:
                    total_subsidy_ratio = total_row['subsidy_income'] / total_revenue
                    total_product_ratio = total_row['product_value'] / total_revenue
                else:
                    total_subsidy_ratio = 0.0
                    total_product_ratio = 0.0
                
                total_row_data = {
                    '产线': '合计',
                    '基金补贴收入': round(total_row['subsidy_income'], 2),
                    '拆解产物价值': round(total_row['product_value'], 2),
                    '基金补贴成本分摊比例': round(total_subsidy_ratio, 4),
                    '拆解产物成本分摊比例': round(total_product_ratio, 4),
                    '直接材料': round(total_row['direct_material'], 2),
                    '直接人工': round(total_row['direct_labor'], 2),
                    '制造费用': round(total_row['manufacturing_cost'], 2),
                    '生产成本小计': round(total_row['production_cost_subtotal'], 2),
                    '基金补贴收入分摊成本': round(total_row['subsidy_allocation_cost'], 2),
                    '拆解产物分摊成本': round(total_row['product_allocation_cost'], 2)
                }
                result_data.append(total_row_data)
                
                if result_data:
                    df = pd.DataFrame(result_data)
                    df.to_excel(writer, sheet_name='生产成本分摊', index=False)
                    
                    # 设置样式
                    ws = writer.sheets['生产成本分摊']
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    
                    # 设置列宽
                    column_widths = {
                        'A': 12, 'B': 18, 'C': 18, 'D': 22, 'E': 22,
                        'F': 15, 'G': 15, 'H': 15, 'I': 18, 'J': 22, 'K': 22
                    }
                    for col_letter, width in column_widths.items():
                        ws.column_dimensions[col_letter].width = width
                    
                    # 设置表头样式
                    for col in range(1, len(df.columns) + 1):
                        cell = ws.cell(row=1, column=col)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                    
                    # 设置数据行样式
                    for row in range(2, len(df) + 2):
                        is_total = ws.cell(row=row, column=1).value == '合计'
                        for col in range(1, len(df.columns) + 1):
                            cell = ws.cell(row=row, column=col)
                            cell.font = data_font
                            cell.alignment = center_alignment
                            
                            if is_total:
                                cell.font = Font(bold=True, name="仿宋")
                                if col == 1:
                                    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                                elif col in [6, 7, 8, 9, 10, 11]:
                                    cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                            elif col in [4, 5]:
                                cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
                    
                    sheets_created += 1
            except Exception as e:
                errors.append(f"生产成本分摊: {str(e)}")
            
            # ========== Sheet 2: 一次拆解产物成本计算 ==========
            try:
                result_data = calculate_disassembly_product_cost(app_data, prediction_period)
                
                if result_data and len(result_data) > 0:
                    export_df = pd.DataFrame(result_data)
                    export_df.to_excel(writer, sheet_name='一次拆解产物成本计算', index=False)
                    
                    # 设置列宽和样式
                    worksheet = writer.sheets['一次拆解产物成本计算']
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    
                    # 列宽设置
                    column_widths = {
                        '拆解产物编码': 18,
                        '拆解产物名称': 30,
                        '分类': 12,
                        '期初库存金额': 18,
                        '期初库存数量': 18,
                        '物料产值': 18,
                        '分类产值总数': 18,
                        '分类拆解产物分摊成本': 22,
                        '本期生产成本': 18,
                        '当期生产重量': 18,
                        '期末单位成本': 18,
                        '销售数量': 18,
                        '一次拆解产物销售成本': 22,
                        '付费处置（数量）': 18,
                        '内转屏处置（数量）': 18,
                        '内转印制板处置（数量）': 20,
                        '内转荧光灯处置（数量）': 20,
                        '深加工-打包铁（数量）': 20,
                        '深加工-塑料一破（数量）': 20,
                        '付费处置（成本）': 18,
                        '内转屏处置（成本）': 18,
                        '内转印制板处置（成本）': 20,
                        '内转荧光灯处置（成本）': 20,
                        '深加工-打包铁（成本）': 20,
                        '深加工-塑料一破（成本）': 20,
                        '期末库存数量': 18,
                        '期末库存成本': 18
                    }
                    
                    for col in range(1, len(export_df.columns) + 1):
                        cell = worksheet.cell(row=1, column=col)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                        
                        # 设置列宽
                        col_letter = get_column_letter(col)
                        col_name = export_df.columns[col - 1]
                        if col_name in column_widths:
                            worksheet.column_dimensions[col_letter].width = column_widths[col_name]
                        else:
                            worksheet.column_dimensions[col_letter].width = 18
                    
                    # 设置数据行样式
                    for row in range(2, len(export_df) + 2):
                        for col in range(1, len(export_df.columns) + 1):
                            cell = worksheet.cell(row=row, column=col)
                            cell.font = data_font
                            cell.alignment = center_alignment
                            
                            # 期末单位成本列高亮
                            col_name = export_df.columns[col - 1]
                            if col_name == '期末单位成本':
                                cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                                cell.font = Font(name="仿宋", bold=True)
                    
                    sheets_created += 1
            except Exception as e:
                errors.append(f"一次拆解产物成本计算: {str(e)}")
            
            # ========== Sheet 3: 深加工产物成本计算 ==========
            try:
                result_data = calculate_deep_processing_product_cost(app_data, prediction_period)
                
                if result_data and len(result_data) > 0:
                    export_df = pd.DataFrame(result_data)
                    export_df.to_excel(writer, sheet_name='深加工产物成本计算', index=False)
                    
                    # 设置列宽和样式
                    worksheet = writer.sheets['深加工产物成本计算']
                    header_fill = PatternFill(start_color="8b5cf6", end_color="8b5cf6", fill_type="solid")
                    
                    # 列宽设置
                    column_widths = {
                        '深加工产物编码': 18,
                        '深加工产物名称': 30,
                        '四机一脑类别': 15,
                        '类别': 12,
                        '期初库存数量': 18,
                        '期初库存金额': 18,
                        '当期生产重量': 18,
                        '物料产值': 18,
                        '分类产值总数': 18,
                        '分类深加工产物分摊成本': 22,
                        '本期生产成本': 18,
                        '期末单位成本': 18,
                        '销售数量': 18,
                        '深加工产物销售成本': 22,
                        '付费处置（数量）': 18,
                        '付费处置（成本）': 18,
                        '期末库存数量': 18,
                        '期末库存成本': 18
                    }
                    
                    for col in range(1, len(export_df.columns) + 1):
                        cell = worksheet.cell(row=1, column=col)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                        
                        # 设置列宽
                        col_letter = get_column_letter(col)
                        col_name = export_df.columns[col - 1]
                        if col_name in column_widths:
                            worksheet.column_dimensions[col_letter].width = column_widths[col_name]
                        else:
                            worksheet.column_dimensions[col_letter].width = 18
                    
                    # 设置数据行样式
                    for row in range(2, len(export_df) + 2):
                        for col in range(1, len(export_df.columns) + 1):
                            cell = worksheet.cell(row=row, column=col)
                            cell.font = data_font
                            cell.alignment = center_alignment
                            
                            # 期末单位成本列高亮
                            col_name = export_df.columns[col - 1]
                            if col_name == '期末单位成本':
                                cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                                cell.font = Font(name="仿宋", bold=True)
                    
                    sheets_created += 1
            except Exception as e:
                errors.append(f"深加工产物成本计算: {str(e)}")
        
        if sheets_created == 0:
            return jsonify({
                'success': False,
                'error': '没有可导出的数据'
            }), 400
        
        output.seek(0)
        filename = f'成本计算汇总_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"导出成本计算汇总失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@cost_forecast_bp.route('/export-all', methods=['GET'])
def export_all_cost_forecast():
    """导出所有成本预测详情页数据到一个Excel文件"""
    try:
        app_data = get_session_data_manager()
        prediction_period = int(request.args.get('prediction_period', 1))
        
        # 获取期间费用的分摊比例参数
        quality_manager_ratio = request.args.get('quality_manager_ratio')
        quality_group_ratio = request.args.get('quality_group_ratio')
        warehouse_group_ratio = request.args.get('warehouse_group_ratio')
        
        if quality_manager_ratio is not None:
            quality_manager_ratio = float(quality_manager_ratio)
        if quality_group_ratio is not None:
            quality_group_ratio = float(quality_group_ratio)
        if warehouse_group_ratio is not None:
            warehouse_group_ratio = float(warehouse_group_ratio)
        
        # 创建Excel文件
        output = io.BytesIO()
        errors = []  # 记录错误信息
        sheets_created = 0  # 跟踪创建的sheet数量
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # 定义统一的样式
            header_font = Font(bold=True, color="FFFFFF", name="仿宋")
            data_font = Font(name="仿宋")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            left_alignment = Alignment(horizontal="left", vertical="center")
            right_alignment = Alignment(horizontal="right", vertical="center")
            
            # 1. 拆解物原料成本
            try:
                manual_data = app_data.get_data('extracted_data_manual')
                if manual_data is not None and not manual_data.empty:
                    cost_data = calculate_material_cost(manual_data)
                    if '类别' in cost_data.columns:
                        export_data = cost_data[cost_data['类别'] == '旧机'].copy()
                    else:
                        export_data = cost_data
                    
                    if not export_data.empty:
                        export_columns = [
                            '序号', '物料代码', '物料描述', '初始数据', '本期计划采购数量',
                            '计划采购单价', '非限制使用的库存', '单位投料成本', '拆解物原料成本'
                        ]
                        available_columns = [col for col in export_columns if col in export_data.columns]
                        export_df = export_data[available_columns].copy()
                        if '非限制使用的库存' in export_df.columns:
                            export_df = export_df.rename(columns={'非限制使用的库存': '本期实际投产数量'})
                        
                        export_df.to_excel(writer, sheet_name='拆解物原料成本', index=False)
                        worksheet = writer.sheets['拆解物原料成本']
                        for col in range(1, len(export_df.columns) + 1):
                            cell = worksheet.cell(row=1, column=col)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_alignment
                            worksheet.column_dimensions[get_column_letter(col)].width = 20
                        sheets_created += 1
            except Exception as e:
                errors.append(f"拆解物原料成本: {str(e)}")
            
            # 2. 直接人工成本
            try:
                result = calculate_direct_labor_cost(app_data, prediction_period)
                if result and (result.get('total_wage', 0) != 0 or result.get('total_fixed_cost', 0) != 0):
                    # 2.1 汇总统计表
                    summary_data = {
                        '统计项目': [
                            '计件总工资', '旧机类别工资', '一次拆解产物工资', '深加工工资（一破/打包铁/屏）',
                            '分摊固定成本（白电）', '分摊固定成本（黑电）', '分摊固定成本（冰箱）',
                            '分摊固定成本（金属打包）', '分摊固定成本（塑料）', '分摊固定成本（屏）',
                            '总分摊固定成本', '直接人工成本'
                        ],
                        '金额(元)': [
                            result.get('total_wage', 0),
                            result.get('part1_wage', 0),
                            result.get('part2_wage', 0),
                            result.get('part3_wage', 0),
                            result.get('category_details', {}).get('白电', {}).get('total_fixed_cost', 0),
                            result.get('category_details', {}).get('黑电', {}).get('total_fixed_cost', 0),
                            result.get('category_details', {}).get('冰箱', {}).get('total_fixed_cost', 0),
                            result.get('category_details', {}).get('金属打包', {}).get('total_fixed_cost', 0),
                            result.get('category_details', {}).get('塑料', {}).get('total_fixed_cost', 0),
                            result.get('category_details', {}).get('屏', {}).get('total_fixed_cost', 0),
                            result.get('total_fixed_cost', 0),
                            result.get('direct_labor_cost', 0)
                        ]
                    }
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name='直接人工-汇总统计', index=False)
                    worksheet = writer.sheets['直接人工-汇总统计']
                    for col in range(1, len(summary_df.columns) + 1):
                        cell = worksheet.cell(row=1, column=col)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                        worksheet.column_dimensions[get_column_letter(col)].width = 25
                    for row in range(2, len(summary_df) + 2):
                        for col in range(1, len(summary_df.columns) + 1):
                            cell = worksheet.cell(row=row, column=col)
                            cell.font = data_font
                            cell.alignment = center_alignment
                    sheets_created += 1
                    
                    # 2.2 计件工资明细表
                    all_details = []
                    if result.get('part1_details'):
                        for item in result['part1_details']:
                            all_details.append({
                                '序号': len(all_details) + 1,
                                '原物料代码': item.get('原物料代码', '') or item.get('物料代码', ''),
                                '原物料名称': item.get('原物料名称', '') or item.get('物料名称', ''),
                                '物料/产物编码': item.get('物料代码', ''),
                                '物料/产物名称': item.get('物料名称', ''),
                                '数量/重量': item.get('数量', 0),
                                '类别': '旧机',
                                '单价': item.get('单价', 0),
                                '工资': item.get('工资', 0)
                            })
                    if result.get('part2_details'):
                        for item in result['part2_details']:
                            all_details.append({
                                '序号': len(all_details) + 1,
                                '原物料代码': item.get('原物料代码', ''),
                                '原物料名称': item.get('原物料名称', ''),
                                '物料/产物编码': item.get('拆解产物编码', ''),
                                '物料/产物名称': item.get('拆解产物名称', ''),
                                '数量/重量': item.get('计算结果(KG)', 0),
                                '类别': '一次拆解产物',
                                '单价': item.get('单价', 0),
                                '工资': item.get('工资', 0)
                            })
                    if result.get('part3_details'):
                        for item in result['part3_details']:
                            all_details.append({
                                '序号': len(all_details) + 1,
                                '原物料代码': item.get('原物料代码', ''),
                                '原物料名称': item.get('原物料名称', ''),
                                '物料/产物编码': item.get('深加工产物编码', ''),
                                '物料/产物名称': item.get('深加工产物名称', ''),
                                '数量/重量': item.get('深加工结果(KG)', 0),
                                '类别': item.get('类别', ''),
                                '单价': item.get('单价', 0),
                                '工资': item.get('工资', 0)
                            })
                    
                    if all_details:
                        details_df = pd.DataFrame(all_details)
                        details_df.to_excel(writer, sheet_name='直接人工-计件工资', index=False)
                        details_ws = writer.sheets['直接人工-计件工资']
                        for col in range(1, len(details_df.columns) + 1):
                            cell = details_ws.cell(row=1, column=col)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_alignment
                            col_letter = get_column_letter(col)
                            if col == 1:  # 序号
                                details_ws.column_dimensions[col_letter].width = 10
                            elif col in [2, 4]:  # 原物料代码、物料/产物编码
                                details_ws.column_dimensions[col_letter].width = 18
                            elif col in [3, 5]:  # 原物料名称、物料/产物名称
                                details_ws.column_dimensions[col_letter].width = 30
                            else:
                                details_ws.column_dimensions[col_letter].width = 18
                        for row in range(2, len(details_df) + 2):
                            for col in range(1, len(details_df.columns) + 1):
                                cell = details_ws.cell(row=row, column=col)
                                cell.font = data_font
                                cell.alignment = center_alignment
                        sheets_created += 1
                    
                    # 2.3 分摊固定成本明细表
                    fixed_cost_details = []
                    category_details = result.get('category_details', {})
                    category_order = ['白电', '黑电', '冰箱', '金属打包', '塑料', '屏']
                    index = 1
                    for category in category_order:
                        category_data = category_details.get(category)
                        if not category_data or not category_data.get('item_allocations'):
                            continue
                        category_wage_sum = 0
                        category_fixed_cost_sum = 0
                        for allocation in category_data['item_allocations']:
                            item = allocation.get('item', {})
                            
                            # 提取物料信息
                            origin_code = ''
                            origin_name = ''
                            code = ''
                            name = ''
                            wage = 0
                            
                            if item.get('物料代码'):
                                code = item.get('物料代码', '')
                                name = item.get('物料名称', '')
                                origin_code = item.get('原物料代码', '') or code
                                origin_name = item.get('原物料名称', '') or name
                                wage = item.get('工资', 0)
                            elif item.get('拆解产物编码'):
                                code = item.get('拆解产物编码', '')
                                name = item.get('拆解产物名称', '')
                                origin_code = item.get('原物料代码', '')
                                origin_name = item.get('原物料名称', '')
                                wage = item.get('工资', 0)
                            elif item.get('深加工产物编码'):
                                code = item.get('深加工产物编码', '')
                                name = item.get('深加工产物名称', '')
                                origin_code = item.get('原物料代码', '')
                                origin_name = item.get('原物料名称', '')
                                wage = item.get('工资', 0)
                            
                            category_wage_sum += wage
                            category_fixed_cost_sum += allocation.get('fixed_cost', 0)
                            fixed_cost_details.append({
                                '序号': index,
                                '原物料代码': origin_code,
                                '原物料名称': origin_name,
                                '类别': category,
                                '物料/产物编码': code,
                                '物料/产物名称': name,
                                '计件工资（元）': wage,
                                '分摊比例': f"{allocation.get('allocation_ratio', 0) * 100:.2f}%",
                                '分摊固定成本（元）': allocation.get('fixed_cost', 0)
                            })
                            index += 1
                        if category_wage_sum > 0 or category_fixed_cost_sum > 0:
                            fixed_cost_details.append({
                                '序号': '',
                                '原物料代码': '',
                                '原物料名称': '',
                                '类别': f'{category}小计',
                                '物料/产物编码': '',
                                '物料/产物名称': '',
                                '计件工资（元）': category_wage_sum,
                                '分摊比例': '',
                                '分摊固定成本（元）': category_fixed_cost_sum
                            })
                    
                    if fixed_cost_details:
                        fixed_cost_df = pd.DataFrame(fixed_cost_details)
                        fixed_cost_df.to_excel(writer, sheet_name='直接人工-分摊固定成本明细', index=False)
                        fixed_cost_ws = writer.sheets['直接人工-分摊固定成本明细']
                        for col in range(1, len(fixed_cost_df.columns) + 1):
                            cell = fixed_cost_ws.cell(row=1, column=col)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_alignment
                            col_letter = get_column_letter(col)
                            if col == 1:  # 序号
                                fixed_cost_ws.column_dimensions[col_letter].width = 10
                            elif col == 4:  # 类别
                                fixed_cost_ws.column_dimensions[col_letter].width = 12
                            elif col in [2, 5]:  # 原物料代码、物料/产物编码
                                fixed_cost_ws.column_dimensions[col_letter].width = 18
                            elif col in [3, 6]:  # 原物料名称、物料/产物名称
                                fixed_cost_ws.column_dimensions[col_letter].width = 30
                            else:
                                fixed_cost_ws.column_dimensions[col_letter].width = 18
                        for row in range(2, len(fixed_cost_df) + 2):
                            for col in range(1, len(fixed_cost_df.columns) + 1):
                                cell = fixed_cost_ws.cell(row=row, column=col)
                                cell.font = data_font
                                cell.alignment = center_alignment
                        sheets_created += 1
                    
                    # 2.4 分类统计表（如果有）
                    if result.get('product_category_stats'):
                        category_stats_data = []
                        for category, stats in result['product_category_stats'].items():
                            category_stats_data.append({
                                '类别': category,
                                '计件工资': stats.get('wage', 0),
                                '分摊固定成本': stats.get('fixed_cost', 0),
                                '直接人工成本': stats.get('wage', 0) + stats.get('fixed_cost', 0)
                            })
                        if category_stats_data:
                            category_stats_df = pd.DataFrame(category_stats_data)
                            category_stats_df.to_excel(writer, sheet_name='直接人工-分类统计', index=False)
                            category_stats_ws = writer.sheets['直接人工-分类统计']
                            for col in range(1, len(category_stats_df.columns) + 1):
                                cell = category_stats_ws.cell(row=1, column=col)
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = center_alignment
                                category_stats_ws.column_dimensions[get_column_letter(col)].width = 20
                            for row in range(2, len(category_stats_df) + 2):
                                for col in range(1, len(category_stats_df.columns) + 1):
                                    cell = category_stats_ws.cell(row=row, column=col)
                                    cell.font = data_font
                                    cell.alignment = center_alignment
                            sheets_created += 1
            except Exception as e:
                errors.append(f"直接人工成本: {str(e)}")
            
            # 3. 间接人工成本
            try:
                result = calculate_indirect_labor_cost(app_data, prediction_period, include_no_opening_columns=True)
                if result and result.get('total_cost', 0) != 0:
                    # 3.1 汇总统计表
                    category_totals = result.get('category_totals', {})
                    category_fixed_costs = result.get('category_fixed_costs', {})
                    all_details = []
                    if result.get('part1_details'):
                        for item in result['part1_details']:
                            all_details.append({
                                '序号': len(all_details) + 1,
                                '原物料代码': item.get('原物料代码', '') or item.get('物料代码', ''),
                                '原物料名称': item.get('原物料名称', '') or item.get('物料名称', ''),
                                '物料/产物编码': item.get('物料代码', ''),
                                '物料/产物名称': item.get('物料名称', ''),
                                '数量/重量': item.get('数量', 0),
                                '数量/重量(不考虑期初库存和库存结余)': item.get(
                                    '数量/重量(不考虑期初库存和库存结余)', item.get('数量', 0)
                                ),
                                '类别': '旧机',
                                '品管提成成本': item.get('品管提成成本', 0),
                                '物流主管提成成本': item.get('物流主管提成成本', 0),
                                '物流卸货提成成本': item.get('物流卸货提成成本', 0),
                                '班组长提成成本': item.get('班组长提成成本', 0),
                                '班组长提成成本(不考虑期初库存和库存结余)': item.get(
                                    '班组长提成成本(不考虑期初库存和库存结余)', item.get('班组长提成成本', 0)
                                ),
                                '生产主管提成成本': item.get('生产主管提成成本', 0),
                                '维修班长提成成本': item.get('维修班长提成成本', 0),
                                '维修员提成成本': item.get('维修员提成成本', 0),
                                '冰箱维修主管提成成本': item.get('冰箱维修主管提成成本', 0),
                                '叉车司磅库管等提成成本': item.get('叉车司磅库管等提成成本', 0),
                                '总成本': item.get('总成本', 0)
                            })
                    if result.get('part2_details'):
                        for item in result['part2_details']:
                            all_details.append({
                                '序号': len(all_details) + 1,
                                '原物料代码': item.get('原物料代码', ''),
                                '原物料名称': item.get('原物料名称', ''),
                                '物料/产物编码': item.get('拆解产物编码', ''),
                                '物料/产物名称': item.get('拆解产物名称', ''),
                                '数量/重量': item.get('计算结果(KG)', 0),
                                '数量/重量(不考虑期初库存和库存结余)': item.get(
                                    '数量/重量(不考虑期初库存和库存结余)', item.get('计算结果(KG)', 0)
                                ),
                                '类别': '一次拆解产物',
                                '品管提成成本': item.get('品管提成成本', 0),
                                '物流主管提成成本': item.get('物流主管提成成本', 0),
                                '物流卸货提成成本': item.get('物流卸货提成成本', 0),
                                '班组长提成成本': item.get('班组长提成成本', 0),
                                '班组长提成成本(不考虑期初库存和库存结余)': item.get(
                                    '班组长提成成本(不考虑期初库存和库存结余)', item.get('班组长提成成本', 0)
                                ),
                                '生产主管提成成本': item.get('生产主管提成成本', 0),
                                '维修班长提成成本': item.get('维修班长提成成本', 0),
                                '维修员提成成本': item.get('维修员提成成本', 0),
                                '冰箱维修主管提成成本': item.get('冰箱维修主管提成成本', 0),
                                '叉车司磅库管等提成成本': item.get('叉车司磅库管等提成成本', 0),
                                '总成本': item.get('总成本', 0)
                            })
                    if result.get('part3_details'):
                        for item in result['part3_details']:
                            all_details.append({
                                '序号': len(all_details) + 1,
                                '原物料代码': item.get('原物料代码', ''),
                                '原物料名称': item.get('原物料名称', ''),
                                '物料/产物编码': item.get('深加工产物编码', ''),
                                '物料/产物名称': item.get('深加工产物名称', ''),
                                '数量/重量': item.get('深加工结果(KG)', 0),
                                '数量/重量(不考虑期初库存和库存结余)': item.get(
                                    '数量/重量(不考虑期初库存和库存结余)', item.get('深加工结果(KG)', 0)
                                ),
                                '类别': item.get('类别', ''),
                                '品管提成成本': item.get('品管提成成本', 0),
                                '物流主管提成成本': item.get('物流主管提成成本', 0),
                                '物流卸货提成成本': item.get('物流卸货提成成本', 0),
                                '班组长提成成本': item.get('班组长提成成本', 0),
                                '班组长提成成本(不考虑期初库存和库存结余)': item.get(
                                    '班组长提成成本(不考虑期初库存和库存结余)', item.get('班组长提成成本', 0)
                                ),
                                '生产主管提成成本': item.get('生产主管提成成本', 0),
                                '维修班长提成成本': item.get('维修班长提成成本', 0),
                                '维修员提成成本': item.get('维修员提成成本', 0),
                                '冰箱维修主管提成成本': item.get('冰箱维修主管提成成本', 0),
                                '叉车司磅库管等提成成本': item.get('叉车司磅库管等提成成本', 0),
                                '总成本': item.get('总成本', 0)
                            })
                    
                    commission_detail_total = sum(item.get('总成本', 0) for item in all_details)
                    summary_rows = [{
                        '类别': '总体汇总',
                        '项目': '间接人工提成汇总',
                        '金额(元)': result.get('total_cost', 0)
                    }, {
                        '类别': '',
                        '项目': '人工提成成本',
                        '金额(元)': commission_detail_total
                    }, {
                        '类别': '',
                        '项目': '分摊固定成本',
                        '金额(元)': result.get('total_fixed_cost', 0)
                    }, {
                        '类别': '',
                        '项目': '其他岗位成本',
                        '金额(元)': result.get('other_positions_cost', 0)
                    }, {
                        '类别': '',
                        '项目': '间接人工总成本',
                        '金额(元)': result.get('indirect_labor_cost', 0)
                    }]
                    summary_df = pd.DataFrame(summary_rows)
                    summary_df.to_excel(writer, sheet_name='间接人工-汇总统计', index=False, startrow=1)
                    summary_ws = writer.sheets['间接人工-汇总统计']
                    summary_ws.column_dimensions['A'].width = 20
                    summary_ws.column_dimensions['B'].width = 35
                    summary_ws.column_dimensions['C'].width = 20
                    title_row = 1
                    summary_ws.merge_cells(f'A{title_row}:C{title_row}')
                    title_cell = summary_ws.cell(row=title_row, column=1)
                    title_cell.value = f'间接人工成本汇总统计（预测期数：{prediction_period}个月）'
                    title_cell.font = Font(bold=True, name="仿宋", size=16, color="FFFFFF")
                    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    title_cell.alignment = Alignment(horizontal="center", vertical="center")
                    summary_ws.row_dimensions[title_row].height = 30
                    header_row = 2
                    for col in range(1, 4):
                        cell = summary_ws.cell(row=header_row, column=col)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                    summary_ws.row_dimensions[header_row].height = 25
                    data_start_row = 3
                    for idx, (_, row_data) in enumerate(summary_df.iterrows()):
                        row = data_start_row + idx
                        for col in range(1, 4):
                            cell = summary_ws.cell(row=row, column=col)
                            cell.font = data_font
                            cell.alignment = left_alignment if col <= 2 else center_alignment
                            if col == 3 and cell.value is None:
                                cell.value = float(row_data['金额(元)']) if row_data['金额(元)'] else 0.0
                                cell.number_format = '#,##0.00'
                    sheets_created += 1
                    
                    # 3.2 人工提成明细表
                    if all_details:
                        details_df = pd.DataFrame(all_details)
                        details_df.to_excel(writer, sheet_name='间接人工-人工提成明细', index=False)
                        details_ws = writer.sheets['间接人工-人工提成明细']
                        for col in range(1, len(details_df.columns) + 1):
                            cell = details_ws.cell(row=1, column=col)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_alignment
                            col_letter = get_column_letter(col)
                            if col == 1:
                                details_ws.column_dimensions[col_letter].width = 10
                            elif col == 2:
                                details_ws.column_dimensions[col_letter].width = 18
                            elif col == 3:
                                details_ws.column_dimensions[col_letter].width = 30
                            else:
                                details_ws.column_dimensions[col_letter].width = 18
                        for row in range(2, len(details_df) + 2):
                            for col in range(1, len(details_df.columns) + 1):
                                cell = details_ws.cell(row=row, column=col)
                                cell.font = data_font
                                cell.alignment = center_alignment
                                if col == len(details_df.columns):
                                    cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                                    cell.font = Font(bold=True, name="仿宋", color="2E7D32")
                        sheets_created += 1
                    
                    # 3.3 分摊固定成本明细表
                    fixed_cost_details = []
                    fixed_cost_index = 1
                    
                    def collect_fixed_cost_details(items, get_code, get_name, get_category, get_origin_code, get_origin_name):
                        nonlocal fixed_cost_index
                        for item in items:
                            fixed_costs = {
                                '品管分摊': item.get('品管分摊固定成本', 0) or 0,
                                '物流主管分摊': item.get('物流主管分摊固定成本', 0) or 0,
                                '生产主管分摊': item.get('生产主管分摊固定成本', 0) or 0,
                                '物流卸货分摊': item.get('物流卸货分摊固定成本', 0) or 0,
                                '维修班长分摊': item.get('维修班长分摊固定成本', 0) or 0,
                                '维修员分摊': item.get('维修员分摊固定成本', 0) or 0,
                                '冰箱维修主管分摊': item.get('冰箱维修主管分摊固定成本', 0) or 0,
                                '叉车司磅库管等分摊': item.get('叉车司磅库管等分摊固定成本', 0) or 0,
                                '白电小组长分摊': 0,
                                '生产班组长(黑电)分摊': 0,
                                '生产班组长(冰箱)分摊': 0,
                                '生产班组长(塑料破碎)分摊': 0
                            }
                            material_name = str(get_name(item) or '').upper()
                            category = get_category(item)
                            if category == '旧机':
                                if '空调' in material_name or '洗衣机' in material_name:
                                    fixed_costs['白电小组长分摊'] = item.get('白电小组长分摊固定成本', 0) or 0
                                elif '电视' in material_name or '电脑' in material_name:
                                    fixed_costs['生产班组长(黑电)分摊'] = item.get('生产班组长(黑电)分摊固定成本', 0) or 0
                                elif '冰箱' in material_name:
                                    fixed_costs['生产班组长(冰箱)分摊'] = item.get('生产班组长(冰箱)分摊固定成本', 0) or 0
                            elif category in ['一次拆解产物', '一破']:
                                fixed_costs['生产班组长(塑料破碎)分摊'] = item.get('生产班组长(塑料破碎)分摊固定成本', 0) or 0
                            plastic_fc_no = float(
                                item.get('生产班组长(塑料破碎)分摊固定成本(不考虑期初库存和库存结余)', 0) or 0
                            )
                            total_fixed_cost = sum(fixed_costs.values())
                            if total_fixed_cost > 0 or plastic_fc_no > 0:
                                fixed_cost_details.append({
                                    '序号': fixed_cost_index,
                                    '原物料代码': get_origin_code(item) or '',
                                    '原物料名称': get_origin_name(item) or '',
                                    '物料/产物编码': get_code(item) or '',
                                    '物料/产物名称': get_name(item) or '',
                                    '类别': category,
                                    '品管分摊': fixed_costs['品管分摊'],
                                    '物流主管分摊': fixed_costs['物流主管分摊'],
                                    '生产主管分摊': fixed_costs['生产主管分摊'],
                                    '物流卸货分摊': fixed_costs['物流卸货分摊'],
                                    '维修班长分摊': fixed_costs['维修班长分摊'],
                                    '维修员分摊': fixed_costs['维修员分摊'],
                                    '冰箱维修主管分摊': fixed_costs['冰箱维修主管分摊'],
                                    '叉车司磅库管等分摊': fixed_costs['叉车司磅库管等分摊'],
                                    '白电小组长分摊': fixed_costs['白电小组长分摊'],
                                    '生产班组长(黑电)分摊': fixed_costs['生产班组长(黑电)分摊'],
                                    '生产班组长(冰箱)分摊': fixed_costs['生产班组长(冰箱)分摊'],
                                    '生产班组长(塑料破碎)分摊': fixed_costs['生产班组长(塑料破碎)分摊'],
                                    '生产班组长(塑料破碎)分摊（不考虑期初库存和库存结余）': plastic_fc_no,
                                    '分摊固定成本合计': total_fixed_cost
                                })
                                fixed_cost_index += 1
                    
                    if result.get('part1_details'):
                        collect_fixed_cost_details(
                            result['part1_details'],
                            lambda item: item.get('物料代码', ''),
                            lambda item: item.get('物料名称', ''),
                            lambda item: '旧机',
                            lambda item: item.get('原物料代码', '') or item.get('物料代码', ''),
                            lambda item: item.get('原物料名称', '') or item.get('物料名称', '')
                        )
                    if result.get('part2_details'):
                        collect_fixed_cost_details(
                            result['part2_details'],
                            lambda item: item.get('拆解产物编码', ''),
                            lambda item: item.get('拆解产物名称', ''),
                            lambda item: '一次拆解产物',
                            lambda item: item.get('原物料代码', ''),
                            lambda item: item.get('原物料名称', '')
                        )
                    if result.get('part3_details'):
                        collect_fixed_cost_details(
                            result['part3_details'],
                            lambda item: item.get('深加工产物编码', ''),
                            lambda item: item.get('深加工产物名称', ''),
                            lambda item: item.get('类别', ''),
                            lambda item: item.get('原物料代码', ''),
                            lambda item: item.get('原物料名称', '')
                        )
                    
                    if fixed_cost_details:
                        fixed_cost_df = pd.DataFrame(fixed_cost_details)
                        fixed_cost_df.to_excel(writer, sheet_name='间接人工-分摊固定成本明细', index=False)
                        fixed_cost_ws = writer.sheets['间接人工-分摊固定成本明细']
                        for col in range(1, len(fixed_cost_df.columns) + 1):
                            cell = fixed_cost_ws.cell(row=1, column=col)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_alignment
                            col_letter = get_column_letter(col)
                            if col == 1:  # 序号
                                fixed_cost_ws.column_dimensions[col_letter].width = 10
                            elif col == 2:  # 原物料代码
                                fixed_cost_ws.column_dimensions[col_letter].width = 18
                            elif col == 3:  # 原物料名称
                                fixed_cost_ws.column_dimensions[col_letter].width = 30
                            elif col == 4:  # 物料/产物编码
                                fixed_cost_ws.column_dimensions[col_letter].width = 18
                            elif col == 5:  # 物料/产物名称
                                fixed_cost_ws.column_dimensions[col_letter].width = 30
                            elif col == 6:  # 类别
                                fixed_cost_ws.column_dimensions[col_letter].width = 12
                            else:
                                fixed_cost_ws.column_dimensions[col_letter].width = 18
                        for row in range(2, len(fixed_cost_df) + 2):
                            for col in range(1, len(fixed_cost_df.columns) + 1):
                                cell = fixed_cost_ws.cell(row=row, column=col)
                                cell.font = data_font
                                cell.alignment = center_alignment
                                if col == len(fixed_cost_df.columns):
                                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                                    cell.font = Font(bold=True, name="仿宋", color="856404")
                        sheets_created += 1
                    
                    # 3.4 其他岗位成本明细表
                    other_positions_details = result.get('other_positions_details', [])
                    if other_positions_details:
                        other_positions_data = []
                        for item in other_positions_details:
                            other_positions_data.append({
                                '序号': len(other_positions_data) + 1,
                                '岗位': item.get('岗位', ''),
                                '人员基础配置': item.get('人员基础配置', 0),
                                '平均工资（元/月/人）': item.get('平均工资（元/月/人）', 0),
                                '奖励/补助（元/月）': item.get('奖励/补助（元/月）', 0),
                                '餐补（元/月/人）': item.get('餐补（元/月/人）', 0),
                                '年终奖（元/人）': item.get('年终奖（元/人）', 0),
                                '养老保险费（元/月/人）': item.get('养老保险费（元/月/人）', 0),
                                '失业保险费（元/月/人）': item.get('失业保险费（元/月/人）', 0),
                                '医疗/生育保险费（元/月/人）': item.get('医疗/生育保险费（元/月/人）', 0),
                                '工伤保险费（元/月/人）': item.get('工伤保险费（元/月/人）', 0),
                                '住房公积金（元/月/人）': item.get('住房公积金（元/月/人）', 0),
                                '月均固定成本（元/月/人）': item.get('月均固定成本（元/月/人）', 0),
                                '岗位成本': item.get('岗位成本', 0)
                            })
                        if other_positions_data:
                            total_personnel = sum(item.get('人员基础配置', 0) for item in other_positions_data)
                            total_cost = sum(item.get('岗位成本', 0) for item in other_positions_data)
                            other_positions_data.append({
                                '序号': '',
                                '岗位': '合计',
                                '人员基础配置': total_personnel,
                                '平均工资（元/月/人）': '',
                                '奖励/补助（元/月）': '',
                                '餐补（元/月/人）': '',
                                '年终奖（元/人）': '',
                                '养老保险费（元/月/人）': '',
                                '失业保险费（元/月/人）': '',
                                '医疗/生育保险费（元/月/人）': '',
                                '工伤保险费（元/月/人）': '',
                                '住房公积金（元/月/人）': '',
                                '月均固定成本（元/月/人）': '',
                                '岗位成本': total_cost
                            })
                            other_positions_df = pd.DataFrame(other_positions_data)
                            other_positions_df.to_excel(writer, sheet_name='间接人工-其他岗位成本明细', index=False)
                            other_positions_ws = writer.sheets['间接人工-其他岗位成本明细']
                            for col in range(1, len(other_positions_df.columns) + 1):
                                cell = other_positions_ws.cell(row=1, column=col)
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = center_alignment
                                other_positions_ws.column_dimensions[get_column_letter(col)].width = 20
                            for row in range(2, len(other_positions_df) + 2):
                                for col in range(1, len(other_positions_df.columns) + 1):
                                    cell = other_positions_ws.cell(row=row, column=col)
                                    cell.font = data_font
                                    cell.alignment = center_alignment
                            sheets_created += 1
            except Exception as e:
                errors.append(f"间接人工成本: {str(e)}")
            
            # 4. 制造费用成本
            try:
                result = calculate_manufacturing_cost(app_data, prediction_period)
                if result.get('success', False):
                    # 4.1 与拆解量相关的费用
                    disassembly_data = []
                    for item in result.get('disassembly_related', []):
                        for detail in item.get('明细', []):
                            disassembly_data.append({
                                '费用类型': item.get('费用类型', ''),
                                '费用种类': item.get('费用种类', ''),
                                '费用名称': item.get('费用名称', ''),
                                '类别': detail.get('category', ''),
                                '数量/重量': detail.get('quantity', 0),
                                '单价': detail.get('unit_price', 0),
                                '成本': detail.get('cost', 0)
                            })
                        if item.get('明细'):
                            disassembly_data.append({
                                '费用类型': item.get('费用类型', ''),
                                '费用种类': item.get('费用种类', ''),
                                '费用名称': f"{item.get('费用名称', '')}小计",
                                '类别': '',
                                '数量/重量': '',
                                '单价': '',
                                '成本': item.get('总成本', 0)
                            })
                    
                    if disassembly_data:
                        disassembly_df = pd.DataFrame(disassembly_data)
                        disassembly_df.to_excel(writer, sheet_name='制造费用-与拆解量相关', index=False)
                        worksheet = writer.sheets['制造费用-与拆解量相关']
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_alignment
                        for col in range(1, len(disassembly_df.columns) + 1):
                            worksheet.column_dimensions[get_column_letter(col)].width = 20
                        sheets_created += 1
                    
                    # 4.2 与电机入库量相关的费用
                    motor_data = []
                    for item in result.get('motor_inventory_related', []):
                        for detail in item.get('明细', []):
                            motor_data.append({
                                '费用类型': item.get('费用类型', ''),
                                '费用种类': item.get('费用种类', ''),
                                '费用名称': item.get('费用名称', ''),
                                '类别': detail.get('category', ''),
                                '数量/重量': detail.get('quantity', 0),
                                '单价': detail.get('unit_price', 0),
                                '成本': detail.get('cost', 0)
                            })
                        if item.get('明细'):
                            motor_data.append({
                                '费用类型': item.get('费用类型', ''),
                                '费用种类': item.get('费用种类', ''),
                                '费用名称': f"{item.get('费用名称', '')}小计",
                                '类别': '',
                                '数量/重量': '',
                                '单价': '',
                                '成本': item.get('总成本', 0)
                            })
                    
                    if motor_data:
                        motor_df = pd.DataFrame(motor_data)
                        motor_df.to_excel(writer, sheet_name='制造费用-与电机入库量相关', index=False)
                        worksheet = writer.sheets['制造费用-与电机入库量相关']
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_alignment
                        for col in range(1, len(motor_df.columns) + 1):
                            worksheet.column_dimensions[get_column_letter(col)].width = 20
                        sheets_created += 1
                    
                    # 4.3 预计月均费用
                    monthly_data = []
                    for item in result.get('monthly_average', []):
                        for detail in item.get('明细', []):
                            monthly_data.append({
                                '费用类型': item.get('费用类型', ''),
                                '费用种类': item.get('费用种类', ''),
                                '费用名称': item.get('费用名称', ''),
                                '类别': detail.get('category', ''),
                                '月均费用': detail.get('monthly_cost', 0),
                                '期数': detail.get('periods', 0),
                                '成本': detail.get('cost', 0)
                            })
                        if item.get('明细'):
                            monthly_data.append({
                                '费用类型': item.get('费用类型', ''),
                                '费用种类': item.get('费用种类', ''),
                                '费用名称': f"{item.get('费用名称', '')}小计",
                                '类别': '',
                                '月均费用': '',
                                '期数': '',
                                '成本': item.get('总成本', 0)
                            })
                    
                    if monthly_data:
                        monthly_df = pd.DataFrame(monthly_data)
                        monthly_df.to_excel(writer, sheet_name='制造费用-预计月均费用', index=False)
                        worksheet = writer.sheets['制造费用-预计月均费用']
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_alignment
                        for col in range(1, len(monthly_df.columns) + 1):
                            worksheet.column_dimensions[get_column_letter(col)].width = 20
                        sheets_created += 1
                    
                    # 4.4 环保费
                    environmental_fee_data = []
                    environmental_fee_info = result.get('environmental_fee', {})
                    if environmental_fee_info and environmental_fee_info.get('明细'):
                        for detail in environmental_fee_info.get('明细', []):
                            environmental_fee_data.append({
                                '拆解产物编码': detail.get('拆解产物编码', ''),
                                '拆解产物名称': detail.get('拆解产物名称', ''),
                                '类别': detail.get('类别', ''),
                                '重量(KG)': detail.get('重量', 0),
                                '单价(元/KG)': detail.get('单价', 0),
                                '费用(元)': detail.get('费用', 0)
                            })
                        if environmental_fee_data:
                            environmental_fee_data.append({
                                '拆解产物编码': '',
                                '拆解产物名称': '环保费小计',
                                '类别': '',
                                '重量(KG)': '',
                                '单价(元/KG)': '',
                                '费用(元)': environmental_fee_info.get('总成本', 0)
                            })
                    
                    if environmental_fee_data:
                        environmental_fee_df = pd.DataFrame(environmental_fee_data)
                        environmental_fee_df.to_excel(writer, sheet_name='制造费用-环保费', index=False)
                        worksheet = writer.sheets['制造费用-环保费']
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_alignment
                        for col in range(1, len(environmental_fee_df.columns) + 1):
                            worksheet.column_dimensions[get_column_letter(col)].width = 20
                        sheets_created += 1
                    
                    # 4.5 分类费用汇总
                    category_mapping = {'电视': 'tv', '电脑': 'computer', '冰箱': 'fridge', '空调': 'ac', '洗衣机': 'washing'}
                    category_stats = {cat: {'disassembly': 0.0, 'motor': 0.0, 'monthly': 0.0, 'environmental': 0.0} 
                                    for cat in category_mapping.keys()}
                    for item in result.get('disassembly_related', []):
                        for detail in item.get('明细', []):
                            cat = detail.get('category', '')
                            if cat in category_stats:
                                category_stats[cat]['disassembly'] += detail.get('cost', 0)
                    for item in result.get('motor_inventory_related', []):
                        for detail in item.get('明细', []):
                            cat = detail.get('category', '')
                            if cat in category_stats:
                                category_stats[cat]['motor'] += detail.get('cost', 0)
                    for item in result.get('monthly_average', []):
                        for detail in item.get('明细', []):
                            cat = detail.get('category', '')
                            if cat in category_stats:
                                category_stats[cat]['monthly'] += detail.get('cost', 0)
                    env_details = environmental_fee_info.get('明细', []) if environmental_fee_info else []
                    for detail in env_details:
                        product_name = detail.get('拆解产物名称', '')
                        fee = detail.get('费用', 0)
                        if '彩电' in product_name or '电视' in product_name:
                            category_stats['电视']['environmental'] += fee
                        elif '电脑' in product_name or '显示器' in product_name or '主机' in product_name or '笔记本' in product_name:
                            category_stats['电脑']['environmental'] += fee
                        elif '冰箱' in product_name or '冰柜' in product_name:
                            category_stats['冰箱']['environmental'] += fee
                        elif '空调' in product_name:
                            category_stats['空调']['environmental'] += fee
                        elif '洗衣机' in product_name:
                            category_stats['洗衣机']['environmental'] += fee
                    
                    category_summary_data = []
                    for cat in ['电视', '电脑', '冰箱', '空调', '洗衣机']:
                        stats = category_stats[cat]
                        total = stats['disassembly'] + stats['motor'] + stats['monthly'] + stats['environmental']
                        category_summary_data.append({
                            '类别': cat,
                            '与拆解量相关': stats['disassembly'],
                            '与电机入库量相关': stats['motor'],
                            '预计月均费用': stats['monthly'],
                            '环保费': stats['environmental'],
                            '合计': total
                        })
                    
                    if category_summary_data:
                        category_summary_df = pd.DataFrame(category_summary_data)
                        category_summary_df.to_excel(writer, sheet_name='制造费用-分类费用汇总', index=False)
                        worksheet = writer.sheets['制造费用-分类费用汇总']
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_alignment
                        for col in range(1, len(category_summary_df.columns) + 1):
                            worksheet.column_dimensions[get_column_letter(col)].width = 20
                        sheets_created += 1
                    
                    # 4.6 汇总
                    environmental_fee_total = environmental_fee_info.get('总成本', 0) if environmental_fee_info else 0
                    summary_data = [{
                        '费用类别': '与拆解量相关',
                        '总成本': sum(item.get('总成本', 0) for item in result.get('disassembly_related', []))
                    }, {
                        '费用类别': '与电机入库量相关',
                        '总成本': sum(item.get('总成本', 0) for item in result.get('motor_inventory_related', []))
                    }, {
                        '费用类别': '预计月均费用',
                        '总成本': sum(item.get('总成本', 0) for item in result.get('monthly_average', []))
                    }, {
                        '费用类别': '环保费',
                        '总成本': environmental_fee_total
                    }, {
                        '费用类别': '制造费用总成本',
                        '总成本': result.get('total_cost', 0)
                    }]
                    if summary_data:
                        summary_df = pd.DataFrame(summary_data)
                        summary_df.to_excel(writer, sheet_name='制造费用-汇总', index=False)
                        worksheet = writer.sheets['制造费用-汇总']
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_alignment
                        for col in range(1, len(summary_df.columns) + 1):
                            worksheet.column_dimensions[get_column_letter(col)].width = 20
                        sheets_created += 1
            except Exception as e:
                errors.append(f"制造费用成本: {str(e)}")
            
            # 5. 公共费用分摊明细
            try:
                result = calculate_screen_cost_allocation(app_data, prediction_period)
                if result.get('success', False):
                    # 5.1 直接人工
                    direct_labor = result.get('direct_labor', {})
                    direct_labor_data = []
                    for item in direct_labor.get('screen_wage_details', []):
                        code = item.get('物料代码') or item.get('拆解产物编码') or item.get('深加工产物编码') or ''
                        name = item.get('物料名称') or item.get('拆解产物名称') or item.get('深加工产物名称') or ''
                        wage = item.get('工资', 0.0)
                        direct_labor_data.append({
                            '类型': '屏工资（计件工资）',
                            '物料代码': code,
                            '物料名称': name,
                            '金额（元）': wage
                        })
                    for item in direct_labor.get('screen_allocation_details', []):
                        item_data = item.get('item', {}) if isinstance(item.get('item'), dict) else {}
                        code = item_data.get('物料代码') or item_data.get('拆解产物编码') or item_data.get('深加工产物编码') or ''
                        name = item_data.get('物料名称') or item_data.get('拆解产物名称') or item_data.get('深加工产物名称') or ''
                        fixed_cost = item.get('fixed_cost', 0.0)
                        direct_labor_data.append({
                            '类型': '屏分摊（固定工资分摊）',
                            '物料代码': code,
                            '物料名称': name,
                            '金额（元）': fixed_cost
                        })
                    direct_labor_data.append({
                        '类型': '屏工资小计',
                        '物料代码': '',
                        '物料名称': '',
                        '金额（元）': direct_labor.get('screen_wage', 0.0)
                    })
                    direct_labor_data.append({
                        '类型': '屏分摊小计',
                        '物料代码': '',
                        '物料名称': '',
                        '金额（元）': direct_labor.get('screen_allocation', 0.0)
                    })
                    
                    if direct_labor_data:
                        direct_labor_df = pd.DataFrame(direct_labor_data)
                        direct_labor_df.to_excel(writer, sheet_name='公共费用分摊-直接人工', index=False)
                        worksheet = writer.sheets['公共费用分摊-直接人工']
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_alignment
                        for col in range(1, len(direct_labor_df.columns) + 1):
                            worksheet.column_dimensions[get_column_letter(col)].width = 20
                        sheets_created += 1
                    
                    # 5.2 制造费用-与拆解量相关的费用
                    manufacturing_cost = result.get('manufacturing_cost', {})
                    disassembly_data = manufacturing_cost.get('disassembly_related_details', [])
                    if disassembly_data:
                        disassembly_df = pd.DataFrame(disassembly_data)
                        disassembly_df.to_excel(writer, sheet_name='公共费用分摊-与拆解量相关', index=False)
                        worksheet = writer.sheets['公共费用分摊-与拆解量相关']
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_alignment
                        for col in range(1, len(disassembly_df.columns) + 1):
                            worksheet.column_dimensions[get_column_letter(col)].width = 20
                        sheets_created += 1
                    
                    # 5.3 制造费用-预计月均费用
                    monthly_data = manufacturing_cost.get('monthly_average_details', [])
                    if monthly_data:
                        monthly_df = pd.DataFrame(monthly_data)
                        monthly_df.to_excel(writer, sheet_name='公共费用分摊-预计月均费用', index=False)
                        worksheet = writer.sheets['公共费用分摊-预计月均费用']
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_alignment
                        for col in range(1, len(monthly_df.columns) + 1):
                            worksheet.column_dimensions[get_column_letter(col)].width = 20
                        sheets_created += 1
                    
                    # 5.4 制造费用-环保费
                    environmental_data = manufacturing_cost.get('environmental_fee_details', [])
                    if environmental_data:
                        environmental_df = pd.DataFrame(environmental_data)
                        environmental_df.to_excel(writer, sheet_name='公共费用分摊-环保费', index=False)
                        worksheet = writer.sheets['公共费用分摊-环保费']
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_alignment
                        for col in range(1, len(environmental_df.columns) + 1):
                            worksheet.column_dimensions[get_column_letter(col)].width = 20
                        sheets_created += 1
                    
                    # 5.5 被减扣数据统计
                    deducted_data = result.get('deducted_data', {})
                    deducted_list = []
                    for item in deducted_data.get('tv_details', []):
                        deducted_list.append({
                            '类别': '电视',
                            '拆解产物编码': item.get('拆解产物编码', ''),
                            '拆解产物名称': item.get('拆解产物名称', ''),
                            '重量（KG）': item.get('重量(KG)', 0.0)
                        })
                    for item in deducted_data.get('computer_details', []):
                        deducted_list.append({
                            '类别': '电脑',
                            '拆解产物编码': item.get('拆解产物编码', ''),
                            '拆解产物名称': item.get('拆解产物名称', ''),
                            '重量（KG）': item.get('重量(KG)', 0.0)
                        })
                    deducted_list.append({
                        '类别': '电视小计',
                        '拆解产物编码': '',
                        '拆解产物名称': '',
                        '重量（KG）': deducted_data.get('tv_weight', 0.0)
                    })
                    deducted_list.append({
                        '类别': '电脑小计',
                        '拆解产物编码': '',
                        '拆解产物名称': '',
                        '重量（KG）': deducted_data.get('computer_weight', 0.0)
                    })
                    deducted_list.append({
                        '类别': '总重量',
                        '拆解产物编码': '',
                        '拆解产物名称': '',
                        '重量（KG）': deducted_data.get('total_weight', 0.0)
                    })
                    
                    if deducted_list:
                        deducted_df = pd.DataFrame(deducted_list)
                        deducted_df.to_excel(writer, sheet_name='公共费用分摊-被减扣数据统计', index=False)
                        worksheet = writer.sheets['公共费用分摊-被减扣数据统计']
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_alignment
                        for col in range(1, len(deducted_df.columns) + 1):
                            worksheet.column_dimensions[get_column_letter(col)].width = 20
                        sheets_created += 1
                    
                    # 5.6 制造费用间接人工分摊
                    indirect_labor_allocation = result.get('indirect_labor_allocation', {})
                    indirect_labor_data = []
                    for detail in indirect_labor_allocation.get('details', []):
                        indirect_labor_data.append({
                            '岗位': detail.get('岗位', ''),
                            '拆解产物编码': detail.get('拆解产物编码', ''),
                            '拆解产物名称': detail.get('拆解产物名称', ''),
                            '分类': detail.get('分类', ''),
                            '物料产值（元）': detail.get('物料产值（元）', 0.0),
                            '分摊比例': detail.get('分摊比例', 0.0),
                            '分摊成本（元）': detail.get('分摊成本（元）', 0.0)
                        })
                    category_totals = indirect_labor_allocation.get('category_totals', {})
                    for category, total in category_totals.items():
                        if total > 0:
                            indirect_labor_data.append({
                                '岗位': f'{category}小计',
                                '拆解产物编码': '',
                                '拆解产物名称': '',
                                '分类': category,
                                '物料产值（元）': 0.0,
                                '分摊比例': 0.0,
                                '分摊成本（元）': total
                            })
                    total_cost = indirect_labor_allocation.get('total_cost', 0.0)
                    if total_cost > 0 or len(indirect_labor_data) > 0:
                        indirect_labor_data.append({
                            '岗位': '总合计',
                            '拆解产物编码': '',
                            '拆解产物名称': '',
                            '分类': '',
                            '物料产值（元）': 0.0,
                            '分摊比例': 0.0,
                            '分摊成本（元）': total_cost
                        })
                    
                    if indirect_labor_data:
                        indirect_labor_df = pd.DataFrame(indirect_labor_data)
                        indirect_labor_df.to_excel(writer, sheet_name='公共费用分摊-制造费用间接人工分摊', index=False)
                        worksheet = writer.sheets['公共费用分摊-制造费用间接人工分摊']
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_alignment
                        for col in range(1, len(indirect_labor_df.columns) + 1):
                            worksheet.column_dimensions[get_column_letter(col)].width = 20
                        sheets_created += 1
                    
                    # 5.7 制造费用公共成本分摊
                    public_cost_allocation = result.get('public_cost_allocation', {})
                    public_cost_data = []
                    for detail in public_cost_allocation.get('details', []):
                        public_cost_data.append({
                            '费用类型': detail.get('费用类型', ''),
                            '费用种类': detail.get('费用种类', ''),
                            '费用名称': detail.get('费用名称', ''),
                            '拆解产物编码': detail.get('拆解产物编码', ''),
                            '拆解产物名称': detail.get('拆解产物名称', ''),
                            '分类': detail.get('分类', ''),
                            '物料产值（元）': detail.get('物料产值（元）', 0.0),
                            '分摊比例': detail.get('分摊比例', 0.0),
                            '分摊成本（元）': detail.get('分摊成本（元）', 0.0)
                        })
                    category_totals = public_cost_allocation.get('category_totals', {})
                    for category, total in category_totals.items():
                        if total > 0:
                            public_cost_data.append({
                                '费用类型': f'{category}小计',
                                '费用种类': '',
                                '费用名称': '',
                                '拆解产物编码': '',
                                '拆解产物名称': '',
                                '分类': category,
                                '物料产值（元）': 0.0,
                                '分摊比例': 0.0,
                                '分摊成本（元）': total
                            })
                    total_cost = public_cost_allocation.get('total_cost', 0.0)
                    if total_cost > 0 or len(public_cost_data) > 0:
                        public_cost_data.append({
                            '费用类型': '总合计',
                            '费用种类': '',
                            '费用名称': '',
                            '拆解产物编码': '',
                            '拆解产物名称': '',
                            '分类': '',
                            '物料产值（元）': 0.0,
                            '分摊比例': 0.0,
                            '分摊成本（元）': total_cost
                        })
                    
                    if public_cost_data:
                        public_cost_df = pd.DataFrame(public_cost_data)
                        public_cost_df.to_excel(writer, sheet_name='公共费用分摊-制造费用公共成本分摊', index=False)
                        worksheet = writer.sheets['公共费用分摊-制造费用公共成本分摊']
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_alignment
                        for col in range(1, len(public_cost_df.columns) + 1):
                            worksheet.column_dimensions[get_column_letter(col)].width = 20
                        sheets_created += 1
                    
                    # 5.8 分摊计算结果
                    allocation = result.get('allocation', {})
                    allocation_data = []
                    cost_details = allocation.get('cost_details', {})
                    screen_allocation = allocation.get('screen_allocation', {})
                    allocation_data.append({
                        '项目': '制造费用间接人工分摊',
                        '明细': '',
                        '金额（元）': cost_details.get('制造费用间接人工分摊', 0.0)
                    })
                    allocation_data.append({
                        '项目': '制造费用公共成本分摊',
                        '明细': '',
                        '金额（元）': cost_details.get('制造费用公共成本分摊', 0.0)
                    })
                    allocation_data.append({
                        '项目': '屏费用分摊结果',
                        '明细': '',
                        '金额（元）': screen_allocation.get('total_cost', 0.0)
                    })
                    direct_labor_details = cost_details.get('直接人工', {})
                    allocation_data.append({
                        '项目': '  └─ 直接人工',
                        '明细': '',
                        '金额（元）': direct_labor_details.get('小计', 0.0)
                    })
                    allocation_data.append({
                        '项目': '    └─ 屏工资',
                        '明细': '',
                        '金额（元）': direct_labor_details.get('屏工资', 0.0)
                    })
                    allocation_data.append({
                        '项目': '    └─ 屏分摊',
                        '明细': '',
                        '金额（元）': direct_labor_details.get('屏分摊', 0.0)
                    })
                    manufacturing_details = cost_details.get('制造费用（屏）', {})
                    allocation_data.append({
                        '项目': '  └─ 制造费用（屏）',
                        '明细': '',
                        '金额（元）': manufacturing_details.get('小计', 0.0)
                    })
                    allocation_data.append({
                        '项目': '    └─ 与拆解量相关的费用',
                        '明细': '',
                        '金额（元）': manufacturing_details.get('与拆解量相关的费用', 0.0)
                    })
                    allocation_data.append({
                        '项目': '    └─ 预计月均费用',
                        '明细': '',
                        '金额（元）': manufacturing_details.get('预计月均费用', 0.0)
                    })
                    allocation_data.append({
                        '项目': '    └─ 环保费',
                        '明细': '',
                        '金额（元）': manufacturing_details.get('环保费', 0.0)
                    })
                    allocation_data.append({
                        '项目': '  └─ 电视（屏）',
                        '明细': '',
                        '金额（元）': screen_allocation.get('tv_allocation', 0.0)
                    })
                    allocation_data.append({
                        '项目': '  └─ 电脑（屏）',
                        '明细': '',
                        '金额（元）': screen_allocation.get('computer_allocation', 0.0)
                    })
                    allocation_data.append({
                        '项目': '总费用',
                        '明细': '',
                        '金额（元）': allocation.get('total_cost', 0.0)
                    })
                    allocation_data.append({
                        '项目': '',
                        '明细': '',
                        '金额（元）': 0.0
                    })
                    allocation_data.append({
                        '项目': '按类别分摊',
                        '明细': '',
                        '金额（元）': 0.0
                    })
                    category_allocation = allocation.get('category_allocation', {})
                    for category in ['冰箱', '空调', '电脑', '电视', '洗衣机']:
                        allocation_data.append({
                            '项目': f'  └─ {category}',
                            '明细': '',
                            '金额（元）': category_allocation.get(category, 0.0)
                        })
                    
                    if allocation_data:
                        allocation_df = pd.DataFrame(allocation_data)
                        allocation_df.to_excel(writer, sheet_name='公共费用分摊-分摊计算结果', index=False)
                        worksheet = writer.sheets['公共费用分摊-分摊计算结果']
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_alignment
                        for col in range(1, len(allocation_df.columns) + 1):
                            worksheet.column_dimensions[get_column_letter(col)].width = 25
                        sheets_created += 1
            except Exception as e:
                errors.append(f"公共费用分摊明细: {str(e)}")
            
            # 6. 期间费用
            try:
                result = calculate_period_cost(
                    app_data, 
                    prediction_period,
                    quality_manager_ratio,
                    quality_group_ratio,
                    warehouse_group_ratio
                )
                if result.get('success', False):
                    export_data = []
                    columns = [
                        '费用明细',
                        '管理-电废事业部-质量管理',
                        '管理-电废事业部-库房',
                        '管理-电废事业部-平台',
                        '管理-电废事业部-基金管理项目组',
                        '管理-电废事业部-回收经营项目组',
                        '管理-消电项目组',
                        '销售-平台',
                        '销售-消电项目组',
                        '屏',
                        '合计',
                    ]
                    value_columns = columns[1:-1]

                    if result.get('salary_row'):
                        salary_row = result['salary_row'].copy()
                        salary_row['费用明细'] = '薪酬费用'
                        row_total = sum(float(salary_row.get(col, 0) or 0) for col in value_columns)
                        salary_row['合计'] = row_total
                        export_data.append(salary_row)

                    if result.get('period_cost_data'):
                        for row in result['period_cost_data']:
                            row_data = row.copy()
                            row_total = sum(
                                float(row_data.get(col, 0) or 0) if row_data.get(col) is not None else 0
                                for col in value_columns
                            )
                            row_data['合计'] = row_total
                            export_data.append(row_data)

                    finance_row = {'费用明细': '财务费用'}
                    for col in value_columns:
                        finance_row[col] = None
                    finance_row['合计'] = 0
                    export_data.append(finance_row)

                    if result.get('totals'):
                        total_row = {'费用明细': '合计'}
                        column_total_sum = 0
                        for col in value_columns:
                            total_value = result['totals'].get(col, 0) or 0
                            total_row[col] = float(total_value)
                            column_total_sum += float(total_value)
                        total_row['合计'] = column_total_sum
                        export_data.append(total_row)
                    
                    if export_data:
                        export_df = pd.DataFrame(export_data)
                        export_df = export_df[columns]
                        export_df.to_excel(writer, sheet_name='期间费用', index=False)
                        worksheet = writer.sheets['期间费用']
                        for col in range(1, len(export_df.columns) + 1):
                            cell = worksheet.cell(row=1, column=col)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_alignment
                            col_letter = get_column_letter(col)
                            col_name = export_df.columns[col - 1]
                            if col_name == '费用明细':
                                worksheet.column_dimensions[col_letter].width = 30
                            elif col_name == '合计':
                                worksheet.column_dimensions[col_letter].width = 18
                            else:
                                worksheet.column_dimensions[col_letter].width = 25
                        sheets_created += 1
            except Exception as e:
                errors.append(f"期间费用: {str(e)}")
            
            # 7. 税金及附加
            try:
                tax_result = calculate_tax_surcharge(app_data, prediction_period)
                if tax_result.get('success', False) and tax_result.get('rows'):
                    tax_categories = TAX_SURCHARGE_CATEGORIES
                    tax_columns = ['项目'] + tax_categories + ['合计']
                    tax_export_data = []
                    for row in tax_result.get('rows', []):
                        values = row.get('values', {}) or {}
                        row_data = {'项目': row.get('项目', '')}
                        for cat in tax_categories:
                            row_data[cat] = float(values.get(cat, 0) or 0)
                        row_total = row.get('row_total')
                        if row_total is None:
                            row_total = sum(row_data[cat] for cat in tax_categories)
                        row_data['合计'] = float(row_total or 0)
                        tax_export_data.append(row_data)
                    
                    if tax_export_data:
                        tax_export_df = pd.DataFrame(tax_export_data)
                        tax_export_df = tax_export_df[tax_columns]
                        tax_export_df.to_excel(writer, sheet_name='税金及附加', index=False)
                        worksheet = writer.sheets['税金及附加']
                        for col in range(1, len(tax_export_df.columns) + 1):
                            cell = worksheet.cell(row=1, column=col)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_alignment
                            col_letter = get_column_letter(col)
                            col_name = tax_export_df.columns[col - 1]
                            if col_name == '项目':
                                worksheet.column_dimensions[col_letter].width = 35
                            elif col_name == '合计':
                                worksheet.column_dimensions[col_letter].width = 18
                            else:
                                worksheet.column_dimensions[col_letter].width = 16
                        sheets_created += 1
            except Exception as e:
                errors.append(f"税金及附加: {str(e)}")
        
        # 检查是否有任何数据被导出
        if sheets_created == 0:
            return jsonify({
                'success': False,
                'error': '没有可导出的数据。' + ('错误信息: ' + '; '.join(errors) if errors else '')
            }), 400
        
        output.seek(0)
        filename = f'成本预测结果_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"导出所有成本预测数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


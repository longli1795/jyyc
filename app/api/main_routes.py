from flask import Blueprint, render_template, send_file, jsonify, send_from_directory, request, current_app
from app.services.status_service import StatusService
from app.services.data_service import DataService
from app.models.app_data import AppDataManager
from app.models.compatibility import AppDataManagerAdapter
from app.utils.auth_utils import login_required, require_can_edit, page_permission_required
import os

def get_session_data_manager():
    """获取会话数据管理器的便利函数"""
    from flask import session
    session_id = session.get('session_id')
    return AppDataManagerAdapter.get_instance(session_id)

main_bp = Blueprint('main', __name__)

def _get_allowed_pages_for_user(user):
    """返回用户可访问的页面 key 列表（管理员为全部，已展开 data_management 子页面）。"""
    from app.core.page_permissions import get_all_page_keys, expand_allowed_pages
    if not user:
        return []
    if user.is_admin:
        return get_all_page_keys()
    ap = getattr(user, 'allowed_pages', None)
    if ap is None:
        return []
    if isinstance(ap, (str, bytes)):
        import json
        try:
            ap = json.loads(ap) if ap else []
        except (TypeError, ValueError):
            return []
    pages = list(ap) if hasattr(ap, '__iter__') and not isinstance(ap, (str, bytes)) else []
    return expand_allowed_pages(pages)


@main_bp.route('/')
@login_required
def main_index():
    """主页"""
    from flask import session, redirect
    from app.utils.auth_utils import get_current_user
    from app.core.page_permissions import user_has_home_access, DATA_MANAGEMENT_LANDING_PATH

    session_id = session.get('session_id')
    status_service = StatusService(session_id)
    status_info = status_service.get_status_info()
    current_user = get_current_user()
    if current_user and not current_user.is_admin:
        allowed = _get_allowed_pages_for_user(current_user)
        if not user_has_home_access(allowed):
            return redirect(DATA_MANAGEMENT_LANDING_PATH)

    user_is_read_only = getattr(current_user, 'is_read_only', False) if current_user else False
    allowed_pages = _get_allowed_pages_for_user(current_user) if current_user else []
    return render_template('main_index.html', status_info=status_info, current_user=current_user, user_is_read_only=user_is_read_only, allowed_pages=allowed_pages)

@main_bp.route('/cost-forecast/')
@page_permission_required
def cost_forecast():
    """间接人工成本页面"""
    return render_template('indirect_labor_cost.html')

@main_bp.route('/cost-forecast/material-cost/')
@page_permission_required
def material_cost_detail():
    """拆解物原料成本详情页"""
    return render_template('material_cost_detail.html')

@main_bp.route('/cost-forecast/piece-rate-wage/')
@page_permission_required
def piece_rate_wage_detail():
    """生产工人计件工资详情页"""
    return render_template('piece_rate_wage_detail.html')

@main_bp.route('/cost-forecast/manufacturing-cost/')
@page_permission_required
def manufacturing_cost_detail():
    """制造费用成本详情页"""
    return render_template('manufacturing_cost_detail.html')

@main_bp.route('/cost-forecast/screen-cost-allocation/')
@page_permission_required
def screen_cost_allocation_detail():
    """公共费用分摊明细页面"""
    return render_template('screen_cost_allocation_detail.html')

@main_bp.route('/cost-forecast/period-cost/')
@page_permission_required
def period_cost_detail():
    """期间费用详情页面"""
    return render_template('period_cost_detail.html')

@main_bp.route('/cost-forecast/tax-surcharge/')
@page_permission_required
def tax_surcharge_detail():
    """税金及附加详情页面"""
    return render_template('tax_surcharge_detail.html')

@main_bp.route('/cost-forecast/production-cost-allocation/')
@page_permission_required
def production_cost_allocation():
    """生产成本分摊详情页面"""
    return render_template('production_cost_allocation.html')

@main_bp.route('/cost-forecast/disassembly-product-cost/')
@page_permission_required
def disassembly_product_cost():
    """一次拆解产物成本计算详情页面"""
    return render_template('disassembly_product_cost.html')

@main_bp.route('/cost-forecast/deep-processing-product-cost/')
@page_permission_required
def deep_processing_product_cost():
    """深加工产物成本计算详情页面"""
    return render_template('deep_processing_product_cost.html')

@main_bp.route('/cost-forecast/disassembly-profit-analysis/')
@page_permission_required
def disassembly_profit_analysis():
    """当期拆解收益测算分析表详情页"""
    return render_template('disassembly_profit_analysis.html')


@main_bp.route('/data-management/')
@page_permission_required
def data_management():
    """数据管理页面"""
    from flask import session
    from app.utils.auth_utils import get_current_user
    cu = get_current_user()
    user_is_read_only = getattr(cu, 'is_read_only', False) if cu else False
    session_id = session.get('session_id')
    status_service = StatusService(session_id)
    status_info = status_service.get_status_info()
    return render_template('data_management.html', current_user=cu, user_is_read_only=user_is_read_only, status_info=status_info)

@main_bp.route('/data-management/mapping')
@page_permission_required
def mapping_management():
    """映射表管理页面"""
    return render_template('mapping_management.html')

@main_bp.route('/data-management/product')
@page_permission_required
def product_management():
    """产品数据管理页面"""
    return render_template('product_management.html')

@main_bp.route('/data-management/deduction')
@page_permission_required
def deduction_management():
    """减扣规则管理页面"""
    return render_template('deduction_management.html')

@main_bp.route('/data-management/deducted-data-editor')
@page_permission_required
def deducted_data_editor():
    """被减扣数据编辑页面"""
    return render_template('deducted_data_editor.html')

@main_bp.route('/data-management/deep-processing')
@page_permission_required
def deep_processing_management():
    """深加工管理页面"""
    return render_template('deep_processing_management.html')

# 营业收入预测模块 - 基础数据管理路由
@main_bp.route('/revenue-forecast/data-management')
@page_permission_required
def revenue_forecast_data_management():
    """营业收入预测-基础数据管理主页"""
    from flask import session
    from app.utils.auth_utils import get_current_user
    cu = get_current_user()
    user_is_read_only = getattr(cu, 'is_read_only', False) if cu else False
    session_id = session.get('session_id')
    status_service = StatusService(session_id)
    status_info = status_service.get_status_info()
    return render_template('data_management.html', current_user=cu, user_is_read_only=user_is_read_only, status_info=status_info)

@main_bp.route('/revenue-forecast/data-management/mapping')
@page_permission_required
def revenue_forecast_mapping_management():
    """营业收入预测-映射表管理页面"""
    return render_template('mapping_management.html')

@main_bp.route('/revenue-forecast/data-management/deduction')
@page_permission_required
def revenue_forecast_deduction_management():
    """营业收入预测-减扣规则管理页面"""
    return render_template('deduction_management.html')

@main_bp.route('/revenue-forecast/data-management/deep-processing')
@page_permission_required
def revenue_forecast_deep_processing_management():
    """营业收入预测-深加工数据管理页面"""
    return render_template('deep_processing_management.html')

@main_bp.route('/revenue-forecast/data-management/product')
@page_permission_required
def revenue_forecast_product_management():
    """营业收入预测-产品拆解系数管理页面"""
    return render_template('product_management.html')

@main_bp.route('/data-management/deducted-data')
@page_permission_required
def data_management_deducted_data():
    """数据管理 - 被减扣数据编辑页面"""
    return render_template('deducted_data_editor.html')

@main_bp.route('/data-management/extracted-data-editor')
@page_permission_required
def extracted_data_editor():
    """提取结果数据编辑页面"""
    return render_template('extracted_data_editor.html')

@main_bp.route('/data-management/extracted-data')
@page_permission_required
def data_management_extracted_data():
    """数据管理 - 提取结果数据编辑页面"""
    return render_template('extracted_data_editor.html')

@main_bp.route('/revenue-forecast/data-management/saleable-data-editor')
@page_permission_required
def revenue_forecast_saleable_data_editor():
    """营业收入预测-可销售量数据编辑页面"""
    return render_template('saleable_data_editor.html')

@main_bp.route('/revenue-forecast/data-management/price')
@page_permission_required
def revenue_forecast_price_management():
    """营业收入预测-销售价格管理页面"""
    return render_template('price_management.html')

@main_bp.route('/data-management/price')
@page_permission_required
def price_management():
    """销售价格管理页面"""
    return render_template('price_management.html')

@main_bp.route('/revenue-forecast/data-management/subsidy')
@page_permission_required
def revenue_forecast_subsidy_management():
    """营业收入预测-基金补贴单价管理页面"""
    return render_template('subsidy_management.html')

@main_bp.route('/revenue-forecast/sales-revenue/')
@page_permission_required
def sales_revenue_page():
    """销售收益详情页"""
    return render_template('sales_revenue.html')

@main_bp.route('/revenue-forecast/subsidy-income/')
@page_permission_required
def subsidy_income_page():
    """基金补贴收入详情页"""
    return render_template('subsidy_income.html')

@main_bp.route('/revenue-forecast/disassembly-product-output-value/')
@page_permission_required
def disassembly_product_output_value():
    """一次拆解产物产值详情页"""
    return render_template('disassembly_product_output_value.html')

@main_bp.route('/revenue-forecast/deep-processing-product-output-value/')
@page_permission_required
def deep_processing_product_output_value():
    """深加工拆解产物产值详情页"""
    return render_template('deep_processing_product_output_value.html')

@main_bp.route('/data-management/subsidy')
@page_permission_required
def subsidy_management():
    """基金补贴单价管理页面"""
    return render_template('subsidy_management.html')

@main_bp.route('/data-management/labor-cost')
@page_permission_required
def labor_cost_management():
    """旧机拆解人工提成单价管理页面"""
    return render_template('labor_cost_management.html')

@main_bp.route('/revenue-forecast/data-management/labor-cost')
@page_permission_required
def revenue_forecast_labor_cost_management():
    """营业收入预测-旧机拆解人工提成单价管理页面"""
    return render_template('labor_cost_management.html')

# 营业成本预测模块 - 基础数据管理路由
@main_bp.route('/cost-forecast/data-management/salary-accounting')
@page_permission_required
def cost_forecast_salary_accounting_management():
    """成本预测-薪酬核算基础数据管理页面"""
    return render_template('salary_accounting_management.html')

@main_bp.route('/cost-forecast/data-management/manufacturing-cost')
@page_permission_required
def cost_forecast_manufacturing_cost_management():
    """成本预测-制造费用基础数据管理页面"""
    return render_template('manufacturing_cost_management.html')

@main_bp.route('/cost-forecast/data-management')
@page_permission_required
def cost_forecast_data_management():
    """成本预测-基础数据管理主页"""
    from flask import session
    from app.utils.auth_utils import get_current_user
    cu = get_current_user()
    user_is_read_only = getattr(cu, 'is_read_only', False) if cu else False
    session_id = session.get('session_id')
    status_service = StatusService(session_id)
    status_info = status_service.get_status_info()
    return render_template('data_management.html', current_user=cu, user_is_read_only=user_is_read_only, status_info=status_info)

@main_bp.route('/cost-forecast/data-management/period-cost')
@page_permission_required
def cost_forecast_period_cost_management():
    """成本预测-期间费用基础数据管理页面"""
    return render_template('period_cost_management.html')

@main_bp.route('/cost-forecast/data-management/tax-surcharge')
@page_permission_required
def cost_forecast_tax_surcharge_management():
    """成本预测-税金及附加基础数据管理页面"""
    return render_template('tax_surcharge_management.html')

@main_bp.route('/logo/<filename>')
def serve_logo(filename):
    """提供Logo文件"""
    try:
        # 获取项目根目录
        current_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        logo_folder = os.path.join(current_dir, 'logo')
        return send_from_directory(logo_folder, filename)
    except Exception as e:
        return f"Error serving logo: {str(e)}", 500


def _get_project_root():
    return os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


@main_bp.route('/docs/operation-manual/')
@main_bp.route('/docs/operation-manual/<path:filename>')
@login_required
def serve_operation_manual(filename='index.html'):
    """提供系统操作手册静态文件"""
    manual_dir = os.path.join(_get_project_root(), 'docs', 'operation_manual')
    # URL 路径始终使用正斜杠；Windows 下 normpath 转反斜杠会导致 send_from_directory 失败
    safe_path = filename.replace('\\', '/').lstrip('/')
    if not safe_path or '..' in safe_path.split('/'):
        return 'Not Found', 404
    target = os.path.normpath(os.path.join(manual_dir, *safe_path.split('/')))
    manual_dir_norm = os.path.normpath(manual_dir)
    if os.path.commonpath([manual_dir_norm, target]) != manual_dir_norm:
        return 'Not Found', 404
    if not os.path.isfile(target):
        return 'Not Found', 404
    return send_from_directory(manual_dir, safe_path)

@main_bp.route('/status')
@login_required
def get_status():
    """获取当前状态"""
    app_data_manager = get_session_data_manager()
    
    # 检查是否有任何数据
    has_data = any([
        app_data_manager.get_data('source_data') is not None,
        app_data_manager.get_data('extracted_data') is not None,
        app_data_manager.get_data('disassembly_data') is not None,
        app_data_manager.get_data('calculated_data') is not None,
        app_data_manager.get_data('deep_processing_data') is not None,
        app_data_manager.get_data('saleable_data') is not None
    ])
    
    return jsonify({
        'status': app_data_manager.get_data('status') or '准备就绪',
        'has_source': app_data_manager.get_data('source_data') is not None,
        'has_extracted': app_data_manager.get_data('extracted_data') is not None,
        'has_disassembly': app_data_manager.get_data('disassembly_data') is not None,
        'has_calculated': app_data_manager.get_data('calculated_data') is not None,
        'has_deep_processing': app_data_manager.get_data('deep_processing_data') is not None,
        'has_saleable': app_data_manager.get_data('saleable_data') is not None,
        'has_data': has_data
    })

@main_bp.route('/api/status')
@login_required
def get_api_status():
    """获取API状态信息"""
    from app.services.status_service import StatusService
    
    try:
        from flask import session
        session_id = session.get('session_id')
        status_service = StatusService(session_id)
        status_info = status_service.get_status_info()
        
        app_data_manager = get_session_data_manager()
        
        # 添加数据计数信息
        status_info.update({
            'has_statistics': any([
                app_data_manager.get_data('extracted_data') is not None,
                app_data_manager.get_data('calculated_data') is not None,
                app_data_manager.get_data('deducted_data_manual') is not None,  # 🔧 架构重构
                app_data_manager.get_data('deep_processing_data') is not None,
                app_data_manager.get_data('saleable_data') is not None
            ]),
            'extracted_data_count': len(app_data_manager.get_data('extracted_data')) if app_data_manager.get_data('extracted_data') is not None else 0,
            'calculated_data_count': len(app_data_manager.get_data('calculated_data')) if app_data_manager.get_data('calculated_data') is not None else 0,
            'deducted_data_count': len(app_data_manager.get_data('deducted_data_manual')) if app_data_manager.get_data('deducted_data_manual') is not None else 0,  # 🔧 架构重构
            'deep_processing_data_count': len(app_data_manager.get_data('deep_processing_data')) if app_data_manager.get_data('deep_processing_data') is not None else 0,
            'saleable_data_count': len(app_data_manager.get_data('saleable_data')) if app_data_manager.get_data('saleable_data') is not None else 0
        })
        
        return jsonify(status_info)
        
    except Exception as e:
        return jsonify({
            'error': str(e),
            'has_statistics': False,
            'extracted_data_count': 0,
            'calculated_data_count': 0,
            'deducted_data_count': 0,
            'deep_processing_data_count': 0,
            'saleable_data_count': 0
        })

@main_bp.route('/upload', methods=['POST'])
@login_required
@require_can_edit
def upload_file():
    """文件上传处理：固化到 data/persistent 并加载到全局会话"""
    try:
        print("📁 收到文件上传请求")

        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '没有选择文件'})

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '没有选择文件'})

        if not file or not file.filename:
            return jsonify({'success': False, 'message': '无效的文件'})

        from flask import session
        from app.services.opening_inventory_store import (
            GLOBAL_OPENING_SESSION_ID,
            save_uploaded_file,
            load_from_disk_into_memory,
            get_meta,
        )
        from app.utils.auth_utils import get_current_user

        user = get_current_user()
        user_id = user.id if user else session.get('user_id')

        ok, msg, meta = save_uploaded_file(file, uploaded_by_user_id=user_id)
        if not ok:
            return jsonify({'success': False, 'message': msg})

        load_ok, load_msg = load_from_disk_into_memory()
        if not load_ok:
            return jsonify({'success': False, 'message': load_msg})

        data_service = DataService(GLOBAL_OPENING_SESSION_ID)
        summary = data_service.get_data_summary()
        rows = summary.get('source_data', {}).get('rows', 0)

        print(f"✅ 期初库存固化并加载成功: {rows} 行")

        return jsonify({
            'success': True,
            'message': f'期初库存已固化并加载，共 {rows} 行数据',
            'data_summary': summary,
            'opening_inventory_meta': get_meta() or meta,
        })

    except Exception as e:
        error_msg = f'文件上传处理失败: {str(e)}'
        print(f"❌ {error_msg}")
        return jsonify({'success': False, 'message': error_msg})


@main_bp.route('/api/opening-inventory/download')
@login_required
def download_opening_inventory():
    """下载已固化的期初库存文件"""
    from app.services.opening_inventory_store import get_persistent_file_path, get_meta

    path = get_persistent_file_path()
    if not path or not os.path.exists(path):
        return jsonify({'success': False, 'message': '暂无固化的期初库存文件'}), 404

    meta = get_meta() or {}
    download_name = meta.get('original_filename') or os.path.basename(path)
    return send_file(path, as_attachment=True, download_name=download_name)

@main_bp.route('/get_data/<data_type>')
@login_required
def get_data_legacy(data_type):
    """获取指定类型的数据 - 兼容旧版API"""
    from app.models.app_data import AppDataManager
    
    try:
        app_data = get_session_data_manager()
        
        # 数据类型映射
        data_mapping = {
            'source': 'source_data',
            'extracted': 'extracted_data', 
            'original': 'disassembly_data',
            'final': 'calculated_data',
            'deducted': 'deducted_data_manual',  # 🔧 架构重构：使用 deducted_data_manual
            'deep_original': 'deep_processing_data',
            'saleable': 'saleable_data'
        }
        
        # 获取实际的数据键名
        actual_key = data_mapping.get(data_type, data_type + '_data')
        data = app_data.get_data(actual_key)
        
        if data is not None and hasattr(data, 'empty') and not data.empty:
            return jsonify(app_data.safe_json_convert(data))
        else:
            return jsonify([])
            
    except Exception as e:
        print(f"获取数据失败: {str(e)}")
        return jsonify([])

@main_bp.route('/favicon.ico')
def favicon():
    """网站图标"""
    try:
        # 尝试发送static目录中的favicon.ico
        static_folder = current_app.static_folder
        if static_folder and os.path.exists(os.path.join(static_folder, 'favicon.ico')):
            return send_from_directory(static_folder, 'favicon.ico')
        else:
            # 如果没有找到，返回204 No Content
            return '', 204
    except Exception:
        return '', 204 

@main_bp.route('/auto_process', methods=['POST'])
@login_required
def auto_process_compatibility():
    """自动处理 - 兼容性路由，重定向到新的API"""
    from app.services.calculation_service import CalculationService
    from flask import session
    
    try:
        session_id = session.get('session_id')
        calculation_service = CalculationService(session_id)
        success, message = calculation_service.auto_process_all()
        
        return jsonify({
            'success': success,
            'message': message
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'自动处理失败: {str(e)}'
        }), 500

@main_bp.route('/export')
@login_required
def export_data():
    """导出数据为Excel文件，与GUI版本格式保持一致"""
    import io
    import pandas as pd
    from datetime import datetime
    from app.api.data_management_api import calculate_disassembly_product_output_value_data
    
    try:
        app_data_manager = get_session_data_manager()
        
        # 检查是否有数据
        extracted_data = app_data_manager.get_data('extracted_data')
        if extracted_data is None or extracted_data.empty:
            return jsonify({'success': False, 'message': '没有数据可导出'})

        # 创建Excel写入器
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 导入样式模块
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # 定义统一的样式
            header_font = Font(bold=True, color="FFFFFF", name="仿宋")
            data_font = Font(name="仿宋")
            bold_data_font = Font(bold=True, name="仿宋")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            left_alignment = Alignment(horizontal="left", vertical="center")
            data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            def apply_sheet_formatting(worksheet, data_df, is_statistics=False):
                """应用工作表格式"""
                # 设置列宽
                if is_statistics:
                    # 统计表格列宽
                    worksheet.column_dimensions['A'].width = 25
                    worksheet.column_dimensions['B'].width = 20
                else:
                    # 数据表格列宽
                    for col in range(1, len(data_df.columns) + 1):
                        col_letter = get_column_letter(col)
                        worksheet.column_dimensions[col_letter].width = 12
                
                # 应用表头样式
                for col in range(1, len(data_df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                
                # 设置数据行样式
                for row in range(2, len(data_df) + 2):
                    for col in range(1, len(data_df.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        
                        if is_statistics:
                            # 统计表格样式
                            if col == 1:
                                # 第一列：统计项目
                                cell.alignment = left_alignment
                                if cell.value and not str(cell.value).startswith('  ') and cell.value != '':
                                    cell.font = bold_data_font
                                else:
                                    cell.font = data_font
                            else:
                                # 第二列：数值
                                cell.font = data_font
                                cell.alignment = center_alignment
                        else:
                            # 数据表格样式
                            cell.font = data_font
                            cell.alignment = data_alignment
                
                # 设置行高
                if not is_statistics:
                    for row in range(1, len(data_df) + 2):
                        worksheet.row_dimensions[row].height = 40

            # 1. 提取结果
            if extracted_data is not None and not extracted_data.empty:
                extracted_data.to_excel(writer, sheet_name='提取结果', index=False)
                worksheet = writer.sheets['提取结果']
                apply_sheet_formatting(worksheet, extracted_data)
                
                # 2. 统计数据
                stats_data = create_extraction_statistics(extracted_data)
                stats_df = pd.DataFrame(stats_data)
                stats_df.to_excel(writer, sheet_name='统计数据', index=False)
                stats_worksheet = writer.sheets['统计数据']
                apply_sheet_formatting(stats_worksheet, stats_df, is_statistics=True)
            
            # 1.5. 提取结果(手工) - 存在手工表即导出（含仅初始化未改的情形）
            extracted_data_manual = app_data_manager.get_data('extracted_data_manual')
            if extracted_data_manual is not None and not extracted_data_manual.empty:
                extracted_data_manual.to_excel(writer, sheet_name='提取结果(手工)', index=False)
                worksheet = writer.sheets['提取结果(手工)']
                apply_sheet_formatting(worksheet, extracted_data_manual)
                
                # 提取结果(手工)统计
                manual_stats_data = create_extraction_statistics(extracted_data_manual)
                manual_stats_df = pd.DataFrame(manual_stats_data)
                manual_stats_df.to_excel(writer, sheet_name='提取结果(手工)统计', index=False)
                stats_worksheet = writer.sheets['提取结果(手工)统计']
                apply_sheet_formatting(stats_worksheet, manual_stats_df, is_statistics=True)

            # 3. 原始数据(未减扣)
            from app.utils.deducted_disassembly_align import (
                align_deducted_inventory_tai_from_disassembly,
            )

            disassembly_data = app_data_manager.get_data('disassembly_data')
            if disassembly_data is not None and not disassembly_data.empty:
                disassembly_data.to_excel(writer, sheet_name='原始数据(未减扣)', index=False)
                worksheet = writer.sheets['原始数据(未减扣)']
                apply_sheet_formatting(worksheet, disassembly_data)

            # 4. 被减扣数据 - 导出未修改前的数据；已修改时用原始备份，未修改时用手工数据
            if app_data_manager.get_data('deducted_data_modified'):
                original_data = app_data_manager.get_data('original_deducted_data')
                deducted_data = original_data if (original_data is not None and not original_data.empty) else app_data_manager.get_data('deducted_data_manual')
            else:
                deducted_data = app_data_manager.get_data('deducted_data_manual')
            if deducted_data is None or deducted_data.empty:
                original_data = app_data_manager.get_data('original_deducted_data')
                if original_data is not None and not original_data.empty:
                    deducted_data = original_data
            if deducted_data is not None and not deducted_data.empty:
                if disassembly_data is not None and not disassembly_data.empty:
                    deducted_data = align_deducted_inventory_tai_from_disassembly(
                        deducted_data,
                        disassembly_data,
                        recalculate_kg=True,
                        recalculate_kg_when_tai_changed=False,
                    )
                deducted_data.to_excel(writer, sheet_name='被减扣数据', index=False)
                worksheet = writer.sheets['被减扣数据']
                apply_sheet_formatting(worksheet, deducted_data)
                
                # 5. 被减扣统计
                deducted_stats = create_deducted_statistics(deducted_data)
                deducted_stats_df = pd.DataFrame(deducted_stats)
                deducted_stats_df.to_excel(writer, sheet_name='被减扣统计', index=False)
                stats_worksheet = writer.sheets['被减扣统计']
                apply_sheet_formatting(stats_worksheet, deducted_stats_df, is_statistics=True)

            # 5.5. 被减扣数据(手工) - 如果存在手工修改的数据
            deducted_data_manual = app_data_manager.get_data('deducted_data_manual')
            if deducted_data_manual is not None and not deducted_data_manual.empty:
                # 检查是否真的有手工修改过的数据
                if app_data_manager.get_data('deducted_data_modified'):
                    _man = deducted_data_manual
                    if disassembly_data is not None and not disassembly_data.empty:
                        _man = align_deducted_inventory_tai_from_disassembly(
                            deducted_data_manual,
                            disassembly_data,
                            recalculate_kg=False,
                            recalculate_kg_when_tai_changed=True,
                        )
                    _man.to_excel(writer, sheet_name='被减扣数据(手工)', index=False)
                    worksheet = writer.sheets['被减扣数据(手工)']
                    apply_sheet_formatting(worksheet, _man)
                    
                    # 5.6. 被减扣数据(手工)统计
                    deducted_manual_stats = create_deducted_statistics(_man)
                    deducted_manual_stats_df = pd.DataFrame(deducted_manual_stats)
                    deducted_manual_stats_df.to_excel(writer, sheet_name='被减扣数据(手工)统计', index=False)
                    stats_worksheet = writer.sheets['被减扣数据(手工)统计']
                    apply_sheet_formatting(stats_worksheet, deducted_manual_stats_df, is_statistics=True)

            # 7. 减扣后数据
            calculated_data = app_data_manager.get_data('calculated_data')
            if calculated_data is not None and not calculated_data.empty:
                calculated_data.to_excel(writer, sheet_name='减扣后数据', index=False)
                worksheet = writer.sheets['减扣后数据']
                apply_sheet_formatting(worksheet, calculated_data)
                
                # 8. 计算统计
                calc_stats = create_calculation_statistics(calculated_data)
                calc_stats_df = pd.DataFrame(calc_stats)
                calc_stats_df.to_excel(writer, sheet_name='计算统计', index=False)
                stats_worksheet = writer.sheets['计算统计']
                apply_sheet_formatting(stats_worksheet, calc_stats_df, is_statistics=True)

            # 9. 深加工数据
            deep_processing_data = app_data_manager.get_data('deep_processing_data')
            if deep_processing_data is not None and not deep_processing_data.empty:
                deep_processing_data.to_excel(writer, sheet_name='深加工数据', index=False)
                worksheet = writer.sheets['深加工数据']
                apply_sheet_formatting(worksheet, deep_processing_data)

            # 10. 可销售量
            saleable_data = app_data_manager.get_data('saleable_data')
            if saleable_data is not None and not saleable_data.empty:
                saleable_data.to_excel(writer, sheet_name='可销售量', index=False)
                worksheet = writer.sheets['可销售量']
                apply_sheet_formatting(worksheet, saleable_data)
                
                # 11. 可销售量统计
                saleable_stats = create_saleable_statistics(saleable_data)
                saleable_stats_df = pd.DataFrame(saleable_stats)
                saleable_stats_df.to_excel(writer, sheet_name='可销售量统计', index=False)
                stats_worksheet = writer.sheets['可销售量统计']
                apply_sheet_formatting(stats_worksheet, saleable_stats_df, is_statistics=True)

            # 12. 可销售量(手工) - 如果存在手工修改的数据
            saleable_data_manual = app_data_manager.get_data('saleable_data_manual')
            if saleable_data_manual is not None and not saleable_data_manual.empty:
                # 检查是否真的有手工修改过的数据
                saleable_data_modified = app_data_manager.get_data('saleable_data_modified')
                print(f"📊 检查手工数据: manual_data存在={saleable_data_manual is not None}, 非空={not saleable_data_manual.empty if saleable_data_manual is not None else False}, 已修改={saleable_data_modified}")
                
                if saleable_data_modified:
                    print(f"📝 导出手工可销售量数据: {len(saleable_data_manual)} 条记录")
                    
                    # 确保手工数据的列顺序与可销售量数据一致
                    if saleable_data is not None and not saleable_data.empty:
                        # 使用可销售量数据的列顺序重新排列手工数据
                        manual_data_reordered = saleable_data_manual.reindex(columns=saleable_data.columns)
                        manual_data_reordered.to_excel(writer, sheet_name='可销售量(手工)', index=False)
                        print(f"✅ 手工数据列顺序已调整为与可销售量数据一致: {list(saleable_data.columns)}")
                        worksheet = writer.sheets['可销售量(手工)']
                        apply_sheet_formatting(worksheet, manual_data_reordered)
                    else:
                        # 如果没有可销售量数据作为参考，则使用原始顺序
                        saleable_data_manual.to_excel(writer, sheet_name='可销售量(手工)', index=False)
                        print("⚠️ 未找到可销售量数据作为列顺序参考，使用原始顺序")
                        worksheet = writer.sheets['可销售量(手工)']
                        apply_sheet_formatting(worksheet, saleable_data_manual)
                    
                    # 13. 可销售量(手工)统计
                    print("📊 生成手工可销售量统计...")
                    saleable_manual_stats = create_saleable_statistics(saleable_data_manual)
                    saleable_manual_stats_df = pd.DataFrame(saleable_manual_stats)
                    saleable_manual_stats_df.to_excel(writer, sheet_name='可销售量(手工)统计', index=False)
                    stats_worksheet = writer.sheets['可销售量(手工)统计']
                    apply_sheet_formatting(stats_worksheet, saleable_manual_stats_df, is_statistics=True)
                    print("✅ 手工可销售量数据导出完成")

            # 14. 销售收益 - 优先使用手工数据，如果不存在则使用系统数据
            revenue_data = None
            revenue_data_source = None
            
            # 检查是否有手工数据且包含收益列
            if saleable_data_manual is not None and not saleable_data_manual.empty and saleable_data_modified:
                if '销售收益(元)' in saleable_data_manual.columns:
                    revenue_data = saleable_data_manual
                    revenue_data_source = '手工'
                    print("📊 使用手工可销售量数据导出销售收益...")
            
            # 如果没有手工数据，使用系统数据
            if revenue_data is None and saleable_data is not None and not saleable_data.empty:
                if '销售收益(元)' in saleable_data.columns:
                    revenue_data = saleable_data
                    revenue_data_source = '系统'
                    print("📊 使用系统可销售量数据导出销售收益...")
            
            # 导出销售收益sheet
            if revenue_data is not None:
                print(f"💰 导出销售收益数据 (数据源: {revenue_data_source})...")
                # 重命名列名用于显示（不影响原始数据）
                revenue_data_export = revenue_data.copy()
                if '销售收益(元)' in revenue_data_export.columns:
                    revenue_data_export = revenue_data_export.rename(columns={'销售收益(元)': '销售收入（元）'})
                revenue_data_export.to_excel(writer, sheet_name='销售收入', index=False)
                worksheet = writer.sheets['销售收入']
                apply_sheet_formatting(worksheet, revenue_data_export)
                print(f"✅ 销售收益数据导出完成 ({len(revenue_data)} 条记录)")
                
                # 15. 销售收入统计
                print("📊 生成销售收入统计...")
                revenue_stats = create_revenue_statistics(revenue_data)
                revenue_stats_df = pd.DataFrame(revenue_stats)
                revenue_stats_df.to_excel(writer, sheet_name='销售收入统计', index=False)
                stats_worksheet = writer.sheets['销售收入统计']
                apply_sheet_formatting(stats_worksheet, revenue_stats_df, is_statistics=True)
                print("✅ 销售收入统计导出完成")
            else:
                print("⚠️ 未找到包含销售收益数据的可销售量数据，跳过销售收益导出")

            # 15.5. 一次拆解产物产值
            try:
                print("📊 开始导出一拆解产物产值数据...")
                success, disassembly_product_data, error_msg = calculate_disassembly_product_output_value_data(app_data_manager)
                if success and disassembly_product_data:
                    # 创建详细数据DataFrame
                    detail_df = pd.DataFrame(disassembly_product_data)
                    # 添加序号列
                    detail_df.insert(0, '序号', range(1, len(detail_df) + 1))
                    detail_df.to_excel(writer, sheet_name='一次拆解产物产值', index=False)
                    worksheet = writer.sheets['一次拆解产物产值']
                    apply_sheet_formatting(worksheet, detail_df)
                    print(f"✅ 一次拆解产物产值数据导出完成 ({len(detail_df)} 条记录)")
                else:
                    print(f"⚠️ 一次拆解产物产值数据为空: {error_msg if error_msg else '无数据'}")
            except Exception as e:
                print(f"⚠️ 导出一拆解产物产值数据失败: {str(e)}")

            # 15.6. 深加工拆解产物产值
            try:
                print("📊 开始导出深加工拆解产物产值数据...")
                # 直接调用GET接口的逻辑获取数据
                deep_processing_data = app_data_manager.get_data('deep_processing_data')
                if deep_processing_data is not None and not deep_processing_data.empty:
                    # 筛选是否减扣 == '否'的记录
                    if '是否减扣' in deep_processing_data.columns:
                        non_deducted_data = deep_processing_data[deep_processing_data['是否减扣'] == '否'].copy()
                        if not non_deducted_data.empty:
                            # 获取价格数据
                            from data.base_data.price_data import load_price_data
                            price_df = load_price_data()
                            if price_df is not None and not price_df.empty:
                                # 创建价格映射
                                price_mapping = {}
                                for _, row in price_df.iterrows():
                                    code = str(row['拆解产物编码']).strip()
                                    price_no_tax = row.get('销售单价-不含税(元/KG)', 0)
                                    if pd.notna(price_no_tax):
                                        price_mapping[code] = float(price_no_tax)
                                
                                # 获取映射表数据
                                from data.base_data.mapping_data import MAPPING_TABLE_DATA as MAPPING_DATA
                                r3_to_category = {}
                                for mapping_item in MAPPING_DATA:
                                    r3_code = str(mapping_item.get('R3系统代码', '')).strip()
                                    category = str(mapping_item.get('类别', '')).strip()
                                    if r3_code and category:
                                        r3_to_category[r3_code] = category
                                
                                # 分类关键词映射规则（四机一脑类别）
                                category_keyword_mapping = {
                                    '电视': ['电视', '彩电', 'CRT其它机壳破碎塑料', '线路板边框破碎塑料', '等离子', '废旧玻璃电子枪', '废旧金属荫罩压块铁', '黑白'],
                                    '电脑': ['电脑', '显示器', '笔记本', '主机', '废旧金属黑色金属-铁及其合金-电子枪'],
                                    '冰箱': ['冰箱', '冰柜'],
                                    '空调': ['空调'],
                                    '洗衣机': ['洗衣机', '双缸']
                                }
                                
                                # 处理数据
                                result_rows = []
                                for idx, row in non_deducted_data.iterrows():
                                    material_name = str(row.get('原物料名称', '')).strip()
                                    
                                    # 四机一脑类别匹配
                                    four_category = None
                                    for cat, keywords in category_keyword_mapping.items():
                                        for keyword in keywords:
                                            if keyword in material_name:
                                                four_category = cat
                                                break
                                        if four_category:
                                            break
                                    
                                    if not four_category:
                                        continue
                                    
                                    deep_product_code = str(row.get('深加工产物编码', '')).strip()
                                    deep_result_kg = row.get('深加工结果(KG)', 0)
                                    try:
                                        deep_result_kg = float(deep_result_kg) if pd.notna(deep_result_kg) else 0
                                    except (ValueError, TypeError):
                                        deep_result_kg = 0
                                    
                                    price_no_tax = price_mapping.get(deep_product_code, 0)
                                    mapping_category = r3_to_category.get(deep_product_code, '')
                                    
                                    if price_no_tax < 0:
                                        material_value = deep_result_kg * 0
                                    else:
                                        material_value = deep_result_kg * price_no_tax
                                    
                                    result_rows.append({
                                        '原物料代码': str(row.get('原物料代码', '')).strip(),
                                        '原物料名称': material_name,
                                        '四机一脑类别': four_category,
                                        '类别': mapping_category,
                                        '一次拆解产物编码': str(row.get('一次拆解产物编码', '')).strip(),
                                        '一次拆解产物名称': str(row.get('一次拆解产物名称', '')).strip(),
                                        '深加工产物编码': deep_product_code,
                                        '深加工产物名称': str(row.get('深加工产物名称', '')).strip(),
                                        '深加工结果(KG)': deep_result_kg,
                                        '销售单价-不含税(元/KG)': price_no_tax,
                                        '物料产值（元）': material_value
                                    })
                                
                                if result_rows:
                                    result_df = pd.DataFrame(result_rows)
                                    result_df.to_excel(writer, sheet_name='深加工拆解产物产值', index=False)
                                    worksheet = writer.sheets['深加工拆解产物产值']
                                    apply_sheet_formatting(worksheet, result_df)
                                    print(f"✅ 深加工拆解产物产值数据导出完成 ({len(result_df)} 条记录)")
                                else:
                                    print("⚠️ 深加工拆解产物产值数据为空")
                            else:
                                print("⚠️ 暂无销售价格数据，跳过深加工拆解产物产值导出")
                        else:
                            print("⚠️ 暂无非减扣的深加工产物数据，跳过深加工拆解产物产值导出")
                    else:
                        print("⚠️ 深加工数据中缺少'是否减扣'列，跳过深加工拆解产物产值导出")
                else:
                    print("⚠️ 暂无深加工数据，跳过深加工拆解产物产值导出")
            except Exception as e:
                print(f"⚠️ 导出深加工拆解产物产值数据失败: {str(e)}")

            # 16. 基金补贴收入
            subsidy_income_data = app_data_manager.get_data('subsidy_income_data')
            if subsidy_income_data is not None and not subsidy_income_data.empty:
                print(f"💰 导出基金补贴收入数据...")
                subsidy_income_data.to_excel(writer, sheet_name='基金补贴收入', index=False)
                worksheet = writer.sheets['基金补贴收入']
                apply_sheet_formatting(worksheet, subsidy_income_data)
                print(f"✅ 基金补贴收入数据导出完成 ({len(subsidy_income_data)} 条记录)")
                
                # 17. 基金补贴收入统计
                print("📊 生成基金补贴收入统计...")
                subsidy_stats = create_subsidy_statistics(subsidy_income_data)
                subsidy_stats_df = pd.DataFrame(subsidy_stats)
                subsidy_stats_df.to_excel(writer, sheet_name='基金补贴收入统计', index=False)
                stats_worksheet = writer.sheets['基金补贴收入统计']
                apply_sheet_formatting(stats_worksheet, subsidy_stats_df, is_statistics=True)
                print("✅ 基金补贴收入统计导出完成")
            else:
                print("⚠️ 未找到基金补贴收入数据，跳过基金补贴收入导出")
            
            # 18. 总收益统计（包含销售收益和基金补贴收入）
            print("📊 生成总收益统计...")
            total_revenue_stats = create_total_revenue_statistics(revenue_data, subsidy_income_data)
            total_revenue_stats_df = pd.DataFrame(total_revenue_stats)
            total_revenue_stats_df.to_excel(writer, sheet_name='总收益统计', index=False)
            stats_worksheet = writer.sheets['总收益统计']
            apply_sheet_formatting(stats_worksheet, total_revenue_stats_df, is_statistics=True)
            print("✅ 总收益统计导出完成")

        output.seek(0)
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"revenue_forecast_{timestamp}.xlsx"
        
        # 发送Excel文件
        from flask import Response
        response = Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        return response
        
    except Exception as e:
        print(f"✗ 导出失败: {e}")
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}'})

def create_extraction_statistics(data):
    """创建提取数据的统计信息"""
    if data is None or data.empty:
        return [{'统计项目': '无数据', '数值': 0}]
    
    stats = []
    
    # 总记录数
    stats.append({'统计项目': '总记录数', '数值': len(data)})
    stats.append({'统计项目': None, '数值': None})  # 空行
    
    # 本期实际投产数量统计
    if '非限制使用的库存' in data.columns:
        stats.append({'统计项目': '本期实际投产数量统计', '数值': None})
        
        # 转换为数值类型
        import pandas as pd
        numeric_stock = pd.to_numeric(data['非限制使用的库存'], errors='coerce')
        valid_stock = numeric_stock.dropna()
        
        if not valid_stock.empty:
            stats.append({'统计项目': '总重量', '数值': valid_stock.sum()})
            stats.append({'统计项目': '平均重量', '数值': valid_stock.mean()})
            stats.append({'统计项目': '最大重量', '数值': valid_stock.max()})
            stats.append({'统计项目': '最小重量', '数值': valid_stock.min()})
            stats.append({'统计项目': '非零记录数', '数值': len(valid_stock[valid_stock > 0])})
    
    return stats

def create_deducted_statistics(data):
    """创建被减扣数据的统计信息"""
    if data is None or data.empty:
        return [{'统计项目': '无数据', '数值': 0}]
    
    stats = []
    
    # 被减扣记录总数
    stats.append({'统计项目': '被减扣记录总数', '数值': len(data)})
    stats.append({'统计项目': None, '数值': None})  # 空行
    
    # 被减扣重量统计
    stats.append({'统计项目': '被减扣重量统计', '数值': None})
    
    # 计算结果统计
    if '计算结果(KG)' in data.columns:
        import pandas as pd
        numeric_weight = pd.to_numeric(data['计算结果(KG)'], errors='coerce')
        valid_weight = numeric_weight.dropna()
        
        if not valid_weight.empty:
            stats.append({'统计项目': '总减扣重量(KG)', '数值': valid_weight.sum()})
            stats.append({'统计项目': '平均重量(KG)', '数值': valid_weight.mean()})
            stats.append({'统计项目': '最大重量(KG)', '数值': valid_weight.max()})
            stats.append({'统计项目': '最小重量(KG)', '数值': valid_weight.min()})
            stats.append({'统计项目': '非零记录数', '数值': len(valid_weight[valid_weight > 0])})
            stats.append({'统计项目': '数值型记录数', '数值': len(valid_weight)})
            stats.append({'统计项目': '非数值型记录数', '数值': len(data) - len(valid_weight)})
    
    return stats

def create_calculation_statistics(data):
    """创建计算数据的统计信息"""
    if data is None or data.empty:
        return [{'统计项目': '无数据', '数值': 0}]
    
    stats = []
    
    # 总计算记录数
    stats.append({'统计项目': '总计算记录数', '数值': len(data)})
    stats.append({'统计项目': None, '数值': None})  # 空行
    
    # 计算结果统计
    stats.append({'统计项目': '计算结果统计', '数值': None})
    
    if '计算结果(KG)' in data.columns:
        import pandas as pd
        numeric_weight = pd.to_numeric(data['计算结果(KG)'], errors='coerce')
        valid_weight = numeric_weight.dropna()
        
        if not valid_weight.empty:
            stats.append({'统计项目': '总重量(KG)', '数值': valid_weight.sum()})
            stats.append({'统计项目': '平均重量(KG)', '数值': valid_weight.mean()})
            stats.append({'统计项目': '最大重量(KG)', '数值': valid_weight.max()})
            stats.append({'统计项目': '最小重量(KG)', '数值': valid_weight.min()})
            stats.append({'统计项目': '非零记录数', '数值': len(valid_weight[valid_weight > 0])})
    
    return stats

def create_saleable_statistics(data):
    """创建可销售量的统计信息"""
    if data is None or data.empty:
        return [{'统计项目': '无数据', '数值': 0}]
    
    stats = []
    
    # 总记录数
    stats.append({'统计项目': '总记录数', '数值': len(data)})
    stats.append({'统计项目': None, '数值': None})  # 空行
    
    # 按类别统计
    stats.append({'统计项目': '按类别统计', '数值': None})
    
    if '类别' in data.columns and '计算结果(KG)' in data.columns:
        # 按类别分组统计
        import pandas as pd
        category_stats = data.groupby('类别').agg({
            '计算结果(KG)': ['sum', 'count']
        }).round(6)
        
        for category in category_stats.index:
            weight_sum = category_stats.loc[category, ('计算结果(KG)', 'sum')]
            count = category_stats.loc[category, ('计算结果(KG)', 'count')]
            
            # 确保weight_sum是数值类型，处理NaN情况
            try:
                if pd.isna(weight_sum):
                    weight_sum = 0.0
                else:
                    weight_sum = float(weight_sum)
                count = int(count) if not pd.isna(count) else 0
                
                stats.append({
                    '统计项目': str(category),
                    '数值': f"{weight_sum:.6f} KG ({count}条)"
                })
            except (ValueError, TypeError):
                stats.append({
                    '统计项目': str(category),
                    '数值': f"0.000000 KG ({count}条)"
                })
    
    # 总重量
    if '计算结果(KG)' in data.columns:
        import pandas as pd
        numeric_weight = pd.to_numeric(data['计算结果(KG)'], errors='coerce')
        valid_weight = numeric_weight.dropna()
        
        if not valid_weight.empty:
            total_weight = valid_weight.sum()
            stats.append({'统计项目': None, '数值': None})  # 空行
            stats.append({'统计项目': '总重量(KG)', '数值': f"{total_weight:.6f}"})
        else:
            stats.append({'统计项目': None, '数值': None})  # 空行
            stats.append({'统计项目': '总重量(KG)', '数值': "0.000000"})
    
    return stats

def create_revenue_statistics(data):
    """创建销售收益的统计信息"""
    if data is None or data.empty:
        return [{'统计项目': '无数据', '数值': 0}]
    
    # 检查是否包含必要的列
    if '销售收益(元)' not in data.columns:
        return [{'统计项目': '无收入数据', '数值': 0}]
    
    stats = []
    import pandas as pd
    
    # 总记录数
    stats.append({'统计项目': '总记录数', '数值': len(data)})
    stats.append({'统计项目': None, '数值': None})  # 空行
    
    # 按类别统计
    stats.append({'统计项目': '按类别收入统计', '数值': None})
    
    if '类别' in data.columns:
        # 准备数值列
        numeric_weight = pd.to_numeric(data['计算结果(KG)'], errors='coerce') if '计算结果(KG)' in data.columns else pd.Series([0] * len(data))
        numeric_revenue = pd.to_numeric(data['销售收益(元)'], errors='coerce')
        numeric_price = pd.to_numeric(data['销售单价(元/KG)'], errors='coerce') if '销售单价(元/KG)' in data.columns else pd.Series([0] * len(data))
        
        # 创建临时DataFrame用于统计
        temp_df = pd.DataFrame({
            '类别': data['类别'],
            '重量': numeric_weight.fillna(0),
            '收益': numeric_revenue.fillna(0),
            '单价': numeric_price.fillna(0)
        })
        
        # 按类别分组统计
        category_stats = temp_df.groupby('类别').agg({
            '重量': 'sum',
            '收益': 'sum',
            '单价': 'mean'
        })
        
        # 计算每个类别的记录数
        category_counts = temp_df.groupby('类别').size()
        
        for category in category_stats.index:
            weight_sum = category_stats.loc[category, '重量']
            revenue_sum = category_stats.loc[category, '收益']
            avg_price = category_stats.loc[category, '单价']
            count = category_counts[category]
            
            # 格式化输出
            stats.append({
                '统计项目': f"  {category}",
                '数值': None
            })
            stats.append({
                '统计项目': f"    记录数",
                '数值': f"{count}条"
            })
            stats.append({
                '统计项目': f"    总重量(KG)",
                '数值': f"{weight_sum:.2f}"
            })
            stats.append({
                '统计项目': f"    总收入(元)",
                '数值': f"{revenue_sum:.2f}"
            })
            stats.append({
                '统计项目': f"    平均单价(元/KG)",
                '数值': f"{avg_price:.2f}"
            })
    
    # 总计
    stats.append({'统计项目': None, '数值': None})  # 空行
    stats.append({'统计项目': '总计', '数值': None})
    
    if '计算结果(KG)' in data.columns:
        numeric_weight = pd.to_numeric(data['计算结果(KG)'], errors='coerce')
        total_weight = numeric_weight.fillna(0).sum()
        stats.append({'统计项目': '  总重量(KG)', '数值': f"{total_weight:.2f}"})
    
    numeric_revenue = pd.to_numeric(data['销售收益(元)'], errors='coerce')
    total_revenue = numeric_revenue.fillna(0).sum()
    stats.append({'统计项目': '  总收入(元)', '数值': f"{total_revenue:.2f}"})
    
    if '销售单价(元/KG)' in data.columns and len(data) > 0:
        numeric_price = pd.to_numeric(data['销售单价(元/KG)'], errors='coerce')
        valid_prices = numeric_price[numeric_price > 0]
        if not valid_prices.empty:
            avg_price = valid_prices.mean()
            stats.append({'统计项目': '  平均单价(元/KG)', '数值': f"{avg_price:.2f}"})
    
    return stats


def create_subsidy_statistics(data):
    """创建基金补贴收入统计信息"""
    if data is None or data.empty:
        return [{'统计项目': '无数据', '数值': 0}]
    
    stats = []
    import pandas as pd
    
    # 总记录数
    stats.append({'统计项目': '总记录数', '数值': len(data)})
    stats.append({'统计项目': None, '数值': None})  # 空行
    
    # 按补贴大类分组统计
    if '补贴大类' in data.columns:
        stats.append({'统计项目': '按类别统计', '数值': None})
        category_groups = data.groupby('补贴大类')
        
        for category, group in category_groups:
            total_quantity = pd.to_numeric(group['当期拆解量(台)'], errors='coerce').sum()
            total_subsidy = pd.to_numeric(group['基金补贴收入(元)'], errors='coerce').sum()
            avg_price = pd.to_numeric(group['补贴单价(元/台)'], errors='coerce').mean() if '补贴单价(元/台)' in group.columns else 0
            
            stats.append({'统计项目': f'  {category}', '数值': None})
            stats.append({'统计项目': f'    拆解量(台)', '数值': f"{total_quantity:.2f}"})
            stats.append({'统计项目': f'    补贴收入(元)', '数值': f"{total_subsidy:.2f}"})
            stats.append({'统计项目': f'    平均单价(元/台)', '数值': f"{avg_price:.2f}"})
        
        stats.append({'统计项目': None, '数值': None})  # 空行
    
    # 总计
    stats.append({'统计项目': '总计', '数值': None})
    
    # 总拆解量统计
    if '当期拆解量(台)' in data.columns:
        numeric_quantity = pd.to_numeric(data['当期拆解量(台)'], errors='coerce')
        total_quantity = numeric_quantity.fillna(0).sum()
        stats.append({'统计项目': '  总拆解量(台)', '数值': f"{total_quantity:.2f}"})
    
    # 基金补贴收入统计
    if '基金补贴收入(元)' in data.columns:
        numeric_subsidy = pd.to_numeric(data['基金补贴收入(元)'], errors='coerce')
        total_subsidy = numeric_subsidy.fillna(0).sum()
        stats.append({'统计项目': '  总补贴收入(元)', '数值': f"{total_subsidy:.2f}"})
        
        avg_subsidy = numeric_subsidy[numeric_subsidy > 0].mean() if len(numeric_subsidy[numeric_subsidy > 0]) > 0 else 0
        stats.append({'统计项目': '  平均补贴收入(元)', '数值': f"{avg_subsidy:.2f}"})
    
    return stats


def create_total_revenue_statistics(saleable_data, subsidy_income_data):
    """创建总收益统计信息（销售收益 + 基金补贴收入）"""
    import pandas as pd
    
    stats = []
    
    # 计算销售收益
    total_sales_revenue = 0
    if saleable_data is not None and not saleable_data.empty and '销售收益(元)' in saleable_data.columns:
        numeric_revenue = pd.to_numeric(saleable_data['销售收益(元)'], errors='coerce')
        total_sales_revenue = numeric_revenue.fillna(0).sum()
    
    # 计算基金补贴收入
    total_subsidy = 0
    if subsidy_income_data is not None and not subsidy_income_data.empty and '基金补贴收入(元)' in subsidy_income_data.columns:
        numeric_subsidy = pd.to_numeric(subsidy_income_data['基金补贴收入(元)'], errors='coerce')
        total_subsidy = numeric_subsidy.fillna(0).sum()
    
    # 计算总收益
    total_revenue = total_sales_revenue + total_subsidy
    
    # 收益构成
    stats.append({'统计项目': '收益构成', '数值': None})
    stats.append({'统计项目': '  销售收益（拆解产物）', '数值': f"{total_sales_revenue:.2f} 元"})
    stats.append({'统计项目': '  基金补贴收入（旧机）', '数值': f"{total_subsidy:.2f} 元"})
    stats.append({'统计项目': None, '数值': None})  # 空行
    
    # 总收益统计
    stats.append({'统计项目': '总收益汇总', '数值': None})
    stats.append({'统计项目': '  总收益合计', '数值': f"{total_revenue:.2f} 元"})
    stats.append({'统计项目': None, '数值': None})  # 空行
    
    # 比例分析
    if total_revenue > 0:
        sales_percentage = (total_sales_revenue / total_revenue * 100)
        subsidy_percentage = (total_subsidy / total_revenue * 100)
        stats.append({'统计项目': '收益比例分析', '数值': None})
        stats.append({'统计项目': '  销售收益占比', '数值': f"{sales_percentage:.2f}%"})
        stats.append({'统计项目': '  基金补贴收入占比', '数值': f"{subsidy_percentage:.2f}%"})
    
    return stats

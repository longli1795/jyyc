from flask import Blueprint, jsonify
import math
import numbers
import pandas as pd
import json
from app.models.app_data import AppDataManager
from app.models.compatibility import AppDataManagerAdapter
from app.utils.auth_utils import page_permission_required

def get_session_data_manager():
    """获取会话数据管理器的便利函数"""
    from flask import session
    session_id = session.get('session_id')
    return AppDataManagerAdapter.get_instance(session_id)

statistics_bp = Blueprint('statistics', __name__)

@statistics_bp.route('/<data_type>', methods=['GET'])
def get_category_statistics(data_type):
    """获取指定数据类型的按类别重量统计"""
    try:
        app_data = get_session_data_manager()
        
        # 数据类型映射
        # 🔧 架构重构：deducted 节点直接使用 deducted_data_manual，不再使用 deducted_data (只读)
        data_mapping = {
            'extracted': 'extracted_data',
            'extracted_manual': 'extracted_data_manual',  # 手工提取数据
            'original': 'disassembly_data',
            'final': 'calculated_data',
            'deducted': 'deducted_data_manual',  # 被减扣数据(手工) - 唯一工作数据源
            'deducted_manual': 'deducted_data_manual',    # 修改后的被减扣数据
            'deep_original': 'deep_processing_data',
            'saleable': 'saleable_data',
            'saleable_manual': 'saleable_data_manual'  # 手工可销售量数据
        }
        
        # 🔧 架构重构：如果 deducted_data_manual 为空且未修改，使用 original_deducted_data 作为显示源
        if data_type == 'deducted':
            deducted_data_manual = app_data.get_data('deducted_data_manual')
            deducted_data_modified = app_data.get_data('deducted_data_modified')
            
            if (deducted_data_manual is None or deducted_data_manual.empty) and not deducted_data_modified:
                # 如果手工数据为空且未修改，使用原始备份作为显示源
                original_data = app_data.get_data('original_deducted_data')
                if original_data is not None and not original_data.empty:
                    data_mapping['deducted'] = 'original_deducted_data'
                    print(f"ℹ️ 被减扣数据统计使用原始备份数据: {len(original_data)} 条记录")
                else:
                    print(f"⚠️ 被减扣数据统计：没有可用数据")
            else:
                print(f"✅ 被减扣数据统计使用手工数据: {len(deducted_data_manual) if deducted_data_manual is not None and not deducted_data_manual.empty else 0} 条记录")
        
        # 重量列映射
        weight_column_mapping = {
            'extracted': '非限制使用的库存',
            'extracted_manual': '非限制使用的库存',  # 手工提取数据
            'original': '计算结果(KG)',
            'final': '计算结果(KG)',
            'deducted': '计算结果(KG)',
            'deducted_manual': '计算结果(KG)',
            'deep_original': '深加工结果(KG)',
            'saleable': '计算结果(KG)',
            'saleable_manual': '计算结果(KG)'  # 手工可销售量数据
        }
        
        actual_key = data_mapping.get(data_type)
        weight_column = weight_column_mapping.get(data_type)
        
        if not actual_key or not weight_column:
            return jsonify({
                'success': False,
                'error': f'不支持的数据类型: {data_type}'
            })
        
        data = app_data.get_data(actual_key)
        
        if data is None or data.empty:
            return jsonify({
                'success': True,
                'total_records': 0,
                'total_weight': 0,
                'categories': []
            })
        
        # 检查必要的列是否存在
        if '类别' not in data.columns or weight_column not in data.columns:
            return jsonify({
                'success': True,
                'total_records': len(data),
                'total_weight': 0,
                'categories': []
            })
        
        # 按类别统计重量
        categories = []
        total_weight = 0
        
        # 转换重量数据为数值类型
        data_copy = data.copy()
        data_copy['weight_numeric'] = pd.to_numeric(data_copy[weight_column], errors='coerce')
        data_copy = data_copy.dropna(subset=['weight_numeric'])
        
        if not data_copy.empty:
            category_stats = data_copy.groupby('类别').agg({
                'weight_numeric': ['sum', 'count']
            }).round(6)
            
            # 计算总重量
            total_weight = float(data_copy['weight_numeric'].sum())
            
            # 构建类别统计结果
            for category in category_stats.index:
                weight_sum = float(category_stats.loc[category, ('weight_numeric', 'sum')])
                count = int(category_stats.loc[category, ('weight_numeric', 'count')])
                percentage = (weight_sum / total_weight * 100) if total_weight > 0 else 0
                
                categories.append({
                    'name': str(category),
                    'weight': float(round(weight_sum, 6)),
                    'count': int(count),
                    'percentage': float(round(percentage, 2))
                })
            
            # 按重量排序
            categories.sort(key=lambda x: x['weight'], reverse=True)
        
        return jsonify({
            'success': True,
            'total_records': int(len(data)),
            'total_weight': float(round(total_weight, 6)),
            'categories': categories
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'统计数据获取失败: {str(e)}'
        })

@statistics_bp.route('/summary', methods=['GET'])
def get_summary():
    """获取所有数据节点的统计汇总"""
    try:
        from app.models.app_data import AppDataManager
        
        app_data = get_session_data_manager()
        
        # 检查数据是否已被清除
        data_cleared = app_data.get_data('__data_cleared__')
        if data_cleared:
            # 数据已被清除，直接返回空统计
            return jsonify({
                'success': True,
                'summary': {
                    'total_revenue': {
                        'sales_revenue': 0.0,
                        'subsidy_income': 0.0,
                        'total': 0.0,
                        'sales_percentage': 0.0,
                        'subsidy_percentage': 0.0
                    }
                }
            })
        
        # 基础数据类型
        data_types = ['extracted', 'original', 'final', 'deducted', 'deep_original', 'saleable']
        
        # 如果提取数据被修改过，则增加手工节点
        if app_data.get_data('extracted_data_modified'):
            data_types.insert(1, 'extracted_manual')  # 在extracted后面插入
        
        # 如果被减扣数据被修改过，则增加手工节点
        if app_data.get_data('deducted_data_modified'):
            data_types.insert(data_types.index('deducted') + 1 if 'deducted' in data_types else 4, 'deducted_manual')
        
        # 如果可销售量数据被修改过，则增加手工节点
        if app_data.get_data('saleable_data_modified'):
            data_types.append('saleable_manual')  # 在saleable后面添加
        
        summary = {}
        
        for data_type in data_types:
            # 调用单个统计函数
            try:
                stats_data = get_category_statistics(data_type)
                if hasattr(stats_data, 'get_json'):
                    stats_json = stats_data.get_json()
                else:
                    stats_json = stats_data
                
                if stats_json.get('success'):
                    summary[data_type] = {
                        'total_records': int(stats_json.get('total_records', 0)),
                        'total_weight': float(stats_json.get('total_weight', 0)),
                        'categories': stats_json.get('categories', [])
                    }
                else:
                    summary[data_type] = {
                        'total_records': 0,
                        'total_weight': 0.0,
                        'categories': []
                    }
            except Exception as e:
                print(f"获取{data_type}统计失败: {str(e)}")
                summary[data_type] = {
                    'total_records': 0,
                    'total_weight': 0.0,
                    'categories': []
                }
        
        # 添加基金补贴收入统计
        try:
            subsidy_income_data = app_data.get_data('subsidy_income_data')
            if subsidy_income_data is not None and not subsidy_income_data.empty:
                # 按类别统计补贴收入
                subsidy_categories = []
                total_subsidy = 0
                
                if '补贴大类' in subsidy_income_data.columns and '基金补贴收入(元)' in subsidy_income_data.columns:
                    category_groups = subsidy_income_data.groupby('补贴大类').agg({
                        '基金补贴收入(元)': 'sum',
                        '当期拆解量(台)': 'sum'
                    })
                    
                    total_subsidy = float(subsidy_income_data['基金补贴收入(元)'].sum())
                    
                    for category in category_groups.index:
                        subsidy_sum = float(category_groups.loc[category, '基金补贴收入(元)'])
                        quantity_sum = float(category_groups.loc[category, '当期拆解量(台)'])
                        percentage = (subsidy_sum / total_subsidy * 100) if total_subsidy > 0 else 0
                        
                        subsidy_categories.append({
                            'name': str(category),
                            'subsidy': round(subsidy_sum, 2),
                            'quantity': round(quantity_sum, 2),
                            'percentage': round(percentage, 2)
                        })
                    
                    # 按补贴金额排序
                    subsidy_categories.sort(key=lambda x: x['subsidy'], reverse=True)
                
                summary['subsidy_income'] = {
                    'total_records': int(len(subsidy_income_data)),
                    'total_subsidy': round(total_subsidy, 2),
                    'categories': subsidy_categories
                }
            else:
                summary['subsidy_income'] = {
                    'total_records': 0,
                    'total_subsidy': 0.0,
                    'categories': []
                }
        except Exception as e:
            print(f"获取补贴收入统计失败: {str(e)}")
            summary['subsidy_income'] = {
                'total_records': 0,
                'total_subsidy': 0.0,
                'categories': []
            }
        
        # 添加总收益统计
        try:
            # 销售收益从可销售量数据中获取
            revenue_data = None
            if app_data.get_data('saleable_data_modified'):
                revenue_data = app_data.get_data('saleable_data_manual')
            if revenue_data is None or revenue_data.empty:
                revenue_data = app_data.get_data('saleable_data')
            
            total_sales_revenue = 0.0
            if revenue_data is not None and not revenue_data.empty and '销售收益(元)' in revenue_data.columns:
                total_sales_revenue = float(pd.to_numeric(revenue_data['销售收益(元)'], errors='coerce').fillna(0).sum())
            
            # 基金补贴收入从独立的 subsidy_income_data 中获取
            subsidy_income_data = app_data.get_data('subsidy_income_data')
            total_subsidy_income = 0.0
            if subsidy_income_data is not None and not subsidy_income_data.empty and '基金补贴收入(元)' in subsidy_income_data.columns:
                total_subsidy_income = float(pd.to_numeric(subsidy_income_data['基金补贴收入(元)'], errors='coerce').fillna(0).sum())
            
            # 总收益 = 销售收益 + 基金补贴收入
            total_revenue = float(total_sales_revenue + total_subsidy_income)
            
            summary['total_revenue'] = {
                'sales_revenue': round(total_sales_revenue, 2),
                'subsidy_income': round(total_subsidy_income, 2),
                'total': round(total_revenue, 2),
                'sales_percentage': round((total_sales_revenue / total_revenue * 100) if total_revenue > 0 else 0, 2),
                'subsidy_percentage': round((total_subsidy_income / total_revenue * 100) if total_revenue > 0 else 0, 2)
            }
            
        except Exception as e:
            print(f"获取总收益统计失败: {str(e)}")
            import traceback
            traceback.print_exc()
            summary['total_revenue'] = {
                'sales_revenue': 0.0,
                'subsidy_income': 0.0,
                'total': 0.0,
                'sales_percentage': 0.0,
                'subsidy_percentage': 0.0
            }
        
        return jsonify({
            'success': True,
            'summary': summary
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'汇总统计获取失败: {str(e)}'
        })

@statistics_bp.route('/deducted-comparison', methods=['GET'])
def get_deducted_comparison_stats():
    """获取被减扣数据对比统计"""
    try:
        app_data = get_session_data_manager()
        stats = app_data.get_deducted_comparison_stats()
        
        return jsonify({
            'success': True,
            'data': stats
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'获取被减扣数据对比统计失败: {str(e)}'
        })

@statistics_bp.route('/export', methods=['GET'])
def export_data():
    """导出数据 - 占位符"""
    return jsonify({'success': True, 'message': '数据导出功能待实现'})


FOUR_MACHINE_CATEGORIES = ['冰箱', '空调', '电脑', '电视', '洗衣机']


def map_material_to_four_category(material_name):
    """映射物料名称到四机一脑分类（与利润测算汇总表一致）"""
    if pd.isna(material_name) or not material_name:
        return None
    name = str(material_name)
    if ('CRT其它机壳破碎塑料' in name or '线路板边框破碎塑料' in name or
            '废旧玻璃电子枪' in name or '废旧金属荫罩压块铁' in name or
            '黑白' in name or '电视' in name or '彩电' in name or '等离子' in name):
        return '电视'
    if ('电脑' in name or '显示器' in name or '笔记本' in name or
            '主机' in name or '废旧金属黑色金属-铁及其合金-电子枪' in name):
        return '电脑'
    if '冰箱' in name or '冰柜' in name:
        return '冰箱'
    if '空调' in name:
        return '空调'
    if '洗衣机' in name or '双缸' in name:
        return '洗衣机'
    return None


def get_disassembly_output_value_ratios(app_data, categories=None):
    """按一次拆解产物产值计算各分类物料产值占比（与期间费用分摊一致）"""
    import traceback
    from app.api.data_management_api import calculate_disassembly_product_output_value_data

    categories = categories or FOUR_MACHINE_CATEGORIES
    category_values = {cat: 0.0 for cat in categories}
    total_value = 0.0

    try:
        success, output_value_data, _ = calculate_disassembly_product_output_value_data(app_data)
        if success and output_value_data:
            for item in output_value_data:
                category = item.get('分类', '')
                material_value = float(item.get('物料产值（元）', 0) or 0)
                if category and category in categories:
                    category_values[category] += material_value
                    total_value += material_value
    except Exception as e:
        print(f"获取产值分摊比例失败: {e}")
        traceback.print_exc()

    ratios = {
        cat: (category_values[cat] / total_value if total_value > 0 else 0.0)
        for cat in categories
    }
    return ratios, category_values, total_value


def get_category_operating_revenue(app_data, categories=None):
    """各分类营业收入 = 产物销售收入 + 基金补贴收入（与利润测算汇总表一致）"""
    import traceback

    categories = categories or FOUR_MACHINE_CATEGORIES
    product_sales_revenue = {cat: 0.0 for cat in categories}
    subsidy_income = {cat: 0.0 for cat in categories}

    try:
        saleable_data_manual = app_data.get_data('saleable_data_manual')
        saleable_data = app_data.get_data('saleable_data')
        saleable_data_modified = app_data.get_data('saleable_data_modified')

        revenue_data = None
        if saleable_data_manual is not None and not saleable_data_manual.empty and saleable_data_modified:
            if '销售收益(元)' in saleable_data_manual.columns:
                revenue_data = saleable_data_manual
        if revenue_data is None and saleable_data is not None and not saleable_data.empty:
            if '销售收益(元)' in saleable_data.columns:
                revenue_data = saleable_data

        if revenue_data is not None and not revenue_data.empty:
            if '销售收益(元)' in revenue_data.columns:
                revenue_data = revenue_data.copy()
                revenue_data['销售收益(元)'] = pd.to_numeric(
                    revenue_data['销售收益(元)'], errors='coerce'
                ).fillna(0)
            if '原物料名称' in revenue_data.columns:
                revenue_data['分类'] = revenue_data['原物料名称'].apply(map_material_to_four_category)
                for category in categories:
                    mask = revenue_data['分类'] == category
                    product_sales_revenue[category] = float(
                        revenue_data.loc[mask, '销售收益(元)'].sum()
                    )
    except Exception as e:
        print(f"获取产物销售收入失败: {e}")
        traceback.print_exc()

    try:
        subsidy_income_data = app_data.get_data('subsidy_income_data')
        if subsidy_income_data is not None and not subsidy_income_data.empty:
            if '补贴大类' in subsidy_income_data.columns and '基金补贴收入(元)' in subsidy_income_data.columns:
                subsidy_income_data = subsidy_income_data.copy()
                subsidy_income_data['基金补贴收入(元)'] = pd.to_numeric(
                    subsidy_income_data['基金补贴收入(元)'], errors='coerce'
                ).fillna(0)

                def map_subsidy_category_to_product_type(subsidy_category):
                    subsidy_category = str(subsidy_category).strip()
                    if subsidy_category in ['整机', '内机', '外机']:
                        return '空调'
                    if subsidy_category in ['笔记本', '显示器', '主机']:
                        return '电脑'
                    if subsidy_category in ['电视机', '冰箱', '洗衣机']:
                        return subsidy_category
                    return subsidy_category

                subsidy_income_data['产品类型'] = subsidy_income_data['补贴大类'].apply(
                    map_subsidy_category_to_product_type
                )
                product_type_groups = subsidy_income_data.groupby('产品类型')['基金补贴收入(元)'].sum()
                for product_type, total_subsidy in product_type_groups.items():
                    product_type = str(product_type).strip()
                    if product_type == '电视机':
                        product_type = '电视'
                    if product_type in categories:
                        subsidy_income[product_type] += float(total_subsidy)
    except Exception as e:
        print(f"获取基金补贴收入失败: {e}")
        traceback.print_exc()

    operating_revenue = {
        cat: product_sales_revenue[cat] + subsidy_income[cat]
        for cat in categories
    }
    return operating_revenue, product_sales_revenue, subsidy_income


PROFIT_SUMMARY_CATEGORIES = ['冰箱', '空调', '电脑', '电视', '洗衣机']


def _empty_profit_summary_data():
    return {
        key: {cat: 0.0 for cat in PROFIT_SUMMARY_CATEGORIES}
        for key in (
            'product_sales_revenue', 'subsidy_income', 'product_sales_cost',
            'subsidy_cost', 'period_cost', 'tax_surcharge'
        )
    }


def _fetch_profit_summary_data(
    app_data,
    prediction_period=1,
    quality_manager_ratio=None,
    quality_group_ratio=None,
    warehouse_group_ratio=None,
):
    """获取利润测算汇总表原始数据（供 API 与导出复用）"""
    from app.api.cost_forecast_api import (
        calculate_disassembly_product_cost,
        calculate_deep_processing_product_cost,
        calculate_period_cost,
        calculate_tax_surcharge,
    )
    import traceback

    if app_data.get_data('__data_cleared__'):
        return _empty_profit_summary_data()

    categories = PROFIT_SUMMARY_CATEGORIES

    # 分类映射函数
    def map_category_py(material_name):
        """映射物料名称到四机一脑分类（Python版本）"""
        if pd.isna(material_name) or not material_name:
            return None
        name = str(material_name)

        if ('CRT其它机壳破碎塑料' in name or '线路板边框破碎塑料' in name or
            '废旧玻璃电子枪' in name or '废旧金属荫罩压块铁' in name or
            '黑白' in name or '电视' in name or '彩电' in name or '等离子' in name):
            return '电视'

        if ('电脑' in name or '显示器' in name or '笔记本' in name or
            '主机' in name or '废旧金属黑色金属-铁及其合金-电子枪' in name):
            return '电脑'

        if '冰箱' in name or '冰柜' in name:
            return '冰箱'

        if '空调' in name:
            return '空调'

        if '洗衣机' in name or '双缸' in name:
            return '洗衣机'

        return None

    # 初始化结果
    result = {
        'product_sales_revenue': {cat: 0.0 for cat in categories},
        'subsidy_income': {cat: 0.0 for cat in categories},
        'product_sales_cost': {cat: 0.0 for cat in categories},
        'subsidy_cost': {cat: 0.0 for cat in categories},
        'period_cost': {cat: 0.0 for cat in categories},
        'tax_surcharge': {cat: 0.0 for cat in categories},
    }

    # ========== 1.1 产物销售收入 ==========
    try:
        # 优先使用手工数据，否则使用系统数据（与销售收益页逻辑一致）
        saleable_data_manual = app_data.get_data('saleable_data_manual')
        saleable_data = app_data.get_data('saleable_data')
        saleable_data_modified = app_data.get_data('saleable_data_modified')
        
        revenue_data = None
        # 检查是否有手工数据且包含收益列，并且已标记为修改（与导出结果逻辑一致）
        if saleable_data_manual is not None and not saleable_data_manual.empty and saleable_data_modified:
            if '销售收益(元)' in saleable_data_manual.columns:
                revenue_data = saleable_data_manual
        # 如果没有手工数据或未修改，使用系统数据
        if revenue_data is None and saleable_data is not None and not saleable_data.empty:
            if '销售收益(元)' in saleable_data.columns:
                revenue_data = saleable_data
        
        if revenue_data is not None and not revenue_data.empty:
            # 确保数值类型
            if '销售收益(元)' in revenue_data.columns:
                revenue_data['销售收益(元)'] = pd.to_numeric(
                    revenue_data['销售收益(元)'], errors='coerce'
                ).fillna(0)
            
            # 按分类汇总销售收益（使用外部定义的map_category_py函数）
            if '原物料名称' in revenue_data.columns:
                revenue_data['分类'] = revenue_data['原物料名称'].apply(map_category_py)
                for category in categories:
                    mask = revenue_data['分类'] == category
                    result['product_sales_revenue'][category] = float(
                        revenue_data.loc[mask, '销售收益(元)'].sum()
                    )
    except Exception as e:
        print(f"获取产物销售收入失败: {str(e)}")
        traceback.print_exc()
    
    # ========== 1.2 基金补贴收入 ==========
    try:
        subsidy_income_data = app_data.get_data('subsidy_income_data')
        if subsidy_income_data is not None and not subsidy_income_data.empty:
            if '补贴大类' in subsidy_income_data.columns and '基金补贴收入(元)' in subsidy_income_data.columns:
                subsidy_income_data['基金补贴收入(元)'] = pd.to_numeric(
                    subsidy_income_data['基金补贴收入(元)'], errors='coerce'
                ).fillna(0)
                
                # 将补贴大类映射到产品类型（五大类：电视机、冰箱、洗衣机、空调、电脑）
                def map_subsidy_category_to_product_type(subsidy_category):
                    """将补贴大类映射到产品类型"""
                    subsidy_category = str(subsidy_category).strip()
                    # 空调的细分补贴大类
                    if subsidy_category in ['整机', '内机', '外机']:
                        return '空调'
                    # 电脑的细分补贴大类
                    elif subsidy_category in ['笔记本', '显示器', '主机']:
                        return '电脑'
                    # 其他直接返回（电视机、冰箱、洗衣机）
                    elif subsidy_category in ['电视机', '冰箱', '洗衣机']:
                        return subsidy_category
                    else:
                        # 默认返回原值（兼容旧数据）
                        return subsidy_category
                
                # 添加产品类型列
                subsidy_income_data['产品类型'] = subsidy_income_data['补贴大类'].apply(map_subsidy_category_to_product_type)
                
                # 按产品类型分组汇总
                product_type_groups = subsidy_income_data.groupby('产品类型')['基金补贴收入(元)'].sum()
                
                # 映射产品类型到四机一脑分类
                for product_type, total_subsidy in product_type_groups.items():
                    product_type = str(product_type).strip()
                    # 映射：电视机 -> 电视
                    if product_type == '电视机':
                        product_type = '电视'
                    
                    if product_type in categories:
                        result['subsidy_income'][product_type] += float(total_subsidy)
    except Exception as e:
        print(f"获取基金补贴收入失败: {str(e)}")
        traceback.print_exc()
    
    # ========== 2.1 产物销售成本 ==========
    try:
        # 获取一次拆解产物成本数据
        disassembly_cost_data = calculate_disassembly_product_cost(app_data, prediction_period)
        if disassembly_cost_data:
            for row in disassembly_cost_data:
                category = row.get('分类', '')
                sales_cost = float(row.get('一次拆解产物销售成本', 0) or 0)
                if category in categories:
                    result['product_sales_cost'][category] += sales_cost
        
        # 获取深加工产物成本数据
        deep_processing_cost_data = calculate_deep_processing_product_cost(app_data, prediction_period)
        if deep_processing_cost_data:
            for row in deep_processing_cost_data:
                category = row.get('四机一脑类别', '')
                sales_cost = float(row.get('深加工产物销售成本', 0) or 0)
                if category in categories:
                    result['product_sales_cost'][category] += sales_cost
    except Exception as e:
        print(f"获取产物销售成本失败: {str(e)}")
        traceback.print_exc()
    
    # ========== 2.2 基金补贴成本 ==========
    try:
        # 读取缓存的生产成本分摊结果
        cache_key = f'production_cost_allocation_result_v2_{prediction_period}'
        cached_result = app_data.get_data(cache_key)
        
        if cached_result is not None and len(cached_result) > 0:
            for row in cached_result:
                category = row.get('产线', '')
                subsidy_allocation_cost = float(row.get('基金补贴收入分摊成本', 0) or 0)
                if category in categories:
                    result['subsidy_cost'][category] += subsidy_allocation_cost
    except Exception as e:
        print(f"获取基金补贴成本失败: {str(e)}")
        traceback.print_exc()
    
    # ========== 5. 期间费用 ==========
    try:
        # 获取期间费用总额（直接调用计算函数）
        from app.api.cost_forecast_api import calculate_period_cost
        
        # 使用传入的分摊比例（如果为None，函数内部会使用默认值）
        period_cost_result = calculate_period_cost(
            app_data, 
            prediction_period,
            quality_manager_ratio,
            quality_group_ratio,
            warehouse_group_ratio
        )
        
        period_cost_total = 0.0
        if period_cost_result.get('success'):
            period_cost_total = float(period_cost_result.get('total_cost', 0) or 0)
        
        # 获取一次拆解产物产值数据用于计算分摊比例
        # 使用与期间费用页面相同的API逻辑，确保数据一致性
        from app.api.data_management_api import calculate_disassembly_product_output_value_data
        
        success, output_value_data, error_message = calculate_disassembly_product_output_value_data(app_data)
        
        if success and output_value_data:
            # 计算各分类的产值和占比
            category_values = {cat: 0.0 for cat in categories}
            total_value = 0.0
            
            for item in output_value_data:
                category = item.get('分类', '')
                material_value = float(item.get('物料产值（元）', 0) or 0)
                
                if category and category in categories:
                    category_values[category] += material_value
                    total_value += material_value
            
            # 计算各分类的期间费用分摊值
            if total_value > 0:
                for category in categories:
                    ratio = category_values[category] / total_value
                    result['period_cost'][category] = period_cost_total * ratio
    except Exception as e:
        print(f"获取期间费用失败: {str(e)}")
        traceback.print_exc()
    
    # ========== 6. 税金及附加 ==========
    try:
        from app.api.cost_forecast_api import calculate_tax_surcharge
        
        tax_result = calculate_tax_surcharge(app_data, prediction_period)
        if tax_result.get('success'):
            for row in tax_result.get('rows', []):
                if row.get('项目') == '合计':
                    for category in categories:
                        result['tax_surcharge'][category] = float(
                            row.get('values', {}).get(category, 0) or 0
                        )
                    break
    except Exception as e:
        print(f"获取税金及附加失败: {str(e)}")
        traceback.print_exc()
    
    # 四舍五入所有数值到2位小数
    for key in result:
        for category in categories:
            result[key][category] = round(result[key][category], 2)

    return result


def _build_profit_summary_export_df(data):
    """将利润测算汇总表原始数据转为与页面一致的导出表格"""
    categories = PROFIT_SUMMARY_CATEGORIES
    columns = ['项目'] + categories + ['小计']

    def gv(key, cat):
        return float((data.get(key) or {}).get(cat, 0) or 0)

    calc = {}
    for cat in categories:
        revenue = gv('product_sales_revenue', cat) + gv('subsidy_income', cat)
        cost = gv('product_sales_cost', cat) + gv('subsidy_cost', cat)
        gross_profit = revenue - cost
        margin = (gross_profit / revenue * 100) if revenue > 0 else 0.0
        operating_profit = gross_profit - gv('period_cost', cat) - gv('tax_surcharge', cat)
        calc[cat] = {
            'revenue': revenue,
            'cost': cost,
            'gross_profit': gross_profit,
            'margin': margin,
            'operating_profit': operating_profit,
        }

    def make_row(label, value_fn, is_percent=False):
        row = {'项目': label}
        if is_percent:
            total_revenue = sum(calc[c]['revenue'] for c in categories)
            total_gross = sum(calc[c]['gross_profit'] for c in categories)
            for cat in categories:
                row[cat] = f"{round(value_fn(cat), 2)}%"
            row['小计'] = (
                f"{round(total_gross / total_revenue * 100, 2)}%"
                if total_revenue > 0 else '0.00%'
            )
        else:
            total = 0.0
            for cat in categories:
                val = round(float(value_fn(cat) or 0), 2)
                row[cat] = val
                total += val
            row['小计'] = round(total, 2)
        return row

    rows = [
        make_row('1.营业收入', lambda c: calc[c]['revenue']),
        make_row('  1.1产物销售收入', lambda c: gv('product_sales_revenue', c)),
        make_row('  1.2基金补贴收入', lambda c: gv('subsidy_income', c)),
        make_row('2.营业成本', lambda c: calc[c]['cost']),
        make_row('  2.1产物销售成本', lambda c: gv('product_sales_cost', c)),
        make_row('  2.2基金补贴成本', lambda c: gv('subsidy_cost', c)),
        make_row('3.项目毛利', lambda c: calc[c]['gross_profit']),
        make_row('4.项目毛利率', lambda c: calc[c]['margin'], is_percent=True),
        make_row('5.期间费用', lambda c: gv('period_cost', c)),
        make_row('6.税金及附加', lambda c: gv('tax_surcharge', c)),
        make_row('7.营业利润', lambda c: calc[c]['operating_profit']),
    ]
    return pd.DataFrame(rows, columns=columns)


def _parse_profit_summary_request_args():
    """解析利润测算汇总表请求参数"""
    from flask import request

    prediction_period = int(request.args.get('prediction_period', 1))
    quality_manager_ratio = request.args.get('quality_manager_ratio')
    quality_group_ratio = request.args.get('quality_group_ratio')
    warehouse_group_ratio = request.args.get('warehouse_group_ratio')

    if quality_manager_ratio is not None:
        quality_manager_ratio = float(quality_manager_ratio)
    if quality_group_ratio is not None:
        quality_group_ratio = float(quality_group_ratio)
    if warehouse_group_ratio is not None:
        warehouse_group_ratio = float(warehouse_group_ratio)

    return prediction_period, quality_manager_ratio, quality_group_ratio, warehouse_group_ratio


@statistics_bp.route('/profit-summary', methods=['GET'])
@page_permission_required
def get_profit_summary():
    """获取利润测算汇总表数据"""
    try:
        app_data = get_session_data_manager()
        prediction_period, qm, qg, wh = _parse_profit_summary_request_args()
        result = _fetch_profit_summary_data(
            app_data, prediction_period, qm, qg, wh
        )
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        print(f"获取利润测算汇总表数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@statistics_bp.route('/profit-summary/export', methods=['GET'])
@page_permission_required
def export_profit_summary():
    """导出利润测算汇总表到 Excel"""
    try:
        from flask import send_file
        import io
        from datetime import datetime
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        app_data = get_session_data_manager()
        prediction_period, qm, qg, wh = _parse_profit_summary_request_args()
        data = _fetch_profit_summary_data(app_data, prediction_period, qm, qg, wh)

        export_df = _build_profit_summary_export_df(data)
        if export_df.empty:
            return jsonify({'success': False, 'error': '没有可导出的数据'}), 400

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            export_df.to_excel(writer, sheet_name='利润测算汇总表', index=False)
            worksheet = writer.sheets['利润测算汇总表']

            header_font = Font(bold=True, color='FFFFFF', name='仿宋')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            center_alignment = Alignment(horizontal='center', vertical='center')
            right_alignment = Alignment(horizontal='right', vertical='center')

            for col in range(1, len(export_df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                col_letter = get_column_letter(col)
                col_name = export_df.columns[col - 1]
                if col_name == '项目':
                    worksheet.column_dimensions[col_letter].width = 22
                else:
                    worksheet.column_dimensions[col_letter].width = 16

            for row_idx in range(2, len(export_df) + 2):
                label_cell = worksheet.cell(row=row_idx, column=1)
                label_cell.alignment = Alignment(horizontal='left', vertical='center')
                for col_idx in range(2, len(export_df.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).alignment = right_alignment

        output.seek(0)
        filename = f'利润测算汇总表_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        print(f"导出利润测算汇总表失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@statistics_bp.route('/analyze-profit-summary', methods=['POST'])
@page_permission_required
def analyze_profit_summary():
    """使用AI分析利润测算汇总表数据"""
    try:
        from flask import request
        from app.services.ai_analysis_service import AIAnalysisService
        import traceback
        
        # 获取请求数据
        request_data = request.get_json()
        if not request_data or 'data' not in request_data:
            return jsonify({
                'success': False,
                'error': '缺少利润测算汇总表数据'
            }), 400
        
        profit_data = request_data['data']
        
        # 验证数据格式
        required_keys = ['product_sales_revenue', 'subsidy_income', 'product_sales_cost', 
                        'subsidy_cost', 'period_cost', 'tax_surcharge']
        for key in required_keys:
            if key not in profit_data:
                return jsonify({
                    'success': False,
                    'error': f'缺少必要的数据字段: {key}'
                }), 400
        
        # 调用AI分析服务
        ai_service = AIAnalysisService()
        success, analysis_result, error_message = ai_service.analyze_profit_summary(profit_data)
        
        if success:
            return jsonify({
                'success': True,
                'analysis': analysis_result
            })
        else:
            return jsonify({
                'success': False,
                'error': error_message or 'AI分析失败'
            }), 500
            
    except Exception as e:
        print(f"AI分析利润测算汇总表失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'分析服务错误: {str(e)}'
        }), 500


@statistics_bp.route('/chat', methods=['POST'])
def chat_with_ai():
    """与AI进行对话（支持多轮对话）"""
    try:
        from flask import request, session
        from app.services.ai_analysis_service import AIAnalysisService
        from app.models.compatibility import AppDataManagerAdapter
        import traceback
        
        # 获取请求数据
        request_data = request.get_json()
        if not request_data or 'message' not in request_data:
            return jsonify({
                'success': False,
                'error': '缺少消息内容'
            }), 400
        
        user_message = request_data['message']
        conversation_history = request_data.get('history', [])  # 对话历史
        
        # 获取会话数据管理器
        session_id = session.get('session_id')
        app_data = AppDataManagerAdapter.get_instance(session_id) if session_id else None
        
        # 获取利润测算汇总表数据（如果前端传递了）
        profit_data = None
        if request_data.get('include_profit_data', False):
            profit_data = request_data.get('profit_data')
        
        # 自动收集所有页面数据作为上下文
        all_page_data = None
        if app_data:
            try:
                ai_service = AIAnalysisService()
                all_page_data = ai_service._collect_all_page_data(app_data)
            except Exception as e:
                print(f"收集所有页面数据失败: {str(e)}")
                import traceback
                traceback.print_exc()
                # 如果收集失败，继续使用利润数据（如果有）
        
        # 构建消息列表（包含历史对话）
        messages = []
        for msg in conversation_history:
            if isinstance(msg, dict) and 'role' in msg and 'content' in msg:
                messages.append({
                    'role': msg['role'],
                    'content': msg['content']
                })
        
        # 添加当前用户消息
        messages.append({
            'role': 'user',
            'content': user_message
        })
        
        # 调用AI对话服务
        ai_service = AIAnalysisService()
        success, ai_reply, error_message = ai_service.chat_with_ai(messages, profit_data, all_page_data)
        
        if success:
            return jsonify({
                'success': True,
                'reply': ai_reply
            })
        else:
            return jsonify({
                'success': False,
                'error': error_message or 'AI对话失败'
            }), 500
            
    except Exception as e:
        print(f"AI对话失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'对话服务错误: {str(e)}'
        }), 500


@statistics_bp.route('/chat/stream', methods=['POST'])
def chat_with_ai_stream():
    """与AI进行流式对话（Server-Sent Events）"""
    try:
        from flask import request, session, Response, stream_with_context
        from app.services.ai_analysis_service import AIAnalysisService
        from app.models.compatibility import AppDataManagerAdapter
        import traceback
        
        # 获取请求数据
        request_data = request.get_json()
        if not request_data or 'message' not in request_data:
            return jsonify({
                'success': False,
                'error': '缺少消息内容'
            }), 400
        
        user_message = request_data['message']
        conversation_history = request_data.get('history', [])
        
        # 获取会话数据管理器
        session_id = session.get('session_id')
        app_data = AppDataManagerAdapter.get_instance(session_id) if session_id else None
        
        # 获取利润测算汇总表数据（如果前端传递了）
        profit_data = None
        if request_data.get('include_profit_data', False):
            profit_data = request_data.get('profit_data')
        
        # 自动收集所有页面数据作为上下文
        all_page_data = None
        if app_data:
            try:
                ai_service = AIAnalysisService()
                all_page_data = ai_service._collect_all_page_data(app_data)
            except Exception as e:
                print(f"收集所有页面数据失败: {str(e)}")
                import traceback
                traceback.print_exc()
                # 如果收集失败，继续使用利润数据（如果有）
        
        # 构建消息列表
        messages = []
        for msg in conversation_history:
            if isinstance(msg, dict) and 'role' in msg and 'content' in msg:
                messages.append({
                    'role': msg['role'],
                    'content': msg['content']
                })
        
        # 添加当前用户消息
        messages.append({
            'role': 'user',
            'content': user_message
        })
        
        # 调用流式AI对话服务
        ai_service = AIAnalysisService()
        
        def generate():
            try:
                for content, is_done in ai_service.stream_chat_with_ai(messages, profit_data, all_page_data):
                    # 先发送content（如果有）
                    if content:
                        # 发送内容片段
                        yield f"data: {json.dumps({'content': content, 'done': False}, ensure_ascii=False)}\n\n"
                    # 然后检查是否完成
                    if is_done:
                        # 发送完成信号（确保所有content都已发送）
                        yield f"data: {json.dumps({'content': '', 'done': True}, ensure_ascii=False)}\n\n"
                        break
            except Exception as e:
                error_msg = str(e)
                yield f"data: {json.dumps({'error': error_msg, 'done': True}, ensure_ascii=False)}\n\n"
        
        return Response(
            stream_with_context(generate()),
            mimetype='text/event-stream',
            headers={
                'Cache-Control': 'no-cache',
                'X-Accel-Buffering': 'no'
            }
        )
        
    except Exception as e:
        print(f"AI流式对话失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'对话服务错误: {str(e)}'
        }), 500


@statistics_bp.route('/analyze-profit-summary/stream', methods=['POST'])
@page_permission_required
def analyze_profit_summary_stream():
    """流式分析利润测算汇总表数据（Server-Sent Events）"""
    try:
        from flask import request, Response, stream_with_context
        from app.services.ai_analysis_service import AIAnalysisService
        import traceback
        
        # 获取请求数据
        request_data = request.get_json()
        if not request_data or 'data' not in request_data:
            return jsonify({
                'success': False,
                'error': '缺少利润测算汇总表数据'
            }), 400
        
        profit_data = request_data['data']
        
        # 验证数据格式
        required_keys = ['product_sales_revenue', 'subsidy_income', 'product_sales_cost', 
                        'subsidy_cost', 'period_cost', 'tax_surcharge']
        for key in required_keys:
            if key not in profit_data:
                return jsonify({
                    'success': False,
                    'error': f'缺少必要的数据字段: {key}'
                }), 400
        
        # 调用流式AI分析服务
        ai_service = AIAnalysisService()
        
        def generate():
            try:
                for content, is_done in ai_service.stream_analyze_profit_summary(profit_data):
                    if content:
                        yield f"data: {json.dumps({'content': content, 'done': False}, ensure_ascii=False)}\n\n"
                    if is_done:
                        yield f"data: {json.dumps({'content': '', 'done': True}, ensure_ascii=False)}\n\n"
                        break
            except Exception as e:
                error_msg = str(e)
                yield f"data: {json.dumps({'error': error_msg, 'done': True}, ensure_ascii=False)}\n\n"
        
        return Response(
            stream_with_context(generate()),
            mimetype='text/event-stream',
            headers={
                'Cache-Control': 'no-cache',
                'X-Accel-Buffering': 'no'
            }
        )
        
    except Exception as e:
        print(f"AI流式分析失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'分析服务错误: {str(e)}'
        }), 500


def _calculate_disassembly_profit_analysis_data(app_data, prediction_period=1, quality_manager_ratio=None, quality_group_ratio=None, warehouse_group_ratio=None, consider_opening_stock=True):
    """计算当期拆解收益测算分析表数据（内部函数）。consider_opening_stock=False 时拆解数量仅取拆解物原料成本页本期实际投产数量（不含被减扣数据 KG 部分）；「直接人工」③ 被减扣用当前手工表；「产物价值」深加工换算仍可用 original_deducted_data 快照（deducted_data_for_product_value）。"""
    try:
        from app.api.cost_forecast_api import (
            calculate_direct_labor_cost,
            calculate_manufacturing_cost,
            calculate_period_cost,
            classify_by_product_name,
            calculate_material_cost,
            resolve_disassembly_category_unit_price,
        )
        from app.api.data_management_api import (
            calculate_disassembly_product_output_value_data,
            calculate_deep_processing_product_output_value_data
        )
        from data.base_data.price_data import load_price_data
        import traceback
        import pandas as pd
        
        # 检查数据是否已被清除
        data_cleared = app_data.get_data('__data_cleared__')
        if data_cleared:
            return jsonify({
                'success': True,
                'data': []
            })
        
        # ========== 步骤1: 获取有效物料代码清单 ==========
        extracted_data = app_data.get_data('extracted_data_manual')
        if extracted_data is None or extracted_data.empty:
            return jsonify({
                'success': True,
                'data': []
            })
        
        # 从"拆解物原料成本"页获取"本期实际投产数量"（即调用calculate_material_cost获取数据）
        # 拆解数量 = 拆解物原料成本页"本期实际投产数量"按"物料代码"统计
        cost_data = calculate_material_cost(extracted_data)
        if cost_data is None or cost_data.empty:
            return jsonify({
                'success': True,
                'data': []
            })
        
        # 筛选有效物料代码（拆解数量 > 0，类别为旧机）
        valid_material_codes = set()
        material_info = {}  # {物料代码: {物料名称, 拆解数量(台), 类别}}
        # 保存第二行数据（单位：KG），按原物料代码统计
        material_info_kg = {}  # {物料代码: {物料名称, 拆解数量(KG), 类别}}
        
        # 被减扣数据（手工）：用于后续产值/费用等计算；与是否追加 KG 拆解数量无关
        deducted_data = app_data.get_data('deducted_data_manual')
        
        # 产物价值专用被减扣数据源：
        # - 不考虑期初库存口径：按“修改前快照”计算（original_deducted_data）
        # - 快照缺失时回退到当前手工数据，保证接口可用
        deducted_data_for_product_value = deducted_data
        if not consider_opening_stock:
            original_deducted_data = app_data.get_data('original_deducted_data')
            if original_deducted_data is not None and not original_deducted_data.empty:
                deducted_data_for_product_value = original_deducted_data
        
        if '类别' in cost_data.columns and '物料代码' in cost_data.columns:
            old_machine_mask = cost_data['类别'] == '旧机'
            old_machine_data = cost_data[old_machine_mask].copy()
            
            if '非限制使用的库存' in old_machine_data.columns:
                old_machine_data['非限制使用的库存'] = pd.to_numeric(
                    old_machine_data['非限制使用的库存'], errors='coerce'
                ).fillna(0)
                
                # 筛选拆解数量 > 0 的物料代码
                valid_mask = old_machine_data['非限制使用的库存'] > 0
                valid_data = old_machine_data[valid_mask]
                
                for _, row in valid_data.iterrows():
                    material_code = str(row.get('物料代码', '')).strip()
                    if material_code:
                        valid_material_codes.add(material_code)
                        if material_code not in material_info:
                            material_info[material_code] = {
                                '物料名称': str(row.get('物料描述', '')).strip(),
                                '拆解数量': float(row.get('非限制使用的库存', 0)),  # 本期实际投产数量（单位：台）
                                '类别': classify_by_product_name(str(row.get('物料描述', ''))) or ''
                            }
        
        if not valid_material_codes:
            return jsonify({
                'success': True,
                'data': []
            })
        
        # ========== 步骤1.1: 从被减扣数据（手工）统计拆解数量（单位：KG） ==========
        # 仅当考虑期初库存和库存结余时追加 KG 维度；不考虑时拆解数量仅取拆解物原料成本页本期实际投产数量
        if consider_opening_stock:
            # "被减扣数据（手工）"sheet，筛选类别为"一次拆解产物"、处置类别为"内转屏处置、内转印制板处置、深加工-打包铁、深加工-塑料一破"，
            # 按"原物料代码"统计"计算结果"列数据（单位：KG）
            # 注意：这些"原物料代码"是独立的物料代码，不需要在valid_material_codes中，即使物料代码相同也作为独立行显示
            if deducted_data is not None and not deducted_data.empty:
                # 筛选类别为"一次拆解产物"，处置类别为指定值
                if '类别' in deducted_data.columns and '处置类别' in deducted_data.columns:
                    mask = (
                        (deducted_data['类别'] == '一次拆解产物') & 
                        (deducted_data['处置类别'].isin(['内转屏处置', '内转印制板处置', '深加工-打包铁', '深加工-塑料一破']))
                    )
                    filtered_deducted = deducted_data[mask].copy()
                    
                    if '原物料代码' in filtered_deducted.columns and '计算结果(KG)' in filtered_deducted.columns:
                        filtered_deducted['计算结果(KG)'] = pd.to_numeric(
                            filtered_deducted['计算结果(KG)'], errors='coerce'
                        ).fillna(0)
                        
                        # 按原物料代码统计计算结果（单位：KG）
                        # 这些"原物料代码"是独立的物料代码，不需要检查是否在valid_material_codes中
                        for _, row in filtered_deducted.iterrows():
                            material_code = str(row.get('原物料代码', '')).strip()
                            if material_code:  # 移除检查 material_code in valid_material_codes
                                qty = float(row.get('计算结果(KG)', 0))
                                if qty > 0:
                                    # 如果物料代码已在material_info_kg中，累加拆解数量（KG）
                                    if material_code in material_info_kg:
                                        material_info_kg[material_code]['拆解数量'] += qty
                                    else:
                                        # 创建独立的物料代码行（单位：KG）
                                        material_info_kg[material_code] = {
                                            '物料名称': str(row.get('原物料名称', '')).strip(),
                                            '拆解数量': qty,  # 单位：KG
                                            '类别': classify_by_product_name(str(row.get('原物料名称', ''))) or ''
                                        }
        
        # ========== 步骤2: 计算各项指标 ==========
        result_data = []
        allowed_material_codes = valid_material_codes if not consider_opening_stock else (valid_material_codes | set(material_info_kg.keys()))
        
        # 2.1 基金补贴收入
        subsidy_income_by_code = {}
        subsidy_income_data = app_data.get_data('subsidy_income_data')
        if subsidy_income_data is not None and not subsidy_income_data.empty:
            if '物料代码' in subsidy_income_data.columns and '基金补贴收入(元)' in subsidy_income_data.columns:
                subsidy_income_data['基金补贴收入(元)'] = pd.to_numeric(
                    subsidy_income_data['基金补贴收入(元)'], errors='coerce'
                ).fillna(0)
                
                for _, row in subsidy_income_data.iterrows():
                    material_code = str(row.get('物料代码', '')).strip()
                    if material_code in valid_material_codes:
                        income = float(row.get('基金补贴收入(元)', 0))
                        subsidy_income_by_code[material_code] = subsidy_income_by_code.get(material_code, 0) + income
        
        # 2.2 产物价值计算
        original_data = app_data.get_data('disassembly_data')
        deep_processing_data = app_data.get_data('deep_processing_data')
        price_df = load_price_data()
        price_mapping = {}
        if price_df is not None and not price_df.empty:
            # 使用拆解产物编码匹配价格（与销售价格管理页一致）
            if '拆解产物编码' in price_df.columns and '销售单价-不含税(元/KG)' in price_df.columns:
                for _, row in price_df.iterrows():
                    code = str(row.get('拆解产物编码', '')).strip()
                    price = float(row.get('销售单价-不含税(元/KG)', 0)) if pd.notna(row.get('销售单价-不含税(元/KG)')) else 0
                    price_mapping[code] = price
        
        product_value_by_code = {}
        if consider_opening_stock:
            # ---------- 考虑期初库存和库存结余：原有口径 ----------
            # 2.2.1 一次拆解产物产值（全部拆解产物）
            primary_product_value_count = 0
            primary_product_value_total = 0.0
            if original_data is not None and not original_data.empty:
                if '类别' in original_data.columns and '原物料代码' in original_data.columns:
                    product_mask = (original_data['类别'] == '拆解产物')
                    product_data = original_data[product_mask].copy()
                    if not product_data.empty and '原物料代码' in product_data.columns and '拆解产物编码' in product_data.columns and '计算结果(KG)' in product_data.columns:
                        product_data['计算结果(KG)'] = pd.to_numeric(
                            product_data['计算结果(KG)'], errors='coerce'
                        ).fillna(0)
                        for _, row in product_data.iterrows():
                            material_code = str(row.get('原物料代码', '')).strip()
                            if material_code and material_code in allowed_material_codes:
                                product_code = str(row.get('拆解产物编码', '')).strip()
                                weight = float(row.get('计算结果(KG)', 0))
                                if weight > 0:
                                    price = price_mapping.get(product_code, 0)
                                    if price < 0:
                                        price = 0
                                    value = weight * price
                                    product_value_by_code[material_code] = product_value_by_code.get(material_code, 0) + value
                                    primary_product_value_count += 1
                                    primary_product_value_total += value
            print(f"一次拆解产物产值计算完成：处理了 {primary_product_value_count} 条记录，总产值 {primary_product_value_total:.2f} 元")
            # 2.2.1.1 减去被减扣数据中类别为"拆解产物"的产值
            deducted_product_value_count = 0
            deducted_product_value_total = 0.0
            if deducted_data is not None and not deducted_data.empty:
                if '类别' in deducted_data.columns:
                    deducted_product_mask = (deducted_data['类别'] == '拆解产物')
                    deducted_product_data = deducted_data[deducted_product_mask].copy()
                    if not deducted_product_data.empty:
                        if '原物料代码' in deducted_product_data.columns and '拆解产物编码' in deducted_product_data.columns and '计算结果(KG)' in deducted_product_data.columns:
                            deducted_product_data['计算结果(KG)'] = pd.to_numeric(
                                deducted_product_data['计算结果(KG)'], errors='coerce'
                            ).fillna(0)
                            for _, row in deducted_product_data.iterrows():
                                material_code = str(row.get('原物料代码', '')).strip()
                                if material_code and material_code in allowed_material_codes:
                                    product_code = str(row.get('拆解产物编码', '')).strip()
                                    weight = float(row.get('计算结果(KG)', 0))
                                    if weight > 0:
                                        price = price_mapping.get(product_code, 0)
                                        if price < 0:
                                            price = 0
                                        value = weight * price
                                        product_value_by_code[material_code] = product_value_by_code.get(material_code, 0) - value
                                        deducted_product_value_count += 1
                                        deducted_product_value_total += value
            print(f"被减扣拆解产物产值计算完成：处理了 {deducted_product_value_count} 条记录，减扣总产值 {deducted_product_value_total:.2f} 元")
            # 2.2.2 深加工产物产值（深加工拆解产物产值页）
            deep_processing_value_count = 0
            deep_processing_value_total = 0.0
            deep_success, deep_result_data, deep_message = calculate_deep_processing_product_output_value_data(app_data)
            if not deep_success:
                print(f"获取深加工拆解产物产值数据失败: {deep_message}")
            elif deep_result_data:
                for row in deep_result_data:
                    material_code = str(row.get('原物料代码', '')).strip()
                    if material_code and material_code in allowed_material_codes:
                        material_value = row.get('物料产值（元）', 0)
                        try:
                            material_value = float(material_value) if pd.notna(material_value) else 0
                        except (ValueError, TypeError):
                            material_value = 0
                        if material_value != 0:
                            product_value_by_code[material_code] = product_value_by_code.get(material_code, 0) + material_value
                            deep_processing_value_count += 1
                            deep_processing_value_total += material_value
            print(f"深加工产物产值计算完成：处理了 {deep_processing_value_count} 条记录，总产值 {deep_processing_value_total:.2f} 元")
        else:
            # ---------- 不考虑期初库存和库存结余：直接销售 + 深加工换算 ----------
            # 一次拆解产物直接销售类：原始数据(未减扣)，类别=拆解产物，处置类别=空白
            direct_count = 0
            direct_total = 0.0
            if original_data is not None and not original_data.empty:
                if '类别' in original_data.columns and '原物料代码' in original_data.columns and '处置类别' in original_data.columns:
                    product_mask = (original_data['类别'] == '拆解产物')
                    disp_mask = (original_data['处置类别'].astype(str).str.strip() == '空白') | (original_data['处置类别'].astype(str).str.strip() == '')
                    product_data = original_data[product_mask & disp_mask].copy()
                    if not product_data.empty and '原物料代码' in product_data.columns and '拆解产物编码' in product_data.columns and '计算结果(KG)' in product_data.columns:
                        product_data['计算结果(KG)'] = pd.to_numeric(product_data['计算结果(KG)'], errors='coerce').fillna(0)
                        for _, row in product_data.iterrows():
                            material_code = str(row.get('原物料代码', '')).strip()
                            if material_code and material_code in allowed_material_codes:
                                product_code = str(row.get('拆解产物编码', '')).strip()
                                weight = float(row.get('计算结果(KG)', 0))
                                if weight > 0:
                                    price = price_mapping.get(product_code, 0)
                                    if price < 0:
                                        price = 0
                                    value = weight * price
                                    product_value_by_code[material_code] = product_value_by_code.get(material_code, 0) + value
                                    direct_count += 1
                                    direct_total += value
            print(f"不考虑期初库存-直接销售类产值：处理了 {direct_count} 条记录，总产值 {direct_total:.2f} 元")
            # 深加工产物价值：被减扣数据 类别=拆解产物，处置类别 in (内转屏处置, 内转印制板处置, 深加工-打包铁, 深加工-塑料一破)，用深加工系数表换算
            from data.base_data.deep_processing_data import DEEP_PROCESSING_DATA
            coeff_by_first = {}
            for r in DEEP_PROCESSING_DATA:
                first = str(r.get('拆解产物编码', '')).strip()
                if first:
                    coeff_by_first.setdefault(first, []).append(r)
            deep_deducted_disposal = ['内转屏处置', '内转印制板处置', '深加工-打包铁', '深加工-塑料一破']
            deep_count = 0
            deep_total = 0.0
            if deducted_data_for_product_value is not None and not deducted_data_for_product_value.empty:
                if '类别' in deducted_data_for_product_value.columns and '处置类别' in deducted_data_for_product_value.columns and '原物料代码' in deducted_data_for_product_value.columns and '拆解产物编码' in deducted_data_for_product_value.columns and '计算结果(KG)' in deducted_data_for_product_value.columns:
                    mask = (deducted_data_for_product_value['类别'] == '拆解产物') & (deducted_data_for_product_value['处置类别'].astype(str).str.strip().isin(deep_deducted_disposal))
                    filtered = deducted_data_for_product_value[mask].copy()
                    filtered['计算结果(KG)'] = pd.to_numeric(filtered['计算结果(KG)'], errors='coerce').fillna(0)
                    for _, row in filtered.iterrows():
                        material_code = str(row.get('原物料代码', '')).strip()
                        if not material_code or material_code not in allowed_material_codes:
                            continue
                        product_code = str(row.get('拆解产物编码', '')).strip()
                        weight = float(row.get('计算结果(KG)', 0))
                        if weight <= 0:
                            continue
                        coeff_list = coeff_by_first.get(product_code, [])
                        for coeff in coeff_list:
                            io_ratio = float(coeff.get('深加工投入产出比例', 0) or 0)
                            coef = float(coeff.get('深加工拆解系数', 0) or 0)
                            deep_product_code = str(coeff.get('深加工产物编码', '')).strip()
                            deep_weight = weight * io_ratio * coef
                            if deep_weight <= 0:
                                continue
                            price = price_mapping.get(deep_product_code, 0)
                            if price < 0:
                                price = 0
                            value = deep_weight * price
                            product_value_by_code[material_code] = product_value_by_code.get(material_code, 0) + value
                            deep_count += 1
                            deep_total += value
            print(f"不考虑期初库存-深加工换算产值：处理了 {deep_count} 条记录，总产值 {deep_total:.2f} 元")
        
        # 2.3 材料成本
        # 先调用calculate_material_cost函数计算拆解物原料成本，然后按物料代码统计
        material_cost_by_code = {}
        try:
            # 计算拆解物原料成本
            cost_data = calculate_material_cost(extracted_data)
            if cost_data is not None and not cost_data.empty:
                if '物料代码' in cost_data.columns and '拆解物原料成本' in cost_data.columns:
                    cost_data['拆解物原料成本'] = pd.to_numeric(
                        cost_data['拆解物原料成本'], errors='coerce'
                    ).fillna(0)
                    
                    for _, row in cost_data.iterrows():
                        material_code = str(row.get('物料代码', '')).strip()
                        if material_code in valid_material_codes:
                            cost = float(row.get('拆解物原料成本', 0))
                            material_cost_by_code[material_code] = material_cost_by_code.get(material_code, 0) + cost
        except Exception as e:
            print(f"计算材料成本失败: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # 2.3.1 KG 行材料成本（仅考虑期初库存和库存结余）
        # 材料成本 = 收入预测「提取结果」sheet「单价」× 拆解数量(KG)，始终使用原始「单价」列
        material_cost_kg_by_code = {}
        if consider_opening_stock and material_info_kg:
            def _normalize_material_code(value):
                """统一物料代码：去空格、去 .0、处理 NaN/None。"""
                if value is None:
                    return ''
                try:
                    if pd.isna(value):
                        return ''
                except Exception:
                    pass
                s = str(value).strip()
                if not s or s.lower() == 'nan':
                    return ''
                if s.endswith('.0'):
                    s = s[:-2]
                return s

            try:
                # 优先使用只读「提取结果」(extracted_data)，与首页导出 Excel 的「提取结果」sheet 一致
                price_source = app_data.get_data('extracted_data')
                if price_source is None or price_source.empty:
                    price_source = extracted_data

                price_by_code = {}
                if price_source is not None and not price_source.empty:
                    if '物料代码' in price_source.columns and '单价' in price_source.columns:
                        price_df = price_source.copy()
                        price_df['单价'] = pd.to_numeric(
                            price_df['单价'], errors='coerce'
                        ).fillna(0)
                        for _, row in price_df.iterrows():
                            code = _normalize_material_code(row.get('物料代码', ''))
                            if not code:
                                continue
                            price = float(row.get('单价', 0))
                            # 同一物料代码多行时，优先保留非零单价
                            if code not in price_by_code or (price > 0 and price_by_code[code] == 0):
                                price_by_code[code] = price

                for material_code, info_kg in material_info_kg.items():
                    code_norm = _normalize_material_code(material_code)
                    unit_price = price_by_code.get(code_norm, 0)
                    qty_kg = float(info_kg.get('拆解数量', 0) or 0)
                    material_cost_kg_by_code[material_code] = unit_price * qty_kg
            except Exception as e:
                print(f"计算KG行材料成本失败: {str(e)}")
                traceback.print_exc()
        
        # 2.4 直接人工成本（按物料代码统计）
        direct_labor_by_code = {}
        try:
            direct_labor_result = calculate_direct_labor_cost(app_data, prediction_period)
            if direct_labor_result and not direct_labor_result.get('error'):
                def _normalize_code(value):
                    """统一物料/产物编码：去空格、去 .0、处理 NaN/None。"""
                    if value is None:
                        return ''
                    try:
                        if pd.isna(value):
                            return ''
                    except Exception:
                        pass
                    s = str(value).strip()
                    if not s or s.lower() == 'nan':
                        return ''
                    if s.endswith('.0'):
                        s = s[:-2]
                    return s

                def _finite_float(x, default=0.0):
                    """避免 NaN/Inf 进入 JSON（标准 JSON 不支持 NaN）。"""
                    try:
                        if x is None:
                            return default
                        try:
                            if pd.isna(x):
                                return default
                        except (TypeError, ValueError):
                            pass
                        v = float(x)
                    except (TypeError, ValueError):
                        return default
                    if not math.isfinite(v):
                        return default
                    return v

                def _pick_wage(item):
                    """按口径读取计件工资：不考虑期初库存和库存结余 → 读取同名列，否则回退到常规 工资。"""
                    if not consider_opening_stock:
                        raw = item.get('工资(不考虑期初库存和库存结余)')
                        if raw is None or raw == '':
                            raw = item.get('工资', 0)
                    else:
                        raw = item.get('工资', 0)
                    try:
                        return float(raw or 0)
                    except (TypeError, ValueError):
                        return 0.0

                def _pick_fixed_cost(allocation):
                    """按口径读取分摊固定成本：不考虑期初库存和库存结余 → fixed_cost_no_opening，否则 fixed_cost。"""
                    if not consider_opening_stock:
                        raw = allocation.get('fixed_cost_no_opening')
                        if raw is None or raw == '':
                            raw = allocation.get('fixed_cost', 0)
                    else:
                        raw = allocation.get('fixed_cost', 0)
                    try:
                        return float(raw or 0)
                    except (TypeError, ValueError):
                        return 0.0

                # 2.4.1 生产工人计件工资统计
                # 口径说明：
                # - 考虑期初：保持改造前行为（物料代码/工资）。
                # - 不考虑期初：按用户规范读取「工资(不考虑期初库存和库存结余)」，按「原物料代码」匹配；
                #               类别 ∈ {旧机, 一次拆解产物, 一破, 打包铁}，排除「屏」。
                # 第一部分：旧机类别 - 按物料/原物料代码统计（两者在 part1_details 中同值）
                if direct_labor_result.get('part1_details'):
                    for item in direct_labor_result['part1_details']:
                        if consider_opening_stock:
                            material_code = _normalize_code(item.get('物料代码', ''))
                        else:
                            material_code = _normalize_code(
                                item.get('原物料代码', '') or item.get('物料代码', '')
                            )
                        if material_code and material_code in allowed_material_codes:
                            wage = _pick_wage(item)
                            direct_labor_by_code[material_code] = direct_labor_by_code.get(material_code, 0) + wage

                # 第二部分：一次拆解产物 - 优先按原物料代码归属，否则在原始数据(未减扣)中匹配原物料代码
                if direct_labor_result.get('part2_details'):
                    for item in direct_labor_result['part2_details']:
                        wage = _pick_wage(item)
                        material_code = _normalize_code(item.get('原物料代码', ''))
                        if material_code and material_code in allowed_material_codes:
                            direct_labor_by_code[material_code] = direct_labor_by_code.get(material_code, 0) + wage
                        else:
                            # 兼容旧数据：用拆解产物编码在原始数据、类别=拆解产物中匹配原物料代码
                            product_code = _normalize_code(item.get('拆解产物编码', ''))
                            if original_data is not None and not original_data.empty:
                                if '类别' in original_data.columns and '拆解产物编码' in original_data.columns:
                                    product_mask = (original_data['类别'] == '拆解产物')
                                    product_data = original_data[product_mask]
                                    matched_rows = product_data[
                                        product_data['拆解产物编码']
                                            .astype(str)
                                            .str.strip()
                                            .str.replace(r'\.0$', '', regex=True) == product_code
                                    ]
                                    for _, row in matched_rows.iterrows():
                                        material_code = _normalize_code(row.get('原物料代码', ''))
                                        if material_code and material_code in allowed_material_codes:
                                            direct_labor_by_code[material_code] = direct_labor_by_code.get(material_code, 0) + wage

                # 第三部分：打包铁、一破计件（排除屏类别）
                if direct_labor_result.get('part3_details'):
                    for item in direct_labor_result['part3_details']:
                        item_category = str(item.get('类别', '')).strip()
                        if item_category not in ('打包铁', '一破'):
                            continue
                        wage = _pick_wage(item)
                        material_code = _normalize_code(item.get('原物料代码', ''))
                        if material_code and material_code in allowed_material_codes:
                            direct_labor_by_code[material_code] = direct_labor_by_code.get(material_code, 0) + wage
                        else:
                            # 兼容旧数据：用深加工产物编码回填原物料代码
                            deep_product_code = _normalize_code(item.get('深加工产物编码', ''))
                            if deep_processing_data is not None and not deep_processing_data.empty:
                                if '深加工产物编码' in deep_processing_data.columns:
                                    matched_rows = deep_processing_data[
                                        deep_processing_data['深加工产物编码']
                                            .astype(str)
                                            .str.strip()
                                            .str.replace(r'\.0$', '', regex=True) == deep_product_code
                                    ]
                                    for _, row in matched_rows.iterrows():
                                        material_code = _normalize_code(row.get('原物料代码', ''))
                                        if material_code and material_code in allowed_material_codes:
                                            direct_labor_by_code[material_code] = direct_labor_by_code.get(material_code, 0) + wage

                # 2.4.2 分摊固定工资、社保、公积金统计
                # 口径说明：
                # - 考虑期初：读取 fixed_cost；不考虑期初：读取 fixed_cost_no_opening。
                # - 类别 ∈ {黑电, 白电, 冰箱, 金属打包, 塑料}，排除「屏」。
                # - 按「原物料代码」匹配（旧机情形下 原物料代码 == 物料代码）。
                category_details = direct_labor_result.get('category_details', {})

                # 黑电、白电、冰箱：按原物料代码统计
                for category in ['黑电', '白电', '冰箱']:
                    if category in category_details:
                        category_data = category_details[category]
                        item_allocations = category_data.get('item_allocations', [])
                        for allocation in item_allocations:
                            allocation_item = allocation.get('item', {})
                            fixed_cost = _pick_fixed_cost(allocation)

                            material_code = _normalize_code(
                                allocation_item.get('原物料代码', '') or allocation_item.get('物料代码', '')
                            )
                            if material_code and material_code in allowed_material_codes:
                                direct_labor_by_code[material_code] = direct_labor_by_code.get(material_code, 0) + fixed_cost

                # 金属打包、塑料：取自直接人工成本页分摊明细，按原物料代码统计
                if '金属打包' in category_details:
                    category_data = category_details['金属打包']
                    item_allocations = category_data.get('item_allocations', [])
                    for allocation in item_allocations:
                        allocation_item = allocation.get('item', {})
                        fixed_cost = _pick_fixed_cost(allocation)
                        material_code = _normalize_code(allocation_item.get('原物料代码', ''))
                        if material_code and material_code in allowed_material_codes:
                            direct_labor_by_code[material_code] = direct_labor_by_code.get(material_code, 0) + fixed_cost
                        else:
                            if not consider_opening_stock:
                                # 不考虑期初：⑤仅按「原物料代码」分摊，禁止编码回填导致重复归集
                                continue
                            deep_product_code = _normalize_code(allocation_item.get('深加工产物编码', ''))
                            if deep_processing_data is not None and not deep_processing_data.empty:
                                if '深加工产物编码' in deep_processing_data.columns:
                                    matched_rows = deep_processing_data[
                                        deep_processing_data['深加工产物编码']
                                            .astype(str)
                                            .str.strip()
                                            .str.replace(r'\.0$', '', regex=True) == deep_product_code
                                    ]
                                    for _, row in matched_rows.iterrows():
                                        material_code = _normalize_code(row.get('原物料代码', ''))
                                        if material_code and material_code in allowed_material_codes:
                                            direct_labor_by_code[material_code] = direct_labor_by_code.get(material_code, 0) + fixed_cost

                if '塑料' in category_details:
                    category_data = category_details['塑料']
                    item_allocations = category_data.get('item_allocations', [])
                    for allocation in item_allocations:
                        allocation_item = allocation.get('item', {})
                        fixed_cost = _pick_fixed_cost(allocation)
                        material_code = _normalize_code(allocation_item.get('原物料代码', ''))
                        if material_code and material_code in allowed_material_codes:
                            direct_labor_by_code[material_code] = direct_labor_by_code.get(material_code, 0) + fixed_cost
                        else:
                            if not consider_opening_stock:
                                # 不考虑期初：⑤仅按「原物料代码」分摊，禁止编码回填导致重复归集
                                continue
                            deep_product_code = _normalize_code(allocation_item.get('深加工产物编码', ''))
                            if deep_product_code:
                                if deep_processing_data is not None and not deep_processing_data.empty:
                                    if '深加工产物编码' in deep_processing_data.columns:
                                        matched_rows = deep_processing_data[
                                            deep_processing_data['深加工产物编码']
                                                .astype(str)
                                                .str.strip()
                                                .str.replace(r'\.0$', '', regex=True) == deep_product_code
                                        ]
                                        for _, row in matched_rows.iterrows():
                                            material_code = _normalize_code(row.get('原物料代码', ''))
                                            if material_code and material_code in allowed_material_codes:
                                                direct_labor_by_code[material_code] = direct_labor_by_code.get(material_code, 0) + fixed_cost
                            product_code = _normalize_code(allocation_item.get('拆解产物编码', ''))
                            if product_code:
                                if original_data is not None and not original_data.empty:
                                    if '类别' in original_data.columns and '拆解产物编码' in original_data.columns:
                                        product_mask = (original_data['类别'] == '拆解产物')
                                        product_data = original_data[product_mask]
                                        matched_rows = product_data[
                                            product_data['拆解产物编码']
                                                .astype(str)
                                                .str.strip()
                                                .str.replace(r'\.0$', '', regex=True) == product_code
                                        ]
                                        for _, row in matched_rows.iterrows():
                                            material_code = _normalize_code(row.get('原物料代码', ''))
                                            if material_code and material_code in allowed_material_codes:
                                                direct_labor_by_code[material_code] = direct_labor_by_code.get(material_code, 0) + fixed_cost

        except Exception as e:
            print(f"计算直接人工成本失败: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # ========== 辅助函数：计算类别总产值和物料代码按类别的产物价值 ==========
        def _calculate_category_total_value(valid_material_codes, original_data, deep_processing_data, price_mapping):
            """计算类别总产值（用于分摊计算）- 使用与产物价值列相同的计算逻辑"""
            categories = ['电视', '冰箱', '空调', '洗衣机', '电脑']
            category_total = {cat: 0.0 for cat in categories}
            
            # 分类关键词映射
            category_keyword_mapping = {
                '电视': ['电视', '彩电', 'CRT其它机壳破碎塑料', '线路板边框破碎塑料', '等离子', '废旧玻璃电子枪', '废旧金属荫罩压块铁', '黑白'],
                '电脑': ['电脑', '显示器', '笔记本', '主机', '废旧金属黑色金属-铁及其合金-电子枪'],
                '冰箱': ['冰箱', '冰柜'],
                '空调': ['空调'],
                '洗衣机': ['洗衣机', '双缸']
            }
            
            # 获取被减扣数据
            deducted_data = app_data.get_data('deducted_data_manual')
            
            # 1. 一次拆解产物产值分类汇总（处置类别为"空白"或空）
            if original_data is not None and not original_data.empty:
                if '类别' in original_data.columns and '原物料代码' in original_data.columns:
                    product_mask = (original_data['类别'] == '拆解产物')
                    # 筛选处置类别为"空白"或空的记录
                    if '处置类别' in original_data.columns:
                        product_mask = product_mask & (
                            (original_data['处置类别'].astype(str).str.strip() == '空白') |
                            (original_data['处置类别'].astype(str).str.strip() == '')
                        )
                    
                    product_data = original_data[product_mask].copy()
                    if '原物料代码' in product_data.columns and '拆解产物编码' in product_data.columns and '计算结果(KG)' in product_data.columns:
                        product_data['计算结果(KG)'] = pd.to_numeric(product_data['计算结果(KG)'], errors='coerce').fillna(0)
                        
                        for _, row in product_data.iterrows():
                            material_code = str(row.get('原物料代码', '')).strip()
                            if material_code in valid_material_codes:
                                material_name = str(row.get('原物料名称', '')).strip()
                                product_code = str(row.get('拆解产物编码', '')).strip()
                                weight = float(row.get('计算结果(KG)', 0))
                                price = price_mapping.get(product_code, 0)
                                if price < 0:
                                    price = 0
                                value = weight * price
                                
                                # 根据原物料名称分类
                                for cat, keywords in category_keyword_mapping.items():
                                    for keyword in keywords:
                                        if keyword in material_name:
                                            category_total[cat] += value
                                            break
                                    else:
                                        continue
                                    break
            
            # 2. 减去被减扣数据中类别为"拆解产物"的产值
            if deducted_data is not None and not deducted_data.empty:
                if '类别' in deducted_data.columns:
                    deducted_product_mask = (deducted_data['类别'] == '拆解产物')
                    deducted_product_data = deducted_data[deducted_product_mask].copy()
                    
                    if not deducted_product_data.empty:
                        if '原物料代码' in deducted_product_data.columns and '拆解产物编码' in deducted_product_data.columns and '计算结果(KG)' in deducted_product_data.columns:
                            deducted_product_data['计算结果(KG)'] = pd.to_numeric(deducted_product_data['计算结果(KG)'], errors='coerce').fillna(0)
                            
                            for _, row in deducted_product_data.iterrows():
                                material_code = str(row.get('原物料代码', '')).strip()
                                if material_code in valid_material_codes:
                                    material_name = str(row.get('原物料名称', '')).strip()
                                    product_code = str(row.get('拆解产物编码', '')).strip()
                                    weight = float(row.get('计算结果(KG)', 0))
                                    if weight > 0:
                                        price = price_mapping.get(product_code, 0)
                                        if price < 0:
                                            price = 0
                                        value = weight * price
                                        
                                        # 根据原物料名称分类
                                        for cat, keywords in category_keyword_mapping.items():
                                            for keyword in keywords:
                                                if keyword in material_name:
                                                    category_total[cat] -= value
                                                    break
                                            else:
                                                continue
                                            break
            
            # 3. 深加工拆解产物产值分类汇总（未减扣的）
            if deep_processing_data is not None and not deep_processing_data.empty:
                if '是否减扣' in deep_processing_data.columns:
                    non_deducted_data = deep_processing_data[deep_processing_data['是否减扣'] == '否'].copy()
                else:
                    non_deducted_data = deep_processing_data.copy()
                
                if not non_deducted_data.empty:
                    for _, row in non_deducted_data.iterrows():
                        material_code = str(row.get('原物料代码', '')).strip()
                        if material_code in valid_material_codes:
                            material_name = str(row.get('原物料名称', '')).strip()
                            deep_product_code = str(row.get('深加工产物编码', '')).strip()
                            deep_result_kg = float(row.get('深加工结果(KG)', 0)) if pd.notna(row.get('深加工结果(KG)')) else 0
                            price_no_tax = price_mapping.get(deep_product_code, 0)
                            if price_no_tax < 0:
                                price_no_tax = 0
                            material_value = deep_result_kg * price_no_tax
                            
                            # 根据原物料名称分类
                            for cat, keywords in category_keyword_mapping.items():
                                for keyword in keywords:
                                    if keyword in material_name:
                                        category_total[cat] += material_value
                                        break
                                else:
                                    continue
                                break
            
            return category_total
        
        def _calculate_category_product_value_by_code(valid_material_codes, original_data, deep_processing_data, deducted_data, price_mapping):
            """计算每个物料代码按类别的产物价值（考虑被减扣数据）"""
            categories = ['电视', '冰箱', '空调', '洗衣机', '电脑']
            result = {}  # {物料代码: {类别: 产值}}
            
            # 分类关键词映射
            category_keyword_mapping = {
                '电视': ['电视', '彩电', 'CRT其它机壳破碎塑料', '线路板边框破碎塑料', '等离子', '废旧玻璃电子枪', '废旧金属荫罩压块铁', '黑白'],
                '电脑': ['电脑', '显示器', '笔记本', '主机', '废旧金属黑色金属-铁及其合金-电子枪'],
                '冰箱': ['冰箱', '冰柜'],
                '空调': ['空调'],
                '洗衣机': ['洗衣机', '双缸']
            }
            
            # 初始化
            for material_code in valid_material_codes:
                result[material_code] = {cat: 0.0 for cat in categories}
            
            # 1. 一次拆解产物产值（从原始数据）
            if original_data is not None and not original_data.empty:
                if '类别' in original_data.columns and '原物料代码' in original_data.columns:
                    product_mask = (original_data['类别'] == '拆解产物')
                    product_data = original_data[product_mask].copy()
                    if '原物料代码' in product_data.columns and '拆解产物编码' in product_data.columns and '计算结果(KG)' in product_data.columns:
                        product_data['计算结果(KG)'] = pd.to_numeric(product_data['计算结果(KG)'], errors='coerce').fillna(0)
                        
                        for _, row in product_data.iterrows():
                            material_code = str(row.get('原物料代码', '')).strip()
                            if material_code in valid_material_codes:
                                material_name = str(row.get('原物料名称', '')).strip()
                                product_code = str(row.get('拆解产物编码', '')).strip()
                                weight = float(row.get('计算结果(KG)', 0))
                                price = price_mapping.get(product_code, 0)
                                if price < 0:
                                    price = 0
                                value = weight * price
                                
                                # 根据原物料名称分类
                                for cat, keywords in category_keyword_mapping.items():
                                    for keyword in keywords:
                                        if keyword in material_name:
                                            result[material_code][cat] += value
                                            break
                                    else:
                                        continue
                                    break
            
            # 1.1 减去被减扣数据中类别为"拆解产物"的产值
            if deducted_data is not None and not deducted_data.empty:
                if '类别' in deducted_data.columns:
                    deducted_product_mask = (deducted_data['类别'] == '拆解产物')
                    deducted_product_data = deducted_data[deducted_product_mask].copy()
                    
                    if not deducted_product_data.empty:
                        if '原物料代码' in deducted_product_data.columns and '拆解产物编码' in deducted_product_data.columns and '计算结果(KG)' in deducted_product_data.columns:
                            deducted_product_data['计算结果(KG)'] = pd.to_numeric(deducted_product_data['计算结果(KG)'], errors='coerce').fillna(0)
                            
                            for _, row in deducted_product_data.iterrows():
                                material_code = str(row.get('原物料代码', '')).strip()
                                if material_code in valid_material_codes:
                                    material_name = str(row.get('原物料名称', '')).strip()
                                    product_code = str(row.get('拆解产物编码', '')).strip()
                                    weight = float(row.get('计算结果(KG)', 0))
                                    if weight > 0:
                                        price = price_mapping.get(product_code, 0)
                                        if price < 0:
                                            price = 0
                                        value = weight * price
                                        
                                        # 根据原物料名称分类，减去被减扣的产值
                                        for cat, keywords in category_keyword_mapping.items():
                                            for keyword in keywords:
                                                if keyword in material_name:
                                                    result[material_code][cat] -= value
                                                    break
                                            else:
                                                continue
                                            break
            
            # 2. 深加工拆解产物产值
            if deep_processing_data is not None and not deep_processing_data.empty:
                if '是否减扣' in deep_processing_data.columns:
                    non_deducted_data = deep_processing_data[deep_processing_data['是否减扣'] == '否'].copy()
                else:
                    non_deducted_data = deep_processing_data.copy()
                
                if not non_deducted_data.empty:
                    for _, row in non_deducted_data.iterrows():
                        material_code = str(row.get('原物料代码', '')).strip()
                        if material_code in valid_material_codes:
                            material_name = str(row.get('原物料名称', '')).strip()
                            deep_product_code = str(row.get('深加工产物编码', '')).strip()
                            deep_result_kg = float(row.get('深加工结果(KG)', 0)) if pd.notna(row.get('深加工结果(KG)')) else 0
                            price_no_tax = price_mapping.get(deep_product_code, 0)
                            if price_no_tax < 0:
                                price_no_tax = 0
                            material_value = deep_result_kg * price_no_tax
                            
                            # 根据原物料名称分类
                            for cat, keywords in category_keyword_mapping.items():
                                for keyword in keywords:
                                    if keyword in material_name:
                                        result[material_code][cat] += material_value
                                        break
                                else:
                                    continue
                                break
            
            return result
        
        # 计算类别总产值和物料代码按类别的产物价值
        # 不考虑期初库存时与 product_value_by_code 口径一致，直接从 product_value_by_code + material_info 按类别汇总
        if consider_opening_stock:
            category_total_value = _calculate_category_total_value(valid_material_codes, original_data, deep_processing_data, price_mapping)
            category_product_value_by_code = _calculate_category_product_value_by_code(valid_material_codes, original_data, deep_processing_data, deducted_data, price_mapping)
        else:
            categories_list = ['电视', '冰箱', '空调', '洗衣机', '电脑']
            category_total_value = {c: 0.0 for c in categories_list}
            category_product_value_by_code = {mc: {c: 0.0 for c in categories_list} for mc in valid_material_codes}
            for material_code in allowed_material_codes:
                val = product_value_by_code.get(material_code, 0.0)
                if val <= 0:
                    continue
                info = material_info.get(material_code, {})
                cat = info.get('类别', '')
                if cat in category_total_value:
                    category_total_value[cat] += val
                    if material_code in category_product_value_by_code:
                        category_product_value_by_code[material_code][cat] += val
        
        # 2.5 制造费用（按物料代码统计）
        manufacturing_cost_by_code = {}
        # 制造费用明细（按物料代码统计）
        manufacturing_cost_details_by_code = {}
        try:
            from app.api.cost_forecast_api import (
                calculate_indirect_labor_cost,
                calculate_screen_cost_allocation
            )
            from app.api.data_management_api import get_manufacturing_cost_dataframe
            
            # 初始化制造费用和明细（包括所有有效的物料代码）
            allowed_material_codes = valid_material_codes | set(material_info_kg.keys())
            for material_code in allowed_material_codes:
                manufacturing_cost_by_code[material_code] = 0.0
                manufacturing_cost_details_by_code[material_code] = {
                    '间接人工提成成本': 0.0,
                    '分摊固定成本明细': 0.0,
                    '与拆解量相关的费用': 0.0,
                    '与电机入库量相关的费用': 0.0,
                    '预计月均费用分摊': 0.0,
                    '环保费': 0.0,
                    '公共费用分摊': 0.0,
                    '屏费用分摊': 0.0,
                    '制造费用间接人工分摊': 0.0,
                    '制造费用公共成本分摊': 0.0,
                }
            
            # 2.5.1 间接人工提成成本
            
            indirect_labor_result = calculate_indirect_labor_cost(
                app_data, prediction_period, include_no_opening_columns=True
            )
            if consider_opening_stock:
                if indirect_labor_result and not indirect_labor_result.get('error'):
                    # 旧机类别：直接按物料代码统计指定的提成成本字段
                    for detail in indirect_labor_result.get('part1_details', []):
                        material_code = str(detail.get('物料代码', '')).strip()
                        if material_code in allowed_material_codes:
                            cost = (
                                float(detail.get('物流主管提成成本', 0) or 0) +
                                float(detail.get('物流卸货提成成本', 0) or 0) +
                                float(detail.get('班组长提成成本', 0) or 0) +
                                float(detail.get('生产主管提成成本', 0) or 0) +
                                float(detail.get('维修班长提成成本', 0) or 0) +
                                float(detail.get('维修员提成成本', 0) or 0) +
                                float(detail.get('冰箱维修主管提成成本', 0) or 0)
                            )
                            
                            manufacturing_cost_by_code[material_code] += cost
                            manufacturing_cost_details_by_code[material_code]['间接人工提成成本'] += cost
                    
                    # 一次拆解产物：在原始数据(未减扣)匹配原物料代码
                    # 按原物料代码汇总，避免重复计算
                    # 先按拆解产物编码去重，避免重复处理
                    processed_part2_keys = set()  # 用于跟踪已处理的拆解产物编码
                    for detail in indirect_labor_result.get('part2_details', []):
                        product_code = str(detail.get('拆解产物编码', '')).strip()
                        if product_code in processed_part2_keys:
                            continue  # 跳过已处理的记录
                        processed_part2_keys.add(product_code)
                        
                        if original_data is not None and not original_data.empty:
                            if '类别' in original_data.columns and '拆解产物编码' in original_data.columns:
                                product_mask = (original_data['类别'] == '拆解产物')
                                product_data = original_data[product_mask]
                                matched_rows = product_data[product_data['拆解产物编码'].astype(str).str.strip() == product_code]
                                
                                # 按原物料代码汇总计算结果(KG)
                                material_code_kg_sum = {}  # {原物料代码: 计算结果(KG)汇总}
                                for _, row in matched_rows.iterrows():
                                    material_code = str(row.get('原物料代码', '')).strip()
                                    if material_code in allowed_material_codes:
                                        result_kg = float(row.get('计算结果(KG)', 0)) if pd.notna(row.get('计算结果(KG)')) else 0
                                        if material_code not in material_code_kg_sum:
                                            material_code_kg_sum[material_code] = 0.0
                                        material_code_kg_sum[material_code] += result_kg
                                
                                # 对每个原物料代码，计算班组长提成成本并累加
                                unit_price = float(detail.get('班组长提成单价', 0) or 0)
                                for material_code, total_kg in material_code_kg_sum.items():
                                    if total_kg > 0:
                                        cost = total_kg * unit_price
                                        manufacturing_cost_by_code[material_code] += cost
                                        manufacturing_cost_details_by_code[material_code]['间接人工提成成本'] += cost
                    
                    # 打包铁、一破：在深加工数据匹配原物料代码
                    # 按原物料代码汇总，避免重复计算
                    # 先按深加工产物编码和类别去重，避免重复处理
                    processed_part3_keys = set()  # 用于跟踪已处理的(深加工产物编码, 类别)组合
                    for detail in indirect_labor_result.get('part3_details', []):
                        deep_product_code = str(detail.get('深加工产物编码', '')).strip()
                        category = detail.get('类别', '')
                        part3_key = (deep_product_code, category)
                        if part3_key in processed_part3_keys:
                            continue  # 跳过已处理的记录
                        processed_part3_keys.add(part3_key)
                        
                        if category in ['打包铁', '一破'] and deep_processing_data is not None and not deep_processing_data.empty:
                            if '深加工产物编码' in deep_processing_data.columns:
                                matched_rows = deep_processing_data[deep_processing_data['深加工产物编码'].astype(str).str.strip() == deep_product_code]

                                # 按原物料代码汇总深加工结果(KG)
                                material_code_kg_sum = {}  # {原物料代码: 深加工结果(KG)汇总}
                                for _, row in matched_rows.iterrows():
                                    material_code = str(row.get('原物料代码', '')).strip()
                                    if material_code in allowed_material_codes:
                                        deep_result_kg = float(row.get('深加工结果(KG)', 0)) if pd.notna(row.get('深加工结果(KG)')) else 0
                                        if material_code not in material_code_kg_sum:
                                            material_code_kg_sum[material_code] = 0.0
                                        material_code_kg_sum[material_code] += deep_result_kg
                                
                                # 对每个原物料代码，计算班组长提成成本并累加
                                unit_price = float(detail.get('班组长提成单价', 0) or 0)
                                for material_code, total_kg in material_code_kg_sum.items():
                                    if total_kg > 0:
                                        cost = total_kg * unit_price
                                        
                                        
                                        
                                        manufacturing_cost_by_code[material_code] += cost
                                        manufacturing_cost_details_by_code[material_code]['间接人工提成成本'] += cost
                                        
                                        
                
                # 2.5.2 分摊固定成本明细
                if indirect_labor_result and not indirect_labor_result.get('error'):
                    # 旧机类别：直接按物料代码统计指定的分摊字段
                    for detail in indirect_labor_result.get('part1_details', []):
                        material_code = str(detail.get('物料代码', '')).strip()
                        if material_code in valid_material_codes:
                            cost = (
                                float(detail.get('物流主管分摊固定成本', 0) or 0) +
                                float(detail.get('生产主管分摊固定成本', 0) or 0) +
                                float(detail.get('物流卸货分摊固定成本', 0) or 0) +
                                float(detail.get('维修班长分摊固定成本', 0) or 0) +
                                float(detail.get('维修员分摊固定成本', 0) or 0) +
                                float(detail.get('冰箱维修主管分摊固定成本', 0) or 0) +
                                float(detail.get('白电小组长分摊固定成本', 0) or 0) +
                                float(detail.get('生产班组长(黑电)分摊固定成本', 0) or 0) +
                                float(detail.get('生产班组长(冰箱)分摊固定成本', 0) or 0)
                            )
                            
                            
                            
                            manufacturing_cost_by_code[material_code] += cost
                            manufacturing_cost_details_by_code[material_code]['分摊固定成本明细'] += cost
                            
                            
                    
                    # 一次拆解产物：在原始数据(未减扣)匹配原物料代码
                    # 按原物料代码汇总，避免重复计算
                    # 先按拆解产物编码去重，避免重复处理
                    processed_part2_fixed_keys = set()  # 用于跟踪已处理的拆解产物编码
                    for detail in indirect_labor_result.get('part2_details', []):
                        product_code = str(detail.get('拆解产物编码', '')).strip()
                        if product_code in processed_part2_fixed_keys:
                            continue  # 跳过已处理的记录
                        processed_part2_fixed_keys.add(product_code)
                        
                        if original_data is not None and not original_data.empty:
                            if '类别' in original_data.columns and '拆解产物编码' in original_data.columns:
                                product_mask = (original_data['类别'] == '拆解产物')
                                product_data = original_data[product_mask]
                                matched_rows = product_data[product_data['拆解产物编码'].astype(str).str.strip() == product_code]
                                
                                # 按原物料代码汇总计算结果(KG)
                                material_code_kg_sum = {}  # {原物料代码: 计算结果(KG)汇总}
                                for _, row in matched_rows.iterrows():
                                    material_code = str(row.get('原物料代码', '')).strip()
                                    if material_code in allowed_material_codes:
                                        result_kg = float(row.get('计算结果(KG)', 0)) if pd.notna(row.get('计算结果(KG)')) else 0
                                        if material_code not in material_code_kg_sum:
                                            material_code_kg_sum[material_code] = 0.0
                                        material_code_kg_sum[material_code] += result_kg
                                
                                # 计算每个原物料代码的分摊固定成本
                                # 分摊固定成本 = (汇总的计算结果(KG) / detail中的计算结果(KG)) × detail中的分摊固定成本
                                detail_kg = float(detail.get('计算结果(KG)', 0) or 0)
                                detail_fixed_cost = float(detail.get('生产班组长(塑料破碎)分摊固定成本', 0) or 0)
                                
                                for material_code, total_kg in material_code_kg_sum.items():
                                    if total_kg > 0 and detail_kg > 0:
                                        # 按比例计算分摊固定成本
                                        cost = (total_kg / detail_kg) * detail_fixed_cost
                                        
                                        
                                        
                                        manufacturing_cost_by_code[material_code] += cost
                                        manufacturing_cost_details_by_code[material_code]['分摊固定成本明细'] += cost
                                        
                                        
                    
                    # 一破：在深加工数据匹配原物料代码
                    # 按原物料代码汇总，避免重复计算
                    # 先按深加工产物编码和类别去重，避免重复处理
                    processed_part3_fixed_keys = set()  # 用于跟踪已处理的(深加工产物编码, 类别)组合
                    for detail in indirect_labor_result.get('part3_details', []):
                        deep_product_code = str(detail.get('深加工产物编码', '')).strip()
                        category = detail.get('类别', '')
                        part3_key = (deep_product_code, category)
                        if part3_key in processed_part3_fixed_keys:
                            continue  # 跳过已处理的记录
                        processed_part3_fixed_keys.add(part3_key)
                        
                        if category == '一破' and deep_processing_data is not None and not deep_processing_data.empty:
                            if '深加工产物编码' in deep_processing_data.columns:
                                matched_rows = deep_processing_data[deep_processing_data['深加工产物编码'].astype(str).str.strip() == deep_product_code]
                                
                                # 按原物料代码汇总深加工结果(KG)
                                material_code_kg_sum = {}  # {原物料代码: 深加工结果(KG)汇总}
                                for _, row in matched_rows.iterrows():
                                    material_code = str(row.get('原物料代码', '')).strip()
                                    if material_code in allowed_material_codes:
                                        deep_result_kg = float(row.get('深加工结果(KG)', 0)) if pd.notna(row.get('深加工结果(KG)')) else 0
                                        if material_code not in material_code_kg_sum:
                                            material_code_kg_sum[material_code] = 0.0
                                        material_code_kg_sum[material_code] += deep_result_kg
                                
                                # 计算每个原物料代码的分摊固定成本
                                # 分摊固定成本 = (汇总的深加工结果(KG) / detail中的深加工结果(KG)) × detail中的分摊固定成本
                                detail_kg = float(detail.get('深加工结果(KG)', 0) or 0)
                                detail_fixed_cost = float(detail.get('生产班组长(塑料破碎)分摊固定成本', 0) or 0)
                                
                                for material_code, total_kg in material_code_kg_sum.items():
                                    if total_kg > 0 and detail_kg > 0:
                                        # 按比例计算分摊固定成本
                                        cost = (total_kg / detail_kg) * detail_fixed_cost
                                        
                                        
                                        
                                        manufacturing_cost_by_code[material_code] += cost
                                        manufacturing_cost_details_by_code[material_code]['分摊固定成本明细'] += cost
                                        
                                        
                
                # 2.5.3 与拆解量相关的费用
                # 拆解物原料成本页"本期实际投产数量" * 按照"物料代码"匹配"制造费用成本"页"与拆解量相关的费用"卡片"类别"对应的各费用名称"单价"之和
                
                
                manufacturing_result = calculate_manufacturing_cost(app_data, prediction_period)
                
                
                if manufacturing_result and not manufacturing_result.get('error'):
                    # 获取制造费用基础数据
                    manufacturing_cost_df = get_manufacturing_cost_dataframe()
                    
                    
                    if manufacturing_cost_df is not None and not manufacturing_cost_df.empty:
                        # 筛选与拆解量相关的费用
                        if '备注' in manufacturing_cost_df.columns:
                            disassembly_related_df = manufacturing_cost_df[
                                manufacturing_cost_df['备注'].astype(str).str.contains('与拆解量相关', case=False, na=False)
                            ].copy()
                            
                            
                            
                            # 获取拆解数量
                            extracted_data = app_data.get_data('extracted_data_manual')
                            
                            
                            if extracted_data is not None and not extracted_data.empty:
                                if '类别' in extracted_data.columns and '物料代码' in extracted_data.columns:
                                    old_machine_data = extracted_data[extracted_data['类别'] == '旧机'].copy()
                                    
                                    
                                    if '非限制使用的库存' in old_machine_data.columns:
                                        old_machine_data['非限制使用的库存'] = pd.to_numeric(
                                            old_machine_data['非限制使用的库存'], errors='coerce'
                                        ).fillna(0)
                                        
                                        # 按物料代码分组，对每个物料代码计算总单价
                                        processed_count = 0
                                        for _, material_row in old_machine_data.iterrows():
                                            material_code = str(material_row.get('物料代码', '')).strip()
                                            if material_code in valid_material_codes:
                                                material_name = str(material_row.get('物料描述', '')).strip()
                                                quantity = float(material_row.get('非限制使用的库存', 0))
                                                
                                                if quantity > 0:
                                                    # 根据物料名称匹配类别
                                                    material_category = classify_by_product_name(material_name)
                                                    if material_category:
                                                        # 找到该类别对应的所有费用行的单价，求和
                                                        total_unit_price = 0.0
                                                        for _, cost_row in disassembly_related_df.iterrows():
                                                            unit_price = resolve_disassembly_category_unit_price(cost_row, material_category)
                                                            if unit_price > 0:
                                                                total_unit_price += unit_price
                                                        
                                                        # 计算：本期实际投产数量 × 总单价
                                                        if total_unit_price > 0:
                                                            cost = quantity * total_unit_price
                                                            manufacturing_cost_by_code[material_code] += cost
                                                            manufacturing_cost_details_by_code[material_code]['与拆解量相关的费用'] += cost
                                                            processed_count += 1
                                                            
                                                            
                                        
                                        
                
                # 2.5.4 与电机入库量相关的费用
                deducted_data = app_data.get_data('deducted_data_manual')
                motor_codes = ['811053046', '811053050', '811304664', '811437999']
                if deducted_data is not None and not deducted_data.empty:
                    if '拆解产物编码' in deducted_data.columns and '原物料代码' in deducted_data.columns:
                        # 筛选电机编码
                        motor_mask = deducted_data['拆解产物编码'].astype(str).str.strip().isin([code.strip() for code in motor_codes])
                        motor_data = deducted_data[motor_mask].copy()
                        
                        if not motor_data.empty and '计算结果(KG)' in motor_data.columns:
                            motor_data['计算结果(KG)'] = pd.to_numeric(motor_data['计算结果(KG)'], errors='coerce').fillna(0)
                            
                            # 获取制造费用基础数据中的单价
                            manufacturing_cost_df = get_manufacturing_cost_dataframe()
                            if manufacturing_cost_df is not None and not manufacturing_cost_df.empty:
                                if '备注' in manufacturing_cost_df.columns:
                                    motor_related_df = manufacturing_cost_df[
                                        manufacturing_cost_df['备注'].astype(str).str.contains('与电机入库量相关', case=False, na=False)
                                    ].copy()
                                    
                                    # 按原物料代码统计
                                    for _, row in motor_data.iterrows():
                                        material_code = str(row.get('原物料代码', '')).strip()
                                        if material_code in valid_material_codes:
                                            weight = float(row.get('计算结果(KG)', 0))
                                            material_name = str(row.get('原物料名称', '')).strip()
                                            
                                            # 根据物料名称匹配类别（空调或洗衣机）
                                            material_category = classify_by_product_name(material_name)
                                            if material_category in ['空调', '洗衣机']:
                                                # 从制造费用基础数据获取单价
                                                for _, cost_row in motor_related_df.iterrows():
                                                    if material_category in cost_row.index:
                                                        unit_price = float(cost_row.get(material_category, 0)) if pd.notna(cost_row.get(material_category)) else 0
                                                        if unit_price > 0:
                                                            cost = weight * unit_price
                                                            manufacturing_cost_by_code[material_code] += cost
                                                            manufacturing_cost_details_by_code[material_code]['与电机入库量相关的费用'] += cost
                                                            break
                
                # 2.5.5 预计月均费用分摊
                # "预计月均费用"列 = "制造费用成本"页面的预计月均费用"卡片，类别为"电视"，"冰箱"，"空调"，"洗衣机"，"电脑"，
                # 统计"成本" x "产物价值"列每个"物料代码"的产值 / "产物价值"列的"电视"，"冰箱"，"空调"，"洗衣机"，"电脑"分类总数
                # 注意：这里使用"分类费用汇总"卡片的汇总值，而不是明细项
                # 类别总产值 = 当期拆解收益测算分析表页面的"类别"为对应类别，产物价值的总和
                
                # 先构建临时的result_data来计算类别总产值（按类别汇总产物价值）
                # 使用与result_data相同的逻辑：从material_info和material_info_kg中获取类别信息
                from app.api.cost_forecast_api import classify_by_product_name
                category_total_value_from_product = {'电视': 0.0, '冰箱': 0.0, '空调': 0.0, '洗衣机': 0.0, '电脑': 0.0}
                
                # 从material_info（单位：台）汇总
                for material_code in valid_material_codes:
                    material_value = product_value_by_code.get(material_code, 0.0)
                    if material_value > 0:
                        info = material_info.get(material_code, {})
                        material_category = info.get('类别', '')
                        if material_category and material_category in category_total_value_from_product:
                            category_total_value_from_product[material_category] += material_value
                
                # 从material_info_kg（单位：KG）汇总
                for material_code in material_info_kg.keys():
                    material_value = product_value_by_code.get(material_code, 0.0)
                    if material_value > 0:
                        info_kg = material_info_kg.get(material_code, {})
                        material_category = info_kg.get('类别', '')
                        if material_category and material_category in category_total_value_from_product:
                            category_total_value_from_product[material_category] += material_value
                
                
                
                if manufacturing_result and not manufacturing_result.get('error'):
                    monthly_average_list = manufacturing_result.get('monthly_average', [])
                    
                    # 计算"分类费用汇总"：按类别汇总所有预计月均费用明细项的成本
                    category_monthly_cost_sum = {'电视': 0.0, '冰箱': 0.0, '空调': 0.0, '洗衣机': 0.0, '电脑': 0.0}
                    for item in monthly_average_list:
                        details = item.get('明细', [])
                        for detail in details:
                            category = detail.get('category', '')
                            if category in category_monthly_cost_sum:
                                cost = float(detail.get('cost', 0) or 0)
                                category_monthly_cost_sum[category] += cost
                    
                    
                    
                    # 使用"分类费用汇总"的汇总值进行分摊
                    for category in ['电视', '冰箱', '空调', '洗衣机', '电脑']:
                        category_total_cost = category_monthly_cost_sum.get(category, 0.0)
                        if category_total_cost > 0:
                            category_total = category_total_value_from_product.get(category, 0.0)
                            if category_total > 0:
                                # 分摊公式：分类费用汇总 × (物料代码产值 / 类别总产值)
                                for material_code in allowed_material_codes:
                                    material_value = product_value_by_code.get(material_code, 0.0)
                                    if material_value > 0:
                                        # 确定物料代码的类别
                                        material_category = None
                                        if material_code in valid_material_codes:
                                            info = material_info.get(material_code, {})
                                            material_category = info.get('类别', '')
                                        elif material_code in material_info_kg:
                                            info_kg = material_info_kg.get(material_code, {})
                                            material_category = info_kg.get('类别', '')
                                        
                                        if material_category == category:
                                            allocated_cost = category_total_cost * (material_value / category_total)
                                            manufacturing_cost_by_code[material_code] += allocated_cost
                                            manufacturing_cost_details_by_code[material_code]['预计月均费用分摊'] += allocated_cost
                                            
                                            
                
                # 2.5.6 环保费
                # "被减扣数据"sheet，筛选"类别"为"拆解产物"，"处置类别"为"付费处置"，"内转荧光灯处置"；
                # 按照"原物料代码"匹配"拆解产物编码"统计"计算结果(KG)"*"制造费用成本"页，"环保费"卡片对应的"拆解产物编码"的"单价(元/KG)"
                # 注意：环保费单价来自价格数据中销售单价<0的记录（与cost_forecast_api.py逻辑一致）
                # 被减扣重量与「被减扣数据（只读）」表同源，与制造费用成本页环保费一致，不用 deducted_data_manual。
                from app.api.data_management_api import _build_deducted_readonly_dataframe
                deducted_data_for_env = _build_deducted_readonly_dataframe(app_data)
                if deducted_data_for_env is not None and not deducted_data_for_env.empty:
                    if '类别' in deducted_data_for_env.columns and '处置类别' in deducted_data_for_env.columns:
                        # 筛选类别="拆解产物"且处置类别="付费处置"或"内转荧光灯处置"
                        env_mask = (
                            (deducted_data_for_env['类别'].astype(str) == '拆解产物') & 
                            (deducted_data_for_env['处置类别'].astype(str).isin(['付费处置', '内转荧光灯处置']))
                        )
                        env_data = deducted_data_for_env[env_mask].copy()
                        
                        if not env_data.empty and '计算结果(KG)' in env_data.columns and '拆解产物编码' in env_data.columns:
                            env_data['计算结果(KG)'] = pd.to_numeric(env_data['计算结果(KG)'], errors='coerce').fillna(0)
                            env_data['拆解产物编码'] = env_data['拆解产物编码'].astype(str).str.strip()
                            
                            # 从价格数据中获取环保费单价（销售单价<0的记录）
                            from data.base_data.price_data import load_price_data
                            price_df = load_price_data()
                            env_fee_price_mapping = {}  # 拆解产物编码 -> 单价(元/KG)
                            
                            if price_df is not None and not price_df.empty:
                                # 确保价格列为数值类型
                                if '销售单价-不含税(元/KG)' in price_df.columns:
                                    price_df['销售单价-不含税(元/KG)'] = pd.to_numeric(
                                        price_df['销售单价-不含税(元/KG)'], errors='coerce'
                                    ).fillna(0)
                                    
                                    # 筛选销售单价<0的记录（环保费）
                                    negative_price_mask = price_df['销售单价-不含税(元/KG)'] < 0
                                    env_fee_price_df = price_df[negative_price_mask].copy()
                                    
                                    if not env_fee_price_df.empty:
                                        # 确保编码列为字符串类型并去除空格
                                        if '拆解产物编码' in env_fee_price_df.columns:
                                            env_fee_price_df['拆解产物编码'] = env_fee_price_df['拆解产物编码'].astype(str).str.strip()
                                            
                                            # 建立拆解产物编码到单价的映射（取绝对值）
                                            for _, price_row in env_fee_price_df.iterrows():
                                                product_code = str(price_row.get('拆解产物编码', '')).strip()
                                                unit_price = float(price_row.get('销售单价-不含税(元/KG)', 0))
                                                if product_code and unit_price < 0:
                                                    # 取绝对值作为单价
                                                    env_fee_price_mapping[product_code] = abs(unit_price)
                            
                            
                            
                            # 按原物料代码和拆解产物编码统计
                            processed_count = 0
                            skipped_count = 0
                            skipped_reasons = {}
                            
                            for _, row in env_data.iterrows():
                                material_code = str(row.get('原物料代码', '')).strip()
                                product_code = str(row.get('拆解产物编码', '')).strip()
                                
                                if material_code in valid_material_codes and product_code:
                                    weight = float(row.get('计算结果(KG)', 0))
                                    
                                    # 按照拆解产物编码匹配单价
                                    unit_price = env_fee_price_mapping.get(product_code, 0.0)
                                    
                                    if weight > 0 and unit_price > 0:
                                        cost = weight * unit_price
                                        manufacturing_cost_by_code[material_code] += cost
                                        manufacturing_cost_details_by_code[material_code]['环保费'] += cost
                                        processed_count += 1
                                        
                                        
                                    else:
                                        skipped_count += 1
                                        reason = f"weight={weight}, unit_price={unit_price}"
                                        if reason not in skipped_reasons:
                                            skipped_reasons[reason] = 0
                                        skipped_reasons[reason] += 1
                                else:
                                    skipped_count += 1
                                    reason = f"material_code_in_valid={material_code in valid_material_codes}, product_code_empty={not product_code}"
                                    if reason not in skipped_reasons:
                                        skipped_reasons[reason] = 0
                                    skipped_reasons[reason] += 1
                            
                            
                
                # 2.5.7 公共费用分摊
                # 使用"产物价值"列的数据计算类别总产值（已在2.5.5中计算）
                screen_allocation_result = calculate_screen_cost_allocation(app_data, prediction_period)
                if screen_allocation_result and not screen_allocation_result.get('error'):
                    category_allocation = screen_allocation_result.get('allocation', {}).get('category_allocation', {})
                    for category, total_cost in category_allocation.items():
                        if category in ['电视', '冰箱', '空调', '洗衣机', '电脑'] and total_cost > 0:
                            # 分摊公式：按类别分摊的费用 × (物料代码产值 / 类别总产值)
                            # 使用"产物价值"列的数据计算类别总产值
                            category_total = category_total_value_from_product.get(category, 0.0)
                            if category_total > 0:
                                for material_code in allowed_material_codes:
                                    material_value = product_value_by_code.get(material_code, 0.0)
                                    if material_value > 0:
                                        material_category = None
                                        if material_code in valid_material_codes:
                                            info = material_info.get(material_code, {})
                                            material_category = info.get('类别', '')
                                        elif material_code in material_info_kg:
                                            info_kg = material_info_kg.get(material_code, {})
                                            material_category = info_kg.get('类别', '')
                                        
                                        if material_category == category:
                                            allocated_cost = total_cost * (material_value / category_total)
                                            manufacturing_cost_by_code[material_code] += allocated_cost
                                            manufacturing_cost_details_by_code[material_code]['公共费用分摊'] += allocated_cost
            else:
                from app.api.disassembly_manufacturing_no_opening import apply_manufacturing_cost_no_opening
                apply_manufacturing_cost_no_opening(
                    app_data,
                    prediction_period,
                    manufacturing_cost_by_code,
                    manufacturing_cost_details_by_code,
                    allowed_material_codes,
                    valid_material_codes,
                    material_info,
                    original_data,
                    deducted_data,
                    product_value_by_code,
                    category_total_value,
                    classify_by_product_name,
                    indirect_labor_result,
                )

                                        
        except Exception as e:
            print(f"计算制造费用失败: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # 2.6 期间费用（按物料代码统计）
        period_cost_by_code = {}
        
        def _calculate_period_cost_by_category(app_data, total_period_cost):
            """
            计算每个分类的期间费用，模拟"期间费用"页"四机一脑分类"卡片的计算逻辑
            
            该函数使用"一次拆解产物产值"数据计算分类占比，与"期间费用"页的计算方式一致
            
            Args:
                app_data: 应用数据管理器
                total_period_cost: 总期间费用
                
            Returns:
                dict: {分类: 期间费用}，例如 {'电视': 1000.0, '冰箱': 800.0, ...}
            """
            from app.api.data_management_api import calculate_disassembly_product_output_value_data
            
            categories = ['电视', '冰箱', '空调', '洗衣机', '电脑']
            category_period_cost = {cat: 0.0 for cat in categories}
            
            if total_period_cost <= 0:
                return category_period_cost
            
            # 获取"一次拆解产物产值"数据（与期间费用页一致）
            success, output_value_data, error_message = calculate_disassembly_product_output_value_data(app_data)
            
            if not success or not output_value_data:
                print(f"获取一次拆解产物产值数据失败: {error_message}")
                return category_period_cost
            
            # 计算各分类产值和总产值
            category_value = {cat: 0.0 for cat in categories}
            total_value = 0.0
            
            for item in output_value_data:
                category = item.get('分类', '')
                material_value = float(item.get('物料产值（元）', 0) or 0)
                
                if category in categories and material_value > 0:
                    category_value[category] += material_value
                    total_value += material_value
            
            # 计算各分类的期间费用 = 总期间费用 × (分类产值 / 总产值)
            if total_value > 0:
                for category in categories:
                    if category_value[category] > 0:
                        ratio = category_value[category] / total_value
                        category_period_cost[category] = total_period_cost * ratio
            
            return category_period_cost
        
        try:
            period_cost_result = calculate_period_cost(
                app_data, prediction_period,
                quality_manager_ratio, quality_group_ratio, warehouse_group_ratio
            )
            if period_cost_result and not period_cost_result.get('error'):
                # 初始化期间费用（包括所有物料代码）
                allowed_material_codes = valid_material_codes | set(material_info_kg.keys())
                for material_code in allowed_material_codes:
                    period_cost_by_code[material_code] = 0.0
                
                # 从期间费用基础数据中提取四机一脑分类的费用
                # 根据文档，期间费用应该从"期间费用"页的"四机一脑分类"卡片获取数据
                # 计算公式：按类别的费用 × (物料代码产值 / 该分类的总产值)
                categories = ['电视', '冰箱', '空调', '洗衣机', '电脑']
                
                # 获取总期间费用（所有列的总和）
                totals = period_cost_result.get('totals', {})
                total_period_cost = sum(totals.values()) if totals else 0.0
                
                if total_period_cost > 0:
                    # 使用"一次拆解产物产值"数据计算各分类的期间费用（与期间费用页一致）
                    category_period_cost = _calculate_period_cost_by_category(app_data, total_period_cost)
                    
                    # 计算每个类别的总产物价值（使用"产物价值"列，即product_value_by_code）
                    # 这是"当期拆解收益测算分析表"页的产物价值列数据
                    category_value_total = {cat: 0.0 for cat in categories}
                    
                    # 从material_info（单位：台）汇总
                    for material_code in valid_material_codes:
                        material_value = product_value_by_code.get(material_code, 0.0)
                        if material_value > 0:
                            info = material_info.get(material_code, {})
                            material_category = info.get('类别', '')
                            if material_category and material_category in categories:
                                category_value_total[material_category] += material_value
                    
                    # 从material_info_kg（单位：KG）汇总
                    for material_code in material_info_kg.keys():
                        material_value = product_value_by_code.get(material_code, 0.0)
                        if material_value > 0:
                            info_kg = material_info_kg.get(material_code, {})
                            material_category = info_kg.get('类别', '')
                            if material_category and material_category in categories:
                                category_value_total[material_category] += material_value
                    
                    # 按物料代码分摊期间费用
                    # 使用"产物价值"列（product_value_by_code）进行物料代码级别的分摊
                    # 对于每个物料代码，根据其类别，使用product_value_by_code[material_code]作为该物料代码在该类别的产值
                    
                    # 处理valid_material_codes（单位：台）
                    for material_code in valid_material_codes:
                        material_value = product_value_by_code.get(material_code, 0.0)
                        if material_value > 0:
                            info = material_info.get(material_code, {})
                            material_category = info.get('类别', '')
                            if material_category and material_category in categories:
                                category_value = category_value_total.get(material_category, 0.0)
                                category_cost = category_period_cost.get(material_category, 0.0)
                                
                                if category_value > 0 and category_cost > 0:
                                    # 分摊公式：按类别的费用 × (物料代码产值 / 该分类的总产值)
                                    allocated_cost = category_cost * (material_value / category_value)
                                    period_cost_by_code[material_code] += allocated_cost
                    
                    # 处理material_info_kg（单位：KG）
                    for material_code in material_info_kg.keys():
                        material_value = product_value_by_code.get(material_code, 0.0)
                        if material_value > 0:
                            info_kg = material_info_kg.get(material_code, {})
                            material_category = info_kg.get('类别', '')
                            if material_category and material_category in categories:
                                category_value = category_value_total.get(material_category, 0.0)
                                category_cost = category_period_cost.get(material_category, 0.0)
                                
                                if category_value > 0 and category_cost > 0:
                                    # 分摊公式：按类别的费用 × (物料代码产值 / 该分类的总产值)
                                    allocated_cost = category_cost * (material_value / category_value)
                                    period_cost_by_code[material_code] += allocated_cost
                                        
        except Exception as e:
            print(f"计算期间费用失败: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # ========== 步骤3: 汇总计算并生成结果 ==========
        # 产物价值 = 一次拆解产物产值 + 深加工产物产值
        # product_value_by_code 已经包含了两部分的累加结果
        print(f"产物价值汇总：共 {len(product_value_by_code)} 个物料代码有产物价值数据")
        total_product_value = sum(product_value_by_code.values())
        print(f"产物价值总计：{total_product_value:.2f} 元")
        
        # ========== 第一类数据：单位"台" ==========
        # 处理来自"拆解物原料成本"页的物料代码（单位：台）
        for material_code in valid_material_codes:
            info = material_info.get(material_code, {})
            subsidy_income = subsidy_income_by_code.get(material_code, 0)
            # 产物价值 = 一次拆解产物产值 + 深加工产物产值（已累加在product_value_by_code中）
            product_value = product_value_by_code.get(material_code, 0)
            material_cost = material_cost_by_code.get(material_code, 0)
            direct_labor = direct_labor_by_code.get(material_code, 0)
            manufacturing_cost = manufacturing_cost_by_code.get(material_code, 0)
            period_cost = period_cost_by_code.get(material_code, 0)
            
            disassembly_qty_tai = info.get('拆解数量', 0)  # 单位：台
            # 单台产物价值 = 产物价值 ÷ 拆解数量（台）
            unit_product_value = product_value / disassembly_qty_tai if disassembly_qty_tai > 0 else 0
            material_price_diff = product_value - material_cost
            unit_material_price_diff = material_price_diff / disassembly_qty_tai if disassembly_qty_tai > 0 else 0
            
            # 汇总指标
            total_revenue = subsidy_income + product_value
            total_cost = material_cost + direct_labor + manufacturing_cost  # 不包含期间费用
            gross_profit = total_revenue - total_cost
            gross_profit_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0
            
            # 获取制造费用明细
            manufacturing_details = manufacturing_cost_details_by_code.get(material_code, {})
            indirect_labor_commission = manufacturing_details.get('间接人工提成成本', 0.0)
            fixed_cost_allocation = manufacturing_details.get('分摊固定成本明细', 0.0)
            indirect_labor_total = indirect_labor_commission + fixed_cost_allocation  # 间接人工 = 提成成本 + 分摊固定成本
            
            
            
            result_data.append({
                '类别': info.get('类别', ''),
                '物料代码': material_code,
                '物料名称': info.get('物料名称', ''),
                '拆解数量': round(disassembly_qty_tai, 2),
                '单位': '台',
                '基金补贴收入': round(subsidy_income, 2),
                '产物价值': round(product_value, 2),
                '单台产物价值': round(unit_product_value, 2),
                '材料成本': round(material_cost, 2),
                '材料价差': round(material_price_diff, 2),
                '单台材料价差': round(unit_material_price_diff, 2),
                '直接人工': round(direct_labor, 6) if not consider_opening_stock else round(direct_labor, 2),
                '间接人工': round(indirect_labor_total, 2),
                '与拆解量相关的费用': round(manufacturing_details.get('与拆解量相关的费用', 0.0), 2),
                '与电机入库量相关的费用': round(manufacturing_details.get('与电机入库量相关的费用', 0.0), 2),
                '预计月均费用分摊': round(manufacturing_details.get('预计月均费用分摊', 0.0), 2),
                '环保费': round(manufacturing_details.get('环保费', 0.0), 2),
                '公共费用分摊': round(manufacturing_details.get('公共费用分摊', 0.0), 2),
                '屏费用分摊': round(manufacturing_details.get('屏费用分摊', 0.0), 2),
                '制造费用间接人工分摊': round(manufacturing_details.get('制造费用间接人工分摊', 0.0), 2),
                '制造费用公共成本分摊': round(manufacturing_details.get('制造费用公共成本分摊', 0.0), 2),
                '制造费用': round(manufacturing_cost, 2),
                '期间费用': round(period_cost, 2),
                '收入总额': round(total_revenue, 2),
                '成本总额': round(total_cost, 2),
                '毛利额': round(gross_profit, 2),
                '毛利率': round(gross_profit_margin, 2)
            })
        
        # ========== 第二类数据：单位"KG" ==========
        # 仅当考虑期初库存和库存结余时输出被减扣数据对应的 KG 行
        if consider_opening_stock:
            # 处理来自"被减扣数据（手工）"的"原物料代码"（单位：KG），作为独立的物料代码行
            for material_code, info_kg in material_info_kg.items():
                # 获取相应的指标（如果存在则使用，不存在则使用0或空值）
                subsidy_income = subsidy_income_by_code.get(material_code, 0)
                product_value = product_value_by_code.get(material_code, 0)
                material_cost = material_cost_kg_by_code.get(material_code, 0)
                direct_labor = direct_labor_by_code.get(material_code, 0)
                manufacturing_cost = manufacturing_cost_by_code.get(material_code, 0)
                period_cost = period_cost_by_code.get(material_code, 0)
                
                disassembly_qty_kg = info_kg.get('拆解数量', 0)  # 单位：KG
                # 单台产物价值 = 产物价值 ÷ 拆解数量（KG）
                unit_product_value_kg = product_value / disassembly_qty_kg if disassembly_qty_kg > 0 else 0
                material_price_diff = product_value - material_cost
                unit_material_price_diff_kg = material_price_diff / disassembly_qty_kg if disassembly_qty_kg > 0 else 0
                
                # 汇总指标
                total_revenue = subsidy_income + product_value
                total_cost = material_cost + direct_labor + manufacturing_cost  # 不包含期间费用
                gross_profit = total_revenue - total_cost
                gross_profit_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0
                
                # 获取制造费用明细
                manufacturing_details = manufacturing_cost_details_by_code.get(material_code, {})
                indirect_labor_commission = manufacturing_details.get('间接人工提成成本', 0.0)
                fixed_cost_allocation = manufacturing_details.get('分摊固定成本明细', 0.0)
                indirect_labor_total = indirect_labor_commission + fixed_cost_allocation  # 间接人工 = 提成成本 + 分摊固定成本
                
                result_data.append({
                    '类别': info_kg.get('类别', ''),
                    '物料代码': material_code,
                    '物料名称': info_kg.get('物料名称', ''),
                    '拆解数量': round(disassembly_qty_kg, 2),
                    '单位': 'KG',
                    '基金补贴收入': round(subsidy_income, 2),
                    '产物价值': round(product_value, 2),
                    '单台产物价值': round(unit_product_value_kg, 2),
                    '材料成本': round(material_cost, 2),
                    '材料价差': round(material_price_diff, 2),
                    '单台材料价差': round(unit_material_price_diff_kg, 2),
                    '直接人工': round(direct_labor, 2),
                    '间接人工': round(indirect_labor_total, 2),
                    '与拆解量相关的费用': round(manufacturing_details.get('与拆解量相关的费用', 0.0), 2),
                    '与电机入库量相关的费用': round(manufacturing_details.get('与电机入库量相关的费用', 0.0), 2),
                    '预计月均费用分摊': round(manufacturing_details.get('预计月均费用分摊', 0.0), 2),
                    '环保费': round(manufacturing_details.get('环保费', 0.0), 2),
                    '公共费用分摊': round(manufacturing_details.get('公共费用分摊', 0.0), 2),
                    '屏费用分摊': round(manufacturing_details.get('屏费用分摊', 0.0), 2),
                    '制造费用间接人工分摊': round(manufacturing_details.get('制造费用间接人工分摊', 0.0), 2),
                    '制造费用公共成本分摊': round(manufacturing_details.get('制造费用公共成本分摊', 0.0), 2),
                    '制造费用': round(manufacturing_cost, 2),
                    '期间费用': round(period_cost, 2),
                    '收入总额': round(total_revenue, 2),
                    '成本总额': round(total_cost, 2),
                    '毛利额': round(gross_profit, 2),
                    '毛利率': round(gross_profit_margin, 2)
                })
        
        # 按物料代码和单位排序（先按物料代码，再按单位：台在前，KG在后）
        result_data.sort(key=lambda x: (x['物料代码'], 0 if x.get('单位') == '台' else 1))

        # 确保 JSON 合法：Flask jsonify 会把 float('nan') 序列化为 NaN，前端 JSON.parse 会报错
        for row in result_data:
            for k, v in list(row.items()):
                if isinstance(v, bool):
                    continue
                if isinstance(v, numbers.Real):
                    fv = float(v)
                    if not math.isfinite(fv):
                        row[k] = 0.0
        
        return {
            'success': True,
            'data': result_data,
            'consider_opening_stock': consider_opening_stock,
        }
        
    except Exception as e:
        print(f"计算当期拆解收益测算分析表数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'error': str(e),
            'data': []
        }


@statistics_bp.route('/disassembly-profit-analysis', methods=['GET'])
def get_disassembly_profit_analysis():
    """获取当期拆解收益测算分析表数据"""
    try:
        from flask import request
        
        app_data = get_session_data_manager()
        prediction_period = int(request.args.get('prediction_period', 1))
        
        # 获取分摊比例参数
        quality_manager_ratio = request.args.get('quality_manager_ratio')
        quality_group_ratio = request.args.get('quality_group_ratio')
        warehouse_group_ratio = request.args.get('warehouse_group_ratio')
        
        if quality_manager_ratio is not None:
            quality_manager_ratio = float(quality_manager_ratio)
        if quality_group_ratio is not None:
            quality_group_ratio = float(quality_group_ratio)
        if warehouse_group_ratio is not None:
            warehouse_group_ratio = float(warehouse_group_ratio)
        
        consider_opening_stock = request.args.get('consider_opening_stock', 'true').lower() == 'true'
        
        result = _calculate_disassembly_profit_analysis_data(
            app_data, prediction_period, quality_manager_ratio, quality_group_ratio, warehouse_group_ratio,
            consider_opening_stock=consider_opening_stock
        )
        
        if not result.get('success'):
            return jsonify({
                'success': False,
                'error': result.get('error', '获取数据失败')
            }), 500
        
        return jsonify(result)
        
    except Exception as e:
        print(f"获取当期拆解收益测算分析表数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@statistics_bp.route('/disassembly-profit-analysis/export', methods=['GET'])
def export_disassembly_profit_analysis():
    """导出当期拆解收益测算分析表数据到Excel"""
    try:
        from flask import request, send_file
        from app.api.cost_forecast_api import (
            calculate_direct_labor_cost, 
            calculate_manufacturing_cost,
            calculate_period_cost,
            classify_by_product_name
        )
        from app.api.data_management_api import (
            calculate_disassembly_product_output_value_data
        )
        from data.base_data.price_data import load_price_data
        import io
        from datetime import datetime
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        app_data = get_session_data_manager()
        prediction_period = int(request.args.get('prediction_period', 1))
        
        # 获取分摊比例参数
        quality_manager_ratio = request.args.get('quality_manager_ratio')
        quality_group_ratio = request.args.get('quality_group_ratio')
        warehouse_group_ratio = request.args.get('warehouse_group_ratio')
        
        if quality_manager_ratio is not None:
            quality_manager_ratio = float(quality_manager_ratio)
        if quality_group_ratio is not None:
            quality_group_ratio = float(quality_group_ratio)
        if warehouse_group_ratio is not None:
            warehouse_group_ratio = float(warehouse_group_ratio)
        
        consider_opening_stock = request.args.get('consider_opening_stock', 'true').lower() == 'true'
        
        # 直接调用内部计算函数获取数据
        result = _calculate_disassembly_profit_analysis_data(
            app_data, prediction_period, quality_manager_ratio, quality_group_ratio, warehouse_group_ratio,
            consider_opening_stock=consider_opening_stock
        )
        
        if not result.get('success'):
            return jsonify({
                'success': False,
                'error': result.get('error', '获取数据失败')
            }), 500
        
        data = result.get('data', [])
        
        if not data:
            return jsonify({
                'success': False,
                'error': '没有可导出的数据'
            }), 400
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 转换为DataFrame
            df = pd.DataFrame(data)
            
            # 写入数据
            df.to_excel(writer, sheet_name='当期拆解收益测算分析表', index=False)
            
            # 设置样式
            worksheet = writer.sheets['当期拆解收益测算分析表']
            
            # 定义样式
            header_font = Font(bold=True, color="FFFFFF", name="仿宋")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            data_alignment = Alignment(horizontal="right", vertical="center")
            left_alignment = Alignment(horizontal="left", vertical="center")
            
            # 设置列宽
            column_widths = {
                '类别': 10,
                '物料代码': 15,
                '物料名称': 30,
                '拆解数量': 12,
                '单位': 8,
                '基金补贴收入': 15,
                '产物价值': 15,
                '单台产物价值': 15,
                '材料成本': 15,
                '材料价差': 15,
                '单台材料价差': 15,
                '直接人工': 15,
                '间接人工': 15,
                '与拆解量相关的费用': 18,
                '与电机入库量相关的费用': 20,
                '预计月均费用分摊': 18,
                '环保费': 15,
                '公共费用分摊': 15,
                '屏费用分摊': 15,
                '制造费用间接人工分摊': 18,
                '制造费用公共成本分摊': 18,
                '制造费用': 15,
                '期间费用': 15,
                '收入总额': 15,
                '成本总额': 15,
                '毛利额': 15,
                '毛利率': 12
            }
            
            # 需要隐藏的列
            hidden_columns = ['间接人工', '与拆解量相关的费用', '与电机入库量相关的费用', '预计月均费用分摊', '环保费', '公共费用分摊', '屏费用分摊', '制造费用间接人工分摊', '制造费用公共成本分摊']
            
            for idx, col in enumerate(df.columns, start=1):
                col_letter = get_column_letter(idx)
                width = column_widths.get(col, 15)
                worksheet.column_dimensions[col_letter].width = width
                
                # 隐藏指定的列
                if col in hidden_columns:
                    worksheet.column_dimensions[col_letter].hidden = True
            
            # 设置表头样式
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            
            # 设置数据行样式
            for row_idx in range(2, len(df) + 2):
                for col_idx, col in enumerate(df.columns, start=1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if col in ['类别', '物料代码', '物料名称']:
                        cell.alignment = left_alignment
                    else:
                        cell.alignment = data_alignment
                    
                    # 合计行加粗
                    if row_idx == len(df) + 1:
                        cell.font = Font(bold=True, name="仿宋")
            
            # 设置行高
            worksheet.row_dimensions[1].height = 30
            for row_idx in range(2, len(df) + 2):
                worksheet.row_dimensions[row_idx].height = 25

            # 毛利率列格式化为百分比
            if '毛利率' in df.columns:
                margin_col_idx = list(df.columns).index('毛利率') + 1
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=margin_col_idx)
                    if cell.value is not None and cell.value != '':
                        cell.value = float(cell.value) / 100
                        cell.number_format = '0.00%'
            
            # ========== Sheet 2: 分类汇总 ==========
            # 计算分类汇总
            category_summary = {}
            for _, row in df.iterrows():
                category = str(row.get('类别', '其他')).strip()
                if not category or category == '':
                    category = '其他'
                
                if category not in category_summary:
                    category_summary[category] = {
                        '类别': category,
                        '记录数': 0,
                        '拆解数量': 0,
                        '基金补贴收入': 0,
                        '产物价值': 0,
                        '材料成本': 0,
                        '材料价差': 0,
                        '直接人工': 0,
                        '间接人工': 0,
                        '与拆解量相关的费用': 0,
                        '与电机入库量相关的费用': 0,
                        '预计月均费用分摊': 0,
                        '环保费': 0,
                        '公共费用分摊': 0,
                        '屏费用分摊': 0,
                        '制造费用间接人工分摊': 0,
                        '制造费用公共成本分摊': 0,
                        '制造费用': 0,
                        '期间费用': 0,
                        '收入总额': 0,
                        '成本总额': 0,
                        '毛利额': 0
                    }
                
                category_summary[category]['记录数'] += 1
                category_summary[category]['拆解数量'] += float(row.get('拆解数量', 0) or 0)
                category_summary[category]['基金补贴收入'] += float(row.get('基金补贴收入', 0) or 0)
                category_summary[category]['产物价值'] += float(row.get('产物价值', 0) or 0)
                category_summary[category]['材料成本'] += float(row.get('材料成本', 0) or 0)
                category_summary[category]['材料价差'] += float(row.get('材料价差', 0) or 0)
                category_summary[category]['直接人工'] += float(row.get('直接人工', 0) or 0)
                category_summary[category]['间接人工'] += float(row.get('间接人工', 0) or 0)
                category_summary[category]['与拆解量相关的费用'] += float(row.get('与拆解量相关的费用', 0) or 0)
                category_summary[category]['与电机入库量相关的费用'] += float(row.get('与电机入库量相关的费用', 0) or 0)
                category_summary[category]['预计月均费用分摊'] += float(row.get('预计月均费用分摊', 0) or 0)
                category_summary[category]['环保费'] += float(row.get('环保费', 0) or 0)
                category_summary[category]['公共费用分摊'] += float(row.get('公共费用分摊', 0) or 0)
                category_summary[category]['屏费用分摊'] += float(row.get('屏费用分摊', 0) or 0)
                category_summary[category]['制造费用间接人工分摊'] += float(row.get('制造费用间接人工分摊', 0) or 0)
                category_summary[category]['制造费用公共成本分摊'] += float(row.get('制造费用公共成本分摊', 0) or 0)
                category_summary[category]['制造费用'] += float(row.get('制造费用', 0) or 0)
                category_summary[category]['期间费用'] += float(row.get('期间费用', 0) or 0)
                category_summary[category]['收入总额'] += float(row.get('收入总额', 0) or 0)
                category_summary[category]['成本总额'] += float(row.get('成本总额', 0) or 0)
                category_summary[category]['毛利额'] += float(row.get('毛利额', 0) or 0)
            
            # 计算毛利率并构建汇总数据
            summary_rows = []
            total_summary = {
                '类别': '合计',
                '记录数': 0,
                '拆解数量': 0,
                '基金补贴收入': 0,
                '产物价值': 0,
                '材料成本': 0,
                '材料价差': 0,
                '直接人工': 0,
                '间接人工': 0,
                '与拆解量相关的费用': 0,
                '与电机入库量相关的费用': 0,
                '预计月均费用分摊': 0,
                '环保费': 0,
                '公共费用分摊': 0,
                '屏费用分摊': 0,
                '制造费用间接人工分摊': 0,
                '制造费用公共成本分摊': 0,
                '制造费用': 0,
                '期间费用': 0,
                '收入总额': 0,
                '成本总额': 0,
                '毛利额': 0,
                '毛利率': 0
            }
            
            # 按类别排序（电视、电脑、冰箱、空调、洗衣机、其他）
            category_order = ['电视', '电脑', '冰箱', '空调', '洗衣机', '其他']
            sorted_categories = sorted(category_summary.keys(), key=lambda x: (
                category_order.index(x) if x in category_order else len(category_order)
            ))
            
            for category in sorted_categories:
                summary = category_summary[category]
                margin = (summary['毛利额'] / summary['收入总额'] * 100) if summary['收入总额'] > 0 else 0
                
                summary_rows.append({
                    '类别': summary['类别'],
                    '记录数': summary['记录数'],
                    '拆解数量': round(summary['拆解数量'], 2),
                    '基金补贴收入': round(summary['基金补贴收入'], 2),
                    '产物价值': round(summary['产物价值'], 2),
                    '材料成本': round(summary['材料成本'], 2),
                    '材料价差': round(summary['材料价差'], 2),
                    '直接人工': round(summary['直接人工'], 6) if not consider_opening_stock else round(summary['直接人工'], 2),
                    '间接人工': round(summary['间接人工'], 2),
                    '与拆解量相关的费用': round(summary['与拆解量相关的费用'], 2),
                    '与电机入库量相关的费用': round(summary['与电机入库量相关的费用'], 2),
                    '预计月均费用分摊': round(summary['预计月均费用分摊'], 2),
                    '环保费': round(summary['环保费'], 2),
                    '公共费用分摊': round(summary['公共费用分摊'], 2),
                    '屏费用分摊': round(summary['屏费用分摊'], 2),
                    '制造费用间接人工分摊': round(summary['制造费用间接人工分摊'], 2),
                    '制造费用公共成本分摊': round(summary['制造费用公共成本分摊'], 2),
                    '制造费用': round(summary['制造费用'], 2),
                    '期间费用': round(summary['期间费用'], 2),
                    '收入总额': round(summary['收入总额'], 2),
                    '成本总额': round(summary['成本总额'], 2),
                    '毛利额': round(summary['毛利额'], 2),
                    '毛利率': round(margin, 2)
                })
                
                # 累计合计
                total_summary['记录数'] += summary['记录数']
                total_summary['拆解数量'] += summary['拆解数量']
                total_summary['基金补贴收入'] += summary['基金补贴收入']
                total_summary['产物价值'] += summary['产物价值']
                total_summary['材料成本'] += summary['材料成本']
                total_summary['材料价差'] += summary['材料价差']
                total_summary['直接人工'] += summary['直接人工']
                total_summary['间接人工'] += summary['间接人工']
                total_summary['与拆解量相关的费用'] += summary['与拆解量相关的费用']
                total_summary['与电机入库量相关的费用'] += summary['与电机入库量相关的费用']
                total_summary['预计月均费用分摊'] += summary['预计月均费用分摊']
                total_summary['环保费'] += summary['环保费']
                total_summary['公共费用分摊'] += summary['公共费用分摊']
                total_summary['屏费用分摊'] += summary['屏费用分摊']
                total_summary['制造费用间接人工分摊'] += summary['制造费用间接人工分摊']
                total_summary['制造费用公共成本分摊'] += summary['制造费用公共成本分摊']
                total_summary['制造费用'] += summary['制造费用']
                total_summary['期间费用'] += summary['期间费用']
                total_summary['收入总额'] += summary['收入总额']
                total_summary['成本总额'] += summary['成本总额']
                total_summary['毛利额'] += summary['毛利额']
            
            # 计算合计的毛利率
            total_summary['毛利率'] = (total_summary['毛利额'] / total_summary['收入总额'] * 100) if total_summary['收入总额'] > 0 else 0
            total_summary['拆解数量'] = round(total_summary['拆解数量'], 2)
            total_summary['基金补贴收入'] = round(total_summary['基金补贴收入'], 2)
            total_summary['产物价值'] = round(total_summary['产物价值'], 2)
            total_summary['材料成本'] = round(total_summary['材料成本'], 2)
            total_summary['材料价差'] = round(total_summary['材料价差'], 2)
            total_summary['直接人工'] = round(total_summary['直接人工'], 2)
            total_summary['间接人工'] = round(total_summary['间接人工'], 2)
            total_summary['与拆解量相关的费用'] = round(total_summary['与拆解量相关的费用'], 2)
            total_summary['与电机入库量相关的费用'] = round(total_summary['与电机入库量相关的费用'], 2)
            total_summary['预计月均费用分摊'] = round(total_summary['预计月均费用分摊'], 2)
            total_summary['环保费'] = round(total_summary['环保费'], 2)
            total_summary['公共费用分摊'] = round(total_summary['公共费用分摊'], 2)
            total_summary['屏费用分摊'] = round(total_summary['屏费用分摊'], 2)
            total_summary['制造费用间接人工分摊'] = round(total_summary['制造费用间接人工分摊'], 2)
            total_summary['制造费用公共成本分摊'] = round(total_summary['制造费用公共成本分摊'], 2)
            total_summary['制造费用'] = round(total_summary['制造费用'], 2)
            total_summary['期间费用'] = round(total_summary['期间费用'], 2)
            total_summary['收入总额'] = round(total_summary['收入总额'], 2)
            total_summary['成本总额'] = round(total_summary['成本总额'], 2)
            total_summary['毛利额'] = round(total_summary['毛利额'], 2)
            total_summary['毛利率'] = round(total_summary['毛利率'], 2)
            
            # 添加合计行
            summary_rows.append(total_summary)
            
            # 创建分类汇总DataFrame
            summary_df = pd.DataFrame(summary_rows)
            summary_df.to_excel(writer, sheet_name='分类汇总', index=False)
            
            # 设置分类汇总表样式
            summary_worksheet = writer.sheets['分类汇总']
            
            # 设置列宽
            summary_column_widths = {
                '类别': 12,
                '记录数': 10,
                '拆解数量': 12,
                '基金补贴收入': 15,
                '产物价值': 15,
                '材料成本': 15,
                '材料价差': 15,
                '单台材料价差': 15,
                '直接人工': 15,
                '间接人工': 15,
                '与拆解量相关的费用': 18,
                '与电机入库量相关的费用': 20,
                '预计月均费用分摊': 18,
                '环保费': 15,
                '公共费用分摊': 15,
                '屏费用分摊': 15,
                '制造费用间接人工分摊': 18,
                '制造费用公共成本分摊': 18,
                '制造费用': 15,
                '期间费用': 15,
                '收入总额': 15,
                '成本总额': 15,
                '毛利额': 15,
                '毛利率': 12
            }
            
            for idx, col in enumerate(summary_df.columns, start=1):
                col_letter = get_column_letter(idx)
                width = summary_column_widths.get(col, 15)
                summary_worksheet.column_dimensions[col_letter].width = width
                
                # 隐藏指定的列
                if col in hidden_columns:
                    summary_worksheet.column_dimensions[col_letter].hidden = True
            
            # 设置表头样式
            for cell in summary_worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            
            # 设置数据行样式
            for row_idx in range(2, len(summary_df) + 2):
                for col_idx, col in enumerate(summary_df.columns, start=1):
                    cell = summary_worksheet.cell(row=row_idx, column=col_idx)
                    if col == '类别':
                        cell.alignment = left_alignment
                    else:
                        cell.alignment = data_alignment
                    
                    # 合计行加粗
                    if row_idx == len(summary_df) + 1:
                        cell.font = Font(bold=True, name="仿宋")
                        if col in ['材料价差', '毛利额']:
                            # 根据数值设置颜色
                            value = summary_df.iloc[-1][col]
                            if value < 0:
                                cell.font = Font(bold=True, name="仿宋", color="DC2626")
                            else:
                                cell.font = Font(bold=True, name="仿宋", color="059669")
            
            # 设置行高
            summary_worksheet.row_dimensions[1].height = 30
            for row_idx in range(2, len(summary_df) + 2):
                summary_worksheet.row_dimensions[row_idx].height = 25

            # 毛利率列格式化为百分比
            if '毛利率' in summary_df.columns:
                margin_col_idx = list(summary_df.columns).index('毛利率') + 1
                for row_idx in range(2, len(summary_df) + 2):
                    cell = summary_worksheet.cell(row=row_idx, column=margin_col_idx)
                    if cell.value is not None and cell.value != '':
                        cell.value = float(cell.value) / 100
                        cell.number_format = '0.00%'
        
        output.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"当期拆解收益测算分析表_{timestamp}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"导出当期拆解收益测算分析表失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500 
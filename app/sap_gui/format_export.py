# -*- coding: utf-8 -*-
"""把 SAP 原表按工厂+库位筛成报送版。"""
from __future__ import print_function

import os
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 与截图筛选一致：N301 指定库位，另含 N302 电路板-电废内转库
PLANT_LOCATIONS = {
    "N301": ("成品库", "屏项目-成品", "危废库", "原料库"),
    "N302": ("电路板-电废内转库",),
}
PLANTS = tuple(PLANT_LOCATIONS.keys())
LOCATIONS = tuple(
    loc for locs in PLANT_LOCATIONS.values() for loc in locs
)

# 报送页靠前的列；其余有值的列跟在后面
LEAD_COLS = (
    "期间",
    "工厂",
    "存储位置",
    "库位描述",
    "物料代码",
    "物料描述",
    "基本计量单位",
    "非限制使用的库存",
    "价值",
    "单价",
    "评估类",
    "物料组",
    "保管员",
)

INK = "1B3A4B"
PAPER = "FAF8F5"
STRIPE = "F3EEE6"
LINE = "D8D0C4"
QTY_COLS = ("非限制使用的库存", "转储库存", "在质量检测中", "限制使用的库存", "冻结库存", "返回冻结的库存")
MONEY_COLS = ("价值", "单价")


def _log(msg):
    from app.sap_gui.paths import log

    log(msg)


def _disp_width(text):
    n = 0.0
    for ch in str(text or ""):
        n += 2.1 if ord(ch) > 127 else 1.05
    return n


def _norm(text):
    return re.sub(r"\s+", "", str(text or "")).strip()


def _to_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value
    s = str(value).strip().replace(",", "").replace(" ", "")
    if s in ("", "-", "None"):
        return None
    try:
        return float(s)
    except ValueError:
        return value


def _location_ok(value, locations):
    text = _norm(value)
    if not text:
        return False
    for loc in locations:
        if text == loc or text.startswith(loc):
            return True
    return False


def _read_sheet(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([c if c is not None else "" for c in row])
    wb.close()
    if not rows:
        raise RuntimeError("空表: " + path)
    headers = [_norm(h) for h in rows[0]]
    # SAP 原表列名与平台标准列对齐
    alias = {"月": "期间", "物料": "物料代码"}
    headers = [alias.get(h, h) for h in headers]
    body = []
    for raw in rows[1:]:
        if not any(_norm(c) for c in raw):
            continue
        rec = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            rec[h] = raw[i] if i < len(raw) else ""
        body.append(rec)
    return headers, body


def _filter_rows(rows):
    kept = []
    for rec in rows:
        plant = _norm(rec.get("工厂", ""))
        locations = PLANT_LOCATIONS.get(plant)
        if not locations:
            continue
        if not _location_ok(rec.get("库位描述", ""), locations):
            continue
        kept.append(rec)
    return kept


def _ordered_headers(all_headers, rows):
    used = []
    seen = set()
    for name in LEAD_COLS:
        if name in all_headers:
            used.append(name)
            seen.add(name)
    for rec in rows:
        for k, v in rec.items():
            if k in seen or not k:
                continue
            if _norm(v):
                used.append(k)
                seen.add(k)
    for name in all_headers:
        if name not in seen and name:
            used.append(name)
            seen.add(name)
    return used


def _write_table(ws, headers, rows, start_row=1):
    thin = Border(
        left=Side(style="thin", color=LINE),
        right=Side(style="thin", color=LINE),
        top=Side(style="thin", color=LINE),
        bottom=Side(style="thin", color=LINE),
    )
    head_fill = PatternFill("solid", fgColor=INK)
    head_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="微软雅黑", size=10, color="243842")
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left = Alignment(horizontal="left", vertical="center", wrap_text=False)
    right = Alignment(horizontal="right", vertical="center")

    for c, name in enumerate(headers, 1):
        cell = ws.cell(start_row, c, name)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[start_row].height = 22
    last = start_row + max(len(rows), 1)
    ws.auto_filter.ref = "{}{}:{}{}".format(
        get_column_letter(1), start_row, get_column_letter(len(headers)), last
    )
    ws.freeze_panes = "A{}".format(start_row + 1)

    qty = set(QTY_COLS)
    money = set(MONEY_COLS)
    for r, rec in enumerate(rows, start_row + 1):
        ws.row_dimensions[r].height = 18
        fill = PatternFill("solid", fgColor=STRIPE if (r - start_row) % 2 == 0 else PAPER)
        for c, name in enumerate(headers, 1):
            raw = rec.get(name, "")
            value = _to_number(raw) if name in qty or name in money else raw
            cell = ws.cell(r, c, value if value not in ("", None) else None)
            cell.font = data_font
            cell.fill = fill
            cell.border = thin
            if name in qty:
                cell.number_format = "#,##0.000"
                cell.alignment = right
            elif name in money:
                cell.number_format = "#,##0.00"
                cell.alignment = right
            elif name in ("期间", "工厂", "存储位置", "评估类", "物料组"):
                cell.alignment = center
            else:
                cell.alignment = left

    for c, name in enumerate(headers, 1):
        width = _disp_width(name) + 3.2
        for rec in rows[:80]:
            width = max(width, min(42, _disp_width(rec.get(name, "")) + 1.6))
        ws.column_dimensions[get_column_letter(c)].width = max(10, width)

    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_title_rows = "1:1"
    ws.page_setup.horizontalCentered = True
    ws.oddFooter.right.text = "第 &P 页 / 共 &N 页"


def polish(src_path, dest_path=None):
    headers, all_rows = _read_sheet(src_path)
    filtered = _filter_rows(all_rows)
    show_cols = _ordered_headers(headers, filtered)
    if dest_path is None:
        root, ext = os.path.splitext(src_path)
        dest_path = root + "_筛选" + ext

    wb = Workbook()
    ws = wb.active
    ws.title = "物料库存"
    _write_table(ws, show_cols, filtered)

    ws2 = wb.create_sheet("全部明细")
    all_cols = _ordered_headers(headers, all_rows)
    _write_table(ws2, all_cols, all_rows)

    wb.save(dest_path)
    _log("polished {} rows -> {}".format(len(filtered), dest_path))
    return dest_path, len(filtered), len(all_rows)


if __name__ == "__main__":
    import glob
    import sys

    from app.sap_gui.paths import exports_dir

    src = sys.argv[1] if len(sys.argv) > 1 else None
    if not src:
        files = [
            p
            for p in glob.glob(os.path.join(exports_dir(), "E01_*.xlsx"))
            if "_筛选" not in os.path.basename(p)
        ]
        if not files:
            raise SystemExit("data/sap_exports 里没有原表，请先取数或指定文件路径")
        src = max(files, key=os.path.getmtime)
    polish(src)

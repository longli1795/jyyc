# -*- coding: utf-8 -*-
"""SAP 取数全流程：导出 → 筛选 → 写入期初库存固化文件。"""
from __future__ import print_function

import os
from datetime import datetime
from typing import Any, Dict, Optional

from app.sap_gui.paths import (
    exports_dir,
    last_month_period,
    log,
    write_job_status,
)


def _ensure_utf8_stdio():
    import sys

    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass


def _notify_flask_reload() -> None:
    """Redis 不可用时，通知正在运行的 Flask 进程从磁盘重载期初库存。"""
    try:
        import urllib.error
        import urllib.request

        port = os.environ.get("FLASK_PORT", "8080")
        url = "http://127.0.0.1:{}/api/opening-inventory/reload-from-disk".format(port)
        req = urllib.request.Request(url, data=b"{}", method="POST")
        req.add_header("Content-Type", "application/json")
        with urllib.request.urlopen(req, timeout=30) as resp:
            log("reload notify HTTP {} {}".format(resp.status, url))
    except Exception as e:
        log("reload notify skipped: {}".format(e))


def ingest_filtered_file(
    filtered_path: str,
    period: str,
    notify_http: bool = True,
    extra_meta: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """把筛选表写入 data/persistent 并加载到全局会话。"""
    from app.services.opening_inventory_store import (
        load_from_disk_into_memory,
        save_from_local_path,
    )

    meta_extra = {
        "source": "sap",
        "sap_period": period,
        "sap_exported_at": datetime.now().isoformat(timespec="seconds"),
    }
    if extra_meta:
        meta_extra.update(extra_meta)

    original = os.path.basename(filtered_path)
    ok, msg, meta = save_from_local_path(
        filtered_path,
        original_filename=original,
        extra_meta=meta_extra,
    )
    if not ok:
        raise RuntimeError(msg)

    load_ok, load_msg = load_from_disk_into_memory()
    if not load_ok:
        raise RuntimeError(load_msg)

    if notify_http:
        _notify_flask_reload()

    return {
        "message": load_msg,
        "meta": meta,
    }


def _assert_export_period(filtered_path: str, period: str) -> None:
    """写入期初库存前确认筛选表期间与本次要取的上个月一致。"""
    from openpyxl import load_workbook

    wb = load_workbook(filtered_path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            raise RuntimeError("筛选表为空，未覆盖期初库存")
        idx = None
        for i, name in enumerate(header):
            h = str(name or "").replace(" ", "")
            if h in ("期间", "月"):
                idx = i
                break
        if idx is None:
            raise RuntimeError("筛选表没有期间列，未覆盖期初库存")
        for raw in rows:
            if raw is None or idx >= len(raw):
                continue
            value = raw[idx]
            if value in (None, ""):
                continue
            got = _norm_period_value(value)
            if got != period:
                raise RuntimeError(
                    "导出期间是 {}，期望 {}，未覆盖期初库存".format(got or value, period)
                )
            return
        raise RuntimeError("筛选表没有期间数据，未覆盖期初库存")
    finally:
        wb.close()


def _norm_period_value(value) -> str:
    if isinstance(value, float):
        year = int(value)
        month = int(round((value - year) * 100))
        if 1 <= month <= 12:
            return "{}.{:02d}".format(year, month)
    s = str(value).strip().replace(" ", "").replace("-", ".").replace("/", ".")
    parts = s.split(".")
    if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
        return "{}.{:02d}".format(int(parts[0]), int(parts[1]))
    return s


def run_pipeline(
    period: Optional[str] = None,
    ingest: bool = True,
    notify_http: bool = True,
) -> Dict[str, Any]:
    """
    从 SAP 导出上个月（或指定期间）库存，筛选后可选写入期初库存。
    失败时不覆盖已成功固化的期初库存文件。
    """
    _ensure_utf8_stdio()
    from app.sap_gui.export_zmmr0010n_y import COMPANY, export_inventory
    from app.sap_gui.format_export import polish

    period = period or last_month_period()
    started = datetime.now().isoformat(timespec="seconds")
    write_job_status(
        running=True,
        phase="exporting",
        period=period,
        message="正在从 SAP 导出 {}".format(period),
        started_at=started,
        finished_at="",
        error="",
    )
    log("pipeline start period={} ingest={}".format(period, ingest))

    try:
        raw_path = export_inventory(period, dest_dir=exports_dir())
        write_job_status(phase="filtering", message="正在筛选工厂/库位", raw_path=raw_path)
        filtered_path, n_keep, n_all = polish(raw_path)
        log("filtered {}/{} -> {}".format(n_keep, n_all, filtered_path))
        _assert_export_period(filtered_path, period)

        result = {
            "success": True,
            "period": period,
            "company": COMPANY,
            "raw_path": raw_path,
            "filtered_path": filtered_path,
            "filtered_rows": n_keep,
            "all_rows": n_all,
            "ingested": False,
            "message": "已导出并筛选 {} 行（原表 {} 行）".format(n_keep, n_all),
        }

        if ingest:
            write_job_status(
                phase="ingesting",
                message="正在写入期初库存",
                filtered_path=filtered_path,
                filtered_rows=n_keep,
                all_rows=n_all,
            )
            ingest_info = ingest_filtered_file(
                filtered_path,
                period,
                notify_http=notify_http,
                extra_meta={"sap_filtered_rows": n_keep, "sap_all_rows": n_all},
            )
            result["ingested"] = True
            result["message"] = ingest_info["message"]
            result["meta"] = ingest_info.get("meta")

        finished = datetime.now().isoformat(timespec="seconds")
        write_job_status(
            running=False,
            phase="done",
            message=result["message"],
            finished_at=finished,
            raw_path=raw_path,
            filtered_path=filtered_path,
            filtered_rows=n_keep,
            all_rows=n_all,
            ingested=bool(ingest),
        )
        log("pipeline done: {}".format(result["message"]))
        return result
    except Exception as e:
        write_job_status(
            running=False,
            phase="error",
            message=str(e),
            error=str(e),
            finished_at=datetime.now().isoformat(timespec="seconds"),
        )
        log("pipeline ERROR {}".format(e))
        raise

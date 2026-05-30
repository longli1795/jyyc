import re
from urllib.parse import quote
from flask import send_file, make_response

def send_excel_file(file_obj, filename):
    """
    发送Excel文件，正确处理中文文件名编码
    支持多种浏览器兼容性
    """
    response = make_response(send_file(
        file_obj,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    ))
    
    # 设置中文文件名的正确编码，支持多种浏览器
    # 生成安全的ASCII文件名作为fallback
    ascii_filename = re.sub(r'[^\x00-\x7F]+', 'data', filename)
    
    # 方案1: RFC 6266标准的UTF-8编码（现代浏览器）
    encoded_filename = quote(filename, safe='().-_~')
    
    # 方案2: 为了更好的兼容性，同时提供多种编码方式
    try:
        # 检查是否包含中文字符
        if any('\u4e00' <= char <= '\u9fff' for char in filename):
            # 包含中文，使用多重编码策略（确保header值是ASCII兼容的）
            disposition_value = f'attachment; filename="{ascii_filename}"; filename*=UTF-8\'\'{encoded_filename}'
        else:
            # 不包含中文，直接使用文件名
            disposition_value = f'attachment; filename="{filename}"'
        
        # 确保Content-Disposition值是ASCII兼容的
        response.headers['Content-Disposition'] = disposition_value.encode('ascii', 'ignore').decode('ascii')
    except Exception as e:
        # 如果编码失败，使用fallback
        response.headers['Content-Disposition'] = f'attachment; filename="{ascii_filename}"'
    
    return response 

def create_deducted_data_excel(deducted_data):
    """创建被减扣数据Excel文件"""
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    import tempfile
    import os
    
    # 创建临时文件
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file.close()
    
    try:
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "被减扣数据"
        
        # 定义表头（包含原物料信息）
        headers = [
            '原物料代码', '原物料名称', '拆解产物编码', '拆解产物名称', 
            '计算结果(KG)', '处置类别', '期间'
        ]
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # 写入数据
        for row_idx, (_, row) in enumerate(deducted_data.iterrows(), 2):
            for col_idx, header in enumerate(headers, 1):
                value = row.get(header, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 数值列右对齐
                if header in ['计算结果(KG)']:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # 调整列宽
        column_widths = {
            'A': 15,  # 原物料代码
            'B': 25,  # 原物料名称
            'C': 15,  # 拆解产物编码
            'D': 25,  # 拆解产物名称
            'E': 18,  # 计算结果(KG)
            'F': 15,  # 处置类别
            'G': 12   # 期间
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 添加说明工作表
        ws_info = wb.create_sheet("使用说明")
        info_data = [
            ["被减扣数据Excel文件说明"],
            [""],
            ["1. 本文件包含被减扣的拆解产物数据"],
            ["2. 可以修改以下列的数值："],
            ["   - 原库存数量(TAI)：原始库存数量"],
            ["   - 单台重量(KG/台)：单台产品重量"],
            ["   - 投入产出比例：拆解投入产出比例"],
            ["   - 拆解系数：产物拆解系数"],
            [""],
            ["3. 修改后重新导入系统可自动重新计算"],
            ["4. 计算公式：计算结果(KG) = 原库存数量(TAI) × 单台重量(KG/台) × 投入产出比例 × 拆解系数"],
            [""],
            ["5. 导入要求："],
            ["   - 文件格式：.xlsx 或 .xls"],
            ["   - 文件大小：不超过16MB"],
            ["   - 必填列：原物料代码、拆解产物编码"],
            [""],
            ["6. 支持的导入模式："],
            ["   - 完全覆盖：清空现有数据，使用导入数据替换"],
            ["   - 合并更新：保留现有数据，更新重复项，新增不存在项"],
            ["   - 追加模式：保留现有数据，仅追加新记录"]
        ]
        
        for row_idx, info_row in enumerate(info_data, 1):
            for col_idx, value in enumerate(info_row, 1):
                cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # 标题行
                    cell.font = Font(bold=True, size=14)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # 调整说明工作表列宽
        ws_info.column_dimensions['A'].width = 60
        
        # 保存文件
        wb.save(temp_file.name)
        return temp_file.name
        
    except Exception as e:
        # 清理临时文件
        if os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
        raise e

def create_deducted_data_template():
    """创建被减扣数据导入模板"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import tempfile
    import os
    
    # 创建临时文件
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file.close()
    
    try:
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "被减扣数据模板"
        
        # 定义表头和示例数据（包含原物料信息）
        headers = [
            '原物料代码', '原物料名称', '拆解产物编码', '拆解产物名称', 
            '计算结果(KG)', '处置类别', '期间'
        ]
        
        # 示例数据
        sample_data = [
            ['100001', '电冰箱', '811052988', '铁及其合金黑白-防爆带', 1.234567, '深加工-打包铁', '202506'],
            ['100002', '洗衣机', '811052896', '铁及其合金空调-外壳及铁管', 6.543210, '深加工-打包铁', '202506']
        ]
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # 写入示例数据
        for row_idx, data_row in enumerate(sample_data, 2):
            for col_idx, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 数值列右对齐
                if col_idx == 5:  # 计算结果(KG) - 现在是第5列
                    cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # 调整列宽
        column_widths = {
            'A': 15,  # 原物料代码
            'B': 25,  # 原物料名称
            'C': 15,  # 拆解产物编码
            'D': 25,  # 拆解产物名称
            'E': 18,  # 计算结果(KG)
            'F': 15,  # 处置类别
            'G': 12   # 期间
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 添加使用说明工作表
        ws_info = wb.create_sheet("导入说明")
        info_data = [
            ["被减扣数据导入模板使用说明"],
            [""],
            ["一、模板说明"],
            ["1. 本模板用于导入被减扣数据"],
            ["2. 红色表头为必填字段"],
            ["3. 请保持表头格式不变"],
            ["4. 示例数据可以删除，填入实际数据"],
            [""],
            ["二、字段说明"],
            ["原物料代码：原物料的编码"],
            ["原物料名称：原物料的名称"],
            ["拆解产物编码：拆解后产物的编码（必填）"],
            ["拆解产物名称：拆解后产物的名称"],
            ["计算结果(KG)：计算得出的重量（数字，可编辑）"],
            ["处置类别：处置方式分类"],
            ["期间：数据所属期间"],
            [""],
            ["三、导入要求"],
            ["1. 文件格式：.xlsx 或 .xls"],
            ["2. 文件大小：不超过16MB"],
            ["3. 拆解产物编码不能为空"],
            ["4. 计算结果请填入有效数字"],
            [""],
            ["四、使用说明"],
            ["1. 可以直接修改计算结果(KG)列的数值"],
            ["2. 导入后系统将使用修改后的数值"],
            ["3. 支持覆盖、合并、追加三种导入模式"]
        ]
        
        for row_idx, info_row in enumerate(info_data, 1):
            for col_idx, value in enumerate(info_row, 1):
                cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # 标题行
                    cell.font = Font(bold=True, size=14, color="FF0000")
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                elif any(keyword in str(value) for keyword in ["一、", "二、", "三、", "四、"]):
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # 调整说明工作表列宽
        ws_info.column_dimensions['A'].width = 70
        
        # 保存文件
        wb.save(temp_file.name)
        return temp_file.name
        
    except Exception as e:
        # 清理临时文件
        if os.path.exists(temp_file.name):
            os.unlink(temp_file.name)
        raise e

def parse_deducted_data_excel(file):
    """解析被减扣数据Excel文件"""
    import pandas as pd
    from io import BytesIO
    
    errors = []
    data = []
    
    try:
        # 读取Excel文件
        if hasattr(file, 'read'):
            file_content = file.read()
            file.seek(0)  # 重置文件指针
            excel_file = BytesIO(file_content)
        else:
            excel_file = file
        
        # 尝试读取Excel文件
        try:
            df = pd.read_excel(excel_file, sheet_name=0)  # 读取第一个工作表
        except Exception as e:
            errors.append(f"Excel文件读取失败: {str(e)}")
            return [], errors
        
        if df.empty:
            errors.append("Excel文件中没有数据")
            return [], errors
        
        # 定义必需的列
        required_columns = ['拆解产物编码']
        numeric_columns = ['序号', '原库存数量(TAI)', '单台重量(KG/台)', '投入产出比例', '拆解系数', '计算结果(KG)']
        
        # 检查必需列
        missing_columns = []
        for col in required_columns:
            if col not in df.columns:
                missing_columns.append(col)
        
        if missing_columns:
            errors.append(f"缺少必需的列: {', '.join(missing_columns)}")
            return [], errors
        
        # 逐行验证数据
        for index, row in df.iterrows():
            row_data = {}
            row_errors = []
            
            # 复制所有列数据
            for col in df.columns:
                value = row[col]
                # 处理NaN值
                if pd.isna(value):
                    value = ''
                elif isinstance(value, float) and value.is_integer():
                    value = int(value)
                row_data[col] = value
            
            # 验证必需字段
            for col in required_columns:
                if not str(row_data.get(col, '')).strip():
                    row_errors.append(f"{col}不能为空")
            
            # 验证数值字段
            for col in numeric_columns:
                if col in row_data and str(row_data[col]).strip():
                    try:
                        float(row_data[col])
                    except (ValueError, TypeError):
                        if col != '原库存数量(TAI)' or str(row_data[col]) != '-':  # 原库存数量可以是'-'
                            row_errors.append(f"{col}必须是有效数字")
            
            # 如果有错误，标记这行数据
            if row_errors:
                row_data['error'] = '; '.join(row_errors)
            
            data.append(row_data)
        
        return data, errors
        
    except Exception as e:
        errors.append(f"文件解析失败: {str(e)}")
        return [], errors 